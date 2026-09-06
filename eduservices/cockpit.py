#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Cockpit CFO sur CAD_SAAD_LIVE.xlsx : masques Cadrage & Pilotage finement
organises + entierement connectes (formules vivantes) :
 - Pilotage : Cap strategique par campus (marque x ville) + SYNTHESE PAR CAMPUS (live)
   + rollup marque + graphes de decision (poids marques, cap->rejoue, marge/campus).
 - Cadrage  : bloc 'Projete 2027 V01 (live) vs cible' + graphe valeur par marque.
 - Helper _CALC_PNL!T = EBITDA signe par ligne -> EBITDA par campus en direct.
Aucune cellule d'entree (anchor) ni source (vues/socle/compta) n'est deplacee."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, DoughnutChart, Reference, Series
from openpyxl.chart.label import DataLabelList

F = "CAD_SAAD_LIVE.xlsx"
wb = openpyxl.load_workbook(F)

NAVY="1F3864"; BLUE="2E75B6"; BLUE_L="DDEBF7"; GREEN="548235"; GREEN_L="E2EFDA"
AMBER="BF8F00"; INPUT="FFF2CC"; GREY_D="3B3B3B"; GREY_L="F2F2F2"; WHITE="FFFFFF"
LIVE="C6E0B4"; TEAL="2E75B6"
thin=Side(style="thin",color="BFBFBF"); box=Border(thin,thin,thin,thin)
def fill(h): return PatternFill("solid",fgColor=h)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEF=Alignment(horizontal="left",vertical="center")
RIG=Alignment(horizontal="right",vertical="center")
H1=Font(size=10,bold=True,color=WHITE)
LAB=Font(size=9,color=GREY_D); LABB=Font(size=10,bold=True,color=NAVY)
VAL=Font(size=9,color="000000"); LIVEF=Font(size=9,bold=True,color="375623")
def S(ws,ref,val=None,font=None,fillc=None,align=None,border=None,fmt=None):
    c=ws[ref]
    if val is not None: c.value=val
    if font: c.font=font
    if fillc: c.fill=fill(fillc)
    if align: c.alignment=align
    if border: c.border=border
    if fmt: c.number_format=fmt
    return c

CAMPUS=[("MBWAY_PAR","MBway","Paris"),("MBWAY_LYO","MBway","Lyon"),("MBWAY_NAN","MBway","Nantes"),
 ("MBWAY_BOR","MBway","Bordeaux"),("ISCOM_PAR","ISCOM","Paris"),("ISCOM_LIL","ISCOM","Lille"),
 ("ISCOM_TLS","ISCOM","Toulouse"),("IPAC_NAN","Ipac","Nantes"),("IPAC_REN","Ipac","Rennes"),
 ("IPAC_MTP","Ipac","Montpellier"),("PIGIER_LYO","Pigier","Lyon"),("PIGIER_BOR","Pigier","Bordeaux"),
 ("TUNON_PAR","Tunon","Paris"),("TUNON_LYO","Tunon","Lyon")]
MARQUES=[("MBWAY","MBway"),("ISCOM","ISCOM"),("IPAC","Ipac"),("PIGIER","Pigier"),("TUNON","Tunon")]

# ============================================================= _CALC_PNL helper T
cp=wb["_CALC_PNL"]
cp["T1"]="EBITDA_SIGNED"
cp["T1"].font=Font(bold=True,color="808080")
for r in range(2,1348):
    cp["T%d"%r]=('=IF(E{r}=1,IF(OR(B{r}="7062",B{r}="706",B{r}="708"),P{r},'
                 'IF(AND(LEFT(B{r},1)="6",B{r}<>"6811"),-P{r},0)),0)').format(r=r)

# ============================================================= PILOTAGE
ps=wb["Pilotage"]
ps._charts=[]  # on repart des graphes
# relabel cap header (marque x ville)
S(ps,"D11","Cap strategique par campus  (marque x ville)  ·  budget acquisition rejoue = somme constante",
  font=Font(size=11,bold=True,color=NAVY))

# ---------- SYNTHESE PAR CAMPUS (live) ----------
S(ps,"D28","Synthese par campus  ·  budget 2027 V01  (tout est vivant)",
  font=Font(size=11,bold=True,color=NAVY))
ps.merge_cells("D28:O28"); ps["D28"].fill=fill(GREEN_L); ps["D28"].alignment=Alignment(indent=1,vertical="center")
for col in "EFGHIJKLMNO": ps["%s28"%col].fill=fill(GREEN_L)
heads=["Campus","Marque","Ville","Effectif","CA 2027","Prix moy.","Part CA","EBITDA",
       "Mrg EBITDA","EBITDA/etud.","Rejoue","CAC marg."]
for i,h in enumerate(heads):
    c=ps.cell(29,4+i,h); c.font=H1; c.fill=fill(GREEN); c.alignment=CEN; c.border=box
ps.row_dimensions[29].height=26
r0=30
for i,(code,mq,ville) in enumerate(CAMPUS):
    r=r0+i; band=WHITE if i%2==0 else GREY_L
    S(ps,"D%d"%r,code,font=Font(size=8,color="808080"),align=LEF,border=box,fillc=band)
    S(ps,"E%d"%r,mq,font=LAB,align=LEF,border=box,fillc=band)
    S(ps,"F%d"%r,ville,font=LAB,align=LEF,border=box,fillc=band)
    S(ps,"G%d"%r,'=SUMIFS(Moteur!$P$2:$P$175,Moteur!$D$2:$D$175,$D%d,Moteur!$B$2:$B$175,"V01")'%r,
      font=VAL,align=RIG,border=box,fillc=band,fmt="# ##0")
    S(ps,"H%d"%r,'=SUMIFS(Moteur!$R$2:$R$175,Moteur!$D$2:$D$175,$D%d,Moteur!$B$2:$B$175,"V01")'%r,
      font=LIVEF,align=RIG,border=box,fillc=LIVE,fmt="# ##0")
    S(ps,"I%d"%r,"=IFERROR(H%d/G%d,0)"%(r,r),font=VAL,align=RIG,border=box,fillc=band,fmt="# ##0")
    S(ps,"J%d"%r,"=IFERROR(H%d/SUM($H$30:$H$43),0)"%r,font=VAL,align=RIG,border=box,fillc=band,fmt="0.0%")
    S(ps,"K%d"%r,('=SUMIFS(_CALC_PNL!$T$2:$T$1347,_CALC_PNL!$A$2:$A$1347,$D%d,'
                  '_CALC_PNL!$D$2:$D$1347,"V01",_CALC_PNL!$C$2:$C$1347,"2027")')%r,
      font=LIVEF,align=RIG,border=box,fillc=LIVE,fmt="# ##0")
    S(ps,"L%d"%r,"=IFERROR(K%d/H%d,0)"%(r,r),font=VAL,align=RIG,border=box,fillc=band,fmt="0.0%")
    S(ps,"M%d"%r,"=IFERROR(K%d/G%d,0)"%(r,r),font=VAL,align=RIG,border=box,fillc=band,fmt="# ##0")
    S(ps,"N%d"%r,'=SUMIFS(Pilotage!$Q$13:$Q$26,Pilotage!$F$13:$F$26,$D%d)'%r,
      font=VAL,align=RIG,border=box,fillc=band,fmt="# ##0")
    S(ps,"O%d"%r,'=SUMIFS(Pilotage!$G$13:$G$26,Pilotage!$F$13:$F$26,$D%d)'%r,
      font=VAL,align=RIG,border=box,fillc=band,fmt="# ##0")
# --- reconciliation : sous-total campus (44) -> siege GRP (45) -> GROUPE (46) ---
GY=Font(bold=True,color=NAVY); GYW=Font(bold=True,color=WHITE)
# 44 sous-total campus
S(ps,"D44","Sous-total campus",font=GY,fillc=BLUE_L,align=LEF,border=box)
for col in ("E","F"): S(ps,"%s44"%col,"",fillc=BLUE_L,border=box)
for col,fmt in [("G","# ##0"),("H","# ##0"),("K","# ##0"),("N","# ##0")]:
    S(ps,"%s44"%col,"=SUM(%s30:%s43)"%(col,col),font=GY,fillc=BLUE_L,align=RIG,border=box,fmt=fmt)
S(ps,"I44","=IFERROR(H44/G44,0)",font=GY,fillc=BLUE_L,align=RIG,border=box,fmt="# ##0")
S(ps,"J44","=SUM(J30:J43)",font=GY,fillc=BLUE_L,align=RIG,border=box,fmt="0.0%")
S(ps,"L44","=IFERROR(K44/H44,0)",font=GY,fillc=BLUE_L,align=RIG,border=box,fmt="0.0%")
S(ps,"M44","=IFERROR(K44/G44,0)",font=GY,fillc=BLUE_L,align=RIG,border=box,fmt="# ##0")
# 45 siege / holding (GRP) : EBITDA seulement (charges de structure groupe)
S(ps,"D45","Siege / holding (GRP)",font=Font(bold=True,color="843C0C"),fillc="FCE4D6",align=LEF,border=box)
for col in ("E","F","G","H","I","J","L","M","N","O"): S(ps,"%s45"%col,"",fillc="FCE4D6",border=box)
S(ps,"K45",'=SUMIFS(_CALC_PNL!$T$2:$T$1347,_CALC_PNL!$A$2:$A$1347,"GRP",'
           '_CALC_PNL!$D$2:$D$1347,"V01",_CALC_PNL!$C$2:$C$1347,"2027")',
  font=Font(bold=True,color="843C0C"),fillc="FCE4D6",align=RIG,border=box,fmt="# ##0")
# 46 GROUPE = campus + siege
S(ps,"D46","GROUPE 2027 V01",font=GYW,fillc=NAVY,align=LEF,border=box)
for col in ("E","F"): S(ps,"%s46"%col,"",fillc=NAVY,border=box)
S(ps,"G46","=G44",font=GYW,fillc=NAVY,align=RIG,border=box,fmt="# ##0")
S(ps,"H46","=H44",font=GYW,fillc=NAVY,align=RIG,border=box,fmt="# ##0")
S(ps,"I46","=IFERROR(H46/G46,0)",font=GYW,fillc=NAVY,align=RIG,border=box,fmt="# ##0")
S(ps,"J46","=SUM(J30:J43)",font=GYW,fillc=NAVY,align=RIG,border=box,fmt="0.0%")
S(ps,"K46","=K44+K45",font=GYW,fillc=NAVY,align=RIG,border=box,fmt="# ##0")
S(ps,"L46","=IFERROR(K46/H46,0)",font=GYW,fillc=NAVY,align=RIG,border=box,fmt="0.0%")
S(ps,"M46","=IFERROR(K46/G46,0)",font=GYW,fillc=NAVY,align=RIG,border=box,fmt="# ##0")
S(ps,"N46","=SUM(N30:N43)",font=GYW,fillc=NAVY,align=RIG,border=box,fmt="# ##0")

# ---------- ROLLUP MARQUE (pour graphes), cols W:Z rows 29-34 ----------
S(ps,"W28","Rollup marque (live)",font=Font(size=10,bold=True,color=NAVY))
for i,h in enumerate(["Code","Marque","CA","EBITDA"]):
    c=ps.cell(29,23+i,h); c.font=Font(size=8,bold=True,color="808080")
for i,(code,lab) in enumerate(MARQUES):
    r=30+i
    ps.cell(r,23,code).font=Font(size=8,color="BFBFBF")
    ps.cell(r,24,lab).font=LAB
    S(ps,"Y%d"%r,'=SUMIFS(Moteur!$R$2:$R$175,Moteur!$E$2:$E$175,$W%d,Moteur!$B$2:$B$175,"V01")'%r,
      align=RIG,fmt="# ##0",font=VAL)
    S(ps,"Z%d"%r,"=SUMIFS($K$30:$K$43,$E$30:$E$43,$X%d)"%r,align=RIG,fmt="# ##0",font=VAL)

# ---------- GRAPHES DECISION ----------
# 1. combo cap : ref(N) vs rejoue(Q) barres + CAC(G) ligne
bar=BarChart(); bar.type="col"; bar.style=10; bar.title="Cap : budget acquisition reference -> rejoue (live)"
bar.height=8.5; bar.width=20
bar.add_data(Reference(ps,min_col=14,min_row=12,max_row=26),titles_from_data=True)   # N BUD_REF
bar.add_data(Reference(ps,min_col=17,min_row=12,max_row=26),titles_from_data=True)   # Q REJOUE live
bar.set_categories(Reference(ps,min_col=6,min_row=13,max_row=26))
bar.y_axis.title="EUR"; bar.gapWidth=40
line=LineChart()
line.add_data(Reference(ps,min_col=7,min_row=12,max_row=26),titles_from_data=True)   # G CAC
line.set_categories(Reference(ps,min_col=6,min_row=13,max_row=26))
line.y_axis.axId=200; line.y_axis.title="CAC (EUR)"; line.y_axis.crosses="max"
bar+=line
ps.add_chart(bar,"D46")

# 2. donut poids marque CA
dn=DoughnutChart(); dn.title="Poids des marques  ·  CA 2027"; dn.height=8.5; dn.width=9
dn.add_data(Reference(ps,min_col=25,min_row=29,max_row=34),titles_from_data=True)     # Y CA
dn.set_categories(Reference(ps,min_col=24,min_row=30,max_row=34))                     # X label
dn.dataLabels=DataLabelList(); dn.dataLabels.showPercent=True
ps.add_chart(dn,"N46")

# 3. EBITDA / etudiant par campus (bar) — efficience
mc=BarChart(); mc.type="bar"; mc.style=12; mc.title="EBITDA / etudiant par campus (2027 V01, live)"
mc.height=9; mc.width=13
mc.add_data(Reference(ps,min_col=13,min_row=29,max_row=43),titles_from_data=True)     # M EBITDA/etud
mc.set_categories(Reference(ps,min_col=4,min_row=30,max_row=43))
mc.legend=None
ps.add_chart(mc,"D63")

# 4. CA vs EBITDA par campus (bar)
ce=BarChart(); ce.type="col"; ce.style=10; ce.title="CA & EBITDA par campus (2027 V01, live)"
ce.height=9; ce.width=15
ce.add_data(Reference(ps,min_col=8,min_row=29,max_row=43),titles_from_data=True)      # H CA
ce.add_data(Reference(ps,min_col=11,min_row=29,max_row=43),titles_from_data=True)     # K EBITDA
ce.set_categories(Reference(ps,min_col=4,min_row=30,max_row=43))
ps.add_chart(ce,"N63")

for col,w in {"V":2,"W":8,"X":14,"Y":13,"Z":13}.items(): ps.column_dimensions[col].width=w

# ============================================================= CADRAGE
cad=wb["cad"]; cad._charts=[]
# bloc Projete 2027 V01 (live) vs cible  (rows 50-54, libres)
S(cad,"G50","Projete 2027 V01 (live)  vs  cible",font=Font(size=11,bold=True,color=NAVY))
cad.merge_cells("G50:K50"); cad["G50"].fill=fill(GREEN_L); cad["G50"].alignment=Alignment(indent=1,vertical="center")
for col in "HIJK": cad["%s50"%col].fill=fill(GREEN_L)
proj=[("G51","CA projete 2027","H51","=Pilotage!$H$46","# ##0"),
      ("G52","EBITDA projete 2027 (apres siege)","H52","=Pilotage!$K$46","# ##0"),
      ("G53","Effectif projete","H53","=Pilotage!$G$46","# ##0"),
      ("G54","Marge EBITDA projetee","H54","=IFERROR(H52/H51,0)","0.0%")]
for lc,lab,vc,f,fmt in proj:
    S(cad,lc,lab,font=LABB,align=LEF,border=box,fillc=GREY_L)
    S(cad,vc,f,font=LIVEF,align=RIG,border=box,fillc=LIVE,fmt=fmt)
# ecart aux cibles a cote (J/K col) : croissance projetee vs H56, marge vs H57
S(cad,"J53","Croissance CA projetee",font=LAB,align=RIG)
S(cad,"K53","=IFERROR(H51/H46-1,0)",font=LIVEF,align=RIG,border=box,fillc=LIVE,fmt="0.0%")
S(cad,"J54","Ecart marge vs cible",font=LAB,align=RIG)
S(cad,"K54","=H54-H57",font=LIVEF,align=RIG,border=box,fillc=LIVE,fmt="+0.0%;-0.0%")

# graphe valeur par marque (reference le rollup Pilotage)
vm=BarChart(); vm.type="col"; vm.style=10; vm.title="Ou se cree la valeur : CA & EBITDA 2027 par marque"
vm.height=8; vm.width=18
vm.add_data(Reference(ps,min_col=25,min_row=29,max_row=34),titles_from_data=True)  # CA
vm.add_data(Reference(ps,min_col=26,min_row=29,max_row=34),titles_from_data=True)  # EBITDA
vm.set_categories(Reference(ps,min_col=24,min_row=30,max_row=34))
cad.add_chart(vm,"M20")

wb.save(F)
print("OK cockpit : Synthese par campus + rollup marque + 5 graphes decision + bloc projete/cible.")
print("charts Pilotage=%d  cad=%d"%(len(ps._charts),len(cad._charts)))
