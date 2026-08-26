#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Genere le KIT DEMO : SQL de chaque rapport + classeur mockup (vrais chiffres).
   Reference que l'utilisateur reproduit cote Tagetik. Une feuille par etape."""
import csv,glob,os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter

def num(s): return float(s.replace(',','.')) if s else 0.0
def load(pat):
    f=glob.glob("tgk_data/*/"+pat)[0]
    with open(f,encoding="utf-8") as fh: return list(csv.DictReader(fh,delimiter=';'))
LAB={'MBWAY':'MBway','ISCOM':'ISCOM','IPAC':'Ipac','PIGIER':'Pigier','TUNON':'Tunon'}
MARQUES=['MBway','ISCOM','Ipac','Pigier','Tunon']

# ================= DONNEES =================
socle=load("AW_002_000002_000001.csv"); cap=load("AW_002_000008_000002.csv"); compta=load("AW_002_000004_000001.csv")
agg=defaultdict(lambda:defaultdict(float))
for r in socle:
    ex=r['EXERCICE']; m=LAB[r['ENTITY'].split('_')[0]]; k=(ex,m)
    for c in ['VOL_LEAD','VOL_CAND','VOL_ADMIS','VOL_NEW','VOL_EFF','VOL_LEAD_PAY','DEPENSE_ACQ','DEPENSE_MARQUE']:
        agg[k][c]+=num(r[c])
    agg[k]['CA']+=num(r['VOL_EFF'])*num(r['REV_STUD'])+num(r['VOL_NEW'])*num(r['REV_FRAIS_INS'])
def G(ex,c): return sum(agg[(ex,m)][c] for m in MARQUES)

# 2027 (verifie), scenarios, bridge, allocation
SCEN={'Cadrage':dict(ca=24120981,ebitda=3875895,eff=3175),
      'Optimiste':dict(ca=26233655,ebitda=6156444,eff=3394),
      'Prudent':dict(ca=22092924,ebitda=1772911,eff=2912)}
BRIDGE=[('CA 2026',22544725,'base'),('Effet volume',1030085,'+'),('Effet prix',522432,'+'),
        ('Effet mix',23739,'+'),('CA 2027 — Cadrage',24120981,'total')]
ALLOC={'Cadrage':(9075831.80,743978.40,2990161.31,12809971.51),
       'Optimiste':(8810615.72,811612.80,2894009.99,12516238.51),
       'Prudent':(9362092.88,642526.80,3092074.35,13096694.02)}

# ================= SQL =================
os.makedirs("tagetik",exist_ok=True)
SQL={
"V_TENDANCE":"""-- Tendance historique : leads, CA, depenses, CAC par exercice x marque (Socle CRM)
CREATE OR REPLACE VIEW V_TENDANCE AS
SELECT EXERCICE, SUBSTR_BEFORE(ENTITY,'_') AS MARQUE,
    SUM(VOL_LEAD)                                       AS LEADS,
    SUM(VOL_NEW)                                        AS INSCRITS,
    SUM(VOL_EFF*REV_STUD + VOL_NEW*REV_FRAIS_INS)       AS CA,
    SUM(DEPENSE_ACQ)                                    AS DEPENSE_ACQ,
    SUM(DEPENSE_MARQUE)                                 AS DEPENSE_MARQUE,
    SUM(DEPENSE_ACQ) / NULLIF(SUM(VOL_NEW),0)           AS CAC
FROM AW_002_000002_000001
GROUP BY EXERCICE, SUBSTR_BEFORE(ENTITY,'_');""",
"V_FUNNEL":"""-- Funnel de conversion : lead -> candidat -> admis -> inscrit + taux (Socle CRM)
CREATE OR REPLACE VIEW V_FUNNEL AS
SELECT EXERCICE, SUBSTR_BEFORE(ENTITY,'_') AS MARQUE,
    SUM(VOL_LEAD)  AS LEADS, SUM(VOL_CAND) AS CANDIDATS,
    SUM(VOL_ADMIS) AS ADMIS, SUM(VOL_NEW)  AS INSCRITS,
    SUM(VOL_CAND)  / NULLIF(SUM(VOL_LEAD),0)  AS TX_LEAD_CAND,
    SUM(VOL_ADMIS) / NULLIF(SUM(VOL_CAND),0)  AS TX_CAND_ADMIS,
    SUM(VOL_NEW)   / NULLIF(SUM(VOL_ADMIS),0) AS TX_ADMIS_INSC
FROM AW_002_000002_000001
GROUP BY EXERCICE, SUBSTR_BEFORE(ENTITY,'_');""",
"V_CAC":"""-- Efficience d'acquisition : CPL (cout par lead payant) et CAC (cout par inscrit)
CREATE OR REPLACE VIEW V_CAC AS
SELECT EXERCICE, SUBSTR_BEFORE(ENTITY,'_') AS MARQUE,
    SUM(DEPENSE_ACQ)   AS DEPENSE_ACQ,
    SUM(VOL_LEAD_PAY)  AS LEADS_PAYANTS,
    SUM(VOL_NEW)       AS INSCRITS,
    SUM(DEPENSE_ACQ) / NULLIF(SUM(VOL_LEAD_PAY),0) AS CPL,
    SUM(DEPENSE_ACQ) / NULLIF(SUM(VOL_NEW),0)      AS CAC
FROM AW_002_000002_000001
GROUP BY EXERCICE, SUBSTR_BEFORE(ENTITY,'_');""",
"V_BRIDGE_CA":"""-- Pont prix / volume / mix : decomposition de la variation de CA 2026 -> 2027, par version.
-- Volume = a prix 2026 ; Prix = a effectif 2026 ; Mix = interaction (residuel). Somme = delta CA.
CREATE OR REPLACE VIEW V_BRIDGE_CA AS
WITH b26 AS (
    SELECT SUBSTR_BEFORE(ENTITY,'_') AS MARQUE, SUM(VOL_EFF) AS EFF,
        SUM(VOL_EFF*REV_STUD + VOL_NEW*REV_FRAIS_INS) AS CA
    FROM AW_002_000002_000001 WHERE EXERCICE='2026'
    GROUP BY SUBSTR_BEFORE(ENTITY,'_')
),
b27 AS ( SELECT MARQUE, VERSION, SUM(EFFECTIF) AS EFF, SUM(CA) AS CA FROM V_MOTEUR GROUP BY MARQUE, VERSION )
SELECT b27.VERSION,
    SUM(b26.CA)                                                                       AS CA_2026,
    SUM((b26.CA/NULLIF(b26.EFF,0)) * (b27.EFF - b26.EFF))                              AS EFFET_VOLUME,
    SUM(b26.EFF * ((b27.CA/NULLIF(b27.EFF,0)) - (b26.CA/NULLIF(b26.EFF,0))))           AS EFFET_PRIX,
    SUM(((b27.CA/NULLIF(b27.EFF,0)) - (b26.CA/NULLIF(b26.EFF,0))) * (b27.EFF-b26.EFF)) AS EFFET_MIX,
    SUM(b27.CA)                                                                       AS CA_2027
FROM b26 JOIN b27 ON b27.MARQUE = b26.MARQUE
GROUP BY b27.VERSION;""",
"Q_SCENARIOS":"""-- Les 3 scenarios cote a cote : CA, effectif, EBITDA, marge (2027).
SELECT
    CASE m.VERSION WHEN 'V01' THEN 'Cadrage' WHEN 'V02' THEN 'Optimiste'
                   WHEN 'V03' THEN 'Prudent' END              AS "Scénario",
    SUM(m.CA)                                                 AS "CA 2027",
    SUM(m.EFFECTIF)                                           AS "Effectif",
    e.EBITDA                                                  AS "EBITDA",
    e.EBITDA / NULLIF(SUM(m.CA),0)                            AS "Marge %"
FROM V_MOTEUR m
JOIN (
    SELECT VERSION,
        SUM(CASE WHEN ACCOUNT IN ('7062','706','708')            THEN AMOUNT
                 WHEN ACCOUNT LIKE '6%' AND ACCOUNT <> '6811'    THEN -AMOUNT ELSE 0 END) AS EBITDA
    FROM V_BUDGET WHERE EXERCICE='2027' GROUP BY VERSION
) e ON e.VERSION = m.VERSION
GROUP BY m.VERSION, e.EBITDA
ORDER BY m.VERSION;""",
"V_CAP_ARBITRAGE":"""-- Arbitrage du budget d'acquisition : 3 caps proposes + cap retenu, par campus.
-- Chaque cap = budget de reference x coefficient de la logique correspondante.
CREATE OR REPLACE VIEW V_CAP_ARBITRAGE AS
SELECT ENTITY AS CAMPUS, SUBSTR_BEFORE(ENTITY,'_') AS MARQUE,
    CAC_MARGINAL,
    BUDGET_ACQ_REF                    AS BUDGET_REFERENCE,
    BUDGET_ACQ_REF * CAP_EFF          AS CAP_EFFICIENT,   -- pilote par le CAC marginal
    BUDGET_ACQ_REF * CAP_MOM          AS CAP_MOMENTUM,    -- pilote par la croissance des leads
    BUDGET_ACQ_REF * CAP_POT          AS CAP_POTENTIEL,   -- pilote par l'intensite marche
    BUDGET_ACQ_REF * CAP_RETENU       AS CAP_RETENU       -- saisie : l'arbitrage du DAF
FROM AW_002_000008_000002;""",
}
for name,sql in SQL.items():
    with open("tagetik/%s.sql"%name,"w",encoding="utf-8") as f: f.write(sql+"\n")
print("SQL ecrits :",", ".join(SQL))

# ================= CLASSEUR =================
INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; OCHRE="B3641C"; OCHRE_BG="F6E8D8"
GREY="7D8B98"; LINE="DBE2E9"; ROWALT="F5F7F9"
F_TITLE=Font(name="Calibri",size=17,bold=True,color=INK)
F_SUB=Font(name="Calibri",size=10.5,italic=True,color=GREY)
F_H=Font(name="Calibri",size=9,bold=True,color="FFFFFF")
F_LBL=Font(name="Calibri",size=10,color=INK)
F_LBLB=Font(name="Calibri",size=10,bold=True,color=INK)
F_NUM=Font(name="Calibri",size=10,color=INK)
F_ACC=Font(name="Calibri",size=10,bold=True,color=TEALD)
F_OCH=Font(name="Calibri",size=10,bold=True,color=OCHRE)
F_NOTE=Font(name="Calibri",size=9,italic=True,color=GREY)
F_EY=Font(name="Consolas",size=8.5,bold=True,color=TEALD)
fill_h=PatternFill("solid",fgColor=TEAL); fill_alt=PatternFill("solid",fgColor=ROWALT)
fill_och=PatternFill("solid",fgColor=OCHRE_BG); fill_tot=PatternFill("solid",fgColor="E2EFEB")
thin=Side(style="thin",color=LINE); bot=Border(bottom=thin)
EUR='#,##0 "€"'; PCT='0.0%'; INT='#,##0'; DEC='#,##0'
wb=openpyxl.Workbook(); wb.remove(wb.active)

def sheet(title):
    ws=wb.create_sheet(title[:31]); ws.sheet_view.showGridLines=False; return ws
def put(ws,r,c,v,font=F_NUM,fmt=None,align="left",fill=None,border=None):
    cell=ws.cell(r,c,v); cell.font=font; cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=False)
    if fmt:cell.number_format=fmt
    if fill:cell.fill=fill
    if border:cell.border=border
    return cell
def header(ws,r,cols,widths):
    for i,(txt,w) in enumerate(zip(cols,widths),1):
        put(ws,r,i,txt,F_H,align="center" if i>1 else "left",fill=fill_h)
        ws.column_dimensions[get_column_letter(i)].width=w
def titleblock(ws,eyebrow,title,sub):
    put(ws,1,1,eyebrow,F_EY); put(ws,2,1,title,F_TITLE); put(ws,3,1,sub,F_SUB)
    ws.row_dimensions[2].height=24

# ---------- 00 SOMMAIRE ----------
ws=sheet("00 · Sommaire & Drills")
titleblock(ws,"KIT DÉMO · EDUSERVICES","Du lead à la marge — kit de production","Une feuille par étape · vrais chiffres · SQL à reproduire sur Tagetik")
rows=[
 ("Temps 1","P&L 2026 réconcilié","T1 · P&L 2026","V_PNL","→ détail compte (Compta)"),
 ("Temps 1","Tendance leads/CA/dépenses","T1 · Tendance","V_TENDANCE","→ CAC (T1½)"),
 ("Temps 1½","Pont prix / volume / mix","T1½ · Bridge PVM","V_BRIDGE_CA","→ Scénarios (T2)"),
 ("Temps 1½","Efficience acquisition (CAC)","T1½ · CAC","V_CAC","↩ résolu en T4"),
 ("Temps 1½","Funnel de conversion","T1½ · Funnel","V_FUNNEL","→ leviers conversion"),
 ("Temps 2","Arbitrage du cap","T2 · Cap arbitrage","V_CAP_ARBITRAGE","→ V_CAP (calcul du cap)"),
 ("Temps 2","Les 3 scénarios","T2 · Scénarios","Q_SCENARIOS","→ P&L par scénario"),
 ("Temps 3","Allocation à la classe","T3 · Allocation","V_ALLOCATION","→ classe→campus→marque→groupe"),
 ("Temps 3½","Contribution campus","T3½ · Contribution","(masque saisie)","→ drill transparence alloc"),
 ("Temps 3½","Ouverture / fermeture","T3½ · Ouv-Ferm","(simulation)","→ Δ marge"),
 ("Temps 4","CAC — où agir","T4 · Où agir","V_CAC","↩ ferme le fil rouge"),
]
header(ws,5,["Temps","Rapport","Feuille du kit","Objet SQL / Tagetik","Drill / hyperlien"],[10,30,20,24,34])
for i,row in enumerate(rows):
    r=6+i; f=fill_alt if i%2 else None
    for c,v in enumerate(row,1):
        fnt=F_ACC if c==4 else (F_LBLB if c==2 else F_LBL)
        put(ws,r,c,v,fnt,fill=f,border=bot)
put(ws,6+len(rows)+1,1,"Légende drill : →  hyperlien vers un autre rapport    ·    ↩  boucle du fil rouge (tension plantée en T1, résolue en T4)",F_NOTE)

# ---------- T1 P&L 2026 ----------
ws=sheet("T1 · P&L 2026")
titleblock(ws,"TEMPS 1 · POSER","P&L 2026 réconcilié","Compta + CRM raccord · la référence du budget")
header(ws,5,["Poste (par nature)","Montant","% CA"],[38,16,10])
pnl=[("Produits (706 / 7062 / 708)",22544725,True),("Enseignants permanents (6411)",-3840917,False),
 ("Autres charges structure & siège",-11590524,False),("Enseignants vacataires (621)",-1581554,False),
 ("Achats directs (604 / 6063)",-1129682,False),("Frais de marque (6236)",-676344,False),
 ("Marketing campus (6231)",-434174,False)]
r=6
for lbl,val,isprod in pnl:
    f=F_LBLB if isprod else F_LBL
    put(ws,r,1,lbl,f,border=bot); put(ws,r,2,val,f,EUR,"right",border=bot)
    put(ws,r,3,val/22544725,F_NUM,PCT,"right",border=bot); r+=1
put(ws,r,1,"EBITDA (avant amortissements)",F_ACC,fill=fill_tot); put(ws,r,2,3291530,F_ACC,EUR,"right",fill=fill_tot)
put(ws,r,3,3291530/22544725,F_ACC,PCT,"right",fill=fill_tot); r+=2
put(ws,r,1,"Contrôle : CA = tendance = référence CAD (22 544 725) · EBITDA = 14,6 %",F_NOTE)
put(ws,r+1,1,"DRILL : clic sur un poste → détail par compte (Compta) et par campus.",F_NOTE)

# ---------- T1 TENDANCE ----------
ws=sheet("T1 · Tendance")
titleblock(ws,"TEMPS 1 · POSER","Tendance — l'hameçon","Les dépenses d'acquisition montent plus vite que les leads")
header(ws,5,["Indicateur","2024","2025","2026","Δ 24→26"],[26,14,14,14,12])
def line(r,lbl,vals,fmt,font=F_NUM,fill=None,pct=False):
    put(ws,r,1,lbl,F_LBLB if fill else F_LBL,fill=fill,border=bot)
    for i,v in enumerate(vals): put(ws,r,2+i,v,font,fmt,"right",fill=fill,border=bot)
    d=(vals[-1]/vals[0]-1) if vals[0] else 0
    put(ws,r,5,d,F_OCH if fill else F_NUM,'+0.0%;-0.0%',"right",fill=fill,border=bot)
line(6,"Leads",[G('2024','VOL_LEAD'),G('2025','VOL_LEAD'),G('2026','VOL_LEAD')],INT)
line(7,"Inscrits",[G('2024','VOL_NEW'),G('2025','VOL_NEW'),G('2026','VOL_NEW')],INT)
line(8,"Chiffre d'affaires",[G('2024','CA'),G('2025','CA'),G('2026','CA')],EUR)
line(9,"Dépenses d'acquisition",[G('2024','DEPENSE_ACQ'),G('2025','DEPENSE_ACQ'),G('2026','DEPENSE_ACQ')],EUR,fill=fill_och)
line(10,"CAC (coût par inscrit)",[G(e,'DEPENSE_ACQ')/G(e,'VOL_NEW') for e in ['2024','2025','2026']],EUR,fill=fill_och)
put(ws,12,1,"L'HAMEÇON : dépenses d'acquisition +21 % vs leads +12 % → le CAC se dégrade. On y revient au Temps 4.",F_OCH)

# ---------- T1½ FUNNEL ----------
ws=sheet("T1½ · Funnel")
titleblock(ws,"TEMPS 1½ · DIAGNOSTIQUER","Funnel de conversion (2026)","Où fuient les candidats — la croissance qu'on récupère sans dépenser")
header(ws,5,["Marque","Leads","Candidats","Admis","Inscrits","Lead→Cand","Cand→Adm","Adm→Insc"],[14,10,11,10,10,11,11,11])
for i,m in enumerate(MARQUES):
    a=agg[('2026',m)]; r=6+i; f=fill_alt if i%2 else None
    put(ws,r,1,m,F_LBLB,fill=f,border=bot)
    for c,v in enumerate([a['VOL_LEAD'],a['VOL_CAND'],a['VOL_ADMIS'],a['VOL_NEW']],2):
        put(ws,r,c,v,F_NUM,INT,"right",fill=f,border=bot)
    put(ws,r,6,a['VOL_CAND']/a['VOL_LEAD'],F_NUM,PCT,"right",fill=f,border=bot)
    put(ws,r,7,a['VOL_ADMIS']/a['VOL_CAND'],F_NUM,PCT,"right",fill=f,border=bot)
    put(ws,r,8,a['VOL_NEW']/a['VOL_ADMIS'],F_ACC,PCT,"right",fill=f,border=bot)
put(ws,12,1,"INSIGHT : Ipac / Pigier / Tunon convertissent moins bien admis→inscrit (~42 % vs 49 %) → levier de rendement.",F_NOTE)

# ---------- T1½ CAC ----------
ws=sheet("T1½ · CAC")
titleblock(ws,"TEMPS 1½ · DIAGNOSTIQUER","Efficience d'acquisition","Chaque euro d'acquisition rapporte-t-il autant qu'avant ? Par marque, non.")
header(ws,5,["Marque","CAC 2024","CAC 2025","CAC 2026","Dérive","Diagnostic"],[14,13,13,13,11,20])
worst=[]
for i,m in enumerate(MARQUES):
    cac={e:agg[(e,m)]['DEPENSE_ACQ']/agg[(e,m)]['VOL_NEW'] for e in ['2024','2025','2026']}
    worst.append((m,cac['2026']))
worst_m=sorted(worst,key=lambda x:-x[1])[0][0]
for i,m in enumerate(MARQUES):
    cac={e:agg[(e,m)]['DEPENSE_ACQ']/agg[(e,m)]['VOL_NEW'] for e in ['2024','2025','2026']}
    r=6+i; bad=(m=="Tunon" or m=="Ipac"); f=fill_och if bad else (fill_alt if i%2 else None)
    put(ws,r,1,m,F_OCH if bad else F_LBLB,fill=f,border=bot)
    for c,e in enumerate(['2024','2025','2026'],2):
        put(ws,r,c,cac[e],F_NUM,EUR,"right",fill=f,border=bot)
    put(ws,r,5,cac['2026']/cac['2024']-1,F_OCH if bad else F_NUM,'+0.0%',"right",fill=f,border=bot)
    diag="⚠ le plus cher" if m=="Tunon" else ("⚠ à surveiller" if m=="Ipac" else "sain")
    put(ws,r,6,diag,F_OCH if bad else F_NOTE,fill=f,border=bot)
put(ws,12,1,"FIL ROUGE : Tunon coûte 564 €/inscrit (vs ~320 € MBway/ISCOM) et se dégrade. → où agir : voir T4.",F_OCH)

# ---------- T1½ BRIDGE ----------
ws=sheet("T1½ · Bridge PVM")
titleblock(ws,"TEMPS 1½ · DIAGNOSTIQUER","Pont prix / volume / mix","Votre CA a bougé — prix, volume ou mix ? La réponse, chiffrée. (2026 → Cadrage 2027)")
header(ws,5,["Composante","Montant","Cumul"],[30,16,16])
cum=0
for i,(lbl,val,typ) in enumerate(BRIDGE):
    r=6+i
    if typ=='base': cum=val; f=fill_tot; fnt=F_LBLB
    elif typ=='total': cum=val; f=fill_tot; fnt=F_ACC
    else: cum+=val; f=None; fnt=F_LBL
    put(ws,r,1,lbl,fnt,fill=f,border=bot)
    put(ws,r,2,val,fnt,('+#,##0 "€";-#,##0 "€"') if typ=='+' else EUR,"right",fill=f,border=bot)
    put(ws,r,3,cum,fnt,EUR,"right",fill=f,border=bot)
put(ws,12,1,"LECTURE : +1,58 M€ de CA = surtout du volume (+1,03 M€) et un peu de prix (+0,52 M€). Contrôle exact.",F_NOTE)

# ---------- T2 SCENARIOS ----------
ws=sheet("T2 · Scénarios")
titleblock(ws,"TEMPS 2 · CONSTRUIRE","Les 3 scénarios","Vous ne présentez plus un budget, mais une fourchette maîtrisée")
header(ws,5,["Scénario","CA 2027","EBITDA","Marge %","Effectif","vs 2026 (CA)"],[16,15,15,11,11,13])
put(ws,6,1,"Réel 2026",F_LBL,fill=fill_alt,border=bot)
put(ws,6,2,22544725,F_LBL,EUR,"right",fill=fill_alt,border=bot); put(ws,6,3,3291530,F_LBL,EUR,"right",fill=fill_alt,border=bot)
put(ws,6,4,3291530/22544725,F_LBL,PCT,"right",fill=fill_alt,border=bot); put(ws,6,5,3036,F_LBL,INT,"right",fill=fill_alt,border=bot)
put(ws,6,6,"référence",F_NOTE,align="right",fill=fill_alt,border=bot)
for i,(nm,d) in enumerate(SCEN.items()):
    r=7+i
    put(ws,r,1,nm,F_LBLB,border=bot); put(ws,r,2,d['ca'],F_NUM,EUR,"right",border=bot)
    put(ws,r,3,d['ebitda'],F_ACC,EUR,"right",border=bot); put(ws,r,4,d['ebitda']/d['ca'],F_ACC,PCT,"right",border=bot)
    put(ws,r,5,d['eff'],F_NUM,INT,"right",border=bot); put(ws,r,6,d['ca']/22544725-1,F_NUM,'+0.0%',"right",border=bot)
put(ws,11,1,"Optimiste : marge 23,5 % (volume + productivité) · Prudent : 8,0 % (prudence commerciale). Le board voit le risque.",F_NOTE)

# ---------- T2 CAP ----------
ws=sheet("T2 · Cap arbitrage")
titleblock(ws,"TEMPS 2 · CONSTRUIRE","Arbitrage du budget d'acquisition","L'outil propose 3 logiques ; le DAF arbitre le cap retenu")
header(ws,5,["Campus","CAC marg.","Réf.","Cap Efficient","Cap Momentum","Cap Potentiel","Cap RETENU"],[13,11,11,13,13,13,13])
for i,r0 in enumerate(cap):
    r=6+i; ref=num(r0['BUDGET_ACQ_REF']); f=fill_alt if i%2 else None
    put(ws,r,1,r0['ENTITY'],F_LBLB,fill=f,border=bot)
    put(ws,r,2,num(r0['CAC_MARGINAL']),F_NUM,'#,##0 "€"',"right",fill=f,border=bot)
    put(ws,r,3,ref,F_NUM,EUR,"right",fill=f,border=bot)
    put(ws,r,4,ref*num(r0['CAP_EFF']),F_NUM,EUR,"right",fill=f,border=bot)
    put(ws,r,5,ref*num(r0['CAP_MOM']),F_NUM,EUR,"right",fill=f,border=bot)
    put(ws,r,6,ref*num(r0['CAP_POT']),F_NUM,EUR,"right",fill=f,border=bot)
    put(ws,r,7,ref*num(r0['CAP_RETENU']),F_ACC,EUR,"right",fill=fill_tot,border=bot)
rr=6+len(cap)+1
put(ws,rr,1,"Efficient = piloté par le CAC marginal · Momentum = croissance des leads · Potentiel = intensité marché.",F_NOTE)
put(ws,rr+1,1,"Cap RETENU = la SAISIE du DAF (arbitrage). DRILL : → V_CAP pour voir le calcul de chaque cap.",F_NOTE)

# ---------- T3 ALLOCATION ----------
ws=sheet("T3 · Allocation")
titleblock(ws,"TEMPS 3 · RÉVÉLER","Les 3 charges allouées, par scénario","La vérité que le reporting ne montrait pas — descend jusqu'à la classe")
header(ws,5,["Scénario","Structure campus","Frais de marque","Holding","Total alloué"],[16,17,16,15,16])
for i,(nm,vals) in enumerate(ALLOC.items()):
    r=6+i; f=fill_alt if i%2 else None
    put(ws,r,1,nm,F_LBLB,fill=f,border=bot)
    for c,v in enumerate(vals,2): put(ws,r,c,v,F_ACC if c==5 else F_NUM,EUR,"right",fill=f,border=bot)
put(ws,10,1,"Query 00 (V_ALLOC_CHARGES). Somme des marques = ces enveloppes (cascade conservative).",F_NOTE)
put(ws,11,1,"DRILL (le clou du spectacle) : classe → campus → marque → charge groupe. Masque ALLOC (déjà construit).",F_OCH)

# ---------- T3½ CONTRIBUTION ----------
ws=sheet("T3½ · Contribution")
titleblock(ws,"TEMPS 3½ · AGIR","Contribution du responsable campus (Lyon)","Le budget à plusieurs mains — chacun son périmètre, ses droits")
header(ws,5,["Registre","Ce qu'il touche","Droit"],[18,44,14])
perim=[("Il SAISIT","Effet volume (effectifs, funnel local) + effet coûts (structure locale)","écriture"),
 ("Il SUBIT","Le prix ET les clés d'allocation — décisions centrales (non exposées)","—"),
 ("Il VOIT","Comment ses charges allouées sont calculées (drill de transparence)","lecture")]
for i,(a,b,c) in enumerate(perim):
    r=6+i; f=fill_alt if i%2 else None
    fnt=F_ACC if a=="Il SAISIT" else (F_OCH if a=="Il SUBIT" else F_LBLB)
    put(ws,r,1,a,fnt,fill=f,border=bot); put(ws,r,2,b,F_LBL,fill=f,border=bot); put(ws,r,3,c,F_NOTE,align="center",fill=f,border=bot)
put(ws,10,1,"Masque de saisie filtré sur MBWAY_LYO. La contribution remonte automatiquement au consolidé.",F_NOTE)
put(ws,11,1,"DRILL : ses charges allouées → transparence jusqu'à la charge groupe (lecture seule). « Il subit, mais il comprend. »",F_OCH)

# ---------- T3½ OUV-FERM ----------
ws=sheet("T3½ · Ouv-Ferm")
titleblock(ws,"TEMPS 3½ · AGIR","Simulation ouverture / fermeture de classe","Une décision structurelle chiffrée — pas au doigt mouillé")
header(ws,5,["","Avant","Après (ex. fermeture 1 classe déficitaire)","Δ"],[22,16,30,14])
demo=[("Nb de classes (structurel)",58,57,-1,INT),("CA groupe",22092924,21980000,-112924,EUR),
 ("Marge complète",1772911,1861000,88089,EUR)]
for i,(lbl,a,b,d,fmt) in enumerate(demo):
    r=6+i; f=fill_alt if i%2 else None
    put(ws,r,1,lbl,F_LBLB,fill=f,border=bot); put(ws,r,2,a,F_NUM,fmt,"right",fill=f,border=bot)
    put(ws,r,3,b,F_NUM,fmt,"right",fill=f,border=bot)
    put(ws,r,4,d,F_ACC if d>0 else F_OCH,('+#,##0 "€";-#,##0 "€"') if fmt==EUR else '+0;-0',"right",fill=f,border=bot)
put(ws,10,1,"VOL_CLASS est figé comme STRUCTUREL (option A) précisément pour être piloté ici, à part des leviers prix/volume.",F_NOTE)
put(ws,11,1,"Chiffres illustratifs — à recalculer live sur le moteur. Fermer une classe déficitaire remonte la marge.",F_NOTE)

# ---------- T4 OU AGIR ----------
ws=sheet("T4 · Où agir")
titleblock(ws,"TEMPS 4 · BOUCLER","CAC — la tension résolue","Souvenez-vous du début : voici pourquoi, et où agir")
header(ws,5,["Marque","CAC 2026","Dérive 24→26","Action recommandée"],[14,13,13,40])
act={"Tunon":"Revoir le mix acquisition / fermer les campus non rentables",
 "Ipac":"Rééquilibrer vers l'organique, améliorer le funnel",
 "Pigier":"Surveiller, tester des canaux moins chers","ISCOM":"Sain — maintenir","MBway":"Sain — maintenir"}
order=sorted(MARQUES,key=lambda m:-(agg[('2026',m)]['DEPENSE_ACQ']/agg[('2026',m)]['VOL_NEW']))
for i,m in enumerate(order):
    cac26=agg[('2026',m)]['DEPENSE_ACQ']/agg[('2026',m)]['VOL_NEW']
    cac24=agg[('2024',m)]['DEPENSE_ACQ']/agg[('2024',m)]['VOL_NEW']
    r=6+i; bad=m in("Tunon","Ipac"); f=fill_och if bad else (fill_alt if i%2 else None)
    put(ws,r,1,m,F_OCH if bad else F_LBLB,fill=f,border=bot)
    put(ws,r,2,cac26,F_NUM,EUR,"right",fill=f,border=bot)
    put(ws,r,3,cac26/cac24-1,F_OCH if bad else F_NUM,'+0.0%',"right",fill=f,border=bot)
    put(ws,r,4,act[m],F_LBL,fill=f,border=bot)
put(ws,12,1,"BOUCLE FERMÉE : la tension plantée au Temps 1 (dépenses > leads) trouve sa réponse — Tunon & Ipac sont les cibles.",F_OCH)

# widths for title sheets col A already set via headers; ensure col A wide on text sheets
for ws in wb.worksheets:
    if ws.column_dimensions['A'].width is None or ws.column_dimensions['A'].width<12:
        ws.column_dimensions['A'].width=24
wb.save("DEMO_KIT.xlsx")
print("OK -> DEMO_KIT.xlsx (%d feuilles)"%len(wb.worksheets))
