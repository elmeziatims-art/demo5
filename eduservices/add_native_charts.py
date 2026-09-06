#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Ajoute de VRAIS graphes natifs (vivants) a cote des images de reference.
Types fiables uniquement : histogramme groupe / barres / empile (pas de combo/donut)."""
import openpyxl
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.series import DataPoint
wb=openpyxl.load_workbook("CAD_SAAD_LIVE.xlsx")
BLUE="2E86DE";GREEN="27AE60";GOLD="C9971B"
MCH=["2E86DE","27AE60","F2A900","E8743B","8E44AD"]
def col(s,h): s.graphicalProperties.solidFill=h; s.graphicalProperties.line.solidFill=h
def dpts(s):
    s.data_points=[DataPoint(idx=i,spPr=GraphicalProperties(solidFill=MCH[i])) for i in range(5)]

# ---------------- cad : trajectoire CA & EBITDA (2027 = live) ----------------
cad=wb["cad"]
cad["B33"]="Trajectoire (données du graphe vivant ci-contre)"; cad["B33"].font=Font(size=9,bold=True,color="15406E")
hdr=["Annee","CA (M EUR)","EBITDA (M EUR)"]
for i,h in enumerate(hdr): cad.cell(34,2+i,h).font=Font(size=8,bold=True,color="8A8FA0")
data=[("2024",20.064725,2.648550),("2025",21.268606,2.977604),("2026",22.544725,3.291530)]
for i,(y,ca,eb) in enumerate(data):
    r=35+i; cad.cell(r,2,y); cad.cell(r,3,ca).number_format="0.00"; cad.cell(r,4,eb).number_format="0.00"
cad["B38"]="2027"; cad["C38"]="=Pilotage!$H$46/1000000"; cad["D38"]="=Pilotage!$K$46/1000000"
cad["C38"].number_format="0.00"; cad["D38"].number_format="0.00"
ch=BarChart(); ch.type="col"; ch.style=10; ch.title="Trajectoire CA & EBITDA (2027 vivant)"
ch.height=6.5; ch.width=11
ch.add_data(Reference(cad,min_col=3,max_col=4,min_row=34,max_row=38),titles_from_data=True)
ch.set_categories(Reference(cad,min_col=2,min_row=35,max_row=38))
col(ch.series[0],BLUE); col(ch.series[1],GREEN); ch.y_axis.numFmt="0.0"
cad.add_chart(ch,"B40")

# ---------------- Pilotage : CA&EBITDA/campus + cap ref->rejoue (vivants) ----------------
ps=wb["Pilotage"]
c1=BarChart(); c1.type="col"; c1.style=10; c1.title="CA & EBITDA par campus (vivant)"; c1.height=7; c1.width=16
c1.add_data(Reference(ps,min_col=8,min_row=29,max_row=43),titles_from_data=True)   # H CA
c1.add_data(Reference(ps,min_col=11,min_row=29,max_row=43),titles_from_data=True)  # K EBITDA
c1.set_categories(Reference(ps,min_col=4,min_row=30,max_row=43))
col(c1.series[0],BLUE); col(c1.series[1],GREEN); c1.y_axis.numFmt="# ##0"
ps.add_chart(c1,"W4")
c2=BarChart(); c2.type="col"; c2.style=10; c2.title="Cap : budget acquisition reference -> rejoue (vivant)"; c2.height=7; c2.width=16
c2.add_data(Reference(ps,min_col=14,min_row=12,max_row=26),titles_from_data=True)  # N BUD_REF
c2.add_data(Reference(ps,min_col=17,min_row=12,max_row=26),titles_from_data=True)  # Q REJOUE
c2.set_categories(Reference(ps,min_col=6,min_row=13,max_row=26))
col(c2.series[0],"BDD7EE"); col(c2.series[1],BLUE); c2.y_axis.numFmt="# ##0"
ps.add_chart(c2,"W20")

# ---------------- 3_Allocation : marge/marque + decompo (vivants, rollup N-W) ----------------
al=wb["3_Allocation"]
# rollup marque : O(15)=libelle, P(16)=CA, Q..V(17-22)=VAC/PERM/ODIR/STRUCT/marque/holding, W(23)=marge, rows 5-9
m1=BarChart(); m1.type="col"; m1.style=12; m1.title="Marge complete par marque (vivant, reagit aux cles)"
m1.height=7; m1.width=12; m1.plotVisOnly=False
m1.add_data(Reference(al,min_col=23,min_row=4,max_row=9),titles_from_data=True)  # W marge
m1.set_categories(Reference(al,min_col=15,min_row=5,max_row=9)); m1.legend=None; dpts(m1.series[0])
m1.y_axis.numFmt="# ##0"; al.add_chart(m1,"B92")
m2=BarChart(); m2.type="col"; m2.grouping="stacked"; m2.overlap=100; m2.plotVisOnly=False
m2.title="Decomposition du cout complet par marque (vivant)"; m2.height=7; m2.width=12
for j in range(6): m2.add_data(Reference(al,min_col=17+j,min_row=4,max_row=9),titles_from_data=True)  # Q..V
m2.set_categories(Reference(al,min_col=15,min_row=5,max_row=9)); m2.y_axis.numFmt="# ##0"
al.add_chart(m2,"G92")

wb.calculation.fullCalcOnLoad=True
wb.save("CAD_SAAD_LIVE.xlsx")
print("OK graphes natifs vivants ajoutes (cad 1 · Pilotage 2 · Allocation 2), images conservees.")
