#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""MASQUE_CAMPUS_ISCOM_PAR.xlsx — 2 onglets :
  • « Données (V_CAMPUS_CLASSE) » : la donnée source, telle que la vue la renvoie
  • « Masque » : le cockpit directeur, en FORMULES (SUMIFS) branchées sur Données
Reproduit le pattern Tagetik (data tab + masque à formules). Iscom Paris, réel."""
import csv,glob
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as GL

INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E3EFEA"; WHITE="FFFFFF"
SOFT="51606D"; FAINT="7D8B98"; RULE="C8D2DA"; OCHRE="B3641C"; OK="1E7A55"; RED="B23A3A"
NAVY="3D4F8F"; NAVYBG="E6E9F3"; CARD="FBFCFB"; YEL="FFF6DA"
AR="Arial"
def F(sz=10,b=False,c=INK): return Font(name=AR,size=sz,bold=b,color=c)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color=RULE); med=Side(style="medium",color=RULE)
box=Border(left=thin,right=thin,top=thin,bottom=thin)
Ln=Alignment("left",vertical="center"); Rn=Alignment("right",vertical="center")
Cn=Alignment("center",vertical="center",wrap_text=True); LW=Alignment("left",vertical="center",wrap_text=True)
EUR='#,##0 "€";-#,##0 "€";"—"'; PCT='0.0%'; PCT0='0%'; NUM='#,##0'; NUM1='#,##0.0'

# ================= 1) calcul des lignes V_CAMPUS_CLASSE (Iscom Paris, 2024-26) =================
COMP=glob.glob('tgk_data/ext_*/AW_002_000004_000001.csv')[0]
SOCLE=glob.glob('tgk_data/ext_*/AW_002_000002_000001.csv')[0]
CAPm={'BAC':32,'MAS':26,'BTS':30}
fam=lambda p:'BAC' if p.startswith('BAC') else('MAS' if p.startswith('MAS') else 'BTS')
def hrs(p,m):
    if p.startswith('BAC') and m=='INIT': return 600
    if p.startswith('BAC'): return 480
    if p.startswith('MAS') and m=='INIT': return 520
    if p.startswith('MAS'): return 420
    return 700
def year(Y):
    camp=defaultdict(lambda: defaultdict(float)); hold=defaultdict(float)
    with open(COMP,encoding='utf-8') as fh:
        rr=csv.reader(fh,delimiter=';'); next(rr)
        for row in rr:
            en,acc,ex=row[0],row[1],row[2]
            if ex!=Y or acc in ('TEC_EBITDA','TEC_PL'): continue
            a=float(row[5].replace(',','.'))
            if en=='GRP':
                if acc in ('6414','6226','626','6281','6331','6333'): hold['HOLDING']+=a
                elif acc=='6236': hold['MARQUE']+=a
                continue
            if acc=='621': camp[en]['VAC']+=a
            elif acc=='6411': camp[en]['PERM']+=a
            elif acc in ('604','6063'): camp[en]['ODIR']+=a
            elif acc=='6231': camp[en]['MKT']+=a
            elif acc in ('6413','645','613','615','616','625','63511'): camp[en]['STRUCT']+=a
    cls=[]
    with open(SOCLE,encoding='utf-8') as fh:
        rr=csv.reader(fh,delimiter=';'); h=next(rr); ix={n:i for i,n in enumerate(h)}
        g=lambda row,n: float(row[ix[n]].replace(',','.'))
        for row in rr:
            if row[ix['EXERCICE']]!=Y: continue
            cls.append(dict(en=row[ix['ENTITY']],mq=row[ix['ENTITY']].split('_')[0],prog=row[ix['PROGRAMME']],
                an=row[ix['AN_ETUDE']],mod=row[ix['MODALITE']],eff=g(row,'VOL_EFF'),vcl=g(row,'VOL_CLASS'),
                new=g(row,'VOL_NEW'),ca=g(row,'VOL_EFF')*g(row,'REV_STUD')+g(row,'VOL_NEW')*g(row,'REV_FRAIS_INS'),
                hrs=g(row,'VOL_CLASS')*hrs(row[ix['PROGRAMME']],row[ix['MODALITE']])))
    E=defaultdict(lambda: defaultdict(float)); M=defaultdict(lambda: defaultdict(float)); G=defaultdict(float)
    for c in cls:
        for k,fk in [('HRS','hrs'),('EFF','eff'),('NEW','new'),('CA','ca')]:
            E[c['en']][k]+=c[fk]; M[c['mq']][k]+=c[fk]; G[k]+=c[fk]
    rows=[]
    for c in cls:
        if c['en']!='ISCOM_PAR': continue
        en,mq=c['en'],c['mq']; cp=camp[en]
        cvac=cp['VAC']*c['hrs']/E[en]['HRS']; cperm=cp['PERM']*c['hrs']/E[en]['HRS']
        codir=cp['ODIR']*c['eff']/E[en]['EFF']+cp['MKT']*c['new']/E[en]['NEW'] if E[en]['NEW'] else cp['ODIR']*c['eff']/E[en]['EFF']
        cstr=cp['STRUCT']*c['eff']/E[en]['EFF']
        share=(M[mq]['EFF']/G['EFF'])*(E[en]['CA']/M[mq]['CA'])*(c['eff']/E[en]['EFF'])
        siege=(hold['MARQUE']+hold['HOLDING'])*share
        cap=CAPm[fam(c['prog'])]
        cvar=cvac+codir; ccompl=cvac+cperm+codir+cstr+siege
        rows.append(dict(EX=Y,EN=en,MQ=mq,PROG=c['prog'],AN=c['an'],MOD=c['mod'],
            VOL_EFF=round(c['eff']),VOL_CLASS=round(c['vcl']),VOL_NEW=round(c['new']),CA=round(c['ca']),
            CAPACITE=cap,PLACES=round(c['vcl'])*cap,COST_VARIABLE=round(cvar),
            CONTRIBUTION=round(c['ca']-cvar),COST_COMPLET=round(ccompl),MARGE_COMPLETE=round(c['ca']-ccompl),
            COST_SIEGE=round(siege),POINT_MORT=round((ccompl/c['vcl'])/((c['ca']-cvar)/c['eff']),1)))
    return rows
DATA=[]
for Y in ('2024','2025','2026'): DATA+=year(Y)

wb=openpyxl.Workbook()

# ================= 2) onglet DONNÉES =================
wd=wb.active; wd.title="Données (V_CAMPUS_CLASSE)"; wd.sheet_view.showGridLines=False
COLS=['EXERCICE','ENTITY','MARQUE','PROGRAMME','AN_ETUDE','MODALITE','VOL_EFF','VOL_CLASS','VOL_NEW','CA',
      'CAPACITE','PLACES','COST_VARIABLE','CONTRIBUTION','COST_COMPLET','MARGE_COMPLETE','COST_SIEGE','POINT_MORT']
KEY={'EXERCICE':'EX','ENTITY':'EN','MARQUE':'MQ','PROGRAMME':'PROG','AN_ETUDE':'AN','MODALITE':'MOD'}
wd.merge_cells("A1:R1"); wd.cell(1,1,"  DONNÉES SOURCE — vue  V_CAMPUS_CLASSE  (wrapper sur V_ALLOCATION)  ·  filtrer ENTITY = campos").font=F(11,True,WHITE); wd.cell(1,1).fill=fill(TEALD); wd.row_dimensions[1].height=22
for j,c in enumerate(COLS,1):
    x=wd.cell(2,j,c); x.font=F(8.5,True,WHITE); x.fill=fill(TEAL); x.alignment=Cn; x.border=box
    wd.column_dimensions[GL(j)].width=11 if j>6 else (10 if j>1 else 9)
r=3
for row in DATA:
    for j,c in enumerate(COLS,1):
        v=row[KEY.get(c,c)]
        if c=='EXERCICE': v=int(v)   # numérique → matche le critère SUMIFS
        x=wd.cell(r,j,v); x.border=box; x.alignment=Ln if j<=6 else Rn
        if c in ('CA','COST_VARIABLE','CONTRIBUTION','COST_COMPLET','MARGE_COMPLETE','COST_SIEGE'): x.number_format=EUR
        elif c=='POINT_MORT': x.number_format=NUM1
        x.font=F(9,False,RED if (c=='MARGE_COMPLETE' and v<0) else INK)
    r+=1
LASTROW=r-1
wd.freeze_panes="A3"

# ================= 3) onglet MASQUE (formules SUMIFS) =================
ws=wb.create_sheet("Masque"); ws.sheet_view.showGridLines=False
for col,w in zip("ABCDEFGHI",[22,11,9,13,13,13,13,11,15]): ws.column_dimensions[col].width=w
DQ="'Données (V_CAMPUS_CLASSE)'!"
A=DQ+"$A$3:$A$%d"%LASTROW; EN=DQ+"$B$3:$B$%d"%LASTROW; PR=DQ+"$D$3:$D$%d"%LASTROW; AN=DQ+"$E$3:$E$%d"%LASTROW
def col(name):  # plage d'une colonne données
    return DQ+"$%s$3:$%s$%d"%(GL(COLS.index(name)+1),GL(COLS.index(name)+1),LASTROW)
def sif(measure,ex,prog=None,an=None):
    f='=SUMIFS(%s,%s,%s,%s,"ISCOM_PAR"'%(col(measure),A,ex,EN)
    if prog: f+=',%s,"%s"'%(PR,prog)
    if an: f+=',%s,"%s"'%(AN,an)
    return f+')'

def band(r,txt,c1=1,c2=9,bg=TEALD,fg=WHITE,sz=12,h=24):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    cc=ws.cell(r,c1,txt); cc.font=F(sz,True,fg); cc.fill=fill(bg); cc.alignment=Ln; ws.row_dimensions[r].height=h

band(1,"  COCKPIT DIRECTEUR — Iscom Paris   ·   MASQUE (formules → onglet Données)",sz=13,h=28)
ws.merge_cells("A2:I2"); ws.cell(2,1,"  Tout est en SUMIFS sur la vue V_CAMPUS_CLASSE. Change le filtre campus/exercice → tout se recalcule.").font=F(9,False,SOFT); ws.row_dimensions[2].height=18

# ① KPI 2026
band(4,"①  Mes chiffres clés 2026",bg=TEAL,sz=11)
kpi=[("Effectif",sif('VOL_EFF','2026'),NUM),("Remplissage moyen","=%s/%s"%(sif('VOL_EFF','2026')[1:],sif('PLACES','2026')[1:]),PCT0),
     ("Chiffre d'affaires",sif('CA','2026'),EUR),("Contribution",sif('CONTRIBUTION','2026'),EUR),
     ("Quote-part siège","=-1*%s"%sif('COST_SIEGE','2026')[1:],EUR),("EBITDA net",sif('MARGE_COMPLETE','2026'),EUR)]
for i,(lab,f,fmt) in enumerate(kpi):
    c0=1+(i%3)*3; r0=5+(i//3)*2
    ws.merge_cells(start_row=r0,start_column=c0,end_row=r0,end_column=c0+2); ws.cell(r0,c0,lab).font=F(8.5,True,SOFT)
    ws.merge_cells(start_row=r0+1,start_column=c0,end_row=r0+1,end_column=c0+2)
    x=ws.cell(r0+1,c0,f); x.number_format=fmt; x.font=F(14,True,TEALD if 'Contribution' in lab or 'EBITDA' in lab else INK)
ws.row_dimensions[6].height=20; ws.row_dimensions[8].height=20

# ② trajectoire
band(10,"②  Ma trajectoire 2024 → 2026",bg=TEAL,sz=11)
for j,hh in enumerate(["","2024","2025","2026"],1):
    x=ws.cell(11,j,hh); x.font=F(9,True,WHITE); x.fill=fill(TEALD); x.alignment=Cn if j>1 else Ln; x.border=box
for k,(lab,meas) in enumerate([("Chiffre d'affaires","CA"),("Effectif","VOL_EFF"),("Contribution","CONTRIBUTION"),("EBITDA net","MARGE_COMPLETE")]):
    r=12+k; ws.cell(r,1,lab).font=F(9.5); ws.cell(r,1).border=box
    for j,Y in enumerate(('2024','2025','2026')):
        x=ws.cell(r,2+j,sif(meas,Y)); x.number_format=EUR if meas!='VOL_EFF' else NUM; x.alignment=Rn; x.border=box; x.font=F(9.5)

# ③ P&L par classe 2026 (SUMIFS par classe)
band(17,"③  Mon P&L par classe 2026  (SUMIFS par programme × année)",bg=NAVY,sz=11)
hdr=["Programme ▸ Année","Rempl.","Eff.","CA","Contribution","Coût complet","Marge complète","Pt mort","Signal"]
for j,hh in enumerate(hdr,1):
    x=ws.cell(18,j,hh); x.font=F(8.5,True,WHITE); x.fill=fill(NAVY); x.alignment=Cn if j>1 else Ln; x.border=box
classes=[('BAC_COM','B1'),('BAC_COM','B2'),('BAC_COM','B3'),('MAS_COM','M1'),('MAS_COM','M2')]
r=19
for prog,an in classes:
    ws.cell(r,1,"%s %s"%(prog,an)).font=F(9.5); ws.cell(r,1).border=box; ws.cell(r,1).alignment=Ln
    eff=sif('VOL_EFF','2026',prog,an)[1:]; pl=sif('PLACES','2026',prog,an)[1:]
    ws.cell(r,2,"=%s/%s"%(eff,pl)).number_format=PCT0
    ws.cell(r,3,sif('VOL_EFF','2026',prog,an)).number_format=NUM
    ws.cell(r,4,sif('CA','2026',prog,an)).number_format=EUR
    ws.cell(r,5,sif('CONTRIBUTION','2026',prog,an)).number_format=EUR; ws.cell(r,5).font=F(9.5,False,OK)
    ws.cell(r,6,sif('COST_COMPLET','2026',prog,an)).number_format=EUR
    ws.cell(r,7,sif('MARGE_COMPLETE','2026',prog,an)).number_format=EUR
    ws.cell(r,8,sif('POINT_MORT','2026',prog,an)).number_format=NUM1
    ws.cell(r,9,'=IF(G%d<0,"🔴 piège",IF(B%d>0.9,"⚠️ saturé","🟢 sain"))'%(r,r)).alignment=Cn; ws.cell(r,9).font=F(9,True)
    for j in range(2,9): ws.cell(r,j).alignment=Rn if j<9 else Cn; ws.cell(r,j).border=box
    r+=1

ws.merge_cells("A%d:I%d"%(r+1,r+2))
ws.cell(r+1,1,"Le masque ne contient AUCUNE valeur en dur : tout est SUMIFS sur la vue V_CAMPUS_CLASSE (onglet Données). "
             "Sous Tagetik : la matrice pointe la vue, mêmes axes (ENTITY, EXERCICE, PROGRAMME, AN_ETUDE), mêmes mesures.").font=F(8.5,False,FAINT)
ws.cell(r+1,1).alignment=LW; ws.row_dimensions[r+1].height=30

out="/home/user/demo5/eduservices/MASQUE_CAMPUS_ISCOM_PAR.xlsx"
wb.save(out); print("SAVED",out,"| lignes données:",len(DATA),"| dernière ligne:",LASTROW)
