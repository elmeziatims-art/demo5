#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Remplace les graphes natifs (moches/non-rendus) par des images dataviz soignees,
scinde le tableau de leviers en Revenus/Couts, rappelle les couts a repartir sur Alloc."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
wb=openpyxl.load_workbook("CAD_SAAD_LIVE.xlsx")
def fill(h):return PatternFill("solid",fgColor=h)
BLUE_L="EAF3FC";GREEN_L="EAF7EF";NAVY="15406E";GOLD="B8860B"
thinb=Side(style="thin",color="D9D9D9")
def img(path,w):
    im=XLImage(path);
    import PIL.Image as P; iw,ih=P.open(path).size
    im.width=w; im.height=int(w*ih/iw); return im

# ================= cad : split leviers + trajectoire =================
cad=wb["cad"]; cad._charts=[]
cad["B14"]="(2)  Leviers  —  🔵 Revenus (→ CA, moteur)   ·   🟢 Couts (→ P&L)"
# bandes couleur : revenus 16-21, couts 22-26 (colonne B = parametre)
for r in range(16,22):
    cad["B%d"%r].fill=fill(BLUE_L); cad["B%d"%r].border=Border(left=Side(style="thick",color="2E86DE"),bottom=thinb,top=thinb)
for r in range(22,27):
    cad["B%d"%r].fill=fill(GREEN_L); cad["B%d"%r].border=Border(left=Side(style="thick",color="27AE60"),bottom=thinb,top=thinb)
# separateur net entre revenus et couts
for col in "BCDEFGH": cad["%s22"%col].border=Border(top=Side(style="medium",color="27AE60"),left=thinb,right=thinb,bottom=thinb)
cad["B27"]="🔵 Leviers 1-6 pilotent le CA (moteur)   ·   🟢 leviers 7-11 pilotent les couts (P&L)"
cad.add_image(img("/tmp/traj.png",760),"M4")

# ================= Pilotage : images cap+poids =================
ps=wb["Pilotage"]; ps._charts=[]
ps.add_image(img("/tmp/pilotage_charts.png",820),"B48")

# ================= 3_Allocation : rappel couts + image =================
al=wb["3_Allocation"]; al._charts=[]
A="Allocation!"
def pool(col): return '=SUMIFS(%s$%s:$%s,%s$C:$C,"2026")'%(A,col,col,A)
al["G4"]="Coûts à répartir (2026)"; al["G4"].font=Font(size=11,bold=True,color=GOLD)
pools=[("Enseignement vacataire (VAC)","V"),("Permanents (PERM)","W"),("Autres directs (ODIR)","X"),
       ("Structure campus","Y"),("Frais de marque (pub)","AC"),("Holding (siège)","Z")]
for i,(lab,col) in enumerate(pools):
    r=5+i
    al["G%d"%r]=lab; al["G%d"%r].font=Font(size=9,color="333333"); al["G%d"%r].alignment=Alignment(horizontal="left")
    al.merge_cells("G%d:J%d"%(r,r))
    al["K%d"%r]=pool(col); al["K%d"%r].number_format="# ##0"; al["K%d"%r].font=Font(size=9,bold=True,color="843C0C")
    al["K%d"%r].alignment=Alignment(horizontal="right"); al["K%d"%r].fill=fill("FBF3DE")
al["G11"]="TOTAL à répartir"; al["G11"].font=Font(size=9,bold=True,color=NAVY); al.merge_cells("G11:J11")
al["K11"]="=SUM(K5:K10)"; al["K11"].number_format="# ##0"; al["K11"].font=Font(size=10,bold=True,color="FFFFFF"); al["K11"].fill=fill(GOLD)
al["K11"].alignment=Alignment(horizontal="right")
# image dataviz alloc a droite (col N, hors maille A-L)
al.add_image(img("/tmp/alloc_charts.png",900),"N4")

wb.save("CAD_SAAD_LIVE.xlsx")
print("OK images embarquees (traj/pilotage/alloc), leviers scindes, couts a repartir rappeles.")
