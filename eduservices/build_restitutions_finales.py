#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""RESTITUTIONS_EDUSERVICES.xlsx — le classeur final : 3 restitutions, chacune
avec sa SPÉCIFICATION embarquée (source · dimensions · éléments) au-dessus du
vrai tableau. Feuilles : Moteur (acquisition + organique) · Alloué · Évolution.
Aucune nouvelle vue : V_CAMPAGNES, V_MOTEUR, V_ALLOCATION."""
import csv, glob, json
from collections import defaultdict, OrderedDict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E2EFEB"; WHITE="FFFFFF"; RULE="C8D2DA"
SOFT="51606D"; OCHRE="B3641C"; OCHRED="8A4A12"; GREEN="1E7A55"; FAINT="7D8B98"; NAVY="3D4F8F"
L0BG="DCE7EE"; L1BG="EAF0F4"; SPECBG="FBF4E6"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
def LI(i): return Alignment("left",vertical="center",indent=i)
RGT=Alignment("right",vertical="center"); CTR=Alignment("center",vertical="center",wrap_text=True); LFT=Alignment("left",vertical="center")
LFTW=Alignment("left",vertical="center",wrap_text=True)
med=Side(style="medium",color=RULE); thin=Side(style="thin",color=RULE)
EUR='#,##0 "€";-#,##0 "€";"-"'; NUM='#,##0'; N1='#,##0.0'; PCT='0.0%'; DEC3='0.000'; DPT='+0.0;-0.0;0.0'
MQ={'MBWAY':'MBway','ISCOM':'Iscom','IPAC':'Ipac','PIGIER':'Pigier','TUNON':'Tunon'}
ORDER=['MBWAY','ISCOM','IPAC','PIGIER','TUNON']
CITY={'LYO':'Lyon','PAR':'Paris','BOR':'Bordeaux','NAN':'Nantes','MTP':'Montpellier','REN':'Rennes','LIL':'Lille','TLS':'Toulouse'}
MODL={'INIT':'Initial','ALT':'Alternance'}
COMP=glob.glob('/home/user/demo5/eduservices/tgk_data/ext_*/AW_002_000004_000001.csv')[0]
SOCLE=glob.glob('/home/user/demo5/eduservices/tgk_data/ext_*/AW_002_000002_000001.csv')[0]
COEF=json.load(open('/tmp/moteur_coeffs.json'))
D_ACQ=0.08; D_MQ=0.10

def spec(ws,r,title,source,dims,elems):
    ws.cell(r,1,title).font=F(14,True,INK); r+=1
    for lab,val,col in [("Source (vue)",source,OCHRED),("Dimensions",dims,OCHRED)]:
        ws.cell(r,1,lab).font=F(9,True,col); ws.cell(r,2,val).font=F(9,False,INK); r+=1
    ws.cell(r,1,"Éléments").font=F(9,True,OCHRED); r+=1
    for lab,comp in elems:
        ws.cell(r,1,"   "+lab).font=F(8.5,True,TEALD); ws.cell(r,2,comp).font=F(8.5,False,SOFT); r+=1
    for j in range(1,9): ws.cell(r,j).fill=fill(SPECBG)
    for j in range(1,9): ws.cell(r,j).border=Border(bottom=med)
    return r+1

def hdr(ws,r,cols):
    for j,h in enumerate(cols,1):
        c=ws.cell(r,j,h); c.font=F(8.5,True,WHITE); c.fill=fill(TEAL); c.alignment=LFT if j==1 else CTR; c.border=Border(bottom=med)

wb=openpyxl.Workbook()

# ========================= FEUILLE 1 — MOTEUR =========================
ws=wb.active; ws.title="Moteur (acq + org)"; ws.sheet_view.showGridLines=False
r=spec(ws,1,"RESTITUTION — MOTEUR (acquisition + organique)","V_CAMPAGNES (calibration) + V_MOTEUR (effet)",
    "Marque ▸ Campus  (élasticité au campus : le budget est décidé là)",
    [("Élasticité acquisition","REND_ACQ"),("Élasticité marque (organique)","REND_BRAND"),
     ("Coût par lead / Conversion","CPL / CONVERSION"),("CAC marginal / Part organique","CAC_MARGINAL / PART_ORG"),
     ("Leads & budgets réf.","PAID_REF / ORG_REF · SPEND_ACQ_REF / SPEND_BRAND_REF"),
     ("Effet +Δ% (membres calculés)","Δbud=SPEND×Δ · Δleads=REF×((1+Δ)^REND−1) · inscrits=×CONV · CA=×CA/inscrit")])
# --- Table CALIBRATION ---
ws.cell(r,1,"① CALIBRATION — les coefficients (V_CAMPAGNES)").font=F(11,True,TEALD); r+=1
hdr(ws,r,["Marque ▸ Campus","Élast. acq","Élast. marque","CPL","Conversion","CAC marginal","Part org.","Budget acq","Budget marque"])
r+=1
for m in ORDER:
    ws.cell(r,1,MQ[m]).font=F(9,True,TEALD);
    for j in range(1,10): ws.cell(r,j).fill=fill(L0BG)
    r+=1
    for e in sorted([e for e in COEF if e.split('_')[0]==m]):
        c=COEF[e]
        ws.cell(r,1,CITY.get(e.split('_')[1],e.split('_')[1])).alignment=LI(1); ws.cell(r,1).font=F(9)
        ws.cell(r,2,round(c['rend_acq'],3)).number_format=DEC3; ws.cell(r,2).font=F(9,True,TEALD)
        ws.cell(r,3,round(c['rend_brand'],3)).number_format=DEC3; ws.cell(r,3).font=F(9,True,NAVY)
        ws.cell(r,4,round(c['cpl'])).number_format=EUR
        ws.cell(r,5,c['conv']).number_format=PCT
        ws.cell(r,6,round(c['cac_marg'])).number_format=EUR
        ws.cell(r,7,c['part_org']).number_format=PCT
        ws.cell(r,8,round(c['sacq'])).number_format=EUR
        ws.cell(r,9,round(c['sbr'])).number_format=EUR
        for j in range(2,10): ws.cell(r,j).alignment=RGT
        for j in range(1,10): ws.cell(r,j).border=Border(bottom=thin)
        r+=1
r+=1
# --- Table EFFET ---
ws.cell(r,1,"② EFFET — +%.0f%% acquisition & +%.0f%% marque (V_CAMPAGNES + V_MOTEUR)"%(D_ACQ*100,D_MQ*100)).font=F(11,True,TEALD); r+=1
hdr(ws,r,["Marque ▸ Campus","Δbud acq","Leads gagnés (acq)","Leads gagnés (org)","Inscrits gagnés","CA gagné","EBITDA 1re année","CAC marginal"])
r+=1
def eff(c):
    dl_a=c['paid']*((1+D_ACQ)**c['rend_acq']-1); dl_o=c['org']*((1+D_MQ)**c['rend_brand']-1)
    ins=(dl_a+dl_o)*c['conv']; dbud=c['sacq']*D_ACQ; ca=ins*c['ca_pnew']
    ebitda=ca-dbud-ins*300; cac=dbud/ins if ins else 0
    return dl_a,dl_o,ins,dbud,ca,ebitda,cac
gt=[0]*6
for m in ORDER:
    es=sorted([e for e in COEF if e.split('_')[0]==m])
    sub=[0]*6
    ws.cell(r,1,MQ[m]).font=F(9,True,TEALD)
    subrow=r; r+=1
    for e in es:
        dl_a,dl_o,ins,dbud,ca,ebitda,cac=eff(COEF[e])
        ws.cell(r,1,CITY.get(e.split('_')[1],e.split('_')[1])).alignment=LI(1); ws.cell(r,1).font=F(9)
        vals=[dbud,dl_a,dl_o,ins,ca,ebitda]
        for j,v in zip((2,3,4,5,6,7),vals):
            ws.cell(r,j,round(v) if j not in(3,4,5) else round(v,1)).number_format=(N1 if j in(3,4,5) else EUR)
            ws.cell(r,j).alignment=RGT
        ws.cell(r,7).font=F(9,False,GREEN); ws.cell(r,8,round(cac)).number_format=EUR; ws.cell(r,8).font=F(9,False,NAVY); ws.cell(r,8).alignment=RGT
        for j in range(1,9): ws.cell(r,j).border=Border(bottom=thin)
        for i,v in enumerate([dbud,dl_a,dl_o,ins,ca,ebitda]): sub[i]+=v
        r+=1
    # sous-total marque
    for j,v in zip((2,3,4,5,6,7),sub):
        ws.cell(subrow,j,round(v) if j not in(3,4,5) else round(v,1)).number_format=(N1 if j in(3,4,5) else EUR); ws.cell(subrow,j).font=F(9,True,TEALD); ws.cell(subrow,j).alignment=RGT
    ws.cell(subrow,8,round(sub[0]/sub[3]) if sub[3] else 0).number_format=EUR; ws.cell(subrow,8).font=F(9,True,NAVY); ws.cell(subrow,8).alignment=RGT
    for j in range(1,9): ws.cell(subrow,j).fill=fill(L0BG); ws.cell(subrow,j).border=Border(top=med,bottom=thin)
    for i in range(6): gt[i]+=sub[i]
ws.cell(r,1,"GROUPE").font=F(11,True,TEAL)
for j,v in zip((2,3,4,5,6,7),gt):
    ws.cell(r,j,round(v) if j not in(3,4,5) else round(v,1)).number_format=(N1 if j in(3,4,5) else EUR); ws.cell(r,j).font=F(11,True,TEAL); ws.cell(r,j).alignment=RGT
ws.cell(r,8,round(gt[0]/gt[3])).number_format=EUR; ws.cell(r,8).font=F(11,True,NAVY); ws.cell(r,8).alignment=RGT
for j in range(1,9): ws.cell(r,j).fill=fill(TEALBG); ws.cell(r,j).border=Border(top=med,bottom=med)
ws.column_dimensions['A'].width=22
for col,w in zip("BCDEFGHI",[11,15,15,13,12,13,11]): ws.column_dimensions[col].width=w+1

# ============ FEUILLES 2 & 3 — ALLOUÉ & ÉVOLUTION (cascade fidèle) ============
def hrs(prog,mod):
    if prog.startswith('BAC') and mod=='INIT': return 600
    if prog.startswith('BAC'): return 480
    if prog.startswith('MAS') and mod=='INIT': return 520
    if prog.startswith('MAS'): return 420
    if mod=='INIT': return 1000
    return 700
def compute(year):
    camp=defaultdict(lambda: defaultdict(float)); hold=defaultdict(float)
    with open(COMP,encoding='utf-8') as fh:
        rr=csv.reader(fh,delimiter=';'); next(rr)
        for row in rr:
            en,acc,ex=row[0],row[1],row[2]
            if ex!=year or acc in ('TEC_EBITDA','TEC_PL'): continue
            a=float(row[5].replace(',','.'))
            if en=='GRP':
                if acc in ('6414','6226','626','6281','6331','6333'): hold['HOLDING']+=a
                elif acc=='6236': hold['MARQUE']+=a
                continue
            if acc=='621': camp[en]['VAC']+=a
            elif acc=='6411': camp[en]['PERM']+=a
            elif acc in ('604','6063'): camp[en]['ODIR']+=a
            elif acc=='6231': camp[en]['MKT']+=a
            elif acc in ('6413','645','613','615','616','625','63511'): camp[en]['STRUCT']+=a
    cls=[]
    with open(SOCLE,encoding='utf-8') as fh:
        rr=csv.reader(fh,delimiter=';'); h=next(rr); ix={n:i for i,n in enumerate(h)}
        g=lambda row,n: float(row[ix[n]].replace(',','.'))
        for row in rr:
            if row[ix['EXERCICE']]!=year: continue
            en=row[ix['ENTITY']]; cls.append(dict(en=en,mq=en.split('_')[0],prog=row[ix['PROGRAMME']],an=row[ix['AN_ETUDE']],mod=row[ix['MODALITE']],
                eff=g(row,'VOL_EFF'),vcl=g(row,'VOL_CLASS'),new=g(row,'VOL_NEW'),ca=g(row,'VOL_EFF')*g(row,'REV_STUD')+g(row,'VOL_NEW')*g(row,'REV_FRAIS_INS'),
                hrs=g(row,'VOL_CLASS')*hrs(row[ix['PROGRAMME']],row[ix['MODALITE']])))
    E=defaultdict(lambda: defaultdict(float)); M=defaultdict(lambda: defaultdict(float)); G=defaultdict(float)
    for c in cls:
        for k,fk in [('HRS','hrs'),('EFF','eff'),('NEW','new'),('CA','ca')]:
            E[c['en']][k]+=c[fk]; M[c['mq']][k]+=c[fk]; G[k]+=c[fk]
    out=[]
    for c in cls:
        en,mq=c['en'],c['mq']; cp=camp[en]
        cvac=cp['VAC']*c['hrs']/E[en]['HRS'] if E[en]['HRS'] else 0; cperm=cp['PERM']*c['hrs']/E[en]['HRS'] if E[en]['HRS'] else 0
        codir=(cp['ODIR']*c['eff']/E[en]['EFF'] if E[en]['EFF'] else 0)+(cp['MKT']*c['new']/E[en]['NEW'] if E[en]['NEW'] else 0)
        cstr=cp['STRUCT']*c['eff']/E[en]['EFF'] if E[en]['EFF'] else 0
        share=(M[mq]['EFF']/G['EFF'])*(E[en]['CA']/M[mq]['CA'] if M[mq]['CA'] else 0)*(c['eff']/E[en]['EFF'] if E[en]['EFF'] else 0)
        siege=(hold['MARQUE']+hold['HOLDING'])*share; propre=c['ca']-(cvac+cperm+codir+cstr); net=propre-siege
        out.append(dict(mq=MQ[mq],campus=CITY.get(en.split('_')[1],en.split('_')[1]),prog=c['prog'],an=c['an'],mod=MODL.get(c['mod'],c['mod']),
            eff=c['eff'],ca=c['ca'],propre=propre,siege=siege,net=net))
    return out
def tree5(rows):
    T=OrderedDict()
    for m in ORDER:
        mm=MQ[m]; sub=[x for x in rows if x['mq']==mm]
        if not sub: continue
        T[mm]=OrderedDict()
        for camp in sorted(set(x['campus'] for x in sub)):
            T[mm][camp]=OrderedDict()
            for prog in sorted(set(x['prog'] for x in sub if x['campus']==camp)):
                T[mm][camp][prog]=OrderedDict()
                for an in sorted(set(x['an'] for x in sub if x['campus']==camp and x['prog']==prog)):
                    T[mm][camp][prog][an]=OrderedDict()
                    for x in [z for z in sub if z['campus']==camp and z['prog']==prog and z['an']==an]:
                        T[mm][camp][prog][an][x['mod']]=x
    return T
def agg(node):
    t=dict(eff=0,ca=0,propre=0,siege=0,net=0)
    def rec(n):
        if 'net' in n and 'prog' in n:
            for k in t: t[k]+=n[k];
            return
        for v in n.values(): rec(v)
    rec(node); return t

# --- Alloué ---
wa=wb.create_sheet("Alloué 2026"); wa.sheet_view.showGridLines=False; wa.sheet_properties.outlinePr.summaryBelow=False
r=spec(wa,1,"RESTITUTION — P&L chargé 2026 (avant / après allocation)","V_ALLOCATION  (Q_RAPPORT_ALLOUE)",
    "Marque ▸ Campus ▸ Programme ▸ Année ▸ Modalité",
    [("Effectif / CA","VOL_EFF / (706+7062+708)"),
     ("EBITDA propre","CA − (621+6411 + 604/6063/6231 + 6413/645/613/615/616/625/63511)"),
     ("Quote-part siège","6236 + 6414/6226/626/6281/6331/6333  (= COST_SIEGE)"),
     ("EBITDA net","CA − tous coûts  (= MARGE_COMPLETE ; dotations 6811 exclues)"),
     ("Marge %","EBITDA net ÷ CA  (membre calculé)")])
hdr(wa,r,["Marque ▸ Campus ▸ Programme ▸ Année ▸ Modalité","Effectif","CA","EBITDA propre","Q-part siège","EBITDA net","Marge %"]); r+=1
def emitA(r,label,v,lvl):
    wa.cell(r,1,label).font=F(9 if lvl<=1 else 8.5,lvl<=1); wa.cell(r,1).alignment=LI(lvl)
    wa.cell(r,2,round(v['eff'])).number_format=NUM; wa.cell(r,3,round(v['ca'])).number_format=EUR
    wa.cell(r,4,round(v['propre'])).number_format=EUR; wa.cell(r,5,round(v['siege'])).number_format=EUR
    mg=v['net']/v['ca'] if v['ca'] else 0
    wa.cell(r,6,round(v['net'])).number_format=EUR; wa.cell(r,6).font=F(9 if lvl<=1 else 8.5,lvl<=1,OCHRED if mg<0.05 else (TEALD if lvl<=1 else INK))
    wa.cell(r,7,mg).number_format=PCT; wa.cell(r,7).font=F(9 if lvl<=1 else 8.5,lvl<=1,OCHRED if mg<0.05 else TEALD)
    for j in (2,3,4,5,6,7): wa.cell(r,j).alignment=RGT
    if lvl==0:
        for j in range(1,8): wa.cell(r,j).fill=fill(L0BG); wa.cell(r,j).border=Border(top=med,bottom=thin)
    elif lvl==1:
        for j in range(1,8): wa.cell(r,j).fill=fill(L1BG)
    else:
        for j in range(1,8): wa.cell(r,j).border=Border(bottom=thin)
    if lvl>0:
        wa.row_dimensions[r].outline_level=min(lvl,4)
        if lvl>=2: wa.row_dimensions[r].hidden=True
def walkA(r,node,lvl):
    for k,child in node.items():
        if 'net' in child: emitA(r,k,child,lvl); r+=1
        else: emitA(r,k,agg(child),lvl); r+=1; r=walkA(r,child,lvl+1)
    return r
TA=tree5(compute('2026')); r=walkA(r,TA,0)
g=dict(eff=0,ca=0,propre=0,siege=0,net=0)
for mq in TA.values():
    a=agg(mq)
    for k in g: g[k]+=a[k]
emitA(r,"GROUPE",g,0)
for j in range(1,8): wa.cell(r,j).font=F(11,True,TEAL); wa.cell(r,j).fill=fill(TEALBG); wa.cell(r,j).border=Border(top=med,bottom=med)
wa.column_dimensions['A'].width=44
for col,w in zip("BCDEFG",[10,13,13,12,13,9]): wa.column_dimensions[col].width=w

# --- Évolution ---
we=wb.create_sheet("Évolution 24-26"); we.sheet_view.showGridLines=False; we.sheet_properties.outlinePr.summaryBelow=False
r=spec(we,1,"RESTITUTION — Évolution de la marge chargée 2024→2026","V_ALLOCATION  (Q_RAPPORT_EVOLUTION, VERSION='ACT')",
    "Marque ▸ Campus ▸ Programme ▸ Année ▸ Modalité  ·  EXERCICE en colonnes",
    [("CA","706 + 7062 + 708"),
     ("EBITDA net","MARGE_COMPLETE (coût complet chargé)"),
     ("Marge nette %","EBITDA net ÷ CA par exercice (membre calculé)"),
     ("Δ (points)","Marge 2026 − Marge 2024")])
hdr(we,r,["Marque ▸ Campus ▸ Programme ▸ Année ▸ Modalité","CA 2026","Marge 2024","Marge 2025","Marge 2026","Δ 26/24 (pt)"]); r+=1
data={y:compute(y) for y in ('2024','2025','2026')}
def key(x): return (x['mq'],x['campus'],x['prog'],x['an'],x['mod'])
idx={y:{key(x):x for x in data[y]} for y in data}
T26=tree5(data['2026'])
def ckeys(node):
    ks=[]
    def rec(n):
        if 'net' in n and 'prog' in n: ks.append(key(n)); return
        for v in n.values(): rec(v)
    rec(node); return ks
def emitE(r,label,node,lvl):
    ks=ckeys(node) if not('net' in node and 'prog' in node) else [key(node)]
    def mg(y):
        ca=sum(idx[y][k]['ca'] for k in ks if k in idx[y]); net=sum(idx[y][k]['net'] for k in ks if k in idx[y])
        return net/ca if ca else None
    m24,m25,m26=mg('2024'),mg('2025'),mg('2026'); ca26=sum(idx['2026'][k]['ca'] for k in ks if k in idx['2026'])
    we.cell(r,1,label).font=F(9 if lvl<=1 else 8.5,lvl<=1); we.cell(r,1).alignment=LI(lvl)
    we.cell(r,2,round(ca26)).number_format=EUR; we.cell(r,2).alignment=RGT
    for j,mv in zip((3,4,5),(m24,m25,m26)):
        if mv is not None:
            c=we.cell(r,j,mv); c.number_format=PCT; c.alignment=RGT; c.font=F(9 if lvl<=1 else 8.5,lvl<=1 and j==5,OCHRED if mv<0.05 else (TEALD if j==5 else SOFT))
    if m24 is not None and m26 is not None:
        c=we.cell(r,6,(m26-m24)*100); c.number_format=DPT; c.alignment=RGT; c.font=F(8.5,False,TEALD if m26>=m24 else OCHRE)
    if lvl==0:
        for j in range(1,7): we.cell(r,j).fill=fill(L0BG); we.cell(r,j).border=Border(top=med,bottom=thin)
    elif lvl==1:
        for j in range(1,7): we.cell(r,j).fill=fill(L1BG)
    else:
        for j in range(1,7): we.cell(r,j).border=Border(bottom=thin)
    if lvl>0:
        we.row_dimensions[r].outline_level=min(lvl,4)
        if lvl>=2: we.row_dimensions[r].hidden=True
def walkE(r,node,lvl):
    for k,child in node.items():
        emitE(r,k,child,lvl); r+=1
        if not('net' in child and 'prog' in child): r=walkE(r,child,lvl+1)
    return r
r=walkE(r,T26,0)
we.column_dimensions['A'].width=44
for col,w in zip("BCDEF",[13,11,11,11,12]): we.column_dimensions[col].width=w

out="/home/user/demo5/eduservices/tagetik/RESTITUTIONS_EDUSERVICES.xlsx"
wb.save(out); print("SAVED",out)
print("  Moteur — CA gagné groupe (effet) =",round(gt[4]))
print("  Alloué — EBITDA net groupe =",round(g['net']))
