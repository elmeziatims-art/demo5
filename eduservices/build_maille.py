#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Maille fine GROUPEE (marque > campus > classe), vivante (reagit aux cles),
theme Allocation (or). Reference 2026. Devient le coeur de l'onglet 3."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import OrderedDict, defaultdict

wb=openpyxl.load_workbook("CAD_SAAD_LIVE.xlsx")
ex=openpyxl.load_workbook("CAD_SAAD_LIVE.xlsx",data_only=True)
al=ex["Allocation"]; H={c.value:c.column_letter for c in al[1] if c.value}
def gv(col,r): return al['%s%d'%(H[col],r)].value

# --- structure ordonnee marque > campus > classes (2026) ---
MORD=["MBWAY","ISCOM","IPAC","PIGIER","TUNON"]
MLAB={"MBWAY":"MBway","ISCOM":"ISCOM","IPAC":"Ipac","PIGIER":"Pigier","TUNON":"Tunon"}
VILLE={"PAR":"Paris","LYO":"Lyon","NAN":"Nantes","BOR":"Bordeaux","LIL":"Lille",
       "TLS":"Toulouse","REN":"Rennes","MTP":"Montpellier"}
tree=OrderedDict((m,OrderedDict()) for m in MORD)
for r in range(2,al.max_row+1):
    if str(gv('EXERCICE',r))!='2026': continue
    m=gv('MARQUE',r); e=gv('ENTITY',r)
    tree[m].setdefault(e,[]).append((gv('PROGRAMME',r),gv('AN_ETUDE',r),gv('MODALITE',r)))

# --- feuille ---
if "3_Maille" in wb.sheetnames: del wb["3_Maille"]
ws=wb.create_sheet("3_Maille")
ws.sheet_view.showGridLines=False
ws.sheet_properties.outlinePr.summaryBelow=False
GOLD="B8860B";GOLD_L="FBF3DE";NAVY="15406E";WHITE="FFFFFF";GREY_L="F7F7F9";PURP="7030A0"
MCOL={"MBWAY":"D4E6FA","ISCOM":"DDF2E3","IPAC":"FBEFD0","PIGIER":"F6E2D4","TUNON":"ECE0F6"}
thin=Side(style="thin",color="E0E0E0");box=Border(thin,thin,thin,thin)
def fill(h):return PatternFill("solid",fgColor=h)
RIG=Alignment(horizontal="right",vertical="center");LEF=Alignment(horizontal="left",vertical="center")
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)

# titre
ws.merge_cells("A1:L1")
ws["A1"]="ALLOCATION · Maille fine du cout complet  —  marque ▸ campus ▸ classe (2026, live)"
ws["A1"].font=Font(size=13,bold=True,color=GOLD); ws["A1"].fill=fill(GOLD_L)
ws["A2"]="Deplie avec les [+] a gauche. Change une cle (onglet Allocation) -> les couts se redistribuent."
ws["A2"].font=Font(size=9,italic=True,color="8A8FA0")
heads=["Marque ▸ Campus ▸ Classe","Effectif","CA","VAC","PERM","ODIR","STRUCT","Frais marque","Holding","Cout complet","Marge complete","Marge %"]
for i,h in enumerate(heads):
    c=ws.cell(3,1+i,h); c.font=Font(size=9,bold=True,color=WHITE); c.fill=fill(GOLD); c.alignment=CEN; c.border=box
ws.row_dimensions[3].height=28

A="Allocation!"
def sif(col,crit):  # SUMIFS live sur Allocation, filtre 2026 + criteres
    s="SUMIFS(%s$%s:$%s"%(A,col,col)
    for cc,val in crit: s+=",%s$%s:$%s,%s"%(A,cc,cc,val)
    return "="+s+',%s$C:$C,"2026")'%A
LC={"eff":"I","ca":"K","VAC":"V","PERM":"W","ODIR":"X","STRUCT":"Y","MARQUE":"AC","HOLD":"Z","MARGE":"AA"}
def costcols(crit):
    return {k:sif(LC[k],crit) for k in ["eff","ca","VAC","PERM","ODIR","STRUCT","MARQUE","HOLD","MARGE"]}
def writerow(r,label,vals,level,fillc,fontc,bold):
    ws.cell(r,1,("   "*level)+label).font=Font(size=9,bold=bold,color=fontc)
    ws.cell(r,1).fill=fill(fillc); ws.cell(r,1).alignment=LEF; ws.cell(r,1).border=box
    order=["eff","ca","VAC","PERM","ODIR","STRUCT","MARQUE","HOLD"]
    for i,k in enumerate(order):
        c=ws.cell(r,2+i,vals[k]); c.number_format="# ##0"; c.font=Font(size=9,bold=bold,color=fontc)
        c.fill=fill(fillc); c.alignment=RIG; c.border=box
    ws.cell(r,10,"=SUM(D%d:I%d)"%(r,r)).number_format="# ##0"  # cout complet
    ws.cell(r,11,vals["MARGE"]).number_format="# ##0"
    ws.cell(r,12,"=IFERROR(K%d/C%d,0)"%(r,r)).number_format="0.0%"
    for col in (10,11,12):
        c=ws.cell(r,col); c.font=Font(size=9,bold=bold,color=fontc if col!=11 else "1E7A46"); c.fill=fill(fillc if col!=11 else "E4F5EA"); c.alignment=RIG; c.border=box
    if level>0: ws.row_dimensions[r].outline_level=level
    if level==2: ws.row_dimensions[r].hidden=True  # classes repliees par defaut

r=4
for m in MORD:
    writerow(r,MLAB[m],costcols([("E",'"%s"'%m)]),0,MCOL[m],NAVY,True); r+=1
    for e,cls in tree[m].items():
        ville=VILLE.get(e.split("_")[-1],e.split("_")[-1])
        writerow(r,ville+"  ("+e+")",costcols([("D",'"%s"'%e)]),1,GREY_L,"333333",True); r+=1
        for (prog,an,mod) in cls:
            lab="%s %s %s"%(prog,an,mod)
            writerow(r,lab,costcols([("D",'"%s"'%e),("F",'"%s"'%prog),("G",'"%s"'%an),("H",'"%s"'%mod)]),2,WHITE,"555555",False); r+=1
# total groupe
ws.cell(r,1,"GROUPE").font=Font(bold=True,color=WHITE); ws.cell(r,1).fill=fill(GOLD); ws.cell(r,1).border=box
for i,k in enumerate(["I","K","V","W","X","Y","AC","Z"]):
    ws.cell(r,2+i,'=SUMIFS(%s$%s:$%s,%s$C:$C,"2026")'%(A,k,k,A)).number_format="# ##0"
    ws.cell(r,2+i).font=Font(bold=True,color=WHITE); ws.cell(r,2+i).fill=fill(GOLD); ws.cell(r,2+i).border=box
ws.cell(r,10,"=SUM(D%d:I%d)"%(r,r)).number_format="# ##0"
ws.cell(r,11,'=SUMIFS(%sAA:AA,%sC:C,"2026")'%(A,A)).number_format="# ##0"
ws.cell(r,12,"=IFERROR(K%d/C%d,0)"%(r,r)).number_format="0.0%"
for col in (10,11,12):
    ws.cell(r,col).font=Font(bold=True,color=WHITE); ws.cell(r,col).fill=fill(GOLD); ws.cell(r,col).border=box

widths={"A":34,"B":9,"C":11,"D":9,"E":9,"F":9,"G":10,"H":11,"I":10,"J":12,"K":13,"L":8}
for c,w in widths.items(): ws.column_dimensions[c].width=w
wb.save("CAD_SAAD_LIVE.xlsx")
print("OK maille fine groupee : %d lignes, classes repliees par defaut."%(r-4))
