#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""DEMO_ACTE1.xlsx — l'histoire + les écrans de l'Acte 1 (cockpit auto-porté).
Feuilles : Histoire · Cockpit · Diagnostic CAC · Structure & Mix."""
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.chart import LineChart,BarChart,Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E4F0EC"; CARD2="F5F7F9"; FAINT="7D8B98"
WHITE="FFFFFF"; BLUE="0000FF"; OCHRE="B3641C"; OCHREBG="F7EAD9"; NAVY="3D4F8F"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
med=Side(style="medium",color=TEAL); thin=Side(style="thin",color="DBE2E9")
CTR=Alignment("center",vertical="center",wrap_text=True); LFT=Alignment("left",vertical="center",wrap_text=True); RGT=Alignment("right",vertical="center")
EUR='#,##0;(#,##0);"-"'; PCT='0.0%'; NUM='#,##0'; EURc='#,##0" €"'; M2='#,##0.00,," M€"'
def Hd(ws,r,labels,fromcol=1):
    for j,h in enumerate(labels,fromcol):
        c=ws.cell(r,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL); c.alignment=CTR if j>fromcol else LFT

wb=openpyxl.Workbook()

# ============================ HISTOIRE ============================
hi=wb.active; hi.title="Histoire"; hi.sheet_view.showGridLines=False
hi["A1"]="EDUSERVICES · BUDGET 2027 — l'Acte 1, du chargement au diagnostic"; hi["A1"].font=F(16,True,INK)
hi["A2"]="Le fil : on POSE (un seul écran, le cockpit), puis on CREUSE (à la demande). Rien n'est affiché en bloc."; hi["A2"].font=F(10,False,TEALD)
hi["A3"]="Vocabulaire : 2024-2025 réalisé · 2026 atterrissage (estimé) · 2027 budget (à construire)."; hi["A3"].font=F(9,False,FAINT,True)
Hd(hi,5,["#","Ce qu'on FAIT","Ce qu'on AFFICHE","Ce qu'on DIT","Drill / suite"])
STO=[
("0","Chargement compta + CRM","(transition, pas d'écran)","« Deux mondes — finance et commercial — entrent sans friction dans le même modèle. »","→ le cockpit"),
("1","On POSE l'état + la tension","COCKPIT d'atterrissage 2026 : bandeau CA/EBITDA/Marge + tuiles KPI (finance & commercial, dont Dépenses) + 1 graphe de tension","« On progresse partout… mais les dépenses d'acquisition montent plus vite que le volume : le CAC se dégrade. C'est le départ du budget 2027. »","clic tuile CAC → Diagnostic"),
("2","On CREUSE le CAC (drill principal)","DIAGNOSTIC CAC : funnel (taux de passage) + CAC par marque (Tunon 564 €)","« Le CAC n'est pas une fatalité : c'est un taux de conversion. On sait où agir — Tunon en priorité. »","clic tuile CAC ; clic marque → funnel campus ; puis → construction 2027"),
("~","2e drill (OPTIONNEL) — le contexte","STRUCTURE & MIX : mix initiale/alternance (OPCO) + CA par marque","« D'où vient le CA et comment il est financé : 2 marques = 70 % du CA, et 3 marques 100 % OPCO — dépendance à surveiller. »","clic tuile CA → Structure & Mix (selon l'audience, hors fil principal)"),
("→","On construit","(Acte 2 — cadrage, moteur, scénarios)","« Maintenant qu'on sait d'où on part et où ça tire, on construit le budget 2027. »","suite de la démo"),
]
r=6
for row in STO:
    for j,val in enumerate(row,1):
        c=hi.cell(r,j,val)
        if j==1: c.font=F(12,True,TEAL); c.alignment=CTR
        elif j==4: c.font=F(9,False,INK,True); c.alignment=LFT
        else: c.font=F(9,False,INK); c.alignment=LFT
        c.border=Border(bottom=thin)
    hi.cell(r,1).fill=fill(TEALBG if row[0] in("1","2") else CARD2)
    hi.row_dimensions[r].height=56; r+=1
hi.cell(r+1,1,"Point clé : après le chargement, on n'affiche QUE le cockpit. Le Diagnostic et le Structure&Mix s'ouvrent APRÈS, quand on creuse.").font=F(9,True,OCHRE,True)
for col,w in zip("ABCDE",[5,26,34,44,26]): hi.column_dimensions[col].width=w

# ============================ COCKPIT (auto-porté) ============================
ck=wb.create_sheet("Cockpit"); ck.sheet_view.showGridLines=False
ck["A1"]="COCKPIT · ATTERRISSAGE 2026"; ck["A1"].font=F(16,True,INK)
ck["A2"]="Point de départ du budget 2027 · réalisé 2024-2025, atterrissage 2026 · commercial + financier, un seul écran"; ck["A2"].font=F(9,False,TEALD)
# colonnes : 2024=B 2025=C 2026=D  YoY=E  sens=F
# bandeau de tête (référence les lignes du cockpit)
def head_stat(col,lab,formula,fmt):
    ck.cell(4,col,lab).font=F(9,True,FAINT); ck.cell(4,col).alignment=LFT
    c=ck.cell(5,col,formula); c.font=F(16,True,TEALD); c.number_format=fmt; c.alignment=LFT
    for rr in (4,5):
        for cc in range(col,col+2): ck.cell(rr,cc).fill=fill(TEALBG)
head_stat(1,"Chiffre d'affaires 2026","=D9",M2); head_stat(3,"EBITDA 2026","=D10",M2); head_stat(5,"Marge EBITDA 2026","=D10/D9",PCT)
hr=7
for j,h in enumerate(["KPI","2024 réalisé","2025 réalisé","2026 atterrissage","YoY 25→26","sens"],1):
    c=ck.cell(hr,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL); c.alignment=CTR if j>1 else LFT
def section(r,txt):
    ck.cell(r,1,txt).font=F(9,True,TEALD)
    for c in range(1,7): ck.cell(r,c).fill=fill(CARD2)
def kpi(r,lab,vals,fmt,pt=False,tension=False,is_input=True):
    lc=OCHRE if tension else INK; yc=OCHRE if tension else TEALD
    ck.cell(r,1,lab).font=F(10,True,lc); ck.cell(r,1).alignment=LFT
    for k,v in enumerate(vals):
        cc=ck.cell(r,2+k,v); cc.number_format=fmt; cc.alignment=RGT
        cc.font=F(10,False,(BLUE if is_input else lc))
    y=ck.cell(r,5)
    if pt: y.value=f"=(D{r}-C{r})*100"; y.number_format='"▲ "0.0" pt"'
    else: y.value=f"=D{r}/C{r}-1"; y.number_format='"▲ "0.0%'
    y.alignment=RGT; y.font=F(10,True,yc)
    ck.cell(r,6,"bon sens" if not tension else "à surveiller").font=F(8,False,yc,True)
section(8,"FINANCE")
kpi(9,"Chiffre d'affaires (€)",[20064725,21268606,22544725],EUR)
kpi(10,"EBITDA (€)",[2648550,2977604,3291530],EUR)
kpi(11,"Marge EBITDA %",["=B10/B9","=C10/C9","=D10/D9"],PCT,pt=True,is_input=False)
section(12,"COMMERCIAL")
kpi(13,"Leads",[15305,16226,17197],NUM)
kpi(14,"Inscrits",[1092,1159,1229],NUM)
kpi(15,"Dépenses acquisition (€)",[358819,394702,434174],EUR,tension=True)
kpi(16,"CAC (€/inscrit)",["=B15/B14","=C15/C14","=D15/D14"],EURc,tension=True,is_input=False)
ck.cell(17,1,"CAC +3,7 % : les dépenses (+10 %) montent plus vite que le volume (+6 %) → chaque inscrit coûte un peu plus. Enjeu 2027 : croître SANS laisser filer le CAC.").font=F(8,True,OCHRE,True)
# combo : inscrits (barres) + CAC (courbe, 2e axe) — la tension en valeurs RÉELLES
tr=19
ck.cell(tr,1,"TENSION — le volume monte, mais le coût unitaire aussi").font=F(10,True,TEALD)
for k,y in enumerate(("2024","2025","2026")): ck.cell(tr+1,2+k,int(y)).font=F(9,True); ck.cell(tr+1,2+k).alignment=CTR
ck.cell(tr+2,1,"Inscrits").font=F(9)
for k,c in enumerate("BCD"): ck.cell(tr+2,2+k,f"={c}14").number_format=NUM
ck.cell(tr+3,1,"CAC (€)").font=F(9)
for k,c in enumerate("BCD"): ck.cell(tr+3,2+k,f"={c}16").number_format=EURc
bar=BarChart(); bar.type="col"; bar.title="Volume (inscrits, barres) vs coût par inscrit (CAC, courbe)"; bar.height=8; bar.width=15
bar.add_data(Reference(ck,min_col=1,min_row=tr+2,max_row=tr+2,max_col=4),titles_from_data=True,from_rows=True)
bar.set_categories(Reference(ck,min_col=2,min_row=tr+1,max_col=4,max_row=tr+1))
bar.series[0].graphicalProperties=GraphicalProperties(solidFill=TEALBG)
bar.y_axis.title="Inscrits"
line=LineChart()
line.add_data(Reference(ck,min_col=1,min_row=tr+3,max_row=tr+3,max_col=4),titles_from_data=True,from_rows=True)
line.series[0].graphicalProperties=GraphicalProperties()
line.series[0].graphicalProperties.line.solidFill=OCHRE
line.series[0].graphicalProperties.line.width=28000
line.y_axis.axId=200; line.y_axis.title="CAC (€)"; line.y_axis.crosses="max"
bar+=line
ck.add_chart(bar,"H7")
ck.cell(tr+5,1,"Les inscrits progressent (barres) mais le CAC grimpe (courbe) : les dépenses montent plus vite que le volume. Clic tuile CAC → Diagnostic.").font=F(8,True,OCHRE,True)
for col,w in zip("ABCDEF",[26,13,13,15,11,12]): ck.column_dimensions[col].width=w

# ============================ DIAGNOSTIC CAC ============================
d=wb.create_sheet("Diagnostic CAC"); d.sheet_view.showGridLines=False
d["A1"]="DIAGNOSTIC CAC  ·  pourquoi le coût d'acquisition se dégrade  ·  2026"; d["A1"].font=F(15,True,INK)
d["A2"]="1er drill depuis la tuile CAC du cockpit. Le funnel explique COMMENT ; le CAC par marque dit OÙ agir."; d["A2"].font=F(9,False,TEALD)
d["A4"]="Funnel de conversion (groupe)"; d["A4"].font=F(11,True,TEALD)
Hd(d,5,["Étape","Volume","Taux de passage"])
FUN=[("Leads",17197,None),("Candidatures",3720,"=B7/B6"),("Admis",2623,"=B8/B7"),("Inscrits",1229,"=B9/B8")]
for i,(lab,vol,tx) in enumerate(FUN):
    r=6+i; d.cell(r,1,lab).font=F(10); d.cell(r,1).alignment=LFT
    d.cell(r,2,vol).font=F(10,False,BLUE); d.cell(r,2).number_format=NUM; d.cell(r,2).alignment=RGT
    if tx: d.cell(r,3,tx).number_format=PCT; d.cell(r,3).font=F(10,True,TEALD); d.cell(r,3).alignment=RGT
d.cell(10,1,"Global lead → inscrit").font=F(9,True,FAINT); d.cell(10,3,"=B9/B6").number_format=PCT; d.cell(10,3).font=F(10,True,OCHRE); d.cell(10,3).alignment=RGT
fch=BarChart(); fch.type="col"; fch.title="Funnel : le volume fond à chaque étape"; fch.legend=None; fch.height=6.5; fch.width=11
fch.add_data(Reference(d,min_col=2,min_row=6,max_row=9)); fch.set_categories(Reference(d,min_col=1,min_row=6,max_row=9)); d.add_chart(fch,"E4")
d["A13"]="CAC par marque — où le coût dérape"; d["A13"].font=F(11,True,TEALD)
Hd(d,14,["Marque","Dépense acq.","Inscrits","CAC (€/inscrit)"])
CAC=[("MBway",166530,510),("ISCOM",108103,338),("IPAC",51871,130),("Pigier",58580,164),("Tunon",49090,87)]
for i,(m,dep,ins) in enumerate(CAC):
    r=15+i; ten=(m=="Tunon"); col=OCHRE if ten else INK
    d.cell(r,1,m).font=F(10,True,col); d.cell(r,1).alignment=LFT
    d.cell(r,2,dep).font=F(10,False,BLUE); d.cell(r,3,ins).font=F(10,False,BLUE); d.cell(r,4,f"=B{r}/C{r}").font=F(10,True,col)
    for cc,fmt in [(2,EUR),(3,NUM),(4,EURc)]: d.cell(r,cc).number_format=fmt; d.cell(r,cc).alignment=RGT
    if ten: d.cell(r,5,"◄ 1,7× la moyenne").font=F(9,True,OCHRE)
cch=BarChart(); cch.type="bar"; cch.title="CAC par marque : Tunon décroche (564 € vs ~320 €)"; cch.legend=None; cch.height=6.5; cch.width=11
cch.add_data(Reference(d,min_col=4,min_row=15,max_row=19)); cch.set_categories(Reference(d,min_col=1,min_row=15,max_row=19)); d.add_chart(cch,"F13")
d.cell(21,1,"Le funnel dit COMMENT (taux) ; le CAC par marque dit OÙ agir → Tunon en priorité. Clic marque → funnel du campus.").font=F(8,True,OCHRE,True)

# --- BRIDGE du CAC (waterfall) : d'où vient la hausse 2025 -> 2026 ---
d.cell(29,1,"Bridge du CAC — d'où vient la hausse 2025 → 2026").font=F(11,True,TEALD)
# inputs (bleu)
for i,(lab,v) in enumerate([("Dépenses 2025",394702),("Dépenses 2026",434174),("Inscrits 2025",1159),("Inscrits 2026",1229)]):
    d.cell(30+i,5,lab).font=F(8,False,FAINT); d.cell(30+i,6,v).font=F(9,False,BLUE); d.cell(30+i,6).number_format=NUM
# F30=dép25 F31=dép26 F32=insc25 F33=insc26
WF=[("CAC 2025",0,"=F30/F32"),
    ("+ Effet dépenses","=F30/F32","=(F31-F30)/F32"),
    ("− Effet volume","=F31/F33","=F31/F32-F31/F33"),
    ("CAC 2026",0,"=F31/F33")]
for i,(cat,base,vis) in enumerate(WF):
    r=30+i
    d.cell(r,1,cat).font=F(9,True); d.cell(r,1).alignment=LFT
    d.cell(r,2,base).number_format=EURc; d.cell(r,2).font=F(8,False,WHITE)
    d.cell(r,3,vis).number_format=EURc; d.cell(r,3).font=F(9,True); d.cell(r,3).alignment=RGT
wf=BarChart(); wf.type="col"; wf.grouping="stacked"; wf.overlap=100; wf.legend=None
wf.title="Bridge du CAC : +34 € (dépenses) − 21 € (volume) = +13 € net"; wf.height=7.5; wf.width=13
wf.add_data(Reference(d,min_col=2,max_col=3,min_row=30,max_row=33))
wf.set_categories(Reference(d,min_col=1,min_row=30,max_row=33))
wf.series[0].graphicalProperties=GraphicalProperties(solidFill="FFFFFF")   # base invisible
for idx,col in enumerate(["9AA7B0",OCHRE,TEAL,NAVY]):
    dp=DataPoint(idx=idx); dp.graphicalProperties=GraphicalProperties(solidFill=col); wf.series[1].data_points.append(dp)
wf.gapWidth=45
wf.y_axis.scaling.min=320; wf.y_axis.scaling.max=385   # axe tronqué -> les pas +34/-21 deviennent lisibles
wf.y_axis.title="CAC (€)"
wf.series[1].dLbls=DataLabelList(); wf.series[1].dLbls.showVal=True   # valeurs sur les barres
d.add_chart(wf,"H29")
d.cell(35,1,"Lecture : les dépenses poussent le CAC de +34 € ; le volume supplémentaire l'amortit de −21 € ; net +13 €. La tension = l'effet dépenses > l'effet volume.").font=F(8,True,OCHRE,True)
for col,w in zip("ABCDEF",[22,13,11,15,15,11]): d.column_dimensions[col].width=w

# ============================ STRUCTURE & MIX ============================
s=wb.create_sheet("Structure & Mix"); s.sheet_view.showGridLines=False
s["A1"]="STRUCTURE & MIX DE FINANCEMENT  ·  2026"; s["A1"].font=F(15,True,INK)
s["A2"]="Ce que le CRM ne montre pas : la dépendance à l'alternance (OPCO). Vraie question DAF — risque réglementaire."; s["A2"].font=F(9,False,TEALD)
s["A4"]="Mix scolarité : initiale vs alternance (OPCO), par marque"; s["A4"].font=F(11,True,TEALD)
Hd(s,5,["Marque","Initiale (706)","Alternance (7062)","% alternance"])
MIX=[("MBway",1995975,7525580),("ISCOM",1310700,4951430),("IPAC",0,2571100),("Pigier",0,2137460),("Tunon",0,1941870)]
for i,(m,ini,alt) in enumerate(MIX):
    r=6+i; full=(ini==0); col=OCHRE if full else INK
    s.cell(r,1,m).font=F(10,True,col); s.cell(r,1).alignment=LFT
    s.cell(r,2,ini).font=F(10,False,BLUE); s.cell(r,3,alt).font=F(10,False,BLUE); s.cell(r,4,f"=C{r}/(B{r}+C{r})").font=F(10,True,col)
    for cc,fmt in [(2,EUR),(3,EUR),(4,PCT)]: s.cell(r,cc).number_format=fmt; s.cell(r,cc).alignment=RGT
    if full: s.cell(r,5,"100 % OPCO").font=F(9,True,OCHRE)
mch=BarChart(); mch.type="col"; mch.grouping="stacked"; mch.overlap=100
mch.title="Mix initiale / alternance : 3 marques 100% OPCO"; mch.height=6.5; mch.width=12
mch.add_data(Reference(s,min_col=2,max_col=3,min_row=5,max_row=11),titles_from_data=True)
mch.set_categories(Reference(s,min_col=1,min_row=6,max_row=11)); s.add_chart(mch,"F4")
s["A13"]="CA total par marque — concentration"; s["A13"].font=F(11,True,TEALD)
Hd(s,14,["Marque","CA total"])
CA=[("MBway",9567455),("ISCOM",6292550),("IPAC",2582800),("Pigier",2152220),("Tunon",1949700)]
for i,(m,ca) in enumerate(CA):
    r=15+i; s.cell(r,1,m).font=F(10,True); s.cell(r,1).alignment=LFT
    s.cell(r,2,ca).font=F(10,False,BLUE); s.cell(r,2).number_format=EUR; s.cell(r,2).alignment=RGT
cca=BarChart(); cca.type="bar"; cca.title="CA par marque : MBway + ISCOM ≈ 70% du CA"; cca.legend=None; cca.height=6; cca.width=12
cca.add_data(Reference(s,min_col=2,min_row=15,max_row=19)); cca.set_categories(Reference(s,min_col=1,min_row=15,max_row=19)); s.add_chart(cca,"F13")
s.cell(21,1,"INSIGHT DAF : groupe massivement financé par l'alternance (OPCO). MBway/ISCOM ~79%, les 3 autres 100%. Dépendance à surveiller.").font=F(8,True,OCHRE,True)
for col,w in zip("ABCDE",[16,15,17,13,13]): s.column_dimensions[col].width=w

out="/home/user/demo5/eduservices/tagetik/DEMO_ACTE1.xlsx"
wb.save(out); print("SAVED",out,"| feuilles:",wb.sheetnames)
