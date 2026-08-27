#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""DEMO_ACTE2.xlsx — la construction du budget 2027.
Feuilles (1re salve) : Histoire A2 · Cadrage & Leviers · Cap arbitrage."""
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E4F0EC"; CARD2="F5F7F9"; FAINT="7D8B98"
WHITE="FFFFFF"; BLUE="0000FF"; OCHRE="B3641C"; OCHREBG="F7EAD9"; NAVY="3D4F8F"; GREEN="1E7A55"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="DBE2E9"); med=Side(style="medium",color=TEAL)
CTR=Alignment("center",vertical="center",wrap_text=True); LFT=Alignment("left",vertical="center",wrap_text=True); RGT=Alignment("right",vertical="center")
EUR='#,##0;(#,##0);"-"'; PCT='+0.0%;-0.0%;"-"'; PCT2='0.0%'; NUM='#,##0'; CO='0.00'; M2='#,##0.00,," M€"'
def Hd(ws,r,labels,fromcol=1):
    for j,h in enumerate(labels,fromcol):
        c=ws.cell(r,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL); c.alignment=CTR if j>fromcol else LFT

wb=openpyxl.Workbook()

# ============================ HISTOIRE A2 ============================
hi=wb.active; hi.title="Histoire A2"; hi.sheet_view.showGridLines=False
hi["A1"]="ACTE 2 — CONSTRUIRE LE BUDGET 2027"; hi["A1"].font=F(16,True,INK)
hi["A2"]="Après le diagnostic (le CAC se dégrade), on construit 2027 pour y répondre. Chaîne : hypothèses → moteur → scénarios → P&L."; hi["A2"].font=F(10,False,TEALD)
hi["A3"]="Le « waouh » de l'acte : un MOTEUR vivant — on bouge un levier, tout recalcule (volumes, CA, marge)."; hi["A3"].font=F(9,False,FAINT,True)
Hd(hi,5,["#","Étape","Ce qu'on montre","Ce qu'on dit","Objet Tagetik"])
STO=[
("1","Cadrage & leviers","La direction pose les hypothèses par scénario (prix, budgets, taux, coûts) — 3 versions V01/V02/V03","« On fixe le cap : voici les hypothèses des 3 scénarios. »","V_CADRAGE_LEVIERS"),
("2","Arbitrage des caps","Budget d'acquisition par campus : 3 logiques proposées (Efficient / Momentum / Potentiel) → Cap retenu (DAF)","« Trois façons de voir chaque campus. On propose, la DAF tranche. Le budget acq en découle. »","V_CAP_ARBITRAGE"),
("3","Le moteur","La chaîne : dépense → leads → funnel → nouveaux → effectifs → CA. Tout est relié.","« Pas de saisie en dur : je bouge une dépense, l'effectif suit. Un vrai moteur. »","V_MOTEUR"),
("4","Levier prix","Ajuster le prix pour tomber pile sur la cible du cadrage","« Il me manque un pouième : +0,29 % de prix, et je tombe sur le cadrage. »","Moteur (cellule prix)"),
("5","3 scénarios","Cadrage / Optimiste / Prudent côte à côte : CA V01 24,1 M€ · V02 26,3 M€ · V03 22,7 M€","« On ne parie pas sur un seul futur. Trois scénarios, même moteur. »","Q_SCENARIOS"),
("6","Bridge de CA","Le pont 2026 → 2027 : effet prix / volume / mix (le vrai bridge canonique)","« La croissance n'est pas un bloc : voilà la part du prix, du volume, du mix. »","V_BRIDGE_CA"),
("7","P&L ② (marge directe)","Le budget construit, AVANT allocation : marge directe par marque","« Voici le budget tel que construit. Belles marges… mais il manque le coût groupe (→ Acte 3). »","V_PNL / FST"),
]
r=6
for row in STO:
    for j,val in enumerate(row,1):
        c=hi.cell(r,j,val)
        if j==1: c.font=F(12,True,TEAL); c.alignment=CTR
        elif j==2: c.font=F(10,True,INK); c.alignment=LFT
        elif j==4: c.font=F(9,False,INK,True); c.alignment=LFT
        else: c.font=F(9,False,INK); c.alignment=LFT
        c.border=Border(bottom=thin)
    hi.cell(r,1).fill=fill(TEALBG if row[0]=="3" else CARD2)
    hi.row_dimensions[r].height=52; r+=1
hi.cell(r+1,1,"Le cœur = étape 3 (le moteur). Étapes 1-2 = les entrées ; 5-7 = les sorties. Le levier prix (4) est le moment 'live'.").font=F(9,True,OCHRE,True)
for col,w in zip("ABCDE",[4,22,40,40,20]): hi.column_dimensions[col].width=w

# ============================ CADRAGE & LEVIERS ============================
cl=wb.create_sheet("Cadrage & Leviers"); cl.sheet_view.showGridLines=False
cl["A1"]="CADRAGE & LEVIERS  ·  le panneau de commande des 3 scénarios"; cl["A1"].font=F(15,True,INK)
cl["A2"]="La direction fixe les hypothèses par version. Le moteur en déduit volumes, CA et marge. Source : V_CADRAGE_LEVIERS."; cl["A2"].font=F(9,False,TEALD)
Hd(cl,4,["Levier","V01 · Cadrage","V02 · Optimiste","V03 · Prudent","Unité"])
# (label, v01, v02, v03, unit) unit: 'pct' or 'pt'
LEV=[("Prix",0.02,0.035,0.02,"pct"),
     ("Budget acquisition",0.08,0.15,-0.05,"pct"),
     ("Budget marque",0.10,0.20,-0.05,"pct"),
     ("Conv. lead → cand.",0.01,0.03,-0.01,"pt"),
     ("Conv. admis → inscrit",0.01,0.025,-0.01,"pt"),
     ("Taux de passage",0.005,0.015,-0.01,"pt"),
     ("Salaires",0.025,0.02,0.03,"pct"),
     ("Coût de structure",0.0,-0.03,0.04,"pct"),
     ("Inflation externe",0.02,0.015,0.03,"pct")]
r=5
for lab,v1,v2,v3,unit in LEV:
    cl.cell(r,1,lab).font=F(10); cl.cell(r,1).alignment=LFT
    for k,v in enumerate([v1,v2,v3]):
        cc=cl.cell(r,2+k, v if unit=="pct" else v*100)
        cc.number_format=PCT if unit=="pct" else '+0.0" pt";-0.0" pt"'
        cc.alignment=RGT; cc.font=F(10,False,BLUE)
    cl.cell(r,5,"%" if unit=="pct" else "points").font=F(8,False,FAINT)
    if r%2==0:
        for j in range(1,6): cl.cell(r,j).fill=fill(CARD2)
    r+=1
# résultat CA moteur
r+=1
cl.cell(r,1,"→ CA 2027 (sortie moteur)").font=F(11,True,TEALD)
for k,v in enumerate([24120315,26307244,22701560]):
    cc=cl.cell(r,2+k,v); cc.number_format=M2; cc.font=F(11,True,TEALD); cc.alignment=RGT; cc.fill=fill(TEALBG)
cl.cell(r,1).fill=fill(TEALBG); cl.cell(r,5).fill=fill(TEALBG)
r+=1
cl.cell(r,1,"vs atterrissage 2026 (22,54 M€)").font=F(9,False,FAINT,True)
for k,v in enumerate([24120315,26307244,22701560]):
    cc=cl.cell(r,2+k,f"={chr(66+k)}{r-1}/22544725-1"); cc.number_format=PCT; cc.font=F(9,True,GREEN if v>22544725 else OCHRE); cc.alignment=RGT
r+=2
cl.cell(r,1,"Lecture : Optimiste pousse budgets & conversions à fond (+16,7 %) ; Prudent réduit l'acquisition (+0,7 %) ; Cadrage = la cible tenue (+7 %).").font=F(8,True,OCHRE,True)
for col,w in zip("ABCDE",[24,15,15,15,9]): cl.column_dimensions[col].width=w

# ============================ CAP ARBITRAGE ============================
ca=wb.create_sheet("Cap arbitrage"); ca.sheet_view.showGridLines=False
ca["A1"]="ARBITRAGE DES CAPS D'ACQUISITION  ·  où mettre le budget, par campus"; ca["A1"].font=F(15,True,INK)
ca["A2"]="3 logiques proposent un coefficient de budget ; la DAF retient (Cap retenu). Budget = réf × retenu. Source : V_CAP_ARBITRAGE."; ca["A2"].font=F(9,False,TEALD)
Hd(ca,4,["Campus","CAC margi. (€)","Croiss. leads","Intensité mkt","Cap Efficient","Cap Momentum","Cap Potentiel","Cap RETENU","Budget réf (€)"])
CAP=[["IPAC_MTP",1034,0.145,0.019,1.163,1.386,0.969,1.0,15178],["IPAC_NAN",1216,0.138,0.021,0.988,1.32,0.881,1.0,21760],
["IPAC_REN",1041,0.145,0.019,1.155,1.386,0.985,1.0,14933],["ISCOM_LIL",979,0.082,0.014,1.228,0.78,1.31,1.0,25220],
["ISCOM_PAR",1540,0.078,0.021,0.78,0.744,0.9,1.0,60800],["ISCOM_TLS",919,0.082,0.014,1.308,0.786,1.387,1.0,22083],
["MBWAY_BOR",890,0.126,0.014,1.351,1.207,1.339,1.0,24923],["MBWAY_LYO",1102,0.119,0.017,1.091,1.138,1.124,1.0,41783],
["MBWAY_NAN",964,0.124,0.015,1.248,1.185,1.288,1.0,31533],["MBWAY_PAR",1475,0.111,0.022,0.815,1.06,0.865,1.0,68291],
["PIGIER_BOR",1368,0.088,0.024,0.879,0.838,0.772,1.0,21968],["PIGIER_LYO",1705,0.085,0.029,0.705,0.813,0.643,1.0,36612],
["TUNON_LYO",1625,0.073,0.022,0.74,0.694,0.862,1.0,18850],["TUNON_PAR",2193,0.07,0.028,0.548,0.665,0.676,1.0,30240]]
r=5
for row in CAP:
    ent=row[0]; ca.cell(r,1,ent.replace('_',' ')).font=F(9); ca.cell(r,1).alignment=LFT
    ca.cell(r,2,row[1]).number_format=EUR
    ca.cell(r,3,row[2]).number_format=PCT2; ca.cell(r,4,row[3]).number_format=PCT2
    for k,cv in enumerate(row[4:7]):  # 3 caps
        cc=ca.cell(r,5+k,cv); cc.number_format=CO
        cc.font=F(9,False, GREEN if cv>=1 else OCHRE)   # >1 = investir, <1 = réduire
    rc=ca.cell(r,8,row[7]); rc.number_format=CO; rc.font=F(10,True,NAVY); rc.fill=fill(TEALBG)
    ca.cell(r,9,row[8]).number_format=EUR
    for j in range(2,10):
        if j not in (5,6,7,8): ca.cell(r,j).font=F(9)
        ca.cell(r,j).alignment=RGT
    r+=1
# total
ca.cell(r,1,"GROUPE").font=F(10,True,WHITE); ca.cell(r,1).fill=fill(TEALD)
ca.cell(r,9,f"=SUM(I5:I{r-1})").number_format=EUR; ca.cell(r,9).font=F(10,True,WHITE); ca.cell(r,9).fill=fill(TEALD); ca.cell(r,9).alignment=RGT; ca.cell(r,9).border=Border(top=med)
for j in (2,3,4,5,6,7,8): ca.cell(r,j).fill=fill(TEALD)
r+=2
ca.cell(r,1,"Vert = coefficient > 1 (investir) · Ochre = < 1 (réduire). Ex : Tunon Paris — les 3 logiques disent RÉDUIRE (CAC 2 193 €, faible croissance).").font=F(8,True,OCHRE,True)
ca.cell(r+1,1,"MBway Bordeaux — les 3 disent INVESTIR (CAC 890 €, +12,6 % leads). En V1 la DAF a retenu 1,00 partout (budget = réf) ; les caps montrent où réallouer.").font=F(8,False,FAINT,True)
for col,w in zip("ABCDEFGHI",[13,13,12,12,12,12,12,11,13]): ca.column_dimensions[col].width=w

out="/home/user/demo5/eduservices/tagetik/DEMO_ACTE2.xlsx"
wb.save(out); print("SAVED",out,"| feuilles:",wb.sheetnames)
