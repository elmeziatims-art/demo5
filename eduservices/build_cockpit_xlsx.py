#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""COCKPIT_EDUSERVICES.xlsx — le dashboard exécutif en Excel, sur le VRAI périmètre.
Socle : 14 campus × 5 marques, construit par socle_reel.py à partir des inducteurs du
référentiel (base_entry, indices ville, rétention, REV par cycle × modalité, croissance
organique par marque) et normalisé sur les ancres groupe.
Onglet « Données » : la seule saisie. Onglet « Cockpit » : 100 % formules.
Aucune cellule fusionnée dans la grille de données."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from openpyxl.utils import get_column_letter, column_index_from_string
from socle_reel import construire, ANCRES

BLUE="2A78D6"; BLUE2="7AABE6"; BLUE3="C3D9F4"; BLUESOFT="DCE9F9"
ORANGE="EB6834"; ORANGESOFT="FBE3D8"; GOOD="0CA30C"; WARN="FAB219"; CRIT="D03B3B"
INK="131922"; INK2="4D5866"; INK3="7C8798"
CANVAS="EEF1F6"; PANEL="FFFFFF"; PANEL2="F7F9FC"; LINE="DFE4EC"; LINE2="EAEEF4"; WHITE="FFFFFF"
UI="Segoe UI"; MONO="Consolas"
def F(sz=10,b=False,c=INK,f=UI): return Font(name=f,size=sz,bold=b,color=c)
def fill(c): return PatternFill("solid",fgColor=c)
def sd(c=LINE,st="thin"): return Side(style=st,color=c)
L=Alignment("left",vertical="center"); R=Alignment("right",vertical="center")
Cn=Alignment("center",vertical="center"); TOP=Alignment("left",vertical="top",wrap_text=True)
def ind(n): return Alignment("left",vertical="center",indent=n)
EUR='#,##0'; PCT='0.0%'; PCT0='0%'; DPCT='"▲ "0.0%;"▼ "0.0%'
DPT='"↗ +"0.0" pt";"↘ −"0.0" pt";"→ "0.0" pt"'; MEUR='#,##0.00,," M€"'
def box(ws,r1,c1,r2,c2,bg=PANEL,bd=LINE):
    for r in range(r1,r2+1):
        for c in range(c1,c2+1):
            x=ws.cell(r,c); x.fill=fill(bg)
            x.border=Border(top=sd(bd) if r==r1 else None,bottom=sd(bd) if r==r2 else None,
                            left=sd(bd) if c==c1 else None,right=sd(bd) if c==c2 else None)

# ================= LE SOCLE =================
ORDRE=["MBway","ISCOM","Ipac Bachelor Factory","Pigier","Tunon"]
SHORT={"MBway":"MBway","ISCOM":"ISCOM","Ipac Bachelor Factory":"Ipac","Pigier":"Pigier","Tunon":"Tunon"}
C=sorted(construire(),key=lambda c:(ORDRE.index(c["marque"]),c["ent"]))
SOCLE=[(c["ent"],SHORT[c["marque"]],round(c["ca"][2024]),round(c["ca"][2025]),round(c["ca"][2026]),
        round(c["eb"][2024]),round(c["eb"][2025]),round(c["eb"][2026]),
        round(c["inscrits"]),round(c["eff"]),c["places"],round(c["mix_alt"],4)) for c in C]
MARQUES=[SHORT[m] for m in ORDRE]
PLACES=sum(c["places"] for c in C)

# ancres de lignes de l'onglet Données
S1=15; S2=S1+len(SOCLE)-1; STOT=S2+1                 # socle 15..28, total 29
A0=STOT+2; A1=A0+2                                    # ③ acquisition : entête A0, dépenses A0+1, inscrits A1
B0=A1+2; B1=B0+1; BR=B1+1                             # ④ bridge : titre 37, entête 38, lignes 39-43
M0=BR+7; M1=M0+1; MR=M1+1                             # ⑤ marques : titre 47, entête 48, lignes 49-53
T0=MR+len(MARQUES)+1; W1=T0+1; WR=W1+1                # ⑥ waterfall : titre 55, entête 56, lignes 57-61
X1=WR+7; XR=X1+1                                      # tension : entête 68, lignes 69-71
D=lambda col,a=S1,b=None: "Données!$%s$%d:$%s$%d"%(col,a,col,b or S2)

wb=openpyxl.Workbook()
# ============================ ONGLET DONNÉES ============================
d=wb.active; d.title="Données"; d.sheet_view.showGridLines=False
for col,w in zip("ABCDEFGHIJKL",[18,12,14,14,14,14,14,14,11,11,10,11]): d.column_dimensions[col].width=w
d.cell(1,1,"SOCLE DE DONNÉES — cockpit EDUSERVICES  ·  14 campus × 5 marques").font=F(12,True,WHITE)
for c in range(1,13): d.cell(1,c).fill=fill(BLUE)
d.row_dimensions[1].height=24
def h(row,txt): d.cell(row,1,txt).font=F(9.5,True,BLUE)
def hdr(row,labels):
    for j,t in enumerate(labels,1):
        c=d.cell(row,j,t); c.font=F(9,True,WHITE); c.fill=fill(INK2); c.alignment=Cn
        c.border=Border(top=sd(INK2),bottom=sd(INK2),left=sd(INK2),right=sd(INK2))

h(3,"① PÉRIMÈTRE (les filtres du cockpit)"); hdr(4,["Filtre","Valeur"])
for i,(k,v) in enumerate([("Scénario","REEL"),("Version","V_FINAL"),("Exercice",2026),("Période","12 — Cumul"),
                          ("Marque","Toutes"),("Campus","Tous"),("Modalité","Toutes")]):
    d.cell(5+i,1,k).font=F(9.5); d.cell(5+i,2,v).font=F(9.5,True)
    d.cell(5+i,1).alignment=L; d.cell(5+i,2).alignment=L

h(13,"② SOCLE CAMPUS — la seule saisie  (source : V_COCKPIT / V_ALLOCATION, grain campus × exercice)")
hdr(14,["Entity","Marque","CA 2024","CA 2025","CA 2026","EBITDA 2024","EBITDA 2025","EBITDA 2026",
        "Inscrits 26","Effectifs 26","Places 26","Mix alt. 26"])
for i,row in enumerate(SOCLE):
    r=S1+i
    for j,v in enumerate(row,1):
        c=d.cell(r,j,v); c.font=F(9.5); c.border=Border(bottom=sd(LINE2))
        c.alignment=L if j<=2 else R
        c.number_format=EUR if 3<=j<=8 else ('#,##0' if j in (9,10,11) else (PCT0 if j==12 else 'General'))
d.cell(STOT,1,"TOTAL GROUPE").font=F(9.5,True)
for j in range(3,12):
    c=d.cell(STOT,j,"=SUM(%s%d:%s%d)"%(get_column_letter(j),S1,get_column_letter(j),S2))
    c.font=F(9.5,True); c.number_format=EUR if j<9 else '#,##0'; c.alignment=R; c.border=Border(top=sd(INK3))

h(A0-1,"③ ACQUISITION (groupe)"); hdr(A0,["Mesure","2024","2025","2026"])
for i,(lab,a,b,c_,fm) in enumerate([("Dépenses d'acquisition (€)",358170,394702,434174,EUR),
                                    ("Inscrits (nouveaux)",1092,1159,1229,'#,##0')]):
    r=A0+1+i
    d.cell(r,1,lab).font=F(9.5); d.cell(r,1).alignment=L
    for j,v in enumerate((a,b,c_),2):
        x=d.cell(r,j,v); x.number_format=fm; x.font=F(9.5); x.alignment=R; x.border=Border(bottom=sd(LINE))

h(B0-1,"④ BRIDGE EBITDA 2025 → 2026  (l'effet coûts est le résidu : le pont boucle toujours)")
hdr(B0,["Étape","Montant (€)"])
d.cell(BR,1,"EBITDA 2025").font=F(9.5);           d.cell(BR,2,"=SUM(%s)"%D("G").replace("Données!",""))
d.cell(BR+1,1,"Effet activité").font=F(9.5);      d.cell(BR+1,2,211706)
d.cell(BR+2,1,"Effet prix / mix").font=F(9.5);    d.cell(BR+2,2,289000)
d.cell(BR+3,1,"Effet coûts (résidu)").font=F(9.5);d.cell(BR+3,2,"=B%d-B%d-B%d-B%d"%(BR+4,BR,BR+1,BR+2))
d.cell(BR+4,1,"EBITDA 2026").font=F(9.5,True);    d.cell(BR+4,2,"=SUM(%s)"%D("H").replace("Données!",""))
for r in range(BR,BR+5):
    d.cell(r,1).alignment=L; d.cell(r,2).number_format=EUR; d.cell(r,2).alignment=R
    d.cell(r,2).font=F(9.5,r==BR+4); d.cell(r,2).border=Border(bottom=sd(LINE))

h(M0-1,"⑤ AGRÉGATS PAR MARQUE — 100 % calculés depuis ②")
hdr(M0,["Marque","CA 2024","CA 2025","CA 2026","EBITDA 2024","EBITDA 2025","EBITDA 2026",
        "Marge 2024","Marge 2025","Marge 2026"])
for i,m in enumerate(MARQUES):
    r=MR+i
    d.cell(r,1,m).font=F(9.5); d.cell(r,1).alignment=L
    for j,src in zip(range(2,8),"CDEFGH"):
        c=d.cell(r,j,'=SUMIF(%s,$A%d,%s)'%(D("B"),r,D(src))); c.number_format=EUR; c.font=F(9.5); c.alignment=R
    for j,(num,den) in zip(range(8,11),[(5,2),(6,3),(7,4)]):
        c=d.cell(r,j,"=%s%d/%s%d"%(get_column_letter(num),r,get_column_letter(den),r))
        c.number_format=PCT; c.font=F(9.5,True); c.alignment=R
    for j in range(1,11): d.cell(r,j).border=Border(bottom=sd(LINE2))

h(T0,"⑥ ZONE TECHNIQUE — séries des graphes (ne pas saisir)")
hdr(W1,["Étape (waterfall)","Socle","Ancre","Hausse","Baisse"])
for i,(cat,so,an,ha,ba) in enumerate([
    ("EBITDA 2025","0","=B%d"%BR,"0","0"),
    ("Activité","=B%d"%BR,"0","=B%d"%(BR+1),"0"),
    ("Prix / mix","=B%d+B%d"%(BR,BR+1),"0","=B%d"%(BR+2),"0"),
    ("Coûts","=B%d"%(BR+4),"0","0","=B%d+B%d+B%d-B%d"%(BR,BR+1,BR+2,BR+4)),
    ("EBITDA 2026","0","=B%d"%(BR+4),"0","0")]):
    r=WR+i
    d.cell(r,1,cat).font=F(9,c=INK3); d.cell(r,1).alignment=L
    for j,v in enumerate((so,an,ha,ba),2):
        c=d.cell(r,j,v); c.number_format=EUR; c.font=F(9,c=INK3); c.alignment=R
hdr(X1,["Exercice","Dépenses (base 100)","Inscrits (base 100)"])
for i,(ex,col) in enumerate([(2024,"B"),(2025,"C"),(2026,"D")]):
    r=XR+i
    d.cell(r,1,ex).font=F(9,c=INK3); d.cell(r,1).alignment=Cn; d.cell(r,1).number_format='0'
    d.cell(r,2,"=%s%d/$B$%d*100"%(col,A0+1,A0+1)).number_format='0.0'
    d.cell(r,3,"=%s%d/$B$%d*100"%(col,A1,A1)).number_format='0.0'
    for j in (2,3): d.cell(r,j).font=F(9,c=INK3); d.cell(r,j).alignment=R
d.cell(XR+5,1,"Règle de lecture : la variation d'une MARGE se lit en POINTS (différence), jamais en % relatif.").font=F(9,True,ORANGE)

# ============================ ONGLET COCKPIT ============================
w=wb.create_sheet("Cockpit"); w.sheet_view.showGridLines=False
w.column_dimensions["A"].width=2; w.column_dimensions["B"].width=22
for col in "CDEFGHIJKLM": w.column_dimensions[col].width=12
w.column_dimensions["N"].width=2
G0=35                                                   # première ligne de la grille (GROUPE)
ROWS=[("GROUPE",0)]
for m in MARQUES:
    ROWS.append((m,1))
    ROWS+= [(s[0],2) for s in SOCLE if s[1]==m]
GEND=G0+len(ROWS)-1                                     # 35..54
CAL=GEND+2                                              # encadré
for r,hh in {1:6,2:30,3:16,4:6,5:20,6:6,7:15,8:28,9:16,10:8,11:20,32:8,33:20,34:18,GEND+1:8,CAL+4:6}.items():
    w.row_dimensions[r].height=hh
for r in range(12,32): w.row_dimensions[r].height=16
for r in range(G0,GEND+1): w.row_dimensions[r].height=17
for r in range(CAL,CAL+3): w.row_dimensions[r].height=15
for r in range(1,CAL+8):
    for c in range(1,16): w.cell(r,c).fill=fill(CANVAS)

box(w,2,2,3,13)
w.cell(2,2,"  Cockpit de performance — EDUSERVICES GROUP").font=F(14,True,INK); w.cell(2,2).alignment=L
w.cell(3,2,"  Constat 2024 → 2026 · 14 campus · 5 marques · socle CRM & comptabilité rapprochés").font=F(9.5,c=INK3)
w.cell(3,2).alignment=L
w.cell(2,13,"RÉEL · CLÔTURE 2026  ").font=F(9,True,BLUE); w.cell(2,13).alignment=R
for i,(lab,ref) in enumerate([("Scénario","=Données!B5"),("Version","=Données!B6"),("Exercice","=Données!B7"),
                              ("Période","=Données!B8"),("Marque","=Données!B9"),("Campus","=Données!B10")]):
    c1=2+i*2
    box(w,5,c1,5,c1+1,bg=(BLUESOFT if lab=="Exercice" else PANEL),bd=(BLUE if lab=="Exercice" else LINE))
    a=w.cell(5,c1,"  "+lab); a.font=F(9,c=INK3); a.alignment=L
    b=w.cell(5,c1+1,ref); b.font=F(9.5,True,BLUE if lab=="Exercice" else INK); b.alignment=L

KPI=[("Chiffre d'affaires","=C%d"%G0,MEUR,"=D%d"%G0,DPCT,False),
     ("EBITDA","=E%d"%G0,MEUR,"=F%d"%G0,DPCT,False),
     ("Marge EBITDA","=H%d"%G0,PCT,"=I%d"%G0,DPT,False),
     ("Inscrits (nouveaux)","=J%d"%G0,'#,##0',"=J%d/Données!C%d-1"%(G0,A1),DPCT,False),
     ("Coût d'acquisition","=Données!D%d/Données!D%d"%(A0+1,A1),'#,##0" €"',
      "=(Données!D%d/Données!D%d)/(Données!C%d/Données!C%d)-1"%(A0+1,A1,A0+1,A1),DPCT,True),
     ("Remplissage moyen","=K%d"%G0,PCT0,'=SUM(%s)-SUM(%s)'%(D("K"),D("J")),'#,##0" places libres"',False)]
for i,(lab,val,fm,dv,dfm,alert) in enumerate(KPI):
    c1=2+i*2
    box(w,7,c1,9,c1+1,bd=(WARN if alert else LINE))
    a=w.cell(7,c1,"  "+lab.upper()); a.font=F(8.5,True,INK3); a.alignment=L
    v=w.cell(8,c1,val); v.font=F(20,True,INK,MONO); v.number_format=fm; v.alignment=ind(1)
    x=w.cell(9,c1,dv); x.font=F(9.5,True,INK3 if i==5 else INK); x.number_format=dfm; x.alignment=ind(1)
    if i<5:
        for op,col in (("greaterThan",GOOD),("lessThan",CRIT)):
            w.conditional_formatting.add("%s9"%get_column_letter(c1),
                CellIsRule(operator=op,formula=["0"],font=Font(name=UI,size=9.5,bold=True,color=col)))
w.conditional_formatting.add("J9",CellIsRule(operator="greaterThan",formula=["0"],
    font=Font(name=UI,size=9.5,bold=True,color=CRIT)))

for c1,c2,title,hint in [(2,5,"Bridge EBITDA 2025 → 2026","M€ · axe tronqué à 3,2"),
                         (6,9,"Marge EBITDA par marque","2024 · 2025 · 2026"),
                         (10,13,"Tension acquisition","base 100 = 2024")]:
    box(w,11,c1,31,c2)
    t=w.cell(11,c1,"  "+title); t.font=F(10.5,True,INK); t.alignment=L
    hn=w.cell(11,c2,hint+"  "); hn.font=F(8.5,c=INK3); hn.alignment=R

wf=BarChart(); wf.type="col"; wf.grouping="stacked"; wf.overlap=100; wf.gapWidth=45
wf.height=7.9; wf.width=10.2; wf.legend=None; wf.y_axis.numFmt='#,##0,," M€"'
wf.visible_cells_only=False
wf.y_axis.scaling.min=3200000; wf.y_axis.scaling.max=4000000; wf.y_axis.majorUnit=200000
wf.x_axis.delete=False; wf.y_axis.delete=False
for col in (2,3,4,5): wf.add_data(Reference(d,min_col=col,max_col=col,min_row=WR,max_row=WR+4),titles_from_data=False)
s_so,s_an,s_ha,s_ba=wf.series
s_so.graphicalProperties.noFill=True; s_so.graphicalProperties.line.noFill=True
for s,cc in ((s_an,BLUE),(s_ha,GOOD),(s_ba,CRIT)):
    s.graphicalProperties.solidFill=cc; s.graphicalProperties.line.solidFill=cc
for s,fm in ((s_an,'#,##0,,.000" M€"'),(s_ha,'"+ "#,##0" €"'),(s_ba,'"− "#,##0" €"')):
    s.dLbls=DataLabelList(); s.dLbls.showVal=True; s.dLbls.numFmt=fm; s.dLbls.dLblPos="ctr"
    s.dLbls.showSerName=False; s.dLbls.showCatName=False; s.dLbls.showLegendKey=False
wf.set_categories(Reference(d,min_col=1,max_col=1,min_row=WR,max_row=WR+4))
w.add_chart(wf,"B12")

mg=BarChart(); mg.type="col"; mg.grouping="clustered"; mg.gapWidth=55; mg.overlap=-10
mg.height=7.9; mg.width=8.4; mg.y_axis.numFmt='0%'; mg.visible_cells_only=False
mg.y_axis.scaling.min=0; mg.y_axis.scaling.max=0.24; mg.y_axis.majorUnit=0.05
mg.x_axis.delete=False; mg.y_axis.delete=False
for col in (8,9,10):
    mg.add_data(Reference(d,min_col=col,max_col=col,min_row=M0,max_row=MR+len(MARQUES)-1),titles_from_data=True)
for s,cc in zip(mg.series,(BLUE3,BLUE2,BLUE)):
    s.graphicalProperties.solidFill=cc; s.graphicalProperties.line.solidFill=cc
mg.series[2].dLbls=DataLabelList(); mg.series[2].dLbls.showVal=True
mg.series[2].dLbls.numFmt='0.0%'; mg.series[2].dLbls.dLblPos="outEnd"
mg.series[2].dLbls.showSerName=False; mg.series[2].dLbls.showCatName=False; mg.series[2].dLbls.showLegendKey=False
mg.set_categories(Reference(d,min_col=1,max_col=1,min_row=MR,max_row=MR+len(MARQUES)-1))
mg.legend.position="b"
w.add_chart(mg,"F12")

tn=LineChart(); tn.height=7.9; tn.width=8.4; tn.visible_cells_only=False
tn.y_axis.numFmt='0'; tn.y_axis.scaling.min=95; tn.y_axis.scaling.max=125; tn.y_axis.majorUnit=10
tn.x_axis.delete=False; tn.y_axis.delete=False
for col in (2,3): tn.add_data(Reference(d,min_col=col,max_col=col,min_row=X1,max_row=XR+2),titles_from_data=True)
for s,cc in zip(tn.series,(ORANGE,BLUE)):
    s.graphicalProperties.line.solidFill=cc; s.graphicalProperties.line.width=25000
    s.marker=Marker(symbol="circle",size=6); s.smooth=False
    s.marker.graphicalProperties.solidFill=cc; s.marker.graphicalProperties.line.solidFill=cc
tn.set_categories(Reference(d,min_col=1,max_col=1,min_row=XR,max_row=XR+2))
tn.legend.position="b"
w.add_chart(tn,"J12")

box(w,33,2,GEND,13)
w.cell(33,2,"  Portefeuille — marque & campus").font=F(10.5,True,INK); w.cell(33,2).alignment=L
w.cell(33,13,"exercice 2026 · variation vs 2025  ").font=F(8.5,c=INK3); w.cell(33,13).alignment=R
for j,t in enumerate(["Marque / campus","CA 2026","Δ CA","EBITDA","Δ EBITDA","Part EBITDA","Marge EBITDA",
                      "Δ marge","Inscrits","Rempl.","Mix alt.","Alerte"],2):
    c=w.cell(34,j,t); c.font=F(8.5,True,INK3); c.fill=fill(PANEL2); c.alignment=L if j==2 else R
    c.border=Border(bottom=sd(LINE),top=sd(LINE))
SRC="Données!$A$%d:$L$%d"%(S1,S2); MK=D("B")
FMT={"C":EUR,"D":DPCT,"E":EUR,"F":DPCT,"G":'0.0%;-0.0%',"H":'0.0%;-0.0%',"I":DPT,
     "J":'#,##0',"K":PCT0,"L":PCT0,"M":"General"}
for i,(name,lvl) in enumerate(ROWS):
    r=G0+i
    w.cell(r,2,name).alignment=ind(lvl)
    w.cell(r,2).font=F(10 if lvl==0 else 9.5,lvl<=1,INK if lvl<2 else INK2)
    if lvl==0:
        f={"C":"=SUM(%s)"%D("E"),"D":"=C%d/SUM(%s)-1"%(r,D("D")),"E":"=SUM(%s)"%D("H"),
           "F":"=E%d/SUM(%s)-1"%(r,D("G")),"G":"=E%d/$E$%d"%(r,G0),"H":"=E%d/C%d"%(r,r),
           "I":"=(H%d-SUM(%s)/SUM(%s))*100"%(r,D("G"),D("D")),"J":"=SUM(%s)"%D("I"),
           "K":"=SUM(%s)/SUM(%s)"%(D("J"),D("K")),
           "L":"=SUMPRODUCT(%s,%s)/SUM(%s)"%(D("J"),D("L"),D("J")),"M":'=IF(H%d<0.08,"FRAGILE","")'%r}
    elif lvl==1:
        S=lambda col: 'SUMIF(%s,$B%d,%s)'%(MK,r,D(col))
        f={"C":"="+S("E"),"D":"=C%d/%s-1"%(r,S("D")),"E":"="+S("H"),"F":"=E%d/%s-1"%(r,S("G")),
           "G":"=E%d/$E$%d"%(r,G0),"H":"=E%d/C%d"%(r,r),"I":"=(H%d-%s/%s)*100"%(r,S("G"),S("D")),
           "J":"="+S("I"),
           "K":"=%s/%s"%(S("J"),S("K")),
           "L":"=SUMPRODUCT((%s=$B%d)*%s*%s)/%s"%(MK,r,D("J"),D("L"),S("J")),
           "M":'=IF(H%d<0.08,"FRAGILE","")'%r}
    else:
        V=lambda n: 'VLOOKUP($B%d,%s,%d,0)'%(r,SRC,n)
        f={"C":"="+V(5),"D":"=C%d/%s-1"%(r,V(4)),"E":"="+V(8),
           "F":'=IF(%s<=0,"n/s",E%d/%s-1)'%(V(7),r,V(7)),"G":"=E%d/$E$%d"%(r,G0),"H":"=E%d/C%d"%(r,r),
           "I":"=(H%d-%s/%s)*100"%(r,V(7),V(4)),"J":"="+V(9),"K":"=%s/%s"%(V(10),V(11)),"L":"="+V(12),
           "M":'=IF(H%d<0.08,"FRAGILE","")'%r}
    for col,formula in f.items():
        j=column_index_from_string(col)
        c=w.cell(r,j,formula); c.number_format=FMT[col]; c.alignment=R if col!="M" else Cn
        c.font=F(9.5,lvl<=1,INK if lvl<2 else INK2,MONO if col!="M" else UI)
    w.cell(r,13).font=F(8,True,ORANGE)
    for j in range(2,14):
        w.cell(r,j).border=Border(bottom=sd(LINE2))
        if lvl==0: w.cell(r,j).fill=fill(PANEL2)
    if lvl==2: w.row_dimensions[r].outlineLevel=2
    elif lvl==1: w.row_dimensions[r].outlineLevel=1
w.sheet_properties.outlinePr.summaryBelow=False
w.conditional_formatting.add("H%d:H%d"%(G0,GEND),
    DataBarRule(start_type="num",start_value=0,end_type="num",end_value=0.24,color="FF"+BLUE,showValue=True))
for col in "DFI":
    for op,cc in (("greaterThan",GOOD),("lessThan",CRIT)):
        w.conditional_formatting.add("%s%d:%s%d"%(col,G0,col,GEND),
            CellIsRule(operator=op,formula=["0"],font=Font(name=MONO,size=9.5,color=cc)))
w.conditional_formatting.add("K%d:K%d"%(G0,GEND),CellIsRule(operator="lessThan",formula=["0.70"],
    font=Font(name=MONO,size=9.5,bold=True,color=WARN)))

box(w,CAL,2,CAL+2,13,bg=PANEL2)
w.merge_cells(start_row=CAL,start_column=2,end_row=CAL+2,end_column=13)
w.cell(CAL,2,"  Le signal du cockpit.  Tunon pèse 8,4 % du CA mais 2,1 % de l'EBITDA. Ipac, le réseau le plus jeune, "
             "pèse 11,2 % du CA pour 9,5 % de l'EBITDA — il monte en charge, avec Rennes et Montpellier à 61 % de "
             "remplissage. À l'inverse Pigier, la plus petite des cinq, est la plus rentable (21,2 %). Et le groupe "
             "n'occupe que 75 % de ses places : 1 032 places ouvertes et vides. Le cadrage 2027 ne peut pas être un "
             "« + x % » uniforme.").font=F(9.5,c=INK2)
w.cell(CAL,2).alignment=TOP
for r in range(CAL,CAL+3): w.cell(r,2).border=Border(left=sd(ORANGE,"thick"))
w.cell(CAL+4,2,"Source : socle CRM et comptabilité rapprochés · vues V_COCKPIT, V_MOTEUR_CAL, V_ALLOCATION · "
               "les marges se lisent en POINTS. Tout le cockpit est en formules : seul l'onglet Données se saisit.").font=F(8.5,c=INK3)
w.cell(CAL+4,2).alignment=L
w.sheet_view.zoomScale=90
wb.active=1
out="/home/user/demo5/eduservices/COCKPIT_EDUSERVICES.xlsx"
wb.save(out)
print("SAVED",out)
print("socle %d..%d  acquisition %d..%d  bridge %d..%d  marques %d..%d  waterfall %d..%d  tension %d..%d"
      %(S1,S2,A0+1,A1+1,BR,BR+4,MR,MR+len(MARQUES)-1,WR,WR+4,XR,XR+2))
print("grille cockpit : lignes %d..%d (%d lignes)  ·  encadré %d"%(G0,GEND,len(ROWS),CAL))
