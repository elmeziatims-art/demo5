#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""RECONCILIATION_CADRAGE.xlsx — la réconciliation réglée pour tomber sur la cible.
Montre Référence / Cible / Construit (après réglage productivité) + le levier utilisé.
CA + EBITDA + Marge à 0 écart ; effectif reste sur-déterminé (assumé)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E2EFEB"; CARD2="F5F7F9"
FAINT="7D8B98"; WHITE="FFFFFF"; OCHRE="B3641C"; OCHRED="8A4A12"; OCHREBG="F6E8D8"
GREEN="1E7A55"; RULE="C8D2DA"; SOFT="51606D"; NAVY="3D4F8F"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
CTR=Alignment("center",vertical="center",wrap_text=True)
LFT=Alignment("left",vertical="center"); RGT=Alignment("right",vertical="center")
thin=Side(style="thin",color=RULE); med=Side(style="medium",color=RULE)
EUR='#,##0 "€";-#,##0 "€";"-"'; PCT='0.0%'; PT='0.0" pt"'; DPT='+0.0" pt";-0.0" pt";0" pt"'
DEUR='+#,##0 "€";-#,##0 "€";0 "€"'; DPCT='+0.0%;-0.0%;0.0%'

wb=openpyxl.Workbook(); ws=wb.active; ws.title="Réconciliation"
ws.sheet_view.showGridLines=False

ws["A1"]="RÉCONCILIATION — Référence · Cible · Construit (réglé)"; ws["A1"].font=F(15,True,INK)
ws["A2"]="Cadrage V01 avec effort de productivité porté à 1,85 %. CA, EBITDA et Marge tombent sur la cible."; ws["A2"].font=F(9,False,TEALD)

# ---- table 1 : indicateurs ----
hr=4
heads=["Indicateur","Référence 2026","Cible","Construit (réglé)","Écart","Écart %"]
for j,h in enumerate(heads,1):
    c=ws.cell(hr,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL)
    c.alignment=LFT if j==1 else CTR; c.border=Border(bottom=med)

# (label, ref, cible, construit, type, fmt)  type: 'eur' 'pct' 'num'
ROWS=[
 ("Chiffre d'affaires",22544725,23671961,23673147,"eur"),
 ("EBITDA",            3291530, 3550794, 3550794, "eur"),
 ("Marge EBITDA",      0.1460,  0.1500,  0.1500,  "pct"),
 ("Effectif",          3036,    3036,    3175,    "num"),
]
r=hr+1
for lab,ref,cib,con,typ in ROWS:
    ws.cell(r,1,lab).font=F(10, lab in ("EBITDA",)); ws.cell(r,1).alignment=LFT
    fmt = EUR if typ=="eur" else (PCT if typ=="pct" else '#,##0')
    ws.cell(r,2,ref).number_format=fmt; ws.cell(r,2).font=F(10,False,SOFT)
    ws.cell(r,3,cib).number_format=fmt; ws.cell(r,3).font=F(10,False,OCHRE)   # cible = ochre
    ws.cell(r,4,con).number_format=fmt; ws.cell(r,4).font=F(10,True,NAVY)     # construit = navy
    # écart construit - cible
    d=con-cib
    if typ=="pct":
        ws.cell(r,5,d).number_format=DPT if False else '+0.0%;-0.0%;0.0%'
        # marge : écart en points
        ws.cell(r,5,(con-cib)).number_format=DPCT
    else:
        ws.cell(r,5,d).number_format=DEUR if typ=="eur" else '+#,##0;-#,##0;0'
    dp = (con-cib)/cib if cib else 0
    ws.cell(r,6,dp).number_format=DPCT
    ok = abs(dp)<0.005
    ecol = GREEN if ok else OCHRED
    ws.cell(r,5).font=F(10,True,ecol); ws.cell(r,6).font=F(10,True,ecol)
    for j in (2,3,4,5,6): ws.cell(r,j).alignment=RGT
    for j in range(1,7): ws.cell(r,j).border=Border(bottom=thin)
    if r%2==1:
        for j in range(1,7):
            if ws.cell(r,j).fill.fgColor.rgb in (None,'00000000'): ws.cell(r,j).fill=fill(CARD2)
    r+=1

r+=1
ws.cell(r,1,"✓ CA, EBITDA et Marge sur la cible (écart ≈ 0). L'effectif reste à +4,6 % : il vient des leviers de croissance et ne peut pas être ramené à 0 sans faire retomber le CA/EBITDA sous cible (cibles sur-déterminées).").font=F(8,True,OCHRE,True)
r+=2

# ---- table 2 : le réglage ----
ws.cell(r,1,"LE RÉGLAGE").font=F(11,True,TEALD); r+=1
h2=["Levier de cadrage","Avant","Après","Effet sur l'EBITDA"]
for j,h in enumerate(h2,1):
    c=ws.cell(r,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL)
    c.alignment=LFT if j==1 else CTR; c.border=Border(bottom=med)
r+=1
ws.cell(r,1,"Effort de productivité (achats + structure)").font=F(10); ws.cell(r,1).alignment=LFT
ws.cell(r,2,0.010).number_format=PCT; ws.cell(r,3,0.0185).number_format=PCT
ws.cell(r,4,69415).number_format='+#,##0 "€"'
ws.cell(r,2).font=F(10,False,SOFT); ws.cell(r,3).font=F(10,True,TEALD); ws.cell(r,4).font=F(10,True,GREEN)
for j in (2,3,4): ws.cell(r,j).alignment=RGT
for j in range(1,5): ws.cell(r,j).border=Border(bottom=thin)
r+=1
ws.cell(r,1,"Sensibilité : 1 pt de productivité = 81 552 € d'EBITDA (base achats+structure+impôts 2027 = 8 155 167 €).  Levier de coût pur → ne touche pas le CA.").font=F(8,False,FAINT,True)
r+=1
ws.cell(r,1,"Alternative équivalente : effectifs permanents (HYP_FTE_PERM) de 4,0 % → 3,34 % libère les mêmes 69 415 €.").font=F(8,False,FAINT,True)

ws.column_dimensions['A'].width=42
for col,w in zip("BCDEF",[16,15,17,14,11]): ws.column_dimensions[col].width=w

out="/home/user/demo5/eduservices/tagetik/RECONCILIATION_CADRAGE.xlsx"
wb.save(out); print("SAVED",out)
