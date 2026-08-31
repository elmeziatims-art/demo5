#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""RESTITUTIONS_EDUSERVICES.xlsx — classeur final.
Repart du modèle VIVANT (LE_MOTEUR_MODELE.xlsx : onglet Moteur + onglet Organique,
avec leurs cases variables/formules intactes) et y AJOUTE 2 restitutions TCD :
Alloué 2026 et Évolution 2024-26 (cascade fidèle V_ALLOCATION, spec embarquée)."""
import csv, glob
from collections import defaultdict, OrderedDict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E2EFEB"; WHITE="FFFFFF"; RULE="C8D2DA"
SOFT="51606D"; OCHRE="B3641C"; OCHRED="8A4A12"; GREEN="1E7A55"; FAINT="7D8B98"; L0BG="DCE7EE"; L1BG="EAF0F4"; SPECBG="FBF4E6"; NAVY="3D4F8F"
# Membres des dimensions du modèle (pour construire les matrices Tagetik)
DIM_MARQUE   = "MBWAY · ISCOM · IPAC · PIGIER · TUNON"
DIM_CAMPUS   = "MBWAY_BOR/LYO/NAN/PAR · ISCOM_LIL/PAR/TLS · IPAC_MTP/NAN/REN · PIGIER_BOR/LYO · TUNON_LYO/PAR   (14 campus)"
DIM_PROG     = "BAC_CCE · BAC_COM · BAC_MGT · BAC_RH · BAC_TOU · BTS_GES · MAS_COM · MAS_MGT"
DIM_AN       = "B1 · B2 · B3 · BTS1 · BTS2 · M1 · M2"
DIM_MOD      = "INIT (Initial) · ALT (Alternance)"
DIM_EX_REEL  = "2024 · 2025 · 2026   (réel ; le budget = 2027)"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
def LI(i): return Alignment("left",vertical="center",indent=i)
RGT=Alignment("right",vertical="center"); CTR=Alignment("center",vertical="center",wrap_text=True); LFT=Alignment("left",vertical="center")
med=Side(style="medium",color=RULE); thin=Side(style="thin",color=RULE)
EUR='#,##0 "€";-#,##0 "€";"-"'; NUM='#,##0'; PCT='0.0%'; DPT='+0.0;-0.0;0.0'
MQ={'MBWAY':'MBway','ISCOM':'Iscom','IPAC':'Ipac','PIGIER':'Pigier','TUNON':'Tunon'}
ORDER=['MBWAY','ISCOM','IPAC','PIGIER','TUNON']
CITY={'LYO':'Lyon','PAR':'Paris','BOR':'Bordeaux','NAN':'Nantes','MTP':'Montpellier','REN':'Rennes','LIL':'Lille','TLS':'Toulouse'}
MODL={'INIT':'Initial','ALT':'Alternance'}
COMP=glob.glob('/home/user/demo5/eduservices/tgk_data/ext_*/AW_002_000004_000001.csv')[0]
SOCLE=glob.glob('/home/user/demo5/eduservices/tgk_data/ext_*/AW_002_000002_000001.csv')[0]

def hrs(prog,mod):
    if prog.startswith('BAC') and mod=='INIT': return 600
    if prog.startswith('BAC'): return 480
    if prog.startswith('MAS') and mod=='INIT': return 520
    if prog.startswith('MAS'): return 420
    if mod=='INIT': return 1000
    return 700
def compute(year):
    camp=defaultdict(lambda: defaultdict(float)); hold=defaultdict(float)
    with open(COMP,encoding='utf-8') as fh:
        rr=csv.reader(fh,delimiter=';'); next(rr)
        for row in rr:
            en,acc,ex=row[0],row[1],row[2]
            if ex!=year or acc in ('TEC_EBITDA','TEC_PL'): continue
            a=float(row[5].replace(',','.'))
            if en=='GRP':
                if acc in ('6414','6226','626','6281','6331','6333'): hold['HOLDING']+=a
                elif acc=='6236': hold['MARQUE']+=a
                continue
            if acc=='621': camp[en]['VAC']+=a
            elif acc=='6411': camp[en]['PERM']+=a
            elif acc in ('604','6063'): camp[en]['ODIR']+=a
            elif acc=='6231': camp[en]['MKT']+=a
            elif acc in ('6413','645','613','615','616','625','63511'): camp[en]['STRUCT']+=a
    cls=[]
    with open(SOCLE,encoding='utf-8') as fh:
        rr=csv.reader(fh,delimiter=';'); h=next(rr); ix={n:i for i,n in enumerate(h)}
        g=lambda row,n: float(row[ix[n]].replace(',','.'))
        for row in rr:
            if row[ix['EXERCICE']]!=year: continue
            en=row[ix['ENTITY']]; cls.append(dict(en=en,mq=en.split('_')[0],prog=row[ix['PROGRAMME']],an=row[ix['AN_ETUDE']],mod=row[ix['MODALITE']],
                eff=g(row,'VOL_EFF'),vcl=g(row,'VOL_CLASS'),new=g(row,'VOL_NEW'),ca=g(row,'VOL_EFF')*g(row,'REV_STUD')+g(row,'VOL_NEW')*g(row,'REV_FRAIS_INS'),
                hrs=g(row,'VOL_CLASS')*hrs(row[ix['PROGRAMME']],row[ix['MODALITE']])))
    E=defaultdict(lambda: defaultdict(float)); M=defaultdict(lambda: defaultdict(float)); G=defaultdict(float)
    for c in cls:
        for k,fk in [('HRS','hrs'),('EFF','eff'),('NEW','new'),('CA','ca')]:
            E[c['en']][k]+=c[fk]; M[c['mq']][k]+=c[fk]; G[k]+=c[fk]
    out=[]
    for c in cls:
        en,mq=c['en'],c['mq']; cp=camp[en]
        cvac=cp['VAC']*c['hrs']/E[en]['HRS'] if E[en]['HRS'] else 0; cperm=cp['PERM']*c['hrs']/E[en]['HRS'] if E[en]['HRS'] else 0
        codir=(cp['ODIR']*c['eff']/E[en]['EFF'] if E[en]['EFF'] else 0)+(cp['MKT']*c['new']/E[en]['NEW'] if E[en]['NEW'] else 0)
        cstr=cp['STRUCT']*c['eff']/E[en]['EFF'] if E[en]['EFF'] else 0
        share=(M[mq]['EFF']/G['EFF'])*(E[en]['CA']/M[mq]['CA'] if M[mq]['CA'] else 0)*(c['eff']/E[en]['EFF'] if E[en]['EFF'] else 0)
        siege=(hold['MARQUE']+hold['HOLDING'])*share; propre=c['ca']-(cvac+cperm+codir+cstr); net=propre-siege
        out.append(dict(mq=MQ[mq],campus=CITY.get(en.split('_')[1],en.split('_')[1]),prog=c['prog'],an=c['an'],mod=MODL.get(c['mod'],c['mod']),
            eff=c['eff'],ca=c['ca'],propre=propre,siege=siege,net=net))
    return out
def tree5(rows):
    T=OrderedDict()
    for m in ORDER:
        mm=MQ[m]; sub=[x for x in rows if x['mq']==mm]
        if not sub: continue
        T[mm]=OrderedDict()
        for camp in sorted(set(x['campus'] for x in sub)):
            T[mm][camp]=OrderedDict()
            for prog in sorted(set(x['prog'] for x in sub if x['campus']==camp)):
                T[mm][camp][prog]=OrderedDict()
                for an in sorted(set(x['an'] for x in sub if x['campus']==camp and x['prog']==prog)):
                    T[mm][camp][prog][an]=OrderedDict()
                    for x in [z for z in sub if z['campus']==camp and z['prog']==prog and z['an']==an]:
                        T[mm][camp][prog][an][x['mod']]=x
    return T
def agg(node):
    t=dict(eff=0,ca=0,propre=0,siege=0,net=0)
    def rec(n):
        if 'net' in n and 'prog' in n:
            for k in t: t[k]+=n[k];
            return
        for v in n.values(): rec(v)
    rec(node); return t
def spec(ws,r,title,source,dims,members,elems):
    ws.sheet_view.showGridLines=False
    ws.cell(r,1,title).font=F(14,True,INK); r+=1
    ws.cell(r,1,"Source (vue)").font=F(9,True,OCHRED); ws.cell(r,2,source).font=F(9,False,INK); r+=1
    ws.cell(r,1,"Dimensions").font=F(9,True,OCHRED); ws.cell(r,2,dims).font=F(9,False,INK); r+=1
    ws.cell(r,1,"Membres des dimensions (matrice Tagetik)").font=F(9,True,NAVY); r+=1
    for dim,mem in members:
        ws.cell(r,1,"   "+dim).font=F(8.5,True,NAVY); ws.cell(r,2,mem).font=F(8.5,False,SOFT); r+=1
    ws.cell(r,1,"Mesures / comptes").font=F(9,True,OCHRED); r+=1
    for lab,comp in elems:
        ws.cell(r,1,"   "+lab).font=F(8.5,True,TEALD); ws.cell(r,2,comp).font=F(8.5,False,SOFT); r+=1
    for j in range(1,9): ws.cell(r,j).fill=fill(SPECBG); ws.cell(r,j).border=Border(bottom=med)
    return r+1
def livespec(ws,r0,source,dims,members,fields):
    """Panneau de spécification (vue + dimensions + membres) sous le contenu vivant."""
    ws.cell(r0,1,"◆ SPÉCIFICATION — pour reconstruire via une matrice Tagetik").font=F(12,True,INK); r=r0+1
    ws.cell(r,1,"Source (vue)").font=F(9,True,OCHRED); ws.cell(r,2,source).font=F(9,False,INK); r+=1
    ws.cell(r,1,"Dimensions").font=F(9,True,OCHRED); ws.cell(r,2,dims).font=F(9,False,INK); r+=1
    ws.cell(r,1,"Membres des dimensions").font=F(9,True,NAVY); r+=1
    for dim,mem in members:
        ws.cell(r,1,"   "+dim).font=F(8.5,True,NAVY); ws.cell(r,2,mem).font=F(8.5,False,SOFT); r+=1
    ws.cell(r,1,"Champs de la vue").font=F(9,True,OCHRED); r+=1
    for lab,comp in fields:
        ws.cell(r,1,"   "+lab).font=F(8.5,True,TEALD); ws.cell(r,2,comp).font=F(8.5,False,SOFT); r+=1
    for j in range(1,3): ws.cell(r,j).border=Border(bottom=med)
    return r+1
def hdr(ws,r,cols):
    for j,h in enumerate(cols,1):
        c=ws.cell(r,j,h); c.font=F(8.5,True,WHITE); c.fill=fill(TEAL); c.alignment=LFT if j==1 else CTR; c.border=Border(bottom=med)

# ---- charge le modèle vivant (2 onglets moteur + organique intacts) ----
wb=openpyxl.load_workbook("/home/user/demo5/eduservices/tagetik/LE_MOTEUR_MODELE.xlsx")
print("onglets vivants repris :",wb.sheetnames)

# ---- spec sous le contenu vivant (sans toucher aux formules au-dessus) ----
wm=wb["Le moteur (modèle)"]
livespec(wm,58,
    "V_CAMPAGNES (calibration : élasticité acquisition, conversion, réfs)  +  socle AW_002_000002_000001 (CA/inscrit réel)",
    "ENTITY (campus) ▸ groupé par MARQUE   —   l'élasticité vit au campus, la marque = sous-total",
    [("MARQUE",DIM_MARQUE),("CAMPUS (ENTITY)",DIM_CAMPUS)],
    [("Leads payants (L.pay 24/25/26)","PAID_REF / LEAD_REF  (V_CAMPAGNES)"),
     ("Budget acquisition (Bud.24/25/26)","SPEND_ACQ_REF  (compte 6231, V_CAMPAGNES)"),
     ("Élasticité","REND_ACQ = régression 3 ans  =SLOPE(LN(leads);LN(budget))"),
     ("Conversion","CONVERSION = inscrits ÷ leads  (V_CAMPAGNES)"),
     ("CA / inscrit","(VOL_NEW·REV_STUD + VOL_NEW·REV_FRAIS_INS) ÷ VOL_NEW  par campus (socle 2026)")])
wo=wb["Budget de marque (organique)"]
livespec(wo,48,
    "V_CAMPAGNES (élasticité marque, réfs organiques)  +  socle AW_002_000002_000001",
    "ENTITY (campus) ▸ groupé par MARQUE",
    [("MARQUE",DIM_MARQUE),("CAMPUS (ENTITY)",DIM_CAMPUS)],
    [("Budget de marque (Bud.mq 24/25/26)","SPEND_BRAND_REF  (compte 6236, V_CAMPAGNES)"),
     ("Leads organiques (Org.24/25/26)","ORG_REF  (V_CAMPAGNES)"),
     ("Élasticité marque","REND_BRAND = régression 3 ans  =SLOPE(LN(organiques);LN(budget marque))")])

# ---- ajoute Alloué ----
wa=wb.create_sheet("Alloué 2026"); wa.sheet_properties.outlinePr.summaryBelow=False
r=spec(wa,1,"RESTITUTION — P&L chargé 2026 (avant / après allocation)","V_ALLOCATION  (Q_RAPPORT_ALLOUE)",
    "Marque ▸ Campus ▸ Programme ▸ Année ▸ Modalité   ·   EXERCICE = 2026",
    [("MARQUE",DIM_MARQUE),("CAMPUS (ENTITY)",DIM_CAMPUS),("PROGRAMME",DIM_PROG),
     ("AN_ETUDE",DIM_AN),("MODALITE",DIM_MOD),("EXERCICE","2026")],
    [("Effectif / CA","VOL_EFF / (706+7062+708)"),
     ("EBITDA propre","CA − (621+6411 + 604/6063/6231 + 6413/645/613/615/616/625/63511)"),
     ("Quote-part siège","6236 + 6414/6226/626/6281/6331/6333  (= COST_SIEGE)"),
     ("EBITDA net","CA − tous coûts  (= MARGE_COMPLETE ; 6811 exclu)"),
     ("Marge %","EBITDA net ÷ CA  (membre calculé)")])
hdr(wa,r,["Marque ▸ Campus ▸ Programme ▸ Année ▸ Modalité","Effectif","CA","EBITDA propre","Q-part siège","EBITDA net","Marge %"]); r+=1
def emitA(r,label,v,lvl):
    wa.cell(r,1,label).font=F(9 if lvl<=1 else 8.5,lvl<=1); wa.cell(r,1).alignment=LI(lvl)
    wa.cell(r,2,round(v['eff'])).number_format=NUM; wa.cell(r,3,round(v['ca'])).number_format=EUR
    wa.cell(r,4,round(v['propre'])).number_format=EUR; wa.cell(r,5,round(v['siege'])).number_format=EUR
    mg=v['net']/v['ca'] if v['ca'] else 0
    wa.cell(r,6,round(v['net'])).number_format=EUR; wa.cell(r,6).font=F(9 if lvl<=1 else 8.5,lvl<=1,OCHRED if mg<0.05 else (TEALD if lvl<=1 else INK))
    wa.cell(r,7,mg).number_format=PCT; wa.cell(r,7).font=F(9 if lvl<=1 else 8.5,lvl<=1,OCHRED if mg<0.05 else TEALD)
    for j in (2,3,4,5,6,7): wa.cell(r,j).alignment=RGT
    if lvl==0:
        for j in range(1,8): wa.cell(r,j).fill=fill(L0BG); wa.cell(r,j).border=Border(top=med,bottom=thin)
    elif lvl==1:
        for j in range(1,8): wa.cell(r,j).fill=fill(L1BG)
    else:
        for j in range(1,8): wa.cell(r,j).border=Border(bottom=thin)
    if lvl>0:
        wa.row_dimensions[r].outline_level=min(lvl,4)
        if lvl>=2: wa.row_dimensions[r].hidden=True
def walkA(r,node,lvl):
    for k,child in node.items():
        if 'net' in child: emitA(r,k,child,lvl); r+=1
        else: emitA(r,k,agg(child),lvl); r+=1; r=walkA(r,child,lvl+1)
    return r
TA=tree5(compute('2026')); r=walkA(r,TA,0)
g=dict(eff=0,ca=0,propre=0,siege=0,net=0)
for mq in TA.values():
    a=agg(mq)
    for k in g: g[k]+=a[k]
emitA(r,"GROUPE",g,0)
for j in range(1,8): wa.cell(r,j).font=F(11,True,TEAL); wa.cell(r,j).fill=fill(TEALBG); wa.cell(r,j).border=Border(top=med,bottom=med)
wa.column_dimensions['A'].width=44
for col,w in zip("BCDEFG",[10,13,13,12,13,9]): wa.column_dimensions[col].width=w

# ---- ajoute Évolution ----
we=wb.create_sheet("Évolution 24-26"); we.sheet_properties.outlinePr.summaryBelow=False
DPCT='+0.0%;-0.0%;0.0%'
r=spec(we,1,"RESTITUTION — Évolution CA · Charges · Marge 2024→2026","V_ALLOCATION  (Q_RAPPORT_EVOLUTION)",
    "Marque ▸ Campus ▸ Programme ▸ Année ▸ Modalité   ·   EXERCICE en colonnes",
    [("MARQUE",DIM_MARQUE),("CAMPUS (ENTITY)",DIM_CAMPUS),("PROGRAMME",DIM_PROG),
     ("AN_ETUDE",DIM_AN),("MODALITE",DIM_MOD),("EXERCICE",DIM_EX_REEL)],
    [("CA","706 + 7062 + 708  (valeur par exercice)"),
     ("Charges","CA − EBITDA net  (coût complet chargé, = MARGE_COMPLETE retranchée)"),
     ("Marge","= CA − Charges  (formule)"),
     ("Écart (€)","= valeur 2026 − valeur 2024  (formule)"),
     ("Évolution (%)","= valeur 2026 ÷ valeur 2024 − 1  (formule)"),
     ("Marge %","= Marge ÷ CA  (formule)")])
# en-têtes sur 2 lignes : groupe puis année
gh=r
for c0,lab in ((2,"CHIFFRE D'AFFAIRES"),(7,"CHARGES (chargées)"),(12,"MARGE")):
    cc=we.cell(gh,c0,lab); cc.font=F(8.5,True,WHITE); cc.fill=fill(TEALD); cc.alignment=CTR
    we.merge_cells(start_row=gh,start_column=c0,end_row=gh,end_column=c0+4)
hdr(we,gh+1,["Marque ▸ Campus ▸ Programme ▸ Année ▸ Modalité",
    "CA 2024","CA 2025","CA 2026","Écart €","Évol %",
    "Ch. 2024","Ch. 2025","Ch. 2026","Écart €","Évol %",
    "Mg 2024","Mg 2025","Mg 2026","Écart €","Marge % 26"])
r=gh+2
data={y:compute(y) for y in ('2024','2025','2026')}
def key(x): return (x['mq'],x['campus'],x['prog'],x['an'],x['mod'])
idx={y:{key(x):x for x in data[y]} for y in data}
T26=tree5(data['2026'])
def ckeys(node):
    ks=[]
    def rec(n):
        if 'net' in n and 'prog' in n: ks.append(key(n)); return
        for v in n.values(): rec(v)
    rec(node); return ks
def emitE(r,label,node,lvl):
    ks=ckeys(node) if not('net' in node and 'prog' in node) else [key(node)]
    def ca(y): return sum(idx[y][k]['ca'] for k in ks if k in idx[y])
    def net(y): return sum(idx[y][k]['net'] for k in ks if k in idx[y])
    we.cell(r,1,label).font=F(9 if lvl<=1 else 8.5,lvl<=1); we.cell(r,1).alignment=LI(lvl)
    # valeurs : CA (B,C,D) et Charges (G,H,I) ; le reste = FORMULES
    for j,y in ((2,'2024'),(3,'2025'),(4,'2026')): we.cell(r,j,round(ca(y))).number_format=EUR
    for j,y in ((7,'2024'),(8,'2025'),(9,'2026')): we.cell(r,j,round(ca(y)-net(y))).number_format=EUR
    we.cell(r,5,f"=D{r}-B{r}").number_format=EUR            # écart CA €
    we.cell(r,6,f"=IFERROR(D{r}/B{r}-1,0)").number_format=DPCT
    we.cell(r,10,f"=I{r}-G{r}").number_format=EUR           # écart charges €
    we.cell(r,11,f"=IFERROR(I{r}/G{r}-1,0)").number_format=DPCT
    we.cell(r,12,f"=B{r}-G{r}").number_format=EUR           # marge = CA − charges (formule)
    we.cell(r,13,f"=C{r}-H{r}").number_format=EUR
    we.cell(r,14,f"=D{r}-I{r}").number_format=EUR
    we.cell(r,15,f"=N{r}-L{r}").number_format=EUR           # écart marge €
    we.cell(r,16,f"=IFERROR(N{r}/D{r},0)").number_format=PCT  # marge % 2026
    for j in range(2,17): we.cell(r,j).alignment=RGT
    # couleurs marge% et évol
    mg26=(net('2026')/ca('2026')) if ca('2026') else 0
    we.cell(r,16).font=F(9 if lvl<=1 else 8.5,lvl<=1,OCHRED if mg26<0.05 else TEALD)
    we.cell(r,14).font=F(9 if lvl<=1 else 8.5,lvl<=1,INK)
    if lvl==0:
        for j in range(1,17): we.cell(r,j).fill=fill(L0BG); we.cell(r,j).border=Border(top=med,bottom=thin)
    elif lvl==1:
        for j in range(1,17): we.cell(r,j).fill=fill(L1BG)
    else:
        for j in range(1,17): we.cell(r,j).border=Border(bottom=thin)
    if lvl>0:
        we.row_dimensions[r].outline_level=min(lvl,4)
        if lvl>=2: we.row_dimensions[r].hidden=True
def walkE(r,node,lvl):
    for k,child in node.items():
        emitE(r,k,child,lvl); r+=1
        if not('net' in child and 'prog' in child): r=walkE(r,child,lvl+1)
    return r
r=walkE(r,T26,0)
we.column_dimensions['A'].width=40
for col in "BCDEFGHIJKLMNOP": we.column_dimensions[col].width=10

# ---- graphe : trajectoire de la marge nette % par marque (courbes) ----
from openpyxl.chart import LineChart, Reference
we.cell(gh,18,"Marge nette %").font=F(9,True,TEALD)
we.cell(gh+1,18,""); we.cell(gh+1,19,2024); we.cell(gh+1,20,2025); we.cell(gh+1,21,2026)
for c in (19,20,21): we.cell(gh+1,c).font=F(8.5,True,SOFT); we.cell(gh+1,c).alignment=CTR
rr=gh+2
for m in ORDER:
    mm=MQ[m]; ks=[key(x) for x in data['2026'] if x['mq']==mm]
    we.cell(rr,18,mm).font=F(9)
    for c,y in ((19,'2024'),(20,'2025'),(21,'2026')):
        ca=sum(idx[y][k]['ca'] for k in ks if k in idx[y]); net=sum(idx[y][k]['net'] for k in ks if k in idx[y])
        we.cell(rr,c, net/ca if ca else 0).number_format=PCT
    rr+=1
ch=LineChart(); ch.title="Marge nette chargée par marque — 2024 → 2026"; ch.style=12; ch.height=8; ch.width=17
ch.y_axis.numFmt='0%'; ch.y_axis.majorGridlines=None; ch.x_axis.delete=False; ch.y_axis.delete=False
dat=Reference(we,min_col=18,max_col=21,min_row=gh+2,max_row=rr-1)
ch.add_data(dat,from_rows=True,titles_from_data=True)
ch.set_categories(Reference(we,min_col=19,max_col=21,min_row=gh+1,max_row=gh+1))
we.add_chart(ch,"R"+str(gh+9))
we.freeze_panes="B"+str(gh+2)

out="/home/user/demo5/eduservices/tagetik/RESTITUTIONS_EDUSERVICES.xlsx"
wb.save(out)
print("SAVED",out," onglets:",wb.sheetnames)
print("  Alloué EBITDA net groupe =",round(g['net']))
