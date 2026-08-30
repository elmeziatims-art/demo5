#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""RAPPORT_EVOLUTION_DEPLIABLE.xlsx — la trajectoire de la marge chargée.
Marque → Campus → Programme → Année → Modalité, dépliable.
Une seule mesure, trois millésimes : marge nette (après allocation à l'effectif)
en 2024 · 2025 · 2026, + Δ en points. On voit qui se redresse et qui s'enfonce.
CA par classe foote à la compta ; net foote au groupe chaque année."""
import csv, glob
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E2EFEB"; CARD2="F5F7F9"
FAINT="7D8B98"; WHITE="FFFFFF"; OCHRE="B3641C"; OCHRED="8A4A12"; RULE="C8D2DA"; SOFT="51606D"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
RGT=Alignment("right",vertical="center")
def LFT(ind=0): return Alignment("left",vertical="center",indent=ind)
thin=Side(style="thin",color=RULE); med=Side(style="medium",color=RULE)
EUR='#,##0;-#,##0;"-"'; PCT='0.0%'
YEARS=[2024,2025,2026]
def bloc(a):
    if a[:1]=='7': return 'CA'
    if a in ('604','6063','621','6231'): return 'DIRECT'
    if a in ('6411','6413','6414','645'): return 'PERSO'
    if a in ('613','615','616','6226','6236','625','626','6281'): return 'STRUCT'
    if a in ('6331','6333','63511'): return 'IMPOT'
    if a=='6811': return 'DOTAT'
    return 'X'
BRAND={'MBWAY':'MBway','ISCOM':'Iscom','IPAC':'Ipac','PIGIER':'Pigier','TUNON':'Tunon'}
BR_ORDER=['MBWAY','ISCOM','IPAC','PIGIER','TUNON']
CITY={'LYO':'Lyon','PAR':'Paris','BOR':'Bordeaux','NAN':'Nantes','MTP':'Montpellier',
      'REN':'Rennes','LIL':'Lille','TLS':'Toulouse'}
MODL={'ALT':'Alternance','INIT':'Initial'}
def campus_label(en): return CITY.get(en.split('_')[1],en.split('_')[1])

# compta : coût EBITDA propre par campus + pool holding, PAR ANNÉE
comp=glob.glob('/home/user/demo5/eduservices/tgk_data/ext_*/AW_002_000004_000001.csv')[0]
camp_cost=defaultdict(lambda:defaultdict(float)); hold_pool=defaultdict(float)
with open(comp,encoding='utf-8') as f:
    rd=csv.reader(f,delimiter=';'); next(rd)
    for row in rd:
        en,acc,ex=row[0],row[1],row[2]
        if acc in ('TEC_EBITDA','TEC_PL'): continue
        try: y=int(ex)
        except: continue
        if y not in YEARS: continue
        b=bloc(acc)
        if b in ('CA','DOTAT','X'): continue
        amt=float(row[5].replace(',','.'))
        if en=='GRP': hold_pool[y]+=amt
        else: camp_cost[en][y]+=amt

# CRM : CA + effectif par classe, PAR ANNÉE
crm=glob.glob('/home/user/demo5/eduservices/tgk_data/ext_*/AW_002_000002_000001.csv')[0]
cls=defaultdict(lambda:{y:{'ca':0.0,'eff':0.0} for y in YEARS})
camp_eff=defaultdict(lambda:defaultdict(float)); eff_group=defaultdict(float)
with open(crm,encoding='utf-8') as f:
    rd=csv.reader(f,delimiter=';'); next(rd)
    for row in rd:
        try: y=int(row[6])
        except: continue
        if y not in YEARS: continue
        en,prog,an,mod=row[2],row[3],row[4],row[5]
        g=lambda i: float(row[i].replace(',','.'))
        eff=g(12); new=g(10); ca=g(15)*eff+g(16)*new
        k=(en,prog,an,mod)
        cls[k][y]['ca']+=ca; cls[k][y]['eff']+=eff
        camp_eff[en][y]+=eff; eff_group[y]+=eff

# noeud : ca[y], cost[y]  (coût chargé complet)
def node(): return {'ca':defaultdict(float),'cost':defaultdict(float)}
def add(n,o):
    for y in YEARS: n['ca'][y]+=o['ca'][y]; n['cost'][y]+=o['cost'][y]
tree=defaultdict(lambda:defaultdict(lambda:defaultdict(lambda:defaultdict(lambda:defaultdict(node)))))
for (en,prog,an,mod),yv in cls.items():
    br=en.split('_')[0]; leaf=tree[br][en][prog][an][mod]
    for y in YEARS:
        eff=yv[y]['eff']; ca=yv[y]['ca']
        cc=camp_cost[en][y]*(eff/camp_eff[en][y]) if camp_eff[en][y] else 0
        ch=hold_pool[y]*(eff/eff_group[y]) if eff_group[y] else 0
        leaf['ca'][y]+=ca; leaf['cost'][y]+=cc+ch
def roll(children):
    n=node()
    for c in children.values(): add(n, c if 'ca' in c else roll(c))
    return n

wb=openpyxl.Workbook(); ws=wb.active; ws.title="Évolution marge chargée"
ws.sheet_view.showGridLines=False
ws.sheet_properties.outlinePr.summaryBelow=False
ws["A1"]="ÉVOLUTION DE LA MARGE CHARGÉE — dépliable  ·  2024 → 2026"; ws["A1"].font=F(15,True,INK)
ws["A2"]="Marge nette après allocation (clé effectif), trois millésimes. Qui se redresse, qui s'enfonce ? Marque → Campus → Programme → Année → Modalité."; ws["A2"].font=F(9,False,TEALD)

hr=4
heads=["Maille","CA 2026","Marge 2024","Marge 2025","Marge 2026","Δ 26/24 (pt)"]
for j,h in enumerate(heads,1):
    c=ws.cell(hr,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL)
    c.alignment=LFT() if j==1 else RGT; c.border=Border(bottom=med)
r=hr+1
def marg(n,y):
    ca=n['ca'][y]; return (ca-n['cost'][y])/ca if ca else None
def wr(lbl,n,lvl,bold=False,color=INK,bg=None,box=None):
    global r
    sz=10 if lvl<=1 else 9
    ws.cell(r,1,lbl).font=F(sz,bold,color); ws.cell(r,1).alignment=LFT(lvl)
    ws.cell(r,2,round(n['ca'][2026])).number_format=EUR; ws.cell(r,2).font=F(sz,bold,color); ws.cell(r,2).alignment=RGT
    ms={}
    for j,y in enumerate(YEARS):
        m=marg(n,y); ms[y]=m
        cc=ws.cell(r,3+j, m if m is not None else "")
        if m is not None:
            cc.number_format=PCT
            col=OCHRED if m<0.05 else (TEALD if y==2026 else SOFT)
            cc.font=F(sz,bold and y==2026,col)
        cc.alignment=RGT
    if ms[2024] is not None and ms[2026] is not None:
        d=(ms[2026]-ms[2024])*100
        cc=ws.cell(r,6,d); cc.number_format='+0.0;-0.0'
        cc.font=F(sz,False,TEALD if d>=0 else OCHRE); cc.alignment=RGT
    if bg:
        for j in range(1,7): ws.cell(r,j).fill=fill(bg)
    if box:
        for j in range(1,7): ws.cell(r,j).border=box
    if lvl>0:
        ws.row_dimensions[r].outline_level=min(lvl,4); ws.row_dimensions[r].hidden=True
    r+=1

grp=node()
for br in BR_ORDER:
    if br not in tree: continue
    nbr=roll(tree[br]); add(grp,nbr)
    wr(BRAND[br],nbr,0,bold=True,color=TEALD,bg=TEALBG,box=Border(top=thin,bottom=thin))
    for en in sorted(tree[br]):
        wr(campus_label(en),roll(tree[br][en]),1,bold=True,color=INK)
        for prog in sorted(tree[br][en]):
            wr(prog,roll(tree[br][en][prog]),2,color=SOFT)
            for an in sorted(tree[br][en][prog]):
                wr(an,roll(tree[br][en][prog][an]),3,color=FAINT)
                for mod in sorted(tree[br][en][prog][an]):
                    wr(MODL.get(mod,mod),tree[br][en][prog][an][mod],4,color=FAINT)
wr("GROUPE",grp,0,bold=True,color=TEAL,bg=TEALBG,box=Border(top=med,bottom=med))
r+=1
for y in YEARS:
    net=grp['ca'][y]-grp['cost'][y]
ws.cell(r,1,"Marge groupe : 2024 = {:.1%} · 2025 = {:.1%} · 2026 = {:.1%}. Δ = points de marge gagnés/perdus vs 2024. Rouge = < 5 %.".format(marg(grp,2024),marg(grp,2025),marg(grp,2026))).font=F(8,True,OCHRE,True)

ws.column_dimensions['A'].width=32
for col,w in zip("BCDEF",[14,12,12,12,13]): ws.column_dimensions[col].width=w
ws.freeze_panes="A5"
out="/home/user/demo5/eduservices/tagetik/RAPPORT_EVOLUTION_DEPLIABLE.xlsx"
wb.save(out)
print("SAVED",out)
for y in YEARS:
    print(y,"marge groupe = {:.1%}  net = {:,.0f}".format(marg(grp,y),grp['ca'][y]-grp['cost'][y]))
