#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""MAQUETTE_CADRAGE_EBITDA.xlsx — onglet Cadrage repensé en logique CFO :
l'EBITDA (€) est l'engagement, la croissance du CA est le garde-fou.
Chiffres de référence 2026 et construit V01 repris de l'export CAD_PIL réel."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E3EFEA"; WHITE="FFFFFF"
SOFT="51606D"; FAINT="7D8B98"; RULE="C8D2DA"; YEL="FFF6DA"; YELB="E8B84B"
OCHRE="B3641C"; OK="1E7A55"; NAVY="3D4F8F"; L0="DCE7EE"; CARD="FBFCFB"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
med=Side(style="medium",color=RULE); thin=Side(style="thin",color=RULE)
box=Border(left=thin,right=thin,top=thin,bottom=thin)
L=Alignment("left",vertical="center"); R=Alignment("right",vertical="center")
C=Alignment("center",vertical="center",wrap_text=True); LW=Alignment("left",vertical="center",wrap_text=True)
EUR='#,##0 "€";-#,##0 "€";"-"'; PCT='0.0%'; PCT2='+0.0%;-0.0%;0.0%'

# ---- valeurs de référence (issues de l'export réel CAD_PIL, V01) ----
REF_CA=23098985; REF_EB=3845790                    # atterrissage 2026
CON_CA=24231704; CON_EB=4111498                    # construit budget 2027 (V01)
MKT_2027=1212886                                   # acquisition + marque 2027
EFF_2027=3176

wb=openpyxl.Workbook(); ws=wb.active; ws.title="Cadrage (EBITDA)"
ws.sheet_view.showGridLines=False
for col,w in zip("ABCDEFG",[34,15,15,15,14,13,15]): ws.column_dimensions[col].width=w

def band(r,txt,c1=1,c2=7,bg=INK,fg=WHITE,sz=11,h=22):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    cc=ws.cell(r,c1,txt); cc.font=F(sz,True,fg); cc.fill=fill(bg); cc.alignment=L
    ws.row_dimensions[r].height=h

# ---------- titre ----------
band(1,"  POSTE DE COMMANDE CFO — Cadrage 2027 · piloté par l'EBITDA",bg=TEALD,sz=13,h=28)
ws.merge_cells("A2:G2")
ws.cell(2,1,"  L'EBITDA (€) est l'engagement. La croissance du CA est le garde-fou — jamais sacrifiée. Cases jaunes = décision CFO.").font=F(9,False,SOFT,True)
ws.row_dimensions[2].height=20

# ---------- ① l'engagement (saisie) ----------
band(4,"①  L'ENGAGEMENT  —  ce que le CFO signe",bg=TEAL,sz=11)
def inp(r,lab,val,fmt,help):
    ws.cell(r,1,lab).font=F(10,True,INK); ws.cell(r,1).alignment=L
    c=ws.cell(r,2,val); c.font=F(11,True,INK); c.fill=fill(YEL); c.number_format=fmt
    c.alignment=C; c.border=Border(*[Side(style="medium",color=YELB)]*4)
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=7)
    ws.cell(r,3,help).font=F(8.5,False,FAINT,True); ws.cell(r,3).alignment=L
    ws.row_dimensions[r].height=20
inp(5,"Amélioration EBITDA vs 2026",0.05,PCT,"le curseur central : de combien on fait progresser le résultat")
inp(6,"Plancher de croissance CA",0.03,PCT,"garde-fou : on refuse de descendre sous ce rythme de croissance")
# EBITDA cible € (dérivé)
ws.cell(7,1,"→ EBITDA cible 2027 (€)").font=F(10,True,TEALD); ws.cell(7,1).alignment=L
cc=ws.cell(7,2,f"={REF_EB}*(1+B5)"); cc.font=F(11,True,TEALD); cc.number_format=EUR; cc.alignment=C
ws.merge_cells("C7:G7")
ws.cell(7,3,"= EBITDA 2026 × (1 + amélioration)  ·  la marge en découle, on ne la fixe pas en rond").font=F(8.5,False,FAINT,True)
ws.row_dimensions[7].height=20

# ---------- ② réconciliation ----------
band(9,"②  RÉCONCILIATION  —  engagement  vs  construit (leviers)",bg=TEAL,sz=11)
hd=["Indicateur","Référence 2026","Engagement","Construit 2027","Écart","","Statut"]
for j,h in enumerate(hd,1):
    c=ws.cell(10,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEALD); c.alignment=C if j>1 else L; c.border=box
ws.merge_cells("E10:F10")
def row(r,lab,ref,eng,con,fmt,statut_formula,strong=False):
    ws.cell(r,1,lab).font=F(10,strong,INK); ws.cell(r,1).alignment=L
    for j,v in ((2,ref),(3,eng),(4,con)):
        c=ws.cell(r,j,v); c.font=F(10,strong,INK); c.number_format=fmt; c.alignment=R; c.border=box
    ws.merge_cells(start_row=r,start_column=5,end_row=r,end_column=6)
    e=ws.cell(r,5,f"=D{r}-C{r}"); e.font=F(10,strong,INK); e.number_format=fmt; e.alignment=R; e.border=box
    s=ws.cell(r,7,statut_formula); s.font=F(9.5,True); s.alignment=C; s.border=box
    ws.row_dimensions[r].height=19
# EBITDA € (l'ancre)
row(11,"EBITDA (€)  ◆ l'ancre",REF_EB,"=B7",CON_EB,EUR,
    '=IF(D11>=C11,"🟢 TENU","🔴 À COMBLER")',strong=True)
for j in range(1,8): ws.cell(11,j).fill=fill(TEALBG)
# marge (informative)
ws.cell(12,1,"Marge EBITDA %").font=F(9.5,False,SOFT); ws.cell(12,1).alignment=L
ws.cell(12,2,f"={REF_EB}/{REF_CA}").number_format=PCT; ws.cell(12,2).alignment=R; ws.cell(12,2).font=F(9.5,c=SOFT); ws.cell(12,2).border=box
ws.cell(12,3,"—").alignment=C; ws.cell(12,3).font=F(9.5,c=FAINT); ws.cell(12,3).border=box
ws.cell(12,4,f"=D11/D13").number_format=PCT; ws.cell(12,4).alignment=R; ws.cell(12,4).font=F(9.5,c=SOFT); ws.cell(12,4).border=box
ws.merge_cells("E12:F12"); ws.cell(12,5,"conséquence, pas objectif").font=F(8,False,FAINT,True); ws.cell(12,5).alignment=C; ws.cell(12,5).border=box
ws.cell(12,7,"").border=box
# CA € (garde-fou)
row(13,"Chiffre d'affaires (€)",REF_CA,f"={REF_CA}*(1+B6)",CON_CA,EUR,
    '=IF(D13>=C13,"🟢 OK","🔴 SOUS PLANCHER")',strong=False)
ws.cell(13,3).font=F(9.5,c=OCHRE)  # engagement = plancher
# croissance CA %
ws.cell(14,1,"Croissance CA").font=F(9.5,False,SOFT); ws.cell(14,1).alignment=L
ws.cell(14,2,"—").alignment=C; ws.cell(14,2).font=F(9.5,c=FAINT); ws.cell(14,2).border=box
c=ws.cell(14,3,"=B6"); c.number_format=PCT2; c.alignment=R; c.font=F(9.5,c=OCHRE); c.border=box
c=ws.cell(14,4,f"=D13/{REF_CA}-1"); c.number_format=PCT2; c.alignment=R; c.font=F(9.5,True,INK); c.border=box
ws.merge_cells("E14:F14")
c=ws.cell(14,5,f"=D14-C14"); c.number_format=PCT2; c.alignment=R; c.font=F(9.5,c=SOFT); c.border=box
s=ws.cell(14,7,'=IF(D14>=C14,"🟢 marge saine","⚠️ vérifier")'); s.font=F(9,True); s.alignment=C; s.border=box
ws.row_dimensions[14].height=19

# ---------- ③ qualité : l'EBITDA est-il sain ? ----------
band(16,"③  QUALITÉ  —  l'EBITDA est-il « sain » ou acheté en affamant la croissance ?",bg=NAVY,sz=10)
q=[("Intensité marketing (% du CA)",f"={MKT_2027}/D13",PCT,"trop bas = on a coupé l'acquisition → cohorte future en danger"),
   ("EBITDA par étudiant (€)",f"=D11/{EFF_2027}",EUR,"la rentabilité unitaire, à surveiller par marque"),
   ("Croissance CA vs plancher",f"=D14-B6",PCT2,"positif = on tient l'engagement SANS sacrifier la croissance")]
r=17
for lab,f_,fmt,note in q:
    ws.cell(r,1,lab).font=F(9.5,False,INK); ws.cell(r,1).alignment=L
    c=ws.cell(r,2,f_); c.number_format=fmt; c.alignment=C; c.font=F(10,True,NAVY); c.border=box
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=7)
    ws.cell(r,3,note).font=F(8.5,False,FAINT,True); ws.cell(r,3).alignment=L
    ws.row_dimensions[r].height=18; r+=1

# ---------- ④ lecture / arbitrage ----------
band(21,"④  LECTURE  —  la décision",bg=TEAL,sz=11)
ws.merge_cells("A22:G24")
verdict=('=IF(AND(D11>=C11,D13>=C13),'
         '"✅ Scénario validé : l EBITDA engagé est tenu ("&TEXT(D11,"#,##0")&" €) ET la croissance du CA ("&TEXT(D14,"+0.0%")&") reste au-dessus du plancher. On signe.",'
         'IF(D11<C11,'
         '"⚠️ EBITDA sous l engagement de "&TEXT(C11-D11,"#,##0")&" € → pousser les leviers (prix, acquisition sur les campus à fort rendement, productivité) ou revoir l ambition.",'
         '"🔴 EBITDA tenu MAIS croissance CA sous le plancher → la marge est achetée en affamant la croissance. Refuser : ré-investir en acquisition."))')
cc=ws.cell(22,1,verdict); cc.font=F(11,True,INK); cc.alignment=LW; cc.fill=fill(CARD)
for rr in (22,):
    for j in range(1,8): ws.cell(rr,j).fill=fill(CARD)
ws.row_dimensions[22].height=58

# note bas de page
ws.merge_cells("A26:G27")
ws.cell(26,1,"Principe : on ancre sur l'EBITDA (valorisation & covenants), on borne la croissance du CA par un plancher (soutenabilité — "
             "dans l'enseignement, couper l'acquisition aujourd'hui réduit la cohorte de demain), et on lit l'ÉCART comme le sujet d'arbitrage "
             "— pas comme une erreur à annuler. La marge se déduit, elle ne se fixe pas en rond.").font=F(8.5,False,FAINT,True)
ws.cell(26,1).alignment=LW; ws.row_dimensions[26].height=44

out="/home/user/demo5/eduservices/MAQUETTE_CADRAGE_EBITDA.xlsx"
wb.save(out); print("SAVED",out)
