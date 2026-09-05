#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""CADRAGE_EBITDA_2027.xlsx — le cadrage inverse, en formules vives.

  ① Base 2026     ce qui est lu en base, rien d'autre (aucune formule)
  ② Cadrage 2027  le modele : constat, cible, tendanciel, ecart, leviers, solde

Tout ce qui est ORANGE se saisit. Tout le reste se recalcule. On bouge un
levier devant le client et le solde se ferme sous ses yeux.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.chart import BarChart, Reference
from cadrage_ebitda import base_reelle, CONSTAT, TENDANCIEL, SCENARIO, CIBLE_MARGE, LIB
from socle_reel import EXERCICES

BLUE="2A78D6"; BLUESOFT="DCE9F9"; ORANGE="EB6834"; ORANGESOFT="FBE3D8"
GOOD="0CA30C"; WARN="FAB219"; CRIT="D03B3B"
INK="131922"; INK2="4D5866"; INK3="7C8798"
CANVAS="EEF1F6"; PANEL="FFFFFF"; PANEL2="F7F9FC"; LINE="DFE4EC"; WHITE="FFFFFF"
UI="Segoe UI"
def F(sz=10,b=False,c=INK): return Font(name=UI,size=sz,bold=b,color=c)
def fill(c): return PatternFill("solid",fgColor=c)
def sd(c=LINE,st="thin"): return Side(style=st,color=c)
L=Alignment("left",vertical="center"); R=Alignment("right",vertical="center")
Cn=Alignment("center",vertical="center"); TOP=Alignment("left",vertical="top",wrap_text=True)
def ind(n): return Alignment("left",vertical="center",indent=n)
EUR='#,##0" €"'; EUR0='#,##0'; PCT1='0.0%'; PCT2='0.00%'; NUM='#,##0'
DEUR='+#,##0" €";−#,##0" €";"—"'; DPT='+0.00" pt";−0.00" pt";"—"'
def box(ws,r1,c1,r2,c2,bg=PANEL,bd=LINE):
    for r in range(r1,r2+1):
        for c in range(c1,c2+1):
            x=ws.cell(r,c); x.fill=fill(bg)
            x.border=Border(top=sd(bd) if r==r1 else None,bottom=sd(bd) if r==r2 else None,
                            left=sd(bd) if c==c1 else None,right=sd(bd) if c==c2 else None)
def canvas(ws,nr,nc=14):
    for r in range(1,nr):
        for c in range(1,nc): ws.cell(r,c).fill=fill(CANVAS)
def hdr(ws,row,labels,start=2):
    for j,t in enumerate(labels,start):
        c=ws.cell(row,j,t); c.font=F(8.5,True,WHITE); c.fill=fill(INK2)
        c.alignment=Cn if j>start else ind(1)
        c.border=Border(top=sd(INK2),bottom=sd(INK2),left=sd(INK2),right=sd(INK2))
def titre(ws,row,num,txt,sous=""):
    c=ws.cell(row,2,"%s  %s"%(num,txt)); c.font=F(11,True,INK); c.alignment=ind(0)
    if sous:
        s=ws.cell(row+1,2,sous); s.font=F(8.5,False,INK3); s.alignment=ind(0)
def saisie(ws,row,col,val,fmt):
    c=ws.cell(row,col,val); c.number_format=fmt; c.alignment=Cn
    c.font=F(11,True,ORANGE); c.fill=fill(ORANGESOFT)
    c.border=Border(*[sd(ORANGE)]*4)
    return c

B=base_reelle()
POSTES=["PERM","VAC","ACQ","AUTRE","SIEGE"]
wb=openpyxl.Workbook()

# ============================ ① BASE 2026 ============================
b=wb.active; b.title="Base 2026"; b.sheet_view.showGridLines=False
canvas(b,40); b.column_dimensions["A"].width=2
for col,w in zip("BCDEF",(34,15,15,15,17)): b.column_dimensions[col].width=w
b.cell(2,2,"LA BASE — telle qu'elle est lue dans AW_002_000002_000001 et AW_002_000004_000001").font=F(12,True,INK)
b.cell(3,2,"Aucune formule sur cette feuille : c'est l'extrait, pas un retraitement.").font=F(9,False,INK3)

box(b,5,2,10,5); hdr(b,5,["Volumes et recettes","2024","2025","2026"])
VOL=[("Effectifs",              "EFF",      NUM),
     ("Inscrits de l'annee",    "INSCRITS", NUM),
     ("Chiffre d'affaires",     "CA",       EUR0)]
for i,(lib,k,fmt) in enumerate(VOL):
    r=6+i; b.cell(r,2,lib).font=F(10); b.cell(r,2).alignment=ind(1)
    for j,ex in enumerate(EXERCICES):
        c=b.cell(r,3+j,B[k][ex]); c.number_format=fmt; c.alignment=R; c.font=F(10)
r=9; b.cell(r,2,"Chiffre d'affaires par eleve").font=F(10,True); b.cell(r,2).alignment=ind(1)
for j,ex in enumerate(EXERCICES):
    c=b.cell(r,3+j,B["CA"][ex]/B["EFF"][ex]); c.number_format=EUR0; c.alignment=R; c.font=F(10,True,BLUE)
b.cell(10,2,"strictement plat sur trois ans : la croissance est 100 % volume").font=F(8.5,False,INK3)
b.cell(10,2).alignment=ind(1)

box(b,12,2,20,6); hdr(b,12,["Postes de cout","2024","2025","2026","€ / eleve 2026"])
UNI={}
for i,k in enumerate(POSTES):
    r=13+i; b.cell(r,2,LIB[k]).font=F(10); b.cell(r,2).alignment=ind(1)
    for j,ex in enumerate(EXERCICES):
        c=b.cell(r,3+j,B[k][ex]); c.number_format=EUR0; c.alignment=R; c.font=F(10)
    u=b.cell(r,6,"=E%d/E$6"%r); u.number_format=EUR0; u.alignment=R; u.font=F(10,True,BLUE)
    UNI[k]="'Base 2026'!$F$%d"%r
r=18
b.cell(r,2,"Cout complet").font=F(10,True); b.cell(r,2).alignment=ind(1)
for j in range(3):
    c=b.cell(r,3+j,"=SUM(%s13:%s17)"%(chr(64+3+j),chr(64+3+j))); c.number_format=EUR0
    c.alignment=R; c.font=F(10,True)
b.cell(19,2,"EBITDA").font=F(10,True); b.cell(19,2).alignment=ind(1)
b.cell(20,2,"Marge").font=F(10,True); b.cell(20,2).alignment=ind(1)
for j in range(3):
    cl=chr(64+3+j)
    c=b.cell(19,3+j,"=%s8-%s18"%(cl,cl)); c.number_format=EUR0; c.alignment=R; c.font=F(10,True,BLUE)
    m=b.cell(20,3+j,"=%s19/%s8"%(cl,cl)); m.number_format=PCT2; m.alignment=R; m.font=F(10,True,BLUE)
b.cell(22,2,"CAF — cout d'acquisition par inscrit").font=F(10,True); b.cell(22,2).alignment=ind(1)
for j,ex in enumerate(EXERCICES):
    c=b.cell(22,3+j,B["ACQ"][ex]/B["INSCRITS"][ex]); c.number_format=EUR0; c.alignment=R; c.font=F(10,True,ORANGE)
b.cell(23,2,"+7 %/an : il faut depenser de plus en plus pour recruter un eleve").font=F(8.5,False,INK3)
b.cell(23,2).alignment=ind(1)
CAF="'Base 2026'!$E$22"; EFF26="'Base 2026'!$E$6"; INS26="'Base 2026'!$E$7"
CAEL26="'Base 2026'!$E$9"

# ============================ ② CADRAGE 2027 ============================
w=wb.create_sheet("Cadrage 2027"); w.sheet_view.showGridLines=False
canvas(w,80); w.column_dimensions["A"].width=2
for col,wd in zip("BCDEFGH",(36,14,13,14,14,14,13)): w.column_dimensions[col].width=wd
w.column_dimensions["I"].width=3
for col,wd in zip(("J","K"),(30,13)): w.column_dimensions[col].width=wd
w.cell(2,2,"CADRAGE 2027 — PILOTÉ PAR L'EBITDA").font=F(15,True,INK)
w.cell(3,2,"On ne construit plus le budget pour découvrir l'EBITDA. On saisit l'EBITDA, "
           "et on cherche ce qu'il faut faire pour l'atteindre.").font=F(9.5,False,INK2)
for c in range(2,9): w.cell(4,c).border=Border(bottom=sd(BLUE))

# ---- ① LE CONSTAT ----------------------------------------------------
titre(w,6,"①","LE CONSTAT — ce que le réseau apporte","La rentrée est engagée. La finance ne la négocie pas : elle en hérite.")
box(w,9,2,13,6); hdr(w,9,["","2026 (base)","hypothèse","2027 (constat)"])
w.cell(10,2,"Effectifs").font=F(10); w.cell(10,2).alignment=ind(1)
w.cell(11,2,"Inscrits de l'année").font=F(10); w.cell(11,2).alignment=ind(1)
w.cell(12,2,"Chiffre d'affaires par élève").font=F(10); w.cell(12,2).alignment=ind(1)
w.cell(13,2,"CHIFFRE D'AFFAIRES 2027").font=F(10,True); w.cell(13,2).alignment=ind(1)
for r,src,fmt in ((10,EFF26,NUM),(11,INS26,NUM),(12,CAEL26,EUR0)):
    c=w.cell(r,3,"="+src); c.number_format=fmt; c.alignment=R; c.font=F(10)
c=w.cell(13,3,"=C10*C12"); c.number_format=EUR0; c.alignment=R; c.font=F(10,True)
saisie(w,10,4,CONSTAT["croiss_eff"],'"+"0.0%')
w.cell(11,4,"taux d'entrée maintenu").font=F(8,False,INK3); w.cell(11,4).alignment=Cn
saisie(w,12,4,CONSTAT["prix"],'"+"0.0%')
c=w.cell(10,5,"=C10*(1+D10)"); c.number_format=NUM
c=w.cell(11,5,"=E10*C11/C10"); c.number_format=NUM
c=w.cell(12,5,"=C12*(1+D12)"); c.number_format=EUR0
c=w.cell(13,5,"=E10*E12"); c.number_format=EUR0; c.font=F(11,True,BLUE)
for r in range(10,14):
    w.cell(r,5).alignment=R
    if not w.cell(r,5).font.b: w.cell(r,5).font=F(10,True)
w.cell(14,2,"L'historique ne montre aucun effet prix : 7 418 € par élève en 2024 comme en 2026.").font=F(8.5,False,INK3)
w.cell(14,2).alignment=ind(1)
CA_CONSTAT="$E$13"; EFF27="$E$10"; INS27="$E$11"

# ---- ② LA CIBLE ------------------------------------------------------
titre(w,17,"②","LA CIBLE — la seule chose que la direction saisit")
box(w,19,2,20,6,PANEL2)
w.cell(19,2,"Marge d'EBITDA visée en 2027").font=F(10,True); w.cell(19,2).alignment=ind(1)
saisie(w,19,4,CIBLE_MARGE,PCT1)
w.cell(20,2,"EBITDA cible").font=F(10,True); w.cell(20,2).alignment=ind(1)
c=w.cell(20,5,"=D19*%s"%CA_CONSTAT); c.number_format=EUR0; c.alignment=R; c.font=F(12,True,BLUE)
w.cell(19,5,"=\"pour mémoire 2026 : \"&TEXT('Base 2026'!$E$20,\"0,00 %\")").font=F(8.5,False,INK3)
w.cell(19,5).alignment=R
CIBLE="$E$20"

# ---- ③ LE TENDANCIEL -------------------------------------------------
titre(w,23,"③","LE TENDANCIEL — le budget au fil de l'eau",
      "Coûts unitaires 2026 reconduits, indexés de leur inflation propre, appliqués au volume 2027. Aucune décision de gestion.")
box(w,26,2,33,8); hdr(w,26,["Poste","€ / unité 2026","indexation","€ / unité 2027","volume 2027","montant 2027","% du CA"])
IDX={"PERM":TENDANCIEL["nao"],"VAC":TENDANCIEL["taux_vac"],"ACQ":TENDANCIEL["derive_caf"],
     "AUTRE":TENDANCIEL["inflation"],"SIEGE":TENDANCIEL["inflation"]}
NOTE={"PERM":"NAO","VAC":"taux horaire","ACQ":"dérive CAF","AUTRE":"inflation","SIEGE":"inflation"}
for i,k in enumerate(POSTES):
    r=27+i
    w.cell(r,2,LIB[k]).font=F(10); w.cell(r,2).alignment=ind(1)
    c=w.cell(r,3,"="+(CAF if k=="ACQ" else UNI[k])); c.number_format=EUR0; c.alignment=R; c.font=F(10)
    saisie(w,r,4,IDX[k],'"+"0.0%')
    c=w.cell(r,5,"=C%d*(1+D%d)"%(r,r)); c.number_format=EUR0; c.alignment=R; c.font=F(10)
    c=w.cell(r,6,"=%s"%(INS27 if k=="ACQ" else EFF27)); c.number_format=NUM; c.alignment=R; c.font=F(9,False,INK3)
    c=w.cell(r,7,"=E%d*F%d"%(r,r)); c.number_format=EUR0; c.alignment=R; c.font=F(10,True)
    c=w.cell(r,8,"=G%d/%s"%(r,CA_CONSTAT)); c.number_format=PCT1; c.alignment=R; c.font=F(9,False,INK3)
w.cell(32,2,"Coût complet").font=F(10,True); w.cell(32,2).alignment=ind(1)
c=w.cell(32,7,"=SUM(G27:G31)"); c.number_format=EUR0; c.alignment=R; c.font=F(10,True)
c=w.cell(32,8,"=G32/%s"%CA_CONSTAT); c.number_format=PCT1; c.alignment=R; c.font=F(9,False,INK3)
w.cell(33,2,"EBITDA TENDANCIEL").font=F(10,True); w.cell(33,2).alignment=ind(1)
c=w.cell(33,7,"=%s-G32"%CA_CONSTAT); c.number_format=EUR0; c.alignment=R; c.font=F(12,True,CRIT)
c=w.cell(33,8,"=G33/%s"%CA_CONSTAT); c.number_format=PCT2; c.alignment=R; c.font=F(11,True,CRIT)
w.cell(34,2,"La marge 2026 était de 16,65 %. Sans décision, elle se dégrade : le levier "
            "d'exploitation qui portait 2024-2026 ne se reconduit pas tout seul.").font=F(8.5,False,INK3)
w.cell(34,2).alignment=ind(1)
TEND="$G$33"

# ---- ④ L'ÉCART -------------------------------------------------------
box(w,37,2,38,8,BLUESOFT,BLUE)
w.cell(37,2,"④  L'ÉCART À COMBLER").font=F(11,True,INK); w.cell(37,2).alignment=ind(0)
w.cell(38,2,"=\"EBITDA cible \"&TEXT(\"\"&%s,\"# ##0 €\")&\"  −  tendanciel \"&TEXT(%s,\"# ##0 €\")"%(CIBLE,TEND)).font=F(9,False,INK2)
w.cell(38,2).alignment=ind(0)
c=w.cell(37,7,"=%s-%s"%(CIBLE,TEND)); c.number_format=EUR0; c.alignment=R; c.font=F(16,True,ORANGE)
w.cell(38,7,"à trouver dans les coûts").font=F(8.5,False,INK3); w.cell(38,7).alignment=R
ECART="$G$37"

# ---- ⑤ LES LEVIERS ---------------------------------------------------
titre(w,41,"⑤","LES LEVIERS — les seules décisions de gestion",
      "Chacun est assis sur une grandeur réelle. Les effets se cumulent en cascade : leur somme est exacte, sans terme croisé.")
box(w,44,2,52,8); hdr(w,44,["Levier","assiette 2027","réglage","effet EBITDA","en points","% de l'écart"])
LEV=[("nao",      "Politique salariale — points retirés à la NAO",  SCENARIO["nao"],        '"−"0.0" pt"'),
     ("encadrement","Encadrement — % d'élèves en plus par ETP",     SCENARIO["encadrement"],'"+"0.0%'),
     ("heures_vac","Vacataires — % d'heures en moins par élève",    SCENARIO["heures_vac"], '"−"0.0%'),
     ("caf",      "Acquisition — % de CAF en moins par inscrit",    SCENARIO["caf"],        '"−"0.0%'),
     ("siege",    "Siège — % de coût en moins par élève",           SCENARIO["siege"],      '"−"0.0%'),
     ("prix",     "Grille tarifaire — revalorisation",              SCENARIO["prix"],       '"+"0.00%')]
# assiettes et effets, en cascade sur la masse salariale
ASSIETTE={"nao":"=G27","encadrement":"=G27*(1-(D27-D45)/(1+D27))","heures_vac":"=G28",
          "caf":"=G29","siege":"=G31","prix":"="+CA_CONSTAT}
EFFET={"nao":        "=G27-C27*(1+D27-D45)*F27",
       "encadrement":"=E46*D46",
       "heures_vac": "=E47*D47",
       "caf":        "=E48*D48",
       "siege":      "=E49*D49",
       "prix":       "=E50*D50"}
for i,(k,lib,val,fmt) in enumerate(LEV):
    r=45+i
    w.cell(r,2,lib).font=F(10); w.cell(r,2).alignment=ind(1)
    c=w.cell(r,5,ASSIETTE[k]); c.number_format=EUR0; c.alignment=R; c.font=F(9,False,INK3)
    saisie(w,r,4,val,fmt)
    c=w.cell(r,6,EFFET[k]); c.number_format=EUR0; c.alignment=R; c.font=F(10,True,GOOD)
    c=w.cell(r,7,"=F%d/%s"%(r,CA_CONSTAT)); c.number_format='0.00" pt"'; c.alignment=R; c.font=F(10)
    c=w.cell(r,8,"=F%d/%s"%(r,ECART)); c.number_format='0%'; c.alignment=R; c.font=F(9,False,INK3)
w.cell(51,2,"TOTAL TROUVÉ").font=F(10,True); w.cell(51,2).alignment=ind(1)
c=w.cell(51,6,"=SUM(F45:F50)"); c.number_format=EUR0; c.alignment=R; c.font=F(11,True,GOOD)
c=w.cell(51,7,"=F51/%s"%CA_CONSTAT); c.number_format='0.00" pt"'; c.alignment=R; c.font=F(10,True)
c=w.cell(51,8,"=F51/%s"%ECART); c.number_format='0%'; c.alignment=R; c.font=F(10,True)
w.conditional_formatting.add("F45:F50",
    DataBarRule(start_type="num",start_value=0,end_type="max",color="FF"+GOOD,showValue=True))
w.cell(52,2,"L'acquisition ne rapporte que 0,02 pt par point de CAF : le cadrage se joue "
            "sur la masse salariale, pas sur le marketing.").font=F(8.5,False,INK3)
w.cell(52,2).alignment=ind(1)

# ---- ⑥ LE SOLDE ------------------------------------------------------
box(w,55,2,56,8,PANEL,INK2)
w.cell(55,2,"⑥  LE SOLDE").font=F(11,True,INK); w.cell(55,2).alignment=ind(0)
w.cell(56,2,"écart à combler − total trouvé. Le cadrage est bouclé quand il tombe à zéro.").font=F(9,False,INK2)
w.cell(56,2).alignment=ind(0)
c=w.cell(55,7,"=%s-F51"%ECART); c.number_format='#,##0" €";−#,##0" €";"0 €"'
c.alignment=R; c.font=F(16,True,INK)
w.cell(56,7,'=IF(ABS(G55)<10000,"cadrage bouclé",IF(G55>0,"il manque encore","au-delà de la cible"))').font=F(8.5,True,INK3)
w.cell(56,7).alignment=R
w.conditional_formatting.add("G55",CellIsRule(operator="between",formula=["-10000","10000"],font=Font(name=UI,size=16,bold=True,color=GOOD)))
w.conditional_formatting.add("G55",CellIsRule(operator="greaterThan",formula=["10000"],font=Font(name=UI,size=16,bold=True,color=CRIT)))

# ---- ⑦ COMPTE D'EXPLOITATION RETENU ----------------------------------
titre(w,59,"⑦","LE COMPTE D'EXPLOITATION RETENU")
box(w,62,2,70,7); hdr(w,62,["","2026 réel","tendanciel 2027","retenu 2027","Δ vs tendanciel"])
RET={"PERM":"=G27-F45-F46","VAC":"=G28-F47","ACQ":"=G29-F48","AUTRE":"=G30","SIEGE":"=G31-F49"}
w.cell(63,2,"Chiffre d'affaires").font=F(10,True); w.cell(63,2).alignment=ind(1)
w.cell(63,3,"='Base 2026'!E8").number_format=EUR0
w.cell(63,4,"="+CA_CONSTAT).number_format=EUR0
w.cell(63,5,"=%s+F50"%CA_CONSTAT).number_format=EUR0
for i,k in enumerate(POSTES):
    r=64+i
    w.cell(r,2,LIB[k]).font=F(10); w.cell(r,2).alignment=ind(1)
    w.cell(r,3,"='Base 2026'!E%d"%(13+i)).number_format=EUR0
    w.cell(r,4,"=G%d"%(27+i)).number_format=EUR0
    w.cell(r,5,RET[k]).number_format=EUR0
w.cell(69,2,"Coût complet").font=F(10,True); w.cell(69,2).alignment=ind(1)
for col in "CDE":
    w.cell(69,ord(col)-64,"=SUM(%s64:%s68)"%(col,col)).number_format=EUR0
w.cell(70,2,"EBITDA").font=F(10,True); w.cell(70,2).alignment=ind(1)
for col in "CDE":
    j=ord(col)-64
    w.cell(70,j,"=%s63-%s69"%(col,col)).number_format=EUR0
    w.cell(71,j,"=%s70/%s63"%(col,col)).number_format=PCT2
w.cell(71,2,"Marge").font=F(10,True); w.cell(71,2).alignment=ind(1)
for r in range(63,72):
    for j in range(3,7):
        c=w.cell(r,j); c.alignment=R
        if not c.font.b: c.font=F(10)
    if r in (69,70,71):
        for j in range(3,7): w.cell(r,j).font=F(10,True,BLUE if r>69 else INK)
    d=w.cell(r,6,"=E%d-D%d"%(r,r) if r!=71 else "=(E71-D71)*100")
    d.number_format=DEUR if r!=71 else '+0.00" pt";−0.00" pt";"—"'
    d.alignment=R; d.font=F(10,True,GOOD if r in (63,70,71) else INK2)
box(w,62,2,71,7)

# ---- lecture rapide, colonne de droite -------------------------------
box(w,6,10,34,11,PANEL2)
w.cell(6,10,"CE QUE LE CADRAGE DIT").font=F(10,True,INK); w.cell(6,10).alignment=ind(1)
NOTES=[("Le CA n'est pas un levier.","Les effectifs de la rentrée sont connus quand le cadrage "
        "s'ouvre. Le CA est un constat ; seul le prix reste discutable."),
       ("Le fil de l'eau détruit de la marge.","Coûts unitaires reconduits et indexés sur un "
        "volume en hausse : 16,65 % → 14,69 %. L'inaction a un coût chiffré."),
       ("Le poids décide de la puissance.","La masse salariale fait 47 % des coûts, "
        "l'acquisition 2 %. Un point sur l'une vaut vingt points sur l'autre."),
       ("Le siège ne se dilue pas.","1 132 € par élève en 2024, 1 146 € en 2026 : il croît "
        "plus vite que le réseau qu'il sert."),
       ("Le prix n'a jamais bougé.","Zéro revalorisation en trois ans. C'est le levier le "
        "plus puissant et le seul jamais actionné.")]
r=8
for t,d in NOTES:
    c=w.cell(r,10,t); c.font=F(9.5,True,BLUE); c.alignment=TOP
    c=w.cell(r+1,10,d); c.font=F(8.5,False,INK2); c.alignment=TOP
    w.merge_cells(start_row=r+1,start_column=10,end_row=r+3,end_column=11)
    w.row_dimensions[r+1].height=13; w.row_dimensions[r+2].height=13; w.row_dimensions[r+3].height=13
    r+=5
for rr in (6,17,23,37,41,55,59): w.row_dimensions[rr].height=20
w.row_dimensions[2].height=24
w.freeze_panes="A6"
wb.save("CADRAGE_EBITDA_2027.xlsx")
print("CADRAGE_EBITDA_2027.xlsx écrit — 2 feuilles")
