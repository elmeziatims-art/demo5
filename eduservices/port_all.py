#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Complete les onglets PIL et ALLOC du DESIGN SUR PLACE avec le contenu du
prototype (Pilotage, 3_Allocation) : 4 cles des B5, couts, maille, synthese, cap.
+ onglets de calcul _CALC_* + 00_Cartographie + feeds baseline + colonnes live.
Ne deplace rien (structures alignees). Briques Tagetik + _TGK_HIDDEN intactes."""
import warnings;warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.utils import get_column_letter as GCL
from tgk_surgery import Book

PROTO="CAD_SAAD_LIVE.xlsx"
P=openpyxl.load_workbook(PROTO)
Pd=openpyxl.load_workbook(PROTO,data_only=True)
ADD_TABS=["_CALC_MOTEUR","_CALC_PNL","_CALC_ALLOC","00_Cartographie"]
INPLACE={"Pilotage":"PIL","3_Allocation":"ALLOC"}   # prototype -> onglet DESIGN
FEEDS={"Socle":29,"Campagne":14,"Moteur":13,"Compta":6,"PNL":9,"Allocation":20}
RAWW={"Socle":29,"Campagne":14,"Moteur":13,"Compta":6,"PNL":7,"Allocation":20}

def remap(f):
    return (f.replace("'3_Allocation'!","ALLOC!").replace("3_Allocation!","ALLOC!")
             .replace("'Pilotage'!","PIL!").replace("Pilotage!","PIL!"))
def put(sh,ref,v):
    if isinstance(v,str) and v.startswith("="): sh.set_formula(ref,remap(v[1:]))
    elif isinstance(v,str): sh.set_text(ref,v)
    elif isinstance(v,bool): sh.set_number(ref,1 if v else 0)
    else: sh.set_number(ref,v)
def port_cells(pws,sh):
    for row in pws.iter_rows():
        for c in row:
            if c.value is not None: put(sh,c.coordinate,c.value)
    for rr,dim in pws.row_dimensions.items():
        if dim.outline_level: sh.set_row_outline(rr,dim.outline_level,bool(dim.hidden))

def build(src,out,embed_baseline=False):
    b=Book(src)
    # 1) noms repointes
    d={}
    LEV={"ACQ_BUD":18,"BRAND_BUD":19,"PRICE":20,"CONV_LEAD":21,"CONV_ADM":22,"PASSAGE":23,
         "INFL_EXT":25,"SALARY":26,"FTE_PERM":27,"PRODUCTIVITY":28,"STRUCT_COST":29,"FEE":31}
    for p,r in LEV.items():
        d["HYP_%s_V01"%p]="cad!$C$%d"%r; d["HYP_%s_V02"%p]="cad!$D$%d"%r; d["HYP_%s_V03"%p]="cad!$E$%d"%r
    for m,r in {"MBWAY":9,"ISCOM":10,"IPAC":11,"PIGIER":12,"TUNON":13}.items():
        d["HYP_PRICE_COEF_%s"%m]="cad!$K$%d"%r
    d["TEC_PL"]="cad!$F$3"; d["TEC_EBITDA"]="cad!$F$4"
    d["SCENARIO_ACTIF"]="cad!$H$3"; d["SCENARIO_CODE"]="cad!$N$1"
    d["ALLOC_GRP_BRAND"]="ALLOC!$C$5"; d["ALLOC_GRP_MARQUE"]="ALLOC!$C$6"
    d["ALLOC_BRAND_CAMP"]="ALLOC!$C$7"; d["ALLOC_CAMP_CLASS"]="ALLOC!$C$8"
    d["HYP_CAP_RETENU"]="PIL!$M$13:$M$26"; d["BUD_REF_CAP"]="PIL!$N$13:$N$26"
    b.add_names(d)
    # 2) cellules scenario dans cad
    cad=b.sheet("cad")
    cad.set_text("G3","Scenario actif :"); cad.set_text("H3","Cadrage")
    cad.set_formula("N1",'IF(H3="Optimiste","V02",IF(H3="Prudent","V03","V01"))')
    # 3) onglets de calcul ajoutes
    for s in ADD_TABS:
        port_cells(P[s],b.add_sheet(s))
    # 4) PIL & ALLOC completes SUR PLACE (contenu prototype, remap)
    for proto,dest in INPLACE.items():
        port_cells(P[proto],b.sheet(dest))
    # 5) feeds baseline + colonnes live
    if embed_baseline:
        for feed,rw in RAWW.items():
            pv=Pd[feed]; fsh=b.sheet(feed)
            for r in range(2,pv.max_row+1):
                if pv.cell(r,1).value in (None,""): continue
                for c in range(1,rw+1):
                    v=pv.cell(r,c).value
                    if v is not None: put(fsh,"%s%d"%(GCL(c),r),v)
    for feed in FEEDS:
        pws=P[feed]; fsh=b.sheet(feed); livecols=set()
        for r in range(2,pws.max_row+1):
            for c in range(1,pws.max_column+1):
                v=pws.cell(r,c).value
                if isinstance(v,str) and v.startswith("="): livecols.add(c)
        for c in sorted(livecols):
            h=pws.cell(1,c).value
            if h is not None: put(fsh,"%s1"%GCL(c),h)
            for r in range(2,pws.max_row+1):
                v=pws.cell(r,c).value
                if v is not None: put(fsh,"%s%d"%(GCL(c),r),v)
    b.set_fullcalc(); b.save(out)
    return len(d)

n=build("DESIGN.xlsm","DESIGN_FULL.xlsm",embed_baseline=True)
build("NAV.xlsx","NAV_FULL.xlsx",embed_baseline=False)
print("OK. zones nommees:",n)
