#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""MAQUETTE_CADRAGE_REVU.xlsx — reproduit FIDELEMENT le layout Cadrage/Pilotage
existant, avec la seule bascule : on saisit un OBJECTIF EBITDA (plus un objectif
CA), le CA devient un constat (evolution vs reference). Les cellules changees
sont reperees par un point (●). Construit = valeurs reelles de l'export V01."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E3EFEA"; WHITE="FFFFFF"
SOFT="51606D"; FAINT="7D8B98"; RULE="C8D2DA"; YEL="FFF6DA"; YELB="E8B84B"
OCHRE="B3641C"; OK="1E7A55"; CHG="FCE9C8"  # surlignage cellule changee
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color=RULE); box=Border(left=thin,right=thin,top=thin,bottom=thin)
Ln=Alignment("left",vertical="center"); Rn=Alignment("right",vertical="center")
Cn=Alignment("center",vertical="center",wrap_text=True); LW=Alignment("left",vertical="center",wrap_text=True)
EUR='#,##0 "€";-#,##0 "€";"—"'; PCT='0.0%'; PCT2='+0.0%;-0.0%;0.0%'; NUM='#,##0'

REF_CA=23098985; REF_EB=3845790; REF_EFF=3114
CON_CA=24231704; CON_EB=4111498; CON_EFF=3176

wb=openpyxl.Workbook()

# ================= ONGLET 1 : CADRAGE REVU =================
ws=wb.active; ws.title="Cadrage (revu)"; ws.sheet_view.showGridLines=False
for col,w in zip("ABCDEFG",[3,26,16,16,16,14,15]): ws.column_dimensions[col].width=w

def band(r,txt,c1=1,c2=7,bg=TEALD,fg=WHITE,sz=12,h=24):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    cc=ws.cell(r,c1,txt); cc.font=F(sz,True,fg); cc.fill=fill(bg); cc.alignment=Ln
    ws.row_dimensions[r].height=h

band(1,"  CADRAGE 2027 — piloté par l'EBITDA  (bascule minimale de ton onglet)",sz=13,h=28)
ws.merge_cells("A2:G2")
ws.cell(2,1,"  ● = les seules cellules à changer vs ta version. Tout le reste est identique.").font=F(9,False,OCHRE,True)
ws.row_dimensions[2].height=18

# --- saisie ---
ws.cell(4,2,"● Objectif EBITDA (amélioration vs 2026) :").font=F(10,True,INK)
c=ws.cell(4,3,0.05); c.fill=fill(YEL); c.number_format=PCT; c.font=F(11,True,INK); c.alignment=Cn
c.border=Border(*[Side(style="medium",color=YELB)]*4)
ws.cell(4,4,"← remplace « Croissance CA cible »").font=F(8.5,False,FAINT,True)
ws.cell(5,2,"(la marge n'est plus une saisie — elle se déduit)").font=F(8.5,False,FAINT,True)

# --- réconciliation ---
band(7,"  Réconciliation — Référence · Objectif · Construit",bg=TEAL,sz=11,h=20)
hd=["","Indicateur","Référence 2026","Objectif","Construit 2027","Écart","Évol. / Statut"]
for j,h in enumerate(hd,1):
    if j==1: continue
    cc=ws.cell(8,j,h); cc.font=F(9,True,WHITE); cc.fill=fill(TEALD); cc.alignment=Cn if j>2 else Ln; cc.border=box
def cell(r,c,v,fmt=None,f=None,al=Rn,bd=True,bg=None):
    x=ws.cell(r,c,v)
    if fmt: x.number_format=fmt
    x.font=f or F(10); x.alignment=al
    if bd: x.border=box
    if bg: x.fill=fill(bg)
    return x

# EBITDA d'abord (l'ancre)
cell(9,2,"◆ EBITDA (€)  — l'objectif",f=F(10,True,TEALD),al=Ln)
cell(9,3,REF_EB,EUR,F(10,True))
cell(9,4,"=C9*(1+$C$4)",EUR,F(10,True,TEALD),bg=CHG)          # ● objectif = 2026 x (1+amél)
cell(9,5,CON_EB,EUR,F(10,True))
cell(9,6,"=E9-D9",EUR,F(10,True))
cell(9,7,'=IF(E9>=D9,"🟢 TENU","🔴 À COMBLER")',None,F(9.5,True),Cn)
for j in range(2,8):
    if j!=4: ws.cell(9,j).fill=fill(TEALBG)
ws.cell(9,4).fill=fill(CHG)
# CA : constat
cell(10,2,"Chiffre d'affaires (€)  — constat",al=Ln)
cell(10,3,REF_CA,EUR)
cell(10,4,"—",None,F(10,c=FAINT),Cn,bg=CHG)                   # ● plus de cible CA
cell(10,5,CON_CA,EUR)
cell(10,6,"=E10-C10",EUR,F(10),bg=CHG)                        # ● écart vs référence
cell(10,7,"=E10/C10-1",PCT2,F(10,True,OK),Rn,bg=CHG)          # ● évolution vs 2026
# Marge : constat
cell(11,2,"Marge EBITDA %  — constat",al=Ln)
cell(11,3,"=C9/C10",PCT,F(9.5,c=SOFT))
cell(11,4,"—",None,F(9.5,c=FAINT),Cn,bg=CHG)                  # ●
cell(11,5,"=E9/E10",PCT,F(9.5,c=SOFT))
cell(11,6,"=E11-C11",PCT2,F(9.5,c=SOFT),bg=CHG)               # ● vs référence
cell(11,7,"=E11-C11",PCT2,F(9.5,c=SOFT),Rn)
# Effectif : constat (déjà comme ça chez toi)
cell(12,2,"Effectif  — constat",al=Ln)
cell(12,3,REF_EFF,NUM)
cell(12,4,"—",None,F(9.5,c=FAINT),Cn)
cell(12,5,CON_EFF,NUM)
cell(12,6,"=E12-C12",NUM,F(9.5))
cell(12,7,"=E12/C12-1",PCT2,F(9.5,c=SOFT))

ws.merge_cells("B14:G15")
ws.cell(14,2,"Récap des changements : ①  C4 = objectif EBITDA en % (ex. 5%).  ②  D9 (EBITDA cible) = C9*(1+C4).  "
             "③  D10 vide + F10 =E10-C10 + G10 =E10/C10-1 (CA en constat).  ④  D11/D13 marge en constat. "
             "Le reste (colonne Construit, effectif) ne bouge pas.").font=F(8.5,False,FAINT,True)
ws.cell(14,2).alignment=LW; ws.row_dimensions[14].height=40

# ================= ONGLET 2 : PILOTAGE (voyant) =================
wp=wb.create_sheet("Pilotage (voyant)"); wp.sheet_view.showGridLines=False
for col,w in zip("ABCDEFG",[3,20,16,16,14,16,4]): wp.column_dimensions[col].width=w
wp.merge_cells("A1:F1"); wp.cell(1,1,"  PILOTAGE — déjà conforme : il montre l'EBITDA + le CA et sa croissance").font=F(12,True,WHITE); wp.cell(1,1).fill=fill(TEALD); wp.row_dimensions[1].height=26
# strip
lab=["EBITDA (après siège)","CA 2027","Marge EBITDA","Effectif","Croissance CA vs 2026"]
val=[CON_EB,CON_CA,None,CON_EFF,None]
fmt=[EUR,EUR,PCT,NUM,PCT2]
for i,(l,v,f) in enumerate(zip(lab,val,fmt)):
    c=2+i
    wp.cell(3,c,l).font=F(8.5,True,SOFT); wp.cell(3,c).alignment=Cn
    x=wp.cell(4,c, v if v is not None else ("=C4/C3" if i==2 else "=C3/%d-1"%REF_CA)); x.number_format=f
    x.font=F(13,True,TEALD if i==0 else INK); x.alignment=Cn; x.fill=fill(TEALBG if i==0 else WHITE); x.border=box
wp.cell(3,2).font=F(8.5,True,TEALD)
# voyant
wp.merge_cells("B6:F6")
v=('="EBITDA vs objectif : "&IF(C3>=\'Cadrage (revu)\'!D9,'
   '"🟢 tenu (+"&TEXT(C3-\'Cadrage (revu)\'!D9,"#,##0")&" €)",'
   '"🔴 à combler ("&TEXT(C3-\'Cadrage (revu)\'!D9,"#,##0")&" €)")')
cc=wp.cell(6,2,v); cc.font=F(11,True,OK); cc.alignment=Ln; cc.fill=fill(TEALBG)
wp.row_dimensions[6].height=24
wp.merge_cells("B8:F9")
wp.cell(8,2,"À ajouter côté Tagetik : UNE cellule voyant qui compare l'EBITDA construit à l'objectif saisi dans le Cadrage "
            "(Cadrage!D9). Optionnel : placer la tuile EBITDA en premier. Rien d'autre ne change sur le Pilotage.").font=F(8.5,False,FAINT,True)
wp.cell(8,2).alignment=LW; wp.row_dimensions[8].height=34

out="/home/user/demo5/eduservices/MAQUETTE_CADRAGE_REVU.xlsx"
wb.save(out); print("SAVED",out)
