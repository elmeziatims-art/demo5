#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""RAPPORT_PNL_DEPLIABLE.xlsx — le P&L qui se plie / déplie.
Groupement Excel natif (les +/-) : Bloc -> Marque -> Campus -> Compte.
Replié : on voit le P&L propre (CA, blocs, EBITDA, Résultat).
Déplié : on descend jusqu'au compte, par marque et campus, sur 2024/25/26.
Source : compta réelle AW_002_000004_000001. Libellés indicatifs (Tagetik = DESC_CONTO0)."""
import csv, glob, os
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E2EFEB"; CARD2="F5F7F9"
FAINT="7D8B98"; WHITE="FFFFFF"; OCHRE="B3641C"; RULE="C8D2DA"; SOFT="51606D"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
RGT=Alignment("right",vertical="center")
def LFT(ind=0): return Alignment("left",vertical="center",indent=ind)
thin=Side(style="thin",color=RULE); med=Side(style="medium",color=RULE)
EUR='#,##0;-#,##0;"-"'

# ---- référentiel comptes : compte -> (bloc, libellé) ----
ACC={
 '706':('CA','Prestations de formation'),
 '7062':('CA','Prestations de formation (contrats pro)'),
 '708':('CA','Produits des activités annexes'),
 '604':('DIRECT',"Achats d'études et prestations"),
 '6063':('DIRECT','Fournitures et petit équipement'),
 '621':('DIRECT','Personnel extérieur (intérim)'),
 '6231':('DIRECT','Annonces et insertions — acquisition'),
 '6411':('PERSO','Salaires'),'6413':('PERSO','Primes'),
 '6414':('PERSO','Indemnités'),'645':('PERSO','Charges sociales'),
 '613':('STRUCT','Locations'),'615':('STRUCT','Entretien et réparations'),
 '616':('STRUCT',"Primes d'assurance"),'6226':('STRUCT','Honoraires'),
 '6236':('STRUCT','Catalogues et imprimés — marque'),
 '625':('STRUCT','Déplacements et réceptions'),'626':('STRUCT','Frais postaux et télécom'),
 '6281':('STRUCT','Cotisations'),
 '6331':('IMPOT',"Taxe d'apprentissage"),'6333':('IMPOT','Formation prof. continue'),
 '63511':('IMPOT','Taxes foncières / CFE'),
 '6811':('DOTAT','Dotations aux amortissements'),
}
# bloc -> (titre, signe)  signe: +1 produit, -1 charge
BLOCS=[('CA',"CHIFFRE D'AFFAIRES",+1),('DIRECT','COÛTS DIRECTS',-1),
       ('PERSO','PERSONNEL',-1),('STRUCT','STRUCTURE',-1),('IMPOT','IMPÔTS & TAXES',-1)]
DOT=('DOTAT','DOTATIONS',-1)
BRAND={'MBWAY':'MBway','ISCOM':'Iscom','IPAC':'Ipac','PIGIER':'Pigier','TUNON':'Tunon','GRP':'Holding (siège)'}
BR_ORDER=['MBWAY','ISCOM','IPAC','PIGIER','TUNON','GRP']
CITY={'LYO':'Lyon','PAR':'Paris','BOR':'Bordeaux','NAN':'Nantes','MTP':'Montpellier',
      'REN':'Rennes','LIL':'Lille','TLS':'Toulouse'}
YEARS=[2024,2025,2026]

def campus_label(en):
    if en=='GRP': return 'Siège'
    p=en.split('_'); return f"{BRAND.get(p[0],p[0])} {CITY.get(p[1],p[1])}"

# ---- lecture compta ----
path=glob.glob('/home/user/demo5/eduservices/tgk_data/ext_*/AW_002_000004_000001.csv')[0]
# data[bloc][brand][entity][account][year] = signed amount
data=lambda:defaultdict(lambda:defaultdict(lambda:defaultdict(lambda:defaultdict(lambda:defaultdict(float)))))
D=data()
with open(path,encoding='utf-8') as f:
    rd=csv.reader(f,delimiter=';'); next(rd)
    for row in rd:
        en,acc,ex=row[0],row[1],row[2]
        if acc in ('TEC_EBITDA','TEC_PL') or acc not in ACC: continue
        try: y=int(ex)
        except: continue
        if y not in YEARS: continue
        amt=float(row[5].replace(',','.'))
        bloc,_=ACC[acc]; sign=dict((b[0],b[2]) for b in BLOCS+[DOT])[bloc]
        br=en.split('_')[0]
        D[bloc][br][en][acc][y]+=amt*sign

def yr_sum(node):
    """somme récursive par année d'un sous-arbre {..:{year:val}}"""
    tot=defaultdict(float)
    def walk(n):
        for k,v in n.items():
            if isinstance(v,dict): walk(v)
            else: tot[k]+=v
    walk(node); return tot

wb=openpyxl.Workbook(); ws=wb.active; ws.title="P&L dépliable"
ws.sheet_view.showGridLines=False
ws.sheet_properties.outlinePr.summaryBelow=False   # la ligne de synthèse est AU-DESSUS du détail

ws["A1"]="P&L DÉPLIABLE  ·  du groupe au compte"; ws["A1"].font=F(15,True,INK)
ws["A2"]="Replié = le P&L propre. Cliquez les + à gauche : Bloc → Marque → Campus → Compte. Source : compta réelle."; ws["A2"].font=F(9,False,TEALD)

hr=4
ws.cell(hr,1,"Poste").font=F(10,True,WHITE); ws.cell(hr,1).fill=fill(TEAL); ws.cell(hr,1).alignment=LFT()
for j,y in enumerate(YEARS):
    c=ws.cell(hr,2+j,y); c.font=F(10,True,WHITE); c.fill=fill(TEAL); c.alignment=RGT
ws.cell(hr,1).border=Border(bottom=med)
for j in range(len(YEARS)): ws.cell(hr,2+j).border=Border(bottom=med)

r=hr+1
def write(lbl,vals,lvl,bold=False,color=INK,bg=None,box=None,ind=None):
    global r
    ind = lvl if ind is None else ind
    c=ws.cell(r,1,lbl); c.font=F(10 if lvl<2 else 9,bold,color); c.alignment=LFT(ind)
    for j,y in enumerate(YEARS):
        cc=ws.cell(r,2+j, round(vals.get(y,0)))
        cc.number_format=EUR; cc.font=F(10 if lvl<2 else 9,bold,color); cc.alignment=RGT
    if bg:
        for j in range(4): ws.cell(r,1+j).fill=fill(bg)
    if box:
        for j in range(4): ws.cell(r,1+j).border=box
    if lvl>0:
        ws.row_dimensions[r].outline_level=min(lvl,3)
        ws.row_dimensions[r].hidden=True   # replié par défaut
    r+=1

# accumulateur pour EBITDA / Résultat
acc_ebitda=defaultdict(float); acc_res=defaultdict(float)

def emit_bloc(bloc,titre):
    tot=yr_sum(D[bloc])
    write(titre,tot,0,bold=True,color=TEALD,bg=TEALBG,box=Border(top=thin,bottom=thin))
    for br in BR_ORDER:
        if br not in D[bloc]: continue
        btot=yr_sum(D[bloc][br])
        if all(abs(v)<1 for v in btot.values()): continue
        write(BRAND[br],btot,1,bold=True,color=INK)
        for en in sorted(D[bloc][br]):
            etot=yr_sum(D[bloc][br][en])
            write(campus_label(en),etot,2,color=SOFT)
            for acc in sorted(D[bloc][br][en]):
                av={y:D[bloc][br][en][acc][y] for y in YEARS}
                _,lab=ACC[acc]
                write(f"{acc} · {lab}",av,3,color=FAINT)
    return tot

for bloc,titre,_ in BLOCS:
    t=emit_bloc(bloc,titre)
    for y in YEARS:
        acc_ebitda[y]+=t.get(y,0); acc_res[y]+=t.get(y,0)

# EBITDA
write("= EBITDA",acc_ebitda,0,bold=True,color=TEAL,bg=TEALBG,box=Border(top=med,bottom=med))
# Dotations
td=emit_bloc('DOTAT','DOTATIONS')
for y in YEARS: acc_res[y]+=td.get(y,0)
# Résultat
write("= RÉSULTAT",acc_res,0,bold=True,color=INK,bg=CARD2,box=Border(top=med,bottom=med))

r+=1
ws.cell(r,1,"Libellés de comptes indicatifs — Tagetik restitue les vrais (DESC_CONTO0). Holding = coûts de siège, sans CA propre.").font=F(8,False,FAINT,True)

ws.column_dimensions['A'].width=46
for col in ('B','C','D'): ws.column_dimensions[col].width=15
ws.freeze_panes="A5"

out="/home/user/demo5/eduservices/tagetik/RAPPORT_PNL_DEPLIABLE.xlsx"
wb.save(out); print("SAVED",out)
