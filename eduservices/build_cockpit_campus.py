#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""COCKPIT_CAMPUS_ISCOM_PAR.xlsx — ce qu'un directeur de campus veut voir EN
PREMIER (étape 2 : consulter). Iscom Paris, données réelles 2024-2026.
Maquette de référence à porter sur Tagetik."""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E3EFEA"; WHITE="FFFFFF"
SOFT="51606D"; FAINT="7D8B98"; RULE="C8D2DA"; OCHRE="B3641C"; OK="1E7A55"; RED="B23A3A"
NAVY="3D4F8F"; NAVYBG="E6E9F3"; CARD="FBFCFB"; L0="DCE7EE"; AMBER="F6ECDD"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color=RULE); med=Side(style="medium",color=RULE)
box=Border(left=thin,right=thin,top=thin,bottom=thin)
Ln=Alignment("left",vertical="center"); Rn=Alignment("right",vertical="center")
Cn=Alignment("center",vertical="center",wrap_text=True); LW=Alignment("left",vertical="center",wrap_text=True)
EUR='#,##0 "€";-#,##0 "€";"—"'; PCT='0.0%'; PCT0='0%'; NUM='#,##0'; NUM1='#,##0.0'
VARSTUD=363
D=json.load(open('/tmp/isc.json'))
def agg(Y):
    d=D[Y]; return dict(ca=sum(c['ca'] for c in d),propre=sum(c['propre'] for c in d),
        siege=sum(c['siege'] for c in d),net=sum(c['net'] for c in d),eff=sum(c['eff'] for c in d),
        cap=sum(c['vcl']*c['cap'] for c in d))
A={Y:agg(Y) for Y in D}

wb=openpyxl.Workbook(); ws=wb.active; ws.title="Cockpit campus"; ws.sheet_view.showGridLines=False
for col,w in zip("ABCDEFGHI",[22,11,9,13,13,13,13,11,15]): ws.column_dimensions[col].width=w

def band(r,txt,c1=1,c2=9,bg=TEALD,fg=WHITE,sz=12,h=24):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    cc=ws.cell(r,c1,txt); cc.font=F(sz,True,fg); cc.fill=fill(bg); cc.alignment=Ln
    ws.row_dimensions[r].height=h

band(1,"  COCKPIT DIRECTEUR DE CAMPUS — Iscom Paris   ·   ce que je consulte en premier",sz=13,h=28)
ws.merge_cells("A2:I2")
ws.cell(2,1,"  Atterrissage 2026. Ma question : où j'en suis, qu'est-ce qui va, qu'est-ce qui coince ?").font=F(9,False,SOFT,True)
ws.row_dimensions[2].height=18

# ① tuiles
band(4,"①  Mes chiffres clés  (atterrissage 2026)",bg=TEAL,sz=11)
a26=A['2026']; rempl=a26['eff']/a26['cap']
tiles=[("Effectif",a26['eff'],NUM,INK),("Remplissage moyen",rempl,PCT0,OCHRE if rempl<0.85 else OK),
       ("Chiffre d'affaires",a26['ca'],EUR,INK),("Ma contribution (avant siège)",a26['propre'],EUR,TEALD),
       ("Quote-part siège (imposée)",-a26['siege'],EUR,SOFT),("EBITDA net",a26['net'],EUR,OK)]
for i,(lab,v,fmt,col) in enumerate(tiles):
    c0=1+i*1 if False else 1+(i%3)*3; r0=5+(i//3)*2
    ws.cell(r0,c0,lab).font=F(8.5,True,SOFT); ws.cell(r0,c0).alignment=Ln
    x=ws.cell(r0+1,c0,v); x.number_format=fmt; x.font=F(15,True,col); x.alignment=Ln
    ws.merge_cells(start_row=r0,start_column=c0,end_row=r0,end_column=c0+2)
    ws.merge_cells(start_row=r0+1,start_column=c0,end_row=r0+1,end_column=c0+2)
ws.row_dimensions[6].height=22; ws.row_dimensions[8].height=22

# ② trajectoire
band(10,"②  Ma trajectoire  2024 → 2026  (est-ce que ça monte ?)",bg=TEAL,sz=11)
hd=["","2024","2025","2026","Évolution"]
for j,hh in enumerate(hd,1):
    c=ws.cell(11,j,hh); c.font=F(9,True,WHITE); c.fill=fill(TEALD); c.alignment=Cn if j>1 else Ln; c.border=box
metrics=[("Chiffre d'affaires","ca",EUR),("Effectif","eff",NUM),("Ma contribution","propre",EUR),("EBITDA net","net",EUR)]
r=12
for lab,key,fmt in metrics:
    ws.cell(r,1,lab).font=F(9.5,False,INK); ws.cell(r,1).alignment=Ln
    for j,Y in enumerate(('2024','2025','2026')):
        x=ws.cell(r,2+j,A[Y][key]); x.number_format=fmt; x.font=F(9.5); x.alignment=Rn; x.border=box
    x=ws.cell(r,5,f"=E{r}/B{r}-1" if False else A['2026'][key]/A['2024'][key]-1); x.number_format='+0.0%;-0.0%'
    x.font=F(9.5,True,OK); x.alignment=Rn; x.border=box
    ws.row_dimensions[r].height=17; r+=1
ch=LineChart(); ch.title="EBITDA net & contribution — trajectoire"; ch.height=5.5; ch.width=9; ch.style=2
ch.y_axis.numFmt='#,##0'; ch.x_axis.delete=False; ch.y_axis.delete=False
dat=Reference(ws,min_col=2,max_col=4,min_row=14,max_row=15)  # contribution + net
ch.add_data(dat,from_rows=True,titles_from_data=False)
ch.set_categories(Reference(ws,min_col=2,max_col=4,min_row=11,max_row=11))
ws.add_chart(ch,"F11")

# ③ P&L par classe
band(17,"③  Mon P&L par classe  (2026) — le détail que je scrute",bg=NAVY,sz=11)
cols=["Programme ▸ Année","Rempl.","Eff.","CA","Contribution","Coût complet","Marge complète","Pt mort","Signal"]
for j,hh in enumerate(cols,1):
    c=ws.cell(18,j,hh); c.font=F(8.5,True,WHITE); c.fill=fill(NAVY); c.alignment=Cn if j>1 else Ln; c.border=box
r=19
LAB={'B1':'B1','B2':'B2','B3':'B3','M1':'M1','M2':'M2'}
for c in sorted(D['2026'],key=lambda x:(x['prog'],x['an'])):
    rempl=c['eff']/(c['vcl']*c['cap']); marge_el=c['rev']-VARSTUD; pm=(c['full']/c['vcl'])/marge_el
    ws.cell(r,1,"%s %s (%s)"%(c['prog'],c['an'],'Init' if c['mod']=='INIT' else 'Alt')).font=F(9.5); ws.cell(r,1).alignment=Ln
    ws.cell(r,2,rempl).number_format=PCT0; ws.cell(r,2).font=F(9.5,False,RED if rempl>0.9 else INK)
    ws.cell(r,3,c['eff']).number_format=NUM
    ws.cell(r,4,c['ca']).number_format=EUR
    ws.cell(r,5,c['contribution']).number_format=EUR; ws.cell(r,5).font=F(9.5,False,OK)
    ws.cell(r,6,c['full']).number_format=EUR
    ws.cell(r,7,c['net']).number_format=EUR; ws.cell(r,7).font=F(9.5,True,RED if c['net']<0 else INK)
    ws.cell(r,8,pm).number_format=NUM1
    sig='🔴 piège' if c['net']<0 else ('⚠️ saturé' if rempl>0.9 else '🟢 sain')
    ws.cell(r,9,sig).font=F(9,True,RED if c['net']<0 else (OCHRE if rempl>0.9 else OK)); ws.cell(r,9).alignment=Cn
    for j in range(2,9): ws.cell(r,j).alignment=Rn if j<9 else Cn; ws.cell(r,j).border=box
    ws.cell(r,1).border=box
    ws.row_dimensions[r].height=17; r+=1
# total
tot=D['2026']; ws.cell(r,1,"TOTAL CAMPUS").font=F(9.5,True,TEALD); ws.cell(r,1).alignment=Ln
ws.cell(r,3,sum(c['eff'] for c in tot)).number_format=NUM; ws.cell(r,4,sum(c['ca'] for c in tot)).number_format=EUR
ws.cell(r,5,sum(c['contribution'] for c in tot)).number_format=EUR; ws.cell(r,6,sum(c['full'] for c in tot)).number_format=EUR
ws.cell(r,7,sum(c['net'] for c in tot)).number_format=EUR; ws.cell(r,7).font=F(9.5,True,TEALD)
for j in range(1,10): ws.cell(r,j).fill=fill(TEALBG); ws.cell(r,j).border=Border(top=med,bottom=med);
for j in (3,4,5,6,7): ws.cell(r,j).alignment=Rn
rtot=r

# ④ signaux
band(rtot+2,"④  Mes signaux — ce qui appelle une décision",bg=OCHRE,sz=11,fg=WHITE)
sigs=[("⚠️  Capacité tendue","MAS_COM M1 (92 %) et M2 (91 %) : quasi pleins. Un push marketing → ouverture d'une 4ᵉ classe à étudier."),
      ("🔴  Piège de la fermeture","BAC_COM B1 : marge complète −12 516 € (paraît déficitaire) MAIS contribution +585 144 €. NE PAS fermer — le fixe se réalloue."),
      ("🎯  Mon objectif","EBITDA net 2026 = 110 581 € (en hausse : 48k→84k→111k). Objectif = amélioration vs 2025.")]
r=rtot+3
for t,txt in sigs:
    ws.cell(r,1,t).font=F(9.5,True,INK); ws.cell(r,1).alignment=LW
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=9)
    ws.cell(r,3,txt).font=F(9,False,SOFT); ws.cell(r,3).alignment=LW
    ws.row_dimensions[r].height=30; r+=1

ws.merge_cells(start_row=r+1,start_column=1,end_row=r+2,end_column=9)
ws.cell(r+1,1,"C'est l'étape « CONSULTER » : le directeur voit sa situation, sa trajectoire et ses signaux. Ensuite → ② AJUSTER (conversion, mix) "
             "puis → ③ SIMULER (ouvrir MAS_COM M1 ? ne pas fermer BAC_COM B1). Maquette de référence à porter sur Tagetik.").font=F(8.5,False,FAINT,True)
ws.cell(r+1,1).alignment=LW; ws.row_dimensions[r+1].height=32

out="/home/user/demo5/eduservices/COCKPIT_CAMPUS_ISCOM_PAR.xlsx"
wb.save(out); print("SAVED",out)
print("2026: CA=%d contribution=%d siège=%d net=%d eff=%d rempl=%.0f%%"%(a26['ca'],a26['propre'],a26['siege'],a26['net'],a26['eff'],rempl*100))
