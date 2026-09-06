#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Rendu du nouveau cad (poste de commande CFO) + cap restructure, avec les
valeurs vivantes injectees pour l'apercu (fichier reel non modifie)."""
import openpyxl, shutil, matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
from openpyxl.utils import column_index_from_string as ci
TMP="/tmp/_prev.xlsx"; shutil.copy("CAD_SAAD_LIVE.xlsx",TMP)
wb=openpyxl.load_workbook(TMP)
cad=wb["cad"]
# injecter valeurs vivantes (depuis replique) pour l'apercu
inj={"E7":24120981,"D7":22544725*1.05,"F7":24120981-22544725*1.05,"G7":24120981/(22544725*1.05)-1,
 "E8":3875895,"D8":22544725*1.05*0.15,"F8":3875895-22544725*1.05*0.15,"G8":3875895/(22544725*1.05*0.15)-1,
 "C9":3291530/22544725,"E9":3875895/24120981,"G9":3875895/24120981-0.15,
 "E10":3175,"F10":3175-3036,"G10":3175/3036-1,"D12":0,"F12":0}
# ACTIF col H = V01 values
for r in range(16,31):
    if cad["E%d"%r].value is not None: cad["H%d"%r]=cad["E%d"%r].value
for k,v in inj.items(): cad[k]=v
wb.save(TMP)
wb=openpyxl.load_workbook(TMP)

def argb(c):
    f=c.fill
    if f and f.patternType=="solid" and f.fgColor and isinstance(f.fgColor.rgb,str):
        rgb=f.fgColor.rgb; return "#"+(rgb[2:] if len(rgb)==8 else rgb)
    return None
def fcol(c):
    x=c.font.color
    if x and x.rgb and isinstance(x.rgb,str): return "#"+(x.rgb[2:] if len(x.rgb)==8 else x.rgb)
    return "#000000"
def fmtv(c):
    v=c.value
    if v is None or (isinstance(v,str) and v.startswith("=")): return ""
    if isinstance(v,(int,float)):
        nf=c.number_format
        if "%" in nf: return ("%+.1f%%" if "+" in nf else "%.1f%%")%(v*100)
        if "0.00" in nf: return "%.2f"%v
        if "##0" in nf: return ("{:+,.0f}" if "+" in nf else "{:,.0f}").format(v).replace(","," ")
        return str(v)
    return str(v)

def render(ws,r0,r1,cols,colw,title,out):
    xs=[0]
    for c in cols: xs.append(xs[-1]+colw.get(c,1.4))
    W=xs[-1]; H=(r1-r0+1)*0.32
    fig,ax=plt.subplots(figsize=(W*0.92,H*0.92+0.8)); ax.set_xlim(0,W); ax.set_ylim(0,H+0.7); ax.axis("off")
    ax.add_patch(Rectangle((0,H+0.28),W,0.42,color="#1F3864"))
    ax.text(0.1,H+0.49,title,color="white",fontsize=12,fontweight="bold",va="center")
    for ri,r in enumerate(range(r0,r1+1)):
        y=H-(ri+1)*0.32
        for cx,c in enumerate(cols):
            cell=ws.cell(r,c); x=xs[cx]; w=xs[cx+1]-xs[cx]; bg=argb(cell)
            ax.add_patch(Rectangle((x,y),w,0.32,facecolor=bg if bg else "none",
                         edgecolor="#DDDDDD",lw=0.3))
            t=fmtv(cell)
            if t:
                ha="left" if cell.alignment.horizontal in(None,"left") else ("right" if cell.alignment.horizontal=="right" else "center")
                tx=x+0.05 if ha=="left" else (x+w-0.05 if ha=="right" else x+w/2)
                ax.text(tx,y+0.16,t[:40],color=fcol(cell),fontsize=7,
                        fontweight="bold" if cell.font.bold else "normal",ha=ha,va="center")
    plt.tight_layout(); plt.savefig(out,dpi=120,bbox_inches="tight"); print("wrote",out)

# cad B1:L33
colsC=[ci(x) for x in "BCDEFGHIJKL"]
cwC={ci("B"):5.4,ci("C"):1.0,ci("D"):1.7,ci("E"):1.6,ci("F"):1.6,ci("G"):1.6,ci("H"):1.6,
     ci("I"):0.3,ci("J"):3.0,ci("K"):1.7,ci("L"):1.6}
render(cad,1,30,colsC,cwC,"cad — POSTE DE COMMANDE CFO (Cadrage)","/tmp/prev_cad.png")

# cap visible only: D,E,F,M,Q  rows 11-26
pil=wb["Pilotage"]
colsP=[ci(x) for x in ["D","E","F","M","Q"]]
cwP={ci("D"):2.4,ci("E"):1.4,ci("F"):2.6,ci("M"):2.4,ci("Q"):2.8}
render(pil,11,26,colsP,cwP,"Pilotage — Cap strategique par campus (vue CFO : rejoue)","/tmp/prev_cap.png")
