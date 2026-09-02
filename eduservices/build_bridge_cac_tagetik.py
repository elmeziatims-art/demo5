#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""BRIDGE_CAC_TAGETIK.xlsx — restit réalisable sur Tagetik.
La matrice tire 4 chiffres bruts (dépenses acq & inscrits, 2025 & 2026) d'une vue.
TOUT le bridge (CAC, effet dépenses, effet volume, waterfall) se calcule À CÔTÉ,
100% en formules qui pointent l'onglet Données. Rien en dur."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E3EFEA"; WHITE="FFFFFF"; NAVY="1F3864"
SOFT="51606D"; FAINT="7D8B98"; RULE="C8D2DA"; OCHRE="B3641C"; OK="1E7A55"; RED="C0392B"; GREEN="1E7A55"; YEL="FFF6DA"
AR="Arial"
def F(sz=10,b=False,c=INK): return Font(name=AR,size=sz,bold=b,color=c)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color=RULE); box=Border(left=thin,right=thin,top=thin,bottom=thin)
Ln=Alignment("left",vertical="center"); Rn=Alignment("right",vertical="center"); Cn=Alignment("center",vertical="center",wrap_text=True)
EUR='#,##0" €"'; NUM='#,##0'; TOP=Alignment("left",vertical="top",wrap_text=True)

wb=openpyxl.Workbook()

# ===================== ONGLET 1 : DONNÉES (ce que la matrice Tagetik tire) =====================
wd=wb.active; wd.title="Données (matrice)"; wd.sheet_view.showGridLines=False
for col,w in zip("ABC",[30,14,14]): wd.column_dimensions[col].width=w
wd.merge_cells("A1:C1"); wd.cell(1,1,"  DONNÉES — matrice Tagetik (2 mesures × 2 exercices)").font=F(11,True,WHITE); wd.cell(1,1).fill=fill(TEALD); wd.row_dimensions[1].height=22
wd.cell(3,1,"Mesure").font=F(9,True,WHITE); wd.cell(3,2,"2025").font=F(9,True,WHITE); wd.cell(3,3,"2026").font=F(9,True,WHITE)
for c in (1,2,3): wd.cell(3,c).fill=fill(TEAL); wd.cell(3,c).alignment=Cn; wd.cell(3,c).border=box
wd.cell(4,1,"Dépenses acquisition (€)").font=F(10); wd.cell(4,2,394702).number_format=EUR; wd.cell(4,3,434174).number_format=EUR
wd.cell(5,1,"Inscrits (nouveaux)").font=F(10); wd.cell(5,2,1159).number_format=NUM; wd.cell(5,3,1229).number_format=NUM
for r in (4,5):
    for c in (1,2,3): wd.cell(r,c).border=box; wd.cell(r,c).alignment=Ln if c==1 else Rn
wd.merge_cells("A7:C10")
wd.cell(7,1,"Source des 2 mesures (au grain groupe, filtre EXERCICE = 2025 / 2026) :\n"
            "• Dépenses acquisition = compte 6231  (ou DEPENSE_ACQ / V_MOTEUR_CAL.SPEND_ACQ)\n"
            "• Inscrits (nouveaux)   = VOL_NEW      (ou V_MOTEUR_CAL.INSCRITS)\n"
            "→ Sous Tagetik : matrice 2 lignes × 2 colonnes. Rien d'autre à charger.").font=F(8.5,c=SOFT)
wd.cell(7,1).alignment=TOP

# raccourcis vers les cellules Données
DEP25="'Données (matrice)'!$B$4"; DEP26="'Données (matrice)'!$C$4"
INS25="'Données (matrice)'!$B$5"; INS26="'Données (matrice)'!$C$5"

# ===================== ONGLET 2 : MASQUE (calculs à côté + waterfall) =====================
ws=wb.create_sheet("Bridge CAC (masque)"); ws.sheet_view.showGridLines=False
for col,w in zip("ABCDEF",[22,13,13,13,13,3]): ws.column_dimensions[col].width=w
ws.merge_cells("A1:E1"); ws.cell(1,1,"  BRIDGE DU CAC — d'où vient la hausse 2025 → 2026").font=F(12,True,WHITE); ws.cell(1,1).fill=fill(NAVY); ws.row_dimensions[1].height=24

# --- bloc calculs (tout en formules sur Données) ---
ws.cell(3,1,"Calcul").font=F(9,True,WHITE); ws.cell(3,2,"Valeur").font=F(9,True,WHITE); ws.cell(3,3,"Formule").font=F(9,True,WHITE)
for c in (1,2,3): ws.cell(3,c).fill=fill(TEAL); ws.cell(3,c).alignment=Cn; ws.cell(3,c).border=box
calc=[
 ("CAC 2025",         "=%s/%s"%(DEP25,INS25),              "Dép.2025 / Inscrits 2025"),
 ("CAC 2026",         "=%s/%s"%(DEP26,INS26),              "Dép.2026 / Inscrits 2026"),
 ("+ Effet dépenses", "=(%s-%s)/%s"%(DEP26,DEP25,INS25),   "(Dép.2026 − Dép.2025) / Inscrits 2025"),
 ("− Effet volume",   "=%s/%s-%s/%s"%(DEP26,INS26,DEP26,INS25), "Dép.2026/Ins.2026 − Dép.2026/Ins.2025"),
]
for i,(lab,f,note) in enumerate(calc):
    r=4+i; ws.cell(r,1,lab).font=F(9.5,lab.startswith('CAC')); ws.cell(r,1).alignment=Ln; ws.cell(r,1).border=box
    x=ws.cell(r,2,f); x.number_format='#,##0.0" €"'; x.font=F(9.5,True,GREEN if '+' in lab else (RED if '−' in lab else INK)); x.alignment=Rn; x.border=box
    ws.cell(r,3,note).font=F(8,c=FAINT); ws.cell(r,3).alignment=Ln; ws.cell(r,3).border=box
# refs calcul
C25="$B$4"; C26="$B$5"; EDEP="$B$6"; EVOL="$B$7"   # dans cet onglet

# --- données waterfall (formules) : Socle | Ancre | Hausse | Baisse ---
hr=10
ws.cell(hr,1,"cat (waterfall)").font=F(8,c=FAINT)
for j,h in enumerate(["Étape","Socle","Ancre","Hausse","Baisse"],1): ws.cell(hr,j,h).font=F(8,True,FAINT)
rows=[
 ("CAC 2025",        "0",                 "=%s"%C25,   "0",              "0"),
 ("+ Effet dépenses","=%s"%C25,           "0",         "=%s"%EDEP,       "0"),
 ("− Effet volume",  "=%s"%C26,           "0",         "0",             "=%s+%s-%s"%(C25,EDEP,C26)),  # hauteur = |effet vol|
 ("CAC 2026",        "0",                 "=%s"%C26,   "0",              "0"),
]
for i,(cat,so,an,ha,ba) in enumerate(rows):
    r=hr+1+i
    ws.cell(r,1,cat).font=F(8,c=FAINT)
    ws.cell(r,2,so); ws.cell(r,3,an); ws.cell(r,4,ha); ws.cell(r,5,ba)
NR=len(rows)
cats=Reference(ws,min_col=1,max_col=1,min_row=hr+1,max_row=hr+NR)

ch=BarChart(); ch.type="col"; ch.grouping="stacked"; ch.overlap=100
ch.title="Évolution du CAC 2025 → 2026 (€/inscrit)"; ch.height=8.5; ch.width=15
ch.y_axis.numFmt='#,##0 "€"'; ch.legend=None; ch.x_axis.delete=False; ch.y_axis.delete=False; ch.gapWidth=40
for colx in (2,3,4,5):
    ch.add_data(Reference(ws,min_col=colx,max_col=colx,min_row=hr+1,max_row=hr+NR),titles_from_data=False)
s_socle,s_ancre,s_hausse,s_baisse=ch.series
s_socle.graphicalProperties.noFill=True; s_socle.graphicalProperties.line.noFill=True
for s,color in ((s_ancre,TEALD),(s_hausse,GREEN),(s_baisse,RED)):
    s.graphicalProperties.solidFill=color; s.graphicalProperties.line.solidFill=color
for s,fmt in ((s_ancre,'#,##0" €"'),(s_hausse,'"+ "#,##0" €"'),(s_baisse,'"− "#,##0" €"')):
    s.dLbls=DataLabelList(); s.dLbls.showVal=True; s.dLbls.numFmt=fmt; s.dLbls.dLblPos="ctr"
    s.dLbls.showSerName=False; s.dLbls.showCatName=False; s.dLbls.showLegendKey=False
ch.set_categories(cats)
ws.add_chart(ch,"D3")

ws.merge_cells("A16:E19")
ws.cell(16,1,"Réalisable sur Tagetik : la MATRICE ne charge que les 4 chiffres (onglet Données). Le CAC, les 2 effets et "
             "les colonnes du waterfall (Socle/Ancre/Hausse/Baisse) sont des CELLULES CALCULÉES à côté — comme tes SUMIFS. "
             "Le graphe lit ces 4 colonnes. Change une dépense ou un nb d'inscrits → tout se recale.").font=F(8.5,c=SOFT)
ws.cell(16,1).alignment=TOP

out="/home/user/demo5/eduservices/BRIDGE_CAC_TAGETIK.xlsx"
wb.save(out); print("SAVED",out)
