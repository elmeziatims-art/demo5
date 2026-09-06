#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Onglet '00_Cartographie' : dictionnaire where-used pour l'equipe Tagetik.
Pour chaque SAISIE -> code/membre Tagetik, stockage, colonne de calcul consommatrice,
vue impactee, regle. Genere a partir du scan des formules + chaines indirectes."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
wb=openpyxl.load_workbook("CAD_SAAD_LIVE.xlsx")
if "00_Cartographie" in wb.sheetnames: del wb["00_Cartographie"]
ws=wb.create_sheet("00_Cartographie", 1)
ws.sheet_view.showGridLines=False
def fill(h):return PatternFill("solid",fgColor=h)
NAVY="15406E";BLUE="1B5FA6";GREEN="2E7D42";GOLD="9A6B00";PURP="7030A0";WHITE="FFFFFF"
thin=Side(style="thin",color="D9D9D9");box=Border(thin,thin,thin,thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)

ws.merge_cells("A1:H1")
ws["A1"]="CARTOGRAPHIE DES SAISIES  →  CALCUL   (data dictionary Tagetik)"
ws["A1"].font=Font(size=15,bold=True,color=NAVY); ws["A1"].fill=fill("EAF0FA")
ws.merge_cells("A2:H2")
ws["A2"]="Pour chaque cellule de saisie : son membre Tagetik, où elle est stockée, et LA/LES colonne(s) de calcul qui la consomment (violet dans les onglets _CALC)."
ws["A2"].font=Font(size=9,italic=True,color="8A8FA0")
heads=["Famille","Paramètre (métier)","Cellule(s) Excel","Code / membre Tagetik","Stockage (table · version · entity)","Colonne calcul consommatrice","Vue / mesure impactée","Règle (résumé)"]
for i,h in enumerate(heads):
    c=ws.cell(3,1+i,h); c.font=Font(size=9,bold=True,color=WHITE); c.fill=fill(NAVY); c.alignment=CEN; c.border=box
ws.row_dimensions[3].height=30

R=[
# Famille, Param, cellules, code, stockage, colonne calc, vue, regle
("Revenus","Variation budget acquisition → leads payants","cad!E/F/G 16","HYP_ACQ_BUD","AW_002_000001 · V01/V02/V03 · GRP","_CALC_MOTEUR!V (ACQ) · _CALC_PNL!G","V_MOTEUR · V_BUDGET","NOUV via ((rejoué/réf)×(1+ACQ))^REND_ACQ ; compte 6231 ×(1+ACQ)"),
("Revenus","Variation budget marque → socle organique","cad!E/F/G 17","HYP_BRAND_BUD","AW_002_000001 · V01/V02/V03 · GRP","_CALC_MOTEUR!W (BRAND) · _CALC_PNL!H","V_MOTEUR · V_BUDGET","ORG_REF×(1+BRAND)^REND_BRAND ; compte 6236 ×(1+BRAND)"),
("Revenus","Hausse tarifaire (prix)","cad!E/F/G 18","HYP_PRICE","AW_002_000001 · V01/V02/V03 · GRP","_CALC_MOTEUR!X (PRICE)","V_MOTEUR","PRIX = REV_STUD ×(1+PRICE×COEF_marque)"),
("Revenus","Gain conversion Lead → Candidature","cad!E/F/G 19","HYP_CONV_LEAD","AW_002_000001 · V01/V02/V03 · GRP","_CALC_MOTEUR!Y (GLC)","V_MOTEUR","facteur (RLC + GLC)"),
("Revenus","Gain conversion Admis → Inscrit","cad!E/F/G 20","HYP_CONV_ADM","AW_002_000001 · V01/V02/V03 · GRP","_CALC_MOTEUR!Z (GCV)","V_MOTEUR","facteur (YLD + GCV)"),
("Revenus","Amélioration du taux de passage","cad!E/F/G 21","HYP_PASSAGE","AW_002_000001 · V01/V02/V03 · GRP","_CALC_MOTEUR!AA (PASS)","V_MOTEUR","EFFECTIF réinscrits = VOL_EFF_INF ×(PASSAGE+PASS)"),
("Revenus","Frais de dossier / nouvel inscrit","cad!E/F/G 30","HYP_FEE","AW_002_000001 · V01/V02/V03 · GRP","_CALC_MOTEUR!AB (FEE)","V_MOTEUR","CA += NOUVEAUX × FEE"),
("Revenus","Coefficient prix par marque (décision)","cad!K7:K11","HYP_PRICE_COEF","AW_002_000001 · GEN · <MARQUE>_REF","_CALC_MOTEUR!AC (PRICE_COEF)","V_MOTEUR","module l'effet prix par marque"),
("Coûts","Inflation des charges externes","cad!E/F/G 23","HYP_INFL_EXT","AW_002_000001 · V01/V02/V03 · GRP","_CALC_PNL!I (INFL)","V_BUDGET","structure/impôts/dotations ×(1+INFL)"),
("Coûts","Politique salariale (masse permanente)","cad!E/F/G 24","HYP_SALARY","AW_002_000001 · V01/V02/V03 · GRP","_CALC_PNL!J (SAL)","V_BUDGET","personnel ×(1+SAL)"),
("Coûts","Variation des effectifs permanents","cad!E/F/G 25","HYP_FTE_PERM","AW_002_000001 · V01/V02/V03 · GRP","_CALC_PNL!K (FTE)","V_BUDGET","personnel ×(1+FTE)"),
("Coûts","Effort de productivité (achats & structure)","cad!E/F/G 26","HYP_PRODUCTIVITY","AW_002_000001 · V01/V02/V03 · GRP","_CALC_PNL!L (PROD)","V_BUDGET","achats & structure ×(1-PROD)"),
("Coûts","Variation des coûts de structure","cad!E/F/G 27","HYP_STRUCT_COST","AW_002_000001 · V01/V02/V03 · GRP","_CALC_PNL!M (STRUCT)","V_BUDGET","structure ×(1+STRUCT)"),
("Cap","Cap retenu par campus (marque×ville)","Pilotage!M13:M26","HYP_CAP_RETENU (par ENTITY)","AW_002_000001 · V01 · <ENTITY>","→ Pilotage!Q (rejoué) → _CALC_MOTEUR!T (REJOUE) · U (BUD_REF)","V_CAP · V_MOTEUR","budget acquisition REJOUÉ (somme groupe constante) → alimente NOUV"),
("Cibles","Croissance CA cible","cad!F3","TEC_PL","AW_002_000001 · GEN","→ cad!D7 (réconciliation)","(pilotage)","CA cible = CA_réf ×(1+croissance)"),
("Cibles","Marge EBITDA cible","cad!H3","TEC_EBITDA","AW_002_000001 · GEN","→ cad!D8 (réconciliation)","(pilotage)","EBITDA cible = CA cible × marge"),
("Contexte","Scénario actif","cad!D3 → code cad!P1","(dimension Version)","—","→ synthèse Pilotage · réconciliation cad · rollup","toutes","filtre la version affichée (V01/V02/V03)"),
("Clés","Clé siège administratif → marque","3_Allocation!C5","ALLOC_GRP_BRAND","AW_002_000001 · GEN","_CALC_ALLOC!AG (D1M) · AH (D1G)","V_ALLOCATION","cascade holding, niveau 1 (groupe→marque)"),
("Clés","Clé pub de marque → marque","3_Allocation!C6","ALLOC_GRP_MARQUE","AW_002_000001 · GEN","_CALC_ALLOC!AP (D1M_K4) · AQ (D1G_K4)","V_ALLOCATION","cascade frais de marque (6236), niveau 1"),
("Clés","Clé marque → campus","3_Allocation!C7","ALLOC_BRAND_CAMP","AW_002_000001 · GEN","_CALC_ALLOC!AE (D2E) · AF (D2M)","V_ALLOCATION","niveau 2 (marque→campus)"),
("Clés","Clé campus → classe","3_Allocation!C8","ALLOC_CAMP_CLASS","AW_002_000001 · GEN","_CALC_ALLOC!AC (D3C) · AD (D3E)","V_ALLOCATION","niveau 3 (campus→classe) + structure campus"),
]
FAMCOL={"Revenus":"EAF3FC","Coûts":"EAF7EF","Cap":"F3E9FB","Cibles":"FFF6DE","Contexte":"F0F0F2","Clés":"FBF3DE"}
FAMTX={"Revenus":BLUE,"Coûts":GREEN,"Cap":PURP,"Cibles":GOLD,"Contexte":"555555","Clés":GOLD}
r=4
for row in R:
    fam=row[0]; bg=FAMCOL[fam]
    for i,val in enumerate(row):
        c=ws.cell(r,1+i,val)
        c.border=box; c.alignment=LEF; c.fill=fill(bg)
        if i==0: c.font=Font(size=9,bold=True,color=FAMTX[fam])
        elif i in (2,3,5): c.font=Font(size=8,bold=True,color=PURP if i==5 else NAVY)  # cellules/code/colonne calc
        else: c.font=Font(size=8,color="333333")
    ws.row_dimensions[r].height=30; r+=1
widths={"A":10,"B":30,"C":16,"D":20,"E":26,"F":30,"G":16,"H":38}
for c,w in widths.items(): ws.column_dimensions[c].width=w
ws.freeze_panes="A4"
wb.save("CAD_SAAD_LIVE.xlsx")
print("OK onglet 00_Cartographie cree (%d saisies documentees)."%len(R))
