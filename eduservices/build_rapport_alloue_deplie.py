#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""RAPPORT_ALLOUE_DEPLIABLE.xlsx — le P&L CHARGÉ, dépliable à tous les niveaux.
Une fois l'allocation faite, on lit la marge à n'importe quelle maille :
   Marque → Campus → Programme → Année → Modalité
CA réel par classe (socle CRM, foote à la compta), coûts chargés = coûts campus
+ quote-part holding, ventilés à l'EFFECTIF (la clé du cadrage : ALLOC_*=VOL_EFF).
EBITDA net foote au groupe (3 291 530 en 2026).

Le reveal fin : le coût chargé par élève est ~uniforme dans un campus, mais le CA
par élève varie selon le prix du programme et la modalité (alternance/initial).
Certaines classes ne couvrent plus leur coût une fois chargées."""
import csv, glob
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E2EFEB"; CARD2="F5F7F9"
FAINT="7D8B98"; WHITE="FFFFFF"; OCHRE="B3641C"; OCHRED="8A4A12"; OCHREBG="F6E8D8"
RULE="C8D2DA"; SOFT="51606D"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
RGT=Alignment("right",vertical="center")
def LFT(ind=0): return Alignment("left",vertical="center",indent=ind)
thin=Side(style="thin",color=RULE); med=Side(style="medium",color=RULE)
EUR='#,##0;-#,##0;"-"'; PCT='0.0%'; NUM='#,##0'

YEAR=2026
# ---- comptes -> bloc (EBITDA = tout sauf dotations) ----
def bloc(a):
    if a[:1]=='7': return 'CA'
    if a in ('604','6063','621','6231'): return 'DIRECT'
    if a in ('6411','6413','6414','645'): return 'PERSO'
    if a in ('613','615','616','6226','6236','625','626','6281'): return 'STRUCT'
    if a in ('6331','6333','63511'): return 'IMPOT'
    if a=='6811': return 'DOTAT'
    return 'X'
BRAND={'MBWAY':'MBway','ISCOM':'Iscom','IPAC':'Ipac','PIGIER':'Pigier','TUNON':'Tunon'}
BR_ORDER=['MBWAY','ISCOM','IPAC','PIGIER','TUNON']
CITY={'LYO':'Lyon','PAR':'Paris','BOR':'Bordeaux','NAN':'Nantes','MTP':'Montpellier',
      'REN':'Rennes','LIL':'Lille','TLS':'Toulouse'}
MODL={'ALT':'Alternance','INIT':'Initial'}
def campus_label(en): p=en.split('_'); return f"{CITY.get(p[1],p[1])}"

# ---- 1) compta : coût EBITDA propre par campus + pool holding ----
comp=glob.glob('/home/user/demo5/eduservices/tgk_data/ext_*/AW_002_000004_000001.csv')[0]
camp_cost=defaultdict(float)   # coût EBITDA (charges, positif) par campus hors holding
hold_pool=0.0
with open(comp,encoding='utf-8') as f:
    rd=csv.reader(f,delimiter=';'); next(rd)
    for row in rd:
        en,acc,ex=row[0],row[1],row[2]
        if acc in ('TEC_EBITDA','TEC_PL'): continue
        if ex!=str(YEAR): continue
        b=bloc(acc)
        if b in ('CA','DOTAT','X'): continue        # EBITDA : hors CA, hors dotations
        amt=float(row[5].replace(',','.'))
        if en=='GRP': hold_pool+=amt
        else: camp_cost[en]+=amt

# ---- 2) CRM : CA et effectif par classe ----
crm=glob.glob('/home/user/demo5/eduservices/tgk_data/ext_*/AW_002_000002_000001.csv')[0]
# cls[(en,prog,an,mod)] = {'ca':..,'eff':..}
cls=defaultdict(lambda:{'ca':0.0,'eff':0.0})
camp_eff=defaultdict(float); eff_group=0.0
with open(crm,encoding='utf-8') as f:
    rd=csv.reader(f,delimiter=';'); next(rd)
    for row in rd:
        if row[6]!=str(YEAR): continue
        en,prog,an,mod=row[2],row[3],row[4],row[5]
        g=lambda i: float(row[i].replace(',','.'))
        eff=g(12); new=g(10); rev_stud=g(15); rev_ins=g(16)
        ca=rev_stud*eff+rev_ins*new
        k=(en,prog,an,mod)
        cls[k]['ca']+=ca; cls[k]['eff']+=eff
        camp_eff[en]+=eff; eff_group+=eff

# ---- 3) coût chargé par classe : campus (par eff) + holding (par eff groupe) ----
def loaded_cost(en,eff):
    camp_share = camp_cost[en]*(eff/camp_eff[en]) if camp_eff[en] else 0
    hold_share = hold_pool*(eff/eff_group) if eff_group else 0
    return camp_share+hold_share

# noeud d'agrégation
def node(): return {'ca':0.0,'eff':0.0,'cost':0.0}
def add(n,ca,eff,cost): n['ca']+=ca; n['eff']+=eff; n['cost']+=cost

# arbre Marque>Campus>Programme>Annee>Modalite
tree=defaultdict(lambda:defaultdict(lambda:defaultdict(lambda:defaultdict(lambda:defaultdict(node)))))
for (en,prog,an,mod),v in cls.items():
    br=en.split('_')[0]
    cost=loaded_cost(en,v['eff'])
    add(tree[br][en][prog][an][mod], v['ca'], v['eff'], cost)

def roll(children):
    n=node()
    for c in children.values():
        cc = c if 'ca' in c else roll(c)
        add(n,cc['ca'],cc['eff'],cc['cost'])
    return n

# ---- écriture ----
wb=openpyxl.Workbook(); ws=wb.active; ws.title="P&L chargé dépliable"
ws.sheet_view.showGridLines=False
ws.sheet_properties.outlinePr.summaryBelow=False

ws["A1"]="P&L CHARGÉ — DÉPLIABLE À TOUS LES NIVEAUX  ·  2026"; ws["A1"].font=F(15,True,INK)
ws["A2"]="Après allocation : la marge à n'importe quelle maille. Coûts chargés ventilés à l'effectif (clé du cadrage). Marque → Campus → Programme → Année → Modalité."; ws["A2"].font=F(9,False,TEALD)

hr=4
heads=["Maille","CA","Effectif","Coût chargé","EBITDA net","Marge"]
for j,h in enumerate(heads,1):
    c=ws.cell(hr,j,h); c.font=F(10,True,WHITE); c.fill=fill(TEAL)
    c.alignment=LFT() if j==1 else RGT; c.border=Border(bottom=med)

r=hr+1
def wr(lbl,n,lvl,bold=False,color=INK,bg=None,box=None):
    global r
    ca,eff,cost=n['ca'],n['eff'],n['cost']; net=ca-cost
    sz=10 if lvl<=1 else 9
    ws.cell(r,1,lbl).font=F(sz,bold,color); ws.cell(r,1).alignment=LFT(lvl)
    ws.cell(r,2,round(ca)).number_format=EUR
    ws.cell(r,3,round(eff)).number_format=NUM
    ws.cell(r,4,-round(cost)).number_format=EUR
    ws.cell(r,5,round(net)).number_format=EUR
    ws.cell(r,6, net/ca if ca else 0).number_format=PCT
    negm = ca and net/ca<0.05
    ws.cell(r,5).font=F(sz,bold or lvl<=1,OCHRED if negm else (TEALD if net>0 else OCHRED))
    ws.cell(r,6).font=F(sz,bold or lvl<=1,OCHRED if negm else TEALD)
    for j in (2,3,4,5,6):
        ws.cell(r,j).alignment=RGT
        if ws.cell(r,j).font.color is None or j in (2,3,4): ws.cell(r,j).font=F(sz,bold,color)
    if bg:
        for j in range(1,7): ws.cell(r,j).fill=fill(bg)
    if box:
        for j in range(1,7): ws.cell(r,j).border=box
    if lvl>0:
        ws.row_dimensions[r].outline_level=min(lvl,4)
        ws.row_dimensions[r].hidden=True
    r+=1

grp=node()
for br in BR_ORDER:
    if br not in tree: continue
    nbr=roll(tree[br]); add(grp,nbr['ca'],nbr['eff'],nbr['cost'])
    wr(BRAND[br],nbr,0,bold=True,color=TEALD,bg=TEALBG,box=Border(top=thin,bottom=thin))
    for en in sorted(tree[br]):
        nen=roll(tree[br][en]); wr(campus_label(en),nen,1,bold=True,color=INK)
        for prog in sorted(tree[br][en]):
            npr=roll(tree[br][en][prog]); wr(prog,npr,2,color=SOFT)
            for an in sorted(tree[br][en][prog]):
                nan=roll(tree[br][en][prog][an]); wr(an,nan,3,color=FAINT)
                for mod in sorted(tree[br][en][prog][an]):
                    wr(MODL.get(mod,mod),tree[br][en][prog][an][mod],4,color=FAINT)

# ligne groupe
ws.cell(r,1,"GROUPE").font=F(11,True,TEAL); ws.cell(r,1).alignment=LFT()
netg=grp['ca']-grp['cost']
ws.cell(r,2,round(grp['ca'])).number_format=EUR; ws.cell(r,2).font=F(11,True,TEAL)
ws.cell(r,3,round(grp['eff'])).number_format=NUM; ws.cell(r,3).font=F(11,True,TEAL)
ws.cell(r,4,-round(grp['cost'])).number_format=EUR; ws.cell(r,4).font=F(11,True,TEAL)
ws.cell(r,5,round(netg)).number_format=EUR; ws.cell(r,5).font=F(11,True,TEAL)
ws.cell(r,6,netg/grp['ca']).number_format=PCT; ws.cell(r,6).font=F(11,True,TEAL)
for j in (2,3,4,5,6): ws.cell(r,j).alignment=RGT
for j in range(1,7): ws.cell(r,j).fill=fill(TEALBG); ws.cell(r,j).border=Border(top=med,bottom=med)
r+=2
ws.cell(r,1,f"EBITDA net groupe = {round(netg):,} €  (foote à la compta : 3 291 530). Coût chargé ventilé à l'effectif ({round(eff_group)} élèves).".replace(',',' ')).font=F(8,False,FAINT,True)
ws.cell(r+1,1,"Rouge = marge nette < 5 % : la classe ne couvre plus son coût chargé. Déplie une marque faible (Pigier/Tunon) pour voir quelles classes plombent.").font=F(8,True,OCHRE,True)

ws.column_dimensions['A'].width=34
for col,w in zip("BCDEF",[14,10,14,14,9]): ws.column_dimensions[col].width=w
ws.freeze_panes="A5"

out="/home/user/demo5/eduservices/tagetik/RAPPORT_ALLOUE_DEPLIABLE.xlsx"
wb.save(out); print("SAVED",out,"| EBITDA net groupe =",round(netg))
