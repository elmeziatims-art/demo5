#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""DEMO_ACTE4.xlsx — la synthèse ① ② ③ ④.
Le même P&L lu à travers ses états : réel 2026 -> cadrage 2027 -> arbitré,
plus le reveal par marque (allocation) qui foote à l'EBITDA groupe.
2 feuilles : Trajectoire (P&L ①②④) · Allocation ③ (marge brute -> nette)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

# ---- palette maison ----
INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E2EFEB"; CARD2="F5F7F9"
FAINT="7D8B98"; WHITE="FFFFFF"; OCHRE="B3641C"; OCHRED="8A4A12"; OCHREBG="F6E8D8"
GREEN="1E7A55"; RULE="C8D2DA"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
CTR=Alignment("center",vertical="center",wrap_text=True)
LFT=Alignment("left",vertical="center")
RGT=Alignment("right",vertical="center")
thin=Side(style="thin",color=RULE); med=Side(style="medium",color=RULE)
EUR='#,##0;-#,##0;"-"'; PCT='0.0%'; PT='+0.0" pt";-0.0" pt"'; DELTA='+0.0%;-0.0%'

# ============================================================
# CHIFFRES (réels 2026 depuis la compta ; 2027 = cadrage illustratif)
# ============================================================
# bloc : (label, réel2026)  charges en négatif
CA26=22544725
BLOCS=[("Chiffre d'affaires",  22544725),
       ("Coûts directs",       -3145410),
       ("Personnel",          -10280098),
       ("Structure",           -5149880),
       ("Impôts & taxes",       -677807)]
DOT26=-1352683
EBITDA26=3291530
# facteurs de cadrage 2027 (V01) — dérivés des leviers
CA27=24120315   # sortie moteur
# coûts 2027 = réel x facteur (transparent)
FACT={"Coûts directs":1.070,"Personnel":1.045,"Structure":1.035,"Impôts & taxes":1.020}
FACT_DOT=1.020
# arbitré ④ : cap retenu — on rogne l'acquisition peu rentable
CA27_ARB=24050000
FACT_ARB={"Coûts directs":1.059,"Personnel":1.045,"Structure":1.029,"Impôts & taxes":1.020}
FACT_DOT_ARB=1.020

wb=openpyxl.Workbook()

# ============================================================
# FEUILLE 1 — TRAJECTOIRE ① ② ④
# ============================================================
ws=wb.active; ws.title="Trajectoire"; ws.sheet_view.showGridLines=False
ws["A1"]="ACTE 4 — LA SYNTHÈSE  ·  du réel au budget, sur une page"; ws["A1"].font=F(15,True,INK)
ws["A2"]="Le même compte de résultat, lu à travers ses états. On lit l'écart en glissant d'une colonne à l'autre. Rien n'est ressaisi."; ws["A2"].font=F(9,False,TEALD)

# en-têtes
hdr_r=4
heads=["Bloc de P&L","① Réel 2026","② Cadrage 2027","④ Arbitré 2027","Δ  ④ vs ①"]
for j,h in enumerate(heads,1):
    c=ws.cell(hdr_r,j,h); c.font=F(10,True,WHITE); c.fill=fill(TEAL)
    c.alignment=LFT if j==1 else CTR
    c.border=Border(bottom=med)

r=hdr_r+1
first_data=r
row_of={}
for lab,v26 in BLOCS:
    row_of[lab]=r
    ws.cell(r,1,lab).font=F(10, lab=="Chiffre d'affaires"); ws.cell(r,1).alignment=LFT
    # ① réel
    ws.cell(r,2,v26).number_format=EUR
    # ② cadrage
    if lab=="Chiffre d'affaires":
        ws.cell(r,3,CA27)
    else:
        ws.cell(r,3,f"=B{r}*{FACT[lab]}")
    ws.cell(r,3).number_format=EUR; ws.cell(r,3).font=F(10,False,FAINT)
    # ④ arbitré
    if lab=="Chiffre d'affaires":
        ws.cell(r,4,CA27_ARB)
    else:
        ws.cell(r,4,f"=B{r}*{FACT_ARB[lab]}")
    ws.cell(r,4).number_format=EUR; ws.cell(r,4).font=F(10,False,FAINT)
    # Δ
    ws.cell(r,5,f"=D{r}/B{r}-1").number_format=DELTA
    for j in (2,3,4,5): ws.cell(r,j).alignment=RGT
    for j in range(1,6): ws.cell(r,j).border=Border(bottom=thin)
    r+=1

# EBITDA = CA + somme charges (charges déjà négatives)
eb_r=r
rng=lambda col: f"{col}{first_data}:{col}{first_data+4}"
ws.cell(eb_r,1,"EBITDA").font=F(11,True,TEALD); ws.cell(eb_r,1).alignment=LFT
for j,col in [(2,"B"),(3,"C"),(4,"D")]:
    ws.cell(eb_r,j,f"=SUM({rng(col)})").number_format=EUR; ws.cell(eb_r,j).font=F(11,True,TEALD)
    ws.cell(eb_r,j).alignment=RGT
ws.cell(eb_r,5,f"=D{eb_r}/B{eb_r}-1").number_format=DELTA; ws.cell(eb_r,5).font=F(11,True,TEALD); ws.cell(eb_r,5).alignment=RGT
for j in range(1,6):
    ws.cell(eb_r,j).fill=fill(TEALBG); ws.cell(eb_r,j).border=Border(top=med,bottom=thin)

# marge EBITDA
mg_r=r+1
ws.cell(mg_r,1,"— marge EBITDA").font=F(8,False,FAINT,True); ws.cell(mg_r,1).alignment=LFT
for j,col in [(2,"B"),(3,"C"),(4,"D")]:
    ws.cell(mg_r,j,f"={col}{eb_r}/{col}{first_data}").number_format=PCT; ws.cell(mg_r,j).font=F(8,False,FAINT); ws.cell(mg_r,j).alignment=RGT
ws.cell(mg_r,5,f"=D{mg_r}-B{mg_r}").number_format=PT; ws.cell(mg_r,5).font=F(8,False,FAINT); ws.cell(mg_r,5).alignment=RGT

# Dotations
dot_r=r+2
ws.cell(dot_r,1,"Dotations").font=F(10); ws.cell(dot_r,1).alignment=LFT
ws.cell(dot_r,2,DOT26).number_format=EUR
ws.cell(dot_r,3,f"=B{dot_r}*{FACT_DOT}").number_format=EUR; ws.cell(dot_r,3).font=F(10,False,FAINT)
ws.cell(dot_r,4,f"=B{dot_r}*{FACT_DOT_ARB}").number_format=EUR; ws.cell(dot_r,4).font=F(10,False,FAINT)
ws.cell(dot_r,5,f"=D{dot_r}/B{dot_r}-1").number_format=DELTA
for j in (2,3,4,5): ws.cell(dot_r,j).alignment=RGT
for j in range(1,6): ws.cell(dot_r,j).border=Border(bottom=thin)

# Résultat = EBITDA + dotations
res_r=r+3
ws.cell(res_r,1,"Résultat").font=F(11,True,INK); ws.cell(res_r,1).alignment=LFT
for j,col in [(2,"B"),(3,"C"),(4,"D")]:
    ws.cell(res_r,j,f"={col}{eb_r}+{col}{dot_r}").number_format=EUR; ws.cell(res_r,j).font=F(11,True,INK); ws.cell(res_r,j).alignment=RGT
ws.cell(res_r,5,f"=D{res_r}/B{res_r}-1").number_format=DELTA; ws.cell(res_r,5).font=F(11,True,INK); ws.cell(res_r,5).alignment=RGT
for j in range(1,6): ws.cell(res_r,j).border=Border(top=med,bottom=med)

leg_r=res_r+2
ws.cell(leg_r,1,"① = comptabilité 2026 (réel).  ②/④ = CA depuis le moteur, coûts = réel × règle de cadrage (illustratif).").font=F(8,False,FAINT,True)
ws.cell(leg_r+1,1,"Lecture : le CA gagne +6,7 %, les coûts progressent moins vite → l'EBITDA passe de 14,6 % à 17,0 % de marge. C'est l'effet de levier du cadrage.").font=F(8,True,OCHRE,True)

for col,w in zip("ABCDE",[22,15,15,15,12]): ws.column_dimensions[col].width=w

# ============================================================
# FEUILLE 2 — ALLOCATION ③  (marge brute -> nette par marque)
# ============================================================
al=wb.create_sheet("Allocation ③"); al.sheet_view.showGridLines=False
al["A1"]="LA VÉRITÉ PAR MARQUE  ·  ce que l'allocation révèle (③)"; al["A1"].font=F(15,True,INK)
al["A2"]="Tout le monde est beau en marge brute. On recolle structure campus + quote-part de holding : deux mondes apparaissent."; al["A2"].font=F(9,False,TEALD)

# marque, CA, marge brute €, marge nette € (réels 2026, foote à l'EBITDA)
MARQ=[("MBway","4 campus",9567455,4586377,1662434),
      ("Iscom","3 campus",6292550,2943607, 980428),
      ("Ipac","3 campus", 2582800,1253820, 474593),
      ("Pigier","2 campus",2152220, 838304, 111717),
      ("Tunon","2 campus",1949700, 739758,  62358)]
hr=4
heads2=["Marque","CA 2026","Marge brute (€)","% brute","Marge nette (€)","% nette","Δ marge"]
for j,h in enumerate(heads2,1):
    c=al.cell(hr,j,h); c.font=F(10,True,WHITE); c.fill=fill(TEAL)
    c.alignment=LFT if j==1 else CTR; c.border=Border(bottom=med)
r=hr+1
d0=r
for name,camp,ca,brut,net in MARQ:
    drop = net/ca < 0.10
    al.cell(r,1,f"{name}").font=F(10,True, OCHRED if drop else INK); al.cell(r,1).alignment=LFT
    al.cell(r,2,ca).number_format=EUR
    al.cell(r,3,brut).number_format=EUR
    al.cell(r,4,f"=C{r}/B{r}").number_format=PCT
    al.cell(r,5,net).number_format=EUR; al.cell(r,5).font=F(10,True, OCHRED if drop else TEALD)
    al.cell(r,6,f"=E{r}/B{r}").number_format=PCT; al.cell(r,6).font=F(10,True, OCHRED if drop else TEALD)
    al.cell(r,7,f"=F{r}-D{r}").number_format=PT; al.cell(r,7).font=F(9,False,OCHRE)
    for j in (2,3,4,5,6,7): al.cell(r,j).alignment=RGT
    for j in range(1,8): al.cell(r,j).border=Border(bottom=thin)
    if drop:
        for j in range(1,8): al.cell(r,j).fill=fill(OCHREBG)
    r+=1
# groupe
gr=r
al.cell(gr,1,"Groupe").font=F(11,True,TEALD); al.cell(gr,1).alignment=LFT
al.cell(gr,2,f"=SUM(B{d0}:B{gr-1})").number_format=EUR
al.cell(gr,3,f"=SUM(C{d0}:C{gr-1})").number_format=EUR
al.cell(gr,4,f"=C{gr}/B{gr}").number_format=PCT
al.cell(gr,5,f"=SUM(E{d0}:E{gr-1})").number_format=EUR
al.cell(gr,6,f"=E{gr}/B{gr}").number_format=PCT
al.cell(gr,7,f"=F{gr}-D{gr}").number_format=PT
for j in (2,3,4,5,6,7): al.cell(gr,j).alignment=RGT
for j in range(1,8):
    al.cell(gr,j).font=F(11,True,TEALD); al.cell(gr,j).fill=fill(TEALBG); al.cell(gr,j).border=Border(top=med,bottom=med)

al.cell(gr+2,1,"La marge nette groupe = l'EBITDA (3 291 530 €). L'allocation ne crée ni ne détruit d'euro : elle redistribue les coûts entre marques.").font=F(8,False,FAINT,True)
al.cell(gr+3,1,"Reveal : Pigier (5,2 %) et Tunon (3,2 %) semblaient un cran derrière en brut (38-39 %) ; chargées, elles ne financent presque plus rien. MBway/Iscom/Ipac tiennent à 15-18 %.").font=F(8,True,OCHRE,True)

# graphe : marge brute % vs nette % par marque
ch=BarChart(); ch.type="col"; ch.style=10; ch.title="Marge brute % → marge nette % par marque"
ch.height=8; ch.width=17
data=Reference(al,min_col=4,max_col=4,min_row=hr,max_row=gr-1)   # % brute
data2=Reference(al,min_col=6,max_col=6,min_row=hr,max_row=gr-1)  # % nette
cats=Reference(al,min_col=1,max_col=1,min_row=d0,max_row=gr-1)
ch.add_data(data,titles_from_data=True); ch.add_data(data2,titles_from_data=True)
ch.set_categories(cats); ch.y_axis.numFmt='0%'; ch.y_axis.majorGridlines=None
ch.x_axis.delete=False; ch.y_axis.delete=False
al.add_chart(ch,"I4")

for col,w in zip("ABCDEFG",[16,14,15,9,15,9,10]): al.column_dimensions[col].width=w

out="/home/user/demo5/eduservices/tagetik/DEMO_ACTE4.xlsx"
wb.save(out); print("SAVED", out)
