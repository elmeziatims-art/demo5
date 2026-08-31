#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""LE_MOTEUR_FIN.xlsx — le moteur d'acquisition DESCENDU jusqu'à la modalité.
TCD dépliable Marque▸Campus▸Programme▸Année▸Modalité. Au CAMPUS : élasticité +
Δ budget (le budget est décidé au campus). En dessous : la VENTILATION des leads
gagnés au prorata historique, convertis au taux de chaque programme, facturés à
son prix. Foote au total campus/groupe (écart 0 vs la maille campus)."""
import json
from collections import OrderedDict, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

D=json.load(open('/tmp/moteur_fin.json')); CS=D['camp']; MAIL=D['mail']; DELTA=D['D']
INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E2EFEB"; WHITE="FFFFFF"; RULE="C8D2DA"
SOFT="51606D"; OCHRE="B3641C"; OCHRED="8A4A12"; GREEN="1E7A55"; FAINT="7D8B98"; NAVY="3D4F8F"
L0BG="DCE7EE"; L1BG="EAF0F4"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
def LI(i): return Alignment("left",vertical="center",indent=i)
RGT=Alignment("right",vertical="center"); CTR=Alignment("center",vertical="center",wrap_text=True); LFT=Alignment("left",vertical="center")
med=Side(style="medium",color=RULE); thin=Side(style="thin",color=RULE)
EUR='#,##0 "€";-#,##0 "€";"-"'; NUM='#,##0'; N1='#,##0.0'; PCT='0.0%'; DEC3='0.000'
MQ={'MBWAY':'MBway','ISCOM':'Iscom','IPAC':'Ipac','PIGIER':'Pigier','TUNON':'Tunon'}
ORDER=['MBWAY','ISCOM','IPAC','PIGIER','TUNON']
CITY={'LYO':'Lyon','PAR':'Paris','BOR':'Bordeaux','NAN':'Nantes','MTP':'Montpellier','REN':'Rennes','LIL':'Lille','TLS':'Toulouse'}
MODL={'INIT':'Initial','ALT':'Alternance'}

# arbre marque>campus>prog>an>mod
tree=OrderedDict()
for m in ORDER:
    mm=MQ[m]; sub=[x for x in MAIL if x['en'].split('_')[0]==m]
    if not sub: continue
    tree[mm]=OrderedDict()
    for en in sorted(set(x['en'] for x in sub)):
        tree[mm][en]=OrderedDict()
        s2=[x for x in sub if x['en']==en]
        for prog in sorted(set(x['prog'] for x in s2)):
            tree[mm][en][prog]=OrderedDict()
            for an in sorted(set(x['an'] for x in s2 if x['prog']==prog)):
                tree[mm][en][prog][an]=OrderedDict()
                for x in [z for z in s2 if z['prog']==prog and z['an']==an]:
                    tree[mm][en][prog][an][x['mod']]=x
def agg(node):
    t=dict(dleads=0.,ins=0.,ca=0.)
    def rec(n):
        if 'ins' in n and 'prog' in n:
            for k in t: t[k]+=n[k];
            return
        for v in n.values(): rec(v)
    rec(node); return t

wb=openpyxl.Workbook(); ws=wb.active; ws.title="Moteur fin (par programme)"
ws.sheet_view.showGridLines=False
ws["A1"]="LE MOTEUR — descendu jusqu'à la modalité (effet d'un +%.0f%% d'acquisition)"%(DELTA*100); ws["A1"].font=F(15,True,INK)
ws["A2"]="MODE D'EMPLOI — au CAMPUS : élasticité + Δ budget (le budget est campus). En dessous : ventilation au prorata historique des leads."; ws["A2"].font=F(9,True,TEALD)
ws.cell(3,1,"Source"); ws.cell(3,1).font=F(9,True,OCHRED); ws.cell(3,2,"V_CAMPAGNES (élasticité) + socle AW_002_000002 (mix/conv/prix par maille)").font=F(9,False,INK)
ws.cell(4,1,"Dimensions"); ws.cell(4,1).font=F(9,True,OCHRED); ws.cell(4,2,"Marque ▸ Campus ▸ Programme ▸ Année ▸ Modalité").font=F(9,False,INK)
ws.cell(5,1,"Chaîne"); ws.cell(5,1).font=F(9,True,OCHRED); ws.cell(5,2,"Δleads campus × part leads maille × conversion maille × prix maille = CA gagné").font=F(9,False,SOFT,True)
for j in range(1,9): ws.cell(6,j).border=Border(bottom=med)
hr=7
head=["Marque ▸ Campus ▸ Programme ▸ Année ▸ Modalité","Élasticité","Δ budget acq","Leads gagnés","Conv.","CA/inscrit","Inscrits gagnés","CA gagné"]
for j,h in enumerate(head,1):
    c=ws.cell(hr,j,h); c.font=F(8.5,True,WHITE); c.fill=fill(TEAL); c.alignment=LFT if j==1 else CTR; c.border=Border(bottom=med)
r=hr+1
def line(r,label,lvl,el=None,dbud=None,dleads=None,conv=None,prix=None,ins=None,ca=None):
    ws.cell(r,1,label).font=F(9 if lvl<=1 else 8.5,lvl<=1); ws.cell(r,1).alignment=LI(lvl)
    if el is not None: c=ws.cell(r,2,round(el,3)); c.number_format=DEC3; c.font=F(9,True,TEALD)
    if dbud is not None: c=ws.cell(r,3,round(dbud)); c.number_format=EUR; c.font=F(9,False,OCHRE)
    if dleads is not None: c=ws.cell(r,4,dleads); c.number_format=N1
    if conv is not None: c=ws.cell(r,5,conv); c.number_format=PCT; c.font=F(8.5,False,SOFT)
    if prix is not None: c=ws.cell(r,6,round(prix)); c.number_format=EUR; c.font=F(8.5,False,SOFT)
    if ins is not None: c=ws.cell(r,7,ins); c.number_format=N1
    if ca is not None: c=ws.cell(r,8,round(ca)); c.number_format=EUR; c.font=F(9 if lvl<=1 else 8.5,lvl<=1,TEALD if lvl<=1 else INK)
    for j in range(2,9): ws.cell(r,j).alignment=RGT
    if lvl==0:
        for j in range(1,9): ws.cell(r,j).fill=fill(L0BG); ws.cell(r,j).border=Border(top=med,bottom=thin)
    elif lvl==1:
        for j in range(1,9): ws.cell(r,j).fill=fill(L1BG)
    else:
        for j in range(1,9): ws.cell(r,j).border=Border(bottom=thin)
    if lvl>0:
        ws.row_dimensions[r].outline_level=min(lvl,4)
        if lvl>=2: ws.row_dimensions[r].hidden=True

gd=dict(dleads=0.,ins=0.,ca=0.)
for mq,camps in tree.items():
    a=agg(camps)
    line(r,mq,0,dleads=a['dleads'],ins=a['ins'],ca=a['ca']); r+=1
    for en,progs in camps.items():
        ac=agg(progs); cs=CS[en]
        line(r,f"{CITY.get(en.split('_')[1],en.split('_')[1])}",1,el=cs['el'],dbud=cs['dbud'],dleads=ac['dleads'],
             conv=ac['ins']/ac['dleads'] if ac['dleads'] else None, prix=ac['ca']/ac['ins'] if ac['ins'] else None, ins=ac['ins'],ca=ac['ca']); r+=1
        for prog,ans in progs.items():
            ap=agg(ans); line(r,prog,2,dleads=ap['dleads'],conv=ap['ins']/ap['dleads'] if ap['dleads'] else None,prix=ap['ca']/ap['ins'] if ap['ins'] else None,ins=ap['ins'],ca=ap['ca']); r+=1
            for an,mods in ans.items():
                aa=agg(mods); line(r,an,3,dleads=aa['dleads'],conv=aa['ins']/aa['dleads'] if aa['dleads'] else None,prix=aa['ca']/aa['ins'] if aa['ins'] else None,ins=aa['ins'],ca=aa['ca']); r+=1
                for mod,x in mods.items():
                    line(r,MODL.get(mod,mod),4,dleads=x['dleads'],conv=x['conv'],prix=x['prix'],ins=x['ins'],ca=x['ca']); r+=1
    for k in gd: gd[k]+=a[k]
line(r,"GROUPE",0,dleads=gd['dleads'],ins=gd['ins'],ca=gd['ca'])
for j in range(1,9): ws.cell(r,j).font=F(11,True,TEAL); ws.cell(r,j).fill=fill(TEALBG); ws.cell(r,j).border=Border(top=med,bottom=med)
r+=2
ws.cell(r,1,"Vérif : la somme des CA gagnés par maille = l'effet campus agrégé, à l'euro près. Le programme n'est pas oublié — il est ventilé.").font=F(8,True,OCHRE,True)
ws.column_dimensions['A'].width=44
for col,w in zip("BCDEFGH",[10,12,11,8,10,12,12]): ws.column_dimensions[col].width=w
ws.freeze_panes=f"A{hr+1}"
out="/home/user/demo5/eduservices/tagetik/LE_MOTEUR_FIN.xlsx"
wb.save(out); print("SAVED",out,"  CA gagné groupe=",round(gd['ca']),"  inscrits=",round(gd['ins'],1))
