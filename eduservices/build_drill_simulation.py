#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""DRILL_SIMULATION.xlsx — à quoi ressemble le drill-through quand on clique.
① Cockpit (extrait)  : la cellule d'origine, marquée comme drillable
② Drill — Pourquoi   : le bridge que renvoie DRILL_EBITDA_POURQUOI.sql
③ Drill — Par compte : ce que renvoie DRILL_EBITDA_PAR_COMPTE.sql
Le bridge est en FORMULES sur les 8 agrégats de la requête : on voit qu'il boucle."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from openpyxl.utils import get_column_letter

BLUE="2A78D6"; BLUESOFT="DCE9F9"; ORANGE="EB6834"; ORANGESOFT="FBE3D8"
GOOD="0CA30C"; WARN="FAB219"; CRIT="D03B3B"
INK="131922"; INK2="4D5866"; INK3="7C8798"
CANVAS="EEF1F6"; PANEL="FFFFFF"; PANEL2="F7F9FC"; LINE="DFE4EC"; LINE2="EAEEF4"; WHITE="FFFFFF"
UI="Segoe UI"; MONO="Consolas"
def F(sz=10,b=False,c=INK,f=UI): return Font(name=f,size=sz,bold=b,color=c)
def fill(c): return PatternFill("solid",fgColor=c)
def sd(c=LINE,st="thin"): return Side(style=st,color=c)
L=Alignment("left",vertical="center"); R=Alignment("right",vertical="center")
Cn=Alignment("center",vertical="center"); TOP=Alignment("left",vertical="top",wrap_text=True)
EUR='#,##0" €"'; NUM='#,##0'; PCT='0.0%'; DPCT='"▲ "0.0%;"▼ "0.0%'
def box(ws,r1,c1,r2,c2,bg=PANEL,bd=LINE):
    for r in range(r1,r2+1):
        for c in range(c1,c2+1):
            x=ws.cell(r,c); x.fill=fill(bg)
            x.border=Border(top=sd(bd) if r==r1 else None,bottom=sd(bd) if r==r2 else None,
                            left=sd(bd) if c==c1 else None,right=sd(bd) if c==c2 else None)
def canvas(ws,nr=60,nc=16):
    for r in range(1,nr):
        for c in range(1,nc): ws.cell(r,c).fill=fill(CANVAS)

# ---- les 8 agrégats que la requête ramène (TUNON_PAR, REEL/V_FINAL/P12) ----
EFF_P,EFF_N     = 122, 124
CA_P,CA_N       = 1069253, 1102400
CVAR_P,CVAR_N   = 42944, 45012
CDIR_P,CDIR_N   = 919509, 961088

wb=openpyxl.Workbook()

# =====================================================================================
# ① COCKPIT (EXTRAIT) — la cellule d'origine
# =====================================================================================
w1=wb.active; w1.title="① Cockpit (extrait)"; w1.sheet_view.showGridLines=False
w1.column_dimensions["A"].width=2; w1.column_dimensions["B"].width=22
for c in "CDEFGHIJ": w1.column_dimensions[c].width=13
w1.column_dimensions["K"].width=42
canvas(w1)
for r,h in {1:8,2:26,3:16,4:10,5:18,6:20,7:20,8:20,9:20,10:20,12:10}.items(): w1.row_dimensions[r].height=h
box(w1,2,2,3,10)
w1.cell(2,2,"  Cockpit › Portefeuille — marque & campus").font=F(13,True,INK); w1.cell(2,2).alignment=L
w1.cell(3,2,"  exercice 2026 · les cellules encadrées en bleu sont des cellules de MATRICE : elles portent le contexte, donc elles sont drillables").font=F(9,c=INK3)
w1.cell(3,2).alignment=L

HD=["Marque / campus","CA 2026","Δ CA","EBITDA 2026","Δ EBITDA","Marge","Δ marge","Effectifs"]
box(w1,5,2,10,9)
for j,t in enumerate(HD,2):
    c=w1.cell(5,j,t); c.font=F(8.5,True,INK3); c.alignment=L if j==2 else R
    c.border=Border(bottom=sd(LINE)); c.fill=fill(PANEL2)
GRID=[("GROUPE",23098985,.0608,3845790,.1036,.1665,.6,2597,0),
      ("Tunon",1847900,.0191,73900,-.2144,.0400,-1.2,309,1),
      ("TUNON_PAR",1102400,.0310,96300,-.0982,.0874,-1.3,124,2),
      ("TUNON_LYO",745500,.0020,-22400,None,-.0300,-1.3,84,2)]
for i,(n,ca,dca,eb,deb,mg,dpt,eff,lvl) in enumerate(GRID):
    r=6+i
    a=w1.cell(r,2,n); a.font=F(10 if lvl==0 else 9.5,lvl<=1,INK if lvl<2 else INK2)
    a.alignment=Alignment("left",vertical="center",indent=lvl)
    vals=[(3,ca,NUM),(4,dca,DPCT),(5,eb,NUM),(6,deb,DPCT),(7,mg,PCT),(8,dpt,'"↗ +"0.0" pt";"↘ −"0.0" pt"'),(9,eff,NUM)]
    for j,v,fm in vals:
        c=w1.cell(r,j,"n/s" if v is None else v); c.number_format=fm; c.alignment=R
        c.font=F(9.5,lvl<=1,INK if lvl<2 else INK2,MONO)
        c.border=Border(bottom=sd(LINE2))
        if j in (4,6,8):                                  # colonnes CALCULÉES : non drillables
            c.fill=fill("F4F5F7")
        else:                                             # colonnes de MATRICE : drillables
            c.border=Border(bottom=sd(LINE2),left=sd(BLUESOFT),right=sd(BLUESOFT))
    w1.cell(r,2).border=Border(bottom=sd(LINE2))
# la cellule cliquée
tc=w1.cell(8,5)
tc.border=Border(top=sd(ORANGE,"medium"),bottom=sd(ORANGE,"medium"),left=sd(ORANGE,"medium"),right=sd(ORANGE,"medium"))
tc.fill=fill(ORANGESOFT); tc.font=F(10,True,INK,MONO)
w1.cell(8,11,"◀  clic droit ici › Drill through").font=F(10,True,ORANGE); w1.cell(8,11).alignment=L
w1.cell(7,11,"colonnes grisées = formules Excel : pas de contexte, pas de drill").font=F(8.5,c=INK3)
w1.cell(9,11,"le contexte transmis : REEL · V_FINAL · 12 · 2026 · TUNON_PAR").font=F(8.5,c=INK3)

box(w1,13,2,16,11,bg=PANEL2)
w1.merge_cells(start_row=13,start_column=2,end_row=16,end_column=11)
w1.cell(13,2,"  Pourquoi cette cellule et pas une autre.  L'audience regarde Tunon depuis le cockpit. TUNON_PAR fait +3,1 % de CA "
             "et −9,8 % d'EBITDA, pendant que le groupe fait +10,4 % : la contradiction est visible, la question se pose d'elle-même. "
             "Et son détail tient sur un écran — 6 étapes de bridge, 9 comptes. Un drill sur le CA groupe ouvrirait des milliers de "
             "lignes et la démonstration tomberait à plat.").font=F(9.5,c=INK2)
w1.cell(13,2).alignment=TOP
for r in range(13,17): w1.cell(r,2).border=Border(left=sd(ORANGE,"thick"))

# =====================================================================================
# ② DRILL — POURQUOI
# =====================================================================================
w2=wb.create_sheet("② Drill — Pourquoi"); w2.sheet_view.showGridLines=False
w2.column_dimensions["A"].width=2; w2.column_dimensions["B"].width=30
w2.column_dimensions["C"].width=15; w2.column_dimensions["D"].width=24; w2.column_dimensions["E"].width=46
for c in "FGHI": w2.column_dimensions[c].width=14
canvas(w2,64,12)
for r,h in {1:8,2:26,3:16,4:8,5:16,6:20,7:8,8:20,9:8,10:18,17:8,18:16}.items(): w2.row_dimensions[r].height=h
for r in range(11,17): w2.row_dimensions[r].height=19

box(w2,2,2,3,9,bg=BLUE,bd=BLUE)
w2.cell(2,2,"  Drill through  ›  Pourquoi l'EBITDA a-t-il varié ?").font=F(13,True,WHITE); w2.cell(2,2).alignment=L
w2.cell(3,2,"  DRILL_EBITDA_POURQUOI.sql — exécutée sur la cellule, paramètres hérités de son contexte").font=F(9,c="D6E6F8")
w2.cell(3,2).alignment=L
# contexte hérité
box(w2,5,2,6,9)
w2.cell(5,2,"  CONTEXTE HÉRITÉ DE LA CELLULE").font=F(8.5,True,INK3); w2.cell(5,2).alignment=L
for i,(k,v) in enumerate([(":SCENARIO","REEL"),(":VERSION","V_FINAL"),(":PERIODE","12"),
                          (":EXERCICE","2026"),(":ENTITY","TUNON_PAR")]):
    cc=3+i if i>0 else 3
    a=w2.cell(5,cc+ (0 if i else 0),k); a.font=F(8,c=INK3); a.alignment=Cn
    b=w2.cell(6,cc,v); b.font=F(10,True,BLUE,MONO); b.alignment=Cn
    b.fill=fill(BLUESOFT); b.border=Border(top=sd(BLUE),bottom=sd(BLUE),left=sd(BLUE),right=sd(BLUE))
for i,k in enumerate([":SCENARIO",":VERSION",":PERIODE",":EXERCICE",":ENTITY"]):
    a=w2.cell(5,3+i,k); a.font=F(8,c=INK3); a.alignment=Cn
w2.cell(8,2,"Cellule d'origine :  Cockpit › Portefeuille › TUNON_PAR › EBITDA 2026").font=F(10,True,INK)
w2.cell(8,2).alignment=L

# --- les 8 agrégats ramenés par la requête (bloc de droite)
box(w2,10,7,16,9)
w2.cell(10,7,"  Agrégats lus").font=F(8.5,True,INK3); w2.cell(10,7).alignment=L
w2.cell(10,8,"2025").font=F(8.5,True,INK3); w2.cell(10,8).alignment=R
w2.cell(10,9,"2026").font=F(8.5,True,INK3); w2.cell(10,9).alignment=R
for i,(lab,p,n,fm) in enumerate([("Effectifs",EFF_P,EFF_N,NUM),("CA",CA_P,CA_N,NUM),
                                 ("Coût variable",CVAR_P,CVAR_N,NUM),("Coûts directs",CDIR_P,CDIR_N,NUM),
                                 ("CA / élève",None,None,'#,##0'),("Coût var. / élève",None,None,'#,##0')]):
    r=11+i
    w2.cell(r,7,lab).font=F(9,c=INK2); w2.cell(r,7).alignment=L
    if p is None:
        f=("=H12/H11","=I12/I11") if lab.startswith("CA /") else ("=H13/H11","=I13/I11")
        w2.cell(r,8,f[0]); w2.cell(r,9,f[1])
    else:
        w2.cell(r,8,p); w2.cell(r,9,n)
    for j in (8,9):
        w2.cell(r,j).number_format=fm; w2.cell(r,j).font=F(9,c=INK2,f=MONO); w2.cell(r,j).alignment=R
        w2.cell(r,j).border=Border(bottom=sd(LINE2))

# --- le bridge, EN FORMULES sur ces agrégats
box(w2,10,2,17,5)
for j,t in zip((2,3,4,5),("Étape","Montant","Détail","Lecture")):
    c=w2.cell(10,j,t); c.font=F(8.5,True,INK3); c.alignment=L if j!=3 else R
    c.border=Border(bottom=sd(LINE)); c.fill=fill(PANEL2)
BR=[("EBITDA 2025","=H12-H13-H14","","point de départ",False),
    ("+ Effet effectifs","=(I11-H11)*(H15-H16)" ,'=H11&" → "&I11&" élèves"',
     "volume, valorisé à la marge variable 2025",True),
    ("+ Effet prix / mix","=(I15-H15)*I11",'=ROUND(H15,0)&" → "&ROUND(I15,0)&" € / élève"',
     "tarif, mix initiale/alternance, mix programmes",True),
    ("− Effet coût variable unitaire","=-(I16-H16)*I11",'=ROUND(H16,0)&" → "&ROUND(I16,0)&" € / élève"',
     "vacataires et achats directs, par élève",True),
    ("− Effet coûts directs","=-(I14-H14)",'=ROUND(H14/1000,0)&" → "&ROUND(I14/1000,0)&" k€"',
     "permanents et structure : ils ne suivent pas l'activité",True),
    ("EBITDA 2026","=I12-I13-I14",'="marge "&ROUND((I12-I13-I14)/I12*100,1)&" %"',
     "doit égaler la cellule cliquée",False)]
for i,(lab,mt,det,lec,eff) in enumerate(BR):
    r=11+i
    a=w2.cell(r,2,lab); a.font=F(10,not eff,INK); a.alignment=L
    c=w2.cell(r,3,mt); c.number_format=EUR; c.alignment=R; c.font=F(10,True,INK,MONO)
    d=w2.cell(r,4,det if det else None); d.font=F(9,c=INK3,f=MONO); d.alignment=R
    e=w2.cell(r,5,"  "+lec); e.font=F(9,c=INK3); e.alignment=L
    for j in range(2,6): w2.cell(r,j).border=Border(bottom=sd(LINE2))
    if not eff:
        for j in range(2,6): w2.cell(r,j).fill=fill(PANEL2)
w2.cell(16,3).font=F(11,True,INK,MONO)
# contrôle : le pont boucle
w2.cell(17,2,"  Contrôle du pont").font=F(9,True,INK3); w2.cell(17,2).alignment=L
ck=w2.cell(17,3,"=C11+C12+C13+C14+C15-C16"); ck.number_format='0.00" €"'; ck.alignment=R
ck.font=F(10,True,GOOD,MONO)
w2.cell(17,5,'="   écart "&ROUND(C17,2)&" € — le pont retombe exactement sur la cellule cliquée"').font=F(9,True,GOOD)
w2.cell(17,5).alignment=L
for rng in ("C12:C15",):
    w2.conditional_formatting.add(rng,CellIsRule(operator="greaterThan",formula=["0"],font=Font(name=MONO,size=10,bold=True,color=GOOD)))
    w2.conditional_formatting.add(rng,CellIsRule(operator="lessThan",formula=["0"],font=Font(name=MONO,size=10,bold=True,color=CRIT)))

# --- waterfall (colonnes techniques masquées en K:O)
w2.cell(10,11,"Étape").font=F(7,c=INK3)
for j,t in enumerate(["Étape","Socle","Ancre","Hausse","Baisse"],11): w2.cell(10,j,t).font=F(7,c=INK3)
WF=[("EBITDA 2025","0","=C11","0","0"),
    ("Effectifs","=C11","0","=C12","0"),
    ("Prix / mix","=C11+C12","0","=C13","0"),
    ("Coût var.","=C11+C12+C13+C14","0","0","=-C14"),
    ("Coûts directs","=C16","0","0","=-C15"),
    ("EBITDA 2026","0","=C16","0","0")]
for i,(cat,so,an,ha,ba) in enumerate(WF):
    r=11+i
    w2.cell(r,11,cat).font=F(7,c=INK3)
    for j,v in enumerate((so,an,ha,ba),12):
        w2.cell(r,j,v).number_format=NUM; w2.cell(r,j).font=F(7,c=INK3)
for col in "KLMNO": w2.column_dimensions[col].hidden=True
ch=BarChart(); ch.type="col"; ch.grouping="stacked"; ch.overlap=100; ch.gapWidth=45
ch.height=8.6; ch.width=19; ch.legend=None; ch.y_axis.numFmt='#,##0 "€"'
ch.y_axis.scaling.min=80000; ch.y_axis.scaling.max=150000; ch.y_axis.majorUnit=10000
ch.x_axis.delete=False; ch.y_axis.delete=False
ch.title="Pourquoi l'EBITDA de TUNON_PAR recule de 10 500 € : la croissance existe, les coûts fixes la mangent"
for col in (12,13,14,15):
    ch.add_data(Reference(w2,min_col=col,max_col=col,min_row=11,max_row=16),titles_from_data=False)
s_so,s_an,s_ha,s_ba=ch.series
s_so.graphicalProperties.noFill=True; s_so.graphicalProperties.line.noFill=True
for s,cc in ((s_an,BLUE),(s_ha,GOOD),(s_ba,CRIT)):
    s.graphicalProperties.solidFill=cc; s.graphicalProperties.line.solidFill=cc
for s,fm in ((s_an,'#,##0" €"'),(s_ha,'"+ "#,##0" €"'),(s_ba,'"− "#,##0" €"')):
    s.dLbls=DataLabelList(); s.dLbls.showVal=True; s.dLbls.numFmt=fm; s.dLbls.dLblPos="ctr"
    s.dLbls.showSerName=False; s.dLbls.showCatName=False; s.dLbls.showLegendKey=False
ch.set_categories(Reference(w2,min_col=11,max_col=11,min_row=11,max_row=16))
w2.add_chart(ch,"B19")

box(w2,38,2,41,9,bg=PANEL2)
w2.merge_cells(start_row=38,start_column=2,end_row=41,end_column=9)
w2.cell(38,2,"  Ce que le drill fait dire au chiffre.  Le campus a gagné 2 élèves et vendu plus cher : +16 825 € et +15 618 €, "
             "soit +32 443 € de croissance réelle. Mais ses coûts directs — permanents et structure — ont pris 41 579 €, "
             "soit +4,5 % quand le CA n'a fait que +3,1 %. Le campus n'a pas un problème de marché, il a un problème de "
             "structure de coûts. Sans ce drill, on aurait discuté du recrutement pendant une heure.").font=F(9.5,c=INK2)
w2.cell(38,2).alignment=TOP
for r in range(38,42): w2.cell(r,2).border=Border(left=sd(ORANGE,"thick"))

# =====================================================================================
# ③ DRILL — PAR COMPTE
# =====================================================================================
w3=wb.create_sheet("③ Drill — Par compte"); w3.sheet_view.showGridLines=False
w3.column_dimensions["A"].width=2; w3.column_dimensions["B"].width=11
w3.column_dimensions["C"].width=32
for c in "DEFG": w3.column_dimensions[c].width=15
w3.column_dimensions["H"].width=16
canvas(w3,40,10)
for r,h in {1:8,2:26,3:16,4:8,5:16,6:20,7:8,8:20,9:8,10:18}.items(): w3.row_dimensions[r].height=h
box(w3,2,2,3,8,bg=BLUE,bd=BLUE)
w3.cell(2,2,"  Drill through  ›  Et dans la compta, quels comptes ont bougé ?").font=F(13,True,WHITE)
w3.cell(2,2).alignment=L
w3.cell(3,2,"  DRILL_EBITDA_PAR_COMPTE.sql — même cellule, même contexte · trié par variation absolue décroissante").font=F(9,c="D6E6F8")
w3.cell(3,2).alignment=L
box(w3,5,2,6,8)
w3.cell(5,2,"  CONTEXTE HÉRITÉ").font=F(8.5,True,INK3); w3.cell(5,2).alignment=L
w3.cell(6,2,"  REEL · V_FINAL · 12 · 2026 · TUNON_PAR").font=F(10,True,BLUE,MONO); w3.cell(6,2).alignment=L
w3.cell(8,2,"Rapprochement : Σ produits − Σ charges = 96 300 € = la cellule").font=F(10,True,INK)
w3.cell(8,2).alignment=L

CPT=[("641","Masse salariale permanente",578000,612000,"charge"),
     ("7062","Scolarité alternance",619500,640000,"produit"),
     ("706","Scolarité initiale",419000,430000,"produit"),
     ("613","Loyers & charges locatives",203500,214000,"charge"),
     ("6236","Quote-part marque",98809,93788,"charge"),
     ("6063","Fournitures & petits équipements",39200,41300,"charge"),
     ("708","Produits annexes",30753,32400,"produit"),
     ("621","Vacataires",36900,38500,"charge"),
     ("604","Achats de prestations",6044,6512,"charge")]
CPT.sort(key=lambda x:-abs(x[3]-x[2]))
for j,t in enumerate(["Compte","Libellé","2025","2026","Variation","Var. %","Nature"],2):
    c=w3.cell(10,j,t); c.font=F(8.5,True,INK3); c.alignment=L if j in (2,3,8) else R
    c.border=Border(bottom=sd(LINE)); c.fill=fill(PANEL2)
box(w3,11,2,19,8)
for j,t in enumerate(["Compte","Libellé","2025","2026","Variation","Var. %","Nature"],2):
    c=w3.cell(10,j,t); c.font=F(8.5,True,INK3); c.alignment=L if j in (2,3,8) else R
    c.border=Border(bottom=sd(LINE)); c.fill=fill(PANEL2)
for i,(cp,lib,p,n,nat) in enumerate(CPT):
    r=11+i
    w3.cell(r,2,cp).font=F(9.5,True,INK,MONO); w3.cell(r,2).alignment=L
    w3.cell(r,3,lib).font=F(9.5,c=INK2); w3.cell(r,3).alignment=L
    w3.cell(r,4,p).number_format=NUM; w3.cell(r,5,n).number_format=NUM
    w3.cell(r,6,"=E%d-D%d"%(r,r)).number_format='+#,##0;−#,##0'
    w3.cell(r,7,"=IFERROR(F%d/ABS(D%d),0)"%(r,r)).number_format='"+ "0.0%;"− "0.0%'
    w3.cell(r,8,nat).font=F(8.5,c=INK3); w3.cell(r,8).alignment=L
    for j in (4,5,6,7): w3.cell(r,j).font=F(9.5,j==6,INK,MONO); w3.cell(r,j).alignment=R
    for j in range(2,9): w3.cell(r,j).border=Border(bottom=sd(LINE2))
r=20
w3.cell(r,3,"EBITDA = Σ produits − Σ charges").font=F(9.5,True)
w3.cell(r,4,"=SUMIF($H$11:$H$19,\"produit\",D11:D19)-SUMIF($H$11:$H$19,\"charge\",D11:D19)").number_format=NUM
w3.cell(r,5,"=SUMIF($H$11:$H$19,\"produit\",E11:E19)-SUMIF($H$11:$H$19,\"charge\",E11:E19)").number_format=NUM
w3.cell(r,6,"=E20-D20").number_format='+#,##0;−#,##0'
for j in (4,5,6):
    w3.cell(r,j).font=F(10,True,INK,MONO); w3.cell(r,j).alignment=R; w3.cell(r,j).border=Border(top=sd(INK3))
# barre de données sur la variation absolue
w3.conditional_formatting.add("F11:F19",DataBarRule(start_type="min",end_type="max",color="FF"+BLUE,showValue=True))
w3.conditional_formatting.add("F11:F19",CellIsRule(operator="lessThan",formula=["0"],font=Font(name=MONO,size=9.5,bold=True,color=CRIT)))

box(w3,22,2,25,8,bg=PANEL2)
w3.merge_cells(start_row=22,start_column=2,end_row=25,end_column=8)
w3.cell(22,2,"  Les deux drills se répondent.  Le bridge disait « effet coûts directs −41 579 € ». Ici on voit lequel : "
             "le compte 641 — masse salariale permanente — arrive en tête avec +34 000 €, suivi des loyers à +10 500 €. "
             "Le premier drill donne la cause économique, le second la ligne comptable. C'est le moment où le CFO comprend "
             "qu'il peut aller du comité de direction jusqu'à l'écriture sans changer d'outil.").font=F(9.5,c=INK2)
w3.cell(22,2).alignment=TOP
for r in range(22,26): w3.cell(r,2).border=Border(left=sd(ORANGE,"thick"))

for ws in (w1,w2,w3): ws.sheet_view.zoomScale=90
wb.active=0
out="/home/user/demo5/eduservices/DRILL_SIMULATION.xlsx"
wb.save(out); print("SAVED",out)

# ---- contrôle arithmétique ----
cae_p,cae_n=CA_P/EFF_P,CA_N/EFF_N; cve_p,cve_n=CVAR_P/EFF_P,CVAR_N/EFF_N
EB_P=CA_P-CVAR_P-CDIR_P; EB_N=CA_N-CVAR_N-CDIR_N
e1=(EFF_N-EFF_P)*(cae_p-cve_p); e2=(cae_n-cae_p)*EFF_N
e3=-(cve_n-cve_p)*EFF_N; e4=-(CDIR_N-CDIR_P)
print("\nEBITDA 2025 %10.0f"%EB_P)
print("  + effet effectifs      %+10.0f   (%d -> %d eleves)"%(e1,EFF_P,EFF_N))
print("  + effet prix / mix     %+10.0f   (%.0f -> %.0f EUR/eleve)"%(e2,cae_p,cae_n))
print("  - effet cout var/eleve %+10.0f   (%.0f -> %.0f EUR/eleve)"%(e3,cve_p,cve_n))
print("  - effet couts directs  %+10.0f   (%.0f -> %.0f, %+.1f%%)"%(e4,CDIR_P,CDIR_N,100*(CDIR_N/CDIR_P-1)))
print("EBITDA 2026 %10.0f   (cellule 96 300)   ecart %.6f"%(EB_N,abs(EB_P+e1+e2+e3+e4-EB_N)))
print("CA %+.1f%%  vs  couts directs %+.1f%%"%(100*(CA_N/CA_P-1),100*(CDIR_N/CDIR_P-1)))
