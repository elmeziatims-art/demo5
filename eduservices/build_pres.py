#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Copie de presentation (.xlsx, sans Tagetik) : met en forme cad/PIL/ALLOC facon
cabinet, masque les onglets techniques, ajoute 00_Notice. Idempotent.
Le fichier garde les formules vivantes (recalcul dans Excel)."""
import warnings;warnings.filterwarnings("ignore")
import openpyxl
from openpyxl.styles import Alignment
from pres_lib import *

SRC="DESIGN3_OP.xlsm"
OUT="Simulation_CFO_EDUSERVICES_2026-2027.xlsx"

def numf(ws,ref,key,color=None,inputcell=False):
    ws[ref].number_format=NF[key]
    ws[ref].alignment=Alignment(horizontal="right",vertical="center")
    if inputcell:
        style_saisie(ws,ref)
    elif color:
        ws[ref].font=font(10,color=color)

# ============================== CAD ==============================
def format_cad(ws):
    clear_zone(ws,1,40,"A","J")        # idempotence : retire fills/bordures residuels
    # gouttiere + largeurs
    set_gutter(ws,"A",2.5)
    ws.column_dimensions["B"].width=34
    for c in ("C","D","E"): ws.column_dimensions[c].width=15.5
    ws.column_dimensions["F"].width=14
    ws.column_dimensions["G"].width=10.5
    ws.column_dimensions["H"].width=3
    ws.column_dimensions["I"].width=23
    ws.column_dimensions["J"].width=11
    # ---- bandeau + sous-titre ----
    style_titre(ws,2,"B","J","POSTE DE COMMANDE CFO   ·   Cadrage CA & EBITDA 2027")
    style_soustitre(ws,3,"B","J",
        "EDUSERVICES GROUP  ·  MBway · ISCOM · Ipac · Pigier · Tunon  ·  Simulation budgétaire  ·  Montants en €")
    ws.row_dimensions[1].height=6
    ws.row_dimensions[4].height=6
    # ---- cadrage (cibles) ligne 5-6 ----
    ws["B5"].value="Scénario actif"; ws["B5"].font=font(10,bold=True)
    ws["B5"].alignment=Alignment(horizontal="left",vertical="center")
    style_saisie(ws,"C5"); ws["C5"].alignment=Alignment(horizontal="center",vertical="center")
    ws["E5"].value="Croissance CA cible"; ws["E6"].value="Marge EBITDA cible"
    for r in (5,6):
        ws["E%d"%r].font=font(10,bold=True)
        ws["E%d"%r].alignment=Alignment(horizontal="right",vertical="center")
    numf(ws,"F5","pct",inputcell=True); numf(ws,"F6","pct",inputcell=True)
    for r in (5,6): ws.row_dimensions[r].height=19
    # ---- sections ligne 8 ----
    style_section(ws,8,"B","G","1 ·  Réconciliation   —   Référence · Cible · Construit (scénario actif)")
    style_section(ws,8,"I","J","Coeff prix / marque")
    # ---- entetes ligne 10 ----
    for col,txt in [("B","Indicateur"),("C","Référence"),("D","Cible"),
                    ("E","Construit"),("F","Écart"),("G","Écart %")]:
        ws[col+"10"].value=txt
    style_entete(ws,10,"B","G")
    ws["I10"].value="Marque"; ws["J10"].value="Coeff prix"
    style_entete(ws,10,"I","J")
    # captions annee ligne 11
    for col,txt in [("C","2026"),("D","2027"),("E","2027")]:
        ws[col+"11"].value=txt; style_note(ws,col+"11")
        ws[col+"11"].alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[11].height=13
    # ---- reconciliation data 13-16 ----
    style_corps(ws,13,"B","G",num_cols=("C","D","E","F","G"))
    style_corps(ws,14,"B","G",num_cols=("C","D","E","F","G"))
    style_corps(ws,15,"B","G",num_cols=("C","D","E","F","G"))
    style_corps(ws,16,"B","G",num_cols=("C","D","E","F","G"))
    # CA (13) & EBITDA (14): euros
    for r in (13,14):
        numf(ws,"C%d"%r,"euro",inputcell=(r in(13,14)))   # reference = saisie dur
        numf(ws,"D%d"%r,"euro",color=TXT_FORMULE)          # cible = formule locale
        numf(ws,"E%d"%r,"euro",color=TXT_LIEN)             # construit = lien
        numf(ws,"F%d"%r,"ecart_eur",color=TXT_FORMULE)
        numf(ws,"G%d"%r,"pct",color=TXT_FORMULE)
    # Marge % (15)
    for col in ("C","D","E","G"): numf(ws,"%s15"%col,"pct",color=TXT_FORMULE)
    numf(ws,"F15","pct",color=TXT_FORMULE)
    ws["E15"].font=font(10,color=TXT_FORMULE)
    # Effectif (16)
    numf(ws,"C16","eff",inputcell=True); numf(ws,"E16","eff",color=TXT_LIEN)
    numf(ws,"F16","ecart",color=TXT_FORMULE); numf(ws,"G16","pct",color=TXT_FORMULE)
    # gras libelles indicateurs
    for r in (13,14,15,16): ws["B%d"%r].font=font(10,bold=True)
    # ---- coeff prix 12-16 ----
    for r in range(12,17):
        style_corps(ws,r,"I","J",num_cols=("J",))
        numf(ws,"J%d"%r,"coeff",inputcell=True)
    # ---- LEVIERS ----
    style_section(ws,19,"B","F","2 ·  Leviers de CROISSANCE / REVENUS")
    # entete 21 (Parametre | Cadrage | Optimiste | Prudent | ACTIF)
    ws["B21"].value="Paramètre"; ws["F21"].value="ACTIF (scénario)"
    style_entete(ws,21,"B","F")
    ws.row_dimensions[22].hidden=True   # ancre technique du MATCH ($C$22:$E$22)
    def lever_rows(rows):
        for r in rows:
            style_corps(ws,r,"B","F",num_cols=("C","D","E","F"))
            ws["B%d"%r].alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
            fee = (r==39)
            key = "euro" if fee else "pct"
            for col in ("C","D","E"): numf(ws,"%s%d"%(col,r),key,inputcell=True)
            numf(ws,"F%d"%r,key,color=TXT_FORMULE)  # ACTIF = formule locale
            ws.row_dimensions[r].height=30
    lever_rows(range(23,29))
    style_section(ws,30,"B","F","3 ·  Leviers de COÛTS")
    lever_rows(range(32,37))
    style_section(ws,38,"B","F","4 ·  Constante  —  frais de dossier")
    lever_rows([39])
    # petites lignes vides
    for r in (7,9,17,18,20,29,31,37): ws.row_dimensions[r].height=6
    # ---- reglages feuille ----
    sheet_setup(ws, freeze="A3", title_rows="1:2")
    ws.print_area="A1:J40"
    assert not ws.merged_cells.ranges, "cad: fusion detectee !"

# ============================== PIL ==============================
def format_pil(ws):
    clear_zone(ws,1,50,"A","O")
    W={"A":13,"B":9,"C":13,"D":11,"E":16,"F":11,"G":10.5,"H":15,"I":11,
       "J":11,"K":13,"L":11,"M":3,"N":9,"O":6}
    for c,w in W.items(): ws.column_dimensions[c].width=w
    style_titre(ws,2,"B","O","PILOTAGE   ·   Cockpit de décision CA / EBITDA 2027")
    style_soustitre(ws,3,"B","O",
        "EDUSERVICES GROUP  ·  14 campus  ·  Scénario : Cadrage (V01)  ·  Montants en €  ·  Exercice 2027")
    ws.row_dimensions[1].height=6
    # ---- bandeau KPI (lignes 6-7) ----
    pf=fill(PALETTE["section_fond"])
    for col in rng_cols("B","O"):
        ws[col+"6"].fill=pf; ws[col+"7"].fill=pf
        ws[col+"7"].border=border(bottom=(PALETTE["bord_fort"],"medium"))
    KPI=[("B","G","CA 2027","euro"),("H","J","EBITDA après siège","euro"),
         ("K","M","Marge EBITDA","pct"),("N","O","Effectif","eff")]
    for c1,c2,lab,key in KPI:
        ws[c1+"6"].value=lab
        for col in rng_cols(c1,c2):
            ws[col+"6"].font=font(9,bold=True,color=PALETTE["note"])
            ws[col+"6"].alignment=Alignment(horizontal="centerContinuous",vertical="center")
            ws[col+"7"].font=font(17,bold=True,color=PALETTE["entete_fond"])
            ws[col+"7"].alignment=Alignment(horizontal="centerContinuous",vertical="center")
        ws[c1+"7"].number_format=NF[key]
    ws.row_dimensions[6].height=18; ws.row_dimensions[7].height=32
    for r in (4,5,8,9,10,11,29,31): ws.row_dimensions[r].height=7
    # ================= 1 · CAP STRATEGIQUE (14-28) =================
    style_section(ws,12,"A","L","1 ·  Cap stratégique par campus   —   capacités & budget d'acquisition")
    CAPH=[("A","Campus"),("B","Marque"),("C","Ville"),("D","CAC marginal"),
          ("E","Croiss. leads"),("F","Intensité mkt"),("G","Cap. effectifs"),
          ("H","Cap. moment."),("I","Cap. potentiel"),("J","Cap retenu"),("K","Budget acq. réf.")]
    for col,txt in CAPH: ws[col+"14"].value=txt
    style_entete(ws,14,"A","K")
    for r in range(15,29):
        style_corps(ws,r,"A","K",num_cols=("D","E","F","G","H","I","J","K"))
        ws.row_dimensions[r].height=17
    zebrer(ws,15,28,"A","K")
    for r in range(15,29):
        ws["D%d"%r].number_format=NF["euro"]
        ws["E%d"%r].number_format=NF["pct"]; ws["F%d"%r].number_format=NF["pct"]
        for col in ("G","H","I"): ws["%s%d"%(col,r)].number_format=NF["coeff"]
        style_saisie(ws,"J%d"%r); ws["J%d"%r].number_format=NF["eff"]
        ws["J%d"%r].alignment=Alignment(horizontal="center",vertical="center")
        ws["K%d"%r].number_format=NF["euro"]
        ws["A%d"%r].font=font(10,bold=True,color=PALETTE["entete_fond"])   # code
    # ================= 2 · SYNTHESE (30-49) =================
    style_section(ws,30,"A","L","2 ·  Synthèse par campus   —   résultats reconstruits (scénario actif)")
    SYNH=[("A","Campus"),("B","Marque"),("C","Ville"),("D","Effectif"),("E","CA 2027"),
          ("F","Prix moyen"),("G","Part CA"),("H","EBITDA campus"),("I","Mrg EBITDA"),
          ("J","EBITDA / étud."),("K","Rejoué"),("L","CAC marg.")]
    for col,txt in SYNH: ws[col+"32"].value=txt
    style_entete(ws,32,"A","L")
    for r in range(33,47):
        style_corps(ws,r,"A","L",num_cols=("D","E","F","G","H","I","J","K","L"))
        ws.row_dimensions[r].height=17
        ws["D%d"%r].number_format=NF["eff"]
        for col in ("E","F","H","J","K","L"): ws["%s%d"%(col,r)].number_format=NF["euro"]
        ws["G%d"%r].number_format=NF["pct"]; ws["I%d"%r].number_format=NF["pct"]
        ws["A%d"%r].font=font(10,color=PALETTE["note"])
        # liens vers autres onglets -> vert
        for col in ("D","E","H"): ws["%s%d"%(col,r)].font=font(10,color=TXT_LIEN)
    zebrer(ws,33,46,"A","L")
    # totaux 47 (sous-total) / 48 (siege) / 49 (groupe)
    ws["A47"].value="Sous-total campus"; ws["A48"].value="Siège / holding (GRP)"; ws["A49"].value="GROUPE 2027"
    for r in (47,48,49):
        style_total(ws,r,"A","L")
        ws["D%d"%r].number_format=NF["eff"]        # Effectif = entier (pas d'€)
        for col in ("E","F","H","J","K","L"): ws["%s%d"%(col,r)].number_format=NF["euro"]
        ws["G%d"%r].number_format=NF["pct"]; ws["I%d"%r].number_format=NF["pct"]
        for col in rng_cols("D","L"):
            ws["%s%d"%(col,r)].alignment=Alignment(horizontal="right",vertical="center")
        ws["A%d"%r].alignment=Alignment(horizontal="left",vertical="center")
    ws.row_dimensions[47].height=18; ws.row_dimensions[48].height=18; ws.row_dimensions[49].height=20
    sheet_setup(ws, freeze="D3", title_rows="1:2")
    ws.print_area="A1:L49"
    assert not ws.merged_cells.ranges, "PIL: fusion detectee !"

# ============================== ALLOC ==============================
def format_alloc(ws):
    clear_zone(ws,1,96,"A","M")
    set_gutter(ws,"A",2.5)
    W={"B":30,"C":12,"D":14,"E":13,"F":13,"G":13,"H":13,"I":13,"J":13,"K":14,"L":14,"M":9}
    for c,w in W.items(): ws.column_dimensions[c].width=w
    style_titre(ws,2,"B","M","ALLOCATION & RENTABILITÉ   ·   coûts complets & marge par maille")
    style_soustitre(ws,3,"B","M",
        "EDUSERVICES GROUP  ·  Marque ▸ Campus ▸ Classe  ·  Exercice 2026  ·  Montants en €")
    ws.row_dimensions[1].height=6
    # ---- cles d'allocation (saisie) 5-9 ----
    style_section(ws,5,"B","M","Clés d'allocation & inducteurs de répartition (saisie)")
    KEYS={6:"Siège administratif  →  marque",7:"Publicité de marque  →  marque",
          8:"Marque  →  campus",9:"Campus  →  classe"}
    for r,lab in KEYS.items():
        ws["B%d"%r].value=lab; ws["B%d"%r].font=font(10,bold=True)
        ws["B%d"%r].alignment=Alignment(horizontal="left",vertical="center")
        style_saisie(ws,"C%d"%r)
        ws["C%d"%r].alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.row_dimensions[r].height=24
    for r in (4,10,11,12,13,15,16): ws.row_dimensions[r].height=7
    # ---- maille ----
    style_section(ws,14,"B","M","Maille fine   —   déplié campus / classe, coût complet et marge")
    HED=[("B","Marque ▸ Campus ▸ Classe"),("C","Effectif"),("D","CA"),("E","VAC"),
         ("F","PERM"),("G","ODIR"),("H","STRUCT"),("I","Frais marque"),("J","Holding"),
         ("K","Coût complet"),("L","Marge complète"),("M","Marge %")]
    for col,txt in HED: ws[col+"17"].value=txt
    style_entete(ws,17,"B","M")
    numcols=("C","D","E","F","G","H","I","J","K","L","M")
    zclass=0
    for r in range(18,96):
        b=ws["B%d"%r].value
        if not isinstance(b,str): continue
        lead=len(b)-len(b.lstrip()); lvl=lead//3
        ws["B%d"%r].value=b.strip()
        ws.row_dimensions[r].hidden=False            # deplie (source Tagetik collapse)
        ws.row_dimensions[r].outline_level=lvl        # groupement +/- (etendu)
        style_corps(ws,r,"B","M",num_cols=numcols)
        ws["C%d"%r].number_format=NF["eff"]
        for col in ("D","E","F","G","H","I","J","K","L"): ws["%s%d"%(col,r)].number_format=NF["euro"]
        ws["M%d"%r].number_format=NF["pct"]
        ws["B%d"%r].alignment=Alignment(horizontal="left",vertical="center",indent=lvl*2)
        ws.row_dimensions[r].height=16
        if r==95 or lvl==0:
            # marque (sous-total) ou GROUPE
            if b.strip()=="GROUPE":
                ws["B%d"%r].value="GROUPE  —  EDUSERVICES"
                style_total(ws,r,"B","M")
                for col in numcols: ws["%s%d"%(col,r)].alignment=Alignment(horizontal="right",vertical="center")
                ws["B%d"%r].alignment=Alignment(horizontal="left",vertical="center")
                ws["C%d"%r].number_format=NF["eff"]; ws["M%d"%r].number_format=NF["pct"]
                ws.row_dimensions[r].height=20
            else:
                for col in rng_cols("B","M"):
                    ws["%s%d"%(col,r)].fill=fill(PALETTE["section_fond"])
                    ws["%s%d"%(col,r)].border=border(top=(PALETTE["bord_int"],))
                ws["B%d"%r].font=font(10,bold=True,color=PALETTE["entete_fond"])
                for col in numcols: ws["%s%d"%(col,r)].font=font(10,bold=True)
                ws.row_dimensions[r].height=18
        elif lvl==1:
            ws["B%d"%r].font=font(10,bold=True)   # campus
        else:
            zclass+=1
            if zclass%2==0:
                for col in rng_cols("B","M"): ws["%s%d"%(col,r)].fill=fill(PALETTE["zebra"])
    from openpyxl.worksheet.properties import Outline
    ws.sheet_properties.outlinePr=Outline(summaryBelow=False,summaryRight=False)
    sheet_setup(ws, freeze="C18", title_rows="1:17")
    ws.print_area="A1:M95"
    assert not ws.merged_cells.ranges, "ALLOC: fusion detectee !"

# ============================== 00_NOTICE ==============================
def create_notice(wb):
    if "00_Notice" in wb.sheetnames: del wb["00_Notice"]
    ws=wb.create_sheet("00_Notice",0)
    clear_zone(ws,1,40,"A","H")
    set_gutter(ws,"A",2.5)
    ws.column_dimensions["B"].width=4
    ws.column_dimensions["C"].width=34
    ws.column_dimensions["D"].width=60
    for c in ("E","F","G","H"): ws.column_dimensions[c].width=3
    style_titre(ws,2,"B","H","NOTICE  &  CONVENTIONS DE LECTURE")
    style_soustitre(ws,3,"B","H",
        "EDUSERVICES GROUP  ·  Simulateur budgétaire CA / EBITDA  ·  Copie de présentation")
    ws.row_dimensions[1].height=6
    def swatch(r,hexfill_,txtcolor,libelle,desc):
        ws["B%d"%r].fill=fill(hexfill_)
        b=side(PALETTE["bord_int"])
        ws["B%d"%r].border=Border(top=b,bottom=b,left=b,right=b)
        ws["C%d"%r].value=libelle; ws["C%d"%r].font=font(10,bold=True,color=txtcolor)
        ws["C%d"%r].alignment=Alignment(horizontal="left",vertical="center")
        ws["D%d"%r].value=desc; ws["D%d"%r].font=font(10,color=PALETTE["corps_txt"])
        ws["D%d"%r].alignment=Alignment(horizontal="left",vertical="center")
        ws.row_dimensions[r].height=20
    style_section(ws,5,"B","H","Couleur du texte des cellules")
    swatch(6,"E4ECFA",TXT_SAISIE,"Saisie en dur","Hypothèse / levier modifiable par le contrôleur de gestion.")
    swatch(7,"FFFFFF",TXT_FORMULE,"Formule locale","Calcul effectué dans l'onglet (agrégat, ratio, écart).")
    swatch(8,"FFFFFF",TXT_LIEN,"Lien inter-onglet","Résultat repris du moteur de calcul (CA, EBITDA, effectif).")
    ws["C9"].value="Aucun lien externe"; ws["C9"].font=font(9,italic=True,color=PALETTE["note"])
    style_section(ws,11,"B","H","Écarts vs cible / référence")
    swatch(12,"1B6B4F","FFFFFF","Écart favorable","Construit ≥ cible (croissance, marge, effectif).")
    swatch(13,"8C2A32","FFFFFF","Écart défavorable","Construit < cible : point de vigilance.")
    style_section(ws,15,"B","H","Onglets de restitution")
    guide=[("cad — Poste de commande CFO","Cadrage des cibles, leviers de croissance & de coûts, réconciliation."),
           ("PIL — Pilotage","Cockpit KPI, cap stratégique par campus, synthèse des résultats."),
           ("ALLOC — Allocation & rentabilité","Maille Marque ▸ Campus ▸ Classe : coût complet et marge.")]
    for i,(t,d) in enumerate(guide):
        r=16+i
        ws["C%d"%r].value=t; ws["C%d"%r].font=font(10,bold=True,color=PALETTE["entete_fond"])
        ws["C%d"%r].alignment=Alignment(horizontal="left",vertical="center")
        ws["D%d"%r].value=d; ws["D%d"%r].font=font(10,color=PALETTE["corps_txt"])
        ws["D%d"%r].alignment=Alignment(horizontal="left",vertical="center")
        ws.row_dimensions[r].height=19
    style_note(ws,"C20","Montants en €  ·  Référence 2026, cible/construit 2027  ·  Scénario par défaut : Cadrage (V01).")
    ws["C21"].value="Les onglets techniques (moteur de calcul, feeds) sont masqués ; les formules restent vives (recalcul dans Excel)."
    ws["C21"].font=font(9,italic=True,color=PALETTE["note"])
    sheet_setup(ws, landscape=True)
    ws.sheet_view.showGridLines=False
    assert not ws.merged_cells.ranges

# ============================== main ==============================
def main():
    wb=openpyxl.load_workbook(SRC)  # openpyxl -> perd les briques Tagetik (voulu)
    format_cad(wb["cad"])
    format_pil(wb["PIL"])
    format_alloc(wb["ALLOC"])
    create_notice(wb)
    # ordre des onglets : Notice, cad, PIL, ALLOC, (techniques masques)
    order=["00_Notice","cad","PIL","ALLOC"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
    # onglets a masquer (techniques + feeds)
    for t in ["_CALC_MOTEUR","_CALC_PNL","_CALC_ALLOC","00_Cartographie",
              "Socle","Campagne","Moteur","Compta","PNL","Allocation"]:
        if t in wb.sheetnames: wb[t].sheet_state="hidden"
    if "_TGK_HIDDEN" in wb.sheetnames:
        del wb["_TGK_HIDDEN"]   # registre Tagetik inutile dans la copie de presentation
    wb.active=wb.sheetnames.index("cad")
    wb.calculation.fullCalcOnLoad=True   # Excel recalcule a l'ouverture (valeurs reelles)
    for name in order:
        wb[name].sheet_view.selection[0].activeCell="A1"
        wb[name].sheet_view.selection[0].sqref="A1"
    wb.save(OUT)
    print("OK ->",OUT,"| onglets:",wb.sheetnames)

if __name__=="__main__":
    main()
