#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""EXEMPLES_COCKPIT.xlsx — 2 exemples propres à reproduire :
  • Bridge CAC en WATERFALL (barres empilées : socle invisible + hausse/baisse/ancres)
  • YoY Marge EBITDA en POINTS (formule =E-D)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E3EFEA"; WHITE="FFFFFF"
SOFT="51606D"; FAINT="7D8B98"; RULE="C8D2DA"; OCHRE="B3641C"; OK="1E7A55"; RED="C0392B"
GREEN="1E7A55"; NAVY="1F3864"; YEL="FFF6DA"
AR="Arial"
def F(sz=10,b=False,c=INK): return Font(name=AR,size=sz,bold=b,color=c)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color=RULE); box=Border(left=thin,right=thin,top=thin,bottom=thin)
Ln=Alignment("left",vertical="center"); Rn=Alignment("right",vertical="center"); Cn=Alignment("center",vertical="center",wrap_text=True)

# ---- décomposition EXACTE du CAC (foote au centime) ----
DEP25,DEP26=394702,434174; INS25,INS26=1159,1229
CAC25=DEP25/INS25                 # 340.55
CAC26=DEP26/INS26                 # 353.27
EFF_DEP=(DEP26-DEP25)/INS25       # +34.06  (à inscrits 2025)
EFF_VOL=DEP26/INS26 - DEP26/INS25 # -21.34  (effet volume)
# contrôle : CAC25 + EFF_DEP + EFF_VOL == CAC26

wb=openpyxl.Workbook()

# ============================================================ 1) BRIDGE CAC
ws=wb.active; ws.title="Bridge CAC (waterfall)"; ws.sheet_view.showGridLines=False
for col,w in zip("ABCDEF",[20,12,12,12,12,3]): ws.column_dimensions[col].width=w
ws.merge_cells("A1:E1"); ws.cell(1,1,"  BRIDGE DU CAC — d'où vient la hausse 2025 → 2026  (waterfall propre)").font=F(12,True,WHITE); ws.cell(1,1).fill=fill(NAVY); ws.row_dimensions[1].height=24

# tableau lisible
ws.cell(3,1,"Étape").font=F(9.5,True,WHITE); ws.cell(3,2,"CAC (€)").font=F(9.5,True,WHITE)
for c in (1,2): ws.cell(3,c).fill=fill(TEAL); ws.cell(3,c).alignment=Cn; ws.cell(3,c).border=box
steps=[("CAC 2025",round(CAC25)),("+ Effet dépenses",round(EFF_DEP)),("− Effet volume",round(EFF_VOL)),("CAC 2026",round(CAC26))]
for i,(lab,v) in enumerate(steps):
    r=4+i; ws.cell(r,1,lab).font=F(9.5,lab.startswith('CAC')); ws.cell(r,1).alignment=Ln; ws.cell(r,1).border=box
    x=ws.cell(r,2,v); x.number_format='#,##0" €"'; x.font=F(9.5,lab.startswith('CAC'),GREEN if '+' in lab else (RED if '−' in lab else INK)); x.alignment=Rn; x.border=box

# --- données waterfall (socle invisible + 3 séries visibles) ---
# colonnes : A cat | B SOCLE(invisible) | C ANCRE(bleu) | D HAUSSE(vert) | E BAISSE(rouge)
hr=9
ws.cell(hr,1,"cat").font=F(8,c=FAINT)
for j,h in enumerate(["Étape","Socle","Ancre","Hausse","Baisse"],1):
    ws.cell(hr,j,h).font=F(8,True,FAINT)
wf=[
 ("CAC 2025", 0,          CAC25, 0,       0),
 ("Effet dépenses", CAC25, 0,     EFF_DEP, 0),          # monte de CAC25 -> CAC25+dep
 ("Effet volume", CAC26,  0,     0,       -EFF_VOL),    # -EFF_VOL est positif (hauteur) ; base=CAC26
 ("CAC 2026", 0,          CAC26, 0,       0),
]
for i,(cat,socle,ancre,hausse,baisse) in enumerate(wf):
    r=hr+1+i
    ws.cell(r,1,cat).font=F(8,c=FAINT)
    ws.cell(r,2,round(socle,2)); ws.cell(r,3,round(ancre,2)); ws.cell(r,4,round(hausse,2)); ws.cell(r,5,round(baisse,2))
NR=len(wf)
cats=Reference(ws,min_col=1,max_col=1,min_row=hr+1,max_row=hr+NR)

ch=BarChart(); ch.type="col"; ch.grouping="stacked"; ch.overlap=100
ch.title="Évolution du CAC 2025 → 2026 (€/inscrit)"
ch.height=8.5; ch.width=15; ch.y_axis.numFmt='#,##0 "€"'; ch.legend=None
ch.x_axis.delete=False; ch.y_axis.delete=False; ch.gapWidth=40
# ordre séries : Socle(2) invisible, Ancre(3), Hausse(4), Baisse(5)
defs=[(2,"Socle",None,None),(3,"Ancre",TEALD,'#,##0" €"'),(4,"Hausse",GREEN,'"+ "#,##0" €"'),(5,"Baisse",RED,'"− "#,##0" €"')]
for colx,name,color,fmt in defs:
    ref=Reference(ws,min_col=colx,max_col=colx,min_row=hr+1,max_row=hr+NR)
    ch.add_data(ref,titles_from_data=False)
s_socle,s_ancre,s_hausse,s_baisse=ch.series
s_socle.graphicalProperties.noFill=True; s_socle.graphicalProperties.line.noFill=True
for s,color in ((s_ancre,TEALD),(s_hausse,GREEN),(s_baisse,RED)):
    s.graphicalProperties.solidFill=color; s.graphicalProperties.line.solidFill=color
# labels (avec signe) sur les 3 visibles, pas sur le socle
for s,fmt in ((s_ancre,'#,##0" €"'),(s_hausse,'"+ "#,##0" €"'),(s_baisse,'"− "#,##0" €"')):
    s.dLbls=DataLabelList(); s.dLbls.showVal=True; s.dLbls.numFmt=fmt; s.dLbls.dLblPos="ctr"
    s.dLbls.showSerName=False; s.dLbls.showCatName=False; s.dLbls.showLegendKey=False
ch.set_categories(cats)
ws.add_chart(ch,"D3")

ws.merge_cells("A15:B18")
ws.cell(15,1,"Technique : une série « Socle » INVISIBLE (noFill) porte chaque barre à la bonne hauteur, "
             "puis 3 séries visibles empilées dessus — Ancre (bleu, CAC 2025 & 2026), Hausse (vert), Baisse (rouge). "
             "Labels formatés en € avec le signe (\"+ \"/\"− \"), nom de série masqué. Rien d'autre à afficher.").font=F(8.5,c=FAINT)
ws.cell(15,1).alignment=Alignment("left",vertical="top",wrap_text=True)

# ============================================================ 2) YoY MARGE EN POINTS
w2=wb.create_sheet("YoY Marge (points)"); w2.sheet_view.showGridLines=False
for col,w in zip("ABCDEF",[22,13,13,13,14,32]): w2.column_dimensions[col].width=w
w2.merge_cells("A1:F1"); w2.cell(1,1,"  YoY d'une MARGE : en POINTS, pas en % relatif").font=F(12,True,WHITE); w2.cell(1,1).fill=fill(NAVY); w2.row_dimensions[1].height=24
hdr=["KPI","2024","2025","2026","YoY 25→26","← formule exacte"]
for j,h in enumerate(hdr,1):
    c=w2.cell(3,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL); c.alignment=Cn if j>1 else Ln; c.border=box
# CA
w2.cell(4,1,"Chiffre d'affaires"); w2.cell(4,2,20552827); w2.cell(4,3,21775820); w2.cell(4,4,23098985)
w2.cell(4,5,"=D4/C4-1").number_format='"▲ "0.0%;"▼ "0.0%'; w2.cell(4,6,'=D4/C4-1   (variation en %)').font=F(8.5,c=FAINT)
# EBITDA
w2.cell(5,1,"EBITDA"); w2.cell(5,2,3136652); w2.cell(5,3,3484818); w2.cell(5,4,3845790)
w2.cell(5,5,"=D5/C5-1").number_format='"▲ "0.0%;"▼ "0.0%'; w2.cell(5,6,'=D5/C5-1   (variation en %)').font=F(8.5,c=FAINT)
# Marge EBITDA % (la ligne clé)
w2.cell(6,1,"Marge EBITDA %").font=F(10,True)
for j,r in ((2,4),(3,4),(4,4)): pass
w2.cell(6,2,"=B5/B4").number_format='0.0%'; w2.cell(6,3,"=C5/C4").number_format='0.0%'; w2.cell(6,4,"=D5/D4").number_format='0.0%'
c=w2.cell(6,5,"=(D6-C6)*100"); c.number_format='"+ "0.0" pt";"− "0.0" pt"'; c.font=F(10,True,OK)
w2.cell(6,6,'=(D6-C6)*100   ← ×100 car la marge est une FRACTION (0,166) → +0,6 pt').font=F(8.5,True,OCHRE)
for r in range(4,7):
    for j in range(2,5):
        if not w2.cell(r,j).number_format or w2.cell(r,j).number_format=='General':
            if r<6: w2.cell(r,j).number_format='#,##0'
        w2.cell(r,j).alignment=Rn
    for j in range(1,6): w2.cell(r,j).border=box
    w2.cell(r,1).alignment=Ln; w2.cell(r,5).alignment=Rn
w2.merge_cells("A8:F10")
w2.cell(8,1,"Pourquoi : une marge est déjà un %. Sa variation « en % » (16,6/16,0−1 = +4,0 %) est trompeuse — "
            "on croit +4 points. La bonne lecture = la DIFFÉRENCE de points : 16,6 % − 16,0 % = +0,6 pt. "
            "ATTENTION : la marge est stockée en FRACTION (0,166), donc =D6-C6 vaut 0,0065 → le format afficherait « 0,0 pt ». "
            "Il faut ×100 : =(D6-C6)*100 = 0,65 → format \"+ \"0.0\" pt\" → + 0,6 pt.").font=F(9,c=SOFT)
w2.cell(8,1).alignment=Alignment("left",vertical="top",wrap_text=True)
# repère cellule jaune sur la formule clé
w2.cell(6,5).fill=fill(YEL)

out="/home/user/demo5/eduservices/EXEMPLES_COCKPIT.xlsx"
wb.save(out)
print("SAVED",out)
print("CAC25=%.2f  +dép=%.2f  -vol=%.2f  CAC26=%.2f  (contrôle=%.2f)"%(CAC25,EFF_DEP,EFF_VOL,CAC26,CAC25+EFF_DEP+EFF_VOL))
print("Marge 2025=%.4f 2026=%.4f  YoY points=%.3f"%(3484818/21775820,3845790/23098985,3845790/23098985-3484818/21775820))
