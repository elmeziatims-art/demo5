#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""RAPPORT_ALLOUE.xlsx & RAPPORT_EVOLUTION.xlsx — forme TAGETIK (à plat).
Une DIMENSION = une COLONNE ; les mesures (marges) sont des colonnes calculées
(pas de dimension Compte). Reproduit fidèlement la cascade de V_ALLOCATION
(vérifié : EBITDA net 2026 = 3 291 530, écart 0). Clés : K1=VOL_EFF, K2=REV_CA,
K3=VOL_EFF, K4=VOL_EFF (extrait cadrage)."""
import csv, glob
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; WHITE="FFFFFF"; RULE="C8D2DA"
SOFT="51606D"; OCHRE="B3641C"; GREEN="1E7A55"; DIMBG="EAF0F4"; FAINT="7D8B98"
AR="Arial"
def F(sz=10,b=False,c=INK): return Font(name=AR,size=sz,bold=b,color=c)
def fill(c): return PatternFill("solid",fgColor=c)
LFT=Alignment("left",vertical="center"); RGT=Alignment("right",vertical="center"); CTR=Alignment("center",vertical="center",wrap_text=True)
med=Side(style="medium",color=RULE); thin=Side(style="thin",color=RULE)
EUR='#,##0;-#,##0;"-"'; NUM='#,##0'; PCT='0.0%'
COMP=glob.glob('/home/user/demo5/eduservices/tgk_data/ext_*/AW_002_000004_000001.csv')[0]
SOCLE=glob.glob('/home/user/demo5/eduservices/tgk_data/ext_*/AW_002_000002_000001.csv')[0]
MQ={'MBWAY':'MBway','ISCOM':'Iscom','IPAC':'Ipac','PIGIER':'Pigier','TUNON':'Tunon'}
CITY={'LYO':'Lyon','PAR':'Paris','BOR':'Bordeaux','NAN':'Nantes','MTP':'Montpellier','REN':'Rennes','LIL':'Lille','TLS':'Toulouse'}
MODL={'INIT':'Initial','ALT':'Alternance'}
def hrs(prog,mod):
    if prog.startswith('BAC') and mod=='INIT': return 600
    if prog.startswith('BAC'): return 480
    if prog.startswith('MAS') and mod=='INIT': return 520
    if prog.startswith('MAS'): return 420
    if mod=='INIT': return 1000
    return 700

def compute(year):
    camp=defaultdict(lambda: defaultdict(float)); hold=defaultdict(float)
    with open(COMP,encoding='utf-8') as fh:
        r=csv.reader(fh,delimiter=';'); next(r)
        for row in r:
            en,acc,ex=row[0],row[1],row[2]
            if ex!=year or acc in ('TEC_EBITDA','TEC_PL'): continue
            a=float(row[5].replace(',','.'))
            if en=='GRP':
                if acc in ('6414','6226','626','6281','6331','6333'): hold['HOLDING']+=a
                elif acc=='6236': hold['MARQUE']+=a
                continue
            if acc=='621': camp[en]['VAC']+=a
            elif acc=='6411': camp[en]['PERM']+=a
            elif acc in ('604','6063'): camp[en]['ODIR']+=a
            elif acc=='6231': camp[en]['MKT']+=a
            elif acc in ('6413','645','613','615','616','625','63511'): camp[en]['STRUCT']+=a
    cls=[]
    with open(SOCLE,encoding='utf-8') as fh:
        r=csv.reader(fh,delimiter=';'); h=next(r); ix={n:i for i,n in enumerate(h)}
        g=lambda row,n: float(row[ix[n]].replace(',','.'))
        for row in r:
            if row[ix['EXERCICE']]!=year: continue
            en=row[ix['ENTITY']]; prog=row[ix['PROGRAMME']]; an=row[ix['AN_ETUDE']]; mod=row[ix['MODALITE']]
            eff=g(row,'VOL_EFF'); vcl=g(row,'VOL_CLASS'); new=g(row,'VOL_NEW')
            ca=eff*g(row,'REV_STUD')+new*g(row,'REV_FRAIS_INS')
            cls.append(dict(en=en,mq=en.split('_')[0],prog=prog,an=an,mod=mod,eff=eff,vcl=vcl,new=new,ca=ca,hrs=vcl*hrs(prog,mod)))
    E=defaultdict(lambda: defaultdict(float)); M=defaultdict(lambda: defaultdict(float)); G=defaultdict(float)
    for c in cls:
        for k,f in [('HRS','hrs'),('EFF','eff'),('NEW','new'),('CLS','vcl'),('CA','ca')]:
            E[c['en']][k]+=c[f]; M[c['mq']][k]+=c[f]; G[k]+=c[f]
    out=[]
    for c in cls:
        en,mq=c['en'],c['mq']; cp=camp[en]
        cvac  = cp['VAC']*c['hrs']/E[en]['HRS'] if E[en]['HRS'] else 0
        cperm = cp['PERM']*c['hrs']/E[en]['HRS'] if E[en]['HRS'] else 0
        codir = (cp['ODIR']*c['eff']/E[en]['EFF'] if E[en]['EFF'] else 0) + (cp['MKT']*c['new']/E[en]['NEW'] if E[en]['NEW'] else 0)
        cstr  = cp['STRUCT']*c['eff']/E[en]['EFF'] if E[en]['EFF'] else 0
        share = (M[mq]['EFF']/G['EFF']) * (E[en]['CA']/M[mq]['CA'] if M[mq]['CA'] else 0) * (c['eff']/E[en]['EFF'] if E[en]['EFF'] else 0)
        cmq=hold['MARQUE']*share; chol=hold['HOLDING']*share
        propre=c['ca']-(cvac+cperm+codir+cstr); siege=cmq+chol; net=propre-siege
        out.append(dict(EXERCICE=year,VERSION='ACT',MARQUE=MQ[mq],CAMPUS=f"{MQ[mq]} {CITY.get(en.split('_')[1],en.split('_')[1])}",
            PROGRAMME=c['prog'],AN_ETUDE=c['an'],MODALITE=MODL.get(c['mod'],c['mod']),
            EFFECTIF=c['eff'],CA=c['ca'],
            COST_VAC=cvac,COST_PERM=cperm,COST_ODIR=codir,COST_STRUCT=cstr,COST_MARQUE=cmq,COST_HOLDING=chol,
            EBITDA_PROPRE=propre,QUOTE_PART_SIEGE=siege,EBITDA_NET=net))
    return out

def sheet(ws,rows,cols):
    # cols = list of (header, key, kind)  kind: dim/eur/num/pct
    ws.sheet_view.showGridLines=False; ws.freeze_panes="A2"
    for j,(hd,k,kind) in enumerate(cols,1):
        c=ws.cell(1,j,hd); c.font=F(8.5,True,WHITE); c.fill=fill(TEAL); c.alignment=LFT if kind=='dim' else CTR; c.border=Border(bottom=med)
    r=2
    for row in rows:
        for j,(hd,k,kind) in enumerate(cols,1):
            if kind=='pct':
                v = row['EBITDA_NET']/row['CA'] if row['CA'] else 0
                c=ws.cell(r,j,v); c.number_format=PCT; c.alignment=RGT; c.font=F(8.5,False,GREEN if v>=0.05 else OCHRE)
            else:
                v=row[k]; c=ws.cell(r,j,round(v) if kind in('eur','num') else v)
                if kind=='dim': c.font=F(8.5); c.alignment=LFT; c.fill=fill(DIMBG)
                else: c.number_format=EUR if kind=='eur' else NUM; c.alignment=RGT; c.font=F(8.5)
        r+=1
    for col,w in zip("ABCDEFGHIJKLMNOPQR",[8,7,14,18,8,10]+[10]*12): ws.column_dimensions[col].width=w

# ===== RAPPORT_ALLOUE (2026, à plat, avant/après) =====
rows26=compute('2026')
wb=openpyxl.Workbook(); ws=wb.active; ws.title="Alloué 2026"
cols_a=[("EXERCICE","EXERCICE","dim"),("VERSION","VERSION","dim"),("MARQUE","MARQUE","dim"),("CAMPUS","CAMPUS","dim"),
        ("PROGRAMME","PROGRAMME","dim"),("AN_ETUDE","AN_ETUDE","dim"),("MODALITE","MODALITE","dim"),
        ("EFFECTIF","EFFECTIF","num"),("CA","CA","eur"),
        ("EBITDA_PROPRE","EBITDA_PROPRE","eur"),("QUOTE_PART_SIEGE","QUOTE_PART_SIEGE","eur"),("EBITDA_NET","EBITDA_NET","eur"),("MARGE_NET_%",None,"pct"),
        ("COST_VAC","COST_VAC","eur"),("COST_PERM","COST_PERM","eur"),("COST_ODIR","COST_ODIR","eur"),
        ("COST_STRUCT","COST_STRUCT","eur"),("COST_MARQUE","COST_MARQUE","eur"),("COST_HOLDING","COST_HOLDING","eur")]
sheet(ws,rows26,cols_a)
wb.save("/home/user/demo5/eduservices/tagetik/RAPPORT_ALLOUE.xlsx")
print("SAVED RAPPORT_ALLOUE.xlsx  lignes=",len(rows26),"  EBITDA_net=",round(sum(r['EBITDA_NET'] for r in rows26)))

# ===== RAPPORT_EVOLUTION (2024-2026, à plat) =====
rows_evo=[]
for y in ('2024','2025','2026'): rows_evo+=compute(y)
wb2=openpyxl.Workbook(); ws2=wb2.active; ws2.title="Évolution 2024-2026"
cols_e=[("EXERCICE","EXERCICE","dim"),("MARQUE","MARQUE","dim"),("CAMPUS","CAMPUS","dim"),
        ("PROGRAMME","PROGRAMME","dim"),("AN_ETUDE","AN_ETUDE","dim"),("MODALITE","MODALITE","dim"),
        ("EFFECTIF","EFFECTIF","num"),("CA","CA","eur"),("EBITDA_NET","EBITDA_NET","eur"),("MARGE_NET_%",None,"pct")]
sheet(ws2,rows_evo,cols_e)
wb2.save("/home/user/demo5/eduservices/tagetik/RAPPORT_EVOLUTION.xlsx")
print("SAVED RAPPORT_EVOLUTION.xlsx  lignes=",len(rows_evo))
for y in ('2024','2025','2026'):
    n=sum(r['EBITDA_NET'] for r in rows_evo if r['EXERCICE']==y); ca=sum(r['CA'] for r in rows_evo if r['EXERCICE']==y)
    print(f"  {y}: EBITDA_net={n:,.0f}  marge={n/ca*100:.1f}%")
