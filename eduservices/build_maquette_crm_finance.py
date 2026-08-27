#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Maquettes : Rapport CRM (vu commercial) et Rapport Finance (vu comptable) — 2026.
Les deux lentilles derrière le CA réconcilié. Formules partout."""
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E4F0EC"; CARD2="F5F7F9"; FAINT="7D8B98"; WHITE="FFFFFF"; BLUE="0000FF"; OCHRE="B3641C"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color="DBE2E9"); med=Side(style="medium",color=TEAL)
CTR=Alignment(horizontal="center",vertical="center",wrap_text=True); LFT=Alignment("left",vertical="center"); RGT=Alignment("right",vertical="center")
EUR='#,##0;(#,##0);"-"'; NUM='#,##0;(#,##0);"-"'; EURc='#,##0" €"'

CITY={"IPAC_MTP":"Montpellier","IPAC_NAN":"Nantes","IPAC_REN":"Rennes","ISCOM_LIL":"Lille","ISCOM_PAR":"Paris","ISCOM_TLS":"Toulouse",
"MBWAY_BOR":"Bordeaux","MBWAY_LYO":"Lyon","MBWAY_NAN":"Nantes","MBWAY_PAR":"Paris","PIGIER_BOR":"Bordeaux","PIGIER_LYO":"Lyon","TUNON_LYO":"Lyon","TUNON_PAR":"Paris"}
CRM={"IPAC":{"IPAC_MTP":[117,778050,40,3600],"IPAC_NAN":[145,1015000,50,4500],"IPAC_REN":[117,778050,40,3600]},
"ISCOM":{"ISCOM_LIL":[244,1746850,100,9000],"ISCOM_PAR":[354,2895760,144,12960],"ISCOM_TLS":[231,1619520,94,8460]},
"MBWAY":{"MBWAY_BOR":[249,1764430,102,9180],"MBWAY_LYO":[324,2484825,132,11880],"MBWAY_NAN":[294,2147500,120,10800],"MBWAY_PAR":[382,3124800,156,14040]},
"PIGIER":{"PIGIER_BOR":[142,895310,72,6480],"PIGIER_LYO":[182,1242150,92,8280]},
"TUNON":{"TUNON_LYO":[117,859950,40,3600],"TUNON_PAR":[138,1081920,47,4230]}}
FIN={"MBWAY":[1995975,7525580,45900],"ISCOM":[1310700,4951430,30420],"IPAC":[0,2571100,11700],"PIGIER":[0,2137460,14760],"TUNON":[0,1941870,7830]}
ORDER=["MBWAY","ISCOM","IPAC","PIGIER","TUNON"]

wb=openpyxl.Workbook()

# ================= RAPPORT CRM =================
cr=wb.active; cr.title="Rapport CRM"; cr.sheet_view.showGridLines=False
cr["A1"]="RAPPORT CRM  ·  le CA vu du commercial  ·  2026"; cr["A1"].font=F(15,True,INK)
cr["A2"]="« Combien d'étudiants, à quel prix. » Source : Q_CA_CONSTITUTION_CRM (socle). Drill depuis la cellule CA CRM du bandeau réconciliation."; cr["A2"].font=F(9,False,TEALD)
hdr=["Marque / Campus","Effectifs","Tarif moyen","CA scolarité","Nouveaux","CA frais insc.","CA total"]
hr=4
for j,h in enumerate(hdr,1):
    c=cr.cell(hr,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL); c.alignment=CTR if j>1 else LFT
r=hr+1
for m in ORDER:
    camp=CRM[m]; first=r
    for e in sorted(camp):
        eff,scol,new,frais=camp[e]
        cr.cell(r,1,"   "+m.title()+" — "+CITY[e]).font=F(9); cr.cell(r,1).alignment=LFT
        cr.cell(r,2,eff).font=F(9,False,BLUE)
        cr.cell(r,3,f"=IFERROR(D{r}/B{r},0)").font=F(9)          # tarif moyen
        cr.cell(r,4,scol).font=F(9,False,BLUE)
        cr.cell(r,5,new).font=F(9,False,BLUE)
        cr.cell(r,6,frais).font=F(9,False,BLUE)
        cr.cell(r,7,f"=D{r}+F{r}").font=F(9)                      # CA total
        for col,fmt in [(2,NUM),(3,EURc),(4,EUR),(5,NUM),(6,EUR),(7,EUR)]:
            cr.cell(r,col).number_format=fmt; cr.cell(r,col).alignment=RGT
        r+=1
    # sous-total marque
    cr.cell(r,1,m.title()).font=F(10,True,TEALD); cr.cell(r,1).alignment=LFT
    cr.cell(r,2,f"=SUM(B{first}:B{r-1})"); cr.cell(r,3,f"=IFERROR(D{r}/B{r},0)")
    cr.cell(r,4,f"=SUM(D{first}:D{r-1})"); cr.cell(r,5,f"=SUM(E{first}:E{r-1})")
    cr.cell(r,6,f"=SUM(F{first}:F{r-1})"); cr.cell(r,7,f"=SUM(G{first}:G{r-1})")
    for col,fmt in [(2,NUM),(3,EURc),(4,EUR),(5,NUM),(6,EUR),(7,EUR)]:
        cr.cell(r,col).number_format=fmt; cr.cell(r,col).alignment=RGT; cr.cell(r,col).font=F(10,True,TEALD)
    for j in range(1,8): cr.cell(r,j).fill=fill(TEALBG)
    r+=1
# total groupe
cr.cell(r,1,"GROUPE").font=F(10,True,WHITE); cr.cell(r,1).fill=fill(TEALD); cr.cell(r,1).alignment=LFT
subtot=[rr for rr in range(hr+1,r) if cr.cell(rr,1).font.bold]  # les sous-totaux marque
# total groupe = somme des sous-totaux marque
def sums(col): return "+".join(f"{col}{rr}" for rr in subtot)
for col in "BDEFG": cr.cell(r,ord(col)-64,f"={sums(col)}")
cr.cell(r,3,f"=IFERROR(D{r}/B{r},0)")
for col,fmt in [(2,NUM),(3,EURc),(4,EUR),(5,NUM),(6,EUR),(7,EUR)]:
    cr.cell(r,col).number_format=fmt; cr.cell(r,col).alignment=RGT; cr.cell(r,col).font=F(10,True,WHITE); cr.cell(r,col).fill=fill(TEALD)
cr.cell(r+2,1,"Dimensions dispo (multidim) : Marque → Campus → Programme → Cycle → Modalité (INIT/ALT). Ici agrégé Marque/Campus. Contrôle CA total = 22 544 725.").font=F(8,False,FAINT,True)
for col,w in zip("ABCDEFG",[26,10,12,14,10,13,14]): cr.column_dimensions[col].width=w
cr.freeze_panes="A5"

# ================= RAPPORT FINANCE =================
fi=wb.create_sheet("Rapport Finance"); fi.sheet_view.showGridLines=False
fi["A1"]="RAPPORT FINANCE  ·  le même CA vu de la compta  ·  2026"; fi["A1"].font=F(15,True,INK)
fi["A2"]="« Ce que dit le grand livre. » Source : Q_CA_CONSTITUTION_COMPTA. Révèle le MIX DE FINANCEMENT (initiale vs alternance OPCO)."; fi["A2"].font=F(9,False,TEALD)
h2=["Marque","Initiale (706)","Alternance OPCO (7062)","Frais dossier (708)","Total"]
hr=4
for j,h in enumerate(h2,1):
    c=fi.cell(hr,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL); c.alignment=CTR if j>1 else LFT
r=hr+1; first=r
for m in ORDER:
    v=FIN[m]
    fi.cell(r,1,m.title()).font=F(10,True); fi.cell(r,1).alignment=LFT
    fi.cell(r,2,v[0]).font=F(9,False,BLUE); fi.cell(r,3,v[1]).font=F(9,False,BLUE); fi.cell(r,4,v[2]).font=F(9,False,BLUE)
    fi.cell(r,5,f"=B{r}+C{r}+D{r}").font=F(10,True)
    for col in (2,3,4,5): fi.cell(r,col).number_format=EUR; fi.cell(r,col).alignment=RGT
    if v[0]==0:  # 100% alternance -> souligner
        fi.cell(r,2,0).font=F(9,False,OCHRE); fi.cell(r,1).font=F(10,True,OCHRE)
    if r%2==0:
        for j in range(1,6): fi.cell(r,j).fill=fill(CARD2)
    r+=1
fi.cell(r,1,"GROUPE").font=F(10,True,WHITE); fi.cell(r,1).fill=fill(TEALD); fi.cell(r,1).alignment=LFT
for col in (2,3,4,5):
    L=chr(64+col); fi.cell(r,col,f"=SUM({L}{first}:{L}{r-1})")
    fi.cell(r,col).number_format=EUR; fi.cell(r,col).alignment=RGT; fi.cell(r,col).font=F(10,True,WHITE); fi.cell(r,col).fill=fill(TEALD)
    fi.cell(r,col).border=Border(top=med)
fi.cell(r+2,1,"INSIGHT : IPAC, Pigier, Tunon = 100 % alternance (OPCO) — aucune formation initiale (706). MBway/ISCOM ont les deux.").font=F(9,True,OCHRE)
fi.cell(r+3,1,"C'est ce que la lentille Finance ajoute au CRM : la nature du financement. Contrôle Total = 22 544 725.").font=F(8,False,FAINT,True)
fi.cell(r+4,1,"Dimensions dispo (multidim) : Compte × Marque × Campus × Exercice. Drill Marque → Campus → écritures.").font=F(8,False,FAINT,True)
for col,w in zip("ABCDE",[14,14,20,15,15]): fi.column_dimensions[col].width=w
fi.freeze_panes="A5"

out="/home/user/demo5/eduservices/tagetik/MAQUETTE_CRM_FINANCE.xlsx"
wb.save(out); print("SAVED",out)
