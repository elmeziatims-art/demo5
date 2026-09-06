#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""RAPPORT_ALLOUE.xlsx & RAPPORT_EVOLUTION.xlsx — RESTITUTIONS type TCD (croisé,
emboîté, replié) à reproduire dans Tagetik. En tête de chaque feuille : le mode
d'emploi (dimensions à emboîter · éléments/comptes par mesure · source = vue).
Chiffres = cascade fidèle de V_ALLOCATION (EBITDA net 2026 = 3 291 530, écart 0)."""
import csv, glob
from collections import defaultdict, OrderedDict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E2EFEB"; WHITE="FFFFFF"; RULE="C8D2DA"
SOFT="51606D"; OCHRE="B3641C"; OCHRED="8A4A12"; GREEN="1E7A55"; FAINT="7D8B98"; NAVY="3D4F8F"
L0BG="DCE7EE"; L1BG="EAF0F4"; CARD="F5F7F9"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
def LI(i): return Alignment("left",vertical="center",indent=i)
RGT=Alignment("right",vertical="center"); CTR=Alignment("center",vertical="center",wrap_text=True); LFT=Alignment("left",vertical="center")
med=Side(style="medium",color=RULE); thin=Side(style="thin",color=RULE)
EUR='#,##0;-#,##0;"-"'; NUM='#,##0'; PCT='0.0%'; DPT='+0.0;-0.0;0.0'
COMP=glob.glob('/home/user/demo5/eduservices/tgk_data/ext_*/AW_002_000004_000001.csv')[0]
SOCLE=glob.glob('/home/user/demo5/eduservices/tgk_data/ext_*/AW_002_000002_000001.csv')[0]
MQ={'MBWAY':'MBway','ISCOM':'Iscom','IPAC':'Ipac','PIGIER':'Pigier','TUNON':'Tunon'}
ORDER=['MBWAY','ISCOM','IPAC','PIGIER','TUNON']
CITY={'LYO':'Lyon','PAR':'Paris','BOR':'Bordeaux','NAN':'Nantes','MTP':'Montpellier','REN':'Rennes','LIL':'Lille','TLS':'Toulouse'}
MODL={'INIT':'Initial','ALT':'Alternance'}
def hrs(prog,mod):
    if prog.startswith('BAC') and mod=='INIT': return 600
    if prog.startswith('BAC'): return 480
    if prog.startswith('MAS') and mod=='INIT': return 520
    if prog.startswith('MAS'): return 420
    if mod=='INIT': return 1000
    return 700

def compute(year):
    camp=defaultdict(lambda: defaultdict(float)); hold=defaultdict(float)
    with open(COMP,encoding='utf-8') as fh:
        r=csv.reader(fh,delimiter=';'); next(r)
        for row in r:
            en,acc,ex=row[0],row[1],row[2]
            if ex!=year or acc in ('TEC_EBITDA','TEC_PL'): continue
            a=float(row[5].replace(',','.'))
            if en=='GRP':
                if acc in ('6414','6226','626','6281','6331','6333'): hold['HOLDING']+=a
                elif acc=='6236': hold['MARQUE']+=a
                continue
            if acc=='621': camp[en]['VAC']+=a
            elif acc=='6411': camp[en]['PERM']+=a
            elif acc in ('604','6063'): camp[en]['ODIR']+=a
            elif acc=='6231': camp[en]['MKT']+=a
            elif acc in ('6413','645','613','615','616','625','63511'): camp[en]['STRUCT']+=a
    cls=[]
    with open(SOCLE,encoding='utf-8') as fh:
        r=csv.reader(fh,delimiter=';'); h=next(r); ix={n:i for i,n in enumerate(h)}
        g=lambda row,n: float(row[ix[n]].replace(',','.'))
        for row in r:
            if row[ix['EXERCICE']]!=year: continue
            en=row[ix['ENTITY']]; cls.append(dict(en=en,mq=en.split('_')[0],prog=row[ix['PROGRAMME']],an=row[ix['AN_ETUDE']],
                mod=row[ix['MODALITE']],eff=g(row,'VOL_EFF'),vcl=g(row,'VOL_CLASS'),new=g(row,'VOL_NEW'),
                ca=g(row,'VOL_EFF')*g(row,'REV_STUD')+g(row,'VOL_NEW')*g(row,'REV_FRAIS_INS'),
                hrs=g(row,'VOL_CLASS')*hrs(row[ix['PROGRAMME']],row[ix['MODALITE']])))
    E=defaultdict(lambda: defaultdict(float)); M=defaultdict(lambda: defaultdict(float)); G=defaultdict(float)
    for c in cls:
        for k,fk in [('HRS','hrs'),('EFF','eff'),('NEW','new'),('CA','ca')]:
            E[c['en']][k]+=c[fk]; M[c['mq']][k]+=c[fk]; G[k]+=c[fk]
    out=[]
    for c in cls:
        en,mq=c['en'],c['mq']; cp=camp[en]
        cvac=cp['VAC']*c['hrs']/E[en]['HRS'] if E[en]['HRS'] else 0
        cperm=cp['PERM']*c['hrs']/E[en]['HRS'] if E[en]['HRS'] else 0
        codir=(cp['ODIR']*c['eff']/E[en]['EFF'] if E[en]['EFF'] else 0)+(cp['MKT']*c['new']/E[en]['NEW'] if E[en]['NEW'] else 0)
        cstr=cp['STRUCT']*c['eff']/E[en]['EFF'] if E[en]['EFF'] else 0
        share=(M[mq]['EFF']/G['EFF'])*(E[en]['CA']/M[mq]['CA'] if M[mq]['CA'] else 0)*(c['eff']/E[en]['EFF'] if E[en]['EFF'] else 0)
        siege=(hold['MARQUE']+hold['HOLDING'])*share
        propre=c['ca']-(cvac+cperm+codir+cstr); net=propre-siege
        out.append(dict(mq=MQ[mq],campus=CITY.get(en.split('_')[1],en.split('_')[1]),prog=c['prog'],an=c['an'],mod=MODL.get(c['mod'],c['mod']),
            eff=c['eff'],ca=c['ca'],propre=propre,siege=siege,net=net))
    return out

def tree5(rows):
    T=OrderedDict()
    for m in ORDER:
        mm=MQ[m]; sub=[r for r in rows if r['mq']==mm]
        if not sub: continue
        T[mm]=OrderedDict()
        for camp in sorted(set(r['campus'] for r in sub)):
            T[mm][camp]=OrderedDict()
            for prog in sorted(set(r['prog'] for r in sub if r['campus']==camp)):
                T[mm][camp][prog]=OrderedDict()
                for an in sorted(set(r['an'] for r in sub if r['campus']==camp and r['prog']==prog)):
                    T[mm][camp][prog][an]=OrderedDict()
                    for r in [x for x in sub if x['campus']==camp and x['prog']==prog and x['an']==an]:
                        T[mm][camp][prog][an][r['mod']]=r
    return T
def agg(node):
    tot=dict(eff=0,ca=0,propre=0,siege=0,net=0)
    def walk(n):
        for v in n.values():
            if 'eff' in v and 'ca' in v and not isinstance(next(iter(v.values()),None),dict):
                for k in tot: tot[k]+=v[k]
            elif isinstance(v,dict): walk(v)
    # simpler: recurse to leaves
    def rec(n):
        if 'net' in n:
            for k in tot: tot[k]+=n[k]
            return
        for v in n.values(): rec(v)
    rec(node); return tot

def docblock(ws,title,source,dims,measures):
    ws.sheet_view.showGridLines=False
    ws["A1"]=title; ws["A1"].font=F(15,True,INK)
    ws["A2"]="MODE D'EMPLOI TAGETIK — à reproduire en croisé dynamique"; ws["A2"].font=F(9,True,TEALD)
    r=3
    ws.cell(r,1,"Source (vue)"); ws.cell(r,1).font=F(9,True,OCHRED); ws.cell(r,2,source).font=F(9,False,INK); r+=1
    ws.cell(r,1,"Dimensions (lignes, à emboîter)"); ws.cell(r,1).font=F(9,True,OCHRED); ws.cell(r,2,dims).font=F(9,False,INK); r+=1
    ws.cell(r,1,"Éléments par mesure (comptes)"); ws.cell(r,1).font=F(9,True,OCHRED); r+=1
    for lab,comp in measures:
        ws.cell(r,1,"   "+lab).font=F(8.5,True,TEALD); ws.cell(r,2,comp).font=F(8.5,False,SOFT); r+=1
    for j in range(1,8): ws.cell(r,j).border=Border(bottom=med)
    return r+1

wb=openpyxl.Workbook()

# ===================== RAPPORT_ALLOUE (TCD 2026) =====================
ws=wb.active; ws.title="Restitution Alloué 2026"
hr=docblock(ws,"RESTITUTION — P&L chargé 2026 (avant / après allocation)","V_ALLOCATION  (datasource Q_RAPPORT_ALLOUE)",
    "Marque ▸ Campus ▸ Programme ▸ Année ▸ Modalité",
    [("Effectif","VOL_EFF"),
     ("CA","706 + 7062 + 708"),
     ("EBITDA propre","CA − (621 + 6411 + 604/6063/6231 + 6413/645/613/615/616/625/63511)"),
     ("Quote-part siège","6236 + 6414/6226/626/6281/6331/6333"),
     ("EBITDA net","CA − tous les coûts ci-dessus   (dotations 6811 exclues)"),
     ("Marge %","EBITDA net ÷ CA   (membre calculé)")])
ws.sheet_properties.outlinePr.summaryBelow=False
head=["Marque ▸ Campus ▸ Programme ▸ Année ▸ Modalité","Effectif","CA","EBITDA propre","Q-part siège","EBITDA net","Marge %"]
for j,h in enumerate(head,1):
    c=ws.cell(hr,j,h); c.font=F(8.5,True,WHITE); c.fill=fill(TEAL); c.alignment=LFT if j==1 else CTR; c.border=Border(bottom=med)
r=hr+1
def emit(ws,r,label,vals,lvl):
    ws.cell(r,1,label).font=F(9 if lvl<=1 else 8.5, lvl<=1); ws.cell(r,1).alignment=LI(lvl)
    ws.cell(r,2,round(vals['eff'])).number_format=NUM
    ws.cell(r,3,round(vals['ca'])).number_format=EUR
    ws.cell(r,4,round(vals['propre'])).number_format=EUR
    ws.cell(r,5,round(vals['siege'])).number_format=EUR
    mg=vals['net']/vals['ca'] if vals['ca'] else 0
    ws.cell(r,6,round(vals['net'])).number_format=EUR; ws.cell(r,6).font=F(9 if lvl<=1 else 8.5, lvl<=1, OCHRED if mg<0.05 else (TEALD if lvl<=1 else INK))
    ws.cell(r,7,mg).number_format=PCT; ws.cell(r,7).font=F(9 if lvl<=1 else 8.5,lvl<=1,OCHRED if mg<0.05 else TEALD)
    for j in (2,3,4,5,6,7): ws.cell(r,j).alignment=RGT
    if lvl==0:
        for j in range(1,8): ws.cell(r,j).fill=fill(L0BG); ws.cell(r,j).border=Border(top=med,bottom=thin)
    elif lvl==1:
        for j in range(1,8): ws.cell(r,j).fill=fill(L1BG)
    else:
        for j in range(1,8): ws.cell(r,j).border=Border(bottom=thin)
    if lvl>0:
        ws.row_dimensions[r].outline_level=min(lvl,4)
        if lvl>=2: ws.row_dimensions[r].hidden=True    # replié : montre Marque + Campus
def walk_emit(ws,r,node,lvl):
    for key,child in node.items():
        if 'net' in child:  # feuille (modalité)
            emit(ws,r,key,child,lvl); r+=1
        else:
            emit(ws,r,key,agg(child),lvl); r+=1
            r=walk_emit(ws,r,child,lvl+1)
    return r
T=tree5(compute('2026'))
r=walk_emit(ws,r,T,0)
# total groupe
g=dict(eff=0,ca=0,propre=0,siege=0,net=0)
for mq in T.values():
    a=agg(mq)
    for k in g: g[k]+=a[k]
ws.cell(r,1,"GROUPE").font=F(11,True,TEAL); ws.cell(r,1).alignment=LI(0)
ws.cell(r,2,round(g['eff'])).number_format=NUM; ws.cell(r,3,round(g['ca'])).number_format=EUR
ws.cell(r,4,round(g['propre'])).number_format=EUR; ws.cell(r,5,round(g['siege'])).number_format=EUR
ws.cell(r,6,round(g['net'])).number_format=EUR; ws.cell(r,7,g['net']/g['ca']).number_format=PCT
for j in (2,3,4,5,6,7): ws.cell(r,j).alignment=RGT
for j in range(1,8): ws.cell(r,j).font=F(11,True,TEAL); ws.cell(r,j).fill=fill(TEALBG); ws.cell(r,j).border=Border(top=med,bottom=med)
ws.column_dimensions['A'].width=44
for col,w in zip("BCDEFG",[10,13,13,12,13,9]): ws.column_dimensions[col].width=w
ws.freeze_panes=f"A{hr+1}"

# ===================== RAPPORT_EVOLUTION (TCD 2024-2026) =====================
w2=wb.create_sheet("Restitution Évolution")
hr2=docblock(w2,"RESTITUTION — Évolution de la marge chargée 2024→2026","V_ALLOCATION  (datasource Q_RAPPORT_EVOLUTION)",
    "Marque ▸ Campus ▸ Programme ▸ Année ▸ Modalité   ·   EXERCICE en colonnes",
    [("CA","706 + 7062 + 708"),
     ("EBITDA net","CA − (621+6411 + 604/6063/6231 + 6413/645/613/615/616/625/63511 + 6236 + 6414/6226/626/6281/6331/6333)"),
     ("Marge nette %","EBITDA net ÷ CA   (membre calculé, par exercice)"),
     ("Δ (points)","Marge 2026 − Marge 2024   (comparaison de colonnes)")])
w2.sheet_properties.outlinePr.summaryBelow=False
head2=["Marque ▸ Campus ▸ Programme ▸ Année ▸ Modalité","CA 2026","Marge 2024","Marge 2025","Marge 2026","Δ 26/24 (pt)"]
for j,h in enumerate(head2,1):
    c=w2.cell(hr2,j,h); c.font=F(8.5,True,WHITE); c.fill=fill(TEAL); c.alignment=LFT if j==1 else CTR; c.border=Border(bottom=med)
# calcule les 3 années et indexe par maille (mq,campus,prog,an,mod)
data={y:compute(y) for y in ('2024','2025','2026')}
def key(r): return (r['mq'],r['campus'],r['prog'],r['an'],r['mod'])
idx={y:{key(r):r for r in data[y]} for y in data}
T26=tree5(data['2026'])
def marg(node_leaves_year):  # returns dict of CA/net summed for a set of leaf keys in a given year
    pass
def collect_keys(node):
    ks=[]
    def rec(n):
        if 'net' in n: ks.append(key(n)); return
        for v in n.values(): rec(v)
    rec(node); return ks
def emit2(w2,r,label,node,lvl):
    ks=collect_keys(node) if 'net' not in node else [key(node)]
    def mg(y):
        ca=sum(idx[y][k]['ca'] for k in ks if k in idx[y]); net=sum(idx[y][k]['net'] for k in ks if k in idx[y])
        return (net/ca) if ca else None
    m24,m25,m26=mg('2024'),mg('2025'),mg('2026')
    ca26=sum(idx['2026'][k]['ca'] for k in ks if k in idx['2026'])
    w2.cell(r,1,label).font=F(9 if lvl<=1 else 8.5,lvl<=1); w2.cell(r,1).alignment=LI(lvl)
    w2.cell(r,2,round(ca26)).number_format=EUR; w2.cell(r,2).alignment=RGT
    for j,mv in zip((3,4,5),(m24,m25,m26)):
        if mv is not None:
            c=w2.cell(r,j,mv); c.number_format=PCT; c.alignment=RGT
            c.font=F(9 if lvl<=1 else 8.5, lvl<=1 and j==5, OCHRED if mv<0.05 else (TEALD if j==5 else SOFT))
    if m24 is not None and m26 is not None:
        c=w2.cell(r,6,(m26-m24)*100); c.number_format=DPT; c.alignment=RGT; c.font=F(8.5,False,TEALD if m26>=m24 else OCHRE)
    if lvl==0:
        for j in range(1,7): w2.cell(r,j).fill=fill(L0BG); w2.cell(r,j).border=Border(top=med,bottom=thin)
    elif lvl==1:
        for j in range(1,7): w2.cell(r,j).fill=fill(L1BG)
    else:
        for j in range(1,7): w2.cell(r,j).border=Border(bottom=thin)
    if lvl>0:
        w2.row_dimensions[r].outline_level=min(lvl,4)
        if lvl>=2: w2.row_dimensions[r].hidden=True
def walk2(w2,r,node,lvl):
    for k,child in node.items():
        emit2(w2,r,k,child,lvl); r+=1
        if 'net' not in child: r=walk2(w2,r,child,lvl+1)
    return r
r=hr2+1
r=walk2(w2,r,T26,0)
w2.column_dimensions['A'].width=44
for col,w in zip("BCDEF",[13,11,11,11,12]): w2.column_dimensions[col].width=w
w2.freeze_panes=f"A{hr2+1}"

out="/home/user/demo5/eduservices/tagetik/RAPPORT_RESTITUTION.xlsx"
wb.save(out)
# contrôle footing
g_net=sum(agg(mq)['net'] for mq in T.values()); g_ca=sum(agg(mq)['ca'] for mq in T.values())
print("SAVED",out)
print(f"  ALLOUÉ 2026 : EBITDA net groupe = {g_net:,.0f}  (cible 3 291 530)  marge {g_net/g_ca*100:.1f}%")
