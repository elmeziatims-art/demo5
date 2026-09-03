#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""DRILL_EBITDA.xlsx — le rapport qui démontre le DRILL, juste après le cockpit.
Chemin : Marque -> Campus -> Programme -> Année d'étude -> Modalité.
Onglet « Données (V_ALLOCATION) » : grain CLASSE, qui remonte EXACTEMENT au socle campus du cockpit.
Onglet « Drill » : 5 listes déroulantes, le niveau courant, le niveau suivant, le détail au grain classe.
À chaque niveau : contribution ET marge complète — elles divergent au grain classe (piège de la fermeture)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

BLUE="2A78D6"; BLUE2="7AABE6"; BLUE3="C3D9F4"; BLUESOFT="DCE9F9"
ORANGE="EB6834"; ORANGESOFT="FBE3D8"; GOOD="0CA30C"; WARN="FAB219"; CRIT="D03B3B"
INK="131922"; INK2="4D5866"; INK3="7C8798"
CANVAS="EEF1F6"; PANEL="FFFFFF"; PANEL2="F7F9FC"; LINE="DFE4EC"; LINE2="EAEEF4"; WHITE="FFFFFF"
UI="Segoe UI"; MONO="Consolas"
def F(sz=10,b=False,c=INK,f=UI): return Font(name=f,size=sz,bold=b,color=c)
def fill(c): return PatternFill("solid",fgColor=c)
def sd(c=LINE): return Side(style="thin",color=c)
L=Alignment("left",vertical="center"); R=Alignment("right",vertical="center")
Cn=Alignment("center",vertical="center"); TOP=Alignment("left",vertical="top",wrap_text=True)
EUR='#,##0'; PCT='0.0%'; PCT0='0%'; DPCT='"▲ "0.0%;"▼ "0.0%'
def box(ws,r1,c1,r2,c2,bg=PANEL,bd=LINE):
    for r in range(r1,r2+1):
        for c in range(c1,c2+1):
            x=ws.cell(r,c); x.fill=fill(bg)
            x.border=Border(top=sd(bd) if r==r1 else None,bottom=sd(bd) if r==r2 else None,
                            left=sd(bd) if c==c1 else None,right=sd(bd) if c==c2 else None)

# =====================================================================================
# 1. LE GRAIN CLASSE — construit pour remonter au socle campus du cockpit
# =====================================================================================
CAP={"BAC":32,"MAS":26,"BTS":30}
HEURES={"BAC":520,"MAS":450,"BTS":560}          # heures d'enseignement par classe et par an
# en initial la famille paie le prix fort ; en alternance le financement OPCO est plafonné
PRIX={("BAC","ALT"):8600,("MAS","ALT"):9800,("BTS","ALT"):7900,
      ("BAC","INIT"):9400,("MAS","INIT"):10500,("BTS","INIT"):8200}
COUT_VAR_ELEVE=363                               # calibré sur la donnée réelle
SIEGE_GROUPE=1250000                             # holding + marque (5,4 % du CA), réaffecté aux classes

# campus : (marque, CA25, CA26, EBITDA25, EBITDA26)  — identiques au socle du cockpit
CAMPUS={
 "MBWAY_PAR":("MBway",3920664,4250000, 678784, 812000),
 "MBWAY_LYO":("MBway",2974742,3180000, 479129, 528000),
 "MBWAY_BOR":("MBway",2149621,2270000, 325885, 350000),
 "ISCOM_PAR":("Iscom",3783019,4010000, 645000, 690000),
 "ISCOM_LIL":("Iscom",2498727,2591085, 363500, 391990),
 "PIGIER_BOR":("Pigier",2462687,2640000, 488000, 545000),
 "PIGIER_LYO":("Pigier",2173095,2310000, 410500, 454900),
 "TUNON_PAR":("Tunon",1069253,1102400, 106800,  96300),
 "TUNON_LYO":("Tunon", 744012, 745500, -12780, -22400),
}
# classes : entity -> [(programme, année, modalité, nb_classes, effectif)]
CLASSES={
 "MBWAY_PAR":[("BAC_MGT","B1","ALT",3,92),("BAC_MGT","B1","INIT",1,30),("BAC_MGT","B2","ALT",4,120),
              ("BAC_MGT","B3","ALT",3,91),("MAS_MGT","M1","ALT",3,74),("MAS_MGT","M2","ALT",2,49),
              ("MAS_MKT","M1","ALT",1,22)],
 "MBWAY_LYO":[("BAC_MGT","B1","ALT",3,84),("BAC_MGT","B2","ALT",3,83),("BAC_MGT","B3","ALT",2,55),
              ("MAS_MGT","M1","ALT",3,68),("MAS_MGT","M2","ALT",3,67)],
 "MBWAY_BOR":[("BAC_MGT","B1","ALT",3,80),("BAC_MGT","B2","ALT",3,76),("BAC_MGT","B3","ALT",2,52),
              ("MAS_MGT","M1","ALT",2,47)],
 "ISCOM_PAR":[("BAC_COM","B1","INIT",2,58),("BAC_COM","B1","ALT",2,58),("BAC_COM","B2","ALT",3,86),
              ("BAC_COM","B3","ALT",3,84),("MAS_COM","M1","ALT",4,92),("MAS_COM","M2","ALT",3,73)],
 "ISCOM_LIL":[("BAC_COM","B1","INIT",2,52),("BAC_COM","B2","ALT",3,78),("BAC_COM","B3","ALT",2,52),
              ("MAS_COM","M1","ALT",3,63),("MAS_COM","M2","ALT",2,46)],
 "PIGIER_BOR":[("BTS_GES","BTS1","ALT",3,85),("BTS_GES","BTS2","ALT",3,82),("BAC_RH","B1","ALT",2,40),
               ("BAC_RH","B2","ALT",2,60),("BAC_RH","B3","ALT",1,30)],
 "PIGIER_LYO":[("BTS_GES","BTS1","ALT",3,71),("BTS_GES","BTS2","ALT",2,50),("BAC_GES","B1","ALT",2,54),
               ("BAC_GES","B2","ALT",2,53),("BAC_GES","B3","ALT",1,32)],
 "TUNON_PAR":[("BAC_TOU","B1","INIT",2,48),("BAC_TOU","B2","ALT",1,27),("BAC_TOU","B3","ALT",1,26),
              ("BTS_TOU","BTS1","ALT",1,23)],
 "TUNON_LYO":[("BAC_TOU","B1","INIT",1,28),("BAC_TOU","B2","ALT",1,24),("BAC_TOU","B3","ALT",1,19),
              ("BTS_TOU","BTS1","ALT",1,13)],
}
cyc=lambda p: p.split("_")[0]
EFF_TOT=sum(e for v in CLASSES.values() for *_,e in v)

rows=[]
for ent,cls in CLASSES.items():
    marque,ca25c,ca26c,eb25c,eb26c=CAMPUS[ent]
    brut=[e*PRIX[(cyc(p),m)] for p,a,m,n,e in cls]
    k=ca26c/sum(brut)
    ca26=[b*k for b in brut]
    var=[e*COUT_VAR_ELEVE for *_,e in [(p,a,m,n,e) for p,a,m,n,e in cls]]
    # coût direct : réparti aux HEURES (une classe coûte une classe, quel que soit son effectif)
    poids=[n*HEURES[cyc(p)] for p,a,m,n,e in cls]
    direct_campus=ca26c-sum(var)-eb26c
    direct=[direct_campus*w/sum(poids) for w in poids]
    # CA 2025 : croissance par classe autour de celle du campus, puis recalage exact
    g=[1+0.012*((i%3)-1) for i in range(len(cls))]
    b25=[c/gi for c,gi in zip(ca26,g)]; k25=ca25c/sum(b25); ca25=[b*k25 for b in b25]
    eb26=[c-v-d for c,v,d in zip(ca26,var,direct)]
    b=[x*(1+0.03*((i%3)-1)) for i,x in enumerate(eb26)]
    kb=eb25c/sum(b); eb25=[x*kb for x in b]
    for (p,a,m,n,e),c26,c25,v,d,x26,x25 in zip(cls,ca26,ca25,var,direct,eb26,eb25):
        rows.append(dict(marque=marque,ent=ent,prog=p,an=a,mod=m,ncl=n,cap=CAP[cyc(p)],eff=e,
                         ca25=round(c25),ca26=round(c26),var=round(v),direct=round(d),
                         siege=round(SIEGE_GROUPE*e/EFF_TOT),eb25=round(x25)))
# contrôles
def chk(lbl,got,want):
    ok="OK " if abs(got-want)<=len(rows) else "ECART"
    print("  %-28s %14s  cible %14s  %s"%(lbl,f"{got:,.0f}",f"{want:,.0f}",ok))
print("CONTRÔLE DE REMONTÉE classe -> campus -> groupe")
chk("CA 2026",sum(r['ca26'] for r in rows),23098985)
chk("CA 2025",sum(r['ca25'] for r in rows),21775820)
chk("EBITDA 2026",sum(r['ca26']-r['var']-r['direct'] for r in rows),3845790)
chk("EBITDA 2025",sum(r['eb25'] for r in rows),3484818)
print("  effectifs %d  ·  classes %d  ·  lignes %d"%(EFF_TOT,sum(r['ncl'] for r in rows),len(rows)))
print("\nCLASSES OÙ CONTRIBUTION > 0 MAIS MARGE COMPLÈTE < 0  (le piège) :")
for r in rows:
    contrib=r['ca26']-r['var']; mc=r['ca26']-r['var']-r['direct']-r['siege']
    if contrib>0 and mc<0:
        print("  %-11s %-8s %-5s %-5s  eff %3d/%3d (%2.0f%%)  contrib %+9s  marge compl. %+8s"
              %(r['ent'],r['prog'],r['an'],r['mod'],r['eff'],r['ncl']*r['cap'],
                100*r['eff']/(r['ncl']*r['cap']),f"{contrib:,.0f}",f"{mc:,.0f}"))
import json,os
json.dump(rows,open("/tmp/drill_rows.json","w"))

# =====================================================================================
# 2. LE CLASSEUR
# =====================================================================================
MARQUES=["MBway","Iscom","Pigier","Tunon"]
ENTS=list(CLASSES.keys())
PROGS=sorted({r['prog'] for r in rows}); ANS=["B1","B2","B3","M1","M2","BTS1","BTS2"]; MODS=["ALT","INIT"]
DIMS=["Marque","Campus","Programme","Année d'étude","Modalité"]
LISTS=[MARQUES,ENTS,PROGS,ANS,MODS]
NR=len(rows); R1,R2=5,5+NR-1                      # lignes de données
D=lambda c: "Données!$%s$%d:$%s$%d"%(c,R1,c,R2)

wb=openpyxl.Workbook()
# ---------------------------------------------------------------- onglet Données
d=wb.active; d.title="Données"; d.sheet_view.showGridLines=False
for col,wd in zip("ABCDEFGHIJKLMNOPQRSTU",[9,12,11,7,7,6,6,7,7,12,12,11,12,11,12,12,12,12,13,7,6]):
    d.column_dimensions[col].width=wd
d.cell(1,1,"DONNÉES DU DRILL — grain CLASSE  (source : V_ALLOCATION / V_CAMPUS_CLASSE)").font=F(12,True,WHITE)
for c in range(1,22): d.cell(1,c).fill=fill(BLUE)
d.row_dimensions[1].height=24
d.cell(3,1,"Une ligne = une cohorte-classe. Tout remonte exactement au socle campus du cockpit.").font=F(9.5,True,BLUE)
HD=["Marque","Campus","Programme","Année","Modalité","Nb cl.","Capa.","Places","Effectif",
    "CA 2025","CA 2026","Coût variable","Coût direct","Coût siège","EBITDA 2025",
    "EBITDA 2026","Contribution","Coût complet","Marge complète","Inclus","Rang"]
for j,t in enumerate(HD,1):
    c=d.cell(4,j,t); c.font=F(8.5,True,WHITE); c.fill=fill(INK2); c.alignment=Cn
    c.border=Border(top=sd(INK2),bottom=sd(INK2),left=sd(INK2),right=sd(INK2))
for i,r in enumerate(rows):
    rr=R1+i
    vals=[r['marque'],r['ent'],r['prog'],r['an'],r['mod'],r['ncl'],r['cap'],
          "=F%d*G%d"%(rr,rr),r['eff'],r['ca25'],r['ca26'],r['var'],r['direct'],r['siege'],r['eb25'],
          "=K%d-L%d-M%d"%(rr,rr,rr),"=K%d-L%d"%(rr,rr),"=L%d+M%d+N%d"%(rr,rr,rr),"=K%d-R%d"%(rr,rr),
          "=IF(AND(OR(Drill!$C$6=\"(Tous)\",$A{0}=Drill!$C$6),OR(Drill!$D$6=\"(Tous)\",$B{0}=Drill!$D$6),"
          "OR(Drill!$E$6=\"(Tous)\",$C{0}=Drill!$E$6),OR(Drill!$F$6=\"(Tous)\",$D{0}=Drill!$F$6),"
          "OR(Drill!$G$6=\"(Tous)\",$E{0}=Drill!$G$6)),1,0)".format(rr),
          "=IF(T%d=1,SUM($T$%d:T%d),\"\")"%(rr,R1,rr)]
    for j,v in enumerate(vals,1):
        c=d.cell(rr,j,v); c.font=F(9); c.border=Border(bottom=sd(LINE2))
        c.alignment=L if j<=5 else R
        c.number_format=EUR if j in (10,11,12,13,14,15,16,17,18,19) else '#,##0'
        if j in (20,21): c.font=F(8,c=INK3)
tot=R2+1
d.cell(tot,1,"TOTAL").font=F(9,True)
for j in (8,9,10,11,12,13,14,15,16,17,18,19):
    c=d.cell(tot,j,"=SUM(%s%d:%s%d)"%(get_column_letter(j),R1,get_column_letter(j),R2))
    c.font=F(9,True); c.number_format=EUR; c.alignment=R; c.border=Border(top=sd(INK3))

# listes de membres (drill) et listes de validation (avec « (Tous) »)
d.cell(4,23,"— membres par niveau —").font=F(8,True,INK3)
for k,(nm,lst) in enumerate(zip(DIMS,LISTS)):
    d.cell(4,29+k,nm).font=F(8,True,INK3)
    for i,v in enumerate(lst): d.cell(R1+i,23+k,v).font=F(8,c=INK3)
    d.cell(R1,31+k+0,None)
for k,(nm,lst) in enumerate(zip(DIMS,LISTS)):
    d.cell(R1-1,31+k,nm+" (validation)").font=F(8,True,INK3)
    d.cell(R1,31+k,"(Tous)").font=F(8,c=INK3)
    for i,v in enumerate(lst): d.cell(R1+1+i,31+k,v).font=F(8,c=INK3)
for k,nm in enumerate(DIMS): d.cell(R1+k,29,nm).font=F(8,c=INK3)   # AC5:AC9 = noms des niveaux

# ---------------------------------------------------------------- onglet Drill
w=wb.create_sheet("Drill"); w.sheet_view.showGridLines=False
w.column_dimensions["A"].width=2; w.column_dimensions["B"].width=34
for col in "CDEFGHIJKLM": w.column_dimensions[col].width=12
w.column_dimensions["N"].width=3; w.column_dimensions["N"].hidden=True
w.column_dimensions["O"].width=3; w.column_dimensions["O"].hidden=True
for r,hh in {1:6,2:28,3:16,4:8,5:15,6:22,7:8,8:18,9:8,10:20,11:18,12:24,13:10,14:20,15:18,29:10,30:20,31:18,47:10}.items():
    w.row_dimensions[r].height=hh
for r in list(range(16,29))+list(range(32,47)): w.row_dimensions[r].height=16
for r in range(1,56):
    for c in range(1,16): w.cell(r,c).fill=fill(CANVAS)

box(w,2,2,3,13)
w.cell(2,2,"  Drill EBITDA — du groupe à la classe").font=F(14,True,INK); w.cell(2,2).alignment=L
w.cell(3,2,"  Marque → Campus → Programme → Année d'étude → Modalité · exercice 2026").font=F(9.5,c=INK3)
w.cell(3,2).alignment=L
w.cell(2,13,"CONTRIBUTION vs COÛT COMPLET  ").font=F(9,True,ORANGE); w.cell(2,13).alignment=R

# --- ① sélection
box(w,5,2,6,13)
w.cell(5,2,"  ① CHEMIN DE DRILL").font=F(9,True,INK3); w.cell(5,2).alignment=L
w.cell(6,2,"  remets « (Tous) » pour remonter d'un niveau").font=F(8.5,c=INK3); w.cell(6,2).alignment=L
for k,nm in enumerate(DIMS):
    cc=3+k
    lb=w.cell(5,cc,nm); lb.font=F(8.5,True,INK3); lb.alignment=Cn
    sel=w.cell(6,cc,"(Tous)"); sel.font=F(10,True,BLUE); sel.alignment=Cn
    sel.fill=fill(BLUESOFT); sel.border=Border(top=sd(BLUE),bottom=sd(BLUE),left=sd(BLUE),right=sd(BLUE))
    col=get_column_letter(31+k)
    dv=DataValidation(type="list",formula1="=Données!$%s$%d:$%s$%d"%(col,R1,col,R1+len(LISTS[k])),allow_blank=False)
    w.add_data_validation(dv); dv.add(sel)
w.cell(6,15,"=5-COUNTIF($C$6:$G$6,\"(Tous)\")+1")           # O6 : niveau courant
w.cell(6,15).font=F(8,c=INK3)
w.cell(8,2,'="②  Vous regardez : "&IF($O$6=1,"le GROUPE (tous périmètres)",'
           'IF($C$6="(Tous)","",$C$6)&IF($D$6="(Tous)",""," › "&$D$6)&IF($E$6="(Tous)",""," › "&$E$6)'
           '&IF($F$6="(Tous)",""," › "&$F$6)&IF($G$6="(Tous)",""," › "&$G$6))').font=F(11,True,INK)
w.cell(8,2).alignment=L

MES=["CA 2026","Δ CA","EBITDA","Marge EBITDA","Contribution","Coût complet","Marge complète",
     "Effectifs","Places","Remplissage","Statut"]
FMT=[EUR,DPCT,EUR,PCT,EUR,EUR,EUR,'#,##0','#,##0',PCT0,"General"]

# --- ② niveau courant
box(w,10,2,12,13)
w.cell(10,2,"  ② NIVEAU COURANT").font=F(9,True,INK3); w.cell(10,2).alignment=L
for j,t in enumerate(MES,3):
    c=w.cell(11,j,t); c.font=F(8.5,True,INK3); c.alignment=R; c.border=Border(bottom=sd(LINE))
w.cell(11,2,"Agrégat de la sélection").font=F(8.5,True,INK3); w.cell(11,2).alignment=L
w.cell(11,2).border=Border(bottom=sd(LINE))
SP=lambda col: "SUMPRODUCT(%s,%s)"%(D("T"),D(col))
cur={"C":"="+SP("K"),"D":"=C12/%s-1"%SP("J"),"E":"="+SP("P"),"F":"=IFERROR(E12/C12,0)",
     "G":"="+SP("Q"),"H":"="+SP("R"),"I":"="+SP("S"),"J":"="+SP("I"),"K":"="+SP("H"),
     "L":"=IFERROR(J12/K12,0)","M":'=IF(E12<0,"DÉFICITAIRE",IF(I12<0,"PIÈGE",IF(L12<0.75,"SOUS-REMPLI","SAIN")))'}
for col,f in cur.items():
    j=openpyxl.utils.column_index_from_string(col)
    c=w.cell(12,j,f); c.font=F(11,True,INK,MONO if col!="M" else UI)
    c.number_format=FMT[j-3]; c.alignment=R if col!="M" else Cn
w.cell(12,2,'=IF($O$6=1,"GROUPE",INDEX($C$6:$G$6,$O$6-1))').font=F(11,True,INK); w.cell(12,2).alignment=L

# --- ③ niveau suivant
box(w,14,2,28,13)
w.cell(14,2,'="  ③ NIVEAU SUIVANT — "&IF($O$6>5,"grain classe atteint, voir le détail ci-dessous",'
            'UPPER(INDEX(Données!$AC$5:$AC$9,$O$6)))').font=F(9,True,INK3)
w.cell(14,2).alignment=L
for j,t in enumerate(MES,3):
    c=w.cell(15,j,t); c.font=F(8.5,True,INK3); c.alignment=R; c.border=Border(bottom=sd(LINE))
w.cell(15,2,"Membre").font=F(8.5,True,INK3); w.cell(15,2).alignment=L; w.cell(15,2).border=Border(bottom=sd(LINE))
MASK="%s*(INDEX(Données!$A$%d:$E$%d,0,$O$6)=$B{r})"%(D("T"),R1,R2)
for i in range(12):
    r=16+i
    w.cell(r,2,'=IF($O$6>5,"",IFERROR(INDEX(Données!$W$%d:$AA$%d,%d,$O$6),""))'%(R1,R1+11,i+1)).font=F(9.5,True)
    w.cell(r,2).alignment=L
    m=MASK.format(r=r)
    nf={"C":"=SUMPRODUCT(%s*%s)"%(m,D("K")),"D":"=C{r}/SUMPRODUCT({m}*{j})-1".format(r=r,m=m,j=D("J")),
        "E":"=SUMPRODUCT(%s*%s)"%(m,D("P")),"F":"=E{r}/C{r}".format(r=r),
        "G":"=SUMPRODUCT(%s*%s)"%(m,D("Q")),"H":"=SUMPRODUCT(%s*%s)"%(m,D("R")),
        "I":"=SUMPRODUCT(%s*%s)"%(m,D("S")),"J":"=SUMPRODUCT(%s*%s)"%(m,D("I")),
        "K":"=SUMPRODUCT(%s*%s)"%(m,D("H")),"L":"=J{r}/K{r}".format(r=r),
        "M":'=IF(E{r}<0,"DÉFICITAIRE",IF(I{r}<0,"PIÈGE",IF(L{r}<0.75,"SOUS-REMPLI","")))'.format(r=r)}
    for col,f in nf.items():
        j=openpyxl.utils.column_index_from_string(col)
        guard=('$B%d=""'%r) if col=="C" else ('OR($C%d="",$C%d=0)'%(r,r))
        c=w.cell(r,j,'=IF(%s,"",%s)'%(guard,f[1:]))
        c.font=F(9.5,False,INK,MONO if col!="M" else UI); c.number_format=FMT[j-3]
        c.alignment=R if col!="M" else Cn
    for j in range(2,14): w.cell(r,j).border=Border(bottom=sd(LINE2))

# --- ④ détail au grain classe
box(w,30,2,46,13)
w.cell(30,2,"  ④ DÉTAIL AU GRAIN CLASSE — le fond du drill").font=F(9,True,INK3); w.cell(30,2).alignment=L
for j,t in enumerate(MES,3):
    c=w.cell(31,j,t); c.font=F(8.5,True,INK3); c.alignment=R; c.border=Border(bottom=sd(LINE))
w.cell(31,2,"Campus · Programme · Année · Modalité").font=F(8.5,True,INK3)
w.cell(31,2).alignment=L; w.cell(31,2).border=Border(bottom=sd(LINE))
for i in range(15):
    r=32+i
    w.cell(r,14,"=IFERROR(MATCH(%d,%s,0),\"\")"%(i+1,D("U"))).font=F(7,c=INK3)
    w.cell(r,2,'=IF($N{r}="","",INDEX({b},$N{r})&" · "&INDEX({c},$N{r})&" · "&INDEX({d},$N{r})&" · "&INDEX({e},$N{r}))'
           .format(r=r,b=D("B"),c=D("C"),d=D("D"),e=D("E"))).font=F(9.5)
    w.cell(r,2).alignment=L
    df={"C":"INDEX(%s,$N%d)"%(D("K"),r),"D":"C{r}/INDEX({j},$N{r})-1".format(r=r,j=D("J")),
        "E":"INDEX(%s,$N%d)"%(D("P"),r),"F":"E{r}/C{r}".format(r=r),
        "G":"INDEX(%s,$N%d)"%(D("Q"),r),"H":"INDEX(%s,$N%d)"%(D("R"),r),
        "I":"INDEX(%s,$N%d)"%(D("S"),r),"J":"INDEX(%s,$N%d)"%(D("I"),r),
        "K":"INDEX(%s,$N%d)"%(D("H"),r),"L":"J{r}/K{r}".format(r=r),
        "M":'IF(E{r}<0,"DÉFICITAIRE",IF(I{r}<0,"PIÈGE",IF(L{r}<0.75,"SOUS-REMPLI","")))'.format(r=r)}
    for col,f in df.items():
        j=openpyxl.utils.column_index_from_string(col)
        c=w.cell(r,j,'=IF($N%d="","",%s)'%(r,f))
        c.font=F(9.5,False,INK2,MONO if col!="M" else UI); c.number_format=FMT[j-3]
        c.alignment=R if col!="M" else Cn
    for j in range(2,14): w.cell(r,j).border=Border(bottom=sd(LINE2))

# --- mises en forme conditionnelles (statuts + signaux)
for rng in ("M12","M16:M28","M32:M46"):
    w.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=['"PIÈGE"'],
        fill=fill(ORANGESOFT),font=Font(name=UI,size=8.5,bold=True,color=ORANGE)))
    w.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=['"DÉFICITAIRE"'],
        fill=fill("FBE0E0"),font=Font(name=UI,size=8.5,bold=True,color=CRIT)))
    w.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=['"SOUS-REMPLI"'],
        fill=fill("FFF3D6"),font=Font(name=UI,size=8.5,bold=True,color="9A6B00")))
for rng in ("I12","I16:I28","I32:I46","E12","E16:E28","E32:E46"):
    w.conditional_formatting.add(rng,CellIsRule(operator="lessThan",formula=["0"],
        font=Font(name=MONO,size=9.5,bold=True,color=CRIT)))
for rng in ("D12","D16:D28","D32:D46"):
    w.conditional_formatting.add(rng,CellIsRule(operator="greaterThan",formula=["0"],font=Font(name=MONO,size=9.5,color=GOOD)))
    w.conditional_formatting.add(rng,CellIsRule(operator="lessThan",formula=["0"],font=Font(name=MONO,size=9.5,color=CRIT)))
for rng in ("L12","L16:L28","L32:L46"):
    w.conditional_formatting.add(rng,CellIsRule(operator="lessThan",formula=["0.75"],
        font=Font(name=MONO,size=9.5,bold=True,color=WARN)))

box(w,48,2,51,13,bg=PANEL2)
w.merge_cells(start_row=48,start_column=2,end_row=51,end_column=13)
w.cell(48,2,"  Comment lire.  La CONTRIBUTION (CA − coûts évitables : vacataires et achats directs) répond à « que perd-on si "
            "on ferme ? ».  La MARGE COMPLÈTE (après permanents, structure campus et siège) répond à « cette classe paie-t-elle "
            "sa part ? ».  Aux niveaux hauts elles disent la même chose. Au grain classe elles se contredisent : « PIÈGE » = "
            "contribution positive mais marge complète négative — la fermer coûterait plus cher que la garder, parce que les "
            "coûts qu'on lui a affectés, eux, ne disparaissent pas.  « DÉFICITAIRE » = son EBITDA lui-même est négatif : elle ne couvre pas ses coûts directs, et là seulement l’arbitrage de fermeture se pose vraiment.").font=F(9.5,c=INK2)
w.cell(48,2).alignment=TOP
for r in range(48,52): w.cell(r,2).border=Border(left=Side(style="thick",color=ORANGE))
w.cell(53,2,"Source : V_ALLOCATION au grain classe · remonte exactement au socle campus du cockpit "
            "(CA 23 098 985 · EBITDA 3 845 790). Coût variable 363 €/élève, coût direct réparti aux HEURES "
            "(une classe coûte une classe, quel que soit son effectif) — c'est ce mécanisme qui crée le piège.").font=F(8.5,c=INK3)
w.cell(53,2).alignment=L
w.sheet_view.zoomScale=90
wb.active=1
out="/home/user/demo5/eduservices/DRILL_EBITDA.xlsx"
wb.save(out); print("\nSAVED",out)
