#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Consolide l'ALLOCATION en UN seul onglet coherent '3_Allocation' :
   cles (saisie) + rollup marque + graphes + maille fine groupee (marque>campus>classe).
   Retire la section 3 redondante de Pilotage. Repointe _CALC_ALLOC sur les nouvelles cles."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from collections import OrderedDict
wb=openpyxl.load_workbook("CAD_SAAD_LIVE.xlsx")
dat=openpyxl.load_workbook("CAD_SAAD_LIVE.xlsx",data_only=True)
al0=dat["Allocation"]; H={c.value:c.column_letter for c in al0[1] if c.value}
def gv(col,r): return al0['%s%d'%(H[col],r)].value

GOLD="B8860B";GOLD_L="FBF3DE";NAVY="15406E";WHITE="FFFFFF";GREY_L="F7F7F9";PURP="7030A0"
INPUT="FFF2CC";LIVE="E4F5EA"
MORD=["MBWAY","ISCOM","IPAC","PIGIER","TUNON"]
MLAB={"MBWAY":"MBway","ISCOM":"ISCOM","IPAC":"Ipac","PIGIER":"Pigier","TUNON":"Tunon"}
MCH={"MBWAY":"2E86DE","ISCOM":"27AE60","IPAC":"F2A900","PIGIER":"E8743B","TUNON":"8E44AD"}
MCOL={"MBWAY":"D4E6FA","ISCOM":"DDF2E3","IPAC":"FBEFD0","PIGIER":"F6E2D4","TUNON":"ECE0F6"}
VILLE={"PAR":"Paris","LYO":"Lyon","NAN":"Nantes","BOR":"Bordeaux","LIL":"Lille","TLS":"Toulouse","REN":"Rennes","MTP":"Montpellier"}
thin=Side(style="thin",color="E0E0E0");box=Border(thin,thin,thin,thin)
gold_s=Side(style="medium",color="E0A800");ibord=Border(gold_s,gold_s,gold_s,gold_s)
def fill(h):return PatternFill("solid",fgColor=h)
RIG=Alignment(horizontal="right",vertical="center");LEF=Alignment(horizontal="left",vertical="center")
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)

# ============ 1) repointer _CALC_ALLOC sur les cles du nouvel onglet ============
TAB="'3_Allocation'!"
K={"K1":TAB+"$C$5","K4":TAB+"$C$6","K2":TAB+"$C$7","K3":TAB+"$C$8"}
ca=wb["_CALC_ALLOC"]
def g(inner): return '=IF($A{r}="","",'+inner+')'
def selk(kc,a,b,c): return 'IF(%s="Chiffre d\'affaires",%s,IF(%s="Effectif",%s,%s))'%(kc,a,kc,b,c)
for r in range(2,2001):
    ca["AC%d"%r]=g(selk(K["K3"],"$I{r}","$G{r}","$H{r}")).format(r=r)
    ca["AD%d"%r]=g(selk(K["K3"],"$N{r}","$L{r}","$M{r}")).format(r=r)
    ca["AE%d"%r]=g(selk(K["K2"],"$N{r}","$L{r}","$M{r}")).format(r=r)
    ca["AF%d"%r]=g(selk(K["K2"],"$S{r}","$Q{r}","$R{r}")).format(r=r)
    ca["AG%d"%r]=g(selk(K["K1"],"$S{r}","$Q{r}","$R{r}")).format(r=r)
    ca["AH%d"%r]=g(selk(K["K1"],"$V{r}","$T{r}","$U{r}")).format(r=r)
    ca["AP%d"%r]=g(selk(K["K4"],"$S{r}","$Q{r}","$R{r}")).format(r=r)
    ca["AQ%d"%r]=g(selk(K["K4"],"$V{r}","$T{r}","$U{r}")).format(r=r)

# ============ 2) onglet 3_Allocation ============
for nm in ("3_Maille","3_Allocation"):
    if nm in wb.sheetnames: del wb[nm]
ws=wb.create_sheet("3_Allocation")
ws.sheet_view.showGridLines=False
ws.sheet_properties.outlinePr.summaryBelow=False
ws.merge_cells("A1:L1")
ws["A1"]="3 · ALLOCATION & RENTABILITE  —  cout complet du groupe a la classe"
ws["A1"].font=Font(size=15,bold=True,color=GOLD); ws["A1"].fill=fill(GOLD_L)
ws["A2"]="« Qui est vraiment rentable, une fois le siege et la pub imputes ? »  ·  change une cle -> tout se redistribue (total groupe constant)"
ws["A2"].font=Font(size=9,italic=True,color="8A8FA0")

# --- cles (saisie) ---
ws["B4"]="Cles d'allocation (saisie)"; ws["B4"].font=Font(size=11,bold=True,color=GOLD)
keys=[("B5","C5","Siege administratif -> marque","Chiffre d'affaires"),
      ("B6","C6","Pub de marque -> marque","Effectif"),
      ("B7","C7","Marque -> campus","Effectif"),
      ("B8","C8","Campus -> classe","Nombre de classes")]
dv=DataValidation(type="list",formula1='"Chiffre d\'affaires,Effectif,Nombre de classes"',allow_blank=False)
ws.add_data_validation(dv)
for lb,vc,lab,val in keys:
    ws[lb]=lab; ws[lb].font=Font(size=9,bold=True,color=NAVY); ws[lb].alignment=LEF
    ws.merge_cells("%s:%s"%(vc,vc.replace("C","E")))
    ws[vc]=val; ws[vc].font=Font(size=10,bold=True,color="7F6000"); ws[vc].alignment=CEN; ws[vc].border=ibord; ws[vc].fill=fill(INPUT)
    for cc in ("D","E"): ws["%s%s"%(cc,vc[1:])].fill=fill(INPUT); ws["%s%s"%(cc,vc[1:])].border=ibord
    dv.add(vc)

# --- rollup marque (source graphes), colonnes N..W rows 4-9 ---
A="Allocation!"
def sif26(col,mq): return '=SUMIFS(%s$%s:$%s,%s$E:$E,"%s",%s$C:$C,"2026")'%(A,col,col,A,mq,A)
ws["N4"]="Rollup marque (live)"; ws["N4"].font=Font(size=8,bold=True,color="8A8FA0")
rh=["","Marque","CA","VAC","PERM","ODIR","STRUCT","Frais marque","Holding","Marge"]
for i,h in enumerate(rh):
    ws.cell(4,14+i,h).font=Font(size=7,bold=True,color="8A8FA0")
LC=["K","V","W","X","Y","AC","Z","AA"]
for i,m in enumerate(MORD):
    r=5+i
    ws.cell(r,14,m).font=Font(size=7,color="D9D9D9")
    ws.cell(r,15,MLAB[m]).font=Font(size=8,color="666666")
    for j,col in enumerate(LC):
        cc=ws.cell(r,16+j,sif26(col,m)); cc.number_format="# ##0"; cc.font=Font(size=8)

# --- graphes ---
def dpts(): return [DataPoint(idx=i,spPr=GraphicalProperties(solidFill=MCH[m])) for i,m in enumerate(MORD)]
mm=BarChart(); mm.type="col"; mm.title="Marge complete par marque (reagit aux cles)"; mm.height=6.5; mm.width=11
mm.add_data(Reference(ws,min_col=23,min_row=4,max_row=9),titles_from_data=True)  # W marge
mm.set_categories(Reference(ws,min_col=15,min_row=5,max_row=9)); mm.legend=None
mm.series[0].data_points=dpts(); mm.y_axis.numFmt="# ##0"; ws.add_chart(mm,"B11")
st=BarChart(); st.type="col"; st.grouping="stacked"; st.overlap=100; st.title="Decomposition du cout complet par marque"
st.height=6.5; st.width=11
for j in range(6): st.add_data(Reference(ws,min_col=18+j,min_row=4,max_row=9),titles_from_data=True)  # VAC..Holding
st.set_categories(Reference(ws,min_col=15,min_row=5,max_row=9)); st.y_axis.numFmt="# ##0"
ws.add_chart(st,"G11")

# --- maille fine groupee (a partir de la ligne 26) ---
R0=26
heads=["Marque ▸ Campus ▸ Classe","Effectif","CA","VAC","PERM","ODIR","STRUCT","Frais marque","Holding","Cout complet","Marge complete","Marge %"]
ws.cell(R0-1,1,"Maille fine — deplie campus/classe avec les [+]").font=Font(size=11,bold=True,color=GOLD)
for i,h in enumerate(heads):
    c=ws.cell(R0,1+i,h); c.font=Font(size=9,bold=True,color=WHITE); c.fill=fill(GOLD); c.alignment=CEN; c.border=box
ws.row_dimensions[R0].height=26
tree=OrderedDict((m,OrderedDict()) for m in MORD)
for r in range(2,al0.max_row+1):
    if str(gv('EXERCICE',r))!='2026': continue
    tree[gv('MARQUE',r)].setdefault(gv('ENTITY',r),[]).append((gv('PROGRAMME',r),gv('AN_ETUDE',r),gv('MODALITE',r)))
def sifm(col,crit):
    s="SUMIFS(%s$%s:$%s"%(A,col,col)
    for cc,vl in crit: s+=",%s$%s:$%s,%s"%(A,cc,cc,vl)
    return "="+s+',%s$C:$C,"2026")'%A
LMAP={"eff":"I","ca":"K","VAC":"V","PERM":"W","ODIR":"X","STRUCT":"Y","MARQUE":"AC","HOLD":"Z","MARGE":"AA"}
def wr(r,label,crit,level,fillc,fontc,bold):
    ws.cell(r,1,("   "*level)+label).font=Font(size=9,bold=bold,color=fontc)
    ws.cell(r,1).fill=fill(fillc); ws.cell(r,1).alignment=LEF; ws.cell(r,1).border=box
    for i,k in enumerate(["eff","ca","VAC","PERM","ODIR","STRUCT","MARQUE","HOLD"]):
        c=ws.cell(r,2+i,sifm(LMAP[k],crit)); c.number_format="# ##0"; c.font=Font(size=9,bold=bold,color=fontc)
        c.fill=fill(fillc); c.alignment=RIG; c.border=box
    ws.cell(r,10,"=SUM(D%d:I%d)"%(r,r)).number_format="# ##0"
    ws.cell(r,11,sifm(LMAP["MARGE"],crit)).number_format="# ##0"
    ws.cell(r,12,"=IFERROR(K%d/C%d,0)"%(r,r)).number_format="0.0%"
    for col in (10,11,12):
        c=ws.cell(r,col); c.font=Font(size=9,bold=bold,color=fontc if col!=11 else "1E7A46")
        c.fill=fill(fillc if col!=11 else LIVE); c.alignment=RIG; c.border=box
    if level>0: ws.row_dimensions[r].outline_level=level
    if level==2: ws.row_dimensions[r].hidden=True
r=R0+1
for m in MORD:
    wr(r,MLAB[m],[("E",'"%s"'%m)],0,MCOL[m],NAVY,True); r+=1
    for e,cls in tree[m].items():
        ville=VILLE.get(e.split("_")[-1],e.split("_")[-1])
        wr(r,ville+"  ("+e+")",[("D",'"%s"'%e)],1,GREY_L,"333333",True); r+=1
        for (prog,an,mod) in cls:
            wr(r,"%s %s %s"%(prog,an,mod),[("D",'"%s"'%e),("F",'"%s"'%prog),("G",'"%s"'%an),("H",'"%s"'%mod)],2,WHITE,"555555",False); r+=1
# total groupe
ws.cell(r,1,"GROUPE").font=Font(bold=True,color=WHITE); ws.cell(r,1).fill=fill(GOLD); ws.cell(r,1).border=box
for i,k in enumerate(["I","K","V","W","X","Y","AC","Z"]):
    ws.cell(r,2+i,'=SUMIFS(%s$%s:$%s,%s$C:$C,"2026")'%(A,k,k,A)).number_format="# ##0"
    ws.cell(r,2+i).font=Font(bold=True,color=WHITE); ws.cell(r,2+i).fill=fill(GOLD); ws.cell(r,2+i).border=box
ws.cell(r,10,"=SUM(D%d:I%d)"%(r,r)).number_format="# ##0"
ws.cell(r,11,'=SUMIFS(%sAA:AA,%sC:C,"2026")'%(A,A)).number_format="# ##0"
ws.cell(r,12,"=IFERROR(K%d/C%d,0)"%(r,r)).number_format="0.0%"
for col in (10,11,12):
    ws.cell(r,col).font=Font(bold=True,color=WHITE); ws.cell(r,col).fill=fill(GOLD); ws.cell(r,col).border=box
for c,w in {"A":34,"B":9,"C":11,"D":9,"E":9,"F":9,"G":10,"H":11,"I":10,"J":12,"K":13,"L":8}.items():
    ws.column_dimensions[c].width=w
for c in "NOPQRSTUVW": ws.column_dimensions[c].hidden=True  # rollup helper masque

# ============ 3) retirer la section 3 (allocation) de Pilotage ============
ps=wb["Pilotage"]
for mr in list(ps.merged_cells.ranges):
    if mr.min_row>=48 and mr.max_row<=72: ps.unmerge_cells(str(mr))
for row in ps.iter_rows(min_row=48,max_row=72,min_col=1,max_col=30):
    for c in row:
        if isinstance(c,openpyxl.cell.cell.MergedCell): continue
        c.value=None; c.fill=PatternFill(); c.font=Font(); c.border=Border(); c.alignment=Alignment()
# retirer les graphes allocation (anchor >= ligne 47), garder cap/donut/CA&EBITDA
def arow(ch):
    try: return ch.anchor._from.row
    except: return 0
ps._charts=[ch for ch in ps._charts if arow(ch)<47]
# retirer la DV des cles (si presente sur C52:C55)
ps.data_validations.dataValidation=[d for d in ps.data_validations.dataValidation if "C52" not in str(d.sqref)]

wb.calculation.fullCalcOnLoad=True
wb.save("CAD_SAAD_LIVE.xlsx")
print("OK onglet 3_Allocation consolide (cles+rollup+2 graphes+maille), Pilotage section 3 retiree.")
print("sheets:",wb.sheetnames)
