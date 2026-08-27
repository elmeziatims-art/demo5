#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Cockpit d'atterrissage 2026 — écran post-chargement. Pas de réconciliation.
6 KPI (finance + commercial) sur 3 ans + tension. Formules partout."""
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.chart import LineChart,Reference
INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E4F0EC"; CARD2="F5F7F9"; FAINT="7D8B98"
WHITE="FFFFFF"; BLUE="0000FF"; OCHRE="B3641C"; OCHREBG="F7EAD9"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
med=Side(style="medium",color=TEAL)
CTR=Alignment("center",vertical="center",wrap_text=True); LFT=Alignment("left",vertical="center"); RGT=Alignment("right",vertical="center")
EUR='#,##0;(#,##0);"-"'; PCT='0.0%'; NUM='#,##0'; EURc='#,##0" €"'; M2='#,##0.00,," M€"'

wb=openpyxl.Workbook()

# ---------- DONNÉES (agrégats annuels, source finance + CRM) ----------
dn=wb.active; dn.title="Données"; dn.sheet_view.showGridLines=False
dn["A1"]="DONNÉES — agrégats annuels groupe (source : compta produits + socle CRM)"; dn["A1"].font=F(11,True,TEALD)
dn["A2"]="2024-2025 réalisé · 2026 atterrissage. Bleu = donnée."; dn["A2"].font=F(8,False,FAINT,True)
head=["Exercice","CA","EBITDA","Leads","Inscrits","Dépense acq."]
for j,h in enumerate(head,1):
    c=dn.cell(4,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL); c.alignment=CTR if j>1 else LFT
DATA=[("2024",20064725,2648550,15305,1092,358819),
      ("2025",21268606,2977604,16226,1159,394702),
      ("2026",22544725,3291530,17197,1229,434174)]
for i,row in enumerate(DATA):
    r=5+i
    dn.cell(r,1,row[0]).font=F(9); dn.cell(r,1).alignment=CTR
    for j in range(1,6):
        cc=dn.cell(r,1+j,row[j]); cc.font=F(9,False,BLUE); cc.alignment=RGT
        cc.number_format=EUR if j in(1,2,5) else NUM
for col,w in zip("ABCDEF",[10,13,12,10,10,13]): dn.column_dimensions[col].width=w
# lignes Données : 2024=5, 2025=6, 2026=7 ; cols CA=B EBITDA=C Leads=D Inscrits=E Dep=F

# ---------- COCKPIT ----------
ck=wb.create_sheet("Cockpit"); ck.sheet_view.showGridLines=False
ck["A1"]="COCKPIT · ATTERRISSAGE 2026"; ck["A1"].font=F(16,True,INK)
ck["A2"]="Point de départ du budget 2027 · réalisé 2024-2025, atterrissage 2026 · commercial + financier, un seul écran"; ck["A2"].font=F(9,False,TEALD)
# bandeau de tête : 3 chiffres finance
def head_stat(col,lab,formula,fmt):
    ck.cell(4,col,lab).font=F(9,True,FAINT); ck.cell(4,col).alignment=LFT
    c=ck.cell(5,col,formula); c.font=F(16,True,TEALD); c.number_format=fmt; c.alignment=LFT
    for rr in (4,5):
        for cc in range(col,col+2): ck.cell(rr,cc).fill=fill(TEALBG)
head_stat(1,"Chiffre d'affaires 2026","=Données!B7",M2)
head_stat(3,"EBITDA 2026","=Données!C7",M2)
head_stat(5,"Marge EBITDA 2026","=Données!C7/Données!B7",PCT)

# grille KPI
hr=7
for j,h in enumerate(["KPI","2024 réalisé","2025 réalisé","2026 atterrissage","YoY 25→26"],1):
    c=ck.cell(hr,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL); c.alignment=CTR if j>1 else LFT
def section(r,txt):
    ck.cell(r,1,txt).font=F(9,True,TEALD)
    for c in range(1,6): ck.cell(r,c).fill=fill(CARD2)
def kpi(r,lab,f2024,f2025,f2026,fmt,yoy_pt=False,tension=False):
    # tension=True -> le KPI monte mais c'est DÉFAVORABLE (CAC). Sinon monter = bon.
    lab_col=OCHRE if tension else INK
    yoy_col=OCHRE if tension else TEALD      # vert = bon sens, ochre = mauvais sens
    ck.cell(r,1,lab).font=F(10,True,lab_col); ck.cell(r,1).alignment=LFT
    for k,fx in enumerate([f2024,f2025,f2026]):
        cc=ck.cell(r,2+k,fx); cc.number_format=fmt; cc.alignment=RGT; cc.font=F(10,False,lab_col)
    y=ck.cell(r,5)
    arrow="▲ " if not yoy_pt else ""
    if yoy_pt: y.value=f"=D{r}-C{r}"; y.number_format='"▲ "0.0" pt"'
    else: y.value=f"=D{r}/C{r}-1"; y.number_format='"▲ "0.0%'
    y.alignment=RGT; y.font=F(10,True,yoy_col)
    ck.cell(r,6,"bon sens" if not tension else "à surveiller").font=F(8,False,yoy_col,True)
section(8,"FINANCE")
kpi(9,"Chiffre d'affaires (€)","=Données!B5","=Données!B6","=Données!B7",EUR)
kpi(10,"EBITDA (€)","=Données!C5","=Données!C6","=Données!C7",EUR)
kpi(11,"Marge EBITDA %","=Données!C5/Données!B5","=Données!C6/Données!B6","=Données!C7/Données!B7",PCT,yoy_pt=True)
section(12,"COMMERCIAL")
kpi(13,"Leads","=Données!D5","=Données!D6","=Données!D7",NUM)
kpi(14,"Inscrits","=Données!E5","=Données!E6","=Données!E7",NUM)
kpi(15,"CAC (€/inscrit)","=Données!F5/Données!E5","=Données!F6/Données!E6","=Données!F7/Données!E7",EURc,tension=True)
ck.cell(16,1,"CAC = seul KPI en tension : le coût d'acquisition monte plus vite que le volume. C'est ce qu'on ira traiter dans le budget 2027.").font=F(8,False,OCHRE,True)

# tendance base 100
tr=18
ck.cell(tr,1,"TENSION — base 100 en 2024 : activité (CA) vs dépenses d'acquisition").font=F(10,True,TEALD)
for k,y in enumerate(("2024","2025","2026")): ck.cell(tr+1,2+k,int(y)).font=F(9,True); ck.cell(tr+1,2+k).alignment=CTR
ck.cell(tr+1,1,"Année").font=F(9,True)
ck.cell(tr+2,1,"Activité (CA)").font=F(9); ck.cell(tr+3,1,"Dépenses acq.").font=F(9)
# activité = CA index (col Données B) ; dépenses = dep index (col Données F)
for k,dcol in enumerate(["B5","B6","B7"]):
    ck.cell(tr+2,2+k,f"=Données!{dcol}/Données!B5*100").number_format='0.0'
for k,dcol in enumerate(["F5","F6","F7"]):
    ck.cell(tr+3,2+k,f"=Données!{dcol}/Données!F5*100").number_format='0.0'
chart=LineChart(); chart.title="Dépenses d'acquisition (+21%) décrochent au-dessus de l'activité (+12%)"; chart.height=8; chart.width=15
chart.add_data(Reference(ck,min_col=1,min_row=tr+2,max_row=tr+3,max_col=4),titles_from_data=True,from_rows=True)
chart.set_categories(Reference(ck,min_col=2,min_row=tr+1,max_col=4,max_row=tr+1))
ck.add_chart(chart,"H7")   # le SEUL graphe, à droite des tuiles, bien visible
ck.cell(tr+5,1,"LE graphe qui compte : l'écart entre les 2 courbes = la dégradation du CAC. C'est ce que le budget 2027 doit corriger.").font=F(8,True,OCHRE,True)
for col,w in zip("ABCDEF",[24,13,13,15,11,12]): ck.column_dimensions[col].width=w

wb.move_sheet("Cockpit",-1)
out="/home/user/demo5/eduservices/tagetik/COCKPIT_ATTERRISSAGE.xlsx"
wb.save(out); print("SAVED",out)
