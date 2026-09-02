#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""BRIDGE_CAC_1FEUILLE.xlsx — TOUT dans une seule feuille :
la matrice (dépenses acq + inscrits, 2025/2026), le bridge calculé juste à côté
(en formules pointant la matrice), les colonnes techniques du waterfall, et le
graphe. Réalisable tel quel sur Tagetik (matrice + cellules calculées + graphe)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E3EFEA"; WHITE="FFFFFF"; NAVY="1F3864"
SOFT="51606D"; FAINT="7D8B98"; RULE="C8D2DA"; OCHRE="B3641C"; OK="1E7A55"; RED="C0392B"; GREEN="1E7A55"; YEL="FFF6DA"
AR="Arial"
def F(sz=10,b=False,c=INK): return Font(name=AR,size=sz,bold=b,color=c)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color=RULE); box=Border(left=thin,right=thin,top=thin,bottom=thin)
Ln=Alignment("left",vertical="center"); Rn=Alignment("right",vertical="center"); Cn=Alignment("center",vertical="center",wrap_text=True)
EUR='#,##0" €"'; NUM='#,##0'; TOP=Alignment("left",vertical="top",wrap_text=True)

wb=openpyxl.Workbook(); ws=wb.active; ws.title="Bridge CAC"; ws.sheet_view.showGridLines=False
for col,w in zip("ABCDEFGHIJ",[20,12,12,4,11,11,11,11,11,3]): ws.column_dimensions[col].width=w

ws.merge_cells("A1:H1"); ws.cell(1,1,"  DIAGNOSTIC CAC — Bridge : d'où vient la hausse 2025 → 2026").font=F(13,True,WHITE); ws.cell(1,1).fill=fill(NAVY); ws.row_dimensions[1].height=26

# ---------- (1) LA MATRICE (chiffres tirés de la vue) ----------
ws.cell(3,1,"① Matrice (vue)").font=F(9.5,True,TEALD)
ws.cell(4,1,"Mesure").font=F(9,True,WHITE); ws.cell(4,2,"2025").font=F(9,True,WHITE); ws.cell(4,3,"2026").font=F(9,True,WHITE)
for c in (1,2,3): ws.cell(4,c).fill=fill(TEAL); ws.cell(4,c).alignment=Cn; ws.cell(4,c).border=box
ws.cell(5,1,"Dépenses acquisition (€)").font=F(9.5); ws.cell(5,2,394702).number_format=EUR; ws.cell(5,3,434174).number_format=EUR
ws.cell(6,1,"Inscrits (nouveaux)").font=F(9.5); ws.cell(6,2,1159).number_format=NUM; ws.cell(6,3,1229).number_format=NUM
for r in (5,6):
    for c in (1,2,3): ws.cell(r,c).border=box; ws.cell(r,c).alignment=Ln if c==1 else Rn
# cellules sources
DEP25,DEP26,INS25,INS26="$B$5","$C$5","$B$6","$C$6"

# ---------- (2) LE BRIDGE CALCULÉ (juste en dessous, formules sur la matrice) ----------
ws.cell(8,1,"② Bridge calculé (formules → la matrice ci-dessus)").font=F(9.5,True,TEALD)
ws.cell(9,1,"Étape").font=F(9,True,WHITE); ws.cell(9,2,"CAC (€)").font=F(9,True,WHITE); ws.cell(9,3,"Formule").font=F(8,True,WHITE)
ws.merge_cells("C9:C9")
for c in (1,2): ws.cell(9,c).fill=fill(TEAL); ws.cell(9,c).alignment=Cn; ws.cell(9,c).border=box
ws.cell(9,3).fill=fill(TEAL); ws.cell(9,3).alignment=Ln; ws.cell(9,3).border=box
calc=[("CAC 2025","=%s/%s"%(DEP25,INS25),"Dép.25 / Ins.25"),
      ("+ Effet dépenses","=(%s-%s)/%s"%(DEP26,DEP25,INS25),"(Dép.26−Dép.25)/Ins.25"),
      ("− Effet volume","=%s/%s-%s/%s"%(DEP26,INS26,DEP26,INS25),"Dép.26/Ins.26 − Dép.26/Ins.25"),
      ("CAC 2026","=%s/%s"%(DEP26,INS26),"Dép.26 / Ins.26")]
for i,(lab,f,note) in enumerate(calc):
    r=10+i
    ws.cell(r,1,lab).font=F(9.5,lab.startswith('CAC')); ws.cell(r,1).alignment=Ln; ws.cell(r,1).border=box
    x=ws.cell(r,2,f); x.number_format='#,##0.0" €"'; x.font=F(9.5,True,GREEN if '+' in lab else (RED if '−' in lab else INK)); x.alignment=Rn; x.border=box
    ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=3)
    ws.cell(r,3,note).font=F(8,c=FAINT); ws.cell(r,3).alignment=Ln; ws.cell(r,3).border=box
CAC25,EDEP,EVOL,CAC26="$B$10","$B$11","$B$12","$B$13"

# ---------- (3) COLONNES TECHNIQUES DU WATERFALL (formules, masquables) ----------
ws.cell(15,1,"③ Colonnes techniques du graphe  (masquables)").font=F(8.5,True,FAINT)
head=["Étape","Socle","Ancre","Hausse","Baisse"]
for j,h in enumerate(head,1):
    c=ws.cell(16,j,h); c.font=F(8,True,FAINT); c.alignment=Cn; c.border=box
wf=[("CAC 2025","0","=%s"%CAC25,"0","0"),
    ("+ Effet dépenses","=%s"%CAC25,"0","=%s"%EDEP,"0"),
    ("− Effet volume","=%s"%CAC26,"0","0","=%s+%s-%s"%(CAC25,EDEP,CAC26)),
    ("CAC 2026","0","=%s"%CAC26,"0","0")]
for i,(cat,so,an,ha,ba) in enumerate(wf):
    r=17+i
    for j,v in enumerate((cat,so,an,ha,ba),1):
        c=ws.cell(r,j,v); c.border=box; c.font=F(8,c=FAINT); c.alignment=Ln if j==1 else Rn
        if j>1: c.number_format='#,##0.0'
NR=len(wf)
cats=Reference(ws,min_col=1,max_col=1,min_row=17,max_row=17+NR-1)

# ---------- (4) LE GRAPHE (à droite) ----------
ch=BarChart(); ch.type="col"; ch.grouping="stacked"; ch.overlap=100
ch.title="Évolution du CAC 2025 → 2026 (€/inscrit)"; ch.height=8.5; ch.width=13.5
ch.y_axis.numFmt='#,##0 "€"'; ch.legend=None; ch.x_axis.delete=False; ch.y_axis.delete=False; ch.gapWidth=40
for colx in (2,3,4,5):
    ch.add_data(Reference(ws,min_col=colx,max_col=colx,min_row=17,max_row=17+NR-1),titles_from_data=False)
s_socle,s_ancre,s_hausse,s_baisse=ch.series
s_socle.graphicalProperties.noFill=True; s_socle.graphicalProperties.line.noFill=True
for s,color in ((s_ancre,TEALD),(s_hausse,GREEN),(s_baisse,RED)):
    s.graphicalProperties.solidFill=color; s.graphicalProperties.line.solidFill=color
for s,fmt in ((s_ancre,'#,##0" €"'),(s_hausse,'"+ "#,##0" €"'),(s_baisse,'"− "#,##0" €"')):
    s.dLbls=DataLabelList(); s.dLbls.showVal=True; s.dLbls.numFmt=fmt; s.dLbls.dLblPos="ctr"
    s.dLbls.showSerName=False; s.dLbls.showCatName=False; s.dLbls.showLegendKey=False
ch.set_categories(cats)
ws.add_chart(ch,"E3")

ws.merge_cells("A22:H24")
ws.cell(22,1,"Une seule feuille : ① la matrice (2 mesures × 2 exercices, tirée de la vue) · ② le bridge en formules juste "
             "en dessous · ③ les colonnes techniques du waterfall (masquables) qui alimentent le graphe. "
             "Change une dépense ou un nb d'inscrits dans ① → CAC, effets et graphe se recalent.").font=F(8.5,c=SOFT)
ws.cell(22,1).alignment=TOP

out="/home/user/demo5/eduservices/BRIDGE_CAC_1FEUILLE.xlsx"
wb.save(out); print("SAVED",out)
