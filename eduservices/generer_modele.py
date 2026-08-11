# -*- coding: utf-8 -*-
"""Construit le modèle Excel de budget/simulation EDUSERVICES GROUP."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/home/user/demo5/eduservices/EDUSERVICES_Budget_Simulation.xlsx"

# ----------------------------------------------------------------------------
# Palette & styles
# ----------------------------------------------------------------------------
NAVY   = "1F3864"   # titres / bandeaux
BLUE2  = "2E5496"   # sous-bandeaux
LIGHT  = "D9E1F2"   # bandes claires
LIGHT2 = "EDF1F9"
GREYH  = "F2F2F2"
YELLOW = "FFF2CC"   # hypothèses clés / cellules à remplir
GREEN_TOT = "E2EFDA"

FONT = "Arial"
CIN   = Font(name=FONT, color="0000FF")            # input codé en dur (bleu)
CINB  = Font(name=FONT, color="0000FF", bold=True)
CFORM = Font(name=FONT, color="000000")            # formule (noir)
CLINK = Font(name=FONT, color="008000")            # lien inter-feuilles (vert)
CHDR  = Font(name=FONT, color="FFFFFF", bold=True) # entête bandeau
CTIT  = Font(name=FONT, color="FFFFFF", bold=True, size=14)
CBOLD = Font(name=FONT, bold=True)
CIT   = Font(name=FONT, italic=True, color="595959", size=9)
CITB  = Font(name=FONT, italic=True, color="595959", size=9, bold=True)
CREG  = Font(name=FONT)

FILL_NAVY  = PatternFill("solid", fgColor=NAVY)
FILL_BLUE2 = PatternFill("solid", fgColor=BLUE2)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT)
FILL_LIGHT2= PatternFill("solid", fgColor=LIGHT2)
FILL_GREYH = PatternFill("solid", fgColor=GREYH)
FILL_YEL   = PatternFill("solid", fgColor=YELLOW)
FILL_TOT   = PatternFill("solid", fgColor=GREEN_TOT)

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
medb = Side(style="medium", color=NAVY)
BTOP = Border(top=medb)

AL = Alignment(horizontal="left", vertical="center", wrap_text=False)
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AR = Alignment(horizontal="right", vertical="center")
ALW = Alignment(horizontal="left", vertical="top", wrap_text=True)

EUR = '#,##0" €";(#,##0)" €";"-"'
EUR2= '#,##0.0" €";(#,##0.0)" €";"-"'
PCT = '0.0%;(0.0%);"-"'
NB  = '#,##0;(#,##0);"-"'
NB1 = '#,##0.0;(#,##0.0);"-"'

def C(ws, ref, val=None, font=None, fill=None, fmt=None, align=None, border=False):
    c = ws[ref]
    if val is not None: c.value = val
    if font: c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if align: c.alignment = align
    if border: c.border = BORDER
    return c

def band(ws, row, first, last, text, fill=FILL_NAVY, font=CHDR, h=20):
    ws.merge_cells(f"{first}{row}:{last}{row}")
    c = ws[f"{first}{row}"]
    c.value = text; c.font = font; c.fill = fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(openpyxl.utils.column_index_from_string(first),
                     openpyxl.utils.column_index_from_string(last)+1):
        ws.cell(row=row, column=col).fill = fill
    ws.row_dimensions[row].height = h

# ----------------------------------------------------------------------------
# Données campus (ILLUSTRATIVES — à remplacer par le réalisé)
# code, marque, ville, eff_n1, new_n1, tarif_n1, admin_etp, loyer_n1, da
# ----------------------------------------------------------------------------
CAMPUS = [
 ("MDS-NAN","MyDigitalSchool","Nantes",   620,260,7200, 9, 480000, 90000),
 ("MDS-PAR","MyDigitalSchool","Paris",    880,360,7900,12, 950000,140000),
 ("MDS-LYO","MyDigitalSchool","Lyon",     540,230,7400, 8, 520000, 80000),
 ("MDS-BOR","MyDigitalSchool","Bordeaux", 360,165,7200, 6, 360000, 55000),
 ("WSS-PAR","Win Sport School","Paris",   430,190,8200, 7, 620000, 70000),
 ("WSS-LYO","Win Sport School","Lyon",    300,140,7800, 5, 380000, 48000),
 ("IHE-NAN","IHECF","Nantes",             410,150,6900, 6, 300000, 40000),
 ("IHE-TLS","IHECF","Toulouse",           280,120,6700, 5, 260000, 35000),
 ("MOD-PAR","MODART International","Paris",350,150,8600, 7, 700000, 95000),
 ("MOD-LYO","MODART International","Lyon", 230,105,8200, 4, 360000, 50000),
 ("ESI-LIL","ESIS","Lille",               320,150,7000, 5, 300000, 42000),
 ("ESI-MTP","ESIS","Montpellier",         260,125,6800, 4, 250000, 33000),
]
# paramètres de référence pour reconstituer le RÉALISÉ N-1
COUT_ENS_N1, COUT_ADMIN_N1 = 52000, 42000
AUTRES_CH_N1, AUTRES_PROD, FRAIS = 600, 150, 90
CAC_N1, RATIO_N1 = 950, 18

def n1_actuals(eff,new,tarif,admin,loyer,da):
    ca  = eff*tarif + new*FRAIS + eff*AUTRES_PROD
    ms  = round(eff/RATIO_N1)*COUT_ENS_N1 + admin*COUT_ADMIN_N1
    mkt = new*CAC_N1
    aut = eff*AUTRES_CH_N1
    ebitda = ca - ms - loyer - mkt - aut
    return ca, ms, mkt, aut, ebitda

MARQUES = []
for row in CAMPUS:
    if row[1] not in MARQUES: MARQUES.append(row[1])

R0 = 4                    # 1ère ligne de données
RN = R0 + len(CAMPUS) - 1 # dernière ligne de données

# refs leviers (feuille Paramètres)
P = lambda a: f"'02_Parametres'!{a}"
P_PROG, P_CROI, P_TAR, P_RAT, P_SAL, P_INF, P_CAC = (
    P("$G$6"),P("$G$7"),P("$G$8"),P("$G$9"),P("$G$10"),P("$G$11"),P("$G$12"))
C_ENS,C_ADM,C_AUT,C_PRO,C_FRA = (
    P("$D$14"),P("$D$15"),P("$D$16"),P("$D$17"),P("$D$18"))

wb = openpyxl.Workbook()

# ============================================================================
# 00_Notice
# ============================================================================
ws = wb.active; ws.title = "00_Notice"; ws.sheet_view.showGridLines = False
for col,w in {"A":2,"B":34,"C":60,"D":22,"E":16}.items(): ws.column_dimensions[col].width=w
ws.merge_cells("B2:E2")
C(ws,"B2","EDUSERVICES GROUP — Modèle de Budget & Simulation",CTIT,FILL_NAVY,align=AL)
ws.row_dimensions[2].height=30
C(ws,"B3","Budget annuel N+1 · piloté par les inducteurs · pré-maquette avant implémentation CCH Tagetik",CIT)
ws.merge_cells("B3:E3")

band(ws,5,"B","E","1. Objet du classeur")
notice = [
 "Ce classeur prépare, sous Excel, l'intégralité d'un budget annuel N+1 pour EDUSERVICES GROUP,",
 "avant sa mise en place dans CCH Tagetik. Il est piloté par les inducteurs métier (effectifs",
 "étudiants, tarifs, ratio d'encadrement, CAC…) et intègre un moteur de simulation par scénario.",
 "",
 "Maille : Campus  →  Marque / École  →  Groupe. Le budget est reconstruit par campus puis",
 "consolidé automatiquement. Le réalisé N-1 sert de base de cadrage et de comparaison.",
]
r=6
for t in notice:
    ws.merge_cells(f"B{r}:E{r}"); C(ws,f"B{r}",t,CREG,align=ALW); r+=1

band(ws,13,"B","E","2. Comment utiliser le modèle")
steps = [
 ("Étape 1","Renseigner le réalisé (feuille 04_Historique) : effectifs, tarifs, loyers, D&A par campus."),
 ("Étape 2","Fixer les hypothèses de la note de cadrage (feuille 02_Parametres) pour chaque scénario."),
 ("Étape 3","Choisir le scénario actif dans la cellule jaune C3 de 02_Parametres (Cadrage / Optimiste / Prudent)."),
 ("Étape 4","Lire le budget consolidé (06_PnL) et le tableau de bord de simulation (07_Simulation)."),
 ("Étape 5","Basculer vers Tagetik en suivant la feuille 08_Mapping_Tagetik (dimensions & comptes)."),
]
r=14
for a,b in steps:
    C(ws,f"B{r}",a,CBOLD,FILL_LIGHT,align=AL,border=True)
    ws.merge_cells(f"C{r}:E{r}"); C(ws,f"C{r}",b,CREG,align=ALW,border=True)
    ws.row_dimensions[r].height=28; r+=1

band(ws,20,"B","E","3. Légende des couleurs (convention modélisation financière)")
leg=[("Saisie / donnée en dur",CIN,None),
     ("Formule / calcul",CFORM,None),
     ("Lien vers une autre feuille",CLINK,None),
     ("Hypothèse clé / cellule à remplir",CBOLD,FILL_YEL)]
r=21
for txt,ft,fl in leg:
    C(ws,f"B{r}","  Exemple  ",ft,fl,align=AC,border=True)
    ws.merge_cells(f"C{r}:E{r}"); C(ws,f"C{r}",txt,CREG,align=AL,border=True); r+=1

band(ws,26,"B","E","4. Avertissement")
ws.merge_cells("B27:E29")
C(ws,"B27","Les montants pré-remplis sont ILLUSTRATIFS (marques, campus, effectifs et euros) et destinés "
          "à faire tourner la démo. Ils doivent être remplacés par les données réelles d'EDUSERVICES "
          "GROUP avant toute exploitation. Sources à documenter dans 04_Historique.",CIT,align=ALW)
C(ws,"B31",f"Date de préparation : 2026-08-11", CIT)

# ============================================================================
# 01_Note_cadrage
# ============================================================================
ws = wb.create_sheet("01_Note_cadrage"); ws.sheet_view.showGridLines=False
for col,w in {"A":2,"B":40,"C":16,"D":58}.items(): ws.column_dimensions[col].width=w
ws.merge_cells("B2:D2"); C(ws,"B2","Note de cadrage budgétaire — Budget N+1",CTIT,FILL_NAVY,align=AL)
ws.row_dimensions[2].height=28
ws.merge_cells("B3:D3")
C(ws,"B3","Hypothèses directrices retenues (scénario CADRAGE). Modifiables en 02_Parametres.",CIT)

band(ws,5,"B","D","1. Objectifs du groupe")
obj=["Poursuivre la croissance des effectifs par le développement du recrutement et l'amélioration",
     "du taux de réinscription, tout en préservant la marge EBITDA.",
     "Maîtriser le coût d'acquisition étudiant (CAC) et la masse salariale via le ratio d'encadrement."]
r=6
for t in obj:
    ws.merge_cells(f"B{r}:D{r}"); C(ws,f"B{r}",t,CREG,align=ALW); r+=1

band(ws,10,"B","D","2. Hypothèses clés (valeur de cadrage)")
C(ws,"B11","Inducteur",CHDR,FILL_BLUE2,align=AL,border=True)
C(ws,"C11","Valeur",CHDR,FILL_BLUE2,align=AC,border=True)
C(ws,"D11","Justification / commentaire",CHDR,FILL_BLUE2,align=AL,border=True)
lignes=[
 ("Taux de réinscription (progression nette)","='02_Parametres'!D6",PCT,
   "Part des étudiants poursuivant leur cursus (net des diplômés et abandons)."),
 ("Croissance des nouvelles inscriptions","='02_Parametres'!D7",PCT,
   "Objectif de recrutement vs N-1, porté par le marketing et l'ouverture de programmes."),
 ("Hausse tarifaire moyenne","='02_Parametres'!D8",PCT,
   "Revalorisation des frais de scolarité, alignée sur le positionnement et l'inflation."),
 ("Ratio d'encadrement (étudiants / ETP enseignant)","='02_Parametres'!D9",NB1,
   "Pilote la masse salariale enseignante. Plus élevé = plus efficient."),
 ("Politique salariale (augmentation)","='02_Parametres'!D10",PCT,
   "Revalorisation moyenne des rémunérations chargées."),
 ("Inflation des charges","='02_Parametres'!D11",PCT,
   "Appliquée aux loyers et autres charges d'exploitation."),
 ("Coût d'acquisition marketing (CAC)","='02_Parametres'!D12",EUR,
   "Dépense marketing moyenne par nouvel inscrit."),
]
r=12
for lib,f,fmt,comm in lignes:
    C(ws,f"B{r}",lib,CREG,align=AL,border=True)
    C(ws,f"C{r}",f,CLINK,fmt=fmt,align=AC,border=True)
    ws.column_dimensions["D"].width=58
    C(ws,f"D{r}",comm,CIT,align=ALW,border=True)
    ws.row_dimensions[r].height=30; r+=1

band(ws,r+1,"B","D","3. Règles de gestion du modèle")
r+=2
regles=[
 "Effectif budgété = Effectif N-1 × taux de réinscription + Nouvelles inscriptions N-1 × (1 + croissance).",
 "Chiffre d'affaires = Effectif × tarif revalorisé + frais de dossier + autres produits par étudiant.",
 "Masse salariale = (Effectif / ratio d'encadrement) × coût enseignant chargé + ETP admin × coût admin,",
 "     coûts revalorisés de la politique salariale.",
 "Marketing = Nouvelles inscriptions × CAC.  Loyers & autres charges revalorisés de l'inflation.",
 "EBITDA = CA − Masse salariale − Loyers − Marketing − Autres charges.  EBIT = EBITDA − D&A.",
]
for t in regles:
    ws.merge_cells(f"B{r}:D{r}"); C(ws,f"B{r}",t,CREG,align=ALW); r+=1

# ============================================================================
# 02_Parametres
# ============================================================================
ws = wb.create_sheet("02_Parametres"); ws.sheet_view.showGridLines=False
for col,w in {"A":2,"B":42,"C":14,"D":13,"E":13,"F":13,"G":15}.items(): ws.column_dimensions[col].width=w
ws.merge_cells("B2:G2"); C(ws,"B2","Paramètres & Scénarios de simulation",CTIT,FILL_NAVY,align=AL)
ws.row_dimensions[2].height=28

C(ws,"B3","Scénario actif :",CBOLD,align=AR)
C(ws,"C3","Cadrage",CINB,FILL_YEL,align=AC,border=True)
dv=DataValidation(type="list",formula1='"Cadrage,Optimiste,Prudent"',allow_blank=False)
ws.add_data_validation(dv); dv.add(ws["C3"])
C(ws,"E3","◄ Choisir ici : le modèle entier bascule automatiquement.",CIT)
ws.merge_cells("E3:G3")

# entête tableau leviers
C(ws,"B5","Levier de simulation",CHDR,FILL_BLUE2,align=AL,border=True)
C(ws,"C5","Unité",CHDR,FILL_BLUE2,align=AC,border=True)
for col,name in (("D","Cadrage"),("E","Optimiste"),("F","Prudent")):
    C(ws,f"{col}5",name,CHDR,FILL_BLUE2,align=AC,border=True)
C(ws,"G5","ACTIF",CHDR,FILL_NAVY,align=AC,border=True)

# levier: (lib, unite, cadrage, opt, prudent, fmt)
levs=[
 ("Taux de réinscription (progression nette)","%",0.62,0.66,0.58,PCT),
 ("Croissance des nouvelles inscriptions","%",0.06,0.12,0.00,PCT),
 ("Hausse tarifaire moyenne","%",0.03,0.04,0.02,PCT),
 ("Ratio d'encadrement (étud./ETP ens.)","nb",18,19,17,NB1),
 ("Politique salariale (augmentation)","%",0.025,0.02,0.03,PCT),
 ("Inflation des charges","%",0.02,0.015,0.03,PCT),
 ("Coût d'acquisition marketing (CAC)","€",900,780,1100,EUR),
]
r=6
for lib,u,cad,opt,pru,fmt in levs:
    C(ws,f"B{r}",lib,CREG,align=AL,border=True)
    C(ws,f"C{r}",u,CREG,align=AC,border=True)
    C(ws,f"D{r}",cad,CIN,fmt=fmt,align=AC,border=True)
    C(ws,f"E{r}",opt,CIN,fmt=fmt,align=AC,border=True)
    C(ws,f"F{r}",pru,CIN,fmt=fmt,align=AC,border=True)
    C(ws,f"G{r}",f"=INDEX(D{r}:F{r},MATCH($C$3,$D$5:$F$5,0))",CFORM,FILL_LIGHT,fmt=fmt,align=AC,border=True)
    r+=1

# constantes de référence
C(ws,"B13","Constantes de référence (communes aux scénarios)",CHDR,FILL_BLUE2,align=AL,border=True)
for col in ("C","D","E","F","G"): C(ws,f"{col}13"," ",fill=FILL_BLUE2,border=True)
consts=[
 ("Coût chargé enseignant (€/ETP)",52000,EUR),
 ("Coût chargé administratif (€/ETP)",42000,EUR),
 ("Autres charges par étudiant (€)",600,EUR),
 ("Autres produits par étudiant (€)",150,EUR),
 ("Frais de dossier par nouvel inscrit (€)",90,EUR),
]
r=14
for lib,val,fmt in consts:
    C(ws,f"B{r}",lib,CREG,align=AL,border=True)
    C(ws,f"C{r}"," ",border=True)
    C(ws,f"D{r}",val,CIN,fmt=fmt,align=AC,border=True)
    for col in ("E","F"): C(ws,f"{col}{r}"," ",border=True)
    C(ws,f"G{r}",f"=D{r}",CFORM,FILL_LIGHT,fmt=fmt,align=AC,border=True)
    r+=1
C(ws,f"B{r+1}","Bleu = saisie modifiable · Gris clair = valeur active utilisée par le moteur.",CIT)
ws.merge_cells(f"B{r+1}:G{r+1}")

# ============================================================================
# 03_Referentiel
# ============================================================================
ws = wb.create_sheet("03_Referentiel"); ws.sheet_view.showGridLines=False
for col,w in {"A":2,"B":14,"C":24,"D":16,"E":10,"F":40}.items(): ws.column_dimensions[col].width=w
ws.merge_cells("B2:F2"); C(ws,"B2","Référentiel — Dimensions (mapping Tagetik)",CTIT,FILL_NAVY,align=AL)
ws.row_dimensions[2].height=28
band(ws,4,"B","F","Dimension ENTITÉ : Groupe → Marque → Campus")
hdr=["Code entité","Marque / École","Campus (ville)","Niveau","Devise"]
for i,h in enumerate(hdr):
    C(ws,f"{get_column_letter(2+i)}5",h,CHDR,FILL_BLUE2,align=AC,border=True)
r=6
C(ws,f"B{r}","GRP-EDU",CINB,FILL_TOT,align=AL,border=True)
C(ws,f"C{r}","EDUSERVICES GROUP",CINB,FILL_TOT,align=AL,border=True)
C(ws,f"D{r}","—",CREG,FILL_TOT,align=AC,border=True)
C(ws,f"E{r}","Groupe",CBOLD,FILL_TOT,align=AC,border=True)
C(ws,f"F{r}","EUR",CREG,FILL_TOT,align=AC,border=True); r+=1
for code,marque,ville,*_ in CAMPUS:
    C(ws,f"B{r}",code,CIN,align=AL,border=True)
    C(ws,f"C{r}",marque,CIN,align=AL,border=True)
    C(ws,f"D{r}",ville,CIN,align=AL,border=True)
    C(ws,f"E{r}","Campus",CREG,align=AC,border=True)
    C(ws,f"F{r}","EUR",CREG,align=AC,border=True); r+=1

band(ws,r+1,"B","F","Dimension COMPTE (Account) — plan simplifié")
r+=2
for i,h in enumerate(["Compte","Libellé","Type","Signe","Rubrique P&L"]):
    C(ws,f"{get_column_letter(2+i)}{r}",h,CHDR,FILL_BLUE2,align=AC,border=True)
r+=1
comptes=[
 ("70600","Scolarité","Produit","+","Chiffre d'affaires"),
 ("70800","Frais de dossier","Produit","+","Chiffre d'affaires"),
 ("70900","Autres produits","Produit","+","Chiffre d'affaires"),
 ("64000","Masse salariale","Charge","-","Charges de personnel"),
 ("61300","Loyers & charges locatives","Charge","-","Charges externes"),
 ("62300","Marketing & acquisition","Charge","-","Charges externes"),
 ("60000","Autres charges d'exploitation","Charge","-","Charges externes"),
 ("68000","Dotations amortissements (D&A)","Charge","-","D&A"),
]
for cpt,lib,typ,sgn,rub in comptes:
    C(ws,f"B{r}",cpt,CIN,align=AL,border=True)
    C(ws,f"C{r}",lib,CREG,align=AL,border=True)
    C(ws,f"D{r}",typ,CREG,align=AC,border=True)
    C(ws,f"E{r}",sgn,CREG,align=AC,border=True)
    C(ws,f"F{r}",rub,CREG,align=AL,border=True); r+=1

# ============================================================================
# 04_Historique  (RÉALISÉ N-1)
# ============================================================================
ws = wb.create_sheet("04_Historique"); ws.sheet_view.showGridLines=False
ws.sheet_view.showGridLines=False
cols_h=["Code","Marque / École","Campus","Effectif N-1","Nouv. inscrits N-1","Tarif moyen N-1",
        "ETP admin","Loyers N-1","D&A","CA N-1","Masse sal. N-1","Marketing N-1","Autres ch. N-1","EBITDA N-1"]
widths=[10,20,14,12,13,13,10,12,11,13,13,12,12,13]
for i,w in enumerate(widths): ws.column_dimensions[get_column_letter(1+i)].width=w
ws.merge_cells("A1:N1"); C(ws,"A1","Réalisé N-1 par campus (base de cadrage)",CTIT,FILL_NAVY,align=AL)
ws.row_dimensions[1].height=26
ws.merge_cells("A2:N2")
C(ws,"A2","Cellules bleues = données réelles à saisir. Valeurs actuelles ILLUSTRATIVES.",CIT)
for i,h in enumerate(cols_h):
    C(ws,f"{get_column_letter(1+i)}3",h,CHDR,FILL_BLUE2,align=AC,border=True)
ws.row_dimensions[3].height=34
r=R0
for code,marque,ville,eff,new,tarif,admin,loyer,da in CAMPUS:
    ca,ms,mkt,aut,ebitda = n1_actuals(eff,new,tarif,admin,loyer,da)
    vals=[(code,None,AL),(marque,None,AL),(ville,None,AL),
          (eff,NB,AC),(new,NB,AC),(tarif,EUR,AC),(admin,NB,AC),(loyer,EUR,AR),(da,EUR,AR),
          (ca,EUR,AR),(ms,EUR,AR),(mkt,EUR,AR),(aut,EUR,AR),(ebitda,EUR,AR)]
    for i,(v,fmt,al) in enumerate(vals):
        C(ws,f"{get_column_letter(1+i)}{r}",v,CIN,fmt=fmt,align=al,border=True)
    r+=1
# totaux
C(ws,f"A{r}","TOTAL GROUPE",CBOLD,FILL_TOT,align=AL,border=True)
C(ws,f"B{r}"," ",fill=FILL_TOT,border=True); C(ws,f"C{r}"," ",fill=FILL_TOT,border=True)
for col in ["D","E","H","I","J","K","L","M","N"]:
    C(ws,f"{col}{r}",f"=SUM({col}{R0}:{col}{RN})",CBOLD,FILL_TOT,
      fmt=(NB if col in("D","E") else EUR),align=(AC if col in("D","E") else AR),border=True)
C(ws,f"F{r}"," ",fill=FILL_TOT,border=True); C(ws,f"G{r}",f"=SUM(G{R0}:G{RN})",CBOLD,FILL_TOT,fmt=NB,align=AC,border=True)
HIST_TOT=r

# ============================================================================
# 05_Moteur  (BUDGET N+1 par campus)
# ============================================================================
ws = wb.create_sheet("05_Moteur"); ws.sheet_view.showGridLines=False
cols_m=["Code","Marque / École","Campus","Réinscrits","Nouv. inscrits","Effectif budget",
        "Tarif moyen","Scolarité","Frais dossier","Autres produits","CA Budget",
        "ETP enseignants","ETP admin","MS enseignants","MS admin","Masse salariale",
        "Loyers","Marketing","Autres charges","EBITDA","D&A","EBIT",
        "Marge EBITDA %","CA / étudiant"]
mw=[10,18,13]+[12]*(len(cols_m)-3)
for i,w in enumerate(mw): ws.column_dimensions[get_column_letter(1+i)].width=w
ws.merge_cells("A1:X1"); C(ws,"A1","Moteur de budget N+1 par campus (calcul par inducteurs)",CTIT,FILL_NAVY,align=AL)
ws.row_dimensions[1].height=26
ws.merge_cells("A2:X2")
C(ws,"A2","100% formules. Réagit aux leviers de 02_Parametres et au réalisé de 04_Historique.",CIT)
for i,h in enumerate(cols_m):
    C(ws,f"{get_column_letter(1+i)}3",h,CHDR,FILL_BLUE2,align=AC,border=True)
ws.row_dimensions[3].height=34

def hcol(letter,row): return f"'04_Historique'!{letter}{row}"

r=R0
for idx,(code,marque,ville,*_ ) in enumerate(CAMPUS):
    hr=R0+idx
    put=lambda col,f,fmt=EUR,al=AR,ft=CFORM: C(ws,f"{col}{r}",f,ft,fmt=fmt,align=al,border=True)
    C(ws,f"A{r}",f"={hcol('A',hr)}",CLINK,align=AL,border=True)
    C(ws,f"B{r}",f"={hcol('B',hr)}",CLINK,align=AL,border=True)
    C(ws,f"C{r}",f"={hcol('C',hr)}",CLINK,align=AL,border=True)
    # effectifs
    put("D",f"={hcol('D',hr)}*{P_PROG}",NB,AC)
    put("E",f"={hcol('E',hr)}*(1+{P_CROI})",NB,AC)
    put("F",f"=D{r}+E{r}",NB,AC)
    # revenus
    put("G",f"={hcol('F',hr)}*(1+{P_TAR})",EUR,AC)
    put("H",f"=F{r}*G{r}")
    put("I",f"=E{r}*{C_FRA}")
    put("J",f"=F{r}*{C_PRO}")
    put("K",f"=H{r}+I{r}+J{r}")
    # masse salariale
    put("L",f"=F{r}/{P_RAT}",NB1,AC)
    put("M",f"={hcol('G',hr)}",NB1,AC)
    put("N",f"=L{r}*{C_ENS}*(1+{P_SAL})")
    put("O",f"=M{r}*{C_ADM}*(1+{P_SAL})")
    put("P",f"=N{r}+O{r}")
    # charges
    put("Q",f"={hcol('H',hr)}*(1+{P_INF})")
    put("R",f"=E{r}*{P_CAC}")
    put("S",f"=F{r}*{C_AUT}*(1+{P_INF})")
    # résultat
    put("T",f"=K{r}-P{r}-Q{r}-R{r}-S{r}")
    put("U",f"={hcol('I',hr)}")
    put("V",f"=T{r}-U{r}")
    put("W",f"=IFERROR(T{r}/K{r},0)",PCT,AC)
    put("X",f"=IFERROR(K{r}/F{r},0)",EUR,AC)
    r+=1
MOT_TOT=r
C(ws,f"A{r}","TOTAL GROUPE",CBOLD,FILL_TOT,align=AL,border=True)
C(ws,f"B{r}"," ",fill=FILL_TOT,border=True); C(ws,f"C{r}"," ",fill=FILL_TOT,border=True)
sum_cols={"D":NB,"E":NB,"F":NB,"H":EUR,"I":EUR,"J":EUR,"K":EUR,"L":NB1,"M":NB1,
          "N":EUR,"O":EUR,"P":EUR,"Q":EUR,"R":EUR,"S":EUR,"T":EUR,"U":EUR,"V":EUR}
for col,fmt in sum_cols.items():
    al = AC if fmt in (NB,NB1) else AR
    C(ws,f"{col}{r}",f"=SUM({col}{R0}:{col}{RN})",CBOLD,FILL_TOT,fmt=fmt,align=al,border=True)
C(ws,f"G{r}"," ",fill=FILL_TOT,border=True)
C(ws,f"W{r}",f"=IFERROR(T{r}/K{r},0)",CBOLD,FILL_TOT,fmt=PCT,align=AC,border=True)
C(ws,f"X{r}",f"=IFERROR(K{r}/F{r},0)",CBOLD,FILL_TOT,fmt=EUR,align=AC,border=True)

# ============================================================================
# 06_PnL  (consolidé)
# ============================================================================
ws = wb.create_sheet("06_PnL"); ws.sheet_view.showGridLines=False
for col,w in {"A":2,"B":34,"C":16,"D":16,"E":15,"F":13}.items(): ws.column_dimensions[col].width=w
ws.merge_cells("B2:F2"); C(ws,"B2","Compte de résultat — Groupe (Budget N+1 vs Réalisé N-1)",CTIT,FILL_NAVY,align=AL)
ws.row_dimensions[2].height=26
C(ws,"B3","Scénario actif :",CBOLD,align=AR)
C(ws,"C3","='02_Parametres'!C3",CITB,align=AC)

hp=["Rubrique","Réalisé N-1","Budget N+1","Écart €","Écart %"]
for i,h in enumerate(hp):
    C(ws,f"{get_column_letter(2+i)}5",h,CHDR,FILL_BLUE2,align=AC,border=True)

MT=f"'05_Moteur'!"; HT=f"'04_Historique'!"
# (libellé, N-1 formule, Budget formule, fmt, gras?)
mn=f"{MOT_TOT}"; hn=f"{HIST_TOT}"
lines=[
 ("Effectifs (nb)",       f"={HT}D{hn}", f"={MT}F{mn}", NB, False),
 ("Chiffre d'affaires",   f"={HT}J{hn}", f"={MT}K{mn}", EUR, True),
 ("  Masse salariale",    f"=-{HT}K{hn}",f"=-{MT}P{mn}",EUR, False),
 ("  Loyers",             f"=-{HT}H{hn}",f"=-{MT}Q{mn}",EUR, False),
 ("  Marketing",          f"=-{HT}L{hn}",f"=-{MT}R{mn}",EUR, False),
 ("  Autres charges",     f"=-{HT}M{hn}",f"=-{MT}S{mn}",EUR, False),
 ("EBITDA",               f"={HT}N{hn}", f"={MT}T{mn}", EUR, True),
 ("  Marge EBITDA %",     f"=IFERROR({HT}N{hn}/{HT}J{hn},0)",f"=IFERROR({MT}T{mn}/{MT}K{mn},0)",PCT,False),
 ("  D&A",                f"=-{HT}I{hn}",f"=-{MT}U{mn}",EUR, False),
 ("EBIT",                 f"={HT}N{hn}-{HT}I{hn}", f"={MT}V{mn}", EUR, True),
]
r=6
for lib,fn1,fbud,fmt,bold in lines:
    ft=CBOLD if bold else CREG
    fl=FILL_LIGHT if bold else None
    C(ws,f"B{r}",lib,ft,fl,align=AL,border=True)
    C(ws,f"C{r}",fn1,(CBOLD if bold else CLINK),fl,fmt=fmt,align=AR,border=True)
    C(ws,f"D{r}",fbud,(CBOLD if bold else CLINK),fl,fmt=fmt,align=AR,border=True)
    if fmt==PCT:
        C(ws,f"E{r}",f"=D{r}-C{r}",ft,fl,fmt=PCT,align=AR,border=True)
        C(ws,f"F{r}"," ",fill=fl,border=True)
    else:
        C(ws,f"E{r}",f"=D{r}-C{r}",ft,fl,fmt=fmt,align=AR,border=True)
        C(ws,f"F{r}",f"=IFERROR(D{r}/C{r}-1,0)",ft,fl,fmt=PCT,align=AR,border=True)
    r+=1

# synthèse par marque
band(ws,r+1,"B","F","Synthèse par marque / école")
r+=2
for i,h in enumerate(["Marque / École","CA Budget","EBITDA Budget","Marge EBITDA %","Δ CA vs N-1"]):
    C(ws,f"{get_column_letter(2+i)}{r}",h,CHDR,FILL_BLUE2,align=AC,border=True)
r+=1
mr0=R0; mrn=RN
for marque in MARQUES:
    crit=f'"{marque}"'
    C(ws,f"B{r}",marque,CREG,align=AL,border=True)
    C(ws,f"C{r}",f'=SUMIF({MT}$B${mr0}:$B${mrn},{crit},{MT}$K${mr0}:$K${mrn})',CFORM,fmt=EUR,align=AR,border=True)
    C(ws,f"D{r}",f'=SUMIF({MT}$B${mr0}:$B${mrn},{crit},{MT}$T${mr0}:$T${mrn})',CFORM,fmt=EUR,align=AR,border=True)
    C(ws,f"E{r}",f"=IFERROR(D{r}/C{r},0)",CFORM,fmt=PCT,align=AR,border=True)
    C(ws,f"F{r}",f'=IFERROR(C{r}/SUMIF({HT}$B${mr0}:$B${mrn},{crit},{HT}$J${mr0}:$J${mrn})-1,0)',
      CFORM,fmt=PCT,align=AR,border=True)
    r+=1
C(ws,f"B{r}","TOTAL",CBOLD,FILL_TOT,align=AL,border=True)
C(ws,f"C{r}",f"=SUM(C{r-len(MARQUES)}:C{r-1})",CBOLD,FILL_TOT,fmt=EUR,align=AR,border=True)
C(ws,f"D{r}",f"=SUM(D{r-len(MARQUES)}:D{r-1})",CBOLD,FILL_TOT,fmt=EUR,align=AR,border=True)
C(ws,f"E{r}",f"=IFERROR(D{r}/C{r},0)",CBOLD,FILL_TOT,fmt=PCT,align=AR,border=True)
C(ws,f"F{r}"," ",fill=FILL_TOT,border=True)

# ============================================================================
# 07_Simulation  (tableau de bord)
# ============================================================================
ws = wb.create_sheet("07_Simulation"); ws.sheet_view.showGridLines=False
for col,w in {"A":2,"B":30,"C":18,"D":18,"E":16,"F":16}.items(): ws.column_dimensions[col].width=w
ws.merge_cells("B2:F2"); C(ws,"B2","Tableau de bord — Simulation budgétaire",CTIT,FILL_NAVY,align=AL)
ws.row_dimensions[2].height=26
ws.merge_cells("B3:F3")
C(ws,"B3","Change le scénario en 02_Parametres!C3 et tous les indicateurs ci-dessous se recalculent.",CIT)
C(ws,"B4","Scénario actif :",CBOLD,align=AR)
C(ws,"C4","='02_Parametres'!C3",CITB,FILL_YEL,align=AC,border=True)

# KPI cards
band(ws,6,"B","F","Indicateurs clés — Budget N+1")
kpis=[
 ("Effectif total (nb)",  f"={MT}F{mn}", NB,  f"=IFERROR({MT}F{mn}/{HT}D{hn}-1,0)"),
 ("Chiffre d'affaires",   f"={MT}K{mn}", EUR, f"=IFERROR({MT}K{mn}/{HT}J{hn}-1,0)"),
 ("EBITDA",               f"={MT}T{mn}", EUR, f"=IFERROR({MT}T{mn}/{HT}N{hn}-1,0)"),
 ("Marge EBITDA %",       f"=IFERROR({MT}T{mn}/{MT}K{mn},0)", PCT, None),
 ("EBIT",                 f"={MT}V{mn}", EUR, None),
 ("CA / étudiant",        f"=IFERROR({MT}K{mn}/{MT}F{mn},0)", EUR, None),
]
C(ws,"B7","Indicateur",CHDR,FILL_BLUE2,align=AL,border=True)
C(ws,"C7","Budget N+1",CHDR,FILL_BLUE2,align=AC,border=True)
C(ws,"D7","Réalisé N-1",CHDR,FILL_BLUE2,align=AC,border=True)
C(ws,"E7","Évolution",CHDR,FILL_BLUE2,align=AC,border=True)
n1map={"Effectif total (nb)":f"={HT}D{hn}","Chiffre d'affaires":f"={HT}J{hn}",
       "EBITDA":f"={HT}N{hn}","Marge EBITDA %":f"=IFERROR({HT}N{hn}/{HT}J{hn},0)",
       "EBIT":f"={HT}N{hn}-{HT}I{hn}","CA / étudiant":f"=IFERROR({HT}J{hn}/{HT}D{hn},0)"}
r=8
for lib,fbud,fmt,eva in kpis:
    C(ws,f"B{r}",lib,CBOLD,align=AL,border=True)
    C(ws,f"C{r}",fbud,CFORM,fmt=fmt,align=AC,border=True)
    C(ws,f"D{r}",n1map[lib],CLINK,fmt=fmt,align=AC,border=True)
    if eva: C(ws,f"E{r}",eva,CFORM,fmt=PCT,align=AC,border=True)
    else:   C(ws,f"E{r}",f"=IFERROR(C{r}/D{r}-1,0)",CFORM,fmt=PCT,align=AC,border=True)
    r+=1

# comparatif leviers des 3 scénarios
band(ws,r+1,"B","F","Comparatif des scénarios (rappel des hypothèses)")
r+=2
C(ws,f"B{r}","Levier",CHDR,FILL_BLUE2,align=AL,border=True)
for col,name in (("C","Cadrage"),("D","Optimiste"),("E","Prudent")):
    C(ws,f"{col}{r}",name,CHDR,FILL_BLUE2,align=AC,border=True)
C(ws,f"F{r}","Actif",CHDR,FILL_NAVY,align=AC,border=True)
r+=1
lev_rows=[("Réinscription",6,PCT),("Croissance recrut.",7,PCT),("Hausse tarifaire",8,PCT),
          ("Ratio encadrement",9,NB1),("Politique salariale",10,PCT),("Inflation charges",11,PCT),
          ("CAC (€)",12,EUR)]
for lib,pr,fmt in lev_rows:
    C(ws,f"B{r}",lib,CREG,align=AL,border=True)
    C(ws,f"C{r}",f"='02_Parametres'!D{pr}",CLINK,fmt=fmt,align=AC,border=True)
    C(ws,f"D{r}",f"='02_Parametres'!E{pr}",CLINK,fmt=fmt,align=AC,border=True)
    C(ws,f"E{r}",f"='02_Parametres'!F{pr}",CLINK,fmt=fmt,align=AC,border=True)
    C(ws,f"F{r}",f"='02_Parametres'!G{pr}",CBOLD,FILL_LIGHT,fmt=fmt,align=AC,border=True)
    r+=1

# ============================================================================
# 08_Mapping_Tagetik
# ============================================================================
ws = wb.create_sheet("08_Mapping_Tagetik"); ws.sheet_view.showGridLines=False
for col,w in {"A":2,"B":28,"C":30,"D":50}.items(): ws.column_dimensions[col].width=w
ws.merge_cells("B2:D2"); C(ws,"B2","Passerelle vers CCH Tagetik",CTIT,FILL_NAVY,align=AL)
ws.row_dimensions[2].height=26
ws.merge_cells("B3:D3")
C(ws,"B3","Correspondance entre le modèle Excel et les objets Tagetik pour l'implémentation.",CIT)
band(ws,5,"B","D","Dimensions")
for i,h in enumerate(["Concept du modèle","Dimension Tagetik","Détail / valeurs"]):
    C(ws,f"{get_column_letter(2+i)}6",h,CHDR,FILL_BLUE2,align=AC,border=True)
maps=[
 ("Campus","Entity","GRP-EDU → marques → campus (voir 03_Referentiel). Hiérarchie de consolidation."),
 ("Compte P&L","Account","Plan de comptes 03_Referentiel (70600, 64000…) rattaché à l'arbre Account."),
 ("Budget N+1 / Réalisé N-1","Category","Category = ACTUAL (N-1) et BUDGET (N+1). Scénarios = versions/Category dédiées."),
 ("Année budgétaire","Period / Year","Budget annuel : Period = FY. Détail mensuel possible via allocation Tagetik."),
 ("Scénario (Cadrage/Opt/Prudent)","Category ou Custom dim.","Un jeu de valeurs par scénario ; en Excel via le sélecteur 02_Parametres!C3."),
 ("Effectifs, ETP, ratio, CAC","Comptes techniques / drivers","Stockés comme comptes non-financiers (statistiques) pour le calcul piloté par inducteurs."),
]
r=7
for a,b,c in maps:
    C(ws,f"B{r}",a,CBOLD,align=ALW,border=True)
    C(ws,f"C{r}",b,CREG,align=ALW,border=True)
    C(ws,f"D{r}",c,CREG,align=ALW,border=True)
    ws.row_dimensions[r].height=32; r+=1
band(ws,r+1,"B","D","Étapes d'implémentation recommandées")
r+=2
impl=[
 "1. Créer/valider les dimensions Entity (campus→marque→groupe) et Account.",
 "2. Charger le réalisé N-1 (Category ACTUAL) depuis 04_Historique.",
 "3. Traduire les inducteurs (02_Parametres) en formules de calcul / règles Tagetik.",
 "4. Rejouer le budget par campus (logique de 05_Moteur) et vérifier vs 06_PnL.",
 "5. Décliner les scénarios en versions et automatiser la simulation.",
]
for t in impl:
    ws.merge_cells(f"B{r}:D{r}"); C(ws,f"B{r}",t,CREG,align=ALW); r+=1

# freeze panes sur feuilles de données
for name,cell in (("04_Historique","D4"),("05_Moteur","D4")):
    wb[name].freeze_panes = cell

# forcer le recalcul complet à l'ouverture (Excel/LibreOffice calculent les valeurs)
try:
    wb.calculation.fullCalcOnLoad = True
except Exception:
    from openpyxl.workbook.properties import CalcProperties
    wb.calculation = CalcProperties(fullCalcOnLoad=True)

wb.save(OUT)
print("Écrit:", OUT)
print("Lignes données:", R0, "->", RN, "| Total Historique:", HIST_TOT, "| Total Moteur:", MOT_TOT)
