#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""LE_MOTEUR_MODELE.xlsx — modèle VIVANT, aligné sur le moteur.
Grain = CAMPUS (le grain réel de l'élasticité), GROUPÉ PAR MARQUE (déplie/replie).
La marque est un sous-total, pas une élasticité inventée. Formules cliquables ;
piloté par les cases jaunes (Δ, CA/inscrit, coût variable).
Onglet 1 : calibration + effet (campus groupé marque) + back-test.
Onglet 2 : budget de marque (organique)."""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

D=json.load(open('/tmp/moteur_full.json')); CAMP=D['camp']
ORG=json.load(open('/tmp/moteur_org.json'))
INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E2EFEB"
FAINT="7D8B98"; WHITE="FFFFFF"; OCHRE="B3641C"; OCHRED="8A4A12"
GREEN="1E7A55"; GREENBG="E4F0E8"; RULE="C8D2DA"; SOFT="51606D"; NAVY="3D4F8F"; MQBG="EEF2F6"
INPUTBG="FFF6DA"; INPUTBD="D9B94A"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
LFT=Alignment("left",vertical="center"); RGT=Alignment("right",vertical="center")
CTR=Alignment("center",vertical="center",wrap_text=True)
def LI(ind): return Alignment("left",vertical="center",indent=ind)
thin=Side(style="thin",color=RULE); med=Side(style="medium",color=RULE); inbd=Side(style="medium",color=INPUTBD)
EUR='#,##0 "€";-#,##0 "€";"-"'; NUM='#,##0'; PCT='0.0%'; DEC3='0.000'; SPCT='+0.0%;-0.0%;0.0%'
NAME={'IPAC_MTP':'Montpellier','IPAC_NAN':'Nantes','IPAC_REN':'Rennes',
 'ISCOM_LIL':'Lille','ISCOM_PAR':'Paris','ISCOM_TLS':'Toulouse',
 'MBWAY_BOR':'Bordeaux','MBWAY_LYO':'Lyon','MBWAY_NAN':'Nantes','MBWAY_PAR':'Paris',
 'PIGIER_BOR':'Bordeaux','PIGIER_LYO':'Lyon','TUNON_LYO':'Lyon','TUNON_PAR':'Paris'}
MQ={'MBWAY':'MBway','ISCOM':'Iscom','IPAC':'Ipac','PIGIER':'Pigier','TUNON':'Tunon'}
ORDER=['MBWAY','ISCOM','IPAC','PIGIER','TUNON']
NAME_ORG={d['campus']:NAME[d['campus']] for d in ORG}

wb=openpyxl.Workbook(); ws=wb.active; ws.title="Le moteur (modèle)"
ws.sheet_view.showGridLines=False
ws.sheet_properties.outlinePr.summaryBelow=False
def put(r,c,v,sz=10,b=False,col=INK,it=False,al=LFT,fmt=None,bg=None,box=None):
    cell=ws.cell(r,c,v); cell.font=F(sz,b,col,it); cell.alignment=al
    if fmt: cell.number_format=fmt
    if bg: cell.fill=fill(bg)
    if box: cell.border=box
    return cell
def head(r,cols):
    for j,h in enumerate(cols,1):
        c=ws.cell(r,j,h); c.font=F(8.5,True,WHITE); c.fill=fill(TEAL); c.alignment=LFT if j==1 else CTR; c.border=Border(bottom=med)
def IN(r,c,v,fmt):
    cell=ws.cell(r,c,v); cell.number_format=fmt; cell.alignment=RGT; cell.font=F(10,True,INK)
    cell.fill=fill(INPUTBG); cell.border=Border(top=inbd,bottom=inbd,left=inbd,right=inbd); return cell

put(1,1,"EDUSERVICES · budget 2027 · le moteur, à nu",sz=9,b=True,col=TEALD)
put(2,1,"Le moteur — par campus, groupé par marque (déplie avec les +)",sz=15,b=True)
put(3,1,"L'élasticité vit au campus (le moteur en a 14, pas 5). La marque = sous-total. Jaune = saisie ; le reste = formule cliquable.",sz=9,col=SOFT)

# --- PARAMÈTRES ---
put(5,1,"PARAMÈTRES",sz=11,b=True,col=OCHRED)
put(6,1,"Δ budget d'acquisition",sz=10); IN(6,2,0.08,PCT); put(6,3,"le geste testé — passez 8→12 %",sz=8,col=FAINT,it=True)
put(7,1,"CA 1ʳᵉ année / inscrit",sz=10); IN(7,2,7608,EUR)
put(8,1,"Coût variable / élève",sz=10); IN(8,2,300,EUR)
DE="$B$6"; RV="$B$7"; CV="$B$8"

# --- TABLE : calibration + effet, campus groupé marque ---
put(10,1,"①  Calibration & effet d'un +Δ% — campus groupé par marque",sz=12,b=True,col=TEALD)
put(11,1,"Élasticité = régression 3 ans =SLOPE(LN(leads);LN(budget)) — la même que la vue V_CAMPAGNES. La marque n'a pas d'élasticité propre (sous-total à blanc).",sz=8,col=FAINT,it=True)
hr=13
head(hr,["Marque ▸ Campus","L.pay24","L.pay25","L.pay26","Bud.24","Bud.25","Bud.26","Élasticité","Conv.",
         "Δ budget","Inscrits gagnés","CA gagné","EBITDA gagné","CAC marg."])
r=hr+1
groups={m:[d for d in CAMP if d['campus'].split('_')[0]==m] for m in ORDER}
gtot_rows=[]
for m in ORDER:
    ds=groups[m]
    sub=r                      # ligne sous-total marque (au-dessus du détail)
    first=r+1; last=r+len(ds)
    # sous-total marque (formules SUM sur le détail en dessous)
    put(sub,1,MQ[m],sz=10,b=True,col=TEALD,al=LI(0))
    put(sub,8,"—",sz=9,col=FAINT,al=RGT)                                     # élasticité : n/a au niveau marque
    ws.cell(sub,9,f"=K{sub}/J{sub}") if False else None
    put(sub,9,f"=SUM(K{first}:K{last})/SUM(J{first}:J{last})" if False else "",)
    # colonnes effet = SUM ; conv marque = inscrits/leads agrégés (via somme)
    for cc in (10,11,12,13):
        L=chr(64+cc); c=ws.cell(sub,cc,f"=SUM({L}{first}:{L}{last})"); c.number_format=(EUR if cc in(10,12,13) else NUM); c.font=F(10,True,TEALD); c.alignment=RGT
    c=ws.cell(sub,14,f"=J{sub}/K{sub}"); c.number_format=EUR; c.font=F(10,True,NAVY); c.alignment=RGT  # CAC marg agrégé
    for j in range(1,15): ws.cell(sub,j).fill=fill(MQBG); ws.cell(sub,j).border=Border(top=med,bottom=thin)
    gtot_rows.append((first,last))
    r=first
    for d in ds:
        put(r,1,NAME[d['campus']],sz=9,al=LI(1))
        put(r,2,d['p24'],sz=9,al=RGT,fmt=NUM,bg=INPUTBG); put(r,3,d['p25'],sz=9,al=RGT,fmt=NUM,bg=INPUTBG); put(r,4,d['p26'],sz=9,al=RGT,fmt=NUM,bg=INPUTBG)
        put(r,5,d['s24'],sz=9,al=RGT,fmt=NUM,bg=INPUTBG); put(r,6,d['s25'],sz=9,al=RGT,fmt=NUM,bg=INPUTBG); put(r,7,d['s26'],sz=9,al=RGT,fmt=NUM,bg=INPUTBG)
        put(r,8,f"=SLOPE(LN(B{r}:D{r}),LN(E{r}:G{r}))",sz=9,b=True,col=TEALD,al=RGT,fmt=DEC3)   # H élast 3 ans
        lead26=d['p26']+d['o26']
        put(r,9,f"={d['n26']}/{lead26}",sz=9,col=SOFT,al=RGT,fmt=PCT)                            # I conv = new26/leads26
        put(r,10,f"=G{r}*{DE}",sz=9,col=OCHRED,al=RGT,fmt=EUR)                                   # J Δbudget
        put(r,11,f"=D{r}*((1+{DE})^H{r}-1)*I{r}",sz=9,al=RGT,fmt=NUM)                            # K inscrits gagnés
        put(r,12,f"=K{r}*{RV}",sz=9,col=TEALD,al=RGT,fmt=EUR)                                    # L CA gagné
        put(r,13,f"=L{r}-J{r}-K{r}*{CV}",sz=9,b=True,col=GREEN,al=RGT,fmt=EUR)                   # M EBITDA
        put(r,14,f"=J{r}/K{r}",sz=9,col=NAVY,al=RGT,fmt=EUR)                                     # N CAC marg
        for j in range(1,15): ws.cell(r,j).border=Border(bottom=thin)
        ws.row_dimensions[r].outline_level=1; ws.row_dimensions[r].hidden=True   # replié par défaut
        r+=1
# GROUPE
put(r,1,"GROUPE",sz=11,b=True,col=TEAL,al=LI(0))
subs=[c for c in range(hr+1, r) if ws.row_dimensions[c].outline_level!=1]  # lignes sous-total marque
# somme sur toutes les lignes campus
allc=[cr for (f,l) in gtot_rows for cr in range(f,l+1)]
for cc in (10,11,12,13):
    L=chr(64+cc); rng="+".join(f"{L}{cr}" for cr in allc)
    c=ws.cell(r,cc,f"={rng}"); c.number_format=(EUR if cc in(10,12,13) else NUM); c.font=F(11,True,TEAL); c.alignment=RGT
c=ws.cell(r,14,f"=J{r}/K{r}"); c.number_format=EUR; c.font=F(11,True,NAVY); c.alignment=RGT
for j in range(1,15): ws.cell(r,j).fill=fill(GREENBG); ws.cell(r,j).border=Border(top=med,bottom=med)
r+=2
put(r,1,"Déplie une marque (le +) : chaque campus a SA vraie élasticité (=SLOPE, comme le moteur). Le CAC marginal marque = Σbudget ÷ Σinscrits (agrégat réel).",sz=8,col=OCHRE,it=True)
r+=3

# --- BACK-TEST (campus, inchangé, en formules) ---
put(r,1,"②  La preuve — back-test (calibré ≤2025, prédit 2026, jusqu'aux inscrits)",sz=12,b=True,col=TEALD); r+=1
put(r,1,"Payants : élasticité 24→25 × budget26. Organiques : leur tendance 24→25. × conversion 2025 → inscrits. Rien de 2026 en entrée.",sz=8,col=SOFT,it=True); r+=2
hr=r
head(hr,["Campus","L.pay24","L.pay25","Bud.24","Bud.25","Bud.26","Élast 24→25","Préd.pay26","Org.24","Org.25","Préd.org26","Leads26 préd","Leads25","Inscrits25","Conv.25","Inscrits26 préd","Inscrits26 réel","Écart"])
r=hr+1; e2=r
for d in CAMP:
    put(r,1,MQ[d['campus'].split('_')[0]]+" "+NAME[d['campus']],sz=8)
    put(r,2,d['p24'],sz=8,al=RGT,fmt=NUM,bg=INPUTBG); put(r,3,d['p25'],sz=8,al=RGT,fmt=NUM,bg=INPUTBG)
    put(r,4,d['s24'],sz=8,al=RGT,fmt=NUM,bg=INPUTBG); put(r,5,d['s25'],sz=8,al=RGT,fmt=NUM,bg=INPUTBG); put(r,6,d['s26'],sz=8,al=RGT,fmt=NUM,bg=INPUTBG)
    put(r,7,f"=LN(C{r}/B{r})/LN(E{r}/D{r})",sz=8,col=SOFT,al=RGT,fmt=DEC3)
    put(r,8,f"=C{r}*(F{r}/E{r})^G{r}",sz=8,col=NAVY,al=RGT,fmt=NUM)
    put(r,9,d['o24'],sz=8,al=RGT,fmt=NUM,bg=INPUTBG); put(r,10,d['o25'],sz=8,al=RGT,fmt=NUM,bg=INPUTBG)
    put(r,11,f"=J{r}*(J{r}/I{r})",sz=8,col=NAVY,al=RGT,fmt=NUM)
    put(r,12,f"=H{r}+K{r}",sz=8,b=True,col=NAVY,al=RGT,fmt=NUM)
    put(r,13,d['l25'],sz=8,al=RGT,fmt=NUM,bg=INPUTBG); put(r,14,d['n25'],sz=8,al=RGT,fmt=NUM,bg=INPUTBG)
    put(r,15,f"=N{r}/M{r}",sz=8,col=SOFT,al=RGT,fmt=PCT)
    put(r,16,f"=L{r}*O{r}",sz=8,b=True,col=NAVY,al=RGT,fmt=NUM)
    put(r,17,d['n26'],sz=8,al=RGT,fmt=NUM)
    put(r,18,f"=P{r}/Q{r}-1",sz=8,b=True,col=GREEN,al=RGT,fmt=SPCT)
    for j in range(1,19): ws.cell(r,j).border=Border(bottom=thin)
    r+=1
l2=r-1
put(r,1,"GROUPE",sz=9,b=True,col=TEALD)
for cc in (12,16,17):
    L=chr(64+cc); ws.cell(r,cc,f"=SUM({L}{e2}:{L}{l2})").number_format=NUM; ws.cell(r,cc).font=F(9,True,TEALD); ws.cell(r,cc).alignment=RGT
ws.cell(r,18,f"=P{r}/Q{r}-1").number_format=SPCT; ws.cell(r,18).font=F(9,True,GREEN); ws.cell(r,18).alignment=RGT
for j in range(1,19): ws.cell(r,j).fill=fill(GREENBG); ws.cell(r,j).border=Border(top=med,bottom=med)

ws.column_dimensions['A'].width=24
for col in "BCDEFGHIJKLMNOPQR": ws.column_dimensions[col].width=9.5
ws.freeze_panes="B4"

# ================= ONGLET 2 — ORGANIQUE (campus) =================
w2=wb.create_sheet("Budget de marque (organique)"); w2.sheet_view.showGridLines=False
def p2(r,c,v,sz=10,b=False,col=INK,it=False,al=LFT,fmt=None,bg=None):
    cell=w2.cell(r,c,v); cell.font=F(sz,b,col,it); cell.alignment=al
    if fmt: cell.number_format=fmt
    if bg: cell.fill=fill(bg)
    return cell
def h2(r,cols):
    for j,h in enumerate(cols,1):
        c=w2.cell(r,j,h); c.font=F(8.5,True,WHITE); c.fill=fill(TEAL); c.alignment=LFT if j==1 else CTR; c.border=Border(bottom=med)
p2(1,1,"EDUSERVICES · le moteur, à nu",sz=9,b=True,col=TEALD)
p2(2,1,"Le budget de marque → les leads organiques",sz=15,b=True)
p2(3,1,"Autre levier : le budget de marque nourrit les organiques, avec sa propre élasticité (plus basse). Régression 3 ans, comme l'acquisition.",sz=9,col=SOFT)
p2(5,1,"Δ budget de marque",sz=10)
c=w2.cell(5,2,0.10); c.number_format=PCT; c.alignment=RGT; c.font=F(10,True,INK); c.fill=fill(INPUTBG); c.border=Border(top=inbd,bottom=inbd,left=inbd,right=inbd)
DM="$B$5"
hr=7
h2(hr,["Campus","Bud.mq24","Bud.mq25","Bud.mq26","Org.24","Org.25","Org.26","Élasticité marque","Org. gagnés (+Δ)"])
r=hr+1; ea=r
for d in ORG:
    p2(r,1,MQ[d['campus'].split('_')[0]]+" "+NAME_ORG[d['campus']],sz=9)
    p2(r,2,d['b24'],sz=9,al=RGT,fmt=EUR,bg=INPUTBG); p2(r,3,d['b25'],sz=9,al=RGT,fmt=EUR,bg=INPUTBG); p2(r,4,d['b26'],sz=9,al=RGT,fmt=EUR,bg=INPUTBG)
    p2(r,5,d['o24'],sz=9,al=RGT,fmt=NUM,bg=INPUTBG); p2(r,6,d['o25'],sz=9,al=RGT,fmt=NUM,bg=INPUTBG); p2(r,7,d['o26'],sz=9,al=RGT,fmt=NUM,bg=INPUTBG)
    p2(r,8,f"=SLOPE(LN(E{r}:G{r}),LN(B{r}:D{r}))",sz=9,b=True,col=TEALD,al=RGT,fmt=DEC3)
    p2(r,9,f"=G{r}*((1+{DM})^H{r}-1)",sz=9,al=RGT,fmt=NUM)
    for j in range(1,10): w2.cell(r,j).border=Border(bottom=thin)
    r+=1
la=r-1
p2(r,1,"GROUPE",sz=9,b=True,col=TEALD)
w2.cell(r,9,f"=SUM(I{ea}:I{la})").number_format=NUM; w2.cell(r,9).font=F(9,True,TEALD); w2.cell(r,9).alignment=RGT
for j in range(1,10): w2.cell(r,j).fill=fill(GREENBG); w2.cell(r,j).border=Border(top=med,bottom=med)
p2(r+2,1,"Les deux leviers (acquisition + marque) alimentent le MÊME funnel → candidatures → inscrits → CA.",sz=9,b=True,col=OCHRE)
w2.column_dimensions['A'].width=22
for col in "BCDEFGHI": w2.column_dimensions[col].width=12
w2.freeze_panes="B4"

wb.calculation.fullCalcOnLoad=True
out="/home/user/demo5/eduservices/tagetik/LE_MOTEUR_MODELE.xlsx"
wb.save(out); print("SAVED",out)
