#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Carte stratégique d'acquisition (bubble) : CAC marginal vs croissance leads,
taille = budget acq. Proposition de graphe à ajouter à l'arbitrage des caps."""
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.chart import BubbleChart,Reference,Series
INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E4F0EC"; CARD2="F5F7F9"; FAINT="7D8B98"
WHITE="FFFFFF"; BLUE="0000FF"; OCHRE="B3641C"; GREEN="1E7A55"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
CTR=Alignment("center",vertical="center",wrap_text=True); LFT=Alignment("left",vertical="center"); RGT=Alignment("right",vertical="center")
EUR='#,##0;(#,##0);"-"'; PCT='0.0%'; NUM='#,##0'
def Hd(ws,r,labels):
    for j,h in enumerate(labels,1):
        c=ws.cell(r,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL); c.alignment=CTR if j>1 else LFT

# campus, CAC marginal, croiss leads, budget acq
CAP=[["IPAC MTP",1034,0.145,15178],["IPAC NAN",1216,0.138,21760],["IPAC REN",1041,0.145,14933],
["ISCOM LIL",979,0.082,25220],["ISCOM PAR",1540,0.078,60800],["ISCOM TLS",919,0.082,22083],
["MBWAY BOR",890,0.126,24923],["MBWAY LYO",1102,0.119,41783],["MBWAY NAN",964,0.124,31533],
["MBWAY PAR",1475,0.111,68291],["PIGIER BOR",1368,0.088,21968],["PIGIER LYO",1705,0.085,36612],
["TUNON LYO",1625,0.073,18850],["TUNON PAR",2193,0.070,30240]]

wb=openpyxl.Workbook(); ws=wb.active; ws.title="Carte acquisition"; ws.sheet_view.showGridLines=False
ws["A1"]="CARTE STRATÉGIQUE D'ACQUISITION  ·  où mettre le budget"; ws["A1"].font=F(15,True,INK)
ws["A2"]="Chaque bulle = un campus. X = CAC marginal (← moins cher) · Y = croissance des leads (↑) · taille = budget acq."; ws["A2"].font=F(9,False,TEALD)
ws["A3"]="Haut-gauche = INVESTIR (CAC bas, forte croissance) · Bas-droite = RÉDUIRE (CAC élevé, faible croissance)."; ws["A3"].font=F(9,True,OCHRE,True)
Hd(ws,5,["Campus","CAC marginal (€)","Croissance leads","Budget acq (€)"])
r=6
for row in CAP:
    ws.cell(r,1,row[0]).font=F(9); ws.cell(r,1).alignment=LFT
    ws.cell(r,2,row[1]).number_format=EUR; ws.cell(r,2).font=F(9,False,BLUE)
    ws.cell(r,3,row[2]).number_format=PCT; ws.cell(r,3).font=F(9,False,BLUE)
    ws.cell(r,4,row[3]).number_format=EUR; ws.cell(r,4).font=F(9,False,BLUE)
    for j in (2,3,4): ws.cell(r,j).alignment=RGT
    r+=1
last=r-1
# bubble chart
ch=BubbleChart(); ch.style=18; ch.title="Carte d'acquisition — CAC marginal (X) vs croissance leads (Y), taille = budget"
ch.height=11; ch.width=20
xv=Reference(ws,min_col=2,min_row=6,max_row=last)
yv=Reference(ws,min_col=3,min_row=6,max_row=last)
sz=Reference(ws,min_col=4,min_row=6,max_row=last)
ch.series.append(Series(values=yv,xvalues=xv,zvalues=sz,title="Campus"))
ch.x_axis.title="CAC marginal (€) — moins cher à gauche"
ch.y_axis.title="Croissance des leads"
ch.x_axis.delete=False; ch.y_axis.delete=False
ws.add_chart(ch,"F5")
ws.cell(last+2,1,"Lecture : MBway Bordeaux (890 € / +12,6 %) = à investir ; Tunon Paris (2 193 € / +7 %) = à réduire. La grosse bulle ISCOM Paris (60,8 k€) est chère (1 540 €) — à surveiller.").font=F(8,True,OCHRE,True)
ws.cell(last+3,1,"NB : openpyxl n'étiquette pas les bulles ; sur Tagetik/Excel tu ajoutes les libellés campus par point.").font=F(8,False,FAINT,True)
for col,w in zip("ABCD",[14,16,15,14]): ws.column_dimensions[col].width=w

out="/home/user/demo5/eduservices/tagetik/CARTE_ACQUISITION.xlsx"
wb.save(out); print("SAVED",out)
