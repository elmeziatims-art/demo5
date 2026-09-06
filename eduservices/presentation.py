#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Refonte presentation Pilotage (demo CFO) :
 flux logique = KPI GROUPE -> Cap -> Synthese -> Allocation (cles PLACEES a cote
 de leur effet, la marge complete qui se redistribue en direct). Graphes soignes.
 Les cles d'allocation sont RELOCALISEES (E50/E51/E52) et les refs _CALC_ALLOC MAJ.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
def dpts(marques):
    return [DataPoint(idx=i,spPr=GraphicalProperties(solidFill=MCOL[c[0]])) for i,c in enumerate(marques)]
F="CAD_SAAD_LIVE.xlsx"; wb=openpyxl.load_workbook(F)
ps=wb["Pilotage"]; ca=wb["_CALC_ALLOC"]

NAVY="1F3864";BLUE="2E75B6";BLUE_L="DDEBF7";GREEN="548235";GREEN_L="E2EFDA"
AMBER="BF8F00";AMBER_L="FFF2CC";INPUT="FFF2CC";GREY_D="3B3B3B";GREY_L="F2F2F2"
WHITE="FFFFFF";LIVE="C6E0B4";RED="C00000";FCE="FCE4D6"
MCOL={"MBWAY":"2E75B6","ISCOM":"548235","IPAC":"BF8F00","PIGIER":"843C0C","TUNON":"7030A0"}
MARQUES=[("MBWAY","MBway"),("ISCOM","ISCOM"),("IPAC","Ipac"),("PIGIER","Pigier"),("TUNON","Tunon")]
thin=Side(style="thin",color="BFBFBF");box=Border(thin,thin,thin,thin)
def fill(h):return PatternFill("solid",fgColor=h)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)
RIG=Alignment(horizontal="right",vertical="center")
def S(ws,ref,val=None,font=None,fillc=None,align=None,border=None,fmt=None):
    c=ws[ref]
    if val is not None:c.value=val
    if font:c.font=font
    if fillc:c.fill=fill(fillc)
    if align:c.alignment=align
    if border:c.border=border
    if fmt:c.number_format=fmt
    return c
H1=Font(size=10,bold=True,color=WHITE);LAB=Font(size=9,color=GREY_D)
LABB=Font(size=10,bold=True,color=NAVY);VAL=Font(size=9,color="000000")
INPF=Font(size=11,bold=True,color="7F6000");LIVEF=Font(size=9,bold=True,color="375623")
SEC=Font(size=12,bold=True,color=NAVY)

# ============================================================ nettoyage haut
ps._charts=[]
for mr in list(ps.merged_cells.ranges): ps.unmerge_cells(str(mr))
# effacer ancienne zone cles (D6:E10) + ancien titre (D2:O3)
for row in ps.iter_rows(min_row=1,max_row=11,min_col=2,max_col=21):
    for c in row:
        c.value=None; c.fill=PatternFill(); c.font=Font(); c.border=Border(); c.alignment=Alignment()
# re-inscrire la liste de validation A1:A3 (elle sert de source dropdown)
for i,v in enumerate(["Chiffre d'affaires","Effectif","Nombre de classes"]):
    ps["A%d"%(i+1)]=v; ps["A%d"%(i+1)].font=Font(size=8,color="D9D9D9")
ps.sheet_view.showGridLines=False

# ============================================================ TITRE
ps.merge_cells("B1:R2")
S(ps,"B1","PILOTAGE  ·  Cockpit de decision CFO  —  Budget 2027",
  font=Font(size=17,bold=True,color=WHITE),fillc=NAVY,align=Alignment(horizontal="left",vertical="center",indent=1))
for r in(1,2):
    for col in "CDEFGHIJKLMNOPQR": ps["%s%d"%(col,r)].fill=fill(NAVY)
ps.row_dimensions[1].height=16; ps.row_dimensions[2].height=16

# ============================================================ BANDEAU KPI GROUPE (live, suit le scenario actif)
S(ps,"B3","Resultat groupe (scenario actif  cad!D3)  —  se recalcule quand tu bouges un parametre",
  font=Font(size=9,italic=True,color="808080"))
cards=[("CA 2027","=H46","# ##0 \"EUR\"",BLUE),
       ("EBITDA (apres siege)","=K46","# ##0 \"EUR\"",GREEN),
       ("Marge EBITDA","=L46","0.0%",NAVY),
       ("Effectif","=G46","# ##0",GREY_D),
       ("Croissance CA vs 2026","=H46/22544725-1","+0.0%;-0.0%",AMBER)]
col0=2  # B
for i,(lab,f,fmt,color) in enumerate(cards):
    c=col0+i*3
    a=ps.cell(4,c); b=ps.cell(4,c+2)
    ps.merge_cells(start_row=4,start_column=c,end_row=4,end_column=c+2)
    ps.merge_cells(start_row=5,start_column=c,end_row=6,end_column=c+2)
    lc=ps.cell(4,c); lc.value=lab; lc.font=Font(size=9,bold=True,color=WHITE); lc.fill=fill(color); lc.alignment=CEN
    for cc in range(c,c+3): ps.cell(4,cc).fill=fill(color)
    vc=ps.cell(5,c); vc.value=f; vc.font=Font(size=16,bold=True,color=color); vc.alignment=CEN
    vc.number_format=fmt; vc.border=box
    for cc in range(c,c+3):
        ps.cell(5,cc).border=box; ps.cell(5,cc).fill=fill("FFFFFF")
ps.row_dimensions[4].height=16; ps.row_dimensions[5].height=20; ps.row_dimensions[6].height=14

# ============================================================ § 1 CAP (table deja en 12-26)
S(ps,"B10","1 ·  Cap strategique par campus (marque x ville)  —  arbitrage du budget d'acquisition",font=SEC)
ps.merge_cells("B10:R10"); ps["B10"].fill=fill(BLUE_L); ps["B10"].alignment=Alignment(indent=1,vertical="center")
for col in "CDEFGHIJKLMNOPQR": ps["%s10"%col].fill=fill(BLUE_L)
S(ps,"B11","Tu montes/baisses le cap d'un campus -> le budget se REJOUE (somme groupe constante). Rouge = CAC eleve = a arbitrer.",
  font=Font(size=9,italic=True,color="808080"))

# ============================================================ § 2 SYNTHESE (deja en 28-46) — retitrer
S(ps,"B28","2 ·  Synthese par campus  (resultats budget, live)",font=SEC)
ps.merge_cells("B28:R28"); ps["B28"].fill=fill(GREEN_L); ps["B28"].alignment=Alignment(indent=1,vertical="center")
for col in "CDEFGHIJKLMNOPQR": ps["%s28"%col].fill=fill(GREEN_L)

# ============================================================ § 3 ALLOCATION — cles + effet
S(ps,"B48","3 ·  Allocation du cout complet  —  cles de repartition  &  effet immediat",font=SEC)
ps.merge_cells("B48:R48"); ps["B48"].fill=fill(AMBER_L); ps["B48"].alignment=Alignment(indent=1,vertical="center")
for col in "CDEFGHIJKLMNOPQR": ps["%s48"%col].fill=fill(AMBER_L)
S(ps,"B49","Change une cle -> la marge complete se REDISTRIBUE entre marques (total groupe constant). C'est l'effet siege.",
  font=Font(size=9,italic=True,color="808080"))
# --- cles (relocalisees) : B51:E53 ---
S(ps,"B51","Parametre",font=H1,fillc=NAVY,align=CEN,border=box)
S(ps,"C51","Clef retenue",font=H1,fillc=NAVY,align=CEN,border=box)
ps.merge_cells("C51:E51")
keys=[("ALLOC_GRP_BRAND","E52","Chiffre d'affaires"),
      ("ALLOC_BRAND_CAMP","E53","Effectif"),
      ("ALLOC_CAMP_CLASS","E54","Nombre de classes")]
dv=DataValidation(type="list",formula1="=$A$1:$A$3",allow_blank=False); ps.add_data_validation(dv)
for i,(name,cell,val) in enumerate(keys):
    r=52+i
    S(ps,"B%d"%r,name,font=LABB,align=LEF,border=box,fillc=GREY_L)
    ps.merge_cells("C%d:E%d"%(r,r))
    S(ps,"C%d"%r,val,font=INPF,align=CEN,border=box,fillc=INPUT)
    for cc in ("D","E"): ps["%s%d"%(cc,r)].fill=fill(INPUT); ps["%s%d"%(cc,r)].border=box
    dv.add("C%d"%r)
KCELL={"K1":"$C$52","K2":"$C$53","K3":"$C$54"}  # GRP_BRAND, BRAND_CAMP, CAMP_CLASS

# --- effet : cout complet & marge par marque (live, reagit aux cles) : B56:K62 ---
heads=["Marque","VAC","PERM","ODIR","STRUCT","SIEGE","Cout complet","Marge complete","Marge %"]
for i,h in enumerate(heads):
    c=ps.cell(56,2+i,h); c.font=H1; c.fill=fill(AMBER); c.alignment=CEN; c.border=box
ps.row_dimensions[56].height=24
alcol={"VAC":"V","PERM":"W","ODIR":"X","STRUCT":"Y","SIEGE":"Z","MARGE":"AA","CA":"K"}
for i,(code,lab) in enumerate(MARQUES):
    r=57+i; band=WHITE if i%2==0 else GREY_L
    S(ps,"B%d"%r,lab,font=LABB,align=LEF,border=box,fillc=band)
    # couts VAC..SIEGE (reference 2026)
    for j,k in enumerate(["VAC","PERM","ODIR","STRUCT","SIEGE"]):
        S(ps,"%s%d"%(chr(67+j),r),'=SUMIFS(Allocation!$%s:$%s,Allocation!$E:$E,"%s",Allocation!$C:$C,"2026")'%(alcol[k],alcol[k],code),
          font=VAL,align=RIG,border=box,fillc=band,fmt="# ##0")
    S(ps,"H%d"%r,"=SUM(C%d:G%d)"%(r,r),font=VAL,align=RIG,border=box,fillc=band,fmt="# ##0")
    S(ps,"I%d"%r,'=SUMIFS(Allocation!$AA:$AA,Allocation!$E:$E,"%s",Allocation!$C:$C,"2026")'%code,
      font=LIVEF,align=RIG,border=box,fillc=LIVE,fmt="# ##0")
    S(ps,"J%d"%r,'=IFERROR(I%d/SUMIFS(Allocation!$K:$K,Allocation!$E:$E,"%s",Allocation!$C:$C,"2026"),0)'%(r,code),
      font=VAL,align=RIG,border=box,fillc=band,fmt="0.0%")
# total row 62
S(ps,"B62","GROUPE",font=Font(bold=True,color=WHITE),fillc=NAVY,align=LEF,border=box)
for col in "CDEFGHI":
    S(ps,"%s62"%col,"=SUM(%s57:%s61)"%(col,col),font=Font(bold=True,color=WHITE),fillc=NAVY,align=RIG,border=box,fmt="# ##0")
S(ps,"J62","=IFERROR(I62/SUMIFS(Allocation!$K:$K,Allocation!$C:$C,\"2026\"),0)",font=Font(bold=True,color=WHITE),fillc=NAVY,align=RIG,border=box,fmt="0.0%")
S(ps,"B55","Reference structurelle 2026 (apres allocation complete du siege) :",font=Font(size=9,italic=True,color="808080"))

# ============================================================ ROLLUP marque (feed charts) X:AA cols 24-27, rows 66-71
S(ps,"W65","Rollup marque (live)",font=Font(size=9,bold=True,color="808080"))
rollh=["Code","Marque","CA","EBITDA","Marge compl."]
for i,h in enumerate(rollh): ps.cell(66,23+i,h).font=Font(size=8,bold=True,color="808080")
for i,(code,lab) in enumerate(MARQUES):
    r=67+i
    ps.cell(r,23,code).font=Font(size=8,color="D9D9D9")
    ps.cell(r,24,lab).font=LAB
    S(ps,"Y%d"%r,'=SUMIFS(Moteur!$R:$R,Moteur!$E:$E,$W%d,Moteur!$B:$B,cad!$D$3)'%r,align=RIG,fmt="# ##0",font=VAL)
    S(ps,"Z%d"%r,"=SUMIFS($K$30:$K$43,$E$30:$E$43,$X%d)"%r,align=RIG,fmt="# ##0",font=VAL)
    S(ps,"AA%d"%r,"=I%d"%(57+i),align=RIG,fmt="# ##0",font=VAL)

# ============================================================ update _CALC_ALLOC key refs
for r in range(2,2001):
    ca["AC%d"%r]=('=IF($A{r}="","",IF(Pilotage!{k}="Chiffre d\'affaires",$I{r},'
                  'IF(Pilotage!{k}="Effectif",$G{r},$H{r})))').format(r=r,k=KCELL["K3"])
    ca["AD%d"%r]=('=IF($A{r}="","",IF(Pilotage!{k}="Chiffre d\'affaires",$N{r},'
                  'IF(Pilotage!{k}="Effectif",$L{r},$M{r})))').format(r=r,k=KCELL["K3"])
    ca["AE%d"%r]=('=IF($A{r}="","",IF(Pilotage!{k}="Chiffre d\'affaires",$N{r},'
                  'IF(Pilotage!{k}="Effectif",$L{r},$M{r})))').format(r=r,k=KCELL["K2"])
    ca["AF%d"%r]=('=IF($A{r}="","",IF(Pilotage!{k}="Chiffre d\'affaires",$S{r},'
                  'IF(Pilotage!{k}="Effectif",$Q{r},$R{r})))').format(r=r,k=KCELL["K2"])
    ca["AG%d"%r]=('=IF($A{r}="","",IF(Pilotage!{k}="Chiffre d\'affaires",$S{r},'
                  'IF(Pilotage!{k}="Effectif",$Q{r},$R{r})))').format(r=r,k=KCELL["K1"])
    ca["AH%d"%r]=('=IF($A{r}="","",IF(Pilotage!{k}="Chiffre d\'affaires",$V{r},'
                  'IF(Pilotage!{k}="Effectif",$T{r},$U{r})))').format(r=r,k=KCELL["K1"])

# ============================================================ GRAPHES soignes
def color_series(s,hexc):
    s.graphicalProperties.solidFill=hexc; s.graphicalProperties.line.solidFill=hexc
# 1) cap combo (near cap, col Q area) -> place at Q11
bar=BarChart(); bar.type="col"; bar.style=10; bar.title="Cap : budget acquisition  reference -> rejoue (live)"
bar.height=7.5; bar.width=15
bar.add_data(Reference(ps,min_col=14,min_row=12,max_row=26),titles_from_data=True)  # N ref
bar.add_data(Reference(ps,min_col=17,min_row=12,max_row=26),titles_from_data=True)  # Q rejoue
bar.set_categories(Reference(ps,min_col=6,min_row=13,max_row=26))
color_series(bar.series[0],"BDD7EE"); color_series(bar.series[1],BLUE)
bar.y_axis.numFmt="# ##0"; bar.gapWidth=40
line=LineChart(); line.add_data(Reference(ps,min_col=7,min_row=12,max_row=26),titles_from_data=True)  # G CAC
line.set_categories(Reference(ps,min_col=6,min_row=13,max_row=26))
line.y_axis.axId=200; line.y_axis.crosses="max"; color_series(line.series[0],RED)
bar+=line
ps.add_chart(bar,"S4")
# 2) donut poids marque CA (near synthese)
dn=DoughnutChart(); dn.title="Poids marques · CA 2027"; dn.height=7.5; dn.width=8
dn.add_data(Reference(ps,min_col=25,min_row=66,max_row=71),titles_from_data=True)   # Y CA
dn.set_categories(Reference(ps,min_col=24,min_row=67,max_row=71))
dn.dataLabels=DataLabelList(); dn.dataLabels.showPercent=True
s=dn.series[0]
s.data_points=dpts(MARQUES)
ps.add_chart(dn,"S30")
# 3) CA & EBITDA par campus
ce=BarChart(); ce.type="col"; ce.style=10; ce.title="CA & EBITDA par campus (live)"; ce.height=7.5; ce.width=15
ce.add_data(Reference(ps,min_col=8,min_row=29,max_row=43),titles_from_data=True)   # H CA
ce.add_data(Reference(ps,min_col=11,min_row=29,max_row=43),titles_from_data=True)  # K EBITDA
ce.set_categories(Reference(ps,min_col=4,min_row=30,max_row=43))
color_series(ce.series[0],BLUE); color_series(ce.series[1],GREEN); ce.y_axis.numFmt="# ##0"
ps.add_chart(ce,"S45")
# 4) EFFET DES CLES : marge complete par marque (reagit aux cles) — a cote des cles
mm=BarChart(); mm.type="col"; mm.style=12; mm.title="Effet des cles : marge complete par marque"
mm.height=7.5; mm.width=13
mm.add_data(Reference(ps,min_col=9,min_row=56,max_row=61),titles_from_data=True)   # I marge
mm.set_categories(Reference(ps,min_col=2,min_row=57,max_row=61))
mm.legend=None
sp=mm.series[0]
sp.data_points=dpts(MARQUES)
mm.y_axis.numFmt="# ##0"
ps.add_chart(mm,"L51")
# 5) decomposition cout complet par marque (stacked)
st=BarChart(); st.type="col"; st.grouping="stacked"; st.overlap=100; st.title="Decomposition du cout complet par marque"
st.height=7.5; st.width=13
for j,k in enumerate(["VAC","PERM","ODIR","STRUCT","SIEGE"]):
    st.add_data(Reference(ps,min_col=3+j,min_row=56,max_row=61),titles_from_data=True)
st.set_categories(Reference(ps,min_col=2,min_row=57,max_row=61))
st.y_axis.numFmt="# ##0"
ps.add_chart(st,"L67")

# widths
for col,w in {"B":20,"C":11,"D":11,"E":11,"F":16,"M":15,"Q":16}.items(): ps.column_dimensions[col].width=w

wb.calculation.fullCalcOnLoad=True
wb.save(F)
print("OK presentation : KPI band + flux Cap/Synthese/Allocation, cles relocalisees pres de l'effet, 5 graphes soignes.")
