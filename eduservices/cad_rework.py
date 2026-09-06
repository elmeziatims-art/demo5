#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""cad : vrai espace entre leviers Revenus / Couts + libelles Cadrage/Optimiste/Prudent.
Recable : couts deplaces (E22-26 -> E23-27) et le scenario actif passe par un code (P1)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
wb=openpyxl.load_workbook("CAD_SAAD_LIVE.xlsx")
cad=wb["cad"]
def fill(h):return PatternFill("solid",fgColor=h)
BLUE="2E86DE";BLUE_L="EAF3FC";GREEN="27AE60";GREEN_L="EAF7EF";AMBER="BF8F00";AMBER_L="FFF2CC"
NAVY="15406E";INPUT="FFF2CC";LIVE="C6E0B4"
thin=Side(style="thin",color="D9D9D9");box=Border(thin,thin,thin,thin)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True);LEF=Alignment(horizontal="left",vertical="center")

# ---------- libelles descriptifs (V01/V02/V03 -> Cadrage/Optimiste/Prudent) ----------
def hdr(r):
    cells={"D":("Reference","3B3B3B","ECEEF1"),"E":("Cadrage","1B5FA6",BLUE_L),
           "F":("Optimiste","2E7D42",GREEN_L),"G":("Prudent","9A6B00",AMBER_L),"H":("ACTIF (scenario)","375623",LIVE)}
    for col,(t,fc,bg) in cells.items():
        c=cad["%s%d"%(col,r)]; c.value=t; c.font=Font(size=9,bold=True,color=fc); c.fill=fill(bg); c.alignment=CEN; c.border=box
hdr(15); hdr(29)
cad["B14"]="(2)  Leviers de CROISSANCE / REVENUS  (-> CA, moteur)"
cad["B14"].font=Font(size=11,bold=True,color=BLUE)

# ---------- scenario actif : dropdown descriptif + code helper P1 ----------
cad["D3"]="Cadrage"
cad["D3"].font=Font(size=12,bold=True,color="1B5FA6"); cad["D3"].fill=fill(INPUT)
cad["D3"].border=Border(*(Side(style="medium",color="E0A800"),)*4)
for d in list(cad.data_validations.dataValidation):
    if "D3" in str(d.sqref): cad.data_validations.dataValidation.remove(d)
dv=DataValidation(type="list",formula1='"Cadrage,Optimiste,Prudent"',allow_blank=False); cad.add_data_validation(dv); dv.add("D3")
cad["O1"]="code scénario ->"; cad["O1"].font=Font(size=8,italic=True,color="9AA0AE"); cad["O1"].alignment=Alignment(horizontal="right")
cad["P1"]='=IF($D$3="Cadrage","V01",IF($D$3="Optimiste","V02","V03"))'
cad["P1"].font=Font(size=8,italic=True,color="9AA0AE")

# ---------- vrai espace + bloc COUTS deplacé en 23-27 ----------
# nettoyer l'ancien bloc couts (22-27)
for r in range(22,28):
    for col in "BCDEFGH":
        c=cad["%s%d"%(col,r)]; c.value=None; c.fill=PatternFill(); c.border=Border()
# ligne 22 = SEPARATEUR + titre COUTS
cad.merge_cells("B22:H22")
cad["B22"]="(2 bis)  Leviers de COUTS  (-> P&L)"
cad["B22"].font=Font(size=11,bold=True,color=GREEN); cad["B22"].fill=fill(GREEN_L)
cad["B22"].alignment=Alignment(horizontal="left",vertical="center",indent=1)
for col in "CDEFGH": cad["%s22"%col].fill=fill(GREEN_L)
cad.row_dimensions[22].height=22
# couts 23-27
COUTS=[("Inflation des charges externes",0,0.02,0.015,0.03),
       ("Politique salariale (masse permanente)",0,0.025,0.02,0.03),
       ("Variation des effectifs permanents",0,0.04,0.03,0.05),
       ("Effort de productivite (achats & structure)",0,0.01,0.03,0),
       ("Variation des couts de structure (loyers, IT, siege)",0,0,-0.03,0.04)]
INPF=Font(size=10,bold=True,color="7F6000");VAL=Font(size=9);LIVEF=Font(size=10,bold=True,color="375623")
gold=Side(style="medium",color="E0A800");ibord=Border(gold,gold,gold,gold)
for i,(lab,rf,v1,v2,v3) in enumerate(COUTS):
    r=23+i; band="FFFFFF" if i%2==0 else "F2F2F2"
    cad["B%d"%r]=lab; cad["B%d"%r].font=Font(size=9,color="3B3B3B"); cad["B%d"%r].alignment=LEF
    cad["B%d"%r].fill=fill(GREEN_L); cad["B%d"%r].border=Border(left=Side(style="thick",color=GREEN),top=thin,bottom=thin)
    cad["C%d"%r]="%"; cad["C%d"%r].font=VAL; cad["C%d"%r].alignment=CEN; cad["C%d"%r].border=box; cad["C%d"%r].fill=fill(band)
    cad["D%d"%r]=rf; cad["D%d"%r].number_format="0.0%"; cad["D%d"%r].font=VAL; cad["D%d"%r].alignment=CEN; cad["D%d"%r].border=box; cad["D%d"%r].fill=fill(band)
    for col,v,bg in [("E",v1,INPUT),("F",v2,GREEN_L),("G",v3,AMBER_L)]:
        c=cad["%s%d"%(col,r)]; c.value=v; c.number_format="0.0%"; c.alignment=CEN
        c.font=INPF if col=="E" else VAL; c.fill=fill(bg); c.border=ibord if col=="E" else box
    cad["H%d"%r]="=INDEX(D%d:G%d,MATCH($D$3,$D$15:$G$15,0))"%(r,r)
    cad["H%d"%r].number_format="0.0%"; cad["H%d"%r].font=LIVEF; cad["H%d"%r].alignment=CEN; cad["H%d"%r].border=box; cad["H%d"%r].fill=fill(LIVE)

# ---------- repointer _CALC_PNL couts E22-26 -> E23-27 ----------
cp=wb["_CALC_PNL"]
rowmap={"I":23,"J":24,"K":25,"L":26,"M":27}  # INFL,SAL,FTE,PROD,STRUCT
for col,newrow in rowmap.items():
    for r in range(2,6001):
        cp["%s%d"%(col,r)]=('=IF($A%d="","",IF($D%d="V01",cad!$E$%d,IF($D%d="V02",cad!$F$%d,cad!$G$%d)))'
                            %(r,r,newrow,r,newrow,newrow))

# ---------- repointer le scenario actif D3 -> code P1 (synthese + reconciliation) ----------
ps=wb["Pilotage"]
for row in ps.iter_rows():
    for c in row:
        if isinstance(c.value,str) and "cad!$D$3" in c.value:
            c.value=c.value.replace("cad!$D$3","cad!$P$1")
for cell in ("E7","E8","E10"):
    v=cad[cell].value
    if isinstance(v,str): cad[cell]=v.replace("$D$3","$P$1")

wb.calculation.fullCalcOnLoad=True
wb.save("CAD_SAAD_LIVE.xlsx")
print("OK cad : espace + bloc Couts en 23-27, libelles descriptifs, code scenario P1, _CALC_PNL repointe.")
