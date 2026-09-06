#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Refonte 2 :
 - Onglet cad : reprend la maquette du simulateur principal 01_Cadrage
   (Poste de commande CFO : scenario actif + colonne ACTIF, reconciliation
   Reference/Cible/Construit/Ecart, coeff prix, leviers, frais). Le moteur
   reste vivant : on REPOINTE _CALC_MOTEUR/_CALC_PNL vers les nouvelles cellules,
   et TOUT suit le scenario actif (cad!D3).
 - Pilotage cap : le CFO voit le budget REJOUE (live) ; les valeurs statiques
   (analytique de la vue) sont masquees a droite.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import Reference
F="CAD_SAAD_LIVE.xlsx"; wb=openpyxl.load_workbook(F)

NAVY="1F3864";BLUE="2E75B6";BLUE_L="DDEBF7";GREEN="548235";GREEN_L="E2EFDA"
AMBER="BF8F00";AMBER_L="FFF2CC";INPUT="FFF2CC";GREY_D="3B3B3B";GREY_L="F2F2F2"
WHITE="FFFFFF";LIVE="C6E0B4";RED="C00000"
thin=Side(style="thin",color="BFBFBF");box=Border(thin,thin,thin,thin)
med=Side(style="medium",color=NAVY)
def fill(h):return PatternFill("solid",fgColor=h)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)
RIG=Alignment(horizontal="right",vertical="center")
def S(ws,ref,val=None,font=None,fillc=None,align=None,border=None,fmt=None):
    c=ws[ref]
    if val is not None:c.value=val
    if font:c.font=font
    if fillc:c.fill=fill(fillc)
    if align:c.alignment=align
    if border:c.border=border
    if fmt:c.number_format=fmt
    return c
TITLE=Font(size=16,bold=True,color=WHITE);SUB=Font(size=9,italic=True,color="D9E1F2")
H1=Font(size=10,bold=True,color=WHITE);LAB=Font(size=9,color=GREY_D)
LABB=Font(size=10,bold=True,color=NAVY);VAL=Font(size=9,color="000000")
INPF=Font(size=10,bold=True,color="7F6000");LIVEF=Font(size=10,bold=True,color="375623")
SEC=Font(size=11,bold=True,color=NAVY)

# =========================================================== 1) NETTOYER cad
cad=wb["cad"]; cad._charts=[]
for mr in list(cad.merged_cells.ranges): cad.unmerge_cells(str(mr))
for row in cad.iter_rows(min_row=1,max_row=70,min_col=1,max_col=20):
    for c in row:
        c.value=None; c.fill=PatternFill(); c.font=Font(); c.border=Border()
        c.number_format="General"; c.alignment=Alignment()
cad.sheet_view.showGridLines=False
for col,w in {"A":2,"B":40,"C":6,"D":11,"E":11,"F":11,"G":11,"H":11,"I":2,
              "J":20,"K":12,"L":11,"M":2}.items(): cad.column_dimensions[col].width=w

# ---- Titre ----
cad.merge_cells("B1:H2")
S(cad,"B1","POSTE DE COMMANDE CFO  —  Cadrage CA & EBITDA 2027",font=TITLE,fillc=NAVY,
  align=Alignment(horizontal="left",vertical="center",indent=1))
for r in(1,2):
    for col in "CDEFGH": cad["%s%d"%(col,r)].fill=fill(NAVY)
# ---- ligne scenario actif + cibles ----
S(cad,"B3","Scenario actif :",font=LABB,align=RIG)
S(cad,"D3","V01",font=Font(size=12,bold=True,color=WHITE),fillc=BLUE,align=CEN,border=box)
dv=DataValidation(type="list",formula1='"V01,V02,V03"',allow_blank=False)
cad.add_data_validation(dv); dv.add("D3")
S(cad,"E3","Croissance CA cible :",font=LABB,align=RIG)
S(cad,"F3",0.05,font=INPF,fillc=INPUT,align=CEN,border=box,fmt="0.0%")
S(cad,"G3","Marge EBITDA cible :",font=LABB,align=RIG)
S(cad,"H3",0.15,font=INPF,fillc=INPUT,align=CEN,border=box,fmt="0.0%")

# ---- (1) Reconciliation Reference / Cible / Construit(actif) / Ecart ----
S(cad,"B5","(1)  Reconciliation  —  Reference  ·  Cible  ·  Construit (scenario actif)",font=SEC)
cad.merge_cells("B5:H5"); cad["B5"].fill=fill(GREEN_L); cad["B5"].alignment=Alignment(indent=1,vertical="center")
for col in "CDEFGH": cad["%s5"%col].fill=fill(GREEN_L)
recon_h=["Indicateur","Reference","Cible","Construit","Ecart","Ecart %"]
for i,h in enumerate(recon_h):
    S(cad,"%s6"%chr(66+i),h,font=H1,fillc=GREEN if i else NAVY,align=CEN,border=box)
# CA (7), EBITDA (8), Marge% (9), Effectif (10)
DV3="cad!$D$3"
S(cad,"B7","Chiffre d'affaires",font=LAB,align=LEF,border=box)
S(cad,"C7",22544725,font=VAL,align=RIG,border=box,fmt="# ##0")
S(cad,"D7","=C7*(1+F3)",font=VAL,align=RIG,border=box,fmt="# ##0")
S(cad,"E7",'=SUMIFS(Moteur!$R$2:$R$175,Moteur!$B$2:$B$175,$D$3)',font=LIVEF,fillc=LIVE,align=RIG,border=box,fmt="# ##0")
S(cad,"F7","=E7-D7",font=VAL,align=RIG,border=box,fmt="+# ##0;-# ##0")
S(cad,"G7","=IFERROR(E7/D7-1,0)",font=VAL,align=RIG,border=box,fmt="+0.0%;-0.0%")
S(cad,"B8","EBITDA (apres siege)",font=LAB,align=LEF,border=box)
S(cad,"C8",3291530,font=VAL,align=RIG,border=box,fmt="# ##0")
S(cad,"D8","=D7*H3",font=VAL,align=RIG,border=box,fmt="# ##0")
S(cad,"E8",'=SUMIFS(_CALC_PNL!$T$2:$T$1347,_CALC_PNL!$D$2:$D$1347,$D$3,_CALC_PNL!$C$2:$C$1347,"2027")',
  font=LIVEF,fillc=LIVE,align=RIG,border=box,fmt="# ##0")
S(cad,"F8","=E8-D8",font=VAL,align=RIG,border=box,fmt="+# ##0;-# ##0")
S(cad,"G8","=IFERROR(E8/D8-1,0)",font=VAL,align=RIG,border=box,fmt="+0.0%;-0.0%")
S(cad,"B9","Marge EBITDA %",font=LAB,align=LEF,border=box)
S(cad,"C9","=IFERROR(C8/C7,0)",font=VAL,align=RIG,border=box,fmt="0.0%")
S(cad,"D9","=H3",font=VAL,align=RIG,border=box,fmt="0.0%")
S(cad,"E9","=IFERROR(E8/E7,0)",font=LIVEF,fillc=LIVE,align=RIG,border=box,fmt="0.0%")
S(cad,"G9","=IFERROR(E9-D9,0)",font=VAL,align=RIG,border=box,fmt="+0.0%;-0.0%")
S(cad,"B10","Effectif total",font=LAB,align=LEF,border=box)
S(cad,"C10",3036,font=VAL,align=RIG,border=box,fmt="# ##0")
S(cad,"E10",'=SUMIFS(Moteur!$P$2:$P$175,Moteur!$B$2:$B$175,$D$3)',font=LIVEF,fillc=LIVE,align=RIG,border=box,fmt="# ##0")
S(cad,"F10","=E10-C10",font=VAL,align=RIG,border=box,fmt="+# ##0;-# ##0")
S(cad,"G10","=IFERROR(E10/C10-1,0)",font=VAL,align=RIG,border=box,fmt="+0.0%;-0.0%")
# reste a trouver
S(cad,"B12","RESTE A TROUVER  —  CA :",font=Font(bold=True,color=RED),align=RIG)
S(cad,"D12","=IF(D7-E7>0,D7-E7,0)",font=Font(bold=True,color=RED),align=RIG,border=box,fmt="# ##0")
S(cad,"E12","EBITDA :",font=Font(bold=True,color=RED),align=RIG)
S(cad,"F12","=IF(D8-E8>0,D8-E8,0)",font=Font(bold=True,color=RED),align=RIG,border=box,fmt="# ##0")

# ---- Coeff prix par marque (droite) : anchors K7:K11 ----
S(cad,"J5","Coeff prix par marque (decision)",font=SEC)
cad.merge_cells("J5:L5"); cad["J5"].fill=fill(BLUE_L); cad["J5"].alignment=Alignment(indent=1,vertical="center")
for col in("K","L"): cad["%s5"%col].fill=fill(BLUE_L)
S(cad,"J6","Marque",font=H1,fillc=NAVY,align=CEN,border=box)
S(cad,"K6","Coeff prix",font=H1,fillc=BLUE,align=CEN,border=box)
for i,(lab,val) in enumerate([("MBway",1.2),("ISCOM",1.15),("Ipac Bachelor Factory",0.95),
                              ("Pigier",0.9),("Tunon",1.05)]):
    r=7+i; band=WHITE if i%2==0 else GREY_L
    S(cad,"J%d"%r,lab,font=LAB,align=LEF,border=box,fillc=band)
    S(cad,"K%d"%r,val,font=INPF,align=CEN,border=box,fillc=INPUT,fmt="0.00")
S(cad,"J13","Coeff prix = seule decision par MARQUE (le prix suit l'indice ville).",
  font=Font(size=8,italic=True,color="808080"),align=LEF)
cad.merge_cells("J13:L14")

# ---- (2) Leviers : Reference / V01 / V02 / V03 / ACTIF ----
S(cad,"B14","(2)  Leviers  —  bascule par scenario (colonne ACTIF suit D3)",font=SEC)
cad.merge_cells("B14:H14"); cad["B14"].fill=fill(GREEN_L); cad["B14"].alignment=Alignment(indent=1,vertical="center")
for col in "CDEFGH": cad["%s14"%col].fill=fill(GREEN_L)
lev_h=["Parametre","Unite","Reference","V01","V02","V03","ACTIF"]
for i,h in enumerate(lev_h):
    c=cad.cell(15,2+i,h); c.font=H1; c.alignment=CEN; c.border=box
    c.fill=fill(BLUE if h=="V01" else GREEN if h=="V02" else AMBER if h=="V03" else NAVY if h!="ACTIF" else GREY_D)
LEVERS=[  # (label, unit, ref, V01, V02, V03)
 ("Variation du budget d'acquisition (-> leads payants)","%",0,0.08,0.15,-0.05),
 ("Variation du budget de marque (-> socle organique)","%",0,0.10,0.20,-0.05),
 ("Hausse tarifaire (prix)","%",0,0.02,0.035,0.02),
 ("Gain taux conversion Lead -> Candidature","pts",0,0.01,0.03,0),
 ("Gain taux conversion Admis -> Inscrit","pts",0,0.01,0.025,0),
 ("Amelioration du taux de passage","pts",0,0.005,0.015,-0.01),
 ("Inflation des charges externes","%",0,0.02,0.015,0.03),
 ("Politique salariale (masse permanente)","%",0,0.025,0.02,0.03),
 ("Variation des effectifs permanents","%",0,0.04,0.03,0.05),
 ("Effort de productivite (achats & structure)","%",0,0.01,0.03,0),
 ("Variation des couts de structure (loyers, IT, siege)","%",0,0,-0.03,0.04)]
for i,(lab,unit,ref,v1,v2,v3) in enumerate(LEVERS):
    r=16+i; band=WHITE if i%2==0 else GREY_L
    pf="0.0%"
    S(cad,"B%d"%r,lab,font=LAB,align=LEF,border=box,fillc=band)
    S(cad,"C%d"%r,unit,font=LAB,align=CEN,border=box,fillc=band)
    S(cad,"D%d"%r,ref,font=VAL,align=CEN,border=box,fillc=band,fmt=pf)
    S(cad,"E%d"%r,v1,font=INPF,align=CEN,border=box,fillc=INPUT,fmt=pf)
    S(cad,"F%d"%r,v2,font=VAL,align=CEN,border=box,fillc=GREEN_L,fmt=pf)
    S(cad,"G%d"%r,v3,font=VAL,align=CEN,border=box,fillc=AMBER_L,fmt=pf)
    S(cad,"H%d"%r,"=INDEX(D%d:G%d,MATCH($D$3,$D$15:$G$15,0))"%(r,r),
      font=LIVEF,align=CEN,border=box,fillc=LIVE,fmt=pf)
S(cad,"B27","Leviers 1-6 -> CA (moteur)  ·  leviers 7-11 -> P&L (budget).",
  font=Font(size=8,italic=True,color="808080"),align=LEF)

# ---- (3) Frais de dossier : E30:G30 ----
S(cad,"B28","(3)  Constante  —  frais de dossier (decision)",font=SEC)
cad.merge_cells("B28:H28"); cad["B28"].fill=fill(GREEN_L); cad["B28"].alignment=Alignment(indent=1,vertical="center")
for col in "CDEFGH": cad["%s28"%col].fill=fill(GREEN_L)
for i,h in enumerate(lev_h):
    c=cad.cell(29,2+i,h); c.font=H1; c.alignment=CEN; c.border=box
    c.fill=fill(BLUE if h=="V01" else GREEN if h=="V02" else AMBER if h=="V03" else NAVY if h!="ACTIF" else GREY_D)
S(cad,"B30","Frais de dossier / nouvel inscrit","","","","","")
S(cad,"B30","Frais de dossier / nouvel inscrit",font=LAB,align=LEF,border=box,fillc=WHITE)
S(cad,"C30","EUR",font=LAB,align=CEN,border=box,fillc=WHITE)
S(cad,"D30",90,font=VAL,align=CEN,border=box,fillc=WHITE,fmt="# ##0")
S(cad,"E30",90,font=INPF,align=CEN,border=box,fillc=INPUT,fmt="# ##0")
S(cad,"F30",90,font=VAL,align=CEN,border=box,fillc=GREEN_L,fmt="# ##0")
S(cad,"G30",90,font=VAL,align=CEN,border=box,fillc=AMBER_L,fmt="# ##0")
S(cad,"H30","=INDEX(D30:G30,MATCH($D$3,$D$15:$G$15,0))",font=LIVEF,align=CEN,border=box,fillc=LIVE,fmt="# ##0")

# =========================================================== 2) REPOINTER le moteur
cm=wb["_CALC_MOTEUR"]
LM={"V":"$E$16","W":"$E$17","X":"$E$18","Y":"$E$19","Z":"$E$20","AA":"$E$21","AB":"$E$30"}
LMc={"V":("E16","F16","G16"),"W":("E17","F17","G17"),"X":("E18","F18","G18"),
     "Y":("E19","F19","G19"),"Z":("E20","F20","G20"),"AA":("E21","F21","G21"),
     "AB":("E30","F30","G30")}
for r in range(2,176):
    for col,(a,b,c) in LMc.items():
        cm["%s%d"%(col,r)]='=IF(B%d="V01",cad!$%s$%s,IF(B%d="V02",cad!$%s$%s,cad!$%s$%s))'%(
            r,a[0],a[1:],r,b[0],b[1:],c[0],c[1:])
    cm["AC%d"%r]=('=IF(F{r}="MBWAY",cad!$K$7,IF(F{r}="ISCOM",cad!$K$8,IF(F{r}="IPAC",cad!$K$9,'
                  'IF(F{r}="PIGIER",cad!$K$10,cad!$K$11))))').format(r=r)

# =========================================================== 3) REPOINTER le P&L
cp=wb["_CALC_PNL"]
LP={"G":("E16","F16","G16"),"H":("E17","F17","G17"),"I":("E22","F22","G22"),
    "J":("E23","F23","G23"),"K":("E24","F24","G24"),"L":("E25","F25","G25"),"M":("E26","F26","G26")}
for r in range(2,1348):
    for col,(a,b,c) in LP.items():
        cp["%s%d"%(col,r)]='=IF(D%d="V01",cad!$%s$%s,IF(D%d="V02",cad!$%s$%s,cad!$%s$%s))'%(
            r,a[0],a[1:],r,b[0],b[1:],c[0],c[1:])

# =========================================================== 4) SYNTHESE & rollup suivent D3
ps=wb["Pilotage"]
for i in range(14):
    r=30+i
    ps["G%d"%r]='=SUMIFS(Moteur!$P$2:$P$175,Moteur!$D$2:$D$175,$D%d,Moteur!$B$2:$B$175,cad!$D$3)'%r
    ps["H%d"%r]='=SUMIFS(Moteur!$R$2:$R$175,Moteur!$D$2:$D$175,$D%d,Moteur!$B$2:$B$175,cad!$D$3)'%r
    ps["K%d"%r]=('=SUMIFS(_CALC_PNL!$T$2:$T$1347,_CALC_PNL!$A$2:$A$1347,$D%d,'
                 '_CALC_PNL!$D$2:$D$1347,cad!$D$3,_CALC_PNL!$C$2:$C$1347,"2027")')%r
ps["K45"]=('=SUMIFS(_CALC_PNL!$T$2:$T$1347,_CALC_PNL!$A$2:$A$1347,"GRP",'
           '_CALC_PNL!$D$2:$D$1347,cad!$D$3,_CALC_PNL!$C$2:$C$1347,"2027")')
for i in range(5):
    r=30+i
    ps["Y%d"%r]='=SUMIFS(Moteur!$R$2:$R$175,Moteur!$E$2:$E$175,$W%d,Moteur!$B$2:$B$175,cad!$D$3)'%r

# =========================================================== 5) CAP : montrer rejoue, masquer statique
# visible : D scenario, E periode, F entity, M cap retenu (input), Q rejoue (live)
# masque (analytique de la vue) : G,H,I,J,K,L,N,O,R,S,T,U
for col in ("G","H","I","J","K","L","N","O","R","S","T","U"):
    ps.column_dimensions[col].hidden=True
for col,w in {"D":12,"E":8,"F":18,"M":16,"Q":18}.items(): ps.column_dimensions[col].width=w
# relabels + mise en avant
S(ps,"M12","Cap retenu\n(decision)",font=H1,fillc=AMBER,align=CEN,border=box)
S(ps,"Q12","Budget acquisition\nREJOUE (live)",font=H1,fillc=GREEN,align=CEN,border=box)
ps.row_dimensions[12].height=30

wb.save(F)
print("OK refonte2 : cad = poste de commande CFO (scenario actif), moteur/PNL repointes,")
print("synthese+rollup suivent D3, cap = rejoue visible / statique masque.")
