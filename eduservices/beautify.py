#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Redesign des onglets 'cad' (Cadrage) et 'Pilotage' de CAD_SAAD_LIVE.xlsx.
On PRESERVE strictement les cellules d'entree (anchors) lues par les formules
vivantes et par Tagetik ; on ne fait qu'embellir autour + ajouter des graphes."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from copy import copy

F = "CAD_SAAD_LIVE.xlsx"
wb = openpyxl.load_workbook(F)

# ---- palette ----
NAVY   = "1F3864"; BLUE = "2E75B6"; BLUE_L = "DDEBF7"
GREEN  = "548235"; GREEN_L = "E2EFDA"
AMBER  = "BF8F00"; AMBER_L = "FFF2CC"
GREY_D = "3B3B3B"; GREY_L = "F2F2F2"; INPUT = "FFF2CC"; INPUTB = "BF8F00"
WHITE  = "FFFFFF"; LIVE = "C6E0B4"

thin = Side(style="thin", color="BFBFBF")
med  = Side(style="medium", color=NAVY)
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def setc(ws, ref, val=None, font=None, fillc=None, align=None, border=None, fmt=None):
    c = ws[ref]
    if val is not None: c.value = val
    if font: c.font = font
    if fillc: c.fill = fill(fillc)
    if align: c.alignment = align
    if border: c.border = border
    if fmt: c.number_format = fmt
    return c

CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEF = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIG = Alignment(horizontal="right", vertical="center")

TITLE = Font(name="Calibri", size=18, bold=True, color=WHITE)
SUB   = Font(name="Calibri", size=10, italic=True, color="D9E1F2")
H1    = Font(name="Calibri", size=11, bold=True, color=WHITE)
LAB   = Font(name="Calibri", size=10, color=GREY_D)
LABB  = Font(name="Calibri", size=10, bold=True, color=NAVY)
VAL   = Font(name="Calibri", size=10, color="000000")
INPF  = Font(name="Calibri", size=10, bold=True, color="7F6000")

# =====================================================================
#  ONGLET  cad  (CADRAGE)
# =====================================================================
ws = wb["cad"]
ws.sheet_view.showGridLines = False
for col, w in {"A":2,"B":3,"G":26,"H":44,"I":13,"J":13,"K":13,"L":2,
               "M":11,"N":11,"O":11,"P":11,"Q":11,"R":11}.items():
    ws.column_dimensions[col].width = w

# --- Titre banner rows 1-4 ---
ws.merge_cells("B2:K3")
setc(ws, "B2", "CADRAGE  ·  Hypotheses budgetaires 2027", font=TITLE,
     fillc=NAVY, align=Alignment(horizontal="left", vertical="center", indent=1))
for r in (2, 3):
    for col in "CDEFGHIJK":
        ws["%s%d" % (col, r)].fill = fill(NAVY)
ws.merge_cells("B4:K4")
setc(ws, "B4", "Masque de saisie CFO  —  V01 Cadrage / V02 Optimiste / V03 Prudent  ·  "
               "les colonnes vertes des vues reagissent en direct", font=SUB,
     fillc=BLUE, align=Alignment(horizontal="left", vertical="center", indent=1))
for col in "CDEFGHIJK":
    ws["%s4" % col].fill = fill(BLUE)
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 20
ws.row_dimensions[4].height = 16

# --- Section 1 : coef prix par marque (I8 existe deja) ---
setc(ws, "G8", "1 · Coefficient prix par marque (decision)", font=Font(size=11, bold=True, color=NAVY))
ws["I8"].value = None  # on remet le titre dans G8, propre
ws.merge_cells("G8:K8")
ws["G8"].fill = fill(BLUE_L); ws["G8"].alignment = Alignment(indent=1, vertical="center")
for col in "HIJK": ws["%s8" % col].fill = fill(BLUE_L)
# header ligne
setc(ws, "H8b" if False else "H8", None)
# table marques H9:I13  (H label, I input)
for i, r in enumerate(range(9, 14)):
    setc(ws, "H%d" % r, font=LABB if False else LAB, align=LEF, border=box,
         fillc=WHITE if i % 2 == 0 else GREY_L)
    setc(ws, "I%d" % r, font=INPF, align=CEN, border=Border(left=med,right=med,top=thin,bottom=thin),
         fillc=INPUT, fmt="0.00")

# --- Section 2 : leviers par version (H21:K31) ---
setc(ws, "G19", "2 · Parametres de projection  (saisie par version)",
     font=Font(size=11, bold=True, color=NAVY))
ws.merge_cells("G19:K19")
ws["G19"].fill = fill(BLUE_L); ws["G19"].alignment = Alignment(indent=1, vertical="center")
for col in "HIJK": ws["%s19" % col].fill = fill(BLUE_L)
# entete versions ligne 20
setc(ws, "G20", "Levier", font=H1, fillc=NAVY, align=CEN, border=box)
setc(ws, "H20", "", font=H1, fillc=NAVY, align=CEN, border=box)
setc(ws, "I20", "V01 · Cadrage",  font=H1, fillc=BLUE,  align=CEN, border=box)
setc(ws, "J20", "V02 · Optimiste", font=H1, fillc=GREEN, align=CEN, border=box)
setc(ws, "K20", "V03 · Prudent",   font=H1, fillc=AMBER, align=CEN, border=box)
ws.row_dimensions[20].height = 26
# lignes leviers 21-31 (% sauf frais ligne 37)
pct_rows = list(range(21, 32))
for i, r in enumerate(pct_rows):
    band = WHITE if i % 2 == 0 else GREY_L
    setc(ws, "H%d" % r, font=LAB, align=LEF, border=box, fillc=band)
    setc(ws, "I%d" % r, font=INPF, align=CEN, border=box, fillc=INPUT,  fmt="0.0%")
    setc(ws, "J%d" % r, font=VAL,  align=CEN, border=box, fillc=GREEN_L, fmt="0.0%")
    setc(ws, "K%d" % r, font=VAL,  align=CEN, border=box, fillc=AMBER_L, fmt="0.0%")

# --- Section 3 : frais de dossier (ligne 36-37) ---
setc(ws, "G35", "3 · Frais de dossier par nouvel inscrit (EUR)",
     font=Font(size=11, bold=True, color=NAVY))
ws.merge_cells("G35:K35")
ws["G35"].fill = fill(BLUE_L); ws["G35"].alignment = Alignment(indent=1, vertical="center")
for col in "HIJK": ws["%s35" % col].fill = fill(BLUE_L)
setc(ws, "I36", "V01",  font=H1, fillc=BLUE,  align=CEN, border=box)
setc(ws, "J36", "V02",  font=H1, fillc=GREEN, align=CEN, border=box)
setc(ws, "K36", "V03",  font=H1, fillc=AMBER, align=CEN, border=box)
setc(ws, "H37", font=LAB, align=LEF, border=box, fillc=WHITE)
setc(ws, "I37", font=INPF, align=CEN, border=box, fillc=INPUT,  fmt="# ##0")
setc(ws, "J37", font=VAL,  align=CEN, border=box, fillc=GREEN_L, fmt="# ##0")
setc(ws, "K37", font=VAL,  align=CEN, border=box, fillc=AMBER_L, fmt="# ##0")

# --- Section 4 : reference & cibles (G45:H57) — cartes KPI ---
setc(ws, "G44", "4 · Reference 2026  &  Cibles 2027", font=Font(size=11, bold=True, color=NAVY))
ws.merge_cells("G44:K44")
ws["G44"].fill = fill(BLUE_L); ws["G44"].alignment = Alignment(indent=1, vertical="center")
for col in "HIJK": ws["%s44" % col].fill = fill(BLUE_L)
setc(ws, "G45", "Reference 2026", font=Font(bold=True, color=WHITE), fillc=GREY_D, align=LEF, border=box)
setc(ws, "H45", "", fillc=GREY_D, border=box)
ref_rows = {46:"# ##0 \"EUR\"", 47:"# ##0 \"EUR\"", 49:"# ##0"}
for r, fmt in ref_rows.items():
    setc(ws, "G%d" % r, font=LABB, align=LEF, border=box, fillc=GREY_L)
    setc(ws, "H%d" % r, font=Font(bold=True, color=NAVY), align=RIG, border=box, fillc=WHITE, fmt=fmt)
setc(ws, "G48", font=LAB, align=LEF); setc(ws, "H48", font=Font(italic=True, color="808080"), align=RIG)
# cibles
setc(ws, "G55", font=LAB); setc(ws, "H55", font=Font(bold=True, color=NAVY))
setc(ws, "G56", "Croissance CA cible", font=LABB, align=LEF, border=box, fillc=AMBER_L)
setc(ws, "H56", font=INPF, align=RIG, border=box, fillc=INPUT, fmt="0.0%")
setc(ws, "G57", "Marge EBITDA cible", font=LABB, align=LEF, border=box, fillc=AMBER_L)
setc(ws, "H57", font=INPF, align=RIG, border=box, fillc=INPUT, fmt="0.0%")

# --- GRAPHE 1 : leviers par version (barres groupees) ---
ch = BarChart(); ch.type = "col"; ch.style = 10
ch.title = "Parametres par version"
ch.height = 8.5; ch.width = 20
data = Reference(ws, min_col=9, max_col=11, min_row=20, max_row=31)
cats = Reference(ws, min_col=8, min_row=21, max_row=31)
ch.add_data(data, titles_from_data=True)
ch.set_categories(cats)
ch.y_axis.numFmt = "0%"; ch.y_axis.majorGridlines = None
ch.gapWidth = 60
ws.add_chart(ch, "M20")

# --- GRAPHE 2 : coef prix par marque ---
ch2 = BarChart(); ch2.type = "bar"; ch2.style = 12
ch2.title = "Coef prix par marque"
ch2.height = 6.5; ch2.width = 10
d2 = Reference(ws, min_col=9, min_row=9, max_row=13)
c2 = Reference(ws, min_col=8, min_row=9, max_row=13)
ch2.add_data(d2, titles_from_data=False)
ch2.set_categories(c2)
ch2.legend = None
ws.add_chart(ch2, "M8")

# =====================================================================
#  ONGLET  Pilotage
# =====================================================================
ps = wb["Pilotage"]
ps.sheet_view.showGridLines = False
widths = {"A":2,"B":3,"C":3,"D":20,"E":20,"F":16,"G":13,"H":12,"I":12,"J":11,
          "K":11,"L":11,"M":12,"N":13,"O":15,"P":2,"Q":15,"R":13,"S":13,"T":13,"U":13}
for col, w in widths.items():
    ps.column_dimensions[col].width = w

# liste helper A1:A3 -> discrete
for r in (1,2,3):
    ps["A%d" % r].font = Font(size=8, color="BFBFBF")

# titre
ps.merge_cells("D2:O3")
setc(ps, "D2", "PILOTAGE  ·  Cap strategique & Cles d'allocation", font=TITLE,
     fillc=NAVY, align=Alignment(horizontal="left", vertical="center", indent=1))
for col in "EFGHIJKLMNO":
    for r in (2,3): ps["%s%d" % (col, r)].fill = fill(NAVY)
ps.row_dimensions[2].height = 20; ps.row_dimensions[3].height = 20

# --- Cles d'allocation D7:E10 ---
setc(ps, "D6", "Cles d'allocation (choix CFO)", font=Font(size=11, bold=True, color=NAVY))
ps.merge_cells("D6:E6"); ps["D6"].fill = fill(BLUE_L); ps["D6"].alignment = Alignment(indent=1, vertical="center")
ps["E6"].fill = fill(BLUE_L)
setc(ps, "D7", "Parametre", font=H1, fillc=NAVY, align=CEN, border=box)
setc(ps, "E7", "Cle retenue", font=H1, fillc=NAVY, align=CEN, border=box)
for r in (8,9,10):
    setc(ps, "D%d" % r, font=LABB, align=LEF, border=box, fillc=GREY_L)
    setc(ps, "E%d" % r, font=INPF, align=CEN, border=box, fillc=INPUT)
dv = DataValidation(type="list", formula1="=$A$1:$A$3", allow_blank=False)
ps.add_data_validation(dv); dv.add("E8"); dv.add("E9"); dv.add("E10")

# --- Table cap D12:U26 ---
setc(ps, "D11", "Cap strategique par campus  ·  budget acquisition rejoue (somme constante)",
     font=Font(size=11, bold=True, color=NAVY))
ps.merge_cells("D11:O11"); ps["D11"].fill = fill(BLUE_L); ps["D11"].alignment = Alignment(indent=1, vertical="center")
for col in "EFGHIJKLMNO": ps["%s11" % col].fill = fill(BLUE_L)
# header row 12
hdr_tag = {"D":"SCENARIO","E":"PERIODE","F":"ENTITY","G":"CAC_MARG","H":"CROISS","I":"INTENSITE",
           "J":"CAP_EFF","K":"CAP_MOM","L":"CAP_POT","M":"CAP_RETENU","N":"BUD_REF","O":"BUD_REJOUE"}
for col in hdr_tag:
    setc(ps, "%s12" % col, font=H1, fillc=NAVY, align=CEN, border=box)
for col in ("Q","R","S","T","U"):
    ps["%s12" % col].font = H1; ps["%s12" % col].fill = fill(GREEN); ps["%s12" % col].alignment = CEN; ps["%s12" % col].border = box
ps.row_dimensions[12].height = 24
# data rows 13-26
numfmt = {"G":"# ##0","H":"0.0%","I":"0.0%","J":"0.00","K":"0.00","L":"0.00",
          "M":"0.00","N":"# ##0","O":"# ##0","Q":"# ##0","R":"# ##0","S":"0.00","T":"0.00","U":"0.00"}
for i, r in enumerate(range(13, 27)):
    band = WHITE if i % 2 == 0 else GREY_L
    for col in ("D","E","F"):
        setc(ps, "%s%d" % (col, r), font=LAB, align=CEN if col!="F" else LEF, border=box, fillc=band)
    for col, fmt in numfmt.items():
        if col in ("Q","R","S","T","U"):
            setc(ps, "%s%d" % (col, r), font=VAL, align=RIG, border=box, fillc=LIVE, fmt=fmt)
        elif col == "M":
            setc(ps, "%s%d" % (col, r), font=INPF, align=RIG, border=box, fillc=INPUT, fmt=fmt)
        else:
            setc(ps, "%s%d" % (col, r), font=VAL, align=RIG, border=box, fillc=band, fmt=fmt)

# conditional color scale sur CAP_RETENU (M13:M26)
from openpyxl.formatting.rule import ColorScaleRule
ps.conditional_formatting.add("M13:M26",
    ColorScaleRule(start_type="min", start_color="F8696B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="63BE7B"))

# --- GRAPHE : budget reference vs rejoue (live) ---
chp = BarChart(); chp.type = "col"; chp.style = 10
chp.title = "Budget acquisition : reference vs rejoue (live)"
chp.height = 8.5; chp.width = 22
dref = Reference(ps, min_col=14, max_col=14, min_row=12, max_row=26)  # N (BUD_REF)
drej = Reference(ps, min_col=17, max_col=17, min_row=12, max_row=26)  # Q (REJOUE live)
chp.add_data(dref, titles_from_data=True)
chp.add_data(drej, titles_from_data=True)
chp.set_categories(Reference(ps, min_col=6, min_row=13, max_row=26))
chp.gapWidth = 50
ps.add_chart(chp, "D28")

# --- GRAPHE : cap retenu par campus ---
chc = BarChart(); chc.type = "bar"; chc.style = 12
chc.title = "Cap retenu par campus"
chc.height = 8.5; chc.width = 11
chc.add_data(Reference(ps, min_col=13, min_row=12, max_row=26), titles_from_data=True)
chc.set_categories(Reference(ps, min_col=6, min_row=13, max_row=26))
chc.legend = None
ps.add_chart(chc, "N28")

wb.save(F)
print("OK - onglets cad & Pilotage redesignes + 4 graphes. Anchors preserves.")
