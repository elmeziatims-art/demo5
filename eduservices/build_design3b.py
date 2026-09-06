#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Phase 5: PIL cap (baseline) + synthese (formules) + KPI ; ALLOC maille + helpers cles."""
import warnings;warnings.filterwarnings("ignore")
import openpyxl,os
from tgk_surgery import Book
Pd=openpyxl.load_workbook("CAD_SAAD_LIVE.xlsx",data_only=True)
b=Book("DESIGN3_OP.xlsm")
def putf(sh,ref,f): sh.set_formula(ref,f)
def putv(sh,ref,v):
    if isinstance(v,str): sh.set_text(ref,v)
    elif v is not None: sh.set_number(ref,v)

# ============== PIL CAP : baseline (proto cap G-N -> DESIGN3 D-K) ==============
pil=b.sheet("PIL"); pcap=Pd["Pilotage"]  # proto cap rows 13-26 : F=entity,G=CAC,H=Croiss,I=Int,J=CapEff,K=CapMom,L=CapPot,N=budget
CAPMAP={"D":7,"E":8,"F":9,"G":10,"H":11,"I":12,"K":14}  # DESIGN3 col -> proto col idx
for i in range(14):
    pr=13+i; dr=15+i
    for dc,pc in CAPMAP.items():
        v=pcap.cell(pr,pc).value
        if v is not None: putv(pil,"%s%d"%(dc,dr),v)
    putv(pil,"J%d"%dr,1)   # cap retenu = 1

# ============== PIL SYNTHESE (formules) rows 33-46 + 47/48/49 + KPI r7 ==============
SC="SCENARIO_CODE"
for i in range(14):
    r=33+i
    putf(pil,"D%d"%r,'SUMIFS(Moteur!$P:$P,Moteur!$D:$D,$A%d,Moteur!$B:$B,%s)'%(r,SC))
    putf(pil,"E%d"%r,'SUMIFS(Moteur!$R:$R,Moteur!$D:$D,$A%d,Moteur!$B:$B,%s)'%(r,SC))
    putf(pil,"F%d"%r,'IFERROR(E%d/D%d,0)'%(r,r))
    putf(pil,"G%d"%r,'IFERROR(E%d/SUM($E$33:$E$46),0)'%r)
    putf(pil,"H%d"%r,'SUMIFS(_CALC_PNL!$T:$T,_CALC_PNL!$A:$A,$A%d,_CALC_PNL!$D:$D,%s,_CALC_PNL!$C:$C,"2027")'%(r,SC))
    putf(pil,"I%d"%r,'IFERROR(H%d/E%d,0)'%(r,r))
    putf(pil,"J%d"%r,'IFERROR(H%d/D%d,0)'%(r,r))
    putf(pil,"K%d"%r,'SUMIFS(PIL!$K$15:$K$28,PIL!$A$15:$A$28,$A%d)*SUMIFS(PIL!$J$15:$J$28,PIL!$A$15:$A$28,$A%d)*(SUM(BUD_REF_CAP)/SUMPRODUCT(BUD_REF_CAP,HYP_CAP_RETENU))'%(r,r))
    putf(pil,"L%d"%r,'SUMIFS(PIL!$D$15:$D$28,PIL!$A$15:$A$28,$A%d)'%r)
# sous-total 47
for c in "DEHK": putf(pil,"%s47"%c,"SUM(%s33:%s46)"%(c,c))
putf(pil,"F47","IFERROR(E47/D47,0)"); putf(pil,"G47","SUM(G33:G46)")
putf(pil,"I47","IFERROR(H47/E47,0)"); putf(pil,"J47","IFERROR(H47/D47,0)")
# siege 48 (EBITDA GRP)
putf(pil,"H48",'SUMIFS(_CALC_PNL!$T:$T,_CALC_PNL!$A:$A,"GRP",_CALC_PNL!$D:$D,%s,_CALC_PNL!$C:$C,"2027")'%SC)
# groupe 49
putf(pil,"D49","D47"); putf(pil,"E49","E47"); putf(pil,"F49","IFERROR(E49/D49,0)")
putf(pil,"G49","SUM(G33:G46)"); putf(pil,"H49","H47+H48"); putf(pil,"I49","IFERROR(H49/E49,0)")
putf(pil,"J49","IFERROR(H49/D49,0)"); putf(pil,"K49","SUM(K33:K46)")
# KPI row 7
putf(pil,"B7","E49"); putf(pil,"H7","H49"); putf(pil,"K7","I49"); putf(pil,"N7","D49")

# ============== ALLOC : helpers cles FR->code + maille ==============
al=b.sheet("ALLOC")
def code(cell): return 'IF(%s="Chiffre d\'affaires","REV_CA",IF(%s="Nombre de classes","VOL_CLASS","VOL_EFF"))'%(cell,cell)
putf(al,"N1",code("$C$6")); putf(al,"N2",code("$C$7")); putf(al,"N3",code("$C$8")); putf(al,"N4",code("$C$9"))
# maille : parser les labels B18..B95
LIBM={"MBway":"MBWAY","ISCOM":"ISCOM","Ipac":"IPAC","Pigier":"PIGIER","Tunon":"TUNON"}
wbv=openpyxl.load_workbook("DESIGN3.xlsm")["ALLOC"]
import re
cur_ent=None
def crit_row(r,label):
    global cur_ent
    s=label.strip()
    m=re.search(r'\(([A-Z_]+)\)',label)
    if label in LIBM:                                   # marque
        return ',Allocation!$E:$E,"%s"'%LIBM[label]
    if m:                                               # campus
        cur_ent=m.group(1)
        return ',Allocation!$D:$D,"%s"'%cur_ent
    if label.startswith("      ") and cur_ent:          # classe
        parts=s.split()
        if len(parts)>=3:
            prog,an,mod=parts[0],parts[1],parts[2]
            return ',Allocation!$D:$D,"%s",Allocation!$F:$F,"%s",Allocation!$G:$G,"%s",Allocation!$H:$H,"%s"'%(cur_ent,prog,an,mod)
    if s=="GROUPE": return ''                            # total groupe
    return None
for r in range(18,96):
    lab=wbv.cell(r,2).value
    if lab is None: continue
    crit=crit_row(r,lab)
    if crit is None: continue
    base=',Allocation!$C:$C,"2026"'+crit
    putf(al,"C%d"%r,'SUMIFS(Allocation!$I:$I%s)'%base)     # Effectif
    putf(al,"D%d"%r,'SUMIFS(Allocation!$K:$K%s)'%base)     # CA
    putf(al,"E%d"%r,'SUMIFS(Allocation!$V:$V%s)'%base)     # VAC live
    putf(al,"F%d"%r,'SUMIFS(Allocation!$W:$W%s)'%base)     # PERM
    putf(al,"G%d"%r,'SUMIFS(Allocation!$X:$X%s)'%base)     # ODIR
    putf(al,"H%d"%r,'SUMIFS(Allocation!$Y:$Y%s)'%base)     # STRUCT
    putf(al,"I%d"%r,'SUMIFS(Allocation!$AC:$AC%s)'%base)   # Frais marque
    putf(al,"J%d"%r,'SUMIFS(Allocation!$Z:$Z%s)'%base)     # Holding
    putf(al,"K%d"%r,'E%d+F%d+G%d+H%d+I%d+J%d'%(r,r,r,r,r,r))  # Cout complet
    putf(al,"L%d"%r,'SUMIFS(Allocation!$AA:$AA%s)'%base)   # Marge complete
    putf(al,"M%d"%r,'IFERROR(L%d/D%d,0)'%(r,r))            # Marge %

b.set_fullcalc(); b.save("_tmp_d3b.xlsx"); os.replace("_tmp_d3b.xlsx","DESIGN3_OP.xlsm")
print("OK phase 5 (PIL cap/synthese/KPI, ALLOC maille).")
