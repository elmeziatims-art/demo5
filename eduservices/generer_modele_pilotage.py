# -*- coding: utf-8 -*-
"""
Modèle de pilotage budgétaire EDUSERVICES GROUP (pré-Tagetik).
Maille fine (marque x campus x programme x niveau x modalité), funnel CRM,
coefficients stratégiques, classes dérivées, décision ouvrir/fermer,
allocation par driver, P&L consolidé + pont Prix/Volume/Mix.
Données ILLUSTRATIVES mais calibrées sur des ordres de grandeur sourcés.
"""
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as GL

OUT = "/home/user/demo5/eduservices/EDUSERVICES_Modele_Pilotage_Budget.xlsx"

# ---------------------------------------------------------------- styles
NAVY,BLUE2,LIGHT,LIGHT2,YEL,TOT = "1F3864","2E5496","D9E1F2","EDF1F9","FFF2CC","E2EFDA"
FONT="Arial"
CIN =Font(name=FONT,color="0000FF")
CINB=Font(name=FONT,color="0000FF",bold=True)
CF  =Font(name=FONT,color="000000")
CFB =Font(name=FONT,color="000000",bold=True)
CL  =Font(name=FONT,color="008000")
CHDR=Font(name=FONT,color="FFFFFF",bold=True)
CTIT=Font(name=FONT,color="FFFFFF",bold=True,size=14)
CB  =Font(name=FONT,bold=True)
CIT =Font(name=FONT,italic=True,color="595959",size=9)
CREG=Font(name=FONT)
FN_NAVY=PatternFill("solid",fgColor=NAVY); FN_BLUE=PatternFill("solid",fgColor=BLUE2)
FN_LIGHT=PatternFill("solid",fgColor=LIGHT); FN_YEL=PatternFill("solid",fgColor=YEL)
FN_TOT=PatternFill("solid",fgColor=TOT)
thin=Side(style="thin",color="BFBFBF"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
AL=Alignment(horizontal="left",vertical="center"); AC=Alignment(horizontal="center",vertical="center",wrap_text=True)
AR=Alignment(horizontal="right",vertical="center"); ALW=Alignment(horizontal="left",vertical="top",wrap_text=True)
EUR='#,##0" €";(#,##0)" €";"-"'; PCT='0.0%;(0.0%);"-"'; NB='#,##0;(#,##0);"-"'; NB1='#,##0.0;(#,##0.0);"-"'; XCOEF='0.00'

def C(ws,ref,val=None,font=None,fill=None,fmt=None,align=None,border=False):
    c=ws[ref]
    if val is not None:c.value=val
    if font:c.font=font
    if fill:c.fill=fill
    if fmt:c.number_format=fmt
    if align:c.alignment=align
    if border:c.border=BORD
    return c
def band(ws,row,a,b,text,fill=FN_NAVY,font=CHDR,h=20):
    ws.merge_cells(f"{a}{row}:{b}{row}")
    for col in range(openpyxl.utils.column_index_from_string(a),openpyxl.utils.column_index_from_string(b)+1):
        ws.cell(row=row,column=col).fill=fill
    cc=ws[f"{a}{row}"]; cc.value=text; cc.font=font; cc.alignment=Alignment(horizontal="left",vertical="center")
    ws.row_dimensions[row].height=h

# ---------------------------------------------------------------- référentiel
# marque -> (domaine, coeff_vol, coeff_prix, base_entry, [villes])
BRANDS={
 "MBway":              ("Management",   1.20,1.00,60,["Paris","Lyon","Nantes","Bordeaux"]),
 "ISCOM":              ("Communication",1.00,1.20,55,["Paris","Lille","Toulouse"]),
 "Ipac Bachelor Factory":("Commerce",   1.30,0.80,50,["Nantes","Rennes","Montpellier"]),
 "Pigier":             ("Commerce/RH",  0.70,1.00,42,["Lyon","Bordeaux"]),
 "Tunon":              ("Tourisme",     1.00,1.10,36,["Paris","Lyon"]),
}
CITY={"Paris":1.30,"Lyon":1.10,"Nantes":1.00,"Bordeaux":0.85,"Lille":0.90,"Toulouse":0.85,"Rennes":0.80,"Montpellier":0.80}
# programmes par marque : (nom, type, [(niveau, modalité)])
PROGS={  # EDUSERVICES ~76% alternance : B1 souvent initial, montée en alternance ensuite
 "MBway":[("Bachelor Management","BAC",[("B1","INIT"),("B2","ALT"),("B3","ALT")]),
          ("Mastère Management","MAST",[("M1","ALT"),("M2","ALT")])],
 "ISCOM":[("Bachelor Communication","BAC",[("B1","INIT"),("B2","ALT"),("B3","ALT")]),
          ("Mastère Communication","MAST",[("M1","ALT"),("M2","ALT")])],
 "Ipac Bachelor Factory":[("Bachelor Commerce","BAC",[("B1","ALT"),("B2","ALT"),("B3","ALT")])],
 "Pigier":[("BTS Gestion","BTS",[("1","ALT"),("2","ALT")]),
           ("Bachelor RH","BAC",[("B1","ALT"),("B3","ALT")])],
 "Tunon":[("Bachelor Tourisme","BAC",[("B1","ALT"),("B2","ALT"),("B3","ALT")])],
}
# constantes calibrées (sourcées)
CAP,SEUIL,HRS,TXH,PEDA,ETPC,FRAIS,PROGREF,PSTRUCT = 30,15,550,60,400,48000,90,0.84,0.08
CONV_N1, ADM_N1, CAC_N1, DA_PCT = 0.372, 0.62, 950, 0.03
SIGN_N1, FALL = 0.82, 0.60   # taux de signature contrat N-1 ; revenu net d'un alternant non signé (% du NPEC)
def sf(mod, sign): return sign + (1-sign)*FALL if mod=="ALT" else 1.0  # facteur de revenu (financement)

def tarif(t,mod):
    if t=="BTS":  return 7000 if mod=="ALT" else 6500
    if t=="MAST": return 9000 if mod=="ALT" else 9500
    return 8000 if mod=="ALT" else 8500   # BAC

def lvl_factor(t,niv):
    return {"BAC":{"B1":1.00,"B2":0.85,"B3":0.75},
            "MAST":{"M1":0.60,"M2":0.54},
            "BTS":{"1":0.80,"2":0.70}}[t][niv]
def is_entry(t,niv): return niv in ("B1","M1","1")
def new_share(t,niv):
    if is_entry(t,niv): return 1.0
    return {"B2":0.10,"B3":0.08,"M2":0.05,"2":0.05}[niv]

rows=[]  # chaque cellule fine
for marque,(dom,cv,cp,base,villes) in BRANDS.items():
    for ville in villes:
        for pnom,ptype,niveaux in PROGS[marque]:
            entry=round(base*CITY[ville])
            for niv,mod in niveaux:
                eff=max(0,round(entry*lvl_factor(ptype,niv)))
                nouv=round(eff*new_share(ptype,niv))
                rein=eff-nouv
                cand=round(nouv/CONV_N1) if nouv>0 else 0
                admis=round(cand*ADM_N1)
                trf=tarif(ptype,mod)
                classes=max(1,math.ceil(eff/CAP)) if eff>0 else 0
                rows.append(dict(marque=marque,ville=ville,prog=pnom,type=ptype,niv=niv,mod=mod,
                    eff=eff,nouv=nouv,rein=rein,cand=cand,admis=admis,tarif=trf,classes=classes))
N=len(rows)
# agrégats campus (entité = marque + ville)
campus=[]
seen=set()
for r in rows:
    key=(r["marque"],r["ville"])
    if key in seen: continue
    seen.add(key)
    cells=[x for x in rows if (x["marque"],x["ville"])==key]
    eff=sum(x["eff"] for x in cells); nouv=sum(x["nouv"] for x in cells)
    ca=sum(x["eff"]*x["tarif"]*sf(x["mod"],SIGN_N1)+x["nouv"]*FRAIS for x in cells)
    ens=sum(x["classes"]*HRS*TXH for x in cells)
    ped=eff*PEDA; mkt=nouv*CAC_N1
    contrib=ca-ens-ped-mkt
    loyer=round(0.11*ca/1000)*1000
    etp=round(eff/25)+2; perm=etp*ETPC
    da=round(DA_PCT*ca/1000)*1000; m2=eff*8
    campus.append(dict(marque=r["marque"],ville=r["ville"],eff=eff,nouv=nouv,ca=ca,ens=ens,ped=ped,
        mkt=mkt,contrib=contrib,loyer=loyer,etp=etp,perm=perm,da=da,m2=m2))
CG=len(campus)
grp_ca_n1=sum(c["ca"] for c in campus)
grp_struct_n1=round(PSTRUCT*grp_ca_n1)

wb=openpyxl.Workbook()

# ================================================================ 00_Notice
ws=wb.active; ws.title="00_Notice"; ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":30,"C":66,"D":20}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:D2"); C(ws,"B2","EDUSERVICES GROUP — Modèle de pilotage budgétaire",CTIT,FN_NAVY,align=AL); ws.row_dimensions[2].height=30
ws.merge_cells("B3:D3"); C(ws,"B3","Budget par inducteurs à maille fine · funnel admissions · simulation · aide à la décision · pré-Tagetik",CIT)
band(ws,5,"B","D","1. La logique du modèle (tout est lié)")
txt=["Le budget se construit du bas vers le haut : chaque cellule fine (marque × campus × programme ×",
 "niveau × modalité) part du RÉALISÉ, à laquelle on applique la NOTE DE CADRAGE (des % de volume et de",
 "prix) modulée par des COEFFICIENTS STRATÉGIQUES par marque. Le funnel CRM (candidatures → inscrits)",
 "projette les nouveaux entrants ; le nombre de CLASSES se déduit de l'effectif ; les coûts en découlent.",
 "On obtient : P&L consolidé, pont Prix/Volume, décisions d'ouverture/fermeture de classes, et allocation",
 "des frais de structure par driver — le tout recalculé instantanément quand on change un levier."]
r=6
for t in txt: ws.merge_cells(f"B{r}:D{r}"); C(ws,f"B{r}",t,CREG,align=ALW); r+=1
band(ws,13,"B","D","2. Les feuilles")
sheets=[("01_Note_cadrage","Objectifs et hypothèses directrices (scénario central)"),
 ("02_Parametres","Sélecteur de scénario + leviers + constantes + driver d'allocation"),
 ("03_Coeff_Strateg","Coefficients stratégiques par marque (on challenge + ou –)"),
 ("04_Referentiel","Dimensions : entités, programmes, comptes (fixe/variable) — mapping Tagetik"),
 ("05_Historique","Réalisé fin par cellule : funnel CRM + effectifs + classes"),
 ("06_Structure","Réalisé par campus : loyers, ETP permanents, D&A, m² (compta/SIRH/immo)"),
 ("07_Moteur","Moteur de budget par cellule (100% formules)"),
 ("08_Allocation","Frais de structure groupe alloués aux campus par driver"),
 ("09_PnL","Compte de résultat consolidé N-1 vs Budget + pont Prix/Volume"),
 ("10_Decision","Ouvrir/fermer une classe, remplissage, point mort, scoring 🟢🟡🔴"),
 ("11_Simulation","Tableau de bord : KPIs, scénarios, levier conversion"),
 ("12_Mapping_Tagetik","Passerelle : dimensions, versions/snapshot, workflow bottom-up")]
r=14
for a,b in sheets:
    C(ws,f"B{r}",a,CB,FN_LIGHT,align=AL,border=True); ws.merge_cells(f"C{r}:D{r}"); C(ws,f"C{r}",b,CREG,align=ALW,border=True); r+=1
band(ws,r+1,"B","D","3. Légende & avertissement"); r+=2
for txt2,ft,fl in [("Saisie / donnée réelle",CIN,None),("Formule / calcul",CF,None),("Lien inter-feuilles",CL,None),("Hypothèse clé à remplir",CB,FN_YEL)]:
    C(ws,f"B{r}","  exemple  ",ft,fl,align=AC,border=True); ws.merge_cells(f"C{r}:D{r}"); C(ws,f"C{r}",txt2,CREG,align=AL,border=True); r+=1
ws.merge_cells(f"B{r+1}:D{r+3}")
C(ws,f"B{r+1}","Marques et campus RÉELS d'EDUSERVICES, mais effectifs, tarifs et montants ILLUSTRATIFS, "
 "calibrés sur des ordres de grandeur sectoriels sourcés (frais de scolarité 8–11 k€, NPEC alternance "
 "~7–10 k€, CAC ~0,8–2 k€, marge EBITDA ~15–22 %, taille de classe ~30). À remplacer par les données réelles.",CIT,align=ALW)

# ================================================================ 01_Note_cadrage
ws=wb.create_sheet("01_Note_cadrage"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":42,"C":16,"D":56}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:D2"); C(ws,"B2","Note de cadrage budgétaire — Budget N+1",CTIT,FN_NAVY,align=AL); ws.row_dimensions[2].height=28
ws.merge_cells("B3:D3"); C(ws,"B3","Valeurs du scénario CADRAGE (modifiables en 02_Parametres).",CIT)
band(ws,5,"B","D","Objectifs")
for i,t in enumerate(["Croître les effectifs par un meilleur recrutement (volume) ET une meilleure conversion des admissions,",
  "tout en revalorisant les tarifs de façon différenciée par marque, et en préservant la marge EBITDA.",
  "Optimiser le remplissage des classes (ni sous-remplies coûteuses, ni surchargées)."]):
    ws.merge_cells(f"B{6+i}:D{6+i}"); C(ws,f"B{6+i}",t,CREG,align=ALW)
band(ws,10,"B","D","Hypothèses directrices (cadrage)")
C(ws,"B11","Levier",CHDR,FN_BLUE,align=AL,border=True); C(ws,"C11","Valeur",CHDR,FN_BLUE,align=AC,border=True); C(ws,"D11","Commentaire",CHDR,FN_BLUE,align=AL,border=True)
hyp=[("Croissance recrutement (volume)","='02_Parametres'!D6",PCT,"Croissance des candidatures, avant coefficient stratégique par marque."),
 ("Hausse tarifaire (prix)","='02_Parametres'!D7",PCT,"Revalorisation moyenne, modulée par marque."),
 ("Gain de conversion admissions","='02_Parametres'!D8",PCT,"Points de conversion candidature→inscrit gagnés (pilotage admissions)."),
 ("Taux de signature contrat (alternance)","='02_Parametres'!D9",PCT,"Part des alternants avec contrat signé → financés OPCO/NPEC. Levier de REVENU."),
 ("Taux de réinscription","='02_Parametres'!D10",PCT,"Progression des étudiants d'une année sur l'autre."),
 ("Inflation des charges","='02_Parametres'!D11",PCT,"Loyers, pédagogie, autres charges."),
 ("Politique salariale","='02_Parametres'!D12",PCT,"Revalorisation des rémunérations chargées."),
 ("Coût d'acquisition (CAC)","='02_Parametres'!D13",EUR,"Dépense marketing par nouvel inscrit.")]
r=12
for lib,f,fmt,com in hyp:
    C(ws,f"B{r}",lib,CREG,align=AL,border=True); C(ws,f"C{r}",f,CL,fmt=fmt,align=AC,border=True); C(ws,f"D{r}",com,CIT,align=ALW,border=True); ws.row_dimensions[r].height=28; r+=1
band(ws,r+1,"B","D","Règles de gestion"); r+=2
for t in ["Nouveaux inscrits = candidatures × (1+croissance×coef marque) × (conversion N-1 + gain conversion).",
 "Réinscrits = réinscrits N-1 × (taux réinscription / référence).  Effectif = réinscrits + nouveaux.",
 "Tarif budget = tarif N-1 × (1 + hausse prix × coef marque).",
 "Nombre de classes = arrondi supérieur (effectif / capacité cible).  Coût enseignement = classes × heures × taux.",
 "Contribution = CA − enseignement − pédagogie variable − marketing.  EBITDA = Σ contributions − loyers − permanents − structure."]:
    ws.merge_cells(f"B{r}:D{r}"); C(ws,f"B{r}",t,CREG,align=ALW); r+=1

# ================================================================ 02_Parametres
ws=wb.create_sheet("02_Parametres"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":40,"C":15,"D":13,"E":13,"F":13,"G":14}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:G2"); C(ws,"B2","Paramètres & scénarios",CTIT,FN_NAVY,align=AL); ws.row_dimensions[2].height=28
C(ws,"B3","Scénario actif :",CB,align=AR); C(ws,"C3","Cadrage",CINB,FN_YEL,align=AC,border=True)
dv=DataValidation(type="list",formula1='"Cadrage,Optimiste,Prudent"',allow_blank=False); ws.add_data_validation(dv); dv.add(ws["C3"])
C(ws,"E3","◄ tout le modèle bascule automatiquement",CIT); ws.merge_cells("E3:G3")
C(ws,"B5","Levier",CHDR,FN_BLUE,align=AL,border=True); C(ws,"C5","Unité",CHDR,FN_BLUE,align=AC,border=True)
for col,n in (("D","Cadrage"),("E","Optimiste"),("F","Prudent")): C(ws,f"{col}5",n,CHDR,FN_BLUE,align=AC,border=True)
C(ws,"G5","ACTIF",CHDR,FN_NAVY,align=AC,border=True)
levs=[("Croissance recrutement (volume)","%",0.04,0.09,-0.02,PCT),
 ("Hausse tarifaire (prix)","%",0.03,0.04,0.02,PCT),
 ("Gain de conversion admissions","pts",0.015,0.04,0.0,PCT),
 ("Taux de signature contrat (alternance)","%",0.85,0.92,0.75,PCT),
 ("Taux de réinscription","%",0.84,0.87,0.80,PCT),
 ("Inflation des charges","%",0.02,0.015,0.03,PCT),
 ("Politique salariale","%",0.025,0.02,0.03,PCT),
 ("Coût d'acquisition (CAC)","€",1300,1150,1600,EUR)]
r=6
for lib,u,cad,opt,pru,fmt in levs:
    C(ws,f"B{r}",lib,CREG,align=AL,border=True); C(ws,f"C{r}",u,CREG,align=AC,border=True)
    C(ws,f"D{r}",cad,CIN,fmt=fmt,align=AC,border=True); C(ws,f"E{r}",opt,CIN,fmt=fmt,align=AC,border=True); C(ws,f"F{r}",pru,CIN,fmt=fmt,align=AC,border=True)
    C(ws,f"G{r}",f"=INDEX(D{r}:F{r},MATCH($C$3,$D$5:$F$5,0))",CF,FN_LIGHT,fmt=fmt,align=AC,border=True); r+=1
C(ws,"B15","Constantes (calibrées, sourcées)",CHDR,FN_BLUE,align=AL,border=True)
for col in ("C","D","E","F","G"): C(ws,f"{col}15"," ",fill=FN_BLUE,border=True)
consts=[("Capacité cible / classe","nb",CAP,NB),("Seuil d'ouverture / classe","nb",SEUIL,NB),
 ("Heures d'enseignement / classe / an","h",HRS,NB),("Taux horaire chargé enseignement","€",TXH,EUR),
 ("Coût pédagogique variable / étudiant","€",PEDA,EUR),("Coût chargé / ETP permanent","€",ETPC,EUR),
 ("Frais de dossier / nouvel inscrit","€",FRAIS,EUR),("Réf. taux de réinscription","%",PROGREF,PCT),
 ("Frais de structure groupe (% CA)","%",PSTRUCT,PCT),
 ("Revenu net d'un alternant NON signé (% NPEC)","%",FALL,PCT),
 ("Réf. taux de signature contrat N-1","%",SIGN_N1,PCT)]
r=16
for lib,u,val,fmt in consts:
    C(ws,f"B{r}",lib,CREG,align=AL,border=True); C(ws,f"C{r}",u,CREG,align=AC,border=True)
    C(ws,f"D{r}",val,CIN,fmt=fmt,align=AC,border=True)
    for col in ("E","F"): C(ws,f"{col}{r}"," ",border=True)
    C(ws,f"G{r}",f"=D{r}",CF,FN_LIGHT,fmt=fmt,align=AC,border=True); r+=1
C(ws,"B28","Driver d'allocation des frais de structure :",CB,align=AR); ws.merge_cells("B28:C28")
C(ws,"D28","Effectifs",CINB,FN_YEL,align=AC,border=True)
dv2=DataValidation(type="list",formula1='"Effectifs,Chiffre d\'affaires,Surface m2"',allow_blank=False); ws.add_data_validation(dv2); dv2.add(ws["D28"])
C(ws,"E28","◄ change la clé de répartition (feuille 08)",CIT); ws.merge_cells("E28:G28")

# ================================================================ 03_Coeff_Strateg
ws=wb.create_sheet("03_Coeff_Strateg"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":26,"C":18,"D":16,"E":16,"F":16,"G":18}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:G2"); C(ws,"B2","Coefficients stratégiques par marque",CTIT,FN_NAVY,align=AL); ws.row_dimensions[2].height=28
ws.merge_cells("B3:G3"); C(ws,"B3","On « serre la vis » plus ou moins selon la marque : % appliqué = % note de cadrage × coefficient.",CIT)
hd=["Marque","Intensité VOLUME","Intensité PRIX","→ Volume appliqué","→ Prix appliqué","Posture"]
for i,h in enumerate(hd): C(ws,f"{GL(2+i)}5",h,CHDR,FN_BLUE,align=AC,border=True)
CO0=6
r=CO0
for marque,(dom,cv,cp,base,villes) in BRANDS.items():
    C(ws,f"B{r}",marque,CREG,align=AL,border=True)
    C(ws,f"C{r}",cv,CIN,fmt=XCOEF,align=AC,border=True); C(ws,f"D{r}",cp,CIN,fmt=XCOEF,align=AC,border=True)
    C(ws,f"E{r}",f"='02_Parametres'!$G$6*C{r}",CF,fmt=PCT,align=AC,border=True)
    C(ws,f"F{r}",f"='02_Parametres'!$G$7*D{r}",CF,fmt=PCT,align=AC,border=True)
    C(ws,f"G{r}",f'=IF(C{r}>=1.15,"Pousser",IF(C{r}<=0.85,"Défendre","Maintenir"))',CF,align=AC,border=True); r+=1
CON=r-1
C(ws,f"B{r+1}","Coefficient 1,00 = on applique le cadrage tel quel · >1 = on pousse · <1 = on freine.",CIT); ws.merge_cells(f"B{r+1}:G{r+1}")

# refs paramètres / coeff
P=lambda a:f"'02_Parametres'!{a}"
PVOL,PPRIX,PDCONV,PSIGN,PPROG,PINFL,PSAL,PCAC = P("$G$6"),P("$G$7"),P("$G$8"),P("$G$9"),P("$G$10"),P("$G$11"),P("$G$12"),P("$G$13")
KCAP,KSEUIL,KHRS,KTXH,KPEDA,KETPC,KFRAIS,KPROGREF,KPSTRUCT,KFALL,KSIGNN1 = (P("$D$16"),P("$D$17"),P("$D$18"),P("$D$19"),
    P("$D$20"),P("$D$21"),P("$D$22"),P("$D$23"),P("$D$24"),P("$D$25"),P("$D$26"))
CVRANGE=f"'03_Coeff_Strateg'!$C${CO0}:$C${CON}"; CPRANGE=f"'03_Coeff_Strateg'!$D${CO0}:$D${CON}"; CMRANGE=f"'03_Coeff_Strateg'!$B${CO0}:$B${CON}"

# ================================================================ 04_Referentiel
ws=wb.create_sheet("04_Referentiel"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":24,"C":16,"D":22,"E":12,"F":10,"G":14}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:G2"); C(ws,"B2","Référentiel — dimensions (mapping Tagetik)",CTIT,FN_NAVY,align=AL); ws.row_dimensions[2].height=28
band(ws,4,"B","G","Dimension ENTITÉ : Groupe → Marque → Campus")
for i,h in enumerate(["Marque","Ville","Domaine","Niveau","Devise","# programmes"]): C(ws,f"{GL(2+i)}5",h,CHDR,FN_BLUE,align=AC,border=True)
r=6
C(ws,f"B{r}","EDUSERVICES GROUP",CINB,FN_TOT,align=AL,border=True); C(ws,f"C{r}","—",CREG,FN_TOT,align=AC,border=True)
C(ws,f"D{r}","Tous",CREG,FN_TOT,align=AC,border=True); C(ws,f"E{r}","Groupe",CB,FN_TOT,align=AC,border=True)
C(ws,f"F{r}","EUR",CREG,FN_TOT,align=AC,border=True); C(ws,f"G{r}"," ",fill=FN_TOT,border=True); r+=1
for marque,(dom,cv,cp,base,villes) in BRANDS.items():
    for ville in villes:
        C(ws,f"B{r}",marque,CIN,align=AL,border=True); C(ws,f"C{r}",ville,CIN,align=AC,border=True)
        C(ws,f"D{r}",dom,CREG,align=AC,border=True); C(ws,f"E{r}","Campus",CREG,align=AC,border=True)
        C(ws,f"F{r}","EUR",CREG,align=AC,border=True); C(ws,f"G{r}",len(PROGS[marque]),CREG,align=AC,border=True); r+=1
band(ws,r+1,"B","G","Dimension COMPTE — nature fixe / variable"); r+=2
for i,h in enumerate(["Compte","Libellé","Rubrique","Nature","Inducteur"]): C(ws,f"{GL(2+i)}{r}",h,CHDR,FN_BLUE,align=AC,border=True)
r+=1
comptes=[("70600","Scolarité","Chiffre d'affaires","—","Effectif × tarif"),
 ("70800","Frais de dossier","Chiffre d'affaires","—","Nouveaux inscrits"),
 ("64100","Enseignement (vacation)","Coût direct","Semi-fixe / classe","Nb classes × heures"),
 ("60700","Pédagogie & supports","Coût direct","Variable","Effectif"),
 ("62300","Marketing & acquisition","Coût direct","Variable","Nouveaux inscrits"),
 ("61300","Loyers & immobilier","Structure","Fixe","Campus / m²"),
 ("64000","Personnel permanent","Structure","Fixe","ETP par campus"),
 ("65000","Frais de structure groupe","Structure","Fixe (alloué)","Driver (effectif/CA/m²)"),
 ("68000","Dotations amortissements","D&A","Fixe","Campus")]
for cpt,lib,rub,nat,ind in comptes:
    C(ws,f"B{r}",cpt,CIN,align=AL,border=True); C(ws,f"C{r}",lib,CREG,align=AL,border=True)
    C(ws,f"D{r}",rub,CREG,align=AC,border=True); C(ws,f"E{r}",nat,CREG,align=AC,border=True); C(ws,f"F{r}",ind,CREG,align=AL,border=True); r+=1

# ================================================================ 05_Historique (cellule)
ws=wb.create_sheet("05_Historique"); ws.sheet_view.showGridLines=False
H_cols=["Marque","Ville","Programme","Type","Niveau","Modalité","Candidatures N-1","Admis N-1",
 "Nouveaux N-1","Réinscrits N-1","Effectif N-1","Tarif N-1","Classes N-1","Conv. N-1"]
Hw=[16,11,20,7,8,9,11,9,10,11,10,10,9,9]
for i,w in enumerate(Hw): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells("A1:N1"); C(ws,"A1","Réalisé N-1 par cellule fine — funnel CRM + scolarité",CTIT,FN_NAVY,align=AL); ws.row_dimensions[1].height=24
ws.merge_cells("A2:N2"); C(ws,"A2","Sources : CRM (candidatures/admis) + SI Scolarité (effectifs/classes) + tarifs. Bleu = à remplacer par le réel.",CIT)
for i,h in enumerate(H_cols): C(ws,f"{GL(1+i)}3",h,CHDR,FN_BLUE,align=AC,border=True)
ws.row_dimensions[3].height=30
HR0=4
for idx,rr in enumerate(rows):
    r=HR0+idx
    vals=[(rr["marque"],AL,None),(rr["ville"],AC,None),(rr["prog"],AL,None),(rr["type"],AC,None),(rr["niv"],AC,None),(rr["mod"],AC,None),
     (rr["cand"],AC,NB),(rr["admis"],AC,NB),(rr["nouv"],AC,NB),(rr["rein"],AC,NB),(rr["eff"],AC,NB),(rr["tarif"],AC,EUR),(rr["classes"],AC,NB)]
    for i,(v,al,fmt) in enumerate(vals): C(ws,f"{GL(1+i)}{r}",v,CIN,fmt=fmt,align=al,border=True)
    C(ws,f"N{r}",f"=IFERROR(I{r}/G{r},0)",CF,fmt=PCT,align=AC,border=True)  # conversion N-1
HRN=HR0+N-1
ws.freeze_panes="G4"

# ================================================================ 06_Structure (campus)
ws=wb.create_sheet("06_Structure"); ws.sheet_view.showGridLines=False
S_cols=["Marque","Ville","Effectif N-1","CA N-1","Contribution N-1","Loyer N-1","ETP perm.","Masse perm. N-1","D&A N-1","Surface m²","EBITDA campus N-1"]
Sw=[16,11,11,13,14,12,10,14,11,10,15]
for i,w in enumerate(Sw): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells("A1:K1"); C(ws,"A1","Réalisé N-1 par campus — structure (compta / SIRH / immobilier)",CTIT,FN_NAVY,align=AL); ws.row_dimensions[1].height=24
ws.merge_cells("A2:K2"); C(ws,"A2","Contribution = agrégée depuis 05/coûts directs. Loyer, ETP, D&A, m² = saisies structurelles (bleu).",CIT)
for i,h in enumerate(S_cols): C(ws,f"{GL(1+i)}3",h,CHDR,FN_BLUE,align=AC,border=True)
ws.row_dimensions[3].height=30
SR0=4
for idx,cc in enumerate(campus):
    r=SR0+idx
    C(ws,f"A{r}",cc["marque"],CIN,align=AL,border=True); C(ws,f"B{r}",cc["ville"],CIN,align=AC,border=True)
    C(ws,f"C{r}",cc["eff"],CIN,fmt=NB,align=AC,border=True); C(ws,f"D{r}",cc["ca"],CIN,fmt=EUR,align=AR,border=True)
    C(ws,f"E{r}",round(cc["contrib"]),CIN,fmt=EUR,align=AR,border=True)
    C(ws,f"F{r}",cc["loyer"],CIN,fmt=EUR,align=AR,border=True); C(ws,f"G{r}",cc["etp"],CIN,fmt=NB,align=AC,border=True)
    C(ws,f"H{r}",f"=G{r}*{KETPC}",CF,fmt=EUR,align=AR,border=True)
    C(ws,f"I{r}",cc["da"],CIN,fmt=EUR,align=AR,border=True); C(ws,f"J{r}",cc["m2"],CIN,fmt=NB,align=AC,border=True)
    C(ws,f"K{r}",f"=E{r}-F{r}-H{r}",CF,fmt=EUR,align=AR,border=True)
SRN=SR0+CG-1
r=SR0+CG
C(ws,f"A{r}","TOTAL GROUPE",CFB,FN_TOT,align=AL,border=True); C(ws,f"B{r}"," ",fill=FN_TOT,border=True)
for col in ["C","D","E","F","G","H","I","J","K"]:
    C(ws,f"{col}{r}",f"=SUM({col}{SR0}:{col}{SRN})",CFB,FN_TOT,fmt=(NB if col in("C","G","J") else EUR),align=(AC if col in("C","G","J") else AR),border=True)
STOT=r

# ================================================================ 07_Moteur
ws=wb.create_sheet("07_Moteur"); ws.sheet_view.showGridLines=False
M_cols=["Marque","Ville","Programme","Niveau","Mod.","Eff. N-1","Nouv N-1","Réins N-1","Cand N-1","Tarif N-1","Cl. N-1",
 "coef Vol","coef Prix","Cand Bud","Conv Bud","Nouv Bud","Réins Bud","Effectif Bud","Tarif Bud","CA Bud",
 "Cl. besoin","Enseign.","Pédago","Marketing","Contribution","Rempl.","Contr/étu","Sig N-1","Sig Bud"]
Mw=[15,10,18,7,6]+[9]*(len(M_cols)-5)
for i,w in enumerate(Mw): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells("A1:AC1"); C(ws,"A1","Moteur de budget par cellule fine (100 % formules — réagit aux leviers & coefficients)",CTIT,FN_NAVY,align=AL); ws.row_dimensions[1].height=24
for i,h in enumerate(M_cols): C(ws,f"{GL(1+i)}2",h,CHDR,FN_BLUE,align=AC,border=True)
ws.row_dimensions[2].height=30
MR0=3
def hc(col,r): return f"'05_Historique'!{col}{r}"
for idx in range(N):
    r=MR0+idx; hr=HR0+idx
    C(ws,f"A{r}",f"={hc('A',hr)}",CL,align=AL,border=True); C(ws,f"B{r}",f"={hc('B',hr)}",CL,align=AC,border=True)
    C(ws,f"C{r}",f"={hc('C',hr)}",CL,align=AL,border=True); C(ws,f"D{r}",f"={hc('E',hr)}",CL,align=AC,border=True); C(ws,f"E{r}",f"={hc('F',hr)}",CL,align=AC,border=True)
    C(ws,f"F{r}",f"={hc('K',hr)}",CL,fmt=NB,align=AC,border=True)   # eff N-1
    C(ws,f"G{r}",f"={hc('I',hr)}",CL,fmt=NB,align=AC,border=True)   # nouv N-1
    C(ws,f"H{r}",f"={hc('J',hr)}",CL,fmt=NB,align=AC,border=True)   # reins N-1
    C(ws,f"I{r}",f"={hc('G',hr)}",CL,fmt=NB,align=AC,border=True)   # cand N-1
    C(ws,f"J{r}",f"={hc('L',hr)}",CL,fmt=EUR,align=AC,border=True)  # tarif N-1
    C(ws,f"K{r}",f"={hc('M',hr)}",CL,fmt=NB,align=AC,border=True)   # classes N-1
    C(ws,f"L{r}",f"=INDEX({CVRANGE},MATCH(A{r},{CMRANGE},0))",CF,fmt=XCOEF,align=AC,border=True)
    C(ws,f"M{r}",f"=INDEX({CPRANGE},MATCH(A{r},{CMRANGE},0))",CF,fmt=XCOEF,align=AC,border=True)
    C(ws,f"N{r}",f"=I{r}*(1+{PVOL}*L{r})",CF,fmt=NB,align=AC,border=True)                       # cand bud
    C(ws,f"O{r}",f"=IFERROR(G{r}/I{r},0)+{PDCONV}",CF,fmt=PCT,align=AC,border=True)              # conv bud
    C(ws,f"P{r}",f"=N{r}*O{r}",CF,fmt=NB,align=AC,border=True)                                   # nouv bud
    C(ws,f"Q{r}",f"=H{r}*({PPROG}/{KPROGREF})",CF,fmt=NB,align=AC,border=True)                   # reins bud
    C(ws,f"R{r}",f"=P{r}+Q{r}",CFB,fmt=NB,align=AC,border=True)                                  # effectif bud
    C(ws,f"S{r}",f"=J{r}*(1+{PPRIX}*M{r})",CF,fmt=EUR,align=AC,border=True)                      # tarif nominal bud
    C(ws,f"AB{r}",f'=IF(E{r}="ALT",{KSIGNN1}+(1-{KSIGNN1})*{KFALL},1)',CF,fmt=XCOEF,align=AC,border=True)  # facteur sig N-1
    C(ws,f"AC{r}",f'=IF(E{r}="ALT",{PSIGN}+(1-{PSIGN})*{KFALL},1)',CF,fmt=XCOEF,align=AC,border=True)      # facteur sig bud
    C(ws,f"T{r}",f"=R{r}*S{r}*AC{r}+P{r}*{KFRAIS}",CF,fmt=EUR,align=AR,border=True)              # CA bud (financement inclus)
    C(ws,f"U{r}",f"=IF(R{r}<=0,0,MAX(1,ROUNDUP(R{r}/{KCAP},0)))",CF,fmt=NB,align=AC,border=True) # classes besoin
    C(ws,f"V{r}",f"=U{r}*{KHRS}*{KTXH}*(1+{PSAL})",CF,fmt=EUR,align=AR,border=True)              # enseignement
    C(ws,f"W{r}",f"=R{r}*{KPEDA}*(1+{PINFL})",CF,fmt=EUR,align=AR,border=True)                   # pédago
    C(ws,f"X{r}",f"=P{r}*{PCAC}",CF,fmt=EUR,align=AR,border=True)                                # marketing
    C(ws,f"Y{r}",f"=T{r}-V{r}-W{r}-X{r}",CFB,fmt=EUR,align=AR,border=True)                       # contribution
    C(ws,f"Z{r}",f"=IFERROR(R{r}/(U{r}*{KCAP}),0)",CF,fmt=PCT,align=AC,border=True)              # remplissage
    C(ws,f"AA{r}",f"=IFERROR(Y{r}/R{r},0)",CF,fmt=EUR,align=AC,border=True)                      # contrib/étu
MRN=MR0+N-1
r=MR0+N
C(ws,f"A{r}","TOTAL",CFB,FN_TOT,align=AL,border=True)
for col in ["B","C","D","E"]: C(ws,f"{col}{r}"," ",fill=FN_TOT,border=True)
for col,fmt in [("F",NB),("G",NB),("H",NB),("P",NB),("Q",NB),("R",NB),("T",EUR),("U",NB),("V",EUR),("W",EUR),("X",EUR),("Y",EUR)]:
    C(ws,f"{col}{r}",f"=SUM({col}{MR0}:{col}{MRN})",CFB,FN_TOT,fmt=fmt,align=(AC if fmt==NB else AR),border=True)
for col in ["I","J","K","L","M","N","O","S","Z","AA","AB","AC"]: C(ws,f"{col}{r}"," ",fill=FN_TOT,border=True)
MTOT=r
ws.freeze_panes="F3"

# refs moteur
MO=lambda col:f"'07_Moteur'!{col}"
def mrng(col): return f"'07_Moteur'!${col}${MR0}:${col}${MRN}"
def msum(col): return f"SUM({mrng(col)})"

# ================================================================ 08_Allocation
ws=wb.create_sheet("08_Allocation"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":16,"C":11,"D":13,"E":11,"F":13,"G":12,"H":13,"I":11,"J":13,"K":11,"L":13}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:L2"); C(ws,"B2","Allocation des frais de structure groupe par driver",CTIT,FN_NAVY,align=AL); ws.row_dimensions[2].height=26
C(ws,"B3","Driver actif :",CB,align=AR); C(ws,"C3","='02_Parametres'!D28",CL,FN_YEL,align=AC,border=True)
C(ws,"E3","Frais de structure groupe à allouer :",CB,align=AR); ws.merge_cells("E3:H3")
C(ws,"I3",f"={KPSTRUCT}*{msum('T')}",CFB,fmt=EUR,align=AR,border=True); ws.merge_cells("I3:J3")
hd=["Marque","Ville","Contribution Bud","Loyer Bud","Masse perm. Bud","Valeur driver","Part driver","Frais alloués","EBITDA campus","D&A","EBIT campus"]
for i,h in enumerate(hd): C(ws,f"{GL(2+i)}5",h,CHDR,FN_BLUE,align=AC,border=True)
ws.row_dimensions[5].height=30
AR0=6
GROUPCOST="$I$3"
for idx,cc in enumerate(campus):
    r=AR0+idx; sr=SR0+idx
    vl=cc["ville"]
    crit=f'{mrng("A")},"{cc["marque"]}",{mrng("B")},"{vl}"'
    C(ws,f"B{r}",cc["marque"],CL,align=AL,border=True); C(ws,f"C{r}",vl,CL,align=AC,border=True)
    C(ws,f"D{r}",f"=SUMIFS({mrng('Y')},{crit})",CF,fmt=EUR,align=AR,border=True)
    C(ws,f"E{r}",f"='06_Structure'!F{sr}*(1+{PINFL})",CL,fmt=EUR,align=AR,border=True)
    C(ws,f"F{r}",f"='06_Structure'!G{sr}*{KETPC}*(1+{PSAL})",CL,fmt=EUR,align=AR,border=True)
    # valeur driver
    effb=f"SUMIFS({mrng('R')},{crit})"
    cab =f"SUMIFS({mrng('T')},{crit})"
    m2 =f"'06_Structure'!J{sr}"
    C(ws,f"G{r}",f"=IF($C$3=\"Effectifs\",{effb},IF($C$3=\"Chiffre d'affaires\",{cab},{m2}))",CF,fmt=NB,align=AR,border=True)
    C(ws,f"H{r}",f"=IFERROR(G{r}/SUM($G${AR0}:$G${AR0+CG-1}),0)",CF,fmt=PCT,align=AC,border=True)
    C(ws,f"I{r}",f"=H{r}*{GROUPCOST}",CF,fmt=EUR,align=AR,border=True)
    C(ws,f"J{r}",f"=D{r}-E{r}-F{r}-I{r}",CFB,fmt=EUR,align=AR,border=True)
    C(ws,f"K{r}",f"='06_Structure'!I{sr}",CL,fmt=EUR,align=AR,border=True)
    C(ws,f"L{r}",f"=J{r}-K{r}",CF,fmt=EUR,align=AR,border=True)
ARN=AR0+CG-1
r=AR0+CG
C(ws,f"B{r}","TOTAL GROUPE",CFB,FN_TOT,align=AL,border=True); C(ws,f"C{r}"," ",fill=FN_TOT,border=True)
for col in ["D","E","F","I","J","K","L"]:
    C(ws,f"{col}{r}",f"=SUM({col}{AR0}:{col}{ARN})",CFB,FN_TOT,fmt=EUR,align=AR,border=True)
C(ws,f"G{r}",f"=SUM(G{AR0}:G{ARN})",CFB,FN_TOT,fmt=NB,align=AR,border=True); C(ws,f"H{r}",f"=SUM(H{AR0}:H{ARN})",CFB,FN_TOT,fmt=PCT,align=AC,border=True)
ATOT=r
AL_=lambda col:f"'08_Allocation'!{col}"

# ================================================================ 09_PnL
ws=wb.create_sheet("09_PnL"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":34,"C":15,"D":15,"E":14,"F":11}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:F2"); C(ws,"B2","Compte de résultat consolidé — Budget N+1 vs Réalisé N-1",CTIT,FN_NAVY,align=AL); ws.row_dimensions[2].height=26
C(ws,"B3","Scénario :",CB,align=AR); C(ws,"C3","='02_Parametres'!C3",CL,align=AC)
for i,h in enumerate(["Rubrique","Réalisé N-1","Budget N+1","Écart €","Écart %"]): C(ws,f"{GL(2+i)}5",h,CHDR,FN_BLUE,align=AC,border=True)
ST=lambda col:f"'06_Structure'!{col}{STOT}"
# N-1 group aggregates
n1_eff=ST("C"); n1_ca=ST("D"); n1_contrib=ST("E"); n1_loyer=ST("F"); n1_perm=ST("H"); n1_da=ST("I")
n1_struct=f"{KPSTRUCT}*{n1_ca}"
lines=[("Effectifs",n1_eff,f"={msum('R')}",NB,False),
 ("Chiffre d'affaires",f"={n1_ca}",f"={msum('T')}",EUR,True),
 ("  Enseignement",f"=-'06_Structure'!{STOT}",None,EUR,False)]  # placeholder replaced below
# build lines explicitly
def put_line(r,lib,n1f,budf,fmt,bold,pct_row=False):
    ft=CFB if bold else CREG; fl=FN_LIGHT if bold else None
    C(ws,f"B{r}",lib,ft,fl,align=AL,border=True)
    C(ws,f"C{r}",n1f,(CFB if bold else CL),fl,fmt=fmt,align=AR,border=True)
    C(ws,f"D{r}",budf,(CFB if bold else CF),fl,fmt=fmt,align=AR,border=True)
    if pct_row:
        C(ws,f"E{r}",f"=D{r}-C{r}",ft,fl,fmt=PCT,align=AR,border=True); C(ws,f"F{r}"," ",fill=fl,border=True)
    else:
        C(ws,f"E{r}",f"=D{r}-C{r}",ft,fl,fmt=fmt,align=AR,border=True)
        C(ws,f"F{r}",f"=IFERROR(D{r}/C{r}-1,0)",ft,fl,fmt=PCT,align=AR,border=True)
# N-1 cost components (recomputed group from Structure not available per-line → use Moteur-consistent N-1 via 06 aggregate)
# For N-1 detail we approximate enseignement/pédago/marketing from Structure contribution identity:
# contribution_N1 = CA_N1 - (enseignement+pédago+marketing)_N1 ; on affiche le bloc agrégé "coûts directs".
r=6
put_line(r,"Effectifs",f"={n1_eff}",f"={msum('R')}",NB,False); r+=1
put_line(r,"Chiffre d'affaires",f"={n1_ca}",f"={msum('T')}",EUR,True); r+=1
put_line(r,"  Coûts directs (enseign.+pédago+market.)",f"=-({n1_ca}-{n1_contrib})",f"=-({msum('V')}+{msum('W')}+{msum('X')})",EUR,False); r+=1
put_line(r,"Marge de contribution",f"={n1_contrib}",f"={msum('Y')}",EUR,True); r+=1
put_line(r,"  Loyers",f"=-{n1_loyer}",f"=-{AL_('$E$'+str(ATOT))}",EUR,False); r+=1
put_line(r,"  Personnel permanent",f"=-{n1_perm}",f"=-{AL_('$F$'+str(ATOT))}",EUR,False); r+=1
put_line(r,"  Frais de structure groupe",f"=-{n1_struct}",f"=-{AL_('$I$'+str(ATOT))}",EUR,False); r+=1
put_line(r,"EBITDA",f"={n1_contrib}-{n1_loyer}-{n1_perm}-{n1_struct}",f"={AL_('$J$'+str(ATOT))}",EUR,True); r+=1
put_line(r,"  Marge EBITDA %",f"=IFERROR(({n1_contrib}-{n1_loyer}-{n1_perm}-{n1_struct})/{n1_ca},0)",
         f"=IFERROR({AL_('$J$'+str(ATOT))}/{msum('T')},0)",PCT,False,pct_row=True); r+=1
put_line(r,"  D&A",f"=-{n1_da}",f"=-{AL_('$K$'+str(ATOT))}",EUR,False); r+=1
put_line(r,"EBIT",f"={n1_contrib}-{n1_loyer}-{n1_perm}-{n1_struct}-{n1_da}",f"={AL_('$L$'+str(ATOT))}",EUR,True); r+=1
# pont Prix/Volume
band(ws,r+1,"B","F","Pont Chiffre d'affaires : N-1 → Budget"); r+=2
for i,h in enumerate(["Effet","Montant"]): C(ws,f"{GL(2+i)}{r}",h,CHDR,FN_BLUE,align=AC,border=True)
r+=1
# pont à 4 effets : Volume / Tarif / Signature (financement) / Frais — réconcilie exactement
volf=f"=SUMPRODUCT(({mrng('R')}-{mrng('F')})*{mrng('J')}*{mrng('AB')})"
tarf=f"=SUMPRODUCT({mrng('R')}*({mrng('S')}-{mrng('J')})*{mrng('AB')})"
sigf=f"=SUMPRODUCT({mrng('R')}*{mrng('S')}*({mrng('AC')}-{mrng('AB')}))"
fraisf=f"=SUMPRODUCT({mrng('P')})*{KFRAIS}-SUMPRODUCT({mrng('G')})*{KFRAIS}"
C(ws,f"B{r}","CA Réalisé N-1",CREG,align=AL,border=True); C(ws,f"C{r}",f"={n1_ca}",CL,fmt=EUR,align=AR,border=True); r+=1
C(ws,f"B{r}","  + Effet Volume (Δ effectif)",CREG,align=AL,border=True); C(ws,f"C{r}",volf,CF,fmt=EUR,align=AR,border=True); r+=1
C(ws,f"B{r}","  + Effet Tarif (hausse prix)",CREG,align=AL,border=True); C(ws,f"C{r}",tarf,CF,fmt=EUR,align=AR,border=True); r+=1
C(ws,f"B{r}","  + Effet Signature (financement OPCO)",CREG,align=AL,border=True); C(ws,f"C{r}",sigf,CF,fmt=EUR,align=AR,border=True); r+=1
C(ws,f"B{r}","  + Effet Frais de dossier",CREG,align=AL,border=True); C(ws,f"C{r}",fraisf,CF,fmt=EUR,align=AR,border=True); r+=1
C(ws,f"B{r}","CA Budget N+1",CFB,FN_TOT,align=AL,border=True); C(ws,f"C{r}",f"={msum('T')}",CFB,FN_TOT,fmt=EUR,align=AR,border=True); r+=1

# ================================================================ 10_Decision
ws=wb.create_sheet("10_Decision"); ws.sheet_view.showGridLines=False
D_cols=["Marque","Ville","Programme","Niveau","Effectif Bud","Cl. actuel","Cl. besoin","Écart",
 "Action","Impact EBITDA €","Remplissage","Contr/étu","Point mort (étu)","Marge sécurité","Statut"]
Dw=[15,10,18,7,10,9,9,7,14,13,10,9,12,12,16]
for i,w in enumerate(Dw): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells("A1:O1"); C(ws,"A1","Aide à la décision — ouverture / fermeture de classes",CTIT,FN_NAVY,align=AL); ws.row_dimensions[1].height=24
ws.merge_cells("A2:O2"); C(ws,"A2","« Fermer & redistribuer » = regrouper si l'effectif tient dans moins de classes (CA conservé). Impact>0 = gain.",CIT)
for i,h in enumerate(D_cols): C(ws,f"{GL(1+i)}3",h,CHDR,FN_BLUE,align=AC,border=True)
ws.row_dimensions[3].height=30
DR0=4
CLASSCOST=f"{KHRS}*{KTXH}*(1+{PSAL})"
for idx in range(N):
    r=DR0+idx; mr=MR0+idx
    C(ws,f"A{r}",f"={MO('A')+str(mr)}",CL,align=AL,border=True); C(ws,f"B{r}",f"={MO('B')+str(mr)}",CL,align=AC,border=True)
    C(ws,f"C{r}",f"={MO('C')+str(mr)}",CL,align=AL,border=True); C(ws,f"D{r}",f"={MO('D')+str(mr)}",CL,align=AC,border=True)
    C(ws,f"E{r}",f"={MO('R')+str(mr)}",CL,fmt=NB,align=AC,border=True)      # effectif bud
    C(ws,f"F{r}",f"={MO('K')+str(mr)}",CL,fmt=NB,align=AC,border=True)      # classes actuel (N-1)
    C(ws,f"G{r}",f"={MO('U')+str(mr)}",CL,fmt=NB,align=AC,border=True)      # classes besoin
    C(ws,f"H{r}",f"=G{r}-F{r}",CF,fmt=NB,align=AC,border=True)             # écart
    C(ws,f"I{r}",f'=IF(H{r}>0,"OUVRIR "&H{r},IF(H{r}<0,"FERMER "&ABS(H{r}),"RAS"))',CF,align=AC,border=True)
    C(ws,f"J{r}",f"=(F{r}-G{r})*{CLASSCOST}",CF,fmt=EUR,align=AR,border=True)  # impact (fermer=+)
    C(ws,f"K{r}",f"={MO('Z')+str(mr)}",CL,fmt=PCT,align=AC,border=True)     # remplissage
    C(ws,f"L{r}",f"={MO('AA')+str(mr)}",CL,fmt=EUR,align=AC,border=True)    # contrib/étu
    # point mort = coût classes / (tarif - pédago/étu)
    C(ws,f"M{r}",f"=IFERROR(G{r}*{CLASSCOST}/({MO('S')+str(mr)}*{MO('AC')+str(mr)}-{KPEDA}*(1+{PINFL})),0)",CF,fmt=NB,align=AC,border=True)
    C(ws,f"N{r}",f"=E{r}-M{r}",CF,fmt=NB,align=AC,border=True)
    C(ws,f"O{r}",f'=IF({MO("Y")+str(mr)}<0,"🔴 Restructurer",IF(K{r}<0.6,"🟡 Surveiller",IF(K{r}>=0.85,"🟢 Développer","🟢 Maintenir")))',CF,align=AL,border=True)
DRN=DR0+N-1
r=DR0+N
C(ws,f"A{r}","TOTAL",CFB,FN_TOT,align=AL,border=True)
for col in ["B","C","D"]: C(ws,f"{col}{r}"," ",fill=FN_TOT,border=True)
for col in ["E","F","G"]: C(ws,f"{col}{r}",f"=SUM({col}{DR0}:{col}{DRN})",CFB,FN_TOT,fmt=NB,align=AC,border=True)
C(ws,f"H{r}",f"=SUM(H{DR0}:H{DRN})",CFB,FN_TOT,fmt=NB,align=AC,border=True)
C(ws,f"I{r}"," ",fill=FN_TOT,border=True)
C(ws,f"J{r}",f"=SUM(J{DR0}:J{DRN})",CFB,FN_TOT,fmt=EUR,align=AR,border=True)
for col in ["K","L","M","N","O"]: C(ws,f"{col}{r}"," ",fill=FN_TOT,border=True)
ws.freeze_panes="E4"

# ================================================================ 11_Simulation
ws=wb.create_sheet("11_Simulation"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":30,"C":16,"D":16,"E":14,"F":14,"G":14}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:G2"); C(ws,"B2","Tableau de bord — simulation",CTIT,FN_NAVY,align=AL); ws.row_dimensions[2].height=26
ws.merge_cells("B3:G3"); C(ws,"B3","Change le scénario (02_Parametres!C3) ou un levier : tous les indicateurs se recalculent.",CIT)
C(ws,"B4","Scénario actif :",CB,align=AR); C(ws,"C4","='02_Parametres'!C3",CL,FN_YEL,align=AC,border=True)
band(ws,6,"B","E","Indicateurs clés — Budget N+1 vs Réalisé N-1")
C(ws,"B7","Indicateur",CHDR,FN_BLUE,align=AL,border=True); C(ws,"C7","Budget",CHDR,FN_BLUE,align=AC,border=True)
C(ws,"D7","N-1",CHDR,FN_BLUE,align=AC,border=True); C(ws,"E7","Évolution",CHDR,FN_BLUE,align=AC,border=True)
EBIT_B=AL_('$J$'+str(ATOT)); ebitda_n1=f"{n1_contrib}-{n1_loyer}-{n1_perm}-{n1_struct}"
kpis=[("Effectif total",f"={msum('R')}",f"={n1_eff}",NB),
 ("Chiffre d'affaires",f"={msum('T')}",f"={n1_ca}",EUR),
 ("Marge de contribution",f"={msum('Y')}",f"={n1_contrib}",EUR),
 ("EBITDA",f"={EBIT_B}",f"={ebitda_n1}",EUR),
 ("Marge EBITDA %",f"=IFERROR({EBIT_B}/{msum('T')},0)",f"=IFERROR(({ebitda_n1})/{n1_ca},0)",PCT),
 ("Nombre de classes",f"={msum('U')}",f"={msum('K')}",NB),
 ("Nouveaux inscrits",f"={msum('P')}",f"={msum('G')}",NB),
 ("Taux d'alternance",f"=IFERROR(SUMIF({mrng('E')},\"ALT\",{mrng('R')})/{msum('R')},0)",f"=IFERROR(SUMIF({mrng('E')},\"ALT\",{mrng('F')})/{msum('F')},0)",PCT),
 ("Taux de signature contrat",f"={PSIGN}",f"={KSIGNN1}",PCT)]
r=8
for lib,bud,n1,fmt in kpis:
    C(ws,f"B{r}",lib,CB,align=AL,border=True); C(ws,f"C{r}",bud,CF,fmt=fmt,align=AC,border=True)
    C(ws,f"D{r}",n1,CF,fmt=fmt,align=AC,border=True)
    C(ws,f"E{r}",(f"=D{r}-C{r}" if fmt==PCT else f"=IFERROR(C{r}/D{r}-1,0)"),CF,fmt=PCT,align=AC,border=True); r+=1
band(ws,r+1,"B","G","Comparatif des scénarios (leviers)"); r+=2
C(ws,f"B{r}","Levier",CHDR,FN_BLUE,align=AL,border=True)
for col,n in (("C","Cadrage"),("D","Optimiste"),("E","Prudent"),("F","Actif")): C(ws,f"{col}{r}",n,CHDR,(FN_NAVY if col=="F" else FN_BLUE),align=AC,border=True)
r+=1
for lib,pr,fmt in [("Croissance recrutement",6,PCT),("Hausse tarifaire",7,PCT),("Gain conversion",8,PCT),
 ("Taux signature contrat",9,PCT),("Taux réinscription",10,PCT),("Inflation charges",11,PCT),("Politique salariale",12,PCT),("CAC",13,EUR)]:
    C(ws,f"B{r}",lib,CREG,align=AL,border=True)
    C(ws,f"C{r}",f"='02_Parametres'!D{pr}",CL,fmt=fmt,align=AC,border=True); C(ws,f"D{r}",f"='02_Parametres'!E{pr}",CL,fmt=fmt,align=AC,border=True)
    C(ws,f"E{r}",f"='02_Parametres'!F{pr}",CL,fmt=fmt,align=AC,border=True); C(ws,f"F{r}",f"='02_Parametres'!G{pr}",CFB,FN_LIGHT,fmt=fmt,align=AC,border=True); r+=1
ws.merge_cells(f"B{r+1}:G{r+1}")
C(ws,f"B{r+1}","Astuce démo : un gain de conversion admissions (levier « Gain conversion ») augmente les inscrits SANS "
 "hausse tarifaire ni dépense volume — souvent le levier le plus rentable.",CIT,align=ALW)

# ================================================================ 12_Mapping_Tagetik
ws=wb.create_sheet("12_Mapping_Tagetik"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":26,"C":26,"D":50}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:D2"); C(ws,"B2","Passerelle vers CCH Tagetik",CTIT,FN_NAVY,align=AL); ws.row_dimensions[2].height=26
band(ws,4,"B","D","Dimensions & objets")
for i,h in enumerate(["Concept du modèle","Objet Tagetik","Détail"]): C(ws,f"{GL(2+i)}5",h,CHDR,FN_BLUE,align=AC,border=True)
maps=[("Marque / Campus","Entity","Hiérarchie Groupe → Marque → Campus (04_Referentiel)."),
 ("Programme / Niveau / Modalité","Dimensions analytiques","Axes de la maille fine ; attributs initial/alternance."),
 ("Comptes P&L","Account","Plan 04_Referentiel, avec tag Fixe/Variable pour l'analyse."),
 ("Effectifs, candidatures, classes","Comptes techniques (statistiques)","Inducteurs non-financiers pour le calcul piloté."),
 ("Réalisé N-1 / Budget N+1","Category","ACTUAL vs BUDGET."),
 ("Scénarios (Cadrage/Opt/Prudent)","Versions","Une version par scénario ; en Excel via 02_Parametres!C3."),
 ("Coefficients stratégiques","Règle de calcul / driver","Cascade cadrage × coef à la maille fine (comme 03/07)."),
 ("Allocation frais de structure","Cost allocation / driver","Répartition par effectif/CA/m² (08_Allocation)."),
 ("Snapshot du budget top-down","Data snapshot / version figée","Fige la V1 top-down avant d'ouvrir le workflow."),
 ("Reprise par les responsables","Workflow bottom-up","Saisie/validation par campus → marque → groupe.")]
r=6
for a,b,c in maps:
    C(ws,f"B{r}",a,CB,align=ALW,border=True); C(ws,f"C{r}",b,CREG,align=ALW,border=True); C(ws,f"D{r}",c,CREG,align=ALW,border=True); ws.row_dimensions[r].height=30; r+=1
band(ws,r+1,"B","D","Séquence de bascule recommandée"); r+=2
for t in ["1. Créer les dimensions (Entity, Account, analytiques) et charger le réalisé N-1 (ACTUAL).",
 "2. Traduire les inducteurs (02/03) en règles de calcul Tagetik (volume, prix, conversion, classes).",
 "3. Générer le budget top-down par cellule (logique 07_Moteur) et contrôler vs 09_PnL.",
 "4. Figer un SNAPSHOT (version) du budget top-down.",
 "5. Ouvrir le WORKFLOW bottom-up : chaque responsable campus challenge/ajuste, la marque valide.",
 "6. Décliner les scénarios en versions et automatiser l'allocation par driver."]:
    ws.merge_cells(f"B{r}:D{r}"); C(ws,f"B{r}",t,CREG,align=ALW); r+=1

# ---------------------------------------------------------------- recalc on load
try: wb.calculation.fullCalcOnLoad=True
except Exception:
    from openpyxl.workbook.properties import CalcProperties
    wb.calculation=CalcProperties(fullCalcOnLoad=True)
wb.save(OUT)
print("OK",OUT)
print(f"cellules fines N={N}  campus CG={CG}")
print(f"[python] CA N-1={grp_ca_n1:,.0f}  EBITDA N-1={sum(c['contrib'] for c in campus)-sum(c['loyer'] for c in campus)-sum(c['perm'] for c in campus)-grp_struct_n1:,.0f}")
