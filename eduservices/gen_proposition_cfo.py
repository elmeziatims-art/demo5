# -*- coding: utf-8 -*-
"""
PROPOSITION CFO — exemple autonome de "back-solve" par cascade de priorite.
Le CFO saisit 2 cibles (croissance CA %, marge EBITDA %). L'outil propose,
levier par levier dans l'ordre de priorite, les valeurs qui atteignent la cible,
avec garde-fou de faisabilite. Tout est en formules (recalcul live, sans macro).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Proposition CFO"
try:
    wb.calculation.fullCalcOnLoad = True
except Exception:
    pass

# ---- styles ----
TITLE = Font(bold=True, size=16, color="1F3864")
H     = Font(bold=True, size=11, color="FFFFFF")
BOLD  = Font(bold=True)
ITAL  = Font(italic=True, size=9, color="595959")
INPUTF= Font(bold=True, color="9C5700")
GREENF= Font(bold=True, color="375623")
subfill = PatternFill("solid", fgColor="2E5496")
hfill   = PatternFill("solid", fgColor="1F3864")
inputfl = PatternFill("solid", fgColor="FFF2CC")   # jaune = saisie
outfl   = PatternFill("solid", fgColor="E2EFDA")   # vert = propose
band    = PatternFill("solid", fgColor="D9E1F2")
thin    = Side(style="thin", color="BFBFBF")
box     = Border(left=thin, right=thin, top=thin, bottom=thin)
center  = Alignment(horizontal="center", vertical="center")
left    = Alignment(horizontal="left", vertical="center")
right   = Alignment(horizontal="right", vertical="center")
wrap    = Alignment(horizontal="left", vertical="center", wrap_text=True)

EUR = '# ##0 "€";[Red]-# ##0 "€"'
PCT = '0.0%'
PTS = '+0.00" pts"'
PC2 = '+0.00"%"'

widths = {"A":2,"B":31,"C":10,"D":8,"E":13,"F":14,"G":13,"H":12,"I":14,"J":2}
for c,w in widths.items():
    ws.column_dimensions[c].width = w

def cell(coord, val, font=None, fill=None, fmt=None, align=None, border=False):
    c = ws[coord]; c.value = val
    if font: c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if align: c.alignment = align
    if border: c.border = box
    return c

def merge(rng, fill=None):
    ws.merge_cells(rng)
    if fill:
        a = rng.split(":")[0]; ws[a].fill = fill

# ================= TITRE =================
merge("B2:I2"); cell("B2","PROPOSITION CFO — atteindre la cible par cascade de leviers", TITLE)
merge("B3:I3"); cell("B3","Le CFO pose son ambition (2 cibles). L'outil propose les leviers, du moins risque au plus "
                          "risque, et alerte si la cible depasse les bornes realistes.", ITAL)

# ================= 1. SAISIE =================
merge("B5:I5", subfill); cell("B5","  ①  SAISIE DU CFO   (cellules jaunes = a saisir)", H, align=left)
rows_in = [
    ("B6","CA de reference 2026",                      "C6",22544725, EUR),
    ("B7","Croissance CA cible",                        "C7",0.06,     PCT),
    ("B8","CA construit actuel (leviers en place)",     "C8",23200000, EUR),
    ("B9","Marge EBITDA cible",                         "C9",0.16,     PCT),
    ("B10","Couts construits @ CA cible (leviers neutres)","C10",20300000, EUR),
]
for lab_c,lab,val_c,val,fmt in rows_in:
    cell(lab_c,lab,BOLD,align=left,border=True); merge(f"{lab_c}:{lab_c[0]}{lab_c[1:]}")
for lab_c,lab,val_c,val,fmt in rows_in:
    r = lab_c[1:]
    ws.merge_cells(f"B{r}:B{r}")
    cell(val_c,val,INPUTF,inputfl,fmt,right,True)
    ws.merge_cells(f"D{r}:I{r}")
    for cc in range(4,10):
        ws.cell(row=int(r),column=cc).border = box

# ================= 2. IMPLICATIONS =================
merge("B12:I12", subfill); cell("B12","  ②  CE QUE CA IMPLIQUE   (calcule)", H, align=left)
der = [
    ("B13","CA cible",           "C13","=C6*(1+C7)",    EUR),
    ("B14","Ecart CA a combler", "C14","=MAX(0,C13-C8)",EUR),
    ("B15","EBITDA cible",       "C15","=C13*C9",       EUR),
    ("B16","EBITDA construit",   "C16","=C13-C10",      EUR),
    ("B17","Marge construite",   "C17","=C16/C13",      PCT),
    ("B18","Couts a reduire",    "C18","=MAX(0,C15-C16)",EUR),
]
for lab_c,lab,val_c,f,fmt in der:
    r = lab_c[1:]
    cell(lab_c,lab,BOLD,align=left,border=True)
    cell(val_c,f,None,band,fmt,right,True)
    ws.merge_cells(f"D{r}:I{r}")
    for cc in range(4,10):
        ws.cell(row=int(r),column=cc).fill = band
        ws.cell(row=int(r),column=cc).border = box

# ---------- helper: cascade table ----------
def cascade(title, top_row, besoin_ref, levers):
    merge(f"B{top_row}:I{top_row}", subfill)
    cell(f"B{top_row}", title, H, align=left)
    hr = top_row+1
    hdr = ["Levier","Borne","Unite","Rendt €/u","Apport max €","Tire €","Applique","Reste €"]
    for i,t in enumerate(hdr):
        col = get_column_letter(2+i)
        cell(f"{col}{hr}", t, H, hfill, None, center, True)
    r0 = hr+1
    n = len(levers)
    for j,(lab,borne,unite,rend,fmta) in enumerate(levers):
        r = r0+j
        cell(f"B{r}", lab, BOLD, align=left, border=True)
        cell(f"C{r}", borne, INPUTF, inputfl, '0.0', center, True)
        cell(f"D{r}", unite, None, None, None, center, True)
        cell(f"E{r}", rend, INPUTF, inputfl, '# ##0', right, True)
        cell(f"F{r}", f"=C{r}*E{r}", None, None, EUR, right, True)          # apport max
        prev_reste = f"$C${besoin_ref}" if j==0 else f"I{r-1}"
        cell(f"G{r}", f"=MIN(F{r},MAX(0,{prev_reste}))", GREENF, outfl, EUR, right, True)  # tire
        cell(f"H{r}", f"=IF(E{r}=0,0,G{r}/E{r})", GREENF, outfl, fmta, center, True)       # applique
        cell(f"I{r}", f"={prev_reste}-G{r}", None, None, EUR, right, True)                  # reste
    return r0, r0+n-1

# ================= 3. MOLETTE CA =================
ca_levers = [
    ("Conversion lead→cand", 2.0,"pts",202902, PTS),
    ("Taux de passage",           1.5,"pts",112724, PTS),
    ("Prix",                      3.0,"%",  225447, PC2),
    ("Budget acquisition",       30.0,"%",  16667,  PC2),
]
ca_first, ca_last = cascade("  ③  MOLETTE CA — cascade de priorite (comble l'ecart CA)", 20, 14, ca_levers)

# synthese CA
sr = ca_last+1
cell(f"B{sr}","CA propose = construit + tire", BOLD, band, None, left, True)
cell(f"F{sr}", f"=C8+SUM(G{ca_first}:G{ca_last})", BOLD, band, EUR, right, True)
cell(f"G{sr}","Residu", BOLD, band, None, center, True)
cell(f"I{sr}", f"=I{ca_last}", BOLD, band, EUR, right, True)
for cc in range(3,6):
    ws.cell(row=sr,column=cc).fill=band; ws.cell(row=sr,column=cc).border=box
vr = sr+1
merge(f"B{vr}:I{vr}")
cell(f"B{vr}",
     f'=IF(I{ca_last}<=1,"✅ Cible CA atteignable dans les bornes",'
     f'"⛔ Cible CA hors bornes  |  max atteignable = "&TEXT(C8+SUM(F{ca_first}:F{ca_last}),"# ##0")&" €  '
     f'(soit +"&TEXT((C8+SUM(F{ca_first}:F{ca_last}))/C6-1,"0.0%")&")  |  ecart residuel = "&TEXT(I{ca_last},"# ##0")&" €")',
     BOLD, None, None, left)

# ================= 4. MOLETTE EBITDA =================
eb_top = vr+2
eb_levers = [
    ("Productivite",       2.0,"%", 80000, PC2),
    ("Couts de structure", 1.5,"%", 90000, PC2),
    ("Masse salariale",    3.0,"%", 33000, PC2),
]
eb_first, eb_last = cascade("  ④  MOLETTE EBITDA — cascade cote couts (comble les couts a reduire)",
                            eb_top, 18, eb_levers)
sr2 = eb_last+1
cell(f"B{sr2}","EBITDA propose = construit + reduction", BOLD, band, None, left, True)
cell(f"F{sr2}", f"=C16+SUM(G{eb_first}:G{eb_last})", BOLD, band, EUR, right, True)
cell(f"G{sr2}","Marge", BOLD, band, None, center, True)
cell(f"I{sr2}", f"=(C16+SUM(G{eb_first}:G{eb_last}))/C13", BOLD, band, PCT, right, True)
for cc in range(3,6):
    ws.cell(row=sr2,column=cc).fill=band; ws.cell(row=sr2,column=cc).border=box
vr2 = sr2+1
merge(f"B{vr2}:I{vr2}")
cell(f"B{vr2}",
     f'=IF(I{eb_last}<=1,"✅ Cible EBITDA atteignable dans les bornes",'
     f'"⛔ Cible EBITDA hors bornes  |  marge max = "&TEXT((C16+SUM(F{eb_first}:F{eb_last}))/C13,"0.0%")&"  '
     f'|  couts residuels a trouver = "&TEXT(I{eb_last},"# ##0")&" €")',
     BOLD, None, None, left)

# ================= NOTICE =================
nr = vr2+2
merge(f"B{nr}:I{nr}")
cell(f"B{nr}","Comment lire : chaque molette tire ses leviers DANS L'ORDRE, chacun jusqu'a sa borne, "
              "jusqu'a combler l'ecart. Le dernier levier ne bouge que partiellement. Si tout est au max et "
              "qu'il reste un ecart -> la cible est hors bornes (garde-fou). Colonnes jaunes = a saisir "
              "(cibles, bornes, rendements) ; colonnes vertes = proposition calculee.", ITAL)
ws.row_dimensions[nr].height = 46
ws[f"B{nr}"].alignment = wrap

ws.sheet_view.showGridLines = False
out = "/home/user/demo5/eduservices/PROPOSITION_CFO_DEMO.xlsx"
wb.save(out)
print("saved", out)
