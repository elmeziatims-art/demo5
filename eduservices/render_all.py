#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Rendu QA des onglets de restitution (valeurs illustratives pour les cellules-
formule ; le vrai calcul se fait dans Excel). Usage: python3 render_all.py [cad|PIL|ALLOC|all]"""
import warnings,sys;warnings.filterwarnings("ignore")
import openpyxl
from pres_lib import render_html, screenshot
F="Simulation_CFO_EDUSERVICES_2026-2027.xlsx"
wb=openpyxl.load_workbook(F)

def inj_cad(ws):
    inj={"C5":"Cadrage","D13":23671961,"E13":23671969,"F13":8,"G13":0.0000003,
     "D14":3550794,"E14":3550797,"F14":3,"G14":0.0000008,
     "C15":0.14600,"D15":0.15,"E15":0.15000,"F15":0.0,"G15":0.0,
     "E16":3175,"F16":139,"G16":0.04578,"C21":"Cadrage","D21":"Optimiste","E21":"Prudent"}
    for r in [23,24,25,26,27,28,32,33,34,35,36,39]: inj["F%d"%r]=ws["C%d"%r].value
    return inj

def inj_pil(ws):
    rows=list(range(15,29)); syn=list(range(33,47))
    K=[ws["K%d"%r].value or 0 for r in rows]; D=[ws["D%d"%r].value or 0 for r in rows]
    tot=sum(K); CA=23671969; EFF=3175; EBC=4350797; SIEGE=-800000
    inj={"B7":CA,"H7":3550797,"K7":0.15,"N7":EFF}
    for i,r in enumerate(syn):
        w=K[i]/tot; ca=round(CA*w); eff=round(EFF*w); eb=round(EBC*w)
        inj["D%d"%r]=eff; inj["E%d"%r]=ca; inj["F%d"%r]=round(ca/eff) if eff else 0
        inj["G%d"%r]=ca/CA; inj["H%d"%r]=eb; inj["I%d"%r]=eb/ca if ca else 0
        inj["J%d"%r]=round(eb/eff) if eff else 0; inj["K%d"%r]=round(K[i]*1.05); inj["L%d"%r]=round(D[i])
    inj.update({"D47":EFF,"E47":CA,"F47":round(CA/EFF),"G47":1.0,"H47":EBC,"I47":EBC/CA,
     "J47":round(EBC/EFF),"K47":round(tot*1.05),"L47":round(sum(D)/len(D)),"H48":SIEGE,
     "D49":EFF,"E49":CA,"F49":round(CA/EFF),"G49":1.0,"H49":3550797,"I49":3550797/CA,
     "J49":round(3550797/EFF),"K49":round(tot*1.05)})
    return inj

def inj_alloc(ws):
    # niveaux
    info=[]
    for r in range(18,96):
        b=ws["B%d"%r].value
        lead=0
        # apres format, l'indentation est via alignment.indent ; on relit le code campus si present
        lvl=2
        al=ws["B%d"%r].alignment
        if al and al.indent is not None: lvl=al.indent//2
        info.append((r,lvl))
    # valeurs feuilles (classes lvl2)
    val={}
    def leaf(i,r):
        ca=80000+((i*53000)%320000)
        eff=max(8,round(ca/7500))
        vac=round(ca*0.26);perm=round(ca*0.20);odir=round(ca*0.08)
        struct=round(ca*0.10);frais=round(ca*0.07);hold=round(ca*0.06)
        cout=vac+perm+odir+struct+frais+hold;marge=ca-cout
        return dict(C=eff,D=ca,E=vac,F=perm,G=odir,H=struct,I=frais,J=hold,K=cout,L=marge,M=marge/ca if ca else 0)
    idx=0
    for k,(r,lvl) in enumerate(info):
        if lvl>=2:
            val[r]=leaf(idx,r); idx+=1
    # rollup parents (somme des feuilles de leur span)
    cols=["C","D","E","F","G","H","I","J","K","L"]
    for k,(r,lvl) in enumerate(info):
        if lvl>=2: continue
        # span jusqu'au prochain row de niveau <= lvl
        s={c:0 for c in cols}
        for j in range(k+1,len(info)):
            rr,ll=info[j]
            if ll<=lvl: break
            if rr in val:
                for c in cols: s[c]+=val[rr][c]
        s["M"]=(s["L"]/s["D"]) if s["D"] else 0
        val[r]=s
    # GROUPE (derniere ligne lvl0) = somme de toutes les feuilles
    leafrows=[r for (r,lvl) in info if lvl>=2]
    groupe=info[-1][0]
    g={c:sum(val[rr][c] for rr in leafrows) for c in cols}
    g["M"]=(g["L"]/g["D"]) if g["D"] else 0
    val[groupe]=g
    inj={}
    for r,d in val.items():
        for c,v in d.items(): inj["%s%d"%(c,r)]=v
    return inj

def render(tab):
    ws=wb[tab]
    inj={"cad":inj_cad,"PIL":inj_pil,"ALLOC":inj_alloc}[tab](ws)
    mr={"cad":40,"PIL":49,"ALLOC":95}[tab]; mc={"cad":10,"PIL":15,"ALLOC":13}[tab]
    html=render_html(ws,inject=inj,max_row=mr,max_col=mc)
    open("/tmp/render/%s.html"%tab,"w").write(html)
    screenshot(html,"/tmp/render/%s.png"%tab)
    print("rendered",tab)

if __name__=="__main__":
    which=sys.argv[1] if len(sys.argv)>1 else "all"
    for t in (["cad","PIL","ALLOC"] if which=="all" else [which]): render(t)
