#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""2 onglets pertinents : Diagnostic CAC (funnel + CAC par marque) et
Structure & Mix (mix initiale/alternance + CA par marque). Graphes = comparaison
et composition (pas des courbes qui montent). Vrais chiffres 2026."""
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.chart import BarChart,Reference
INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E4F0EC"; CARD2="F5F7F9"; FAINT="7D8B98"
WHITE="FFFFFF"; BLUE="0000FF"; OCHRE="B3641C"; OCHREBG="F7EAD9"; NAVY="3D4F8F"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
CTR=Alignment("center",vertical="center",wrap_text=True); LFT=Alignment("left",vertical="center"); RGT=Alignment("right",vertical="center")
EUR='#,##0;(#,##0);"-"'; PCT='0.0%'; NUM='#,##0'; EURc='#,##0" €"'
def hdr(ws,r,labels,fromcol=1):
    for j,h in enumerate(labels,fromcol):
        c=ws.cell(r,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL); c.alignment=CTR if j>fromcol else LFT

wb=openpyxl.Workbook()

# ============ ONGLET 1 : DIAGNOSTIC CAC ============
d=wb.active; d.title="Diagnostic CAC"; d.sheet_view.showGridLines=False
d["A1"]="DIAGNOSTIC CAC  ·  pourquoi le coût d'acquisition se dégrade  ·  2026"; d["A1"].font=F(15,True,INK)
d["A2"]="1er drill depuis la tuile CAC du cockpit. Le funnel explique le CAC ; le CAC par marque dit OÙ agir."; d["A2"].font=F(9,False,TEALD)

# --- funnel ---
d["A4"]="Funnel de conversion (groupe)"; d["A4"].font=F(11,True,TEALD)
hdr(d,5,["Étape","Volume","Taux de passage"])
FUN=[("Leads",17197,None),("Candidatures",3720,"=B7/B6"),("Admis",2623,"=B8/B7"),("Inscrits",1229,"=B9/B8")]
for i,(lab,vol,tx) in enumerate(FUN):
    r=6+i
    d.cell(r,1,lab).font=F(10); d.cell(r,1).alignment=LFT
    d.cell(r,2,vol).font=F(10,False,BLUE); d.cell(r,2).number_format=NUM; d.cell(r,2).alignment=RGT
    if tx: d.cell(r,3,tx).number_format=PCT; d.cell(r,3).font=F(10,True,TEALD); d.cell(r,3).alignment=RGT
d.cell(10,1,"Global lead → inscrit").font=F(9,True,FAINT)
d.cell(10,3,"=B9/B6").number_format=PCT; d.cell(10,3).font=F(10,True,OCHRE); d.cell(10,3).alignment=RGT
fch=BarChart(); fch.type="col"; fch.title="Funnel : le volume fond à chaque étape"; fch.legend=None; fch.height=6.5; fch.width=11
fch.add_data(Reference(d,min_col=2,min_row=6,max_row=9)); fch.set_categories(Reference(d,min_col=1,min_row=6,max_row=9))
d.add_chart(fch,"E4")

# --- CAC par marque ---
d["A13"]="CAC par marque — où le coût dérape"; d["A13"].font=F(11,True,TEALD)
hdr(d,14,["Marque","Dépense acq.","Inscrits","CAC (€/inscrit)"])
CAC=[("MBway",166530,510),("ISCOM",108103,338),("IPAC",51871,130),("Pigier",58580,164),("Tunon",49090,87)]
for i,(m,dep,ins) in enumerate(CAC):
    r=15+i; ten=(m=="Tunon")
    col=OCHRE if ten else INK
    d.cell(r,1,m).font=F(10,True,col); d.cell(r,1).alignment=LFT
    d.cell(r,2,dep).font=F(10,False,BLUE); d.cell(r,3,ins).font=F(10,False,BLUE)
    d.cell(r,4,f"=B{r}/C{r}").font=F(10,True,col)
    for cc,fmt in [(2,EUR),(3,NUM),(4,EURc)]: d.cell(r,cc).number_format=fmt; d.cell(r,cc).alignment=RGT
    if ten: d.cell(r,5,"◄ 1,7× la moyenne").font=F(9,True,OCHRE)
cch=BarChart(); cch.type="bar"; cch.title="CAC par marque : Tunon décroche (564 € vs ~320 €)"; cch.legend=None; cch.height=6.5; cch.width=11
cch.add_data(Reference(d,min_col=4,min_row=15,max_row=19)); cch.set_categories(Reference(d,min_col=1,min_row=15,max_row=19))
d.add_chart(cch,"F13")
d.cell(21,1,"Le funnel dit COMMENT (taux de passage) ; le CAC par marque dit OÙ agir → Tunon en priorité. Clic marque → funnel du campus.").font=F(8,True,OCHRE,True)
for col,w in zip("ABCDE",[22,13,11,15,16]): d.column_dimensions[col].width=w

# ============ ONGLET 2 : STRUCTURE & MIX ============
s=wb.create_sheet("Structure & Mix"); s.sheet_view.showGridLines=False
s["A1"]="STRUCTURE & MIX DE FINANCEMENT  ·  2026"; s["A1"].font=F(15,True,INK)
s["A2"]="Ce que le CRM ne montre pas : la dépendance à l'alternance (OPCO). Vraie question DAF — risque réglementaire."; s["A2"].font=F(9,False,TEALD)

s["A4"]="Mix scolarité : initiale vs alternance (OPCO), par marque"; s["A4"].font=F(11,True,TEALD)
hdr(s,5,["Marque","Initiale (706)","Alternance (7062)","% alternance"])
MIX=[("MBway",1995975,7525580),("ISCOM",1310700,4951430),("IPAC",0,2571100),("Pigier",0,2137460),("Tunon",0,1941870)]
for i,(m,ini,alt) in enumerate(MIX):
    r=6+i; full=(ini==0)
    col=OCHRE if full else INK
    s.cell(r,1,m).font=F(10,True,col); s.cell(r,1).alignment=LFT
    s.cell(r,2,ini).font=F(10,False,BLUE); s.cell(r,3,alt).font=F(10,False,BLUE)
    s.cell(r,4,f"=C{r}/(B{r}+C{r})").font=F(10,True,col)
    for cc,fmt in [(2,EUR),(3,EUR),(4,PCT)]: s.cell(r,cc).number_format=fmt; s.cell(r,cc).alignment=RGT
    if full: s.cell(r,5,"100 % OPCO").font=F(9,True,OCHRE)
mch=BarChart(); mch.type="col"; mch.grouping="stacked"; mch.overlap=100
mch.title="Mix initiale / alternance : 3 marques 100% OPCO"; mch.height=6.5; mch.width=12
mch.add_data(Reference(s,min_col=2,max_col=3,min_row=5,max_row=11),titles_from_data=True)
mch.set_categories(Reference(s,min_col=1,min_row=6,max_row=11))
s.add_chart(mch,"F4")

s["A13"]="CA total par marque — concentration"; s["A13"].font=F(11,True,TEALD)
hdr(s,14,["Marque","CA total"])
CA=[("MBway",9567455),("ISCOM",6292550),("IPAC",2582800),("Pigier",2152220),("Tunon",1949700)]
for i,(m,ca) in enumerate(CA):
    r=15+i
    s.cell(r,1,m).font=F(10,True); s.cell(r,1).alignment=LFT
    s.cell(r,2,ca).font=F(10,False,BLUE); s.cell(r,2).number_format=EUR; s.cell(r,2).alignment=RGT
cca=BarChart(); cca.type="bar"; cca.title="CA par marque : MBway + ISCOM = 70% du CA"; cca.legend=None; cca.height=6; cca.width=12
cca.add_data(Reference(s,min_col=2,min_row=15,max_row=19)); cca.set_categories(Reference(s,min_col=1,min_row=15,max_row=19))
s.add_chart(cca,"F13")
s.cell(21,1,"INSIGHT DAF : le groupe est massivement financé par l'alternance (OPCO). MBway/ISCOM ~79% alt, les 3 autres 100%. Dépendance à surveiller.").font=F(8,True,OCHRE,True)
for col,w in zip("ABCDE",[16,15,17,13,13]): s.column_dimensions[col].width=w

out="/home/user/demo5/eduservices/tagetik/ONGLETS_DIAGNOSTIC_STRUCTURE.xlsx"
wb.save(out); print("SAVED",out)
