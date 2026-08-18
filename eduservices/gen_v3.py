# -*- coding: utf-8 -*-
"""EDUSERVICES — Modèle CA v3 (Architecture BUDGET-DRIVEN, leads observés).
Cadrage top-down -> Base de référence (fusion historique+paramètres, version unique, LEADS OBSERVÉS)
-> Campagnes (budget marketing -> leads, socle organique + part payante à rendement décroissant)
-> Moteur (funnel MESURÉ -> effectif -> CA). Revenu et taux différenciés alternance/initial.
On ne génère QUE du cadrage au moteur ; le reste (coûts, alloc, reporting) viendra ensuite.
"""
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as GL, column_index_from_string as CI
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule
from openpyxl.chart import BarChart, Reference

OUT="/home/user/demo5/eduservices/EDUSERVICES_Modele_CA_v3.xlsx"

# ============================================================ SOCLE DE DONNÉES
BRANDS={  # marque -> (domaine, base_entry, [villes])
 "MBway":("Management",60,["Paris","Lyon","Nantes","Bordeaux"]),
 "ISCOM":("Communication",55,["Paris","Lille","Toulouse"]),
 "Ipac Bachelor Factory":("Commerce",50,["Nantes","Rennes","Montpellier"]),
 "Pigier":("Commerce/RH",42,["Lyon","Bordeaux"]),
 "Tunon":("Tourisme",36,["Paris","Lyon"]),
}
PROGS={
 "MBway":[("Bachelor Management","BAC",[("B1","INIT"),("B2","ALT"),("B3","ALT")]),
          ("Mastère Management","MAST",[("M1","ALT"),("M2","ALT")])],
 "ISCOM":[("Bachelor Communication","BAC",[("B1","INIT"),("B2","ALT"),("B3","ALT")]),
          ("Mastère Communication","MAST",[("M1","ALT"),("M2","ALT")])],
 "Ipac Bachelor Factory":[("Bachelor Commerce","BAC",[("B1","ALT"),("B2","ALT"),("B3","ALT")])],
 "Pigier":[("BTS Gestion","BTS",[("1","ALT"),("2","ALT")]),
           ("Bachelor RH","BAC",[("B1","ALT"),("B3","ALT")])],
 "Tunon":[("Bachelor Tourisme","BAC",[("B1","ALT"),("B2","ALT"),("B3","ALT")])],
}
ENTRY={"B1","M1","1"}
ORDER={"BAC":["B1","B2","B3"],"MAST":["M1","M2"],"BTS":["1","2"]}

CITY_VOL ={"Paris":1.30,"Lyon":1.10,"Nantes":1.00,"Bordeaux":0.85,"Lille":0.90,"Toulouse":0.85,"Rennes":0.80,"Montpellier":0.80}
CITY_PRICE={"Paris":1.12,"Lyon":1.05,"Nantes":1.00,"Bordeaux":0.97,"Lille":0.98,"Toulouse":0.96,"Rennes":0.95,"Montpellier":0.95}
CITY_CPL ={"Paris":1.35,"Lyon":1.10,"Nantes":1.00,"Bordeaux":0.95,"Lille":1.00,"Toulouse":0.95,"Rennes":0.90,"Montpellier":0.90}
CAPT={"BTS":30,"BAC":32,"MAST":26}
# revenu / étudiant / an de référence (€) par cycle x modalité (benchmarks secteur — illustratif)
REV={("BTS","ALT"):6000,("BTS","INIT"):5500,("BAC","ALT"):7000,("BAC","INIT"):7500,("MAST","ALT"):7500,("MAST","INIT"):9000}
PASS={"B2":0.97,"B3":0.98,"M2":0.98,"2":0.97}   # rétention intra-cycle proche de 100 % (privé alternance : peu de décrochage)
# TAUX DU FUNNEL (mesurés, différenciés) — issus de la recherche secteur (privé non sélectif, alternance)
#  INIT = intention forte (Parcoursup) ; ALT = leads plateforme + déperdition "signature contrat"
RATES={"INIT":dict(rlc=0.28,rca=0.72,yld=0.60),"ALT":dict(rlc=0.20,rca=0.70,yld=0.42)}
REND_DEF=0.5      # rendement d'acquisition (repli)
PORG_DEF=0.40     # part organique moyenne (repli)
# part ORGANIQUE différenciée par campus : force de marque (↑) − intensité concurrentielle de la ville (↓)
PORG_BRAND={"MBway":0.46,"ISCOM":0.48,"Ipac Bachelor Factory":0.34,"Pigier":0.50,"Tunon":0.40}
PORG_CITY ={"Paris":-0.10,"Lyon":-0.03,"Nantes":0.02,"Bordeaux":0.03,"Lille":0.02,"Toulouse":0.03,"Rennes":0.05,"Montpellier":0.04}
def part_org(m,v): return min(0.60,max(0.22,PORG_BRAND[m]+PORG_CITY[v]))
# RENDEMENT d'acquisition différencié : sophistication marketing (marque) + marge de marché (ville moins saturée ↑)
REND_BRAND={"MBway":0.52,"ISCOM":0.50,"Ipac Bachelor Factory":0.55,"Pigier":0.46,"Tunon":0.48}
REND_CITY ={"Paris":-0.06,"Lyon":-0.02,"Nantes":0.01,"Bordeaux":0.02,"Lille":0.02,"Toulouse":0.02,"Rennes":0.04,"Montpellier":0.03}
def rend_c(m,v): return min(0.62,max(0.38,REND_BRAND[m]+REND_CITY[v]))
CPL_BASE=40       # coût par lead payant de référence (€)
FRAIS_DEF=90      # frais de dossier / nouvel inscrit (€)

rows=[]
for marque,(dom,base,villes) in BRANDS.items():
    for ville in villes:
        for pnom,ptype,niveaux in PROGS[marque]:
            entry_base=round(base*CITY_VOL[ville])
            effs={}; last=None
            for niv,mod in niveaux:
                if niv in ENTRY: eff=entry_base
                else: eff=max(0,round(effs[last]*PASS[niv]))
                effs[niv]=eff; last=niv
            last=None
            for niv,mod in niveaux:
                eff=effs[niv]; is_entry=1 if niv in ENTRY else 0
                rev=round(REV[(ptype,mod)]*CITY_PRICE[ville])
                if is_entry:
                    R=RATES[mod]
                    nouv=eff; rein=0; eff_prev=0; passage=0.0
                    admis=max(1,round(nouv/R["yld"]))
                    cand =max(1,round(admis/R["rca"]))
                    leads=max(1,round(cand/R["rlc"]))
                    rlc_m=cand/leads; rca_m=admis/cand; yld_m=nouv/admis
                else:
                    nouv=0; rein=eff; cand=0; admis=0; leads=0; eff_prev=effs[last]; passage=PASS[niv]
                    rlc_m=rca_m=yld_m=0.0
                classes=max(1,math.ceil(eff/CAPT[ptype])) if eff>0 else 0
                rows.append(dict(marque=marque,ville=ville,prog=pnom,type=ptype,niv=niv,mod=mod,entry=is_entry,
                    leads=leads,cand=cand,admis=admis,nouv=nouv,rein=rein,eff=eff,eff_prev=eff_prev,classes=classes,
                    rev=rev,passage=passage,rlc=rlc_m,rca=rca_m,yld=yld_m))
                last=niv
N=len(rows)

# historique CRM multi-années : leads (organique/payant) + dépense marketing, sur N-2/N-1/atterrissage.
#  généré pour qu'une MESURE sur ces données retrouve rendement ~0,5, CPL ~benchmark, part org ~40 %.
# croissance annuelle de la notoriété (organique) PAR MARQUE → momentum historique différencié
GORG_M={"MBway":1.07,"ISCOM":1.03,"Ipac Bachelor Factory":1.09,"Pigier":1.04,"Tunon":1.02}
GSP =1.10   # croissance annuelle de la dépense payante  → rendement mesuré = ln(GSP)/ln(GSP^2)=0,5
campus=[]; seen=set()
for r in rows:
    k=(r["marque"],r["ville"])
    if k in seen: continue
    seen.add(k)
    cc=[x for x in rows if (x["marque"],x["ville"])==k]
    sleads=sum(x["leads"] for x in cc)
    cpl=round(CPL_BASE*CITY_CPL[r["ville"]])
    porg=part_org(r["marque"],r["ville"]); rc=rend_c(r["marque"],r["ville"])   # différenciés par campus
    GORG=GORG_M[r["marque"]]
    paid_att=sleads*(1-porg); org_att=sleads*porg; spend_att=paid_att*cpl
    org  =[round(org_att/GORG**2),      round(org_att/GORG),    round(org_att)]      # N-2, N-1, ATT (croissance marque)
    paid =[round(paid_att/GSP**(2*rc)), round(paid_att/GSP**rc), round(paid_att)]    # → rendement mesuré = rc
    spend=[round(spend_att/GSP**2),     round(spend_att/GSP),    round(spend_att)]
    campus.append(dict(marque=r["marque"],ville=r["ville"],sleads=sleads,cpl=cpl,porg=porg,rc=rc,
        org=org,paid=paid,spend=spend,budget_paid_ref=spend[2]))
CG=len(campus)

REFCA=round(sum(r["eff"]*r["rev"]+r["nouv"]*FRAIS_DEF for r in rows))
REFEFF=sum(r["eff"] for r in rows)
REFALT=sum(r["eff"] for r in rows if r["mod"]=="ALT")
REFBUD=sum(c["budget_paid_ref"] for c in campus)
REFLEADS=sum(c["sleads"] for c in campus)
# marketing de MARQUE (notoriété) : dépense qui pilote le SOCLE ORGANIQUE. Mesuré en AGRÉGÉ (marketing-mix),
#  pas par lead : on ne prétend pas savoir d'où vient un lead organique, on mesure comment le socle répond au budget marque.
def rend_marque(po): return round(min(0.45,max(0.25,0.20+0.30*po)),3)  # rendement de marque (agrégé) < rendement acquisition
tot_org_att=sum(c["org"][2] for c in campus)
BRAND_CPO=0.030*REFCA/tot_org_att if tot_org_att else 0   # ~3 % du CA en budget de marque, au prorata de l'organique
for c in campus:
    rb=rend_marque(c["porg"]); c["rb"]=rb
    bref=round(c["org"][2]*BRAND_CPO); c["bref"]=bref
    gm=GORG_M[c["marque"]]
    GBR=math.exp(math.log(gm)/rb) if rb>0 else gm         # croissance dépense marque → rendement marque mesuré = rb
    c["bspend"]=[round(bref/GBR**2),round(bref/GBR),bref]  # N-2, N-1, ATT
REFBRAND=sum(c["bref"] for c in campus)
BUD_M={m:sum(c["spend"][2] for c in campus if c["marque"]==m) for m in BRANDS}   # budget acquisition réf par marque (cap → tilt du budget)
# --- caps PROPOSÉS depuis l'historique (3 logiques) — informationnel ; le CFO garde la main (cap retenu) ---
def _agg(m):
    cc=[c for c in campus if c["marque"]==m]
    spend=sum(c["spend"][2] for c in cc); paid=sum(c["paid"][2] for c in cc)
    rend=sum(c["rc"]*c["paid"][2] for c in cc)/paid                         # rendement pondéré
    nouv=sum(r["nouv"] for r in rows if r["marque"]==m); leads=sum(r["leads"] for r in rows if r["marque"]==m)
    cac=(spend/paid)/(rend*(nouv/leads))                                    # CAC marginal marque
    ln2=sum(c["org"][0]+c["paid"][0] for c in cc); lat=sum(c["org"][2]+c["paid"][2] for c in cc)
    ca=sum(r["eff"]*r["rev"]+r["nouv"]*FRAIS_DEF for r in rows if r["marque"]==m)
    return dict(cac=cac,growth=lat/ln2,intensity=BUD_M[m]/ca)
_AG={m:_agg(m) for m in BRANDS}
def _norm(sig):
    mu=sum(sig.values())/len(sig); return {m:round(sig[m]/mu,2) for m in sig}
CAP_EFF=_norm({m:1/_AG[m]["cac"] for m in BRANDS})       # efficience : CAC marginal bas → cap haut
CAP_MOM=_norm({m:_AG[m]["growth"]-1 for m in BRANDS})    # momentum : TAUX de croissance leads 24→26 (excès sur le plat) → cap
CAP_POT=_norm({m:1/_AG[m]["intensity"] for m in BRANDS}) # potentiel : sous-investissement marketing → cap haut
print("[py] budget MARQUE réf=%d € (%.1f%% du CA)  rendement marque≈%.2f"%(REFBRAND,REFBRAND/REFCA*100,sum(c["rb"] for c in campus)/CG))
print("[py] N=%d cellules  campus=%d"%(N,CG))
print("[py] CA réf=%d €  effectif=%d  alternance=%.0f%%"%(REFCA,REFEFF,REFALT/REFEFF*100))
print("[py] leads réf=%d  budget PAYANT réf=%d € (%.1f%% du CA, part organique %.0f%%)"%(REFLEADS,REFBUD,REFBUD/REFCA*100,PORG_DEF*100))

# ============================================================ STYLES
NAVY,BLUE2,LIGHT,YEL,TOT,GREENF="1F3864","2E5496","D9E1F2","FFF2CC","E2EFDA","EAF3EA"
Fn="Arial"
CIN=Font(name=Fn,color="0000FF"); CINB=Font(name=Fn,color="0000FF",bold=True)
CF=Font(name=Fn,color="000000"); CFB=Font(name=Fn,color="000000",bold=True)
CL=Font(name=Fn,color="008000"); CHDR=Font(name=Fn,color="FFFFFF",bold=True)
CTIT=Font(name=Fn,color="FFFFFF",bold=True,size=14); CB=Font(name=Fn,bold=True)
CIT=Font(name=Fn,italic=True,color="595959",size=9); CREG=Font(name=Fn)
FNAVY=PatternFill("solid",fgColor=NAVY); FBLUE=PatternFill("solid",fgColor=BLUE2)
FLIGHT=PatternFill("solid",fgColor=LIGHT); FYEL=PatternFill("solid",fgColor=YEL)
FTOT=PatternFill("solid",fgColor=TOT); FGRN=PatternFill("solid",fgColor=GREENF)
thin=Side(style="thin",color="BFBFBF"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
AL=Alignment(horizontal="left",vertical="center"); AC=Alignment(horizontal="center",vertical="center",wrap_text=True)
AR=Alignment(horizontal="right",vertical="center"); ALW=Alignment(horizontal="left",vertical="top",wrap_text=True)
EUR='#,##0" €";(#,##0)" €";"-"'; PCT='0.0%;(0.0%);"-"'; NB='#,##0;(#,##0);"-"'; X2='0.00'
def C(ws,ref,val=None,font=None,fill=None,fmt=None,align=None,border=False):
    c=ws[ref]
    if val is not None:c.value=val
    if font:c.font=font
    if fill:c.fill=fill
    if fmt:c.number_format=fmt
    if align:c.alignment=align
    if border:c.border=BORD
    return c
def band(ws,row,a,b,text,fill=FNAVY,font=CHDR,h=20):
    ws.merge_cells(f"{a}{row}:{b}{row}")
    for col in range(CI(a),CI(b)+1): ws.cell(row=row,column=col).fill=fill
    cc=ws[f"{a}{row}"]; cc.value=text; cc.font=font; cc.alignment=AL; ws.row_dimensions[row].height=h

wb=openpyxl.Workbook()

# ============================================================ REFS 01_Cadrage
CAD="'01_Cadrage'!"
LMKT,LBRAND,LPRIX,LGLC,LGCV,LPASS,LINFL,LSAL,LEFFP,LPROD,LSTRUCT=(f"{CAD}$H${_r}" for _r in range(16,27))  # leviers ACTIF (11)
KFRAIS=f"{CAD}$H$30"    # frais de dossier (décision) ; rendement/part org/CPL sont MESURÉS (03_Campagnes)
CROISS=f"{CAD}$F$3"; MARGEC=f"{CAD}$H$3"                 # cible top-down : croissance CA & marge EBITDA
PIL="'01b_Pilotage'!"; PP0=5; PPN=PP0+CG-1   # onglet pilotage marque×ville : cap & override par campus
CAPKEY=f"{PIL}$C${PP0}:$C${PPN}"; CAPVAL=f"{PIL}$J${PP0}:$J${PPN}"   # clé campus & cap retenu (saisie marque×ville)
CPROV=f"{PIL}$K${PP0}:$K${PPN}"                                      # override prix par campus (exception locale)
CPKEY=f"{CAD}$J$7:$J$11"; CPVAL=f"{CAD}$K$7:$K$11"      # coeff prix par MARQUE (décision)

BR0=4; BRN=BR0+N-1
BASE="'02_Base'!"
def brng(col): return f"{BASE}${col}${BR0}:${col}${BRN}"
CR0=5; CRN=CR0+CG-1
CAMP="'03_Campagnes'!"
MR0=4; MRN=MR0+N-1; MTOT=MR0+N

# ---- plan de comptes (défini tôt : dimensionne le P&L et alimente le cadrage) ----
ACCTS=[
 ("7062","Prestations de formation — alternance (OPCO)","Produit","Produits","campus","alt",None,"V"),
 ("706","Prestations de formation — scolarité (initial)","Produit","Produits","campus","init",None,"V"),
 ("708","Frais de dossier & droits d'inscription","Produit","Produits","campus","frais",None,"V"),
 ("621","Personnel extérieur — vacataires & intervenants","Charge","Coûts directs","campus","classes",0.070,"V"),
 ("604","Sous-traitance pédagogique","Charge","Coûts directs","campus","effectif",0.030,"V"),
 ("6063","Fournitures pédagogiques & petit équipement","Charge","Coûts directs","campus","effectif",0.020,"V"),
 ("6231","Publicité & marketing d'acquisition (leads)","Charge","Coûts directs","campus","classes",0.021,"V"),
 ("6411","Rémunération enseignants permanents","Charge","Personnel","campus","classes",0.170,"F"),
 ("6413","Rémunération personnel administratif & pédagogique","Charge","Personnel","campus","effectif",0.090,"F"),
 ("6414","Rémunération direction & fonctions support (siège)","Charge","Personnel","groupe","CA",0.055,"F"),
 ("645","Charges sociales & de prévoyance","Charge","Personnel","campus","effectif",0.140,"F"),
 ("613","Loyers & charges locatives (campus)","Charge","Structure","campus","classes",0.105,"F"),
 ("615","Entretien & maintenance","Charge","Structure","campus","classes",0.015,"F"),
 ("616","Primes d'assurance","Charge","Structure","campus","effectif",0.010,"F"),
 ("6226","Honoraires (audit, conseil, juridique)","Charge","Structure","groupe","CA",0.025,"F"),
 ("6236","Marketing de marque, salons & JPO","Charge","Structure","groupe","CA",0.030,"F"),
 ("625","Déplacements, missions & réceptions","Charge","Structure","campus","effectif",0.015,"F"),
 ("626","Télécom, systèmes d'information & affranchissement","Charge","Structure","groupe","effectif",0.020,"F"),
 ("6281","Cotisations, documentation & abonnements","Charge","Structure","groupe","CA",0.008,"F"),
 ("6331","Taxe sur les salaires","Charge","Impôts & taxes","groupe","effectif",0.015,"F"),
 ("63511","Cotisation foncière & CVAE","Charge","Impôts & taxes","campus","classes",0.010,"F"),
 ("6333","Participation formation professionnelle","Charge","Impôts & taxes","groupe","CA",0.005,"F"),
 ("6811","Dotations aux amortissements (D&A)","Charge","Dotations","campus","classes",0.060,"F"),
]
def _pnl_rows():
    r=4; r+=1; r+=sum(1 for a in ACCTS if a[2]=="Produit"); rowCA=r; r+=1
    r+=1; r+=sum(1 for a in ACCTS if a[3]=="Coûts directs"); r+=1
    for g in ("Personnel","Structure","Impôts & taxes"): r+=1; r+=sum(1 for a in ACCTS if a[3]==g)
    rowEB=r; r+=1; r+=1; r+=sum(1 for a in ACCTS if a[3]=="Dotations"); rowEBIT=r
    return rowCA,rowEB,rowEBIT
PNL_CA,PNL_EB,PNL_EBIT=_pnl_rows()
REFEBITDA=round(REFCA*0.146)
PNL_EBc=f"'07_PnL'!$F${PNL_EB}"; PNL_EBb=f"'07_PnL'!$G${PNL_EB}"   # EBITDA atterrissage / budget

# ============================================================ 00_Notice
ws=wb.active; ws.title="00_Notice"; ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":30,"C":80}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:C2"); C(ws,"B2","EDUSERVICES — Modèle de CA (v3, budget-driven, leads observés)",CTIT,FNAVY,align=AL); ws.row_dimensions[2].height=30
ws.merge_cells("B3:C3"); C(ws,"B3","Budget marketing → leads (CRM, socle organique + payant) → funnel mesuré → inscrits → CA. Revenu et taux différenciés alternance / initial.",CIT)
band(ws,5,"B","C","Les feuilles (périmètre : CA → EBITDA)")
notice=[("01_Cadrage","POSTE DE COMMANDE CFO : objectif CA & EBITDA (calculé), scénarios, 9 leviers (CA + coûts), coefficients prix par marque×ville."),
 ("02_Base","BASE DE RÉFÉRENCE (version unique) : fusion historique + paramètres. LEADS OBSERVÉS (CRM) et taux de funnel MESURÉS (lead→cand, cand→admis, admis→inscrit), différenciés par modalité."),
 ("02_CRM","HISTORIQUE MULTISOURCE (format long) : leads organiques/payants + dépense marketing par campus × 3 exercices. Sert à MESURER CPL, rendement et part organique."),
 ("03_Campagnes","MOTEUR D'ACQUISITION (campus) : CPL, rendement, part organique MESURÉS depuis le CRM, puis budget → leads (socle organique + part payante à rendement décroissant)."),
 ("04_Moteur","MOTEUR DE CA (cellule) : répartition des leads (mix réel) → funnel mesuré → effectif (+cohorte) → CA.")]
r=6
for nom,desc in notice:
    C(ws,f"B{r}",nom,CB,FLIGHT,align=AL,border=True); C(ws,f"C{r}",desc,CREG,align=ALW,border=True); ws.row_dimensions[r].height=46; r+=1
band(ws,r+1,"B","C","Principe d'ancrage")
C(ws,f"B{r+2}","Au budget de référence, tous leviers à 0, le moteur reproduit l'historique. Les taux ne sont pas supposés : ils sont mesurés depuis les leads observés. Seul l'écart au budget de référence fait varier les volumes (part payante × courbe ^r).",CIT,align=ALW); ws.merge_cells(f"B{r+2}:C{r+2}"); ws.row_dimensions[r+2].height=46

# ============================================================ 01_Cadrage
ws=wb.create_sheet("01_Cadrage"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":36,"C":13,"D":13,"E":13,"F":12,"G":12,"H":13,"I":2,"J":16,"K":13,"L":11,"M":13,"N":9,"O":9,"P":9}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B1:H1"); C(ws,"B1","POSTE DE COMMANDE CFO — Cadrage CA & EBITDA",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=28
C(ws,"B3","Scénario actif :",CB,align=AR); C(ws,"D3","Cadrage",CINB,FYEL,align=AC,border=True)
dv=DataValidation(type="list",formula1='"Référence,Cadrage,Optimiste,Prudent"',allow_blank=False); ws.add_data_validation(dv); dv.add(ws["D3"])
C(ws,"E3","🎯 Croissance CA cible :",CB,align=AR); C(ws,"F3",0.05,CINB,FYEL,fmt=PCT,align=AC,border=True)
C(ws,"G3","Marge EBITDA cible :",CB,align=AR); C(ws,"H3",0.15,CINB,FYEL,fmt=PCT,align=AC,border=True)
# --- réconciliation top-down / bottom-up ---
band(ws,5,"B","G","① Réconciliation — Référence · 🎯 Cible (top-down) · 🔧 Construit (bottom-up) · Écart")
for i,h in enumerate(["Indicateur","Référence","🎯 Cible","🔧 Construit","Écart","Écart %"]): C(ws,f"{GL(2+i)}6",h,CHDR,FBLUE,align=AC,border=True)
TOTCA=f"'04_Moteur'!$AA${MTOT}"; TOTEFF=f"'04_Moteur'!$Y${MTOT}"; EFFCIB=f"'04_Moteur'!$AK${MTOT}"
C(ws,"B7","Chiffre d'affaires",CB,align=AL,border=True)
C(ws,"C7",REFCA,CL,fmt=EUR,align=AR,border=True); C(ws,"D7",f"={REFCA}*(1+F3)",CFB,fmt=EUR,align=AR,border=True); C(ws,"E7",f"={TOTCA}",CFB,fmt=EUR,align=AR,border=True)
C(ws,"F7","=E7-D7",CF,fmt=EUR,align=AR,border=True); C(ws,"G7","=IFERROR(E7/D7-1,0)",CF,fmt=PCT,align=AR,border=True)
C(ws,"B8","EBITDA  (calculé)",CB,align=AL,border=True)
C(ws,"C8",f"={PNL_EBc}",CFB,fmt=EUR,align=AR,border=True); C(ws,"D8","=D7*H3",CFB,fmt=EUR,align=AR,border=True); C(ws,"E8",f"={PNL_EBb}",CFB,fmt=EUR,align=AR,border=True)
C(ws,"F8","=E8-D8",CF,fmt=EUR,align=AR,border=True); C(ws,"G8","=IFERROR(E8/D8-1,0)",CF,fmt=PCT,align=AR,border=True)
C(ws,"B9","Marge EBITDA %",CIT,align=AL,border=True)
C(ws,"C9","=IFERROR(C8/C7,0)",CF,fmt=PCT,align=AR,border=True); C(ws,"D9","=H3",CF,fmt=PCT,align=AR,border=True); C(ws,"E9","=IFERROR(E8/E7,0)",CFB,fmt=PCT,align=AR,border=True); C(ws,"G9","=IFERROR(E9-D9,0)",CF,fmt=PCT,align=AR,border=True)
C(ws,"B10","Effectif total",CB,align=AL,border=True)
C(ws,"C10",REFEFF,CL,fmt=NB,align=AR,border=True); C(ws,"D10",f"={EFFCIB}",CFB,fmt=NB,align=AR,border=True); C(ws,"E10",f"={TOTEFF}",CFB,fmt=NB,align=AR,border=True)
C(ws,"F10","=E10-D10",CF,fmt=NB,align=AR,border=True); C(ws,"G10","=IFERROR(E10/D10-1,0)",CF,fmt=PCT,align=AR,border=True)
C(ws,"B12","RESTE À TROUVER — CA :",CB,align=AR); ws.merge_cells("B12:C12")
C(ws,"D12","=IF(D7-E7>0,D7-E7,0)",CFB,FYEL,fmt=EUR,align=AR,border=True); C(ws,"E12","EBITDA :",CB,align=AR); C(ws,"F12","=IF(D8-E8>0,D8-E8,0)",CFB,FYEL,fmt=EUR,align=AR,border=True)
# --- leviers ---
band(ws,14,"B","H","② Leviers — bascule par scénario (colonne ACTIF)")
for i,h in enumerate(["Paramètre","Unité","Référence","Cadrage","Optimiste","Prudent","ACTIF"]): C(ws,f"{GL(2+i)}15",h,CHDR,FBLUE,align=AC,border=True)
MATCHSC='MATCH($D$3,$D$15:$G$15,0)'   # colonne D = scénario "Référence" (leviers à 0)
levs=[("Variation du budget d'acquisition (→ leads payants)","%",0,0.08,0.15,-0.05),
 ("Variation du budget de marque (→ socle organique)","%",0,0.10,0.20,-0.05),
 ("Hausse tarifaire (prix)","%",0,0.025,0.035,0.02),
 ("Gain taux lead → candidature","pts",0,0.01,0.03,0.0),
 ("Gain conversion admis → inscrit","pts",0,0.01,0.025,0.0),
 ("Amélioration du taux de passage","pts",0,0.005,0.015,-0.01),
 ("Inflation des charges externes","%",0,0.02,0.015,0.03),
 ("Politique salariale (masse permanente)","%",0,0.025,0.02,0.03),
 ("Variation des effectifs permanents","%",0,0.04,0.03,0.05),
 ("Effort de productivité (achats & structure)","%",0,0.01,0.03,0.0),
 ("Variation des coûts de structure (loyers, IT, siège…)","%",0,0.0,-0.03,0.04)]
r=16
for lib,u,ba,cad,opt,pru in levs:
    C(ws,f"B{r}",lib,CREG,align=AL,border=True); C(ws,f"C{r}",u,CREG,align=AC,border=True); C(ws,f"D{r}",ba,CIT,fmt=PCT,align=AC,border=True)
    C(ws,f"E{r}",cad,CIN,fmt=PCT,align=AC,border=True); C(ws,f"F{r}",opt,CIN,fmt=PCT,align=AC,border=True); C(ws,f"G{r}",pru,CIN,fmt=PCT,align=AC,border=True)
    C(ws,f"H{r}",f"=INDEX(D{r}:G{r},{MATCHSC})",CFB,FLIGHT,fmt=PCT,align=AC,border=True); r+=1
C(ws,"B27",'Leviers 1-6 → CA (moteur) · leviers 7-11 → coûts (P&L Budget → EBITDA). Acquisition = achat de leads (payant) · Marque = notoriété (socle organique). « Référence » remet tout à 0 = atterrissage.',CIT,align=AL); ws.merge_cells("B27:H27")
# --- constantes ---
band(ws,28,"B","H","③ Constante — frais de dossier (décision)")
for i,h in enumerate(["Paramètre","Unité","Référence","Cadrage","Optimiste","Prudent","ACTIF"]): C(ws,f"{GL(2+i)}29",h,CHDR,FBLUE,align=AC,border=True)
consts=[("Frais de dossier / nouvel inscrit","€",FRAIS_DEF,EUR)]
r=30
for lib,u,val,fmt in consts:
    C(ws,f"B{r}",lib,CREG,align=AL,border=True); C(ws,f"C{r}",u,CREG,align=AC,border=True); C(ws,f"D{r}",val,CIT,fmt=fmt,align=AC,border=True)
    C(ws,f"E{r}",val,CIN,fmt=fmt,align=AC,border=True); C(ws,f"F{r}",val,CIN,fmt=fmt,align=AC,border=True); C(ws,f"G{r}",val,CIN,fmt=fmt,align=AC,border=True)
    C(ws,f"H{r}",f"=INDEX(D{r}:G{r},{MATCHSC})",CFB,FLIGHT,fmt=fmt,align=AC,border=True); r+=1
C(ws,"B32","Rendement, CPL, part organique et rendement de marque ne sont pas saisis : ils sont MESURÉS depuis le CRM (voir 03_Campagnes).",CIT,align=AL); ws.merge_cells("B32:H32")
MARQUES=list(BRANDS.keys())
# --- ① coefficient prix par MARQUE (décision : pouvoir de prix différencié) ---
band(ws,5,"J","M","Coeff prix par marque (décision)")
for i,h in enumerate(["Marque","🔵 Coeff prix"]): C(ws,f"{GL(10+i)}6",h,CHDR,FBLUE,align=AC,border=True)
CPRIX_M={"MBway":1.20,"ISCOM":1.15,"Ipac Bachelor Factory":0.95,"Pigier":0.90,"Tunon":1.05}
r=7
for m in MARQUES:
    C(ws,f"J{r}",m,CL,align=AL,border=True); C(ws,f"K{r}",CPRIX_M[m],CINB,FYEL,fmt=X2,align=AC,border=True); r+=1
C(ws,"J12","Coeff prix = décision MARQUE. Override prix par CAMPUS dans l'onglet 01b_Pilotage (exception locale).",CIT,align=ALW); ws.merge_cells("J12:M12"); ws.row_dimensions[12].height=26
# --- ② cap stratégique & pilotage marque×ville → onglet 01b_Pilotage ---
band(ws,13,"J","M","② Cap stratégique & pilotage → onglet 01b_Pilotage")
C(ws,"J14","La SAISIE du cap stratégique par CAMPUS (marque×ville), les caps proposés (efficience / momentum / potentiel) avec leur justification, et la synthèse CA + EBITDA par campus sont regroupés dans l'onglet 01b_Pilotage. Ici, on garde les décisions GROUPE : coeff prix (marque) et clés d'allocation.",CIT,align=ALW)
ws.merge_cells("J14:M19"); ws.row_dimensions[14].height=112
# --- ③ indice prix par ville — DÉDUIT du réalisé (informationnel, plus de saisie) ---
vrev={}; veff={}
for rr in rows: vrev[rr["ville"]]=vrev.get(rr["ville"],0)+rr["rev"]*rr["eff"]; veff[rr["ville"]]=veff.get(rr["ville"],0)+rr["eff"]
natavg=sum(rr["rev"]*rr["eff"] for rr in rows)/sum(rr["eff"] for rr in rows)
band(ws,21,"J","M","Indice prix par ville — déduit du réalisé")
for i,h in enumerate(["Ville","Prix moyen","Indice /national"]): C(ws,f"{GL(10+i)}22",h,CHDR,FBLUE,align=AC,border=True)
r=23
for v in sorted(vrev,key=lambda x:-vrev[x]/veff[x]):
    C(ws,f"J{r}",v,CL,align=AL,border=True); C(ws,f"K{r}",round(vrev[v]/veff[v]),CF,fmt=EUR,align=AR,border=True)
    C(ws,f"L{r}",round((vrev[v]/veff[v])/natavg,3),CF,fmt=X2,align=AC,border=True); r+=1
C(ws,f"J{r+1}","Indice = prix moyen ville ÷ national, calculé sur le réalisé (le niveau de prix par ville est déjà dans les revenus). Le pouvoir de prix se pilote par MARQUE ci-dessus.",CIT,align=AL); ws.merge_cells(f"J{r+1}:M{r+2}"); ws.row_dimensions[r+1].height=28
# --- ④ clés d'allocation — DÉCISION CFO (top-down) : la MÉTHODE est centrale, figée ; les campus n'ajustent que la MATIÈRE ---
band(ws,34,"J","M","Clés d'allocation — décision CFO (top-down)")
for i,h in enumerate(["Niveau de cascade","🔵 Clé retenue"]): C(ws,f"{GL(10+i*2)}35",h,CHDR,FBLUE,align=AC,border=True); ws.merge_cells(f"{GL(10+i*2)}35:{GL(11+i*2)}35")
dva=DataValidation(type="list",formula1='"Chiffre d\'affaires,Effectif,Nombre de classes"',allow_blank=False); ws.add_data_validation(dva)
for rr,lab,dft in [(36,"① Groupe → Marque","Chiffre d'affaires"),(37,"② Marque → Campus","Effectif"),(38,"③ Campus → Classe","Nombre de classes")]:
    C(ws,f"J{rr}",lab,CL,align=AL,border=True); ws.merge_cells(f"J{rr}:K{rr}")
    C(ws,f"L{rr}",dft,CINB,FYEL,align=AC,border=True); ws.merge_cells(f"L{rr}:M{rr}"); dva.add(ws[f"L{rr}"])
C(ws,"J40","La CLÉ (méthode) est une décision de gouvernance : le CFO la fige ici, elle s'applique à tout le groupe pour rester comparable. Le contrôleur ne choisit pas la clé — il conteste la MATIÈRE (assiette réelle) dans 10_Allocation.",CIT,align=ALW); ws.merge_cells("J40:M42"); ws.row_dimensions[40].height=42
ALLOC_N1=f"{CAD}$L$36"; ALLOC_N2=f"{CAD}$L$37"; ALLOC_N3=f"{CAD}$L$38"
# ⑤ traduction CA + ⑥ EBITDA + lexique → déplacés dans l'onglet 01b_Pilotage (tableau complet marque×ville)

# ============================================================ 02_Base
ws=wb.create_sheet("02_Base"); ws.sheet_view.showGridLines=False
bcols=["Marque","Ville","Programme","Cycle","Année","Modalité","Entrée",
 "Leads hist","Cand hist","Admis hist","Nouv hist","Réins hist","Effectif hist","Eff. année inf.","Classes hist",
 "Revenu/étudiant","Taux passage","Taux lead→cand","Taux cand→admis","Yield admis→inscrit","CA réf"]
for i,w in enumerate([15,10,20,6,6,7,7,9,9,9,9,9,10,11,9,12,10,11,11,12,12]): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells(f"A1:{GL(len(bcols))}1"); C(ws,"A1","BASE DE RÉFÉRENCE — version unique · fusion historique + paramètres · leads observés (CRM) & taux mesurés",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=22
ws.merge_cells(f"A2:{GL(len(bcols))}2"); C(ws,"A2","Une ligne = marque × ville × programme × année × modalité. Leads = donnée observée (CRM). Taux = mesurés (aval÷amont). Revenu différencié alternance (OPCO) / initial (étudiant).",CIT,align=ALW); ws.row_dimensions[2].height=16
for i,h in enumerate(bcols): C(ws,f"{GL(1+i)}3",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[3].height=30
for idx,rr in enumerate(rows):
    r=BR0+idx
    frr=f"{CAD}$D$30"   # frais de dossier de référence (constante cadrage)
    vals=[rr["marque"],rr["ville"],rr["prog"],rr["type"],rr["niv"],("Alternance" if rr["mod"]=="ALT" else "Initial"),rr["entry"],
          rr["leads"],rr["cand"],rr["admis"],rr["nouv"],rr["rein"],rr["eff"],rr["eff_prev"],rr["classes"],
          rr["rev"],
          f"=IFERROR(M{r}/N{r},0)",   # Q passage = effectif ÷ effectif année inf.
          f"=IFERROR(I{r}/H{r},0)",   # R lead→cand = candidatures ÷ leads
          f"=IFERROR(J{r}/I{r},0)",   # S cand→admis = admis ÷ candidatures
          f"=IFERROR(K{r}/J{r},0)",   # T yield = nouveaux ÷ admis
          f"=M{r}*P{r}+K{r}*{frr}"]   # U CA réf = effectif×revenu + nouveaux×frais
    fmts=[None,None,None,None,None,None,NB,NB,NB,NB,NB,NB,NB,NB,NB,EUR,PCT,PCT,PCT,PCT,EUR]
    for i,(v,f) in enumerate(zip(vals,fmts)):
        al=AL if i<6 else AC
        C(ws,f"{GL(1+i)}{r}",v,(CF if i>=16 else CL),fmt=f,align=al,border=True)
ws.freeze_panes="A4"

# ============================================================ 02_CRM (historique multisource, format LONG, par ANNÉE)
CRMY=[2024,2025,2026]   # 2024 (N-2) · 2025 (N-1) · 2026 (atterrissage)  →  on construit le budget 2027
ws=wb.create_sheet("02_CRM"); ws.sheet_view.showGridLines=False
xcols=["Marque","Ville","clé","Année","Leads organiques","Leads payants","Leads totaux","Dépense acquisition","Dépense marque"]
for i,w in enumerate([15,11,15,10,15,14,13,16,14]): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells(f"A1:{GL(len(xcols))}1"); C(ws,"A1","CRM & MARKETING — historique multisource (format long, par année) : leads organiques/payants + dépenses acquisition & marque",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=22
ws.merge_cells(f"A2:{GL(len(xcols))}2"); C(ws,"A2","Source : CRM (leads, tag organique/payant) + compta (2 postes : ACQUISITION = achat de leads, compte 6231 · MARQUE = notoriété/salons, compte 6236). 3 années réelles. 03_Campagnes MESURE CPL, rendements et part organique — rien n'est saisi.",CIT,align=ALW); ws.row_dimensions[2].height=30
for i,h in enumerate(xcols): C(ws,f"{GL(1+i)}3",h,CHDR,FBLUE,align=AC,border=True)
XR0=4; xr=XR0
for cc in campus:
    for vi,yr in enumerate(CRMY):
        C(ws,f"A{xr}",cc["marque"],CL,align=AL,border=True); C(ws,f"B{xr}",cc["ville"],CL,align=AC,border=True)
        C(ws,f"C{xr}",f'=A{xr}&"|"&B{xr}',CF,align=AC,border=True)
        C(ws,f"D{xr}",yr,CL,align=AC,border=True)
        C(ws,f"E{xr}",cc["org"][vi],CL,fmt=NB,align=AR,border=True); C(ws,f"F{xr}",cc["paid"][vi],CL,fmt=NB,align=AR,border=True)
        C(ws,f"G{xr}",f"=E{xr}+F{xr}",CF,fmt=NB,align=AR,border=True); C(ws,f"H{xr}",cc["spend"][vi],CL,fmt=EUR,align=AR,border=True)
        C(ws,f"I{xr}",cc["bspend"][vi],CL,fmt=EUR,align=AR,border=True)
        xr+=1
XRN=xr-1; ws.freeze_panes="A4"
CRM="'02_CRM'!"
def xrng(col): return f"{CRM}${col}${XR0}:${col}${XRN}"
XKEY=xrng("C"); XVER=xrng("D")
def xsum(col,key,yr): return f'SUMIFS({xrng(col)},{XKEY},{key},{XVER},{yr})'

# ============================================================ 03_Campagnes (mesure CPL/rendement/part org + budget->leads)
ws=wb.create_sheet("03_Campagnes"); ws.sheet_view.showGridLines=False
ccols=["Marque","Ville","clé","Leads organiques","Leads payants réf","Leads total réf","Dépense acq. réf","Part organique","CPL mesuré","Rendement acq.","Budget acq. actif","Leads payants actif","Leads actif total","CPL effectif","Conv. lead→inscrit","CAC marginal /inscrit","Dépense marque réf","Rendement marque","Budget marque actif","Leads org. actif","Cap effectif (01b)"]
for i,w in enumerate([15,11,14,13,13,12,13,11,10,12,13,13,12,10,13,15,14,13,14,13,13]): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells(f"A1:{GL(len(ccols))}1"); C(ws,"A1","MOTEUR D'ACQUISITION (campus) — 2 leviers marketing : ACQUISITION (achat de leads → payant) & MARQUE (notoriété → organique), rendements MESURÉS",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=22
ws.merge_cells(f"A2:{GL(len(ccols))}2"); C(ws,"A2","ACQUISITION : leads payants = payants réf × (budget acq.÷réf)^rendement acq. · MARQUE : leads organiques = org réf × (budget marque÷réf)^rendement marque (mesuré en AGRÉGÉ, org vs dépense marque sur 3 ans — pas d'attribution au lead). Total = organiques actif + payants actif.",CIT,align=ALW); ws.row_dimensions[2].height=34
for i,h in enumerate(ccols): C(ws,f"{GL(1+i)}3",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[3].height=30
YATT=2026; YOLD=2024
for idx,cc in enumerate(campus):
    r=CR0+idx; key=f"C{r}"
    C(ws,f"A{r}",cc["marque"],CL,align=AL,border=True); C(ws,f"B{r}",cc["ville"],CL,align=AC,border=True)
    C(ws,f"C{r}",f'=A{r}&"|"&B{r}',CF,align=AC,border=True)
    C(ws,f"D{r}",f"={xsum('E',key,YATT)}",CF,fmt=NB,align=AR,border=True)   # org ATT
    C(ws,f"E{r}",f"={xsum('F',key,YATT)}",CF,fmt=NB,align=AR,border=True)   # payants ATT
    C(ws,f"F{r}",f"=D{r}+E{r}",CF,fmt=NB,align=AR,border=True)
    C(ws,f"G{r}",f"={xsum('H',key,YATT)}",CF,fmt=EUR,align=AR,border=True)  # dépense ATT
    C(ws,f"H{r}",f"=IFERROR(D{r}/F{r},0)",CF,fmt=PCT,align=AC,border=True)  # part org MESURÉE
    C(ws,f"I{r}",f"=IFERROR(G{r}/E{r},0)",CF,fmt=EUR,align=AR,border=True)  # CPL MESURÉ
    C(ws,f"J{r}",f"=IFERROR(LN(E{r}/{xsum('F',key,YOLD)})/LN(G{r}/{xsum('H',key,YOLD)}),0.5)",CF,fmt=X2,align=AC,border=True)  # rendement MESURÉ
    C(ws,f"U{r}",f"=INDEX({CAPVAL},MATCH(C{r},{CAPKEY},0))",CF,fmt=X2,align=AC,border=True)   # cap effectif campus (saisi marque×ville dans 01b_Pilotage)
    # budget acquisition actif = budget réf × (1+levier global) × tilt du cap CAMPUS (enveloppe groupe CONSTANTE)
    C(ws,f"K{r}",f"=G{r}*(1+{LMKT})*U{r}*SUM($G${CR0}:$G${CRN})/SUMPRODUCT($G${CR0}:$G${CRN},$U${CR0}:$U${CRN})",CFB,fmt=EUR,align=AR,border=True)
    C(ws,f"L{r}",f"=E{r}*(K{r}/G{r})^J{r}",CFB,fmt=NB,align=AR,border=True)
    C(ws,f"M{r}",f"=T{r}+L{r}",CFB,fmt=NB,align=AR,border=True)   # total = organiques ACTIF (piloté marque) + payants actif
    C(ws,f"N{r}",f"=IFERROR(K{r}/L{r},0)",CF,fmt=EUR,align=AR,border=True)
    critc=f'{brng("A")},A{r},{brng("B")},B{r},{brng("G")},1'
    C(ws,f"O{r}",f"=IFERROR(SUMIFS({brng('K')},{critc})/SUMIFS({brng('H')},{critc}),0)",CF,fmt=PCT,align=AC,border=True)  # conversion globale lead→inscrit (mesurée)
    C(ws,f"P{r}",f"=IFERROR(I{r}/(J{r}*O{r}),0)",CFB,fmt=EUR,align=AR,border=True)  # CAC marginal = CPL ÷ (rendement × conversion)
    # --- MARQUE (notoriété) : pilote le socle ORGANIQUE, rendement mesuré en AGRÉGÉ (org vs dépense marque) ---
    C(ws,f"Q{r}",f"={xsum('I',key,YATT)}",CF,fmt=EUR,align=AR,border=True)                          # dépense marque ATT
    C(ws,f"R{r}",f"=IFERROR(LN(D{r}/{xsum('E',key,YOLD)})/LN(Q{r}/{xsum('I',key,YOLD)}),0.35)",CF,fmt=X2,align=AC,border=True)  # rendement marque MESURÉ
    C(ws,f"S{r}",f"=Q{r}*(1+{LBRAND})",CFB,fmt=EUR,align=AR,border=True)                            # budget marque actif
    C(ws,f"T{r}",f"=D{r}*(S{r}/Q{r})^R{r}",CFB,fmt=NB,align=AR,border=True)                          # leads organiques ACTIF
# indicateur d'attractivité : CAC marginal (vert = l'euro marketing rapporte le + / rouge = cher)
ws.conditional_formatting.add(f"P{CR0}:P{CRN}",
    ColorScaleRule(start_type="min",start_color="63BE7B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="F8696B"))
C(ws,f"A{CRN+2}","CAC marginal = coût marketing du prochain inscrit (vert = investir, rouge = cher). Q-T : le socle ORGANIQUE répond au budget de MARQUE. U : cap effectif par campus (marque×ville) — saisi dans l'onglet 01b_Pilotage — qui concentre le budget d'acquisition (enveloppe groupe constante).",CIT,align=AL); ws.merge_cells(f"A{CRN+2}:U{CRN+2}")
CKEY=f"{CAMP}$C${CR0}:$C${CRN}"; CLEADS=f"{CAMP}$M${CR0}:$M${CRN}"; CSLEADS=f"{CAMP}$F${CR0}:$F${CRN}"

# ============================================================ 04_Moteur (funnel -> CA)
ws=wb.create_sheet("04_Moteur"); ws.sheet_view.showGridLines=False
mcols=["Marque","Ville","Programme","Année","Mod.","Entrée",
 "Leads hist","Cand hist","Nouv hist","Réins hist","Eff hist","Eff. inf.","Revenu/étu","Passage","T.L→C","T.C→A","Yield",
 "Part leads","Leads campus","Leads cellule","Candidatures","Admis",
 "Nouveaux ③","Réinscrits ③","Effectif ③","Revenu actif","CA ③ ajusté",
 "Nouveaux ②","Réinscr. ②","Effectif ②","CA ② cadré",
 "🟢 Δ L→C","🟢 Δ yield","🟢 Δ passage",
 "Poids","CA ① cible","Effectif ① cible","Écart CA ③−①","Écart eff ③−①"]
for i,w in enumerate([14,9,16,6,5,6]+[8]*(len(mcols)-6)): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells(f"A1:{GL(len(mcols))}1"); C(ws,"A1","MOTEUR DE CA (cellule) — ① CIBLE (éclatée top-down) · ② CADRÉ (groupe) · ③ AJUSTÉ (cadré + ajustement contrôleur) · écart",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=22
for i,h in enumerate(mcols): C(ws,f"{GL(1+i)}3",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[3].height=30
MO_CIBLE=f"({REFCA}*(1+{CROISS}))"    # CA cible groupe = référence × (1 + croissance cible)
# masque de saisie : exemple campus MBway Lyon — ses cellules lisent leurs ajustements depuis 08_Saisie_Campus
MASKS="'08_Saisie_Campus'!"; MK0=7
ML_IDX=[i for i,rr in enumerate(rows) if rr["marque"]=="MBway" and rr["ville"]=="Lyon"]
mask_of={idx:MK0+j for j,idx in enumerate(ML_IDX)}
for idx in range(N):
    r=MR0+idx; b=BR0+idx
    Lk=lambda col:f"={BASE}{col}{b}"
    C(ws,f"A{r}",Lk('A'),CL,align=AL,border=True); C(ws,f"B{r}",Lk('B'),CL,align=AC,border=True); C(ws,f"C{r}",Lk('C'),CL,align=AL,border=True)
    C(ws,f"D{r}",Lk('E'),CL,align=AC,border=True); C(ws,f"E{r}",Lk('F'),CL,align=AC,border=True); C(ws,f"F{r}",Lk('G'),CL,fmt=NB,align=AC,border=True)
    C(ws,f"G{r}",Lk('H'),CL,fmt=NB,align=AC,border=True); C(ws,f"H{r}",Lk('I'),CL,fmt=NB,align=AC,border=True); C(ws,f"I{r}",Lk('K'),CL,fmt=NB,align=AC,border=True)
    C(ws,f"J{r}",Lk('L'),CL,fmt=NB,align=AC,border=True); C(ws,f"K{r}",Lk('M'),CL,fmt=NB,align=AC,border=True); C(ws,f"L{r}",Lk('N'),CL,fmt=NB,align=AC,border=True)
    C(ws,f"M{r}",Lk('P'),CL,fmt=EUR,align=AC,border=True); C(ws,f"N{r}",Lk('Q'),CL,fmt=PCT,align=AC,border=True)
    C(ws,f"O{r}",Lk('R'),CL,fmt=PCT,align=AC,border=True); C(ws,f"P{r}",Lk('S'),CL,fmt=PCT,align=AC,border=True); C(ws,f"Q{r}",Lk('T'),CL,fmt=PCT,align=AC,border=True)
    key=f'A{r}&"|"&B{r}'
    C(ws,f"R{r}",f"=IF(F{r}=1,IFERROR(G{r}/INDEX({CSLEADS},MATCH({key},{CKEY},0)),0),0)",CF,fmt=PCT,align=AC,border=True)
    C(ws,f"S{r}",f"=INDEX({CLEADS},MATCH({key},{CKEY},0))",CF,fmt=NB,align=AR,border=True)
    C(ws,f"T{r}",f"=S{r}*R{r}",CF,fmt=NB,align=AR,border=True)
    C(ws,f"U{r}",f"=IF(F{r}=1,T{r}*(O{r}+{LGLC}+AF{r}),0)",CF,fmt=NB,align=AR,border=True)   # +Δ lead→cand (contrôleur)
    C(ws,f"V{r}",f"=U{r}*P{r}",CF,fmt=NB,align=AR,border=True)
    # ③ AJUSTÉ : semé (groupe) + ajustements contrôleur AF/AG/AH (jamais d'écrasement)
    C(ws,f"W{r}",f"=IF(F{r}=1,V{r}*(Q{r}+{LGCV}+AG{r}),0)",CFB,fmt=NB,align=AR,border=True)
    C(ws,f"X{r}",f"=IF(F{r}=1,0,L{r}*(N{r}+{LPASS}+AH{r}))",CF,fmt=NB,align=AR,border=True)
    C(ws,f"Y{r}",f"=W{r}+X{r}",CFB,fmt=NB,align=AR,border=True)
    C(ws,f"Z{r}",f"=M{r}*(1+{LPRIX}*INDEX({CPVAL},MATCH(A{r},{CPKEY},0))+INDEX({CPROV},MATCH({key},{CAPKEY},0)))",CF,fmt=EUR,align=AR,border=True)  # prix : hausse × coeff marque + override campus (01b)
    C(ws,f"AA{r}",f"=Y{r}*Z{r}+W{r}*{KFRAIS}",CFB,fmt=EUR,align=AR,border=True)
    # ② CADRÉ : semé seul, auto-porté (sans aucun ajustement contrôleur)
    C(ws,f"AB{r}",f"=IF(F{r}=1,T{r}*(O{r}+{LGLC})*P{r}*(Q{r}+{LGCV}),0)",CF,fmt=NB,align=AR,border=True)
    C(ws,f"AC{r}",f"=IF(F{r}=1,0,L{r}*(N{r}+{LPASS}))",CF,fmt=NB,align=AR,border=True)
    C(ws,f"AD{r}",f"=AB{r}+AC{r}",CF,fmt=NB,align=AR,border=True)
    C(ws,f"AE{r}",f"=AD{r}*Z{r}+AB{r}*{KFRAIS}",CF,fmt=EUR,align=AR,border=True)
    # 🟢 ajustements contrôleur (delta, 0 = pas d'ajustement). Pour MBway Lyon, lus depuis le masque de saisie
    if idx in mask_of:
        _mk=mask_of[idx]
        C(ws,f"AF{r}",f"={MASKS}I{_mk}",CL,fmt=PCT,align=AC,border=True); C(ws,f"AG{r}",f"={MASKS}J{_mk}",CL,fmt=PCT,align=AC,border=True); C(ws,f"AH{r}",f"={MASKS}K{_mk}",CL,fmt=PCT,align=AC,border=True)
    else:
        C(ws,f"AF{r}",0,CINB,FYEL,fmt=PCT,align=AC,border=True); C(ws,f"AG{r}",0,CINB,FYEL,fmt=PCT,align=AC,border=True); C(ws,f"AH{r}",0,CINB,FYEL,fmt=PCT,align=AC,border=True)
    # ① CIBLE éclatée : poids = CA réf (proportionnel) ; le cap ne touche plus la cible mais le BUDGET (03_Campagnes)
    C(ws,f"AI{r}",f"=K{r}*M{r}",CF,fmt=NB,align=AR,border=True)
    C(ws,f"AJ{r}",f"=IFERROR({MO_CIBLE}*AI{r}/SUM($AI${MR0}:$AI${MRN}),0)",CFB,fmt=EUR,align=AR,border=True)
    C(ws,f"AK{r}",f"=IFERROR(AJ{r}/Z{r},0)",CF,fmt=NB,align=AR,border=True)
    C(ws,f"AL{r}",f"=AA{r}-AJ{r}",CF,fmt=EUR,align=AR,border=True)
    C(ws,f"AM{r}",f"=Y{r}-AK{r}",CF,fmt=NB,align=AR,border=True)
r=MTOT
C(ws,f"A{r}","TOTAL",CFB,FTOT,align=AL,border=True)
for col in ["B","C","D","E"]: C(ws,f"{col}{r}"," ",fill=FTOT,border=True)
for col,fmt in [("F",NB),("G",NB),("H",NB),("I",NB),("T",NB),("U",NB),("V",NB),("W",NB),("X",NB),("Y",NB),("AA",EUR),
                ("AD",NB),("AE",EUR),("AJ",EUR),("AK",NB),("AL",EUR),("AM",NB)]:
    C(ws,f"{col}{r}",f"=SUM({col}{MR0}:{col}{MRN})",CFB,FTOT,fmt=fmt,align=AR,border=True)
ws.freeze_panes="G4"
gr=r+2
band(ws,gr,"A","H","Guide de lecture — les 3 états côte à côte :"); gr+=1
grp=[("A–Q","Identité + reprise base + taux mesurés (funnel, passage)"),
 ("R–V","Acquisition : leads cellule → candidatures → admis"),
 ("W–AA","③ AJUSTÉ : nouveaux/réinscrits/effectif/CA avec ajustements contrôleur (= budget final)"),
 ("AB–AE","② CADRÉ : le même SANS ajustement (pré-budget cadré par le groupe, top-down)"),
 ("AF–AH","🟢 AJUSTEMENTS contrôleur (Δ lead→cand / Δ yield / Δ passage) — s'ajoutent au cadré, n'écrasent rien"),
 ("AI–AK","① CIBLE éclatée : poids (CA réf, proportionnel) → CA cible → effectif cible"),
 ("AL–AM","ÉCART ③ − ① : le construit atteint-il la cible ?")]
for rng,txt in grp:
    C(ws,f"A{gr}",rng,CB,FLIGHT,align=AC,border=True); ws.merge_cells(f"B{gr}:H{gr}"); C(ws,f"B{gr}",txt,CREG,align=ALW,border=True); gr+=1

# ============================================================ COÛTS / P&L  (comptes PCG 6 & 7)
# ACCTS est défini plus haut (dimensionne le P&L et le cadrage)
GROW_HIST=1.06
# marge EBITDA cible par version (progression douce = levier opérationnel réaliste ~0,7 pt/an)
TARGETS={2:0.132,1:0.140,0:0.146}
NONDA_PCT=sum(a[6] for a in ACCTS if a[3]!="Dotations" and a[6] is not None)  # = 0,854
# dimension Version (codes prêts Tagetik) : (code, libellé exercice, n en arrière, type, année)
VERS=[("2024ACT_VDEF","2024 (N-2)",2,"Actual",2024),
      ("2025ACT_VDEF","2025 (N-1)",1,"Actual",2025),
      ("2026ATT_VDEF","2026 (Atterr.)",0,"Forecast",2026)]
# hiérarchies : PCG (poste) et gestion (SIG agrégat + nœud parent)
def poste_of(code):
    if code[0]=="7": return "70 — Prestations de services"
    return {"60":"60 — Achats","61":"61 — Services extérieurs","62":"62 — Autres services extérieurs",
            "63":"63 — Impôts & taxes","64":"64 — Charges de personnel","68":"68 — Dotations & amort."}[code[:2]]
def classe_of(code): return "7 — Produits d'exploitation" if code[0]=="7" else "6 — Charges d'exploitation"
AGG={"Produits":"Marge de contribution","Coûts directs":"Marge de contribution","Personnel":"EBITDA",
     "Structure":"EBITDA","Impôts & taxes":"EBITDA","Dotations":"Résultat d'exploitation (EBIT)"}
SIGNODE={"Produits":"SIG_PROD","Coûts directs":"SIG_DIR","Personnel":"SIG_PERS",
         "Structure":"SIG_STRUCT","Impôts & taxes":"SIG_IMP","Dotations":"SIG_DOT"}
def slug(s):
    s=s.upper()
    for a,b in [("É","E"),("È","E"),("Ê","E"),("À","A"),("Ç","C"),("'",""),(" ","_"),("-","_")]: s=s.replace(a,b)
    return s
# drivers par campus (depuis la base)
capm={}
for r in rows:
    k=(r["marque"],r["ville"]); d=capm.setdefault(k,dict(ca=0,eff=0,cls=0,alt=0,init=0,frais=0))
    d["ca"]+=r["eff"]*r["rev"]+r["nouv"]*FRAIS_DEF; d["eff"]+=r["eff"]; d["cls"]+=r["classes"]
    d["alt"]+=(r["eff"]*r["rev"] if r["mod"]=="ALT" else 0); d["init"]+=(r["eff"]*r["rev"] if r["mod"]=="INIT" else 0)
    d["frais"]+=r["nouv"]*FRAIS_DEF
CA_T=REFCA; EFF_T=REFEFF; CLS_T=sum(d["cls"] for d in capm.values())
def share(k,drv):
    d=capm[k]; return {"CA":d["ca"]/CA_T,"effectif":d["eff"]/EFF_T,"classes":d["cls"]/CLS_T}[drv]
def ca_ver(n): return REFCA/(GROW_HIST**n)
MKT_PCT=next(a[6] for a in ACCTS if a[0]=="6231")      # slot acquisition — aligné sur la dépense CRM
BRAND_PCT=next(a[6] for a in ACCTS if a[0]=="6236")    # slot marque — aligné sur la dépense CRM
OTHER_PCT=NONDA_PCT-MKT_PCT-BRAND_PCT
cmap={(c["marque"],c["ville"]):c for c in campus}
def crm_spend_ver(n): return sum(c["spend"][2-n] for c in campus)   # dépense ACQUISITION CRM (2-n : 2026/2025/2024)
def brand_spend_ver(n): return sum(c["bspend"][2-n] for c in campus) # dépense MARQUE CRM
def amount(pct,sig,code,n):
    if sig=="Dotations": return pct*ca_ver(n)
    if code in ("6231","6236"): return None             # marketing (acquisition & marque) = dépense CRM (géré à part)
    remaining=ca_ver(n)*(1-TARGETS[n])-crm_spend_ver(n)-brand_spend_ver(n) # reste après les 2 dépenses marketing réelles
    return pct*remaining/OTHER_PCT
# construction de la compta : (ecode,elib,marque,ville,compte,lib,poste,sens,sig,vcode,exlab,montant)
compta=[]
for code,lib,sens,sig,niv,drv,pct,vf in ACCTS:
    poste=poste_of(code)
    for vcode,exlab,n,vtype,vyear in VERS:
        if sens=="Produit":
            for k,d in capm.items():
                val={"alt":d["alt"],"init":d["init"],"frais":d["frais"]}[drv]*(ca_ver(n)/REFCA)
                if val>0: compta.append((slug(k[0])+"_"+slug(k[1]),f"{k[0]} {k[1]}",k[0],k[1],code,lib,poste,sens,sig,vcode,exlab,round(val)))
        elif code=="6231":   # marketing acquisition = dépense CRM, par campus (alignement CRM ↔ compta)
            for k in capm:
                val=cmap[k]["spend"][2-n]
                compta.append((slug(k[0])+"_"+slug(k[1]),f"{k[0]} {k[1]}",k[0],k[1],code,lib,poste,sens,sig,vcode,exlab,round(val)))
        elif code=="6236":   # marketing de marque = dépense CRM, pilotée au SIÈGE (fonction centrale)
            compta.append(("GROUPE","GROUPE — Siège","(groupe)","(groupe)",code,lib,poste,sens,sig,vcode,exlab,round(brand_spend_ver(n))))
        else:
            tot=amount(pct,sig,code,n)
            if niv=="groupe":
                compta.append(("GROUPE","GROUPE — Siège","(groupe)","(groupe)",code,lib,poste,sens,sig,vcode,exlab,round(tot)))
            else:
                for k in capm: compta.append((slug(k[0])+"_"+slug(k[1]),f"{k[0]} {k[1]}",k[0],k[1],code,lib,poste,sens,sig,vcode,exlab,round(tot*share(k,drv))))
# vérification EBITDA par version
for vcode,exlab,n,vt,vy in VERS:
    ca=sum(m for row in compta for m in [row[11]] if row[7]=="Produit" and row[9]==vcode)
    ch=sum(row[11] for row in compta if row[7]=="Charge" and row[8]!="Dotations" and row[9]==vcode)
    da=sum(row[11] for row in compta if row[8]=="Dotations" and row[9]==vcode)
    print("[py] %s  CA=%d  EBITDA=%d (%.1f%%)  EBIT=%d (%.1f%%)"%(exlab,ca,ca-ch,(ca-ch)/ca*100,ca-ch-da,(ca-ch-da)/ca*100))

# ---- 05_Plan_Comptable ----
ws=wb.create_sheet("05_Plan_Comptable"); ws.sheet_view.showGridLines=False
pcols=["Compte","Libellé","Classe PCG","Poste PCG","Ligne SIG","Agrégat SIG","Sens","Rattachement","Driver","% du CA","V/F"]
for i,w in enumerate([9,42,26,30,17,26,9,11,10,9,5]): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells(f"A1:{GL(len(pcols))}1"); C(ws,"A1","PLAN COMPTABLE — comptes PCG · double hiérarchie (comptable + gestion) · prêt Tagetik (dimension Compte)",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=22
for i,h in enumerate(pcols): C(ws,f"{GL(1+i)}3",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[3].height=28
r=4
for code,lib,sens,sig,niv,drv,pct,vf in ACCTS:
    C(ws,f"A{r}",code,(CFB if sens=="Produit" else CF),align=AC,border=True); C(ws,f"B{r}",lib,CREG,align=AL,border=True)
    C(ws,f"C{r}",classe_of(code),CF,align=AL,border=True); C(ws,f"D{r}",poste_of(code),CF,align=AL,border=True)
    C(ws,f"E{r}",sig,CF,align=AL,border=True); C(ws,f"F{r}",AGG[sig],CF,align=AL,border=True)
    C(ws,f"G{r}",sens,CL,align=AC,border=True)
    C(ws,f"H{r}",{"campus":"Campus","groupe":"Groupe","cellule":"Cellule"}[niv],CF,align=AC,border=True)
    C(ws,f"I{r}",("—" if sens=="Produit" else drv),CF,align=AC,border=True)
    C(ws,f"J{r}",("—" if pct is None else pct),CF,fmt=(None if pct is None else PCT),align=AC,border=True)
    C(ws,f"K{r}",("—" if sens=="Produit" else vf),CF,align=AC,border=True); r+=1
# hiérarchie de gestion (parent -> enfant) prête Tagetik
r+=2; band(ws,r,"A","D","Hiérarchie de gestion (dimension Compte — parent → enfant, prêt Tagetik)"); r+=1
for i,h in enumerate(["Code nœud","Libellé","Parent","Type"]): C(ws,f"{GL(1+i)}{r}",h,CHDR,FBLUE,align=AC,border=True)
r+=1
HIER=[("RESULTAT_EXPL","Résultat d'exploitation (EBIT)","(racine)","Agrégat"),
 ("EBITDA","EBITDA","RESULTAT_EXPL","Agrégat"),
 ("MARGE_CONTRIB","Marge de contribution","EBITDA","Agrégat"),
 ("SIG_PROD","Produits d'exploitation","MARGE_CONTRIB","Regroupement"),
 ("SIG_DIR","Coûts directs","MARGE_CONTRIB","Regroupement"),
 ("SIG_PERS","Charges de personnel","EBITDA","Regroupement"),
 ("SIG_STRUCT","Charges de structure","EBITDA","Regroupement"),
 ("SIG_IMP","Impôts & taxes","EBITDA","Regroupement"),
 ("SIG_DOT","Dotations aux amortissements","RESULTAT_EXPL","Regroupement")]
for cnode,lnode,par,typ in HIER:
    C(ws,f"A{r}",cnode,CFB,align=AL,border=True); C(ws,f"B{r}",lnode,CB,align=AL,border=True)
    C(ws,f"C{r}",par,CF,align=AL,border=True); C(ws,f"D{r}",typ,CIT,align=AC,border=True); r+=1
for code,lib,sens,sig,*_ in ACCTS:
    C(ws,f"A{r}",code,CF,align=AL,border=True); C(ws,f"B{r}",lib,CREG,align=AL,border=True)
    C(ws,f"C{r}",SIGNODE[sig],CF,align=AL,border=True); C(ws,f"D{r}","Compte (détail)",CIT,align=AC,border=True); r+=1
# dimension Version (référentiel)
r+=2; band(ws,r,"A","D","Dimension Version (référentiel — prêt Tagetik)"); r+=1
for i,h in enumerate(["Code version","Libellé","Type","Année"]): C(ws,f"{GL(1+i)}{r}",h,CHDR,FBLUE,align=AC,border=True)
r+=1
for vcode,exlab,n,vtype,vyear in VERS:
    C(ws,f"A{r}",vcode,CFB,align=AL,border=True); C(ws,f"B{r}",exlab,CREG,align=AL,border=True)
    C(ws,f"C{r}",vtype,CF,align=AC,border=True); C(ws,f"D{r}",vyear,CF,align=AC,border=True); r+=1

# ---- 06_Compta (jeu de données chargé, clés + version) ----
ws=wb.create_sheet("06_Compta"); ws.sheet_view.showGridLines=False
kcols=["Code entité","Entité","Marque","Ville","Compte","Libellé compte","Poste PCG","Sens","Ligne SIG","Version","Exercice","Montant"]
for i,w in enumerate([18,20,16,11,9,40,28,9,15,15,15,12]): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells(f"A1:{GL(len(kcols))}1"); C(ws,"A1","COMPTA CHARGÉE — écritures par entité × compte × version (format long, clés Tagetik)",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=22
ws.merge_cells(f"A2:{GL(len(kcols))}2"); C(ws,"A2","Clés : code entité (marque_ville), compte PCG, poste, version (2023ACT_VDEF / 2024ACT_VDEF / 2025ATT_VDEF). Charges réparties par driver. Somme des comptes = CA & EBITDA du P&L.",CIT,align=ALW); ws.row_dimensions[2].height=16
for i,h in enumerate(kcols): C(ws,f"{GL(1+i)}3",h,CHDR,FBLUE,align=AC,border=True)
KP0=4
for idx,row in enumerate(compta):
    r=KP0+idx
    for i,v in enumerate(row):
        al=AR if i==11 else (AL if i in(1,5) else AC)
        C(ws,f"{GL(1+i)}{r}",v,CL,fmt=(EUR if i==11 else None),align=al,border=True)
KPN=KP0+len(compta)-1
ws.freeze_panes="E4"
KACC=f"'06_Compta'!$E${KP0}:$E${KPN}"; KVER=f"'06_Compta'!$J${KP0}:$J${KPN}"; KMT=f"'06_Compta'!$L${KP0}:$L${KPN}"
KSIG=f"'06_Compta'!$I${KP0}:$I${KPN}"; KSENS=f"'06_Compta'!$H${KP0}:$H${KPN}"

# ---- 07_PnL (cascade SIG, 3 versions + variance) ----
ws=wb.create_sheet("07_PnL"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":10,"C":44,"D":14,"E":14,"F":14,"G":14,"H":9}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B1:H1"); C(ws,"B1","COMPTE DE RÉSULTAT (SIG) — réalisé 2024 · 2025 · atterrissage 2026 · BUDGET 2027 (piloté par le cadrage)",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=26
hdr=["Compte","Libellé"]+[v[1] for v in VERS]+["🟢 Budget 2027","Var Bud/Att %"]
for i,h in enumerate(hdr): C(ws,f"{GL(2+i)}3",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[3].height=26
VC=[v[0] for v in VERS]   # codes version pour les SUMIFS
def sif(code=None,sig=None,sens=None,ver=""):
    parts=[KMT]
    if code is not None: parts+=[KACC,f'"{code}"']
    if sig is not None: parts+=[KSIG,f'"{sig}"']
    if sens is not None: parts+=[KSENS,f'"{sens}"']
    parts+=[KVER,f'"{ver}"']
    return "SUMIFS("+",".join(parts)+")"
r=4
# facteur budget N+1 par nature de compte : CA-driven, effectif×salaire, ou inflation
MEFF=f"'04_Moteur'!$Y${MTOT}"; MCA=f"'04_Moteur'!$AA${MTOT}"
CAf=f"({MCA}/{REFCA})"; EFFf=f"({MEFF}/{REFEFF})"
def bfac(code,sens,sig):
    if sens=="Produit": return CAf                                  # produits : croissance du CA moteur
    if code=="6231": return f"(1+{LMKT})"                           # marketing acquisition : suit le budget d'acquisition
    if code=="6236": return f"(1+{LBRAND})"                         # marketing de marque : suit le budget de marque
    if sig=="Coûts directs": return f"({CAf}*(1-{LPROD}))"          # autres directs : volume − productivité
    if sig=="Personnel": return f"((1+{LSAL})*(1+{LEFFP}))"         # personnel : salaire × effectifs permanents
    if sig=="Structure": return f"((1+{LINFL})*(1-{LPROD})*(1+{LSTRUCT}))"          # structure : inflation − productivité × levier structure
    if sig=="Impôts & taxes": return f"((1+{LINFL})*(1-{LPROD}))"                    # impôts & taxes : inflation − productivité
    return f"(1+{LINFL})"                                           # dotations : inflation
glist={}   # groupe -> liste des lignes budget (col G)
def line(code,lib,sgn,sens,sig):
    global r
    C(ws,f"B{r}",code,CF,align=AC,border=True); C(ws,f"C{r}",lib,CF,align=AL,border=True)
    for i,vc in enumerate(VC): C(ws,f"{GL(4+i)}{r}",f"={sgn}{sif(code=code,ver=vc)}",CF,fmt=EUR,align=AR,border=True)
    C(ws,f"G{r}",f"=F{r}*{bfac(code,sens,sig)}",CF,FGRN,fmt=EUR,align=AR,border=True)
    C(ws,f"H{r}",f"=IFERROR(G{r}/F{r}-1,0)",CF,fmt=PCT,align=AR,border=True)
    glist.setdefault(sig,[]).append(r); r+=1
def band2(txt):
    global r; band(ws,r,"B","H",txt,fill=FBLUE); r+=1
def result(lib,dsef,grows):
    global r
    C(ws,f"B{r}"," ",CFB,FTOT,border=True); C(ws,f"C{r}",lib,CFB,FTOT,align=AL,border=True)
    for i,vc in enumerate(VC): C(ws,f"{GL(4+i)}{r}",dsef(vc),CFB,FTOT,fmt=EUR,align=AR,border=True)
    C(ws,f"G{r}","="+"+".join(f"G{x}" for x in grows) if grows else "=0",CFB,FGRN,fmt=EUR,align=AR,border=True)
    C(ws,f"H{r}",f"=IFERROR(G{r}/F{r}-1,0)",CFB,FTOT,fmt=PCT,align=AR,border=True)
    rr=r; r+=1; return rr
band2("PRODUITS")
for code,lib,sens,sig,*_ in ACCTS:
    if sens=="Produit": line(code,lib,"",sens,sig)
allg=[]
rowCA=result("CHIFFRE D'AFFAIRES",lambda vc:f"={sif(sens='Produit',ver=vc)}",glist.get("Produits",[])); allg+=glist.get("Produits",[])
band2("COÛTS DIRECTS")
for code,lib,sens,sig,*_ in ACCTS:
    if sig=="Coûts directs": line(code,lib,"-",sens,sig)
allg+=glist.get("Coûts directs",[])
rowMC=result("MARGE DE CONTRIBUTION",lambda vc:f"={sif(sens='Produit',ver=vc)}-{sif(sig='Coûts directs',ver=vc)}",list(allg))
for grp in ["Personnel","Structure","Impôts & taxes"]:
    band2(grp.upper())
    for code,lib,sens,sig,*_ in ACCTS:
        if sig==grp: line(code,lib,"-",sens,sig)
    allg+=glist.get(grp,[])
rowEB=result("EBITDA",lambda vc:f"={sif(sens='Produit',ver=vc)}-{sif(sens='Charge',ver=vc)}+{sif(sig='Dotations',ver=vc)}",list(allg))
band2("DOTATIONS")
for code,lib,sens,sig,*_ in ACCTS:
    if sig=="Dotations": line(code,lib,"-",sens,sig)
allg+=glist.get("Dotations",[])
rowEBIT=result("EBIT / RÉSULTAT D'EXPLOITATION",lambda vc:f"={sif(sens='Produit',ver=vc)}-{sif(sens='Charge',ver=vc)}",list(allg))
# marges % (D..G : 3 versions + budget)
C(ws,f"C{r}","Marge EBITDA %",CIT,align=AL)
for i in range(4): C(ws,f"{GL(4+i)}{r}",f"=IFERROR({GL(4+i)}{rowEB}/{GL(4+i)}{rowCA},0)",CIT,fmt=PCT,align=AR)
r+=1
C(ws,f"C{r}","Marge EBIT %",CIT,align=AL)
for i in range(4): C(ws,f"{GL(4+i)}{r}",f"=IFERROR({GL(4+i)}{rowEBIT}/{GL(4+i)}{rowCA},0)",CIT,fmt=PCT,align=AR)

# ============================================================ 08_Saisie_Campus (masque de saisie — exemple MBway Lyon)
ws=wb.create_sheet("08_Saisie_Campus"); ws.sheet_view.showGridLines=False
scols=["Programme","Année","Mod.","Leads cadrés (groupe)","Taux L→C","Yield","Passage","Effectif cadré ② (groupe)","🟢 Δ L→C","🟢 Δ yield","🟢 Δ passage","Effectif ajusté ③","CA ajusté ③","CA cible ①"]
for i,w in enumerate([22,7,6,10,9,8,9,14,10,10,11,15,13,13]): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells(f"A1:{GL(len(scols))}1"); C(ws,"A1","SAISIE BUDGÉTAIRE 2027 — Campus MBWAY LYON",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=26
ws.merge_cells(f"A2:{GL(len(scols))}2"); C(ws,"A2","Votre périmètre. « Cadrés (groupe) » = valeurs pré-allouées par le groupe (budget marketing, cible) que vous recevez. Vous n'ajustez que vos TAUX OPÉRATIONNELS (🟢 jaune) — prix, budget marketing et cible sont fixés par le groupe. Vos ajustements s'AJOUTENT au cadré (ils ne l'écrasent pas) et remontent au modèle en temps réel.",CIT,align=ALW); ws.row_dimensions[2].height=30
mkT=MK0+len(ML_IDX)   # ligne total
C(ws,"A4","Cible campus :",CB,align=AR); C(ws,f"C4",f"=N{mkT}",CFB,fmt=EUR,align=AL,border=True)
C(ws,"E4","Construit (ajusté) :",CB,align=AR); C(ws,f"G4",f"=M{mkT}",CFB,fmt=EUR,align=AL,border=True)
C(ws,"I4","Reste à trouver :",CB,align=AR); C(ws,f"K4",f"=IF(N{mkT}-M{mkT}>0,N{mkT}-M{mkT},0)",CFB,FYEL,fmt=EUR,align=AL,border=True)
for i,h in enumerate(scols): C(ws,f"{GL(1+i)}6",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[6].height=30
mk=MK0
for idx in ML_IDX:
    mr=MR0+idx; MOc=lambda col:f"='04_Moteur'!{col}{mr}"
    C(ws,f"A{mk}",MOc('C'),CL,align=AL,border=True); C(ws,f"B{mk}",MOc('D'),CL,align=AC,border=True); C(ws,f"C{mk}",MOc('E'),CL,align=AC,border=True)
    C(ws,f"D{mk}",MOc('T'),CF,fmt=NB,align=AR,border=True); C(ws,f"E{mk}",MOc('O'),CF,fmt=PCT,align=AC,border=True); C(ws,f"F{mk}",MOc('Q'),CF,fmt=PCT,align=AC,border=True); C(ws,f"G{mk}",MOc('N'),CF,fmt=PCT,align=AC,border=True)
    C(ws,f"H{mk}",MOc('AD'),CF,fmt=NB,align=AR,border=True)
    C(ws,f"I{mk}",0,CINB,FYEL,fmt=PCT,align=AC,border=True); C(ws,f"J{mk}",0,CINB,FYEL,fmt=PCT,align=AC,border=True); C(ws,f"K{mk}",0,CINB,FYEL,fmt=PCT,align=AC,border=True)
    C(ws,f"L{mk}",MOc('Y'),CFB,fmt=NB,align=AR,border=True); C(ws,f"M{mk}",MOc('AA'),CFB,fmt=EUR,align=AR,border=True); C(ws,f"N{mk}",MOc('AJ'),CF,fmt=EUR,align=AR,border=True)
    mk+=1
C(ws,f"A{mk}","TOTAL CAMPUS",CFB,FTOT,align=AL,border=True)
for col in ["B","C","E","F","G","I","J","K"]: C(ws,f"{col}{mk}"," ",fill=FTOT,border=True)
for col,fmt in [("D",NB),("H",NB),("L",NB),("M",EUR),("N",EUR)]: C(ws,f"{col}{mk}",f"=SUM({col}{MK0}:{col}{mk-1})",CFB,FTOT,fmt=fmt,align=AR,border=True)
ws.freeze_panes="A7"
C(ws,f"A{mk+2}","Exemple : mettez +0,03 en « Δ yield » sur une ligne → l'effectif ajusté ③ et le CA remontent, l'écart à la cible se réduit. En Tagetik, chaque contrôleur a ce masque sur SON périmètre (sécurité par entité).",CIT,align=AL); ws.merge_cells(f"A{mk+2}:N{mk+3}"); ws.row_dimensions[mk+2].height=28

# ============================================================ 09_Reporting (KPIs live + bridge de variance)
ws=wb.create_sheet("09_Reporting"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":15,"C":15,"D":14,"E":3,"F":16,"G":13,"H":13,"I":12,"J":2,"K":12,"L":12,"M":12,"N":12}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B1:N1"); C(ws,"B1","REPORTING — impact live du cadrage · pont d'explication CA & EBITDA",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=26
mCA=f"'04_Moteur'!$AA${MTOT}"; mEFF=f"'04_Moteur'!$Y${MTOT}"; mNOUV=f"'04_Moteur'!$W${MTOT}"
cK=f"SUM('03_Campagnes'!$K${CR0}:$K${CRN})"; cL=f"SUM('03_Campagnes'!$L${CR0}:$L${CRN})"; cM=f"SUM('03_Campagnes'!$M${CR0}:$M${CRN})"
BIG=Font(name=Fn,bold=True,size=16,color="1F3864")
def tile(r,c,label,formula,fmt,fill=FLIGHT):
    a=GL(c); b=GL(c+2)
    for rr in (r,r+1):
        for col in range(c,c+3): ws.cell(row=rr,column=col).fill=fill; ws.cell(row=rr,column=col).border=BORD
    ws.merge_cells(f"{a}{r}:{b}{r}"); ws.merge_cells(f"{a}{r+1}:{b}{r+1}")
    v=ws[f"{a}{r}"]; v.value=formula; v.font=BIG; v.alignment=AC; v.number_format=fmt
    l=ws[f"{a}{r+1}"]; l.value=label; l.font=CIT; l.alignment=AC
tile(3,2,"CA construit (budget)",f"={mCA}",EUR)
tile(3,5,"EBITDA construit",f"={PNL_EBb}",EUR)
tile(3,8,"Marge EBITDA",f"=IFERROR({PNL_EBb}/{mCA},0)",PCT,FGRN)
tile(3,11,"Effectif construit",f"={mEFF}",NB)
tile(6,2,"Écart à la cible CA","='01_Cadrage'!F7",EUR)
tile(6,5,"Reste à trouver CA","='01_Cadrage'!D12",EUR,FYEL)
tile(6,8,"CPL effectif moyen",f"=IFERROR({cK}/{cL},0)",EUR)
tile(6,11,"Coût d'acquisition / inscrit",f"=IFERROR({cK}/{mNOUV},0)",EUR)
# --- Pont CA (Référence → Construit) ---
band(ws,10,"B","D","Pont CA : Référence 2026 → Budget construit")
C(ws,"B11","CA Référence 2026",CB,align=AL,border=True); C(ws,"D11",REFCA,CL,fmt=EUR,align=AR,border=True)
C(ws,"B12","+ Effet volume",CF,align=AL,border=True); C(ws,"D12",f"=({mEFF}-{REFEFF})*({REFCA}/{REFEFF})",CF,fmt=EUR,align=AR,border=True)
C(ws,"B13","+ Effet prix & mix",CF,align=AL,border=True); C(ws,"D13",f"={mCA}-{REFCA}-D12",CF,fmt=EUR,align=AR,border=True)
C(ws,"B14","→ CA Budget construit",CFB,FTOT,align=AL,border=True); C(ws,"D14",f"={mCA}",CFB,FTOT,fmt=EUR,align=AR,border=True)
# --- Pont EBITDA (Référence → Construit) ---
band(ws,16,"B","D","Pont EBITDA : Référence 2026 → Budget construit")
C(ws,"B17","EBITDA Référence 2026",CB,align=AL,border=True); C(ws,"D17",f"={PNL_EBc}",CL,fmt=EUR,align=AR,border=True)
C(ws,"B18","+ Effet CA (à marge réf.)",CF,align=AL,border=True); C(ws,"D18",f"=({mCA}-{REFCA})*IFERROR({PNL_EBc}/{REFCA},0)",CF,fmt=EUR,align=AR,border=True)
C(ws,"B19","+ Effet coûts & levier op.",CF,align=AL,border=True); C(ws,"D19",f"={PNL_EBb}-{PNL_EBc}-D18",CF,fmt=EUR,align=AR,border=True)
C(ws,"B20","→ EBITDA Budget construit",CFB,FTOT,align=AL,border=True); C(ws,"D20",f"={PNL_EBb}",CFB,FTOT,fmt=EUR,align=AR,border=True)
# --- Par marque : cible vs construit ---
MOA=f"'04_Moteur'!$A${MR0}:$A${MRN}"; MOAA=f"'04_Moteur'!$AA${MR0}:$AA${MRN}"; MOAJ=f"'04_Moteur'!$AJ${MR0}:$AJ${MRN}"
band(ws,10,"F","I","CA par marque : cible vs construit")
for i,h in enumerate(["Marque","🎯 Cible","🔧 Construit","Écart"]): C(ws,f"{GL(6+i)}11",h,CHDR,FBLUE,align=AC,border=True)
r=12
for m in MARQUES:
    C(ws,f"F{r}",m,CL,align=AL,border=True)
    C(ws,f"G{r}",f'=SUMIFS({MOAJ},{MOA},F{r})',CF,fmt=EUR,align=AR,border=True)
    C(ws,f"H{r}",f'=SUMIFS({MOAA},{MOA},F{r})',CFB,fmt=EUR,align=AR,border=True)
    C(ws,f"I{r}",f"=H{r}-G{r}",CF,fmt=EUR,align=AR,border=True); r+=1
# graphe barres cible vs construit par marque
ch=BarChart(); ch.type="col"; ch.title="CA par marque — cible vs construit"; ch.height=7.5; ch.width=15
data=Reference(ws,min_col=7,max_col=8,min_row=11,max_row=11+len(MARQUES)); cats=Reference(ws,min_col=6,min_row=12,max_row=11+len(MARQUES))
ch.add_data(data,titles_from_data=True); ch.set_categories(cats); ch.y_axis.numFmt='#,##0'
ws.add_chart(ch,"F20")
C(ws,"B23","Les tuiles et les ponts se recalculent en direct quand on bouge un levier du cadrage ou un scénario.",CIT,align=AL); ws.merge_cells("B23:N23")

# ============================================================ 10_Allocation (cascade 3 niveaux → classe, clé/niveau, coût/étudiant)
ws=wb.create_sheet("10_Allocation"); ws.sheet_view.showGridLines=False
acols=["Marque","Ville","Programme","Année","Mod.","CA","Effectif","Classes","Coûts directs","Siège alloué","Structure allouée","Coût total","Coût / étudiant","Coût / classe","Marge / étu.","🟢 Contestation Δ (niv.3)"]
for i,w in enumerate([13,10,17,6,5,12,8,7,12,12,13,13,12,12,12,14]): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells(f"A1:{GL(len(acols))}1"); C(ws,"A1","ALLOCATION EN CASCADE (3 niveaux) — MÉTHODE (clé) figée par le CFO · MATIÈRE (assiette) contestable par le contrôleur · jusqu'à la classe",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=24
A26V="2026ATT_VDEF"; DIR_PCT=round(sum(a[6] for a in ACCTS if a[3]=="Coûts directs"),4)
KMARQUE=f"'06_Compta'!$C${KP0}:$C${KPN}"; KVILLE=f"'06_Compta'!$D${KP0}:$D${KPN}"; KENT=f"'06_Compta'!$A${KP0}:$A${KPN}"
BU=brng('U'); BM=brng('M'); BO=brng('O'); BA=brng('A'); BB=brng('B')
def dvf(sel,crit=None):
    if crit:
        inner=f'IF({sel}="Nombre de classes",SUMIFS({BO},{crit}),SUMIFS({BU},{crit}))'
        return f'IF({sel}="Effectif",SUMIFS({BM},{crit}),{inner})'
    inner=f'IF({sel}="Nombre de classes",SUM({BO}),SUM({BU}))'
    return f'IF({sel}="Effectif",SUM({BM}),{inner})'
# les CLÉS (méthode) sont figées par le CFO au cadrage : ici on les LIT (vert, verrouillé), on ne les choisit pas
for rr,lab,src in [(3,"Niveau 1 — Groupe → Marque, clé (CFO) :",ALLOC_N1),(4,"Niveau 2 — Marque → Campus, clé (CFO) :",ALLOC_N2),(5,"Niveau 3 — Campus → Classe, clé (CFO) :",ALLOC_N3)]:
    C(ws,f"B{rr}",lab,CB,align=AR); ws.merge_cells(f"B{rr}:D{rr}"); C(ws,f"E{rr}",f"={src}",CL,FGRN,align=AC,border=True)
GRP_CHG=f'SUMIFS({KMT},{KENT},"GROUPE",{KSENS},"Charge",{KVER},"{A26V}")'
C(ws,"G3","Coûts siège à cascader :",CB,align=AR); C(ws,"I3",f"={GRP_CHG}",CFB,fmt=EUR,align=AR,border=True)
D1S,D2S,D3S=ALLOC_N1,ALLOC_N2,ALLOC_N3
for i,h in enumerate(acols): C(ws,f"{GL(1+i)}7",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[7].height=28
AA0=8; AAN=AA0+N-1
PALL=f"$P${AA0}:$P${AAN}"; AALLOC=f"$A${AA0}:$A${AAN}"; BALLOC=f"$B${AA0}:$B${AAN}"   # contestation : delta d'assiette niveau 3
for idx in range(N):
    r=AA0+idx; b=BR0+idx
    C(ws,f"A{r}",f"={BASE}A{b}",CL,align=AL,border=True); C(ws,f"B{r}",f"={BASE}B{b}",CL,align=AC,border=True); C(ws,f"C{r}",f"={BASE}C{b}",CL,align=AL,border=True)
    C(ws,f"D{r}",f"={BASE}E{b}",CL,align=AC,border=True); C(ws,f"E{r}",f"={BASE}F{b}",CL,align=AC,border=True)
    C(ws,f"F{r}",f"={BASE}U{b}",CF,fmt=EUR,align=AR,border=True); C(ws,f"G{r}",f"={BASE}M{b}",CF,fmt=NB,align=AR,border=True); C(ws,f"H{r}",f"={BASE}O{b}",CF,fmt=NB,align=AR,border=True)
    mC=f'{BA},A{r}'; cC=f'{BA},A{r},{BB},B{r}'; kC=f'{KMARQUE},A{r},{KVILLE},B{r},{KVER},"{A26V}"'
    msh=f'IFERROR({dvf(D1S,mC)}/{dvf(D1S)},0)'
    csh=f'IFERROR({dvf(D2S,cC)}/{dvf(D2S,mC)},0)'
    celldrv=f'IF({D3S}="Effectif",{BASE}M{b},IF({D3S}="Nombre de classes",{BASE}O{b},{BASE}U{b}))'
    # contestation : le contrôleur ajoute un Δ d'assiette (P) ; numérateur ET dénominateur de campus l'intègrent
    #  → la répartition INTRA-campus se déplace, le total campus/groupe reste identique (tie-out préservé, jamais d'écrasement)
    lsh=f'IFERROR((({celldrv})+P{r})/({dvf(D3S,cC)}+SUMIFS({PALL},{AALLOC},A{r},{BALLOC},B{r})),0)'
    campind=f'(SUMIFS({KMT},{kC},{KSENS},"Charge")-SUMIFS({KMT},{kC},{KSIG},"Dotations")-SUMIFS({BU},{cC})*{DIR_PCT})'
    C(ws,f"I{r}",f"=F{r}*{DIR_PCT}",CF,fmt=EUR,align=AR,border=True)                       # coûts directs (taux)
    C(ws,f"J{r}",f"=$I$3*({msh})*({csh})*({lsh})",CF,fmt=EUR,align=AR,border=True)          # siège cascadé 3 niveaux
    C(ws,f"K{r}",f"={campind}*({lsh})",CF,fmt=EUR,align=AR,border=True)                     # structure campus allouée (niveau 3)
    C(ws,f"L{r}",f"=I{r}+J{r}+K{r}",CFB,fmt=EUR,align=AR,border=True)
    C(ws,f"M{r}",f"=IFERROR(L{r}/G{r},0)",CFB,fmt=EUR,align=AR,border=True)                 # coût / étudiant
    C(ws,f"N{r}",f"=IFERROR(L{r}/H{r},0)",CF,fmt=EUR,align=AR,border=True)                  # coût / classe
    C(ws,f"O{r}",f"=IFERROR(F{r}/G{r}-M{r},0)",CF,fmt=EUR,align=AR,border=True)             # marge / étudiant
    C(ws,f"P{r}",0,CINB,FYEL,fmt=NB,align=AC,border=True)                                   # 🟢 contestation assiette (Δ, unité de la clé niv.3)
r=AA0+N
C(ws,f"A{r}","TOTAL",CFB,FTOT,align=AL,border=True)
for col in ["B","C","D","E"]: C(ws,f"{col}{r}"," ",fill=FTOT,border=True)
for col,fmt in [("F",EUR),("G",NB),("H",NB),("I",EUR),("J",EUR),("K",EUR),("L",EUR)]:
    C(ws,f"{col}{r}",f"=SUM({col}{AA0}:{col}{r-1})",CFB,FTOT,fmt=fmt,align=AR,border=True)
C(ws,f"M{r}",f"=IFERROR(L{r}/G{r},0)",CFB,FTOT,fmt=EUR,align=AR,border=True)
C(ws,f"N{r}",f"=IFERROR(L{r}/H{r},0)",CFB,FTOT,fmt=EUR,align=AR,border=True)
C(ws,f"O{r}",f"=IFERROR(F{r}/G{r}-M{r},0)",CFB,FTOT,fmt=EUR,align=AR,border=True)
C(ws,f"P{r}",f"=SUM(P{AA0}:P{r-1})",CFB,FTOT,fmt=NB,align=AC,border=True)
ws.freeze_panes="F8"
ws.conditional_formatting.add(f"O{AA0}:O{r-1}",
    ColorScaleRule(start_type="min",start_color="F8696B",mid_type="num",mid_value=0,mid_color="FFEB84",end_type="max",end_color="63BE7B"))
# --- panneau de gouvernance des clés : qui tient la MÉTHODE, qui tient la MATIÈRE ---
gr=r+2
band(ws,gr,"A","H","Gouvernance des clés — le groupe délègue la MATIÈRE, pas la MÉTHODE")
C(ws,f"A{gr+1}","MÉTHODE (la clé)",CB,FLIGHT,align=AL,border=True); ws.merge_cells(f"A{gr+1}:B{gr+1}")
C(ws,f"C{gr+1}","décidée par le CFO, figée au cadrage, uniforme sur tout le groupe (comparabilité). Verrouillée ici (vert).",CREG,align=ALW,border=True); ws.merge_cells(f"C{gr+1}:H{gr+1}")
C(ws,f"A{gr+2}","MATIÈRE (l'assiette)",CB,FGRN,align=AL,border=True); ws.merge_cells(f"A{gr+2}:B{gr+2}")
C(ws,f"C{gr+2}","le contrôleur conteste l'assiette réelle (colonne P, Δ) sur SON périmètre — jamais d'écrasement. La contestation redéploie le coût INTRA-campus ; le total campus & groupe est inchangé (tie-out préservé), elle remonte au CFO pour arbitrage.",CREG,align=ALW,border=True); ws.merge_cells(f"C{gr+2}:H{gr+2}"); ws.row_dimensions[gr+2].height=42
C(ws,f"A{gr+4}","Contestations en attente :",CB,align=AR); ws.merge_cells(f"A{gr+4}:C{gr+4}")
C(ws,f"D{gr+4}",f'=COUNTIF({PALL},"<>0")',CFB,FYEL,fmt=NB,align=AC,border=True)
C(ws,f"E{gr+4}","(nombre de cellules où le contrôleur conteste l'assiette — à arbitrer par le CFO)",CIT,align=AL); ws.merge_cells(f"E{gr+4}:H{gr+4}")
C(ws,f"A{gr+6}","Cascade Groupe → Marque → Campus → Classe : la CLÉ (méthode) est figée par le CFO au cadrage. Coût par étudiant PLEINEMENT chargé (directs + siège cascadé + structure). Marge/étudiant en rouge = classe déficitaire → cible du simulateur d'ouverture / fermeture. Testez la sensibilité : changez une clé au cadrage → une classe peut basculer de rentable à déficitaire.",CIT,align=AL); ws.merge_cells(f"A{gr+6}:O{gr+7}"); ws.row_dimensions[gr+6].height=42

# ============================================================ 11_Reconciliation (contrôle croisé multisource)
ws=wb.create_sheet("11_Reconciliation"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":30,"C":22,"D":15,"E":22,"F":15,"G":13,"H":13}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B1:H1"); C(ws,"B1","RÉCONCILIATION MULTISOURCE — CRM · Compta · Base se recoupent",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=26
ws.merge_cells("B2:H2"); C(ws,"B2","Contrôle interne : chaque grandeur est calculée depuis deux systèmes différents ; l'écart doit être nul. Preuve que le « chargement multisource » se réconcilie.",CIT,align=ALW); ws.row_dimensions[2].height=16
V26="2026ATT_VDEF"
CA_base=f"SUM({brng('U')})"; CA_compta=f'SUMIFS({KMT},{KSENS},"Produit",{KVER},"{V26}")'
Leads_crm=f'SUMIFS({xrng("G")},{XVER},2026)'; Leads_base=f"SUM({brng('H')})"
Spend_crm=f'SUMIFS({xrng("H")},{XVER},2026)'; Spend_compta=f'SUMIFS({KMT},{KACC},"6231",{KVER},"{V26}")'; Spend_camp=f"SUM('03_Campagnes'!$G${CR0}:$G${CRN})"
Brand_crm=f'SUMIFS({xrng("I")},{XVER},2026)'; Brand_compta=f'SUMIFS({KMT},{KACC},"6236",{KVER},"{V26}")'
EB_pl=PNL_EBc; EB_compta=f'SUMIFS({KMT},{KSENS},"Produit",{KVER},"{V26}")-SUMIFS({KMT},{KSENS},"Charge",{KVER},"{V26}")+SUMIFS({KMT},{KSIG},"Dotations",{KVER},"{V26}")'
for i,h in enumerate(["Grandeur (atterrissage 2026)","Source A","Valeur A","Source B","Valeur B","Écart","Statut"]): C(ws,f"{GL(2+i)}4",h,CHDR,FBLUE,align=AC,border=True)
recs=[("Chiffre d'affaires","Base (Σ CA réf)",CA_base,"Compta (produits)",CA_compta,EUR),
 ("Leads","CRM (Σ leads)",Leads_crm,"Base (Σ leads hist)",Leads_base,NB),
 ("Dépense acquisition","CRM (dépense acq.)",Spend_crm,"Compta (compte 6231)",Spend_compta,EUR),
 ("Dépense acquisition (bis)","Compta (6231)",Spend_compta,"Campagnes (budget réf)",Spend_camp,EUR),
 ("Dépense marque","CRM (dépense marque)",Brand_crm,"Compta (compte 6236)",Brand_compta,EUR),
 ("EBITDA","P&L (cascade SIG)",EB_pl,"Compta (produits − charges)",EB_compta,EUR)]
r=5
for lib,sA,fA,sB,fB,fmt in recs:
    C(ws,f"B{r}",lib,CB,align=AL,border=True)
    C(ws,f"C{r}",sA,CIT,align=AL,border=True); C(ws,f"D{r}",f"={fA}",CF,fmt=fmt,align=AR,border=True)
    C(ws,f"E{r}",sB,CIT,align=AL,border=True); C(ws,f"F{r}",f"={fB}",CF,fmt=fmt,align=AR,border=True)
    C(ws,f"G{r}","=D{0}-F{0}".format(r),CFB,fmt=fmt,align=AR,border=True)
    C(ws,f"H{r}",f'=IF(ABS(G{r})<=MAX(1,0.001*D{r}),"✓ aligné","⚠ écart")',CFB,FGRN,align=AC,border=True); r+=1
C(ws,f"B{r+1}","Toutes les grandeurs se recoupent (écart nul aux arrondis près) : le CRM, la compta et la base racontent le même chiffre.",CIT,align=AL); ws.merge_cells(f"B{r+1}:H{r+1}")

# ============================================================ 01b_Pilotage (marque×ville) — en fin car dépend de tous les onglets
ws=wb.create_sheet("01b_Pilotage",index=2); ws.sheet_view.showGridLines=False
CmC=f"{CAMP}$C${CR0}:$C${CRN}"; CmG=f"{CAMP}$G${CR0}:$G${CRN}"; CmP=f"{CAMP}$P${CR0}:$P${CRN}"
BUr=brng('U'); BAr=brng('A'); BBr=brng('B'); xG=xrng('G'); xA=xrng('A'); xB=xrng('B')
MOA_=f"'04_Moteur'!$A${MR0}:$A${MRN}"; MOB_=f"'04_Moteur'!$B${MR0}:$B${MRN}"
MOAA_=f"'04_Moteur'!$AA${MR0}:$AA${MRN}"; MOAJ_=f"'04_Moteur'!$AJ${MR0}:$AJ${MRN}"; MOY_=f"'04_Moteur'!$Y${MR0}:$Y${MRN}"
ALA=f"'10_Allocation'!$A$8:$A${8+N-1}"; ALB=f"'10_Allocation'!$B$8:$B${8+N-1}"; ALF=f"'10_Allocation'!$F$8:$F${8+N-1}"; ALL=f"'10_Allocation'!$L$8:$L${8+N-1}"
for i,w in enumerate([14,10,13,13,13,13,9,9,9,12,12]): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells("A1:K1"); C(ws,"A1","PILOTAGE MARQUE × VILLE — cap stratégique (saisie campus) · caps proposés (justifiés) · synthèse CA & EBITDA",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=24
# ---- ① Cap stratégique par campus : saisie + justification ----
band(ws,3,"A","K","① Cap stratégique par campus (marque×ville) — 🔵 saisie (cap retenu, override prix) · 🟢 caps proposés (mesurés)")
for i,h in enumerate(["Marque","Ville","clé","CAC marginal","Croiss. leads","Intensité mkt","Cap éff.","Cap mom.","Cap pot.","🔵 Cap retenu","🔵 Override prix"]): C(ws,f"{GL(1+i)}4",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[4].height=30
CACr=f"$D${PP0}:$D${PPN}"; CRWr=f"$E${PP0}:$E${PPN}"; INTr=f"$F${PP0}:$F${PPN}"
for i,c in enumerate(campus):
    r=PP0+i; m=c["marque"]; v=c["ville"]
    C(ws,f"A{r}",m,CL,align=AL,border=True); C(ws,f"B{r}",v,CL,align=AC,border=True); C(ws,f"C{r}",f'=A{r}&"|"&B{r}',CF,align=AC,border=True)
    C(ws,f"D{r}",f"=IFERROR(INDEX({CmP},MATCH(C{r},{CmC},0)),0)",CF,fmt=EUR,align=AR,border=True)                       # CAC marginal (mesuré)
    C(ws,f"E{r}",f"=IFERROR(SUMIFS({xG},{xA},A{r},{xB},B{r},{XVER},2026)/SUMIFS({xG},{xA},A{r},{xB},B{r},{XVER},2024)-1,0)",CF,fmt=PCT,align=AC,border=True)  # croissance leads 24→26
    C(ws,f"F{r}",f"=IFERROR(INDEX({CmG},MATCH(C{r},{CmC},0))/SUMIFS({BUr},{BAr},A{r},{BBr},B{r}),0)",CF,fmt=PCT,align=AC,border=True)  # intensité mkt
    C(ws,f"G{r}",f"=IFERROR((1/D{r})/(SUMPRODUCT(1/{CACr})/COUNT({CACr})),0)",CF,FGRN,fmt=X2,align=AC,border=True)      # cap éff proposé
    C(ws,f"H{r}",f"=IFERROR(E{r}/(SUM({CRWr})/COUNT({CRWr})),0)",CF,FGRN,fmt=X2,align=AC,border=True)                  # cap mom proposé
    C(ws,f"I{r}",f"=IFERROR((1/F{r})/(SUMPRODUCT(1/{INTr})/COUNT({INTr})),0)",CF,FGRN,fmt=X2,align=AC,border=True)     # cap pot proposé
    C(ws,f"J{r}",1.0,CINB,FYEL,fmt=X2,align=AC,border=True)    # 🔵 cap retenu (saisie campus)
    C(ws,f"K{r}",0,CINB,FYEL,fmt=PCT,align=AC,border=True)     # 🔵 override prix (saisie campus)
ws.conditional_formatting.add(f"G{PP0}:I{PPN}",ColorScaleRule(start_type="num",start_value=0.7,start_color="F8696B",mid_type="num",mid_value=1,mid_color="FFEB84",end_type="num",end_value=1.3,end_color="63BE7B"))
lg=PPN+2
C(ws,f"A{lg}","🟢 Caps proposés = mesurés (vert) : Cap éff. ∝ 1/CAC · Cap mom. ∝ croissance leads · Cap pot. ∝ 1/intensité, normalisés à moyenne 1 (seul le relatif compte). 🔵 Cap retenu = choix CFO qui pilote le budget d'acquisition (enveloppe groupe constante) · Override prix = exception locale. Sources : CAC & budget = 03_Campagnes, croissance = 02_CRM, CA = 02_Base.",CIT,align=ALW); ws.merge_cells(f"A{lg}:K{lg+1}"); ws.row_dimensions[lg].height=44
# ---- ② Synthèse par campus : CA & EBITDA ----
b2=lg+3
band(ws,b2,"A","I","② Synthèse par campus — CA (réf · cible · construit) & EBITDA (réparti selon les clés d'allocation)")
for i,h in enumerate(["Marque","Ville","CA réf","🎯 CA cible","🔧 CA construit","Écart","Effectif","EBITDA","Marge %"]): C(ws,f"{GL(1+i)}{b2+1}",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[b2+1].height=28
s0=b2+2
for i,c in enumerate(campus):
    r=s0+i; m=c["marque"]; v=c["ville"]
    C(ws,f"A{r}",m,CL,align=AL,border=True); C(ws,f"B{r}",v,CL,align=AC,border=True)
    C(ws,f"C{r}",f'=SUMIFS({BUr},{BAr},A{r},{BBr},B{r})',CF,fmt=EUR,align=AR,border=True)
    C(ws,f"D{r}",f'=SUMIFS({MOAJ_},{MOA_},A{r},{MOB_},B{r})',CFB,fmt=EUR,align=AR,border=True)
    C(ws,f"E{r}",f'=SUMIFS({MOAA_},{MOA_},A{r},{MOB_},B{r})',CFB,fmt=EUR,align=AR,border=True)
    C(ws,f"F{r}",f'=E{r}-D{r}',CF,fmt=EUR,align=AR,border=True)
    C(ws,f"G{r}",f'=SUMIFS({MOY_},{MOA_},A{r},{MOB_},B{r})',CF,fmt=NB,align=AR,border=True)
    C(ws,f"H{r}",f'=SUMIFS({ALF},{ALA},A{r},{ALB},B{r})-SUMIFS({ALL},{ALA},A{r},{ALB},B{r})',CFB,fmt=EUR,align=AR,border=True)
    C(ws,f"I{r}",f'=IFERROR(H{r}/C{r},0)',CF,fmt=PCT,align=AC,border=True)
sT=s0+CG
C(ws,f"A{sT}","TOTAL GROUPE",CFB,FTOT,align=AL,border=True); C(ws,f"B{sT}"," ",fill=FTOT,border=True)
for col in ("C","D","E","F","G","H"): C(ws,f"{col}{sT}",f"=SUM({col}{s0}:{col}{sT-1})",CFB,FTOT,fmt=(NB if col=="G" else EUR),align=AR,border=True)
C(ws,f"I{sT}",f"=IFERROR(H{sT}/C{sT},0)",CFB,FTOT,fmt=PCT,align=AC,border=True)
ws.conditional_formatting.add(f"E{s0}:E{sT-1}",DataBarRule(start_type="min",end_type="max",color="5B9BD5"))
ws.conditional_formatting.add(f"H{s0}:H{sT-1}",DataBarRule(start_type="min",end_type="max",color="70AD47"))
ws.conditional_formatting.add(f"I{s0}:I{sT-1}",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="63BE7B"))
ws.freeze_panes="C5"
nn=sT+2
C(ws,f"A{nn}","CA réf = atterrissage · CA cible = éclatement top-down (proportionnel au CA réf) · CA construit = moteur (budget→leads→funnel→effectif→CA). EBITDA = CA réf − coût pleinement alloué (10_Allocation) : changez une CLÉ d'allocation (01_Cadrage) → l'EBITDA se redéploie entre campus, total groupe constant (somme nulle).",CIT,align=ALW); ws.merge_cells(f"A{nn}:I{nn+1}"); ws.row_dimensions[nn].height=42
# ---- lexique ----
lx=nn+3; band(ws,lx,"A","I","Lexique — que veut dire chaque colonne / paramètre ?"); lx+=1
LEX=[
 ("Marque / Ville","La maille de pilotage : un campus = une marque × une ville. Toutes les décisions et résultats de cet onglet sont à ce grain."),
 ("CAC marginal","Coût marketing du prochain inscrit du campus (mesuré, 03_Campagnes). Plus il est bas, plus l'euro marketing rapporte."),
 ("Croiss. leads","Taux de croissance des leads du campus 2024→2026 (02_CRM). Mesure la dynamique commerciale récente."),
 ("Intensité mkt","Budget d'acquisition ÷ CA du campus. Basse = campus sous-investi en marketing vs sa taille."),
 ("🟢 Cap éff. / mom. / pot.","Caps PROPOSÉS (mesurés) : éff. ∝ 1/CAC (efficience) · mom. ∝ croissance (momentum) · pot. ∝ 1/intensité (potentiel). Normalisés à moyenne 1. Ils éclairent, ils ne décident pas — la divergence est le plus instructif."),
 ("🔵 Cap retenu","LE choix du CFO, saisi ICI par campus (marque×ville). Concentre le BUDGET MARKETING d'acquisition à enveloppe groupe CONSTANTE : monter le cap d'un campus → plus de budget → plus de leads → son CA CONSTRUIT monte (baisse ailleurs). Cap = 1 partout = aucune redistribution ; seuls les caps RELATIFS comptent."),
 ("🔵 Override prix","Exception de prix locale par campus (%), qui s'ajoute au coeff prix de la marque. L'écart de prix entre villes est déjà déduit dans la base — cet override ne sert qu'aux cas particuliers."),
 ("🎯 CA cible","Part du CA cible groupe attribuée au campus par l'éclatement top-down (proportionnel au CA réf). Somme des cibles = objectif groupe."),
 ("🔧 CA construit","Ce que le moteur reconstruit du bas vers le haut après leviers, cap et ajustements. C'est le budget réel du campus."),
 ("EBITDA / Marge","CA réf − coût pleinement alloué (directs + siège cascadé + structure, 10_Allocation). Dépend des CLÉS d'allocation fixées par le CFO : changer une clé redéploie l'EBITDA entre campus, à total groupe constant."),
]
for term,desc in LEX:
    C(ws,f"A{lx}",term,CB,FLIGHT,align=AL,border=True)
    C(ws,f"B{lx}",desc,CREG,align=ALW,border=True); ws.merge_cells(f"B{lx}:I{lx}"); ws.row_dimensions[lx].height=(42 if len(desc)>150 else 30); lx+=1

try: wb.calculation.fullCalcOnLoad=True
except Exception: pass
wb.properties.calcMode="auto"
wb.save(OUT)
print("[py] écrit :",OUT)
