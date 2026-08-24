#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Complete le Design Tagetik DESIGN3 -> 100% operationnel (simulable), ADDITIF.
Ne deplace rien: remplit les placeholders 1000 par des formules, ajoute _CALC_*,
00_Cartographie, feeds baseline + colonnes live (2000/6000). Noms sur les cellules
DESIGN3. Rejoue recalcule en interne (pas de colonne cap rejoue). Cles FR->code."""
import warnings;warnings.filterwarnings("ignore")
import re,openpyxl
from openpyxl.utils import get_column_letter as GCL
from xml.sax.saxutils import escape
from tgk_surgery import Book
P=openpyxl.load_workbook("CAD_SAAD_LIVE.xlsx")           # prototype (formules)
Pd=openpyxl.load_workbook("CAD_SAAD_LIVE.xlsx",data_only=True)  # valeurs
b=Book("DESIGN3.xlsm")
NM_MAX={"Socle":2000,"Campagne":2000,"Moteur":2000,"Compta":2000,"Allocation":2000,"PNL":6000}
RAWW={"Socle":29,"Campagne":14,"Moteur":13,"Compta":6,"PNL":7,"Allocation":20}

def remap_calc(f):
    f=re.sub(r'SUMIFS\(Pilotage!\$Q:\$Q,Pilotage!\$F:\$F,(\$A\d+)\)',
             r'SUMIFS(PIL!$K:$K,PIL!$A:$A,\1)*SUMIFS(PIL!$J:$J,PIL!$A:$A,\1)*(SUM(BUD_REF_CAP)/SUMPRODUCT(BUD_REF_CAP,HYP_CAP_RETENU))',f)
    f=f.replace("Pilotage!$N:$N,Pilotage!$F:$F","PIL!$K:$K,PIL!$A:$A")
    f=f.replace("Pilotage!$F:$F","PIL!$A:$A").replace("Pilotage!","PIL!")
    f=f.replace("'3_Allocation'!","ALLOC!").replace("3_Allocation!","ALLOC!")
    return f
def putf(sh,ref,f): sh.set_formula(ref,f)
def putv(sh,ref,v):
    if isinstance(v,str): sh.set_text(ref,v)
    else: sh.set_number(ref,v)

# ================= 1) ZONES NOMMEES =================
d={"TEC_PL":"cad!$F$5","TEC_EBITDA":"cad!$F$6","SCENARIO_ACTIF":"cad!$C$5","SCENARIO_CODE":"cad!$P$1"}
LEVROW={"ACQ_BUD":23,"BRAND_BUD":24,"PRICE":25,"CONV_LEAD":26,"CONV_ADM":27,"PASSAGE":28,
        "INFL_EXT":32,"SALARY":33,"FTE_PERM":34,"PRODUCTIVITY":35,"STRUCT_COST":36,"FEE":39}
for p,r in LEVROW.items():
    d["HYP_%s_V01"%p]="cad!$C$%d"%r; d["HYP_%s_V02"%p]="cad!$D$%d"%r; d["HYP_%s_V03"%p]="cad!$E$%d"%r
for m,r in {"MBWAY":12,"ISCOM":13,"IPAC":14,"PIGIER":15,"TUNON":16}.items():
    d["HYP_PRICE_COEF_%s"%m]="cad!$J$%d"%r
d["HYP_CAP_RETENU"]="PIL!$J$15:$J$28"; d["BUD_REF_CAP"]="PIL!$K$15:$K$28"
d["ALLOC_GRP_BRAND"]="ALLOC!$N$1"; d["ALLOC_GRP_MARQUE"]="ALLOC!$N$2"
d["ALLOC_BRAND_CAMP"]="ALLOC!$N$3"; d["ALLOC_CAMP_CLASS"]="ALLOC!$N$4"
b.add_names(d)

# ================= 2) cad : defauts calibres + formules =================
cad=b.sheet("cad")
putf(cad,"P1",'IF(C5="Optimiste","V02",IF(C5="Prudent","V03","V01"))')  # SCENARIO_CODE
putv(cad,"F5",0.05); putv(cad,"F6",0.15)                               # cibles
for i,co in enumerate([1.2,1.15,0.95,0.9,1.05]): putv(cad,"J%d"%(12+i),co)  # coeff
LEVVAL={23:(0.08,0.15,-0.05),24:(0.1,0.2,-0.05),25:(0.002855,0.035,0.02),26:(0.01,0.03,0),
        27:(0.01,0.025,0),28:(0.005,0.015,-0.01),32:(0.02,0.015,0.03),33:(0.025,0.02,0.03),
        34:(0.04,0.03,0.05),35:(0.018697,0.03,0),36:(0,-0.03,0.04),39:(90,90,90)}
for r,(v1,v2,v3) in LEVVAL.items():
    putv(cad,"C%d"%r,v1); putv(cad,"D%d"%r,v2); putv(cad,"E%d"%r,v3)
    putf(cad,"F%d"%r,"INDEX(C%d:E%d,MATCH(SCENARIO_ACTIF,$C$22:$E$22,0))"%(r,r))   # ACTIF
# reconciliation
putv(cad,"C13",22544725); putv(cad,"C14",3291530); putv(cad,"C16",3036)
putf(cad,"D13","C13*(1+TEC_PL)"); putf(cad,"D14","D13*TEC_EBITDA"); putf(cad,"D15","TEC_EBITDA")
putf(cad,"E13","SUMIFS(Moteur!$R:$R,Moteur!$B:$B,SCENARIO_CODE)")
putf(cad,"E14",'SUMIFS(_CALC_PNL!$T:$T,_CALC_PNL!$D:$D,SCENARIO_CODE,_CALC_PNL!$C:$C,"2027")')
putf(cad,"E15","IFERROR(E14/E13,0)")
putf(cad,"E16","SUMIFS(Moteur!$P:$P,Moteur!$B:$B,SCENARIO_CODE)")
for r in (13,14,16):
    putf(cad,"F%d"%r,"E%d-D%d"%(r,r)); putf(cad,"G%d"%r,"IFERROR(E%d/D%d-1,0)"%(r,r))
putf(cad,"F15","E15-D15"); putf(cad,"G15","IFERROR(E15-D15,0)")

# ================= 3) onglets _CALC + 00_Cartographie =================
for s in ["_CALC_MOTEUR","_CALC_PNL","_CALC_ALLOC","00_Cartographie"]:
    ws=P[s]; sh=b.add_sheet(s)
    for row in ws.iter_rows():
        for c in row:
            v=c.value
            if v is None: continue
            if isinstance(v,str) and v.startswith("="): sh.set_formula(c.coordinate,remap_calc(v[1:]))
            elif isinstance(v,str): sh.set_text(c.coordinate,v)
            else: sh.set_number(c.coordinate,v)

# ================= 4) feeds : baseline + colonnes live =================
for feed,rw in RAWW.items():
    pv=Pd[feed]; pf=P[feed]; fsh=b.sheet(feed)
    for r in range(2,pv.max_row+1):
        if pv.cell(r,1).value in (None,""): continue
        for c in range(1,rw+1):
            v=pv.cell(r,c).value
            if v is not None: putv(fsh,"%s%d"%(GCL(c),r),v)
    livecols=set()
    for r in range(2,pf.max_row+1):
        for c in range(1,pf.max_column+1):
            v=pf.cell(r,c).value
            if isinstance(v,str) and v.startswith("="): livecols.add(c)
    for c in sorted(livecols):
        h=pf.cell(1,c).value
        if h is not None: putv(fsh,"%s1"%GCL(c),h if not str(h).startswith("=") else "live")
        f2=pf.cell(2,c).value
        if isinstance(f2,str) and f2.startswith("="):
            base=remap_calc(f2[1:])
            for r in range(2,NM_MAX[feed]+1):
                fr=re.sub(r'(?<![A-Za-z0-9$])(\$?[A-Z]{1,3}\$?)2\b',lambda m:m.group(1)+str(r),base) if r!=2 else base
                fsh.set_formula("%s%d"%(GCL(c),r),fr)

b.set_fullcalc()
import os
b.save("_tmp_d3.xlsx"); os.replace("_tmp_d3.xlsx","DESIGN3_OP.xlsm")
print("OK phase 1-4 (noms, cad, _CALC, feeds).")
