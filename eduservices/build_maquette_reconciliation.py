#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Maquette : matrice de réconciliation CRM ‖ Compta par marque/campus (2026).
Une ligne = un campus ; blocs CRM et Compta côte à côte ; écart par ligne.
Aucune cellule fusionnée. Formules pour écart et totaux."""
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E4F0EC"; CARD2="F5F7F9"; FAINT="7D8B98"
WHITE="FFFFFF"; BLUE="0000FF"; OCHRE="B3641C"; OCHREBG="F7EAD9"; NAVY="3D4F8F"; NAVYBG="E6E9F4"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
med=Side(style="medium",color=TEAL)
CTR=Alignment("center",vertical="center",wrap_text=True); LFT=Alignment("left",vertical="center"); RGT=Alignment("right",vertical="center")
EUR='#,##0;(#,##0);"-"'; NUM='#,##0;(#,##0);"-"'

CITY={"IPAC_MTP":"Montpellier","IPAC_NAN":"Nantes","IPAC_REN":"Rennes","ISCOM_LIL":"Lille","ISCOM_PAR":"Paris","ISCOM_TLS":"Toulouse",
"MBWAY_BOR":"Bordeaux","MBWAY_LYO":"Lyon","MBWAY_NAN":"Nantes","MBWAY_PAR":"Paris","PIGIER_BOR":"Bordeaux","PIGIER_LYO":"Lyon","TUNON_LYO":"Lyon","TUNON_PAR":"Paris"}
# campus: [eff, CRM_init, CRM_alt, CRM_frais, 706, 7062, 708]
D={"MBWAY_BOR":[249,371025,1393405,9180,371025,1393405,9180],"MBWAY_LYO":[324,519750,1965075,11880,519750,1965075,11880],
"MBWAY_NAN":[294,450000,1697500,10800,450000,1697500,10800],"MBWAY_PAR":[382,655200,2469600,14040,655200,2469600,14040],
"ISCOM_LIL":[244,367500,1379350,9000,367500,1379350,9000],"ISCOM_PAR":[354,604800,2290960,12960,604800,2290960,12960],
"ISCOM_TLS":[231,338400,1281120,8460,338400,1281120,8460],
"IPAC_MTP":[117,0,778050,3600,0,778050,3600],"IPAC_NAN":[145,0,1015000,4500,0,1015000,4500],"IPAC_REN":[117,0,778050,3600,0,778050,3600],
"PIGIER_BOR":[142,0,895310,6480,0,895310,6480],"PIGIER_LYO":[182,0,1242150,8280,0,1242150,8280],
"TUNON_LYO":[117,0,859950,3600,0,859950,3600],"TUNON_PAR":[138,0,1081920,4230,0,1081920,4230]}
MK={"MBWAY":["MBWAY_BOR","MBWAY_LYO","MBWAY_NAN","MBWAY_PAR"],"ISCOM":["ISCOM_LIL","ISCOM_PAR","ISCOM_TLS"],
"IPAC":["IPAC_MTP","IPAC_NAN","IPAC_REN"],"PIGIER":["PIGIER_BOR","PIGIER_LYO"],"TUNON":["TUNON_LYO","TUNON_PAR"]}
ORDER=["MBWAY","ISCOM","IPAC","PIGIER","TUNON"]

wb=openpyxl.Workbook(); ws=wb.active; ws.title="Réconciliation"; ws.sheet_view.showGridLines=False
ws["A1"]="MATRICE DE RÉCONCILIATION  ·  CRM ‖ Compta par campus  ·  2026"; ws["A1"].font=F(15,True,INK)
ws["A2"]="Les deux lentilles côte à côte, découpées pareil. Écart par ligne → on voit direct chez qui. Source : Q_RECONCILIATION_DETAIL. Drill ligne → transactionnel CRM / écritures Compta."; ws["A2"].font=F(9,False,TEALD)

# bloc-headers (SANS fusion : on colore les cellules du bloc)
br=4
ws.cell(br,3,"— CRM (commercial) —").font=F(9,True,TEALD);
for c in (3,4,5): ws.cell(br,c).fill=fill(TEALBG); ws.cell(br,c).alignment=CTR
ws.cell(br,6,"— COMPTA (grand livre) —").font=F(9,True,NAVY)
for c in (6,7,8): ws.cell(br,c).fill=fill(NAVYBG); ws.cell(br,c).alignment=CTR
# col-headers
hdr=["Marque / Campus","Effectif","CA initial","CA alternance","CA frais insc.","Initial (706)","Alternance (7062)","Frais dossier (708)","Écart"]
hr=5
for j,h in enumerate(hdr,1):
    c=ws.cell(hr,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL); c.alignment=CTR if j>1 else LFT
r=hr+1; subtot=[]
for m in ORDER:
    first=r
    for e in MK[m]:
        v=D[e]
        ws.cell(r,1,"   "+m.title()+" — "+CITY[e]).font=F(9); ws.cell(r,1).alignment=LFT
        for k in range(7):   # eff, init, alt, frais, 706,7062,708
            cc=ws.cell(r,2+k,v[k]); cc.font=F(9,False,BLUE); cc.number_format=(NUM if k==0 else EUR); cc.alignment=RGT
        ws.cell(r,9,f"=(F{r}+G{r}+H{r})-(C{r}+D{r}+E{r})").font=F(9); ws.cell(r,9).number_format=EUR; ws.cell(r,9).alignment=RGT
        r+=1
    # sous-total marque
    ws.cell(r,1,m.title()).font=F(10,True,TEALD); ws.cell(r,1).alignment=LFT
    for col in "BCDEFGH": ws.cell(r,ord(col)-64,f"=SUM({col}{first}:{col}{r-1})")
    ws.cell(r,9,f"=SUM(I{first}:I{r-1})")
    for col in range(2,10):
        cc=ws.cell(r,col); cc.number_format=(NUM if col==2 else EUR); cc.alignment=RGT; cc.font=F(10,True,TEALD); cc.fill=fill(TEALBG)
    subtot.append(r); r+=1
# total groupe
ws.cell(r,1,"GROUPE").font=F(10,True,WHITE); ws.cell(r,1).fill=fill(TEALD); ws.cell(r,1).alignment=LFT
for col in "BCDEFGHI":
    ws.cell(r,ord(col)-64,"="+"+".join(f"{col}{s}" for s in subtot))
for col in range(2,10):
    cc=ws.cell(r,col); cc.number_format=(NUM if col==2 else EUR); cc.alignment=RGT; cc.font=F(10,True,WHITE); cc.fill=fill(TEALD); cc.border=Border(top=med)
grp=r
ws.cell(grp+2,1,"Écart 0 partout en 2026 (réconcilié au centime). Un écart non nul se lirait dans la colonne Écart → clic sur la ligne → transactionnel.").font=F(8,False,FAINT,True)
ws.cell(grp+3,1,"IPAC / Pigier / Tunon : CA initial = 0 (100 % alternance). Multidim : roule Campus → Marque → Groupe ; colonnes = mesures.").font=F(8,False,FAINT,True)
widths=[26,9,12,13,12,12,14,14,10]
for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width=w
ws.freeze_panes="A6"

out="/home/user/demo5/eduservices/tagetik/MAQUETTE_RECONCILIATION.xlsx"
wb.save(out); print("SAVED",out)
