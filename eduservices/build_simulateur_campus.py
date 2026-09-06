#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""SIMULATEUR_CAMPUS_MBWAY_BOR.xlsx — le simulateur À LA MAIN du directeur de
campus : ouvrir / fermer une classe, effet immédiat sur la contribution, point
mort (marginal + complet). Données réelles MBway Bordeaux 2026 (socle + compta,
CA 706+7062+708 réconcilié)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E3EFEA"; WHITE="FFFFFF"
SOFT="51606D"; FAINT="7D8B98"; RULE="C8D2DA"; YEL="FFF6DA"; YELB="E8B84B"
OCHRE="B3641C"; OK="1E7A55"; RED="B23A3A"; NAVY="3D4F8F"; CARD="FBFCFB"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color=RULE); med=Side(style="medium",color=RULE)
box=Border(left=thin,right=thin,top=thin,bottom=thin)
Ln=Alignment("left",vertical="center"); Rn=Alignment("right",vertical="center")
Cn=Alignment("center",vertical="center",wrap_text=True); LW=Alignment("left",vertical="center",wrap_text=True)
EUR='#,##0 "€";-#,##0 "€";"—"'; PCT='0.0%'; NUM='#,##0'; NUM1='#,##0.0'

# ---- données réelles MBWAY_BOR 2026 ----
CA_BASE=1773610; COST_BASE=1118034            # CA (706+7062+708) et coûts directs
CONTRIB_BASE=CA_BASE-COST_BASE                 # 655 576
N_CLASSES=10; N_ELEVES=249
COST_PER_CLASS=round(COST_BASE/N_CLASSES)      # coût complet par classe ≈ 111 803

wb=openpyxl.Workbook(); ws=wb.active; ws.title="Simulateur classe"
ws.sheet_view.showGridLines=False
for col,w in zip("ABCDEFGH",[34,15,14,13,13,13,13,13]): ws.column_dimensions[col].width=w

def band(r,txt,c1=1,c2=8,bg=TEALD,fg=WHITE,sz=12,h=24):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    cc=ws.cell(r,c1,txt); cc.font=F(sz,True,fg); cc.fill=fill(bg); cc.alignment=Ln
    ws.row_dimensions[r].height=h
def inp(r,c,lab,val,fmt,help=None):
    ws.cell(r,1,lab).font=F(10,True,INK); ws.cell(r,1).alignment=Ln
    x=ws.cell(r,c,val); x.fill=fill(YEL); x.number_format=fmt; x.font=F(11,True,INK); x.alignment=Cn
    x.border=Border(*[Side(style="medium",color=YELB)]*4)
    if help: ws.cell(r,c+1,help).font=F(8.5,False,FAINT,True); ws.cell(r,c+1).alignment=Ln
    ws.row_dimensions[r].height=19

band(1,"  SIMULATEUR DIRECTEUR DE CAMPUS — MBway Bordeaux  ·  ouvrir / fermer une classe",sz=13,h=28)
ws.merge_cells("A2:H2")
ws.cell(2,1,"  À la main du directeur : je teste une décision, je vois l'effet immédiat sur ma contribution et mon point mort. Jaune = saisie.").font=F(9,False,SOFT,True)
ws.row_dimensions[2].height=18

# ① contexte
band(4,"①  Mon campus aujourd'hui  (pré-rempli — projection groupe)",bg=TEAL,sz=11)
ctx=[("Chiffre d'affaires",CA_BASE,EUR),("Coûts directs",COST_BASE,EUR),
     ("Contribution (avant siège)",CONTRIB_BASE,EUR),("Marge de contribution",CONTRIB_BASE/CA_BASE,PCT),
     ("Étudiants",N_ELEVES,NUM),("Classes",N_CLASSES,NUM),("Coût complet moyen / classe",COST_PER_CLASS,EUR)]
r=5
for lab,v,fmt in ctx:
    ws.cell(r,1,lab).font=F(9.5,False,SOFT); ws.cell(r,1).alignment=Ln
    x=ws.cell(r,2,v); x.number_format=fmt; x.font=F(10,True,TEALD if 'Contribution' in lab else INK); x.alignment=Rn; x.border=box
    ws.row_dimensions[r].height=17; r+=1
# objectif
ws.cell(5,4,"Objectif contribution (lettre de cadrage) :").font=F(9.5,True,OCHRE); ws.cell(5,4).alignment=Rn
inp(5,7,"",0.05,PCT)  # amélioration
ws.cell(5,7).comment=None
ws.cell(6,4,"→ cible contribution :").font=F(9.5,True,OCHRE); ws.cell(6,4).alignment=Rn
c=ws.cell(6,7,f"={CONTRIB_BASE}*(1+G5)"); c.number_format=EUR; c.font=F(10,True,OCHRE); c.alignment=Rn; c.border=box

# ② la décision
band(9,"②  La décision à tester  (ma main)",bg=TEAL,sz=11)
ws.cell(10,1,"Classe concernée (libellé)").font=F(10,True,INK); ws.cell(10,1).alignment=Ln
x=ws.cell(10,3,"BAC_MGT B1"); x.fill=fill(YEL); x.font=F(11,True,INK); x.alignment=Cn; x.border=Border(*[Side(style="medium",color=YELB)]*4)
inp(11,3,"Ouvrir (+1)  /  Fermer (−1)",1,NUM,"tape 1 pour ouvrir, -1 pour fermer")
inp(12,3,"Élèves dans la classe",25,NUM,"combien tu remplis (ou combien il y a si tu fermes)")
inp(13,3,"Heures d'enseignement / an",600,NUM,"BAC INIT 600 · BAC ALT 480 · MAS ALT 420")
inp(14,3,"Taux horaire prof (€, chargé)",50,EUR,"vacataire chargé ≈ 45-60 €/h")
inp(15,3,"Coût ressource additionnelle (€)",0,EUR,"nouveau permanent / nouvelle salle si besoin (sinon 0)")
inp(16,3,"Prix / élève (€)",7275,EUR,"tarif du programme")
inp(17,3,"Coût variable / élève (€)",300,EUR,"pédago, supports, sous-traitance")

# ③ effet immédiat
band(19,"③  Effet immédiat",bg=NAVY,sz=11)
def line(r,lab,formula,fmt,strong=False,c=INK):
    ws.cell(r,1,lab).font=F(10,strong,INK); ws.cell(r,1).alignment=Ln
    x=ws.cell(r,3,formula); x.number_format=fmt; x.font=F(11 if strong else 10,strong,c); x.alignment=Rn; x.border=box
    ws.row_dimensions[r].height=18
line(20,"Coût de la classe (enseignement + ressource)","=C13*C14+C15",EUR)
line(21,"Marge par élève (prix − variable)","=C16-C17",EUR)
line(22,"◆ POINT MORT marginal (élèves)","=IFERROR((C13*C14+C15)/(C16-C17),0)",NUM1,True,TEALD)
line(23,"◆ POINT MORT complet — couvrir le coût moyen/classe","=IFERROR($B$11/(C16-C17),0)",NUM1,True,OCHRE)
# B11 helper = coût complet par classe
ws.cell(11,2,COST_PER_CLASS)  # helper (masqué visuellement à droite)
ws.cell(11,2).font=F(8,False,WHITE)
line(24,"Contribution de la classe (élèves × marge − coût)","=C11*(C12*(C16-C17)-(C13*C14+C15))",EUR,True,INK)
# voyant
ws.merge_cells("A25:H25")
verdict=('=IF(C11=1,'
         'IF(C12>=(C13*C14+C15)/(C16-C17),'
         '"🟢 OUVRIR : classe rentable — "&TEXT(C12,"0")&" élèves > point mort "&TEXT((C13*C14+C15)/(C16-C17),"0.0")&" → contribution "&TEXT(C12*(C16-C17)-(C13*C14+C15),"+#,##0")&" €",'
         '"🔴 OUVRIR risqué : "&TEXT(C12,"0")&" élèves < point mort "&TEXT((C13*C14+C15)/(C16-C17),"0.0")&" → la classe perd "&TEXT(C12*(C16-C17)-(C13*C14+C15),"#,##0")&" €"),'
         'IF(C12*(C16-C17)<(C13*C14+C15),'
         '"🟢 FERMER : cette classe était déficitaire ("&TEXT(C12,"0")&" élèves < point mort) → tu récupères "&TEXT((C13*C14+C15)-C12*(C16-C17),"+#,##0")&" €",'
         '"⚠️ FERMER coûte : cette classe était rentable → tu perds "&TEXT(C12*(C16-C17)-(C13*C14+C15),"#,##0")&" € de contribution"))')
cc=ws.cell(25,1,verdict); cc.font=F(11,True,INK); cc.alignment=LW; cc.fill=fill(CARD)
for j in range(1,9): ws.cell(25,j).fill=fill(CARD)
ws.row_dimensions[25].height=46

# ④ échelle de rentabilité + graphe
band(27,"④  Échelle de rentabilité de la classe  (contribution selon le remplissage)",bg=TEAL,sz=11)
ws.cell(28,1,"Élèves").font=F(9,True,SOFT)
steps=[5,10,15,20,25,30,32]
for j,n in enumerate(steps):
    ws.cell(28,2+j,n).font=F(9,True,SOFT); ws.cell(28,2+j).alignment=Cn; ws.cell(28,2+j).border=box
ws.cell(29,1,"Contribution (€)").font=F(9,True,INK)
for j,n in enumerate(steps):
    c=ws.cell(29,2+j,f"={n}*(C16-C17)-(C13*C14+C15)"); c.number_format=EUR; c.alignment=Cn; c.border=box
    c.font=F(9,True,OK)
ws.row_dimensions[28].height=16; ws.row_dimensions[29].height=17
ch=LineChart(); ch.title="Contribution de la classe selon le nb d'élèves (croise 0 au point mort)"
ch.height=6.5; ch.width=16; ch.style=2; ch.y_axis.numFmt='#,##0 €'; ch.x_axis.delete=False; ch.y_axis.delete=False
dat=Reference(ws,min_col=2,max_col=8,min_row=29,max_row=29)
ch.add_data(dat,from_rows=True,titles_from_data=False)
ch.set_categories(Reference(ws,min_col=2,max_col=8,min_row=28,max_row=28))
ws.add_chart(ch,"A31")

# ⑤ effet campus
band(46,"⑤  Effet sur mon campus  (vs objectif)",bg=NAVY,sz=11)
line(47,"Contribution campus — avant",f"={CONTRIB_BASE}",EUR)
line(48,"+ effet de la décision","=C11*(C12*(C16-C17)-(C13*C14+C15))",EUR)
line(49,"= Contribution campus — après",f"={CONTRIB_BASE}+C11*(C12*(C16-C17)-(C13*C14+C15))",EUR,True,TEALD)
ws.merge_cells("A50:H50")
v2=('="vs objectif ("&TEXT(G6,"#,##0")&" €) : "&IF('
    f'{CONTRIB_BASE}+C11*(C12*(C16-C17)-(C13*C14+C15))>=G6,'
    '"🟢 tenu (+"&TEXT('+f'{CONTRIB_BASE}'+'+C11*(C12*(C16-C17)-(C13*C14+C15))-G6,"#,##0")&" €)",'
    '"🔴 à combler ("&TEXT('+f'{CONTRIB_BASE}'+'+C11*(C12*(C16-C17)-(C13*C14+C15))-G6,"#,##0")&" €)")')
cc=ws.cell(50,1,v2); cc.font=F(11,True,INK); cc.alignment=LW; cc.fill=fill(CARD)
for j in range(1,9): ws.cell(50,j).fill=fill(CARD)
ws.row_dimensions[50].height=22

ws.merge_cells("A52:H53")
ws.cell(52,1,"Lecture : le point mort MARGINAL (≈ coût prof d'une classe) est bas — en alternance, une classe est rentable dès quelques élèves. "
             "Le point mort COMPLET (couvrir le coût moyen chargé d'une classe) est le vrai seuil de « bonne santé ». La contrainte réelle du "
             "directeur n'est pas le coût d'enseigner, c'est de REMPLIR la classe (recrutement) et la CAPACITÉ des salles.").font=F(8.5,False,FAINT,True)
ws.cell(52,1).alignment=LW; ws.row_dimensions[52].height=42

out="/home/user/demo5/eduservices/SIMULATEUR_CAMPUS_MBWAY_BOR.xlsx"
wb.save(out); print("SAVED",out)
print("Contribution base=%d marge=%.1f%% coût/classe=%d"%(CONTRIB_BASE,CONTRIB_BASE/CA_BASE*100,COST_PER_CLASS))
