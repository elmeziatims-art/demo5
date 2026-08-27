#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Maquettes de rapports (Acte 1-2) — rendu soigne pour inspirer la restitution
Tagetik. Chaque rapport de presentation lit une feuille de donnees qui reproduit
fidelement la vue SQL qui l'alimente. Formules partout (recalcul vivant)."""
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart,Reference,Series

# ---------- palette demo ----------
INK="152230"; TEAL="0D7A62"; TEALD="0A5A48"; TEALBG="E4F0EC"
OCHRE="B3641C"; OCHREBG="F7EAD9"; CARD2="F5F7F9"; LINE="DBE2E9"; FAINT="7D8B98"
WHITE="FFFFFF"; BLUE="0000FF"
AR="Arial"
def F(sz=10,b=False,c=INK,it=False): return Font(name=AR,size=sz,bold=b,color=c,italic=it)
def fill(c): return PatternFill("solid",fgColor=c)
thin=Side(style="thin",color=LINE); med=Side(style="medium",color=TEAL)
def bord(**k): return Border(**{s:thin for s in k})
CTR=Alignment(horizontal="center",vertical="center",wrap_text=True)
LFT=Alignment(horizontal="left",vertical="center")
RGT=Alignment(horizontal="right",vertical="center")
EUR='#,##0;(#,##0);"-"'; PCT='0.0%;(0.0%);"-"'; NUM='#,##0;(#,##0);"-"'; EURc='#,##0" €"'

wb=openpyxl.Workbook()

# =====================================================================
# DONNEES P&L (image de V_PNL, VERSION=ACT) — comptes x annee
# =====================================================================
# compte : (2024,2025,2026)
ACC={'706':(2942930,3119504,3306675),'7062':(17023353,18044754,19127440),'708':(98442,104348,110610),
'604':(618583,646928,677808),'6063':(412387,431286,451874),'621':(1443361,1509499,1581554),'6231':(358819,394702,434174),
'6411':(3505311,3665931,3840917),'6413':(1855752,1940786,2033426),'6414':(1134071,1186037,1242649),'645':(2886726,3019002,3163106),
'613':(2165045,2264254,2372329),'615':(309291,323463,338902),'616':(206194,215645,225937),'6226':(515487,539108,564841),
'6236':(499926,580167,676344),'625':(309292,323464,338906),'626':(412389,431286,451872),'6281':(164956,172514,180749),
'6331':(309292,323465,338904),'6333':(103097,107822,112968),'63511':(206196,215643,225935),'6811':(1203886,1276115,1352683)}
LIB={'706':"Prestations de formation - scolarité (initial)",'7062':"Prestations de formation - alternance (OPCO)",
'708':"Frais de dossier & droits d'inscription",'604':"Sous-traitance pédagogique",'6063':"Fournitures pédagogiques & petit équipement",
'621':"Personnel extérieur - vacataires",'6231':"Publicité & marketing d'acquisition",'6411':"Rémunération enseignants permanents",
'6413':"Rémunération personnel administratif",'6414':"Rémunération direction & siège",'645':"Charges sociales & prévoyance",
'613':"Loyers & charges locatives",'615':"Entretien & maintenance",'616':"Primes d'assurance",'6226':"Honoraires",
'6236':"Marketing de marque, salons & JPO",'625':"Déplacements & missions",'626':"Télécom & SI",'6281':"Cotisations & abonnements",
'6331':"Taxe sur les salaires",'6333':"Participation formation",'63511':"Cotisation foncière & CVAE",'6811':"Dotations aux amortissements"}
NOEUD={'706':'PRODUITS','7062':'PRODUITS','708':'PRODUITS','604':'COUTS_DIRECTS','6063':'COUTS_DIRECTS','621':'COUTS_DIRECTS','6231':'COUTS_DIRECTS',
'6411':'PERSONNEL','6413':'PERSONNEL','6414':'PERSONNEL','645':'PERSONNEL','613':'STRUCTURE','615':'STRUCTURE','616':'STRUCTURE',
'6226':'STRUCTURE','6236':'STRUCTURE','625':'STRUCTURE','626':'STRUCTURE','6281':'STRUCTURE','6331':'IMPOTS_TAXES','6333':'IMPOTS_TAXES',
'63511':'IMPOTS_TAXES','6811':'DOTATIONS'}
ROLE={'706':'PRODUIT','7062':'PRODUIT','708':'PRODUIT','604':'DIRECT','6063':'DIRECT','621':'DIRECT','6231':'DIRECT',
'6411':'DIRECT','6413':'STRUCT_CAMP','6414':'HOLDING','645':'STRUCT_CAMP','613':'STRUCT_CAMP','615':'STRUCT_CAMP','616':'STRUCT_CAMP',
'6226':'HOLDING','6236':'FRAIS_MARQUE','625':'STRUCT_CAMP','626':'HOLDING','6281':'HOLDING','6331':'HOLDING','6333':'HOLDING',
'63511':'STRUCT_CAMP','6811':'DOTATION'}
order=['706','7062','708','604','6063','621','6231','6411','6413','6414','645','613','615','616','6226','6236','625','626','6281','6331','6333','63511','6811']

dp=wb.active; dp.title="Données P&L"
dp["A1"]="DONNÉES  ·  image de la vue V_PNL (VERSION = ACT, réel 2024-2026)"; dp["A1"].font=F(11,True,TEALD)
dp["A2"]="Grain compte × exercice. Sur Tagetik, ces montants viennent de la dimension Compte ; les nœuds SIG et l'EBITDA sont remontés par le FST."; dp["A2"].font=F(8,False,FAINT,True)
hdr=["Compte","Libellé","Nœud P&L (FST)","Rôle allocation","2024","2025","2026"]
for j,h in enumerate(hdr,1):
    c=dp.cell(4,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL); c.alignment=CTR; c.border=bord(bottom=1)
r=5
for a in order:
    dp.cell(r,1,a).font=F(9,False,BLUE); dp.cell(r,1).alignment=LFT
    dp.cell(r,2,LIB[a]).font=F(9); dp.cell(r,2).alignment=LFT
    dp.cell(r,3,NOEUD[a]).font=F(8,False,FAINT); dp.cell(r,3).alignment=LFT
    dp.cell(r,4,ROLE[a]).font=F(8,False,FAINT); dp.cell(r,4).alignment=LFT
    for k,v in enumerate(ACC[a]):
        cc=dp.cell(r,5+k,v); cc.font=F(9,False,BLUE); cc.number_format=EUR; cc.alignment=RGT
    if r%2==0:
        for j in range(1,8): dp.cell(r,j).fill=fill(CARD2)
    r+=1
for col,w in zip("ABCDEFG",[9,42,16,15,13,13,13]): dp.column_dimensions[col].width=w
dp.freeze_panes="A5"; dp.sheet_view.showGridLines=False
DPMAX=r-1  # last data row
def dpcol(y): return {'2024':'E','2025':'F','2026':'G'}[y]

# =====================================================================
# RAPPORT P&L ①  (presentation, formules SUMIFS sur Données P&L)
# =====================================================================
pl=wb.create_sheet("P&L ①")
pl.sheet_view.showGridLines=False
pl["A1"]="RAPPORT P&L ①  ·  Compte de résultat comparatif"; pl["A1"].font=F(15,True,INK)
pl["A2"]="Cible du hyperlink depuis la tuile EBITDA du cockpit · POV = Groupe · réel 2024-2026"; pl["A2"].font=F(9,False,TEALD)
pl["A3"]="SOURCE : FST natif Tagetik (010-EBITDA) sur V_PNL. Ici reproduit par formules SUMIFS sur la feuille « Données P&L »."; pl["A3"].font=F(8,False,FAINT,True)
cols=["","2024","2025","2026","Δ 25→26"]
hr=5
for j,h in enumerate(cols,1):
    c=pl.cell(hr,j,h); c.font=F(10,True,WHITE); c.fill=fill(TEAL); c.alignment=CTR if j>1 else LFT; c.border=bord(bottom=1)
pl.cell(hr,1,"en €").font=F(9,True,WHITE)

def sifs(noeud,col): return f"SUMIFS('Données P&L'!${col}$5:${col}${DPMAX},'Données P&L'!$C$5:$C${DPMAX},\"{noeud}\")"
rows=[("PRODUITS","node",False),("Coûts directs","COUTS_DIRECTS",True),("Personnel","PERSONNEL",True),
      ("Structure","STRUCTURE",True),("Impôts & taxes","IMPOTS_TAXES",True),
      ("EBITDA","ebitda",False),("Dotations (D&A)","DOTATIONS",True),("RÉSULTAT D'EXPLOITATION","ebit",False),
      ("Marge EBITDA %","marge",False)]
r=hr+1; ebitda_row=None; prod_row=None
rowmap={}
for lab,tag,neg in rows:
    pl.cell(r,1,lab)
    isbold = tag in("node","ebitda","ebit","marge") or lab=="PRODUITS"
    for k,col in enumerate(["E","F","G"]):
        cc=pl.cell(r,2+k)
        if tag=="node" or lab=="PRODUITS":
            cc.value=f"={sifs('PRODUITS',col)}"; prod_row=r
        elif tag in("COUTS_DIRECTS","PERSONNEL","STRUCTURE","IMPOTS_TAXES","DOTATIONS"):
            cc.value=f"=-{sifs(tag,col)}"
        elif tag=="ebitda":
            cc.value=f"={get_column_letter(2+k)}{prod_row}+SUM({get_column_letter(2+k)}{prod_row+1}:{get_column_letter(2+k)}{prod_row+4})"; ebitda_row=r
        elif tag=="ebit":
            cc.value=f"={get_column_letter(2+k)}{ebitda_row}+{get_column_letter(2+k)}{ebitda_row+1}"
        elif tag=="marge":
            cc.value=f"={get_column_letter(2+k)}{ebitda_row}/{get_column_letter(2+k)}{prod_row}"
        cc.number_format=PCT if tag=="marge" else EUR
        cc.alignment=RGT; cc.font=F(10,isbold)
    # Δ col
    d=pl.cell(r,5)
    if tag=="marge":
        d.value=f"=G{r}-F{r}"; d.number_format='0.0" pt";(0.0)" pt"'
    else:
        d.value=f"=IFERROR((G{r}-F{r})/F{r},0)"; d.number_format=PCT
    d.alignment=RGT; d.font=F(9,False,OCHRE if not isbold else INK,True)
    pl.cell(r,1).font=F(10,isbold, TEALD if tag in("ebitda","ebit") else INK)
    pl.cell(r,1).alignment=LFT
    if tag=="ebitda":
        for j in range(1,6): pl.cell(r,j).fill=fill(TEALBG); pl.cell(r,j).border=Border(top=med,bottom=med)
    if lab=="PRODUITS":
        for j in range(1,6): pl.cell(r,j).fill=fill(CARD2)
    if tag=="ebit":
        for j in range(1,6): pl.cell(r,j).border=bord(top=1)
    rowmap[tag]=r; r+=1
note=r+1
pl.cell(note,1,"Contrôle : EBITDA 2024/25/26 = 2 648 550 / 2 977 604 / 3 291 530  ·  drill possible : PRODUITS → 2 drill-through Constitution du CA (CRM ‖ Compta).").font=F(8,False,FAINT,True)
pl.cell(note+1,1,"À reproduire nativement : hiérarchie Compte + FST EBITDA. Colonne 2027 (V01/V02/V03) s'ajoute aux P&L ②③④.").font=F(8,False,OCHRE,True)
for col,w in zip("ABCDE",[30,15,15,15,13]): pl.column_dimensions[col].width=w

# =====================================================================
# DONNEES CRM (image V_FUNNEL + V_CAC) — marque x annee
# =====================================================================
CRM={'2024':{'MBWAY':(5736,1298,920,452,3326,137627),'ISCOM':(3794,857,608,300,2132,89342),'IPAC':(1966,393,276,116,1237,42867),'PIGIER':(2492,498,350,146,1293,48413),'TUNON':(1317,263,185,78,909,40570)},
'2025':{'MBWAY':(6079,1374,973,482,3487,151390),'ISCOM':(4022,909,644,318,2233,98275),'IPAC':(2086,416,292,123,1307,47155),'PIGIER':(2642,528,370,154,1350,53255),'TUNON':(1397,279,196,82,948,44627)},
'2026':{'MBWAY':(6444,1458,1032,510,3659,166530),'ISCOM':(4263,964,683,338,2337,108103),'IPAC':(2210,442,309,130,1381,51871),'PIGIER':(2800,560,392,164,1410,58580),'TUNON':(1480,296,207,87,988,49090)}}
Mk=['MBWAY','ISCOM','IPAC','PIGIER','TUNON']
dc=wb.create_sheet("Données CRM")
dc.sheet_view.showGridLines=False
dc["A1"]="DONNÉES  ·  image des vues V_FUNNEL + V_CAC (socle CRM)"; dc["A1"].font=F(11,True,TEALD)
dc["A2"]="Grain marque × exercice (mesures additives). Sur Tagetik, maille fine campus/programme/cycle/modalité — les taux, CPL et CAC se calculent dans la matrice."; dc["A2"].font=F(8,False,FAINT,True)
h2=["Exercice","Marque","Leads","Candidats","Admis","Inscrits","Leads payants","Dépense acq. (€)"]
for j,h in enumerate(h2,1):
    c=dc.cell(4,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL); c.alignment=CTR; c.border=bord(bottom=1)
r=5; crm_rows={}
for y in ('2024','2025','2026'):
    for m in Mk:
        v=CRM[y][m]
        dc.cell(r,1,y).font=F(9,False,BLUE); dc.cell(r,1).alignment=CTR
        dc.cell(r,2,m).font=F(9); dc.cell(r,2).alignment=LFT
        for k,val in enumerate(v):
            cc=dc.cell(r,3+k,val); cc.font=F(9,False,BLUE); cc.number_format=NUM if k<5 else EUR; cc.alignment=RGT
        crm_rows[(y,m)]=r
        if r%2==0:
            for j in range(1,9): dc.cell(r,j).fill=fill(CARD2)
        r+=1
DCMAX=r-1
for col,w in zip("ABCDEFGH",[10,10,10,11,9,10,13,15]): dc.column_dimensions[col].width=w
dc.freeze_panes="A5"

# =====================================================================
# RAPPORT FUNNEL & CAC  (un seul rapport ; le funnel explique le CAC)
# =====================================================================
fc=wb.create_sheet("Funnel & CAC")
fc.sheet_view.showGridLines=False
fc["A1"]="RAPPORT FUNNEL & CAC  ·  le coût d'acquisition, expliqué"; fc["A1"].font=F(15,True,INK)
fc["A2"]="Cible du hyperlink depuis la tuile CAC du cockpit · POV = Groupe, drill marque → campus · exercice 2026"; fc["A2"].font=F(9,False,TEALD)
fc["A3"]="SOURCE : matrice multidim sur V_FUNNEL + V_CAC (feuille « Données CRM »). Taux, CPL et CAC = formules (jamais pré-agrégés)."; fc["A3"].font=F(8,False,FAINT,True)
H=["Marque","Leads","→ Cand.","%","→ Admis","%","→ Inscrits","%","Dépense acq.","CPL","CAC (€/inscrit)"]
hr=5
for j,h in enumerate(H,1):
    c=fc.cell(hr,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL); c.alignment=CTR if j>1 else LFT; c.border=bord(bottom=1)
def cref(m,idx):
    row=crm_rows[('2026',m)]; return f"'Données CRM'!{get_column_letter(3+idx)}{row}"
r=hr+1; first=r
for m in Mk:
    fc.cell(r,1,m.title()).font=F(10,True); fc.cell(r,1).alignment=LFT
    fc.cell(r,2,f"={cref(m,0)}")                       # leads
    fc.cell(r,3,f"={cref(m,1)}")                       # cand
    fc.cell(r,4,f"=IFERROR({cref(m,1)}/{cref(m,0)},0)")
    fc.cell(r,5,f"={cref(m,2)}")                       # admis
    fc.cell(r,6,f"=IFERROR({cref(m,2)}/{cref(m,1)},0)")
    fc.cell(r,7,f"={cref(m,3)}")                       # inscrits
    fc.cell(r,8,f"=IFERROR({cref(m,3)}/{cref(m,2)},0)")
    fc.cell(r,9,f"={cref(m,5)}")                       # depense
    fc.cell(r,10,f"=IFERROR({cref(m,5)}/{cref(m,4)},0)")   # CPL=dep/leads_pay
    fc.cell(r,11,f"=IFERROR({cref(m,5)}/{cref(m,3)},0)")   # CAC=dep/inscrits
    for j in range(2,12):
        cc=fc.cell(r,j); cc.alignment=RGT; cc.font=F(10)
        cc.number_format=PCT if j in(4,6,8) else (EUR if j in(9,10,11) else NUM)
    if r%2==0:
        for j in range(1,12): fc.cell(r,j).fill=fill(CARD2)
    r+=1
# ligne Groupe (somme)
fc.cell(r,1,"GROUPE").font=F(10,True,WHITE); fc.cell(r,1).fill=fill(TEALD); fc.cell(r,1).alignment=LFT
for j,col in [(2,'B'),(3,'C'),(5,'E'),(7,'G'),(9,'I')]:
    fc.cell(r,j,f"=SUM({col}{first}:{col}{r-1})")
fc.cell(r,4,f"=IFERROR(C{r}/B{r},0)"); fc.cell(r,6,f"=IFERROR(E{r}/C{r},0)"); fc.cell(r,8,f"=IFERROR(G{r}/E{r},0)")
# CPL & CAC groupe : recalcul sur totaux (depense somme / leadspay somme ; / inscrits somme)
fc.cell(r,10,f"=IFERROR(I{r}/SUM('Données CRM'!G{crm_rows[('2026','MBWAY')]}:G{crm_rows[('2026','TUNON')]}),0)")
fc.cell(r,11,f"=IFERROR(I{r}/G{r},0)")
for j in range(2,12):
    cc=fc.cell(r,j); cc.alignment=RGT; cc.font=F(10,True,WHITE); cc.fill=fill(TEALD)
    cc.number_format=PCT if j in(4,6,8) else (EUR if j in(9,10,11) else NUM)
    cc.border=Border(top=med)
grp=r
# highlight Tunon CAC (fil rouge) en ochre
for i,m in enumerate(Mk):
    if m=='TUNON':
        cc=fc.cell(first+i,11); cc.fill=fill(OCHREBG); cc.font=F(10,True,OCHRE)
        cc2=fc.cell(first+i,1); cc2.font=F(10,True,OCHRE)
note=grp+2
fc.cell(note,1,"FIL ROUGE : Tunon ≈ 564 €/inscrit contre ~327 € MBway — même funnel, coût très supérieur → c'est LÀ qu'on agira (Acte 6).").font=F(9,True,OCHRE)
fc.cell(note+1,1,"Le funnel (taux de passage) explique le CAC : un rapport unique, on ne les sépare pas. Drill cellule campus → funnel du campus (hyperlink POV).").font=F(8,False,FAINT,True)
fc.cell(note+2,1,"CAC groupe 2026 = 353 € (2024 : 329 · 2025 : 341). Dépense +21 % vs inscrits +12 % → dégradation, déjà vue au cockpit.").font=F(8,False,FAINT,True)
for col,w in zip("ABCDEFGHIJK",[10,9,9,7,9,7,10,7,13,9,14]): fc.column_dimensions[col].width=w
fc.freeze_panes="B6"

# =====================================================================
# COCKPIT (écran post-chargement) — tuiles KPI 3 ans + réconciliation + tendance
# =====================================================================
ck=wb.create_sheet("Cockpit"); ck.sheet_view.showGridLines=False
ck["A1"]="COCKPIT D'OUVERTURE  ·  seul écran affiché après le chargement"; ck["A1"].font=F(15,True,INK)
ck["A2"]="Source : V_COCKPIT (cross-source CRM+compta), niveau Groupe. Tableau de bord (pas une matrice) : tuiles KPI + sparklines 3 ans + bandeau réconciliation."; ck["A2"].font=F(9,False,TEALD)
# bandeau réconciliation
ck["A5"]="RÉCONCILIATION  (le héros)"; ck["A5"].font=F(10,True,WHITE)
for col in "ABCDE": ck[col+"5"].fill=fill(TEAL); ck[col+"5"].font=F(10,True,WHITE)
ck["A6"]="CA CRM (socle)"; ck["A6"].font=F(9,False,FAINT)
ck["C6"]="CA Compta (grand livre)"; ck["C6"].font=F(9,False,FAINT)
ck["E6"]="Écart"; ck["E6"].font=F(9,False,FAINT)
prodG=f"SUMIFS('Données P&L'!$G$5:$G${DPMAX},'Données P&L'!$C$5:$C${DPMAX},\"PRODUITS\")"
ck["A7"]=f"={prodG}"; ck["C7"]=f"={prodG}"; ck["E7"]="=A7-C7"
for cc,col in [("A7",INK),("C7",INK),("E7",TEALD)]:
    ck[cc].font=F(14,True,col); ck[cc].number_format=EUR; ck[cc].alignment=LFT
ck["A8"]="En réel les 2 côtés viennent de sources différentes ; écart 0 vérifié (groupe, marque, campus)."; ck["A8"].font=F(8,False,FAINT,True)
# grille KPI 3 ans
hr=10
for j,h in enumerate(["KPI","2024","2025","2026","YoY 25→26"],1):
    c=ck.cell(hr,j,h); c.font=F(9,True,WHITE); c.fill=fill(TEAL); c.alignment=CTR if j>1 else LFT
def sif(n,col): return f"SUMIFS('Données P&L'!${col}$5:${col}${DPMAX},'Données P&L'!$C$5:$C${DPMAX},\"{n}\")"
def caF(col): return sif("PRODUITS",col)
def ebF(col): return f"({caF(col)}-{sif('COUTS_DIRECTS',col)}-{sif('PERSONNEL',col)}-{sif('STRUCTURE',col)}-{sif('IMPOTS_TAXES',col)})"
def crmS(y,dc): a=crm_rows[(y,'MBWAY')]; b=crm_rows[(y,'TUNON')]; return f"SUM('Données CRM'!{dc}{a}:{dc}{b})"
cols3=['E','F','G']; yrs=['2024','2025','2026']
KPI=[("Chiffre d'affaires (€)",[caF(c) for c in cols3],EUR,False),
     ("EBITDA (€)",[ebF(c) for c in cols3],EUR,False),
     ("Marge EBITDA %",[f"{ebF(c)}/{caF(c)}" for c in cols3],PCT,False),
     ("Leads",[crmS(y,'C') for y in yrs],NUM,False),
     ("Inscrits",[crmS(y,'F') for y in yrs],NUM,False),
     ("CAC (€/inscrit)",[f"{crmS(y,'H')}/{crmS(y,'F')}" for y in yrs],EURc,True)]
r=hr+1; first_kpi=r
for lab,series,fmt,ten in KPI:
    ck.cell(r,1,lab).font=F(10,True,OCHRE if ten else INK); ck.cell(r,1).alignment=LFT
    for k in range(3):
        cc=ck.cell(r,2+k,f"={series[k]}"); cc.number_format=fmt; cc.alignment=RGT; cc.font=F(10,False,OCHRE if ten else INK)
    d=ck.cell(r,5)
    if fmt==PCT: d.value=f"=D{r}-C{r}"; d.number_format='0.0" pt"'
    else: d.value=f"=IFERROR((D{r}-C{r})/C{r},0)"; d.number_format=PCT
    d.font=F(9,False,OCHRE if ten else INK,True); d.alignment=RGT
    r+=1
ck.cell(r+1,1,"Sparkline à poser sur B:D de chaque ligne (série 3 ans). CAC = seul KPI en tension (défavorable quand il monte).").font=F(8,False,FAINT,True)
# tendance base 100
tr=r+3
ck.cell(tr,1,"TENDANCE — base 100 en 2024 (source V_TENDANCE)").font=F(10,True,TEALD)
for k,y in enumerate(yrs): ck.cell(tr+1,2+k,int(y)).font=F(9,True); ck.cell(tr+1,2+k).alignment=CTR
ck.cell(tr+1,1,"Année").font=F(9,True)
ck.cell(tr+2,1,"Activité (CA)").font=F(9); ck.cell(tr+3,1,"Dépenses acq.").font=F(9)
for k,c in enumerate(['B','C','D']):
    ck.cell(tr+2,2+k,f"={c}{first_kpi}/$B${first_kpi}*100").number_format='0.0'
    dep=[crmS(y,'H') for y in yrs]
    ck.cell(tr+3,2+k,f"=({dep[k]})/({dep[0]})*100").number_format='0.0'
chart=LineChart(); chart.title="Activité vs Dépenses d'acquisition (base 100)"; chart.height=7; chart.width=13
data=Reference(ck,min_col=1,min_row=tr+2,max_row=tr+3,max_col=4)
cats=Reference(ck,min_col=2,min_row=tr+1,max_col=4,max_row=tr+1)
chart.add_data(data,titles_from_data=True,from_rows=True); chart.set_categories(cats)
ck.add_chart(chart,"G10")
for col,w in zip("ABCDE",[24,13,13,13,12]): ck.column_dimensions[col].width=w

# =====================================================================
# LISEZ-MOI
# =====================================================================
lm=wb.create_sheet("Lisez-moi")
lm.sheet_view.showGridLines=False
lm["A1"]="MAQUETTES DE RAPPORTS  ·  Actes 1 & 2  ·  EDUSERVICES 2027"; lm["A1"].font=F(15,True,INK)
lm["A2"]="Rendu cible pour la restitution Tagetik. Chaque rapport indique la vue qui l'alimente et ce qui reste à construire."; lm["A2"].font=F(10,False,TEALD)
rows=[
 ("",""),
 ("RAPPORT","SOURCE / À CONSTRUIRE"),
 ("Cockpit  (feuille « Cockpit »)","SEUL écran après le chargement. V_COCKPIT (cross-source). Tuiles KPI 3 ans + sparklines + réconciliation + graphe tendance. Tuile CA → 2 drill-through Constitution du CA."),
 ("P&L ①  (feuille « P&L ① »)","FST natif 010-EBITDA sur V_PNL. Hiérarchie Compte fournie (MAPPING_COMPTES.csv). Colonne 2027 → P&L ②③④."),
 ("Funnel & CAC  (feuille)","Matrice multidim sur V_FUNNEL + V_CAC (refaites plates + additives). Un seul rapport. Rien de plus à construire côté SQL."),
 ("",""),
 ("NAVIGATION (drill-through vs hyperlink)",""),
 ("Cockpit · tuile CA","→ drill-through ×2 : Constitution du CA (CRM ‖ Compta)"),
 ("Cockpit · tuile EBITDA / Marge","→ hyperlink : Rapport P&L ①  (POV groupe)"),
 ("Cockpit · tuile CAC","→ hyperlink : Rapport Funnel & CAC  (POV groupe)"),
 ("P&L ① · ligne marque","→ hyperlink : P&L au POV marque → campus"),
 ("Funnel & CAC · cellule campus","→ hyperlink : funnel du campus"),
 ("",""),
 ("LÉGENDE","Bleu = donnée saisie (image de la vue) · Noir = formule · Vert teal = sous-total / EBITDA · Ochre = point de tension"),
]
r=4
for a,b in rows:
    ca=lm.cell(r,1,a); cb=lm.cell(r,2,b)
    if b in("SOURCE / À CONSTRUIRE",) or a in("NAVIGATION (drill-through vs hyperlink)","LÉGENDE"):
        ca.font=F(10,True,WHITE); ca.fill=fill(TEAL); cb.font=F(10,True,WHITE); cb.fill=fill(TEAL)
    else:
        ca.font=F(10,True,INK); cb.font=F(9,False,INK)
    ca.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
    cb.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
    r+=1
lm.column_dimensions["A"].width=32; lm.column_dimensions["B"].width=88

# ordre des feuilles : le récit d'abord, les données ensuite
desired=["Lisez-moi","Cockpit","P&L ①","Funnel & CAC","Données P&L","Données CRM"]
wb._sheets.sort(key=lambda s: desired.index(s.title))

out="/home/user/demo5/eduservices/tagetik/MAQUETTES_RAPPORTS.xlsx"
wb.save(out); print("SAVED",out)
