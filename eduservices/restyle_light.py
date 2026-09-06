#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Re-skin CLAIR & lumineux (finance dataviz) de CAD_SAAD_LIVE.xlsx :
 - fini les aplats sombres : bandeaux/entetes -> teintes claires, texte colore ;
 - accents marque EDUSERVICES ; blanc/air ; cellules a saisir en JAUNE conservees.
 On ne touche QUE la mise en forme (remap de couleurs), jamais les formules."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from copy import copy
F="CAD_SAAD_LIVE.xlsx"; wb=openpyxl.load_workbook(F)

# ---- palette claire : ancien(sombre) -> nouveau(clair) ----
REMAP={
 "1F3864":"E8F0FB",  # navy -> bleu tres clair
 "2E75B6":"D4E6FA",  # bleu -> clair
 "2E86DE":"D4E6FA",
 "548235":"DDF2E3",  # vert -> clair
 "BF8F00":"FBEFD0",  # ambre -> or clair
 "3B3B3B":"ECEEF1",  # gris fonce -> clair
 "843C0C":"F6E2D4",  # brun -> clair
 "7030A0":"ECE0F6",  # violet -> clair
 "C6E0B4":"E4F5EA",  # 'live' -> vert plus clair
 "375623":"1E7A46",  # (texte live) reste lisible
 "DDEBF7":"EAF3FC",
 "E2EFDA":"E9F7EE",
 "FCE4D6":"FCEDE3",
 "E9D9F2":"F0E6F9",
}
KEEP={"FFF2CC","F2F2F2","FFFFFF"}  # jaune saisie, gris bande, blanc : inchanges
# texte accent par famille (pour remplacer le blanc)
DARK_TXT="15406E"     # bleu ardoise fonce, lisible sur toutes les teintes claires
TITLE_TXT="0F5FB0"    # bleu vif pour les titres

def hex6(rgb):
    if not isinstance(rgb,str): return None
    return rgb[-6:].upper()

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            # --- fill ---
            f=c.fill
            if f and f.patternType=="solid" and f.fgColor and isinstance(f.fgColor.rgb,str):
                h=hex6(f.fgColor.rgb)
                if h in REMAP:
                    c.fill=PatternFill("solid",fgColor=REMAP[h])
            # --- font : blanc -> fonce lisible ---
            if c.font and c.font.color and isinstance(c.font.color.rgb,str):
                fh=hex6(c.font.color.rgb)
                if fh=="FFFFFF":
                    big=(c.font.size or 10)>=15
                    c.font=Font(name=c.font.name,size=c.font.size,bold=c.font.bold,
                                italic=c.font.italic,color=TITLE_TXT if big else DARK_TXT)

# ---- accents « dataviz » sur les bandeaux titres ----
accent=Side(style="thick",color="2E86DE")
def band_accent(ws,cellrange_row,cols):
    for col in cols:
        cc=ws["%s%d"%(col,cellrange_row)]
        cc.border=Border(bottom=accent)
band_accent(wb["cad"],2,list("BCDEFGH"))
band_accent(wb["Pilotage"],2,[chr(66+i) for i in range(17)])  # B..R

# ---- input : bord or un peu plus vif, garde le jaune ----
gold=Side(style="medium",color="E0A800")
ibord=Border(left=gold,right=gold,top=gold,bottom=gold)
for ws,cells in [(wb["cad"],['D3','F3','H3']+['K%d'%r for r in range(7,12)]
                   +[c+'%d'%r for r in range(16,27) for c in 'EFG']+['E30','F30','G30']),
                 (wb["Pilotage"],['M%d'%r for r in range(13,27)]+['C52','C53','C54','C55'])]:
    for ref in cells:
        ws[ref].border=ibord

wb.save(F)
print("OK re-skin clair applique (cad + Pilotage + a-cotes). Formules intactes, saisie en jaune.")
