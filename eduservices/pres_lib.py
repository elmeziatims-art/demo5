#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Bibliotheque de mise en forme cabinet-grade + rendu QA (HTML->Chromium).
UNE seule source de verite : PALETTE + fonctions de style. Idempotent.
Rendu : lit le classeur openpyxl et le restitue en HTML fidele (couleurs, bordures,
formats FR, centerContinuous), capture PNG via Chromium. Valeurs illustratives
injectees pour les cellules-formule (le vrai calcul se fait dans Excel)."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

# ============================ PALETTE (unique) ============================
PALETTE = {
    "titre_fond":   "1C2733",   # anthracite
    "titre_txt":    "FFFFFF",
    "entete_fond":  "14586F",   # bleu petrole
    "entete_txt":   "FFFFFF",
    "total_fond":   "1C2733",
    "total_txt":    "FFFFFF",
    "section_fond": "EAF0F3",   # sous-totaux / bandes de section
    "zebra":        "F6F8F9",
    "saisie_fond":  "E4ECFA",
    "saisie_bord":  "4A6FA5",
    "saisie_txt":   "1F3D7A",
    "fav":          "1B6B4F",   # ecart favorable
    "defav":        "8C2A32",   # ecart defavorable
    "note":         "5B6770",
    "bord_int":     "CBD5DA",
    "bord_fort":    "1C2733",
    "corps_txt":    "1C2733",   # texte courant (quasi noir)
}
# convention couleur texte (00_Notice)
TXT_SAISIE = "1F3D7A"   # bleu = saisie en dur
TXT_FORMULE = "1C2733"  # noir = formule locale
TXT_LIEN = "1B6B4F"     # vert = lien vers autre onglet

# ============================ FORMATS NOMBRES (FR) ============================
NF = {
    "euro":   '#,##0\\ €;(#,##0\\ €);-',
    "keuro":  '#,##0\\ "k€";(#,##0\\ "k€");-',
    "pct":    '0.0\\ %',
    "pct0":   '0\\ %',
    "ecart":  '+#,##0;-#,##0;-',
    "ecart_eur": '+#,##0\\ €;-#,##0\\ €;-',
    "eff":    '#,##0;(#,##0);-',
    "num2":   '#,##0.00;(#,##0.00);-',
    "coeff":  '0.00',
    "texte":  '@',
}

ARIAL = "Arial"

# ============================ FABRIQUES DE STYLE ============================
def _c(hex6): return "FF"+hex6

def font(sz=10, bold=False, italic=False, color=None):
    return Font(name=ARIAL, size=sz, bold=bold, italic=italic,
                color=_c(color) if color else _c(PALETTE["corps_txt"]))

def fill(hex6):
    return PatternFill("solid", fgColor=_c(hex6))

def side(hex6, style="thin"):
    return Side(style=style, color=_c(hex6))

def border(bottom=None, top=None, left=None, right=None):
    return Border(bottom=side(*bottom) if bottom else None,
                  top=side(*top) if top else None,
                  left=side(*left) if left else None,
                  right=side(*right) if right else None)

BORD_INT = lambda: border(bottom=(PALETTE["bord_int"],), top=(PALETTE["bord_int"],),
                          left=(PALETTE["bord_int"],), right=(PALETTE["bord_int"],))

def rng_cols(c1, c2):
    a = column_index_from_string(c1); b = column_index_from_string(c2)
    return [get_column_letter(i) for i in range(a, b+1)]

# ---- bandeau de titre (centerContinuous, SANS fusion) ----
def _clear_span_values(ws, row, cols):
    for col in cols:
        ws[col+str(row)].value = None

def style_titre(ws, row, c1, c2, text, height=30):
    cols = rng_cols(c1, c2)
    _clear_span_values(ws, row, cols)
    ws[c1+str(row)] = text
    f = font(14, bold=True, color=PALETTE["titre_txt"])
    pf = fill(PALETTE["titre_fond"])
    for i, col in enumerate(cols):
        cell = ws[col+str(row)]
        cell.font = f; cell.fill = pf
        cell.alignment = Alignment(horizontal="centerContinuous", vertical="center")
    ws.row_dimensions[row].height = height

def style_soustitre(ws, row, c1, c2, text, height=16):
    cols = rng_cols(c1, c2)
    _clear_span_values(ws, row, cols)
    ws[c1+str(row)] = text
    f = font(9, italic=True, color=PALETTE["note"])
    for col in cols:
        cell = ws[col+str(row)]
        cell.font = f
        cell.alignment = Alignment(horizontal="centerContinuous", vertical="center")
    ws.row_dimensions[row].height = height

def style_section(ws, row, c1, c2, text, height=20):
    cols = rng_cols(c1, c2)
    _clear_span_values(ws, row, cols)
    ws[c1+str(row)] = text
    f = font(10, bold=True, color=PALETTE["entete_fond"])
    pf = fill(PALETTE["section_fond"])
    bt = border(top=(PALETTE["bord_fort"], "medium"))
    for col in cols:
        cell = ws[col+str(row)]
        cell.font = f; cell.fill = pf
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = bt
    ws.row_dimensions[row].height = height

def style_entete(ws, row, c1, c2, height=28):
    cols = rng_cols(c1, c2)
    f = font(10, bold=True, color=PALETTE["entete_txt"])
    pf = fill(PALETTE["entete_fond"])
    bb = border(bottom=(PALETTE["bord_fort"], "medium"))
    for col in cols:
        cell = ws[col+str(row)]
        cell.font = f; cell.fill = pf
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bb
    ws.row_dimensions[row].height = height

def style_total(ws, row, c1, c2, height=None):
    cols = rng_cols(c1, c2)
    f = font(10, bold=True, color=PALETTE["total_txt"])
    pf = fill(PALETTE["total_fond"])
    bt = border(top=(PALETTE["bord_fort"], "medium"))
    for col in cols:
        cell = ws[col+str(row)]
        cell.font = f; cell.fill = pf; cell.border = bt
        # alignement: garde gauche pour 1re col (libelle), droite sinon si nombre
    if height: ws.row_dimensions[row].height = height

def style_corps(ws, row, c1, c2, num_cols=(), align_num="right"):
    """corps de tableau: police 10, bordures internes, alignement texte gauche/nombre droite."""
    cols = rng_cols(c1, c2)
    f = font(10, color=PALETTE["corps_txt"])
    bi = BORD_INT()
    for col in cols:
        cell = ws[col+str(row)]
        cell.font = f; cell.border = bi
        if col in num_cols:
            cell.alignment = Alignment(horizontal=align_num, vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

def style_saisie(ws, ref):
    cell = ws[ref]
    cell.fill = fill(PALETTE["saisie_fond"])
    cell.font = font(10, color=PALETTE["saisie_txt"])
    b = side(PALETTE["saisie_bord"])
    cell.border = Border(bottom=b, top=b, left=b, right=b)

def style_note(ws, ref, text=None):
    cell = ws[ref]
    if text is not None: cell.value = text
    cell.font = font(9, italic=True, color=PALETTE["note"])
    cell.alignment = Alignment(horizontal="left", vertical="center")

def set_nf(ws, ref, key):
    ws[ref].number_format = NF[key]

def zebrer(ws, r1, r2, c1, c2, skip=()):
    cols = rng_cols(c1, c2)
    pf = fill(PALETTE["zebra"])
    for r in range(r1, r2+1):
        if r in skip: continue
        if (r - r1) % 2 == 1:
            for col in cols:
                cell = ws[col+str(r)]
                if cell.fill is None or cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = pf
                else:
                    cell.fill = pf

# ---- largeurs sur contenu AFFICHE ----
def fr_number(v, fmt):
    """formate un nombre a la francaise selon une cle NF (pour mesure + rendu)."""
    if v is None: return ""
    if isinstance(v, str): return v
    neg = v < 0; a = abs(v)
    if "pct" in fmt:
        s = ("%.1f" % (a*100)) if fmt == "pct" else ("%.0f" % (a*100))
        s = s.replace(".", ",") + " %"
        return ("-"+s) if neg else s
    if fmt == "coeff":
        return ("%.2f" % a).replace(".", ",")
    if fmt == "num2":
        ent = int(a); dec = a-ent
        s = format(ent, ",d").replace(",", " ") + ("%.2f" % dec)[1:].replace(".", ",")
        return ("("+s+")") if neg else s
    # entiers avec espace + suffixe
    ent = int(round(a))
    s = format(ent, ",d").replace(",", " ")
    if v == 0: return "-"
    if "euro" in fmt: s += " €"
    if "keuro" in fmt: s = format(int(round(a/1000)), ",d").replace(",", " ") + " k€"
    if "ecart" in fmt:
        base = format(int(round(abs(v))), ",d").replace(",", " ")
        if "eur" in fmt: base += " €"
        return ("-"+base) if neg else ("+"+base)
    if neg: return "("+s+")"
    return s

def largeur(displayed_list):
    m = max((len(s) for s in displayed_list if s), default=1)
    return max(10, min(42, round(m*1.15 + 3, 1)))

def set_gutter(ws, col="A", width=2):
    ws.column_dimensions[col].width = width

def clear_zone(ws, r1, r2, c1, c2):
    """reset fill/bordure/police/alignement d'une zone -> idempotence, retire les
    fills residuels de l'onglet source."""
    cols = rng_cols(c1, c2)
    nofill = PatternFill(fill_type=None)
    for r in range(r1, r2+1):
        for col in cols:
            cell = ws[col+str(r)]
            cell.fill = nofill
            cell.border = Border()
            cell.font = font(10)
            cell.alignment = Alignment(horizontal="left", vertical="center")

def sheet_setup(ws, freeze=None, title_rows=None, landscape=True):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100
    ws.sheet_view.topLeftCell = "A1"
    if freeze: ws.freeze_panes = freeze
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True if ws.sheet_properties.pageSetUpPr else None
    try:
        from openpyxl.worksheet.properties import PageSetupProperties
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    except Exception: pass
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.4
    if title_rows: ws.print_title_rows = title_rows
    ws.oddFooter.left.text = "&A"; ws.oddFooter.right.text = "Page &P/&N"

# ============================ RENDU HTML -> PNG ============================
def _argb_to_css(argb):
    if not argb: return None
    s = str(argb)
    if len(s) == 8: return "#"+s[2:]
    if len(s) == 6: return "#"+s
    return None

def _fmt_key_from_numfmt(nf):
    if nf is None: return None
    for k, v in NF.items():
        if v == nf: return k
    if "%" in nf: return "pct"
    if "€" in nf and "k" in nf: return "keuro"
    if "€" in nf: return "euro"
    return None

def render_html(ws, inject=None, max_row=None, max_col=None):
    """genere un HTML fidele de la feuille. inject: dict {ref: value} pour les
    cellules-formule (valeurs illustratives)."""
    inject = inject or {}
    mr = max_row or ws.max_row
    mc = max_col or ws.max_column
    # largeurs colonnes -> px
    def col_px(col):
        w = ws.column_dimensions[col].width
        if not w: w = 8.43
        return int(round(w*7 + 5))
    colspx = {get_column_letter(c): col_px(get_column_letter(c)) for c in range(1, mc+1)}
    html = ['<html><head><meta charset="utf-8"><style>',
            'body{margin:0;background:#ffffff;}',
            'table{border-collapse:collapse;font-family:Arial,Helvetica,sans-serif;table-layout:fixed;}',
            'td{overflow:hidden;padding:0 4px;box-sizing:border-box;}',
            '</style></head><body><table>']
    # colgroup
    html.append("<colgroup>")
    for c in range(1, mc+1):
        html.append('<col style="width:%dpx">' % colspx[get_column_letter(c)])
    html.append("</colgroup>")
    for r in range(1, mr+1):
        if ws.row_dimensions[r].hidden:   # respecte les lignes masquees
            continue
        h = ws.row_dimensions[r].height
        hpx = int(round((h or 15)*96/72))
        html.append('<tr style="height:%dpx">' % hpx)
        c = 1
        while c <= mc:
            col = get_column_letter(c); ref = col+str(r)
            if ws.column_dimensions[col].hidden:   # colonnes masquees
                c += 1; continue
            cell = ws[ref]
            val = cell.value
            if isinstance(val, str) and val.startswith("="):
                val = inject.get(ref, "")
            elif ref in inject:
                val = inject[ref]
            al = cell.alignment
            # run centerContinuous -> colspan
            span = 1
            if al and al.horizontal == "centerContinuous":
                cc = c+1
                while cc <= mc:
                    ncell = ws[get_column_letter(cc)+str(r)]
                    if ncell.alignment and ncell.alignment.horizontal == "centerContinuous" and (ncell.value in (None, "")):
                        span += 1; cc += 1
                    else:
                        break
            # styles
            styles = []
            fnt = cell.font
            if fnt:
                if fnt.b: styles.append("font-weight:bold")
                if fnt.i: styles.append("font-style:italic")
                styles.append("font-size:%dpx" % int(round((fnt.sz or 10)*96/72)))
                col_css = _argb_to_css(fnt.color.rgb if fnt.color else None)
                if col_css: styles.append("color:%s" % col_css)
            fl = cell.fill
            if fl and fl.patternType == "solid" and fl.fgColor is not None:
                bg = _argb_to_css(fl.fgColor.rgb)
                if bg: styles.append("background:%s" % bg)
            # alignement
            ha = "left"
            if al and al.horizontal in ("right", "center", "centerContinuous"):
                ha = "center" if al.horizontal == "centerContinuous" else al.horizontal
            styles.append("text-align:%s" % ha)
            styles.append("vertical-align:%s" % ("middle" if (al and al.vertical == "center") else "bottom"))
            if al and al.wrap_text: styles.append("white-space:normal")
            else: styles.append("white-space:nowrap")
            # bordures
            bd = cell.border
            for s_name, css_name in (("top","border-top"),("bottom","border-bottom"),("left","border-left"),("right","border-right")):
                sd = getattr(bd, s_name, None)
                if sd and sd.style:
                    w = 2 if sd.style in ("medium","thick") else 1
                    col_b = _argb_to_css(sd.color.rgb if sd.color else None) or "#CBD5DA"
                    styles.append("%s:%dpx solid %s" % (css_name, w, col_b))
            # valeur affichee
            disp = val
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                fk = _fmt_key_from_numfmt(cell.number_format)
                disp = fr_number(val, fk) if fk else (format(val, ",").replace(",", " "))
            disp = "" if disp is None else str(disp)
            disp = disp.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            spanattr = (' colspan="%d"' % span) if span > 1 else ""
            html.append('<td%s style="%s">%s</td>' % (spanattr, ";".join(styles), disp))
            c += span
        html.append("</tr>")
    html.append("</table></body></html>")
    return "".join(html)

import os as _os
def _chrome_path():
    for p in ("/opt/pw-browsers/chromium-1194/chrome-linux/chrome",
              "/opt/pw-browsers/chromium_headless_shell-1194/chrome-linux/headless_shell"):
        if _os.path.exists(p): return p
    return None

def screenshot(html, png, width=1920):
    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        b = p.chromium.launch(executable_path=_chrome_path(),
                              args=["--no-sandbox", "--force-color-profile=srgb"])
        pg = b.new_page(viewport={"width": width, "height": 1200}, device_scale_factor=2)
        pg.set_content(html, wait_until="networkidle")
        el = pg.query_selector("table")
        box = el.bounding_box()
        pg.set_viewport_size({"width": min(width, int(box["width"])+40),
                              "height": int(box["height"])+40})
        el.screenshot(path=png)
        b.close()
    return png
