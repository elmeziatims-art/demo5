#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""COCKPIT_DRILL_EDUSERVICES.xlsx — le classeur complet de la démo.
  ① Données          socle campus (14 × 5 marques) — la seule saisie
  ② Données drill    grain cohorte-classe (60 lignes) + listes de membres
  ③ Cockpit          KPI, bridge, marge par marque, tension, portefeuille
  ④ Drill            cascade Marque → Campus → Programme → Année → Modalité
  ⑤ Drill — Pourquoi le pont EBITDA, tel que la requête le renvoie sur une cellule
  ⑥ Drill — Par compte le retour à la compta, même contexte
Tout dérive de socle_reel.py : une seule source de vérité."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from openpyxl.utils import get_column_letter, column_index_from_string
from socle_reel import (construire_classes, ANCRES, POSTES, COUT_VAR_ELEVE, SIEGE_GROUPE)

BLUE="2A78D6"; BLUE2="7AABE6"; BLUE3="C3D9F4"; BLUESOFT="DCE9F9"
ORANGE="EB6834"; ORANGESOFT="FBE3D8"; GOOD="0CA30C"; WARN="FAB219"; CRIT="D03B3B"
INK="131922"; INK2="4D5866"; INK3="7C8798"
CANVAS="EEF1F6"; PANEL="FFFFFF"; PANEL2="F7F9FC"; LINE="DFE4EC"; LINE2="EAEEF4"; WHITE="FFFFFF"
UI="Segoe UI"; MONO="Consolas"
def F(sz=10,b=False,c=INK,f=UI): return Font(name=f,size=sz,bold=b,color=c)
def fill(c): return PatternFill("solid",fgColor=c)
def sd(c=LINE,st="thin"): return Side(style=st,color=c)
L=Alignment("left",vertical="center"); R=Alignment("right",vertical="center")
Cn=Alignment("center",vertical="center"); TOP=Alignment("left",vertical="top",wrap_text=True)
def ind(n): return Alignment("left",vertical="center",indent=n)
EUR='#,##0'; PCT='0.0%'; PCT0='0%'; NUM='#,##0'
DPCT='"▲ "0.0%;"▼ "0.0%'; DPT='"↗ +"0.0" pt";"↘ −"0.0" pt";"→ "0.0" pt"'; MEUR='#,##0.00,," M€"'
def box(ws,r1,c1,r2,c2,bg=PANEL,bd=LINE):
    for r in range(r1,r2+1):
        for c in range(c1,c2+1):
            x=ws.cell(r,c); x.fill=fill(bg)
            x.border=Border(top=sd(bd) if r==r1 else None,bottom=sd(bd) if r==r2 else None,
                            left=sd(bd) if c==c1 else None,right=sd(bd) if c==c2 else None)
def canvas(ws,nr,nc=16):
    for r in range(1,nr):
        for c in range(1,nc): ws.cell(r,c).fill=fill(CANVAS)
def hdr(ws,row,labels,start=1):
    for j,t in enumerate(labels,start):
        c=ws.cell(row,j,t); c.font=F(8.5,True,WHITE); c.fill=fill(INK2); c.alignment=Cn
        c.border=Border(top=sd(INK2),bottom=sd(INK2),left=sd(INK2),right=sd(INK2))

# ===================== DONNÉES =====================
ORDRE=["MBway","ISCOM","Ipac Bachelor Factory","Pigier","Tunon"]
SHORT={"MBway":"MBway","ISCOM":"ISCOM","Ipac Bachelor Factory":"Ipac","Pigier":"Pigier","Tunon":"Tunon"}
CAMP,LG=construire_classes()
CAMP=sorted(CAMP,key=lambda c:(ORDRE.index(c["marque"]),c["ent"]))
LG=sorted(LG,key=lambda l:(ORDRE.index(l["marque"]),l["ent"],l["prog"],l["an"],l["mod"]))
CBY={c["ent"]:c for c in CAMP}

def entiers(vals,total):
    """arrondi à l'entier en conservant le total (plus forts restes)"""
    base=[int(v) for v in vals]; reste=total-sum(base)
    ordre=sorted(range(len(vals)),key=lambda i:-(vals[i]-base[i]))
    for i in ordre[:max(0,reste)]: base[i]+=1
    return base
EFF_C=dict(zip([c["ent"] for c in CAMP],entiers([c["eff"] for c in CAMP],ANCRES["EFF"])))
INS_C=dict(zip([c["ent"] for c in CAMP],entiers([c["inscrits"] for c in CAMP],ANCRES["INSCRITS"])))
EFF_L={}
for c in CAMP:
    sel=[l for l in LG if l["ent"]==c["ent"]]
    for l,v in zip(sel,entiers([x["eff"] for x in sel],EFF_C[c["ent"]])): EFF_L[id(l)]=v

SOCLE=[(c["ent"],SHORT[c["marque"]],round(c["ca"][2024]),round(c["ca"][2025]),round(c["ca"][2026]),
        round(c["eb"][2024]),round(c["eb"][2025]),round(c["eb"][2026]),
        INS_C[c["ent"]],EFF_C[c["ent"]],c["places"],round(c["mix_alt"],4)) for c in CAMP]
MARQUES=[SHORT[m] for m in ORDRE]
S1=15; S2=S1+len(SOCLE)-1; STOT=S2+1
A0=STOT+2; A1=A0+2
B0=A1+2; B1=B0+1; BR=B1+1
M0=BR+7; M1=M0+1; MR=M1+1
T0=MR+len(MARQUES)+1; W1=T0+1; WR=W1+1
X1=WR+7; XR=X1+1
D=lambda col: "Données!$%s$%d:$%s$%d"%(col,S1,col,S2)

wb=openpyxl.Workbook()
d=wb.active; d.title="Données"; d.sheet_view.showGridLines=False
for col,w in zip("ABCDEFGHIJKL",[18,12,14,14,14,14,14,14,11,11,10,11]): d.column_dimensions[col].width=w
d.cell(1,1,"① SOCLE DE DONNÉES — 14 campus × 5 marques  ·  la seule saisie du classeur").font=F(12,True,WHITE)
for c in range(1,13): d.cell(1,c).fill=fill(BLUE)
d.row_dimensions[1].height=24
def h(row,txt): d.cell(row,1,txt).font=F(9.5,True,BLUE)
h(3,"① PÉRIMÈTRE (les filtres du cockpit)"); hdr(d,4,["Filtre","Valeur"])
for i,(k,v) in enumerate([("Scénario","REEL"),("Version","V_FINAL"),("Exercice",2026),("Période","12 — Cumul"),
                          ("Marque","Toutes"),("Campus","Tous"),("Modalité","Toutes")]):
    d.cell(5+i,1,k).font=F(9.5); d.cell(5+i,2,v).font=F(9.5,True)
    d.cell(5+i,1).alignment=L; d.cell(5+i,2).alignment=L
h(13,"② SOCLE CAMPUS  (source : V_COCKPIT / V_ALLOCATION, grain campus × exercice)")
hdr(d,14,["Entity","Marque","CA 2024","CA 2025","CA 2026","EBITDA 2024","EBITDA 2025","EBITDA 2026",
          "Inscrits 26","Effectifs 26","Places 26","Mix alt. 26"])
for i,row in enumerate(SOCLE):
    r=S1+i
    for j,v in enumerate(row,1):
        c=d.cell(r,j,v); c.font=F(9.5); c.border=Border(bottom=sd(LINE2)); c.alignment=L if j<=2 else R
        c.number_format=EUR if 3<=j<=8 else ('#,##0' if j in (9,10,11) else (PCT0 if j==12 else 'General'))
d.cell(STOT,1,"TOTAL GROUPE").font=F(9.5,True)
for j in range(3,12):
    c=d.cell(STOT,j,"=SUM(%s%d:%s%d)"%(get_column_letter(j),S1,get_column_letter(j),S2))
    c.font=F(9.5,True); c.number_format=EUR if j<9 else '#,##0'; c.alignment=R; c.border=Border(top=sd(INK3))
h(A0-1,"③ ACQUISITION (groupe)"); hdr(d,A0,["Mesure","2024","2025","2026"])
for i,(lab,a,b,c_,fm) in enumerate([("Dépenses d'acquisition (€)",358170,394702,434174,EUR),
                                    ("Inscrits (nouveaux)",1092,1159,1229,'#,##0')]):
    r=A0+1+i
    d.cell(r,1,lab).font=F(9.5); d.cell(r,1).alignment=L
    for j,v in enumerate((a,b,c_),2):
        x=d.cell(r,j,v); x.number_format=fm; x.font=F(9.5); x.alignment=R; x.border=Border(bottom=sd(LINE))
h(B0-1,"④ BRIDGE EBITDA GROUPE 2025 → 2026  (l'effet coûts est le résidu : le pont boucle toujours)")
hdr(d,B0,["Étape","Montant (€)"])
for i,(lab,v) in enumerate([("EBITDA 2025","=SUM($G$%d:$G$%d)"%(S1,S2)),("Effet activité",211706),
        ("Effet prix / mix",289000),("Effet coûts (résidu)","=B%d-B%d-B%d-B%d"%(BR+4,BR,BR+1,BR+2)),
        ("EBITDA 2026","=SUM($H$%d:$H$%d)"%(S1,S2))]):
    r=BR+i
    d.cell(r,1,lab).font=F(9.5,i==4); d.cell(r,1).alignment=L
    c=d.cell(r,2,v); c.number_format=EUR; c.alignment=R; c.font=F(9.5,i==4); c.border=Border(bottom=sd(LINE))
h(M0-1,"⑤ AGRÉGATS PAR MARQUE — 100 % calculés depuis ②")
hdr(d,M0,["Marque","CA 2024","CA 2025","CA 2026","EBITDA 2024","EBITDA 2025","EBITDA 2026",
          "Marge 2024","Marge 2025","Marge 2026"])
for i,m in enumerate(MARQUES):
    r=MR+i
    d.cell(r,1,m).font=F(9.5); d.cell(r,1).alignment=L
    for j,src in zip(range(2,8),"CDEFGH"):
        c=d.cell(r,j,'=SUMIF(%s,$A%d,%s)'%(D("B"),r,D(src))); c.number_format=EUR; c.font=F(9.5); c.alignment=R
    for j,(num,den) in zip(range(8,11),[(5,2),(6,3),(7,4)]):
        c=d.cell(r,j,"=%s%d/%s%d"%(get_column_letter(num),r,get_column_letter(den),r))
        c.number_format=PCT; c.font=F(9.5,True); c.alignment=R
    for j in range(1,11): d.cell(r,j).border=Border(bottom=sd(LINE2))
h(T0,"⑥ ZONE TECHNIQUE — séries des graphes (ne pas saisir)")
hdr(d,W1,["Étape (waterfall)","Socle","Ancre","Hausse","Baisse"])
for i,(cat,so,an,ha,ba) in enumerate([("EBITDA 2025","0","=B%d"%BR,"0","0"),
        ("Activité","=B%d"%BR,"0","=B%d"%(BR+1),"0"),
        ("Prix / mix","=B%d+B%d"%(BR,BR+1),"0","=B%d"%(BR+2),"0"),
        ("Coûts","=B%d"%(BR+4),"0","0","=B%d+B%d+B%d-B%d"%(BR,BR+1,BR+2,BR+4)),
        ("EBITDA 2026","0","=B%d"%(BR+4),"0","0")]):
    r=WR+i
    d.cell(r,1,cat).font=F(9,c=INK3); d.cell(r,1).alignment=L
    for j,v in enumerate((so,an,ha,ba),2):
        c=d.cell(r,j,v); c.number_format=EUR; c.font=F(9,c=INK3); c.alignment=R
hdr(d,X1,["Exercice","Dépenses (base 100)","Inscrits (base 100)"])
for i,(ex,col) in enumerate([(2024,"B"),(2025,"C"),(2026,"D")]):
    r=XR+i
    d.cell(r,1,ex).font=F(9,c=INK3); d.cell(r,1).alignment=Cn; d.cell(r,1).number_format='0'
    d.cell(r,2,"=%s%d/$B$%d*100"%(col,A0+1,A0+1)).number_format='0.0'
    d.cell(r,3,"=%s%d/$B$%d*100"%(col,A1,A1)).number_format='0.0'
    for j in (2,3): d.cell(r,j).font=F(9,c=INK3); d.cell(r,j).alignment=R
# ===================== ② DONNÉES DRILL — grain cohorte =====================
dd=wb.create_sheet("Données drill"); dd.sheet_view.showGridLines=False
LARG=[9,12,11,7,7,6,6,7,8,8,12,12,11,11,12,12,11,12,12,12,12,12,7,6]
for col,w in zip("ABCDEFGHIJKLMNOPQRSTUVWX",LARG): dd.column_dimensions[col].width=w
dd.cell(1,1,"② DONNÉES DU DRILL — grain COHORTE-CLASSE  (V_ALLOCATION / V_CAMPUS_CLASSE)").font=F(12,True,WHITE)
for c in range(1,25): dd.cell(1,c).fill=fill(BLUE)
dd.row_dimensions[1].height=24
dd.cell(3,1,"Une ligne = une cohorte-classe. Tout remonte exactement au socle campus de l'onglet ①.").font=F(9.5,True,BLUE)
hdr(dd,4,["Marque","Campus","Programme","Année","Modalité","Nb cl.","Capa.","Places",
          "Effectif 26","Effectif 25","CA 2025","CA 2026","Coût var. 25","Coût var. 26",
          "Coût dir. 25","Coût dir. 26","Coût siège","EBITDA 2025","EBITDA 2026",
          "Contribution","Coût complet","Marge complète","Inclus","Rang"])
K1=5; K2=K1+len(LG)-1
DD=lambda col: "'Données drill'!$%s$%d:$%s$%d"%(col,K1,col,K2)
for i,l in enumerate(LG):
    r=K1+i
    vals=[SHORT[l["marque"]],l["ent"],l["prog"],l["an"],l["mod"],l["ncl"],l["cap"],"=F%d*G%d"%(r,r),
          EFF_L[id(l)],round(l["eff25"]),round(l["ca25"]),round(l["ca26"]),
          round(l["cvar25"]),round(l["cvar26"]),round(l["cdir25"]),round(l["cdir26"]),
          round(l["siege26"]),"=K%d-M%d-O%d"%(r,r,r),"=L%d-N%d-P%d"%(r,r,r),
          "=L%d-N%d"%(r,r),"=N%d+P%d+Q%d"%(r,r,r),"=L%d-U%d"%(r,r),
          "=IF(AND(OR(Drill!$C$6=\"(Tous)\",$A{0}=Drill!$C$6),OR(Drill!$D$6=\"(Tous)\",$B{0}=Drill!$D$6),"
          "OR(Drill!$E$6=\"(Tous)\",$C{0}=Drill!$E$6),OR(Drill!$F$6=\"(Tous)\",$D{0}=Drill!$F$6),"
          "OR(Drill!$G$6=\"(Tous)\",$E{0}=Drill!$G$6)),1,0)".format(r),
          "=IF(W%d=1,SUM($W$%d:W%d),\"\")"%(r,K1,r)]
    for j,v in enumerate(vals,1):
        c=dd.cell(r,j,v); c.font=F(9); c.border=Border(bottom=sd(LINE2)); c.alignment=L if j<=5 else R
        c.number_format=EUR if j>=11 else '#,##0'
        if j in (23,24): c.font=F(8,c=INK3)
dd.cell(K2+1,1,"TOTAL").font=F(9,True)
for j in list(range(6,23)):
    c=dd.cell(K2+1,j,"=SUM(%s%d:%s%d)"%(get_column_letter(j),K1,get_column_letter(j),K2))
    c.font=F(9,True); c.number_format=EUR if j>=11 else '#,##0'; c.alignment=R; c.border=Border(top=sd(INK3))
DIMS=["Marque","Campus","Programme","Année d'étude","Modalité"]
LISTS=[MARQUES,[c["ent"] for c in CAMP],sorted({l["prog"] for l in LG}),
       sorted({l["an"] for l in LG},key=lambda a:("BM".find(a[0]),a)),sorted({l["mod"] for l in LG})]
MCOL=26; VCOL=34; NIV=32                       # membres Z.., niveaux AF, validation AH..
dd.cell(4,MCOL,"— membres —").font=F(8,True,INK3)
for k,(nm,lst) in enumerate(zip(DIMS,LISTS)):
    dd.cell(K1+k,NIV,nm).font=F(8,c=INK3)
    for i,v in enumerate(lst): dd.cell(K1+i,MCOL+k,v).font=F(8,c=INK3)
    dd.cell(K1-1,VCOL+k,nm).font=F(8,True,INK3)
    dd.cell(K1,VCOL+k,"(Tous)").font=F(8,c=INK3)
    for i,v in enumerate(lst): dd.cell(K1+1+i,VCOL+k,v).font=F(8,c=INK3)
NMAX=max(len(x) for x in LISTS)
dd.cell(K2+3,1,"Clés de ventilation comptable (onglet ⑥) — les postes du coût direct n'évoluent pas au même rythme").font=F(9,True,BLUE)
hdr(dd,K2+4,["Compte","Libellé","Assiette","Part 2025","Part 2026"])
SW=sum(w*g for _,_,w,g in POSTES)
VENT=[("706","Scolarité initiale","CA × 0,97 × (1−mix alt.)",None,None),
      ("7062","Scolarité alternance","CA × 0,97 × mix alt.",None,None),
      ("708","Produits annexes","CA",0.03,0.03),
      ("621","Vacataires","Coût variable",0.85,0.85),
      ("604","Achats de prestations","Coût variable",0.15,0.15)]+[
      (p,lib,"Coût direct",w,w*g/SW) for p,lib,w,g in POSTES]
for i,(cp,lib,ass,p25,p26) in enumerate(VENT):
    r=K2+5+i
    for j,v in enumerate((cp,lib,ass,p25,p26),1):
        c=dd.cell(r,j,v); c.font=F(9,c=INK2); c.alignment=L if j<=3 else R
        if isinstance(v,float): c.number_format='0.00%'
        c.border=Border(bottom=sd(LINE2))
VR=K2+5

# ===================== ③ COCKPIT =====================
w=wb.create_sheet("Cockpit"); w.sheet_view.showGridLines=False
w.column_dimensions["A"].width=2; w.column_dimensions["B"].width=22
for col in "CDEFGHIJKLM": w.column_dimensions[col].width=12
w.column_dimensions["N"].width=2
G0=35
ROWS=[("GROUPE",0)]
for m in MARQUES:
    ROWS.append((m,1)); ROWS+=[(s[0],2) for s in SOCLE if s[1]==m]
GEND=G0+len(ROWS)-1; CAL=GEND+2
for r,hh in {1:6,2:30,3:16,4:6,5:20,6:6,7:15,8:28,9:16,10:8,11:20,32:8,33:20,34:18,GEND+1:8,CAL+4:6}.items():
    w.row_dimensions[r].height=hh
for r in range(12,32): w.row_dimensions[r].height=16
for r in range(G0,GEND+1): w.row_dimensions[r].height=17
canvas(w,CAL+8)
box(w,2,2,3,13)
w.cell(2,2,"  ③ Cockpit de performance — EDUSERVICES GROUP").font=F(14,True,INK); w.cell(2,2).alignment=L
w.cell(3,2,"  Constat 2024 → 2026 · 14 campus · 5 marques · socle CRM & comptabilité rapprochés").font=F(9.5,c=INK3)
w.cell(3,2).alignment=L
w.cell(2,13,"RÉEL · CLÔTURE 2026  ").font=F(9,True,BLUE); w.cell(2,13).alignment=R
for i,(lab,ref) in enumerate([("Scénario","=Données!B5"),("Version","=Données!B6"),("Exercice","=Données!B7"),
                              ("Période","=Données!B8"),("Marque","=Données!B9"),("Campus","=Données!B10")]):
    c1=2+i*2
    box(w,5,c1,5,c1+1,bg=(BLUESOFT if lab=="Exercice" else PANEL),bd=(BLUE if lab=="Exercice" else LINE))
    a=w.cell(5,c1,"  "+lab); a.font=F(9,c=INK3); a.alignment=L
    b=w.cell(5,c1+1,ref); b.font=F(9.5,True,BLUE if lab=="Exercice" else INK); b.alignment=L
KPI=[("Chiffre d'affaires","=C%d"%G0,MEUR,"=D%d"%G0,DPCT,False),
     ("EBITDA","=E%d"%G0,MEUR,"=F%d"%G0,DPCT,False),
     ("Marge EBITDA","=H%d"%G0,PCT,"=I%d"%G0,DPT,False),
     ("Inscrits (nouveaux)","=J%d"%G0,'#,##0',"=J%d/Données!C%d-1"%(G0,A1),DPCT,False),
     ("Coût d'acquisition","=Données!D%d/Données!D%d"%(A0+1,A1),'#,##0" €"',
      "=(Données!D%d/Données!D%d)/(Données!C%d/Données!C%d)-1"%(A0+1,A1,A0+1,A1),DPCT,True),
     ("Remplissage moyen","=K%d"%G0,PCT0,'=SUM(%s)-SUM(%s)'%(D("K"),D("J")),'#,##0" places libres"',False)]
for i,(lab,val,fm,dv,dfm,alert) in enumerate(KPI):
    c1=2+i*2
    box(w,7,c1,9,c1+1,bd=(WARN if alert else LINE))
    a=w.cell(7,c1,"  "+lab.upper()); a.font=F(8.5,True,INK3); a.alignment=L
    v=w.cell(8,c1,val); v.font=F(20,True,INK,MONO); v.number_format=fm; v.alignment=ind(1)
    x=w.cell(9,c1,dv); x.font=F(9.5,True,INK3 if i==5 else INK); x.number_format=dfm; x.alignment=ind(1)
    if i<5:
        for op,cc in (("greaterThan",GOOD),("lessThan",CRIT)):
            w.conditional_formatting.add("%s9"%get_column_letter(c1),
                CellIsRule(operator=op,formula=["0"],font=Font(name=UI,size=9.5,bold=True,color=cc)))
w.conditional_formatting.add("J9",CellIsRule(operator="greaterThan",formula=["0"],
    font=Font(name=UI,size=9.5,bold=True,color=CRIT)))
for c1,c2,title,hint in [(2,5,"Bridge EBITDA 2025 → 2026","M€ · axe tronqué à 3,2"),
                         (6,9,"Marge EBITDA par marque","2024 · 2025 · 2026"),
                         (10,13,"Tension acquisition","base 100 = 2024")]:
    box(w,11,c1,31,c2)
    t=w.cell(11,c1,"  "+title); t.font=F(10.5,True,INK); t.alignment=L
    hn=w.cell(11,c2,hint+"  "); hn.font=F(8.5,c=INK3); hn.alignment=R
def waterfall(ws,ref_sheet,r0,anchor,ymin,ymax,unit,fmt_anchor,title=None,width=10.2):
    ch=BarChart(); ch.type="col"; ch.grouping="stacked"; ch.overlap=100; ch.gapWidth=45
    ch.height=7.9; ch.width=width; ch.legend=None; ch.visible_cells_only=False
    ch.y_axis.numFmt=unit; ch.y_axis.scaling.min=ymin; ch.y_axis.scaling.max=ymax
    ch.y_axis.majorUnit=(ymax-ymin)/7; ch.x_axis.delete=False; ch.y_axis.delete=False
    if title: ch.title=title
    for col in (2,3,4,5):
        ch.add_data(Reference(ref_sheet,min_col=col,max_col=col,min_row=r0,max_row=r0+4),titles_from_data=False)
    a,b,c_,e=ch.series
    a.graphicalProperties.noFill=True; a.graphicalProperties.line.noFill=True
    for s,cc in ((b,BLUE),(c_,GOOD),(e,CRIT)):
        s.graphicalProperties.solidFill=cc; s.graphicalProperties.line.solidFill=cc
    for s,fm in ((b,fmt_anchor),(c_,'"+ "#,##0" €"'),(e,'"− "#,##0" €"')):
        s.dLbls=DataLabelList(); s.dLbls.showVal=True; s.dLbls.numFmt=fm; s.dLbls.dLblPos="ctr"
        s.dLbls.showSerName=False; s.dLbls.showCatName=False; s.dLbls.showLegendKey=False
    ch.set_categories(Reference(ref_sheet,min_col=1,max_col=1,min_row=r0,max_row=r0+4))
    ws.add_chart(ch,anchor)
waterfall(w,d,WR,"B12",3200000,4000000,'#,##0,," M€"','#,##0,,.000" M€"')
mg=BarChart(); mg.type="col"; mg.grouping="clustered"; mg.gapWidth=55; mg.overlap=-10
mg.height=7.9; mg.width=8.4; mg.y_axis.numFmt='0%'; mg.visible_cells_only=False
mg.y_axis.scaling.min=0; mg.y_axis.scaling.max=0.24; mg.y_axis.majorUnit=0.05
mg.x_axis.delete=False; mg.y_axis.delete=False
for col in (8,9,10): mg.add_data(Reference(d,min_col=col,max_col=col,min_row=M0,max_row=MR+len(MARQUES)-1),titles_from_data=True)
for s,cc in zip(mg.series,(BLUE3,BLUE2,BLUE)):
    s.graphicalProperties.solidFill=cc; s.graphicalProperties.line.solidFill=cc
mg.series[2].dLbls=DataLabelList(); mg.series[2].dLbls.showVal=True
mg.series[2].dLbls.numFmt='0.0%'; mg.series[2].dLbls.dLblPos="outEnd"
mg.series[2].dLbls.showSerName=False; mg.series[2].dLbls.showCatName=False; mg.series[2].dLbls.showLegendKey=False
mg.set_categories(Reference(d,min_col=1,max_col=1,min_row=MR,max_row=MR+len(MARQUES)-1))
mg.legend.position="b"; w.add_chart(mg,"F12")
tn=LineChart(); tn.height=7.9; tn.width=8.4; tn.visible_cells_only=False
tn.y_axis.numFmt='0'; tn.y_axis.scaling.min=95; tn.y_axis.scaling.max=125; tn.y_axis.majorUnit=10
tn.x_axis.delete=False; tn.y_axis.delete=False
for col in (2,3): tn.add_data(Reference(d,min_col=col,max_col=col,min_row=X1,max_row=XR+2),titles_from_data=True)
for s,cc in zip(tn.series,(ORANGE,BLUE)):
    s.graphicalProperties.line.solidFill=cc; s.graphicalProperties.line.width=25000
    s.marker=Marker(symbol="circle",size=6); s.smooth=False
    s.marker.graphicalProperties.solidFill=cc; s.marker.graphicalProperties.line.solidFill=cc
tn.set_categories(Reference(d,min_col=1,max_col=1,min_row=XR,max_row=XR+2))
tn.legend.position="b"; w.add_chart(tn,"J12")
box(w,33,2,GEND,13)
w.cell(33,2,"  Portefeuille — marque & campus").font=F(10.5,True,INK); w.cell(33,2).alignment=L
w.cell(33,13,"exercice 2026 · variation vs 2025  ").font=F(8.5,c=INK3); w.cell(33,13).alignment=R
for j,t in enumerate(["Marque / campus","CA 2026","Δ CA","EBITDA","Δ EBITDA","Part EBITDA","Marge EBITDA",
                      "Δ marge","Inscrits","Rempl.","Mix alt.","Alerte"],2):
    c=w.cell(34,j,t); c.font=F(8.5,True,INK3); c.fill=fill(PANEL2); c.alignment=L if j==2 else R
    c.border=Border(bottom=sd(LINE),top=sd(LINE))
SRC="Données!$A$%d:$L$%d"%(S1,S2); MK=D("B")
FMT={"C":EUR,"D":DPCT,"E":EUR,"F":DPCT,"G":'0.0%;-0.0%',"H":'0.0%;-0.0%',"I":DPT,
     "J":'#,##0',"K":PCT0,"L":PCT0,"M":"General"}
for i,(name,lvl) in enumerate(ROWS):
    r=G0+i
    w.cell(r,2,name).alignment=ind(lvl)
    w.cell(r,2).font=F(10 if lvl==0 else 9.5,lvl<=1,INK if lvl<2 else INK2)
    if lvl==0:
        f={"C":"=SUM(%s)"%D("E"),"D":"=C%d/SUM(%s)-1"%(r,D("D")),"E":"=SUM(%s)"%D("H"),
           "F":"=E%d/SUM(%s)-1"%(r,D("G")),"G":"=E%d/$E$%d"%(r,G0),"H":"=E%d/C%d"%(r,r),
           "I":"=(H%d-SUM(%s)/SUM(%s))*100"%(r,D("G"),D("D")),"J":"=SUM(%s)"%D("I"),
           "K":"=SUM(%s)/SUM(%s)"%(D("J"),D("K")),
           "L":"=SUMPRODUCT(%s,%s)/SUM(%s)"%(D("J"),D("L"),D("J")),"M":'=IF(H%d<0.08,"FRAGILE","")'%r}
    elif lvl==1:
        S=lambda col: 'SUMIF(%s,$B%d,%s)'%(MK,r,D(col))
        f={"C":"="+S("E"),"D":"=C%d/%s-1"%(r,S("D")),"E":"="+S("H"),"F":"=E%d/%s-1"%(r,S("G")),
           "G":"=E%d/$E$%d"%(r,G0),"H":"=E%d/C%d"%(r,r),"I":"=(H%d-%s/%s)*100"%(r,S("G"),S("D")),
           "J":"="+S("I"),"K":"=%s/%s"%(S("J"),S("K")),
           "L":"=SUMPRODUCT((%s=$B%d)*%s*%s)/%s"%(MK,r,D("J"),D("L"),S("J")),
           "M":'=IF(H%d<0.08,"FRAGILE","")'%r}
    else:
        V=lambda n: 'VLOOKUP($B%d,%s,%d,0)'%(r,SRC,n)
        f={"C":"="+V(5),"D":"=C%d/%s-1"%(r,V(4)),"E":"="+V(8),
           "F":'=IF(%s<=0,"n/s",E%d/%s-1)'%(V(7),r,V(7)),"G":"=E%d/$E$%d"%(r,G0),"H":"=E%d/C%d"%(r,r),
           "I":"=(H%d-%s/%s)*100"%(r,V(7),V(4)),"J":"="+V(9),"K":"=%s/%s"%(V(10),V(11)),"L":"="+V(12),
           "M":'=IF(H%d<0.08,"FRAGILE","")'%r}
    for col,formula in f.items():
        j=column_index_from_string(col)
        c=w.cell(r,j,formula); c.number_format=FMT[col]; c.alignment=R if col!="M" else Cn
        c.font=F(9.5,lvl<=1,INK if lvl<2 else INK2,MONO if col!="M" else UI)
    w.cell(r,13).font=F(8,True,ORANGE)
    for j in range(2,14):
        w.cell(r,j).border=Border(bottom=sd(LINE2))
        if lvl==0: w.cell(r,j).fill=fill(PANEL2)
    if lvl==2: w.row_dimensions[r].outlineLevel=2
    elif lvl==1: w.row_dimensions[r].outlineLevel=1
w.sheet_properties.outlinePr.summaryBelow=False
w.conditional_formatting.add("H%d:H%d"%(G0,GEND),
    DataBarRule(start_type="num",start_value=0,end_type="num",end_value=0.24,color="FF"+BLUE,showValue=True))
for col in "DFI":
    for op,cc in (("greaterThan",GOOD),("lessThan",CRIT)):
        w.conditional_formatting.add("%s%d:%s%d"%(col,G0,col,GEND),
            CellIsRule(operator=op,formula=["0"],font=Font(name=MONO,size=9.5,color=cc)))
w.conditional_formatting.add("K%d:K%d"%(G0,GEND),CellIsRule(operator="lessThan",formula=["0.70"],
    font=Font(name=MONO,size=9.5,bold=True,color=WARN)))
box(w,CAL,2,CAL+2,13,bg=PANEL2)
w.merge_cells(start_row=CAL,start_column=2,end_row=CAL+2,end_column=13)
w.cell(CAL,2,"  Le signal du cockpit.  Tunon pèse 8,4 % du CA mais 2,1 % de l'EBITDA. Ipac, le réseau le plus jeune, "
             "pèse 11,2 % du CA pour 9,5 % de l'EBITDA — il monte en charge, Rennes et Montpellier encore à 61 % de "
             "remplissage. À l'inverse Pigier, la plus petite des cinq, est la plus rentable (21,2 %). Et le groupe "
             "n'occupe que 75 % de ses places : plus de 1 000 places ouvertes et vides. Le cadrage 2027 ne peut pas "
             "être un « + x % » uniforme — c'est ici que le drill prend le relais.").font=F(9.5,c=INK2)
w.cell(CAL,2).alignment=TOP
for r in range(CAL,CAL+3): w.cell(r,2).border=Border(left=sd(ORANGE,"thick"))
w.cell(CAL+4,2,"Source : socle CRM et comptabilité rapprochés · les marges se lisent en POINTS. "
               "Tout le cockpit est en formules : seuls les onglets ① et ② se saisissent.").font=F(8.5,c=INK3)
w.sheet_view.zoomScale=90

# ===================== ④ DRILL — cascade =====================
dr=wb.create_sheet("Drill"); dr.sheet_view.showGridLines=False
dr.column_dimensions["A"].width=2; dr.column_dimensions["B"].width=34
for col in "CDEFGHIJKLM": dr.column_dimensions[col].width=12
dr.column_dimensions["N"].width=3; dr.column_dimensions["N"].hidden=True
dr.column_dimensions["O"].width=3; dr.column_dimensions["O"].hidden=True
NDET=20; DET0=32
for r,hh in {1:6,2:28,3:16,4:8,5:15,6:22,7:8,8:18,9:8,10:20,11:18,12:24,13:10,14:20,15:18,
             29:10,30:20,31:18}.items(): dr.row_dimensions[r].height=hh
for r in list(range(16,29))+list(range(DET0,DET0+NDET)): dr.row_dimensions[r].height=16
canvas(dr,DET0+NDET+10)
box(dr,2,2,3,13)
dr.cell(2,2,"  ④ Drill EBITDA — du groupe à la classe").font=F(14,True,INK); dr.cell(2,2).alignment=L
dr.cell(3,2,"  Marque → Campus → Programme → Année d'étude → Modalité · exercice 2026").font=F(9.5,c=INK3)
dr.cell(3,2).alignment=L
dr.cell(2,13,"CONTRIBUTION vs COÛT COMPLET  ").font=F(9,True,ORANGE); dr.cell(2,13).alignment=R
box(dr,5,2,6,13)
dr.cell(5,2,"  ① CHEMIN DE DRILL").font=F(9,True,INK3); dr.cell(5,2).alignment=L
dr.cell(6,2,"  remets « (Tous) » pour remonter d'un niveau").font=F(8.5,c=INK3); dr.cell(6,2).alignment=L
for k,nm in enumerate(DIMS):
    cc=3+k
    lb=dr.cell(5,cc,nm); lb.font=F(8.5,True,INK3); lb.alignment=Cn
    sel=dr.cell(6,cc,"(Tous)"); sel.font=F(10,True,BLUE); sel.alignment=Cn; sel.fill=fill(BLUESOFT)
    sel.border=Border(top=sd(BLUE),bottom=sd(BLUE),left=sd(BLUE),right=sd(BLUE))
    col=get_column_letter(VCOL+k)
    dv=DataValidation(type="list",allow_blank=False,
        formula1="='Données drill'!$%s$%d:$%s$%d"%(col,K1,col,K1+len(LISTS[k])))
    dr.add_data_validation(dv); dv.add(sel)
dr.cell(6,15,'=5-COUNTIF($C$6:$G$6,"(Tous)")+1').font=F(8,c=INK3)
dr.cell(8,2,'="②  Vous regardez : "&IF($O$6=1,"le GROUPE (tous périmètres)",'
            'IF($C$6="(Tous)","",$C$6)&IF($D$6="(Tous)",""," › "&$D$6)&IF($E$6="(Tous)",""," › "&$E$6)'
            '&IF($F$6="(Tous)",""," › "&$F$6)&IF($G$6="(Tous)",""," › "&$G$6))').font=F(11,True,INK)
dr.cell(8,2).alignment=L
MES=["CA 2026","Δ CA","EBITDA","Marge EBITDA","Contribution","Coût complet","Marge complète",
     "Effectifs","Places","Remplissage","Statut"]
FMD=[EUR,DPCT,EUR,PCT,EUR,EUR,EUR,'#,##0','#,##0',PCT0,"General"]
def entete(ws,row,first):
    c=ws.cell(row,2,first); c.font=F(8.5,True,INK3); c.alignment=L; c.border=Border(bottom=sd(LINE))
    for j,t in enumerate(MES,3):
        c=ws.cell(row,j,t); c.font=F(8.5,True,INK3); c.alignment=R; c.border=Border(bottom=sd(LINE))
box(dr,10,2,12,13); dr.cell(10,2,"  ② NIVEAU COURANT").font=F(9,True,INK3); dr.cell(10,2).alignment=L
entete(dr,11,"Agrégat de la sélection")
SP=lambda col: "SUMPRODUCT(%s,%s)"%(DD("W"),DD(col))
cur={"C":"="+SP("L"),"D":"=C12/%s-1"%SP("K"),"E":"="+SP("S"),"F":"=IFERROR(E12/C12,0)",
     "G":"="+SP("T"),"H":"="+SP("U"),"I":"="+SP("V"),"J":"="+SP("I"),"K":"="+SP("H"),
     "L":"=IFERROR(J12/K12,0)",
     "M":'=IF(E12<0,"DÉFICITAIRE",IF(I12<0,"PIÈGE",IF(L12<0.7,"SOUS-REMPLI","SAIN")))'}
for col,f in cur.items():
    j=column_index_from_string(col)
    c=dr.cell(12,j,f); c.font=F(11,True,INK,MONO if col!="M" else UI)
    c.number_format=FMD[j-3]; c.alignment=R if col!="M" else Cn
dr.cell(12,2,'=IF($O$6=1,"GROUPE",INDEX($C$6:$G$6,$O$6-1))').font=F(11,True,INK); dr.cell(12,2).alignment=L
box(dr,14,2,28,13)
dr.cell(14,2,'="  ③ NIVEAU SUIVANT — "&IF($O$6>5,"grain classe atteint, voir le détail ci-dessous",'
             "UPPER(INDEX('Données drill'!$%s$%d:$%s$%d,$O$6)))"%(get_column_letter(NIV),K1,get_column_letter(NIV),K1+4)).font=F(9,True,INK3)
dr.cell(14,2).alignment=L
entete(dr,15,"Membre")
MSK="%s*(INDEX('Données drill'!$A$%d:$E$%d,0,$O$6)=$B{r})"%(DD("W"),K1,K2)
for i in range(NMAX):
    r=16+i
    dr.cell(r,2,'=IF($O$6>5,"",IFERROR(INDEX(\'Données drill\'!$%s$%d:$%s$%d,%d,$O$6),""))'
            %(get_column_letter(MCOL),K1,get_column_letter(MCOL+4),K1+NMAX-1,i+1)).font=F(9.5,True)
    dr.cell(r,2).alignment=L
    m=MSK.format(r=r)
    nf={"C":"SUMPRODUCT(%s*%s)"%(m,DD("L")),"D":"C{r}/SUMPRODUCT({m}*{j})-1".format(r=r,m=m,j=DD("K")),
        "E":"SUMPRODUCT(%s*%s)"%(m,DD("S")),"F":"E{r}/C{r}".format(r=r),
        "G":"SUMPRODUCT(%s*%s)"%(m,DD("T")),"H":"SUMPRODUCT(%s*%s)"%(m,DD("U")),
        "I":"SUMPRODUCT(%s*%s)"%(m,DD("V")),"J":"SUMPRODUCT(%s*%s)"%(m,DD("I")),
        "K":"SUMPRODUCT(%s*%s)"%(m,DD("H")),"L":"J{r}/K{r}".format(r=r),
        "M":'IF(E{r}<0,"DÉFICITAIRE",IF(I{r}<0,"PIÈGE",IF(L{r}<0.7,"SOUS-REMPLI","")))'.format(r=r)}
    for col,f in nf.items():
        j=column_index_from_string(col)
        guard=('$B%d=""'%r) if col=="C" else ('OR($C%d="",$C%d=0)'%(r,r))
        c=dr.cell(r,j,'=IF(%s,"",%s)'%(guard,f))
        c.font=F(9.5,False,INK,MONO if col!="M" else UI); c.number_format=FMD[j-3]
        c.alignment=R if col!="M" else Cn
    for j in range(2,14): dr.cell(r,j).border=Border(bottom=sd(LINE2))
box(dr,30,2,DET0+NDET-1,13)
dr.cell(30,2,"  ④ DÉTAIL AU GRAIN CLASSE — le fond du drill").font=F(9,True,INK3); dr.cell(30,2).alignment=L
entete(dr,31,"Campus · Programme · Année · Modalité")
for i in range(NDET):
    r=DET0+i
    dr.cell(r,14,'=IFERROR(MATCH(%d,%s,0),"")'%(i+1,DD("X"))).font=F(7,c=INK3)
    dr.cell(r,2,'=IF($N{r}="","",INDEX({b},$N{r})&" · "&INDEX({c},$N{r})&" · "&INDEX({d},$N{r})&" · "&INDEX({e},$N{r}))'
            .format(r=r,b=DD("B"),c=DD("C"),d=DD("D"),e=DD("E"))).font=F(9.5)
    dr.cell(r,2).alignment=L
    df={"C":"INDEX(%s,$N%d)"%(DD("L"),r),"D":"C{r}/INDEX({j},$N{r})-1".format(r=r,j=DD("K")),
        "E":"INDEX(%s,$N%d)"%(DD("S"),r),"F":"E{r}/C{r}".format(r=r),
        "G":"INDEX(%s,$N%d)"%(DD("T"),r),"H":"INDEX(%s,$N%d)"%(DD("U"),r),
        "I":"INDEX(%s,$N%d)"%(DD("V"),r),"J":"INDEX(%s,$N%d)"%(DD("I"),r),
        "K":"INDEX(%s,$N%d)"%(DD("H"),r),"L":"J{r}/K{r}".format(r=r),
        "M":'IF(E{r}<0,"DÉFICITAIRE",IF(I{r}<0,"PIÈGE",IF(L{r}<0.7,"SOUS-REMPLI","")))'.format(r=r)}
    for col,f in df.items():
        j=column_index_from_string(col)
        c=dr.cell(r,j,'=IF($N%d="","",%s)'%(r,f))
        c.font=F(9.5,False,INK2,MONO if col!="M" else UI); c.number_format=FMD[j-3]
        c.alignment=R if col!="M" else Cn
    for j in range(2,14): dr.cell(r,j).border=Border(bottom=sd(LINE2))
for rng in ("M12","M16:M%d"%(15+NMAX),"M%d:M%d"%(DET0,DET0+NDET-1)):
    for val,bg,fg in (("PIÈGE",ORANGESOFT,ORANGE),("DÉFICITAIRE","FBE0E0",CRIT),("SOUS-REMPLI","FFF3D6","9A6B00")):
        dr.conditional_formatting.add(rng,CellIsRule(operator="equal",formula=['"%s"'%val],
            fill=fill(bg),font=Font(name=UI,size=8.5,bold=True,color=fg)))
for col in ("E","I"):
    for rng in ("%s12"%col,"%s16:%s%d"%(col,col,15+NMAX),"%s%d:%s%d"%(col,DET0,col,DET0+NDET-1)):
        dr.conditional_formatting.add(rng,CellIsRule(operator="lessThan",formula=["0"],
            font=Font(name=MONO,size=9.5,bold=True,color=CRIT)))
for rng in ("D12","D16:D%d"%(15+NMAX),"D%d:D%d"%(DET0,DET0+NDET-1)):
    for op,cc in (("greaterThan",GOOD),("lessThan",CRIT)):
        dr.conditional_formatting.add(rng,CellIsRule(operator=op,formula=["0"],font=Font(name=MONO,size=9.5,color=cc)))
for rng in ("L12","L16:L%d"%(15+NMAX),"L%d:L%d"%(DET0,DET0+NDET-1)):
    dr.conditional_formatting.add(rng,CellIsRule(operator="lessThan",formula=["0.7"],
        font=Font(name=MONO,size=9.5,bold=True,color=WARN)))
CD0=DET0+NDET+1
box(dr,CD0,2,CD0+3,13,bg=PANEL2)
dr.merge_cells(start_row=CD0,start_column=2,end_row=CD0+3,end_column=13)
dr.cell(CD0,2,"  Comment lire.  La CONTRIBUTION (CA − coûts évitables) répond à « que perd-on si on ferme ? ». "
              "La MARGE COMPLÈTE (après permanents, structure et siège) répond à « cette classe paie-t-elle sa part ? ». "
              "Aux niveaux hauts elles disent la même chose ; au grain classe elles se contredisent. "
              "« PIÈGE » = marge complète négative mais EBITDA positif : la fermer coûterait plus cher que la garder. "
              "« DÉFICITAIRE » = EBITDA négatif, elle ne couvre pas ses coûts directs — là seulement l'arbitrage se pose.").font=F(9.5,c=INK2)
dr.cell(CD0,2).alignment=TOP
for r in range(CD0,CD0+4): dr.cell(r,2).border=Border(left=sd(ORANGE,"thick"))
dr.sheet_view.zoomScale=90

# ===================== ⑤ DRILL — POURQUOI =====================
pq=wb.create_sheet("Drill — Pourquoi"); pq.sheet_view.showGridLines=False
pq.column_dimensions["A"].width=2; pq.column_dimensions["B"].width=30
pq.column_dimensions["C"].width=15; pq.column_dimensions["D"].width=24; pq.column_dimensions["E"].width=44
for c in "FGHI": pq.column_dimensions[c].width=14
for c in "KLMNO": pq.column_dimensions[c].hidden=True
canvas(pq,60,12)
for r,hh in {1:8,2:26,3:16,4:8,5:16,6:20,7:8,8:20,9:8,10:18,17:8,18:16}.items(): pq.row_dimensions[r].height=hh
for r in range(11,17): pq.row_dimensions[r].height=19
box(pq,2,2,3,9,bg=BLUE,bd=BLUE)
pq.cell(2,2,"  ⑤ Drill through  ›  Pourquoi l'EBITDA a-t-il varié ?").font=F(13,True,WHITE); pq.cell(2,2).alignment=L
pq.cell(3,2,"  requête exécutée sur la cellule EBITDA du cockpit — paramètres hérités de son contexte").font=F(9,c="D6E6F8")
pq.cell(3,2).alignment=L
box(pq,5,2,6,9)
pq.cell(5,2,"  CONTEXTE HÉRITÉ DE LA CELLULE").font=F(8.5,True,INK3); pq.cell(5,2).alignment=L
for i,(k,v) in enumerate([(":SCENARIO","=Données!B5"),(":VERSION","=Données!B6"),(":PERIODE","12"),
                          (":EXERCICE","=Données!B7"),(":ENTITY","TUNON_PAR")]):
    cc=3+i
    a=pq.cell(5,cc,k); a.font=F(8,c=INK3); a.alignment=Cn
    b=pq.cell(6,cc,v); b.font=F(10,True,BLUE,MONO); b.alignment=Cn; b.fill=fill(BLUESOFT)
    b.border=Border(top=sd(BLUE),bottom=sd(BLUE),left=sd(BLUE),right=sd(BLUE))
col=get_column_letter(VCOL+1)
dvE=DataValidation(type="list",allow_blank=False,
    formula1="='Données drill'!$%s$%d:$%s$%d"%(col,K1+1,col,K1+len(CAMP)))
pq.add_data_validation(dvE); dvE.add(pq.cell(6,7))
pq.cell(8,2,'="Cellule d\'origine :  Cockpit › Portefeuille › "&$G$6&" › EBITDA "&Données!B7').font=F(10,True,INK)
pq.cell(8,2).alignment=L
# agrégats lus par la requête
box(pq,10,7,16,9)
pq.cell(10,7,"  Agrégats lus").font=F(8.5,True,INK3); pq.cell(10,7).alignment=L
pq.cell(10,8,"2025").font=F(8.5,True,INK3); pq.cell(10,8).alignment=R
pq.cell(10,9,"2026").font=F(8.5,True,INK3); pq.cell(10,9).alignment=R
SE=lambda col: 'SUMIF(%s,$G$6,%s)'%(DD("B"),DD(col))
AGG=[("Effectifs","="+SE("J"),"="+SE("I"),NUM),("CA","="+SE("K"),"="+SE("L"),NUM),
     ("Coût variable","="+SE("M"),"="+SE("N"),NUM),("Coûts directs","="+SE("O"),"="+SE("P"),NUM),
     ("CA / élève","=H12/H11","=I12/I11",NUM),("Coût var. / élève","=H13/H11","=I13/I11",NUM)]
for i,(lab,p,n,fm) in enumerate(AGG):
    r=11+i
    pq.cell(r,7,lab).font=F(9,c=INK2); pq.cell(r,7).alignment=L
    pq.cell(r,8,p); pq.cell(r,9,n)
    for j in (8,9):
        pq.cell(r,j).number_format=fm; pq.cell(r,j).font=F(9,c=INK2,f=MONO); pq.cell(r,j).alignment=R
        pq.cell(r,j).border=Border(bottom=sd(LINE2))
# le pont
box(pq,10,2,17,5)
for j,t in zip((2,3,4,5),("Étape","Montant","Détail","Lecture")):
    c=pq.cell(10,j,t); c.font=F(8.5,True,INK3); c.alignment=L if j!=3 else R
    c.border=Border(bottom=sd(LINE)); c.fill=fill(PANEL2)
BR2=[('="EBITDA "&(Données!B7-1)',"=H12-H13-H14","","point de départ",False),
     ("+ Effet effectifs","=(I11-H11)*(H15-H16)",'=H11&" → "&I11&" élèves"',
      "volume, valorisé à la marge variable de l'exercice précédent",True),
     ("+ Effet prix / mix","=(I15-H15)*I11",'=ROUND(H15,0)&" → "&ROUND(I15,0)&" € / élève"',
      "tarif, mix initiale/alternance, mix programmes",True),
     ("− Effet coût variable unitaire","=-(I16-H16)*I11",'=ROUND(H16,0)&" → "&ROUND(I16,0)&" € / élève"',
      "vacataires et achats directs, par élève",True),
     ("− Effet coûts directs","=-(I14-H14)",'=ROUND(H14/1000,0)&" → "&ROUND(I14/1000,0)&" k€"',
      "permanents et structure : ils ne suivent pas l'activité",True),
     ('="EBITDA "&Données!B7',"=I12-I13-I14",'="marge "&ROUND((I12-I13-I14)/I12*100,1)&" %"',
      "doit égaler la cellule cliquée",False)]
for i,(lab,mt,det,lec,eff) in enumerate(BR2):
    r=11+i
    a=pq.cell(r,2,lab); a.font=F(10,not eff,INK); a.alignment=L
    c=pq.cell(r,3,mt); c.number_format='#,##0" €"'; c.alignment=R; c.font=F(10,True,INK,MONO)
    if det: pq.cell(r,4,det)
    pq.cell(r,4).font=F(9,c=INK3,f=MONO); pq.cell(r,4).alignment=R
    e=pq.cell(r,5,"  "+lec); e.font=F(9,c=INK3); e.alignment=L
    for j in range(2,6):
        pq.cell(r,j).border=Border(bottom=sd(LINE2))
        if not eff: pq.cell(r,j).fill=fill(PANEL2)
pq.cell(16,3).font=F(11,True,INK,MONO)
pq.cell(17,2,"  Contrôle du pont").font=F(9,True,INK3); pq.cell(17,2).alignment=L
ck=pq.cell(17,3,"=C11+C12+C13+C14+C15-C16"); ck.number_format='0.00" €"'; ck.alignment=R; ck.font=F(10,True,GOOD,MONO)
pq.cell(17,5,'="   écart "&ROUND(C17,2)&" € — le pont retombe exactement sur la cellule cliquée"').font=F(9,True,GOOD)
pq.cell(17,5).alignment=L
for op,cc in (("greaterThan",GOOD),("lessThan",CRIT)):
    pq.conditional_formatting.add("C12:C15",CellIsRule(operator=op,formula=["0"],
        font=Font(name=MONO,size=10,bold=True,color=cc)))
for j,t in enumerate(["Étape","Socle","Ancre","Hausse","Baisse"],11): pq.cell(10,j,t).font=F(7,c=INK3)
for i,(cat,so,an,ha,ba) in enumerate([("EBITDA N-1","0","=C11","0","0"),
        ("Effectifs","=C11","0","=C12","0"),("Prix / mix","=C11+C12","0","=C13","0"),
        ("Coût var.","=C11+C12+C13+C14","0","0","=-C14"),
        ("Coûts directs","=C16","0","0","=-C15"),("EBITDA N","0","=C16","0","0")]):
    r=11+i
    pq.cell(r,11,cat).font=F(7,c=INK3)
    for j,v in enumerate((so,an,ha,ba),12): pq.cell(r,j,v).number_format=NUM; pq.cell(r,j).font=F(7,c=INK3)
ch=BarChart(); ch.type="col"; ch.grouping="stacked"; ch.overlap=100; ch.gapWidth=45
ch.height=8.6; ch.width=19; ch.legend=None; ch.visible_cells_only=False
ch.y_axis.numFmt='#,##0 "€"'; ch.x_axis.delete=False; ch.y_axis.delete=False
ch.title="Décomposition de la variation d'EBITDA"
for c_ in (12,13,14,15): ch.add_data(Reference(pq,min_col=c_,max_col=c_,min_row=11,max_row=16),titles_from_data=False)
a,b,c3,e3=ch.series
a.graphicalProperties.noFill=True; a.graphicalProperties.line.noFill=True
for s,cc in ((b,BLUE),(c3,GOOD),(e3,CRIT)):
    s.graphicalProperties.solidFill=cc; s.graphicalProperties.line.solidFill=cc
for s,fm in ((b,'#,##0" €"'),(c3,'"+ "#,##0" €"'),(e3,'"− "#,##0" €"')):
    s.dLbls=DataLabelList(); s.dLbls.showVal=True; s.dLbls.numFmt=fm; s.dLbls.dLblPos="ctr"
    s.dLbls.showSerName=False; s.dLbls.showCatName=False; s.dLbls.showLegendKey=False
ch.set_categories(Reference(pq,min_col=11,max_col=11,min_row=11,max_row=16))
pq.add_chart(ch,"B19")
box(pq,38,2,41,9,bg=PANEL2)
pq.merge_cells(start_row=38,start_column=2,end_row=41,end_column=9)
pq.cell(38,2,"  Ce que le drill fait dire au chiffre.  Le pont sépare ce qui vient de l'activité (effectifs, prix, mix) "
             "de ce qui vient des coûts. Quand l'effet coûts directs dépasse à lui seul la croissance, le campus n'a pas "
             "un problème de marché mais un problème de structure de coûts — et la discussion change de nature. "
             "Change le campus en G6 : tout le pont se recalcule, et l'écart de contrôle reste nul.").font=F(9.5,c=INK2)
pq.cell(38,2).alignment=TOP
for r in range(38,42): pq.cell(r,2).border=Border(left=sd(ORANGE,"thick"))
pq.sheet_view.zoomScale=90

# ===================== ⑥ DRILL — PAR COMPTE =====================
pc=wb.create_sheet("Drill — Par compte"); pc.sheet_view.showGridLines=False
pc.column_dimensions["A"].width=2; pc.column_dimensions["B"].width=11; pc.column_dimensions["C"].width=32
for c in "DEFG": pc.column_dimensions[c].width=15
pc.column_dimensions["H"].width=16
canvas(pc,44,10)
for r,hh in {1:8,2:26,3:16,4:8,5:16,6:20,7:8,8:20,9:8,10:18}.items(): pc.row_dimensions[r].height=hh
box(pc,2,2,3,8,bg=BLUE,bd=BLUE)
pc.cell(2,2,"  ⑥ Drill through  ›  Et dans la compta, quels comptes ont bougé ?").font=F(13,True,WHITE)
pc.cell(2,2).alignment=L
pc.cell(3,2,"  même cellule, même contexte · trié par variation absolue décroissante").font=F(9,c="D6E6F8")
pc.cell(3,2).alignment=L
box(pc,5,2,6,8)
pc.cell(5,2,"  CONTEXTE HÉRITÉ").font=F(8.5,True,INK3); pc.cell(5,2).alignment=L
pc.cell(6,2,"=Données!B5&\" · \"&Données!B6&\" · 12 · \"&Données!B7&\" · \"&'Drill — Pourquoi'!$G$6").font=F(10,True,BLUE,MONO)
pc.cell(6,2).alignment=L
pc.cell(8,2,'="Rapprochement :  Σ produits − Σ charges  =  EBITDA du campus  =  la cellule"').font=F(10,True,INK)
pc.cell(8,2).alignment=L
PQ="'Drill — Pourquoi'!"
MIXA="SUMIFS(%s,%s,%s$G$6,%s,\"ALT\")/SUMIF(%s,%s$G$6,%s)"%(DD("I"),DD("B"),PQ,DD("E"),DD("B"),PQ,DD("I"))
ASSIETTE={"CA":(PQ+"$H$12",PQ+"$I$12"),"CV":(PQ+"$H$13",PQ+"$I$13"),"CD":(PQ+"$H$14",PQ+"$I$14")}
LIGNES=[("706","Scolarité initiale","produit","CA","=%s*0.97*(1-%s)"),
        ("7062","Scolarité alternance","produit","CA","=%s*0.97*(%s)"),
        ("708","Produits annexes","produit","CA","=%s*0.03"),
        ("621","Vacataires","charge","CV","=%s*0.85"),
        ("604","Achats de prestations","charge","CV","=%s*0.15")]
for i,(p,lib,_,g) in enumerate(POSTES):
    LIGNES.append((p,lib,"charge","CD","=%s*'Données drill'!$D$"+str(VR+5+i)))
for j,t in enumerate(["Compte","Libellé","2025","2026","Variation","Var. %","Nature"],2):
    c=pc.cell(10,j,t); c.font=F(8.5,True,INK3); c.fill=fill(PANEL2); c.alignment=L if j in (2,3,8) else R
    c.border=Border(bottom=sd(LINE),top=sd(LINE))
box(pc,11,2,11+len(LIGNES)-1,8)
for j,t in enumerate(["Compte","Libellé","2025","2026","Variation","Var. %","Nature"],2):
    c=pc.cell(10,j,t); c.font=F(8.5,True,INK3); c.fill=fill(PANEL2); c.alignment=L if j in (2,3,8) else R
    c.border=Border(bottom=sd(LINE),top=sd(LINE))
for i,(cp,lib,nat,ass,tpl) in enumerate(LIGNES):
    r=11+i
    pc.cell(r,2,cp).font=F(9.5,True,INK,MONO); pc.cell(r,2).alignment=L
    pc.cell(r,3,lib).font=F(9.5,c=INK2); pc.cell(r,3).alignment=L
    for k,(j,ex) in enumerate(((4,0),(5,1))):
        base=ASSIETTE[ass][ex]
        if "(1-%s)" in tpl or "(%s)" in tpl.replace("(1-%s)",""):
            f=tpl%(base,MIXA) if tpl.count("%s")==2 else tpl%base
        else:
            f=tpl%base
        if ass=="CD":
            f=("=%s*'Données drill'!$%s$%d"%(base,"D" if ex==0 else "E",VR+5+(i-5)))
        pc.cell(r,j,f)
    pc.cell(r,6,"=E%d-D%d"%(r,r)); pc.cell(r,7,"=IFERROR(F%d/ABS(D%d),0)"%(r,r))
    pc.cell(r,8,nat).font=F(8.5,c=INK3); pc.cell(r,8).alignment=L
    for j in (4,5): pc.cell(r,j).number_format=NUM
    pc.cell(r,6).number_format='+#,##0;−#,##0'; pc.cell(r,7).number_format='"+ "0.0%;"− "0.0%'
    for j in (4,5,6,7): pc.cell(r,j).font=F(9.5,j==6,INK,MONO); pc.cell(r,j).alignment=R
    for j in range(2,9): pc.cell(r,j).border=Border(bottom=sd(LINE2))
RT=11+len(LIGNES)
pc.cell(RT,3,"EBITDA = Σ produits − Σ charges").font=F(9.5,True)
for j,col in ((4,"D"),(5,"E")):
    pc.cell(RT,j,'=SUMIF($H$11:$H$%d,"produit",%s11:%s%d)-SUMIF($H$11:$H$%d,"charge",%s11:%s%d)'
            %(RT-1,col,col,RT-1,RT-1,col,col,RT-1)).number_format=NUM
pc.cell(RT,6,"=E%d-D%d"%(RT,RT)).number_format='+#,##0;−#,##0'
for j in (4,5,6):
    pc.cell(RT,j).font=F(10,True,INK,MONO); pc.cell(RT,j).alignment=R; pc.cell(RT,j).border=Border(top=sd(INK3))
pc.conditional_formatting.add("F11:F%d"%(RT-1),DataBarRule(start_type="min",end_type="max",color="FF"+BLUE,showValue=True))
pc.conditional_formatting.add("F11:F%d"%(RT-1),CellIsRule(operator="lessThan",formula=["0"],
    font=Font(name=MONO,size=9.5,bold=True,color=CRIT)))
CC=RT+2
box(pc,CC,2,CC+3,8,bg=PANEL2)
pc.merge_cells(start_row=CC,start_column=2,end_row=CC+3,end_column=8)
pc.cell(CC,2,"  Les deux drills se répondent.  L'onglet ⑤ donne la cause économique — « effet coûts directs » ; celui-ci "
             "dit quelle ligne comptable la porte. Les postes du coût direct n'évoluent pas au même rythme : la masse "
             "salariale permanente suit la politique salariale, les loyers l'indexation, la quote-part marque les "
             "arbitrages du siège. Les clés de ventilation sont en bas de l'onglet ② et se modifient.").font=F(9.5,c=INK2)
pc.cell(CC,2).alignment=TOP
for r in range(CC,CC+4): pc.cell(r,2).border=Border(left=sd(ORANGE,"thick"))
pc.sheet_view.zoomScale=90

wb.active=2
out="/home/user/demo5/eduservices/COCKPIT_DRILL_EDUSERVICES.xlsx"
wb.save(out)
print("SAVED",out)
print("onglets :",wb.sheetnames)
