# -*- coding: utf-8 -*-
"""Modèle de pilotage EDUSERVICES v2 — cohortes, marketing mesuré sur l'historique,
params programme×année, seuil=point mort, cadrage top-down, simulateur de décisions, sensibilité."""
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as GL, column_index_from_string as CI
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, LineChart, Reference
exec(open("/home/user/demo5/eduservices/gen_v2_data.py").read().split('if __name__')[0])

OUT="/home/user/demo5/eduservices/EDUSERVICES_Modele_Pilotage_Budget.xlsx"
# ---------- styles ----------
NAVY,BLUE2,LIGHT,YEL,TOT,RISK="1F3864","2E5496","D9E1F2","FFF2CC","E2EFDA","FCE4E4"
F="Arial"
CIN=Font(name=F,color="0000FF"); CINB=Font(name=F,color="0000FF",bold=True)
CF=Font(name=F,color="000000"); CFB=Font(name=F,color="000000",bold=True)
CL=Font(name=F,color="008000"); CHDR=Font(name=F,color="FFFFFF",bold=True)
CTIT=Font(name=F,color="FFFFFF",bold=True,size=14); CB=Font(name=F,bold=True)
CIT=Font(name=F,italic=True,color="595959",size=9); CREG=Font(name=F)
FNAVY=PatternFill("solid",fgColor=NAVY); FBLUE=PatternFill("solid",fgColor=BLUE2)
FLIGHT=PatternFill("solid",fgColor=LIGHT); FYEL=PatternFill("solid",fgColor=YEL)
FTOT=PatternFill("solid",fgColor=TOT); FRISK=PatternFill("solid",fgColor=RISK)
thin=Side(style="thin",color="BFBFBF"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
AL=Alignment(horizontal="left",vertical="center"); AC=Alignment(horizontal="center",vertical="center",wrap_text=True)
AR=Alignment(horizontal="right",vertical="center"); ALW=Alignment(horizontal="left",vertical="top",wrap_text=True)
EUR='#,##0" €";(#,##0)" €";"-"'; PCT='0.0%;(0.0%);"-"'; NB='#,##0;(#,##0);"-"'; NB1='#,##0.0;(#,##0.0);"-"'; X2='0.00'
def C(ws,ref,val=None,font=None,fill=None,fmt=None,align=None,border=False):
    c=ws[ref]
    if val is not None:c.value=val
    if font:c.font=font
    if fill:c.fill=fill
    if fmt:c.number_format=fmt
    if align:c.alignment=align
    if border:c.border=BORD
    return c
def band(ws,row,a,b,text,fill=FNAVY,font=CHDR,h=20):
    ws.merge_cells(f"{a}{row}:{b}{row}")
    for col in range(CI(a),CI(b)+1): ws.cell(row=row,column=col).fill=fill
    cc=ws[f"{a}{row}"]; cc.value=text; cc.font=font; cc.alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[row].height=h

wb=openpyxl.Workbook()

# distinct programmes list (nom -> type,domaine) preserving order
PROGN=[];
for (m,pnom,ptype,dom) in prog_list: PROGN.append((pnom,ptype,dom))
# param prog×année : liste unique (programme,niveau)
PPA=[]
seen=set()
for r in rows:
    k=(r["prog"],r["niv"])
    if k in seen: continue
    seen.add(k); PPA.append((r["prog"],r["type"],r["niv"],r["mod"],r["cap"],r["heures"],r["taux"],r["pedago"],r["cacv"],r["passage"],
                             next(d for (mm,pn,pt,d) in prog_list if pn==r["prog"])))
# campus aggregates (entité marque+ville)
campus=[]; seenc=set()
for r in rows:
    key=(r["marque"],r["ville"])
    if key in seenc: continue
    seenc.add(key); cells=[x for x in rows if (x["marque"],x["ville"])==key]
    eff=sum(x["eff"] for x in cells); nouv=sum(x["nouv"] for x in cells)
    ca=sum(x["eff"]*x["tarif"]+x["nouv"]*FRAIS for x in cells)
    ens=sum(x["classes"]*x["heures"]*x["taux"] for x in cells)
    ped=sum(x["eff"]*x["pedago"] for x in cells)
    mkt=sum(x["nouv"]*x["cacv"] for x in cells)   # variable par programme uniquement (le global est fixe, en structure)
    autres=eff*AUTRES_ETU
    contrib=ca-ens-ped-mkt-autres
    loyer=round(0.11*ca/1000)*1000; etp=round(eff/30)+1; da=round(DA_PCT*ca/1000)*1000; m2=eff*8
    campus.append(dict(marque=r["marque"],ville=r["ville"],eff=eff,nouv=nouv,ca=ca,contrib=contrib,loyer=loyer,etp=etp,da=da,m2=m2))
CG=len(campus)
MARQUES=list(BRANDS.keys())
grp_ca_n1=sum(c["ca"] for c in campus)
grp_ebitda_n1=sum(c["contrib"] for c in campus)-sum(c["loyer"] for c in campus)-sum(c["etp"]*ETPC for c in campus)-STRUCT_FIXE
print("[py] CA N-1=%.0f EBITDA N-1=%.0f (%.1f%%)"%(grp_ca_n1,grp_ebitda_n1,grp_ebitda_n1/grp_ca_n1*100))

# ============================================================ refs paramètres (tout est piloté depuis 01_Cadrage)
CAD="'01_Cadrage'!"
SCEN=f"{CAD}$D$3"                     # scénario actif (Cadrage/Optimiste/Prudent)
PMKT,PPRIX,PDCONV,PPASS,PINFL,PSAL=(f"{CAD}$H${_r}" for _r in (16,17,18,19,20,21))   # leviers % (ACTIF)
KETPC,KFRAIS,KSTRUCT,KAUTRES,KELAST,KCONV=(f"{CAD}$H${_r}" for _r in (23,24,25,26,27,28))  # constantes (ACTIF) — sécurisation & recouvrement retirés
DRIVER=f"{CAD}$D$34"                  # driver d'allocation vers CAMPUS (l'allocation 09 s'appuie dessus)
DRIVER_MARQUE=f"{CAD}$D$33"; DRIVER_PROG=f"{CAD}$D$35"  # drivers marque & programme (cascade)
_NC=sum(len(v[4]) for v in BRANDS.values())   # nb de campus (coeff marque×campus)
CO0,CON=7,7+_NC-1
CVRANGE=f"{CAD}$L${CO0}:$L${CON}"; CPRANGE=f"{CAD}$M${CO0}:$M${CON}"; CMRANGE=f"{CAD}$O${CO0}:$O${CON}"

# ============================================================ 00_Notice
ws=wb.active; ws.title="00_Notice"; ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":30,"C":66,"D":20}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:D2"); C(ws,"B2","EDUSERVICES GROUP — Pilotage budgétaire (v2)",CTIT,FNAVY,align=AL); ws.row_dimensions[2].height=30
ws.merge_cells("B3:D3"); C(ws,"B3","Cohortes · marketing mesuré sur l'historique · maille programme×année · cadrage cible · simulateur de décisions",CIT)
band(ws,5,"B","D","Les feuilles")
sh=[("01_Cadrage","POSTE DE COMMANDE CFO : cadrage top-down (N-2 · atterrissage · objectif · budget · écart) + leviers %/scénarios + coefficients + graphes"),
 ("DATA_Referentiel","RÉFÉRENTIEL Tagetik : dimensions CODE · DESCRIPTION · hiérarchies (Entité, Compte, Version, Programme, Année, Modalité, Période)"),
 ("DATA_Chargement","TABLE DE FAITS Tagetik (format long) : entité×programme×année×modalité × compte × version — structure allouée jusqu'à la classe"),
 ("REPORTING","DASHBOARD CFO (natif) : synthèse groupe, P&L, profitabilité par marque, funnel, trajectoire — sourcé sur le réalisé"),
 ("04_Referentiel","Dimensions : entités, comptes (fixe/variable)"),
 ("05_Param_Prog_Annee","DONNÉES DE RÉFÉRENCE par programme×année (capacité, heures, taux, pédago, CAC variable, passage)"),
 ("06_Historique","Réalisé N-1 par cellule (funnel, cohorte) + historique marketing N-2/N-1 → élasticité mesurée"),
 ("07_Structure","Réalisé par campus (loyers, ETP, D&A, m²)"),
 ("08_Moteur","Moteur budget par cellule (cohortes, marketing→volume, seuil=point mort)"),
 ("09_Allocation","Frais de structure alloués par driver"),
 ("10_PnL","P&L consolidé N-1 vs Budget + pont à 4 effets"),
 ("11_Simulateur","Décisions ouvrir/fermer/redistribuer (maille PROMO) → EBITDA AVANT → APRÈS"),
 ("11b_Mutualisation","Mutualisation inter-sections (maille CAMPUS) : sections économisables & économie potentielle"),
 ("11c_Cascade","Allocation EN CASCADE sur un VRAI campus (MBway Lyon) + effet d'une fermeture (bénéfique vs entraînement)"),
 ("12_Sensibilite","Impact € sur l'EBITDA par levier (sens unique)"),
 ("13_Simulation","KPIs Budget vs N-1 : effectif, CA, EBITDA, taux d'alternance"),
 ("14_Mapping_Tagetik","Passerelle Tagetik")]
r=6
for a,b in sh:
    C(ws,f"B{r}",a,CB,FLIGHT,align=AL,border=True); ws.merge_cells(f"C{r}:D{r}"); C(ws,f"C{r}",b,CREG,align=ALW,border=True); r+=1
band(ws,r+1,"B","D","Légende"); r+=2
for t,ft,fl in [("Saisie / donnée réelle",CIN,None),("Formule",CF,None),("Lien inter-feuilles",CL,None),("Hypothèse à remplir",CB,FYEL)]:
    C(ws,f"B{r}"," exemple ",ft,fl,align=AC,border=True); ws.merge_cells(f"C{r}:D{r}"); C(ws,f"C{r}",t,CREG,align=AL,border=True); r+=1
ws.merge_cells(f"B{r+1}:D{r+2}")
C(ws,f"B{r+1}","Marques/campus réels EDUSERVICES ; montants ILLUSTRATIFS calibrés sur des ordres de grandeur sourcés "
 "(scolarité 8-11 k€, NPEC ~7-10 k€, marge EBITDA ~20 %, classe ~30). À remplacer par le réel.",CIT,align=ALW)
r+=4
band(ws,r,"B","D","Mode d'emploi — par où commencer"); r+=1
steps=[("1","01_Cadrage : choisis le SCÉNARIO et règle les leviers (%) ; pose l'objectif ; lis l'écart et le reste à trouver. Tout le budget se recalcule."),
 ("2","08_Moteur : le budget se construit cellule par cellule (cohortes → marketing→volume → tarif → financement)."),
 ("3","10_PnL & 13_Simulation : lis le compte de résultat consolidé et les KPIs (Budget vs N-1)."),
 ("4","01_Cadrage (haut) : l'objectif vs le budget construit donne l'écart ; 12_Sensibilite : quel levier actionner pour combler l'écart."),
 ("5","11 / 11b / 11c : simule des décisions (ouvrir / fermer / regrouper / mutualiser) — BAC À SABLE, sans impact sur le budget officiel.")]
for n,t in steps:
    C(ws,f"B{r}",n,CB,FLIGHT,align=AC,border=True); ws.merge_cells(f"C{r}:D{r}"); C(ws,f"C{r}",t,CREG,align=ALW,border=True); r+=1
r+=1; band(ws,r,"B","D","3 façons d'utiliser le modèle"); r+=1
for a,b in [("Construire (bottom-up)","Le budget monte tout seul depuis les drivers et l'historique — aucune saisie de résultat."),
 ("Challenger","Un directeur de campus ajuste SES effectifs / curseurs → version bottom-up à confronter au cadrage."),
 ("Simuler","Tester une ouverture / fermeture / mutualisation et voir l'impact EBITDA avant de trancher.")]:
    C(ws,f"B{r}",a,CB,align=AL,border=True); ws.merge_cells(f"C{r}:D{r}"); C(ws,f"C{r}",b,CREG,align=ALW,border=True); r+=1
r+=1; ws.merge_cells(f"B{r}:D{r+1}")
C(ws,f"B{r}","💡 Astuce : survole le TITRE d'un onglet pour lire son OBJET, et une ENTÊTE de colonne pour son explication. 🔵 bleu = à saisir · 🟡 jaune = hypothèse à remplir.",CIT,align=ALW)

# ============================================================ (02_Leviers & 03_Coeff_Strateg fusionnés dans 01_Cadrage — voir plus bas)

# ============================================================ 04_Referentiel
ws=wb.create_sheet("04_Referentiel"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":24,"C":16,"D":22,"E":14,"F":12}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:F2"); C(ws,"B2","Référentiel — dimensions",CTIT,FNAVY,align=AL); ws.row_dimensions[2].height=26
band(ws,4,"B","F","Entités : Groupe → Marque → Campus")
for i,h in enumerate(["Marque","Ville","Domaine","Niveau","Devise"]): C(ws,f"{GL(2+i)}5",h,CHDR,FBLUE,align=AC,border=True)
r=6; C(ws,f"B{r}","EDUSERVICES GROUP",CINB,FTOT,align=AL,border=True); C(ws,f"C{r}","—",CREG,FTOT,align=AC,border=True)
C(ws,f"D{r}","Tous",CREG,FTOT,align=AC,border=True); C(ws,f"E{r}","Groupe",CB,FTOT,align=AC,border=True); C(ws,f"F{r}","EUR",CREG,FTOT,align=AC,border=True); r+=1
for marque,(dom,cv,cp,base,villes) in BRANDS.items():
    for ville in villes:
        C(ws,f"B{r}",marque,CIN,align=AL,border=True); C(ws,f"C{r}",ville,CIN,align=AC,border=True)
        C(ws,f"D{r}",dom,CREG,align=AC,border=True); C(ws,f"E{r}","Campus",CREG,align=AC,border=True); C(ws,f"F{r}","EUR",CREG,align=AC,border=True); r+=1
band(ws,r+1,"B","F","Comptes (nature fixe/variable)"); r+=2
for i,h in enumerate(["Compte","Libellé","Rubrique","Nature","Inducteur"]): C(ws,f"{GL(2+i)}{r}",h,CHDR,FBLUE,align=AC,border=True)
r+=1
for cpt,lib,rub,nat,ind in [("70600","Scolarité","CA","—","Effectif × tarif"),("70800","Frais de dossier","CA","—","Nouveaux"),
 ("64100","Enseignement","Coût direct","Semi-fixe/classe","Classes × heures"),("60700","Pédagogie","Coût direct","Variable","Effectif"),
 ("62300","Marketing","Coût direct","Variable","Nouveaux / budget"),("61300","Loyers","Structure","Fixe","Campus"),
 ("64000","Personnel permanent","Structure","Fixe","ETP"),("65000","Structure groupe","Structure","Fixe (alloué)","Driver"),("68000","D&A","D&A","Fixe","Campus")]:
    C(ws,f"B{r}",cpt,CIN,align=AL,border=True); C(ws,f"C{r}",lib,CREG,align=AL,border=True); C(ws,f"D{r}",rub,CREG,align=AC,border=True)
    C(ws,f"E{r}",nat,CREG,align=AC,border=True); C(ws,f"F{r}",ind,CREG,align=AL,border=True); r+=1

# ============================================================ 05_Param_Prog_Annee
ws=wb.create_sheet("05_Param_Prog_Annee"); ws.sheet_view.showGridLines=False
hh=["Clé (prog|année)","Programme","Année","Mod.","Capacité","Heures/classe","Taux horaire","Coût pédago/étu","CAC variable","Taux de passage"]
hw=[20,20,7,6,10,12,11,13,12,13]
for i,w in enumerate(hw): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells("A1:J1"); C(ws,"A1","Données de référence par programme × année (issues du réalisé — à charger)",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=24
ws.merge_cells("A2:J2"); C(ws,"A2","Bleu = saisie. Capacité, heures, taux, pédago, CAC variable et taux de passage varient par programme ET par année.",CIT)
for i,h in enumerate(hh): C(ws,f"{GL(1+i)}3",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[3].height=30
PA0=4; r=PA0
for (prog,ptype,niv,mod,cap,heures,taux,ped,cacv,passage,dom) in PPA:
    C(ws,f"A{r}",f"{prog}|{niv}",CIN,align=AL,border=True); C(ws,f"B{r}",prog,CIN,align=AL,border=True); C(ws,f"C{r}",niv,CIN,align=AC,border=True); C(ws,f"D{r}",mod,CIN,align=AC,border=True)
    C(ws,f"E{r}",cap,CIN,fmt=NB,align=AC,border=True); C(ws,f"F{r}",heures,CIN,fmt=NB,align=AC,border=True); C(ws,f"G{r}",taux,CIN,fmt=EUR,align=AC,border=True)
    C(ws,f"H{r}",ped,CIN,fmt=EUR,align=AC,border=True); C(ws,f"I{r}",cacv,CIN,fmt=EUR,align=AC,border=True)
    C(ws,f"J{r}",(passage if passage else None),CIN,fmt=PCT,align=AC,border=True); r+=1
PAN=r-1
PA=lambda col:f"'05_Param_Prog_Annee'!${col}${PA0}:${col}${PAN}"
PAKEY=PA("A")

# ============================================================ 06_Historique (+ bloc marketing)
ws=wb.create_sheet("06_Historique"); ws.sheet_view.showGridLines=False
hc=["Marque","Ville","Programme","Type","Année","Mod.","Entrée","Cand N-1","Nouv N-1","Réins N-1","Effectif N-1","Eff. année inf. N-1","Tarif N-1","Classes N-1"]
hw=[15,10,18,6,7,6,7,9,9,9,10,13,9,9]
for i,w in enumerate(hw): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells("A1:N1"); C(ws,"A1","Réalisé N-1 par cellule (funnel + cohorte)",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=22
for i,h in enumerate(hc): C(ws,f"{GL(1+i)}3",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[3].height=30
HR0=4; r=HR0
for rr in rows:
    vals=[rr["marque"],rr["ville"],rr["prog"],rr["type"],rr["niv"],rr["mod"],rr["entry"],rr["cand"],rr["nouv"],rr["rein"],rr["eff"],rr["eff_prev"],rr["tarif"],rr["classes"]]
    for i,v in enumerate(vals):
        fmt=NB if i in (7,8,9,10,11,13) else (EUR if i==12 else None)
        al=AC if i>=4 else AL
        C(ws,f"{GL(1+i)}{r}",v,CIN,fmt=fmt,align=al,border=True)
    r+=1
HRN=HR0+N-1
# bloc marketing par programme (N-2/N-1 -> élasticité)
mrow=HRN+3
C(ws,f"A{mrow}","Historique marketing par programme (issu du réalisé) → élasticité mesurée",CB,align=AL); ws.merge_cells(f"A{mrow}:F{mrow}")
mrow+=1
for i,h in enumerate(["Programme","Cand N-2","Cand N-1","Marketing N-2","Marketing N-1","Élasticité mesurée"]): C(ws,f"{GL(1+i)}{mrow}",h,CHDR,FBLUE,align=AC,border=True)
ME0=mrow+1; r=ME0
for (m,pnom,ptype,dom) in prog_list:
    h=mkt_hist[(m,pnom)]
    C(ws,f"A{r}",pnom,CIN,align=AL,border=True); C(ws,f"B{r}",h["cand_n2"],CIN,fmt=NB,align=AC,border=True); C(ws,f"C{r}",h["cand_n1"],CIN,fmt=NB,align=AC,border=True)
    C(ws,f"D{r}",h["mkt_n2"],CIN,fmt=EUR,align=AR,border=True); C(ws,f"E{r}",h["mkt_n1"],CIN,fmt=EUR,align=AR,border=True)
    C(ws,f"F{r}",f"=IFERROR((C{r}/B{r}-1)/(E{r}/D{r}-1),{KELAST})",CF,fmt=X2,align=AC,border=True); r+=1
MEN=r-1
MEKEY=f"'06_Historique'!$A${ME0}:$A${MEN}"; MEELA=f"'06_Historique'!$F${ME0}:$F${MEN}"
def hc_(col,rr): return f"'06_Historique'!{col}{rr}"

# ============================================================ 07_Structure
ws=wb.create_sheet("07_Structure"); ws.sheet_view.showGridLines=False
sc=["Marque","Ville","Effectif N-1","CA N-1","Contribution N-1","Loyer N-1","ETP perm.","Masse perm. N-1","D&A N-1","Surface m²","EBITDA campus N-1"]
sw=[15,10,11,13,14,12,10,14,11,10,15]
for i,w in enumerate(sw): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells("A1:K1"); C(ws,"A1","Réalisé N-1 par campus (structure)",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=22
for i,h in enumerate(sc): C(ws,f"{GL(1+i)}3",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[3].height=30
SR0=4; r=SR0
for cc in campus:
    C(ws,f"A{r}",cc["marque"],CIN,align=AL,border=True); C(ws,f"B{r}",cc["ville"],CIN,align=AC,border=True)
    C(ws,f"C{r}",cc["eff"],CIN,fmt=NB,align=AC,border=True); C(ws,f"D{r}",cc["ca"],CIN,fmt=EUR,align=AR,border=True)
    C(ws,f"E{r}",round(cc["contrib"]),CIN,fmt=EUR,align=AR,border=True); C(ws,f"F{r}",cc["loyer"],CIN,fmt=EUR,align=AR,border=True)
    C(ws,f"G{r}",cc["etp"],CIN,fmt=NB,align=AC,border=True); C(ws,f"H{r}",f"=G{r}*{KETPC}",CF,fmt=EUR,align=AR,border=True)
    C(ws,f"I{r}",cc["da"],CIN,fmt=EUR,align=AR,border=True); C(ws,f"J{r}",cc["m2"],CIN,fmt=NB,align=AC,border=True)
    C(ws,f"K{r}",f"=E{r}-F{r}-H{r}",CF,fmt=EUR,align=AR,border=True); r+=1
SRN=SR0+CG-1; r=SR0+CG
C(ws,f"A{r}","TOTAL",CFB,FTOT,align=AL,border=True); C(ws,f"B{r}"," ",fill=FTOT,border=True)
for col in ["C","D","E","F","G","H","I","J","K"]:
    C(ws,f"{col}{r}",f"=SUM({col}{SR0}:{col}{SRN})",CFB,FTOT,fmt=(NB if col in("C","G","J") else EUR),align=(AC if col in("C","G","J") else AR),border=True)
STOT=r
STC=lambda col:f"'07_Structure'!{col}{STOT}"

# ============================================================ 08_Moteur
ws=wb.create_sheet("08_Moteur"); ws.sheet_view.showGridLines=False
mcols=["Marque","Ville","Programme","Année","Mod","Entrée","Eff N-1","Nouv N-1","Réins N-1","Cand N-1","EffInf N-1","Tarif N-1","Cl N-1","clé",
 "Cap","Heures","Taux","Pédago","CACvar","Passage","Coût/cl","cMkt","cPrix","Élast","Effort","Cand Bud","Conv","Nouv Bud","Réins Bud","Effectif Bud",
 "Tarif Bud","(réservé)","(réservé2)","CA Bud","Cl besoin","Enseign","Pédago€","Mktg","Contrib","Rempl","Contr/étu","Pt mort"]
for i,w in enumerate([15,9,17,6,5,6]+[9]*(len(mcols)-6)): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells("A1:AP1"); C(ws,"A1","Moteur de budget par cellule (cohortes · marketing→volume mesuré · seuil=point mort)",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=22
for i,h in enumerate(mcols): C(ws,f"{GL(1+i)}2",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[2].height=28
MR0=3
for idx in range(N):
    r=MR0+idx; hr=HR0+idx
    lk=lambda c:f"={hc_(c,hr)}"
    C(ws,f"A{r}",lk('A'),CL,align=AL,border=True); C(ws,f"B{r}",lk('B'),CL,align=AC,border=True); C(ws,f"C{r}",lk('C'),CL,align=AL,border=True)
    C(ws,f"D{r}",lk('E'),CL,align=AC,border=True); C(ws,f"E{r}",lk('F'),CL,align=AC,border=True); C(ws,f"F{r}",lk('G'),CL,fmt=NB,align=AC,border=True)
    C(ws,f"G{r}",lk('K'),CL,fmt=NB,align=AC,border=True); C(ws,f"H{r}",lk('I'),CL,fmt=NB,align=AC,border=True); C(ws,f"I{r}",lk('J'),CL,fmt=NB,align=AC,border=True)
    C(ws,f"J{r}",lk('H'),CL,fmt=NB,align=AC,border=True); C(ws,f"K{r}",lk('L'),CL,fmt=NB,align=AC,border=True); C(ws,f"L{r}",lk('M'),CL,fmt=EUR,align=AC,border=True); C(ws,f"M{r}",lk('N'),CL,fmt=NB,align=AC,border=True)
    C(ws,f"N{r}",f'=C{r}&"|"&D{r}',CF,align=AC,border=True)
    C(ws,f"O{r}",f"=INDEX({PA('E')},MATCH(N{r},{PAKEY},0))",CF,fmt=NB,align=AC,border=True)
    C(ws,f"P{r}",f"=INDEX({PA('F')},MATCH(N{r},{PAKEY},0))",CF,fmt=NB,align=AC,border=True)
    C(ws,f"Q{r}",f"=INDEX({PA('G')},MATCH(N{r},{PAKEY},0))",CF,fmt=EUR,align=AC,border=True)
    C(ws,f"R{r}",f"=INDEX({PA('H')},MATCH(N{r},{PAKEY},0))",CF,fmt=EUR,align=AC,border=True)
    C(ws,f"S{r}",f"=INDEX({PA('I')},MATCH(N{r},{PAKEY},0))",CF,fmt=EUR,align=AC,border=True)
    C(ws,f"T{r}",f"=IFERROR(INDEX({PA('J')},MATCH(N{r},{PAKEY},0)),0)",CF,fmt=PCT,align=AC,border=True)
    C(ws,f"U{r}",f"=P{r}*Q{r}*(1+{PSAL})",CF,fmt=EUR,align=AC,border=True)
    C(ws,f"V{r}",f'=INDEX({CVRANGE},MATCH(A{r}&"|"&B{r},{CMRANGE},0))',CF,fmt=X2,align=AC,border=True)
    C(ws,f"W{r}",f'=INDEX({CPRANGE},MATCH(A{r}&"|"&B{r},{CMRANGE},0))',CF,fmt=X2,align=AC,border=True)
    C(ws,f"X{r}",f"=IFERROR(INDEX({MEELA},MATCH(C{r},{MEKEY},0)),{KELAST})",CF,fmt=X2,align=AC,border=True)
    C(ws,f"Y{r}",f"={PMKT}*V{r}",CF,fmt=PCT,align=AC,border=True)
    C(ws,f"Z{r}",f"=IF(F{r}=1,J{r}*(1+X{r}*Y{r}),0)",CF,fmt=NB,align=AC,border=True)
    C(ws,f"AA{r}",f"=IF(F{r}=1,IFERROR(H{r}/J{r},{KCONV})+{PDCONV},0)",CF,fmt=PCT,align=AC,border=True)  # conversion mesurée par cellule + gain
    C(ws,f"AB{r}",f"=Z{r}*AA{r}",CF,fmt=NB,align=AC,border=True)
    C(ws,f"AC{r}",f"=IF(F{r}=1,0,K{r}*(T{r}+{PPASS}))",CF,fmt=NB,align=AC,border=True)
    C(ws,f"AD{r}",f"=AB{r}+AC{r}",CFB,fmt=NB,align=AC,border=True)
    C(ws,f"AE{r}",f"=L{r}*(1+{PPRIX}*W{r})",CF,fmt=EUR,align=AC,border=True)
    C(ws,f"AF{r}","=1",CF,fmt=X2,align=AC,border=True)   # financement alternance = 1 (tarif plein) — sécurisation retirée
    C(ws,f"AG{r}","=1",CF,fmt=X2,align=AC,border=True)
    C(ws,f"AH{r}",f"=AD{r}*AE{r}*AF{r}+AB{r}*{KFRAIS}",CF,fmt=EUR,align=AR,border=True)
    C(ws,f"AI{r}",f"=IF(AD{r}<=0,0,MAX(1,ROUNDUP(AD{r}/O{r},0)))",CF,fmt=NB,align=AC,border=True)
    C(ws,f"AJ{r}",f"=AI{r}*U{r}",CF,fmt=EUR,align=AR,border=True)
    C(ws,f"AK{r}",f"=AD{r}*R{r}*(1+{PINFL})",CF,fmt=EUR,align=AR,border=True)
    C(ws,f"AL{r}",f"=H{r}*S{r}*(1+Y{r})",CF,fmt=EUR,align=AR,border=True)  # marketing variable (achat de leads) × (1+effort) ; le global est en structure fixe
    C(ws,f"AM{r}",f"=AH{r}-AJ{r}-AK{r}-AL{r}-AD{r}*{KAUTRES}*(1+{PINFL})",CFB,fmt=EUR,align=AR,border=True)
    C(ws,f"AN{r}",f"=IFERROR(AD{r}/(AI{r}*O{r}),0)",CF,fmt=PCT,align=AC,border=True)
    C(ws,f"AO{r}",f"=IFERROR(AM{r}/AD{r},0)",CF,fmt=EUR,align=AC,border=True)
    C(ws,f"AP{r}",f"=IFERROR(AI{r}*U{r}/(AE{r}*AF{r}-R{r}*(1+{PINFL})),0)",CF,fmt=NB,align=AC,border=True)
MRN=MR0+N-1; r=MR0+N
C(ws,f"A{r}","TOTAL",CFB,FTOT,align=AL,border=True)
for col in ["B","C","D","E"]: C(ws,f"{col}{r}"," ",fill=FTOT,border=True)
for col,fmt in [("F",NB),("G",NB),("H",NB),("AB",NB),("AC",NB),("AD",NB),("AH",EUR),("AI",NB),("AJ",EUR),("AK",EUR),("AL",EUR),("AM",EUR)]:
    C(ws,f"{col}{r}",f"=SUM({col}{MR0}:{col}{MRN})",CFB,FTOT,fmt=fmt,align=(AC if fmt==NB else AR),border=True)
MTOT=r; ws.freeze_panes="G3"
# --- guide de lecture visible des blocs de colonnes (pour qui n'ouvre pas les notes) ---
gr=MTOT+2
band(ws,gr,"A","H","Guide de lecture — de gauche à droite, chaque bloc construit le budget de la cellule :"); gr+=1
mgroups=[("A–F","Identité & type : marque, campus, programme, année, modalité, année d'entrée (1 = on recrute)"),
 ("G–M","Réalisé N-1 repris : effectif, nouveaux, réinscrits, candidatures, effectif année inférieure, tarif, classes"),
 ("N–T","Paramètres programme×année (feuille 05) : capacité, heures, taux, pédago, CAC variable, taux de passage"),
 ("U–Y","Coûts unitaires & coefficients : coût/classe, coeff marketing, coeff prix, élasticité, effort"),
 ("Z–AD","COHORTE → VOLUME : candidatures budget → conversion → nouveaux + réinscrits = EFFECTIF budget"),
 ("AE–AH","Tarif & CA budget : tarif budget, CA budget"),
 ("AI–AL","Coûts budget : classes nécessaires, enseignement, pédagogie, marketing"),
 ("AM–AP","RÉSULTATS : marge de CONTRIBUTION, remplissage, contribution/étudiant, point mort (seuil d'ouverture)")]
for rng,txt in mgroups:
    C(ws,f"A{gr}",rng,CB,FLIGHT,align=AC,border=True); ws.merge_cells(f"B{gr}:H{gr}"); C(ws,f"B{gr}",txt,CREG,align=ALW,border=True); gr+=1
MO=lambda col:f"'08_Moteur'!{col}"
def mrng(col): return f"'08_Moteur'!${col}${MR0}:${col}${MRN}"
def msum(col): return f"SUM({mrng(col)})"
# ============================================================ 09_Allocation
ws=wb.create_sheet("09_Allocation"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":16,"C":11,"D":13,"E":11,"F":13,"G":12,"H":11,"I":13,"J":13,"K":11,"L":13}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:L2"); C(ws,"B2","Allocation des frais de structure par driver",CTIT,FNAVY,align=AL); ws.row_dimensions[2].height=24
C(ws,"B3","Driver actif :",CB,align=AR); C(ws,"C3",f"={DRIVER}",CL,FYEL,align=AC,border=True)
C(ws,"E3","Frais de structure (fixe) à allouer :",CB,align=AR); ws.merge_cells("E3:G3")
C(ws,"H3",f"={KSTRUCT}*(1+{PINFL})",CFB,fmt=EUR,align=AR,border=True); ws.merge_cells("H3:I3")
for i,h in enumerate(["Marque","Ville","Contribution","Loyer","Masse perm.","Driver","Part","Alloué","EBITDA campus","D&A","EBIT"]):
    C(ws,f"{GL(2+i)}5",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[5].height=28
AR0=6; GC="$H$3"
for idx,cc in enumerate(campus):
    r=AR0+idx; sr=SR0+idx; vl=cc["ville"]
    crit=f'{mrng("A")},"{cc["marque"]}",{mrng("B")},"{vl}"'
    C(ws,f"B{r}",cc["marque"],CL,align=AL,border=True); C(ws,f"C{r}",vl,CL,align=AC,border=True)
    C(ws,f"D{r}",f"=SUMIFS({mrng('AM')},{crit})",CF,fmt=EUR,align=AR,border=True)
    C(ws,f"E{r}",f"='07_Structure'!F{sr}*(1+{PINFL})",CL,fmt=EUR,align=AR,border=True)
    C(ws,f"F{r}",f"='07_Structure'!G{sr}*{KETPC}*(1+{PSAL})",CL,fmt=EUR,align=AR,border=True)
    effb=f"SUMIFS({mrng('AD')},{crit})"; cab=f"SUMIFS({mrng('AH')},{crit})"; m2=f"'07_Structure'!J{sr}"
    C(ws,f"G{r}",f"=IF($C$3=\"Effectifs\",{effb},IF($C$3=\"Chiffre d'affaires\",{cab},{m2}))",CF,fmt=NB,align=AR,border=True)
    C(ws,f"H{r}",f"=IFERROR(G{r}/SUM($G${AR0}:$G${AR0+CG-1}),0)",CF,fmt=PCT,align=AC,border=True)
    C(ws,f"I{r}",f"=H{r}*{GC}",CF,fmt=EUR,align=AR,border=True)
    C(ws,f"J{r}",f"=D{r}-E{r}-F{r}-I{r}",CFB,fmt=EUR,align=AR,border=True)
    C(ws,f"K{r}",f"='07_Structure'!I{sr}",CL,fmt=EUR,align=AR,border=True)
    C(ws,f"L{r}",f"=J{r}-K{r}",CF,fmt=EUR,align=AR,border=True)
ARN=AR0+CG-1; r=AR0+CG
C(ws,f"B{r}","TOTAL",CFB,FTOT,align=AL,border=True); C(ws,f"C{r}"," ",fill=FTOT,border=True)
for col in ["D","E","F","I","J","K","L"]: C(ws,f"{col}{r}",f"=SUM({col}{AR0}:{col}{ARN})",CFB,FTOT,fmt=EUR,align=AR,border=True)
C(ws,f"G{r}",f"=SUM(G{AR0}:G{ARN})",CFB,FTOT,fmt=NB,align=AR,border=True); C(ws,f"H{r}"," ",fill=FTOT,border=True)
ATOT=r; ALc=lambda col:f"'09_Allocation'!{col}{ATOT}"

# ============================================================ 10_PnL
ws=wb.create_sheet("10_PnL"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":34,"C":15,"D":15,"E":14,"F":11}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:F2"); C(ws,"B2","Compte de résultat consolidé — Budget vs N-1",CTIT,FNAVY,align=AL); ws.row_dimensions[2].height=24
C(ws,"B3","Scénario :",CB,align=AR); C(ws,"C3","='01_Cadrage'!D3",CL,align=AC)
for i,h in enumerate(["Rubrique","Réalisé N-1","Budget N+1","Écart €","Écart %"]): C(ws,f"{GL(2+i)}5",h,CHDR,FBLUE,align=AC,border=True)
n1_eff=STC("C"); n1_ca=STC("D"); n1_ctb=STC("E"); n1_loy=STC("F"); n1_perm=STC("H"); n1_da=STC("I"); n1_str=str(STRUCT_FIXE)
def pl(r,lib,n1,bud,fmt,bold,pct=False):
    ft=CFB if bold else CREG; fl=FLIGHT if bold else None
    C(ws,f"B{r}",lib,ft,fl,align=AL,border=True); C(ws,f"C{r}",n1,(CFB if bold else CL),fl,fmt=fmt,align=AR,border=True); C(ws,f"D{r}",bud,(CFB if bold else CF),fl,fmt=fmt,align=AR,border=True)
    if pct: C(ws,f"E{r}",f"=D{r}-C{r}",ft,fl,fmt=PCT,align=AR,border=True); C(ws,f"F{r}"," ",fill=fl,border=True)
    else: C(ws,f"E{r}",f"=D{r}-C{r}",ft,fl,fmt=fmt,align=AR,border=True); C(ws,f"F{r}",f"=IFERROR(D{r}/C{r}-1,0)",ft,fl,fmt=PCT,align=AR,border=True)
r=6
pl(r,"Effectifs",f"={n1_eff}",f"={msum('AD')}",NB,False); r+=1
pl(r,"Chiffre d'affaires",f"={n1_ca}",f"={msum('AH')}",EUR,True); r+=1
pl(r,"  Coûts directs (ens.+pédago+mktg+autres)",f"=-({n1_ca}-{n1_ctb})",f"=-({msum('AJ')}+{msum('AK')}+{msum('AL')}+{KAUTRES}*(1+{PINFL})*{msum('AD')})",EUR,False); r+=1
pl(r,"Marge de contribution",f"={n1_ctb}",f"={msum('AM')}",EUR,True); r+=1
pl(r,"  Loyers",f"=-{n1_loy}",f"=-{ALc('E')}",EUR,False); r+=1
pl(r,"  Personnel permanent",f"=-{n1_perm}",f"=-{ALc('F')}",EUR,False); r+=1
pl(r,"  Frais de structure groupe",f"=-{n1_str}",f"=-{ALc('I')}",EUR,False); r+=1
pl(r,"EBITDA",f"={n1_ctb}-{n1_loy}-{n1_perm}-{n1_str}",f"={ALc('J')}",EUR,True); r+=1
pl(r,"  Marge EBITDA %",f"=IFERROR(({n1_ctb}-{n1_loy}-{n1_perm}-{n1_str})/{n1_ca},0)",f"=IFERROR({ALc('J')}/{msum('AH')},0)",PCT,False,pct=True); r+=1
pl(r,"  D&A",f"=-{n1_da}",f"=-{ALc('K')}",EUR,False); r+=1
pl(r,"EBIT",f"={n1_ctb}-{n1_loy}-{n1_perm}-{n1_str}-{n1_da}",f"={ALc('L')}",EUR,True); r+=1
band(ws,r+1,"B","F","Pont Chiffre d'affaires (N-1 → Budget)"); r+=2
for i,h in enumerate(["Effet","Montant"]): C(ws,f"{GL(2+i)}{r}",h,CHDR,FBLUE,align=AC,border=True)
r+=1
vol=f"=SUMPRODUCT(({mrng('AD')}-{mrng('G')})*{mrng('L')})"
tar=f"=SUMPRODUCT({mrng('AD')}*({mrng('AE')}-{mrng('L')}))"
fra=f"=SUMPRODUCT({mrng('AB')})*{KFRAIS}-SUMPRODUCT({mrng('H')})*{KFRAIS}"
for lib,f2,fl in [("CA Réalisé N-1",f"={n1_ca}",None),("  + Effet Volume",vol,None),("  + Effet Tarif",tar,None),
 ("  + Effet Frais",fra,None),("CA Budget N+1",f"={msum('AH')}",FTOT)]:
    C(ws,f"B{r}",lib,(CFB if fl else CREG),fl,align=AL,border=True); C(ws,f"C{r}",f2,(CFB if fl else CF),fl,fmt=EUR,align=AR,border=True); r+=1

# ============================================================ 01_Cadrage (poste de commande CFO)  [tab position 1]
ws=wb.create_sheet("01_Cadrage",1); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":40,"C":8,"D":14,"E":13,"F":14,"G":13,"H":14,"I":2,"J":16,"K":11,"L":12,"M":10,"N":11,"O":15}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:H2"); C(ws,"B2","EDUSERVICES — Cadrage & pilotage TOP-DOWN (poste de commande CFO)",CTIT,FNAVY,align=AL); ws.row_dimensions[2].height=30
C(ws,"B3","Scénario actif :",CB,align=AR); C(ws,"D3","Cadrage",CINB,FYEL,align=AC,border=True)
dvs=DataValidation(type="list",formula1='"Cadrage,Optimiste,Prudent"',allow_blank=False); ws.add_data_validation(dvs); dvs.add(ws["D3"])
C(ws,"E3","◄ bascule tout le budget",CIT); ws.merge_cells("E3:H3")
ATT_CA=round(grp_ca_n1); ATT_EB=round(grp_ebitda_n1)
N1_CA=round(grp_ca_n1/1.05); N1_EB=round(grp_ebitda_n1/1.05)          # réalisé N-1 ≈ atterrissage / croissance
N2_CA=round(grp_ca_n1/1.05/1.06); N2_EB=round(grp_ebitda_n1/1.05/1.12)
# ---- choix de la BASE DE PROJECTION ----
C(ws,"B4","Base de projection :",CB,align=AR); C(ws,"D4","Atterrissage N",CINB,FYEL,align=AC,border=True)
dvb=DataValidation(type="list",formula1='"Atterrissage N,Réalisé N-1"',allow_blank=False); ws.add_data_validation(dvb); dvb.add(ws["D4"])
C(ws,"E4","→ facteur base :",CIT,align=AR); C(ws,"F4",'=IF($D$4="Réalisé N-1",$D$7/$E$7,1)',CF,fmt=X2,align=AC,border=True); BF="$F$4"
C(ws,"G4","(le budget construit projette depuis la base retenue)",CIT,align=AL); ws.merge_cells("G4:H4")
# ---- ① CADRAGE TOP-DOWN ----
band(ws,5,"B","H","① Cadrage top-down  —  N-2 · réalisé N-1 · atterrissage N · budget construit (base retenue) · objectif · écart")
for i,h in enumerate(["Indicateur","N-2 (réel)","Réalisé N-1","Atterrissage N","Budget construit","🟡 Objectif","Écart"]): C(ws,f"{GL(2+i)}6",h,CHDR,FBLUE,align=AC,border=True)
C(ws,"B7","Chiffre d'affaires",CB,align=AL,border=True)
C(ws,"C7",N2_CA,CL,fmt=EUR,align=AR,border=True); C(ws,"D7",N1_CA,CL,fmt=EUR,align=AR,border=True); C(ws,"E7",ATT_CA,CL,fmt=EUR,align=AR,border=True)
C(ws,"F7",f"={msum('AH')}*{BF}",CFB,fmt=EUR,align=AR,border=True); C(ws,"G7",round(grp_ca_n1*1.06),CINB,FYEL,fmt=EUR,align=AR,border=True); C(ws,"H7","=F7-G7",CF,fmt=EUR,align=AR,border=True)
C(ws,"B8","EBITDA",CB,align=AL,border=True)
C(ws,"C8",N2_EB,CL,fmt=EUR,align=AR,border=True); C(ws,"D8",N1_EB,CL,fmt=EUR,align=AR,border=True); C(ws,"E8",ATT_EB,CL,fmt=EUR,align=AR,border=True)
C(ws,"F8",f"={ALc('J')}*{BF}",CFB,fmt=EUR,align=AR,border=True); C(ws,"G8",round(grp_ca_n1*1.06*0.16),CINB,FYEL,fmt=EUR,align=AR,border=True); C(ws,"H8","=F8-G8",CF,fmt=EUR,align=AR,border=True)
C(ws,"B9","Marge EBITDA %",CB,align=AL,border=True)
for col in ["C","D","E","F","G"]: C(ws,f"{col}9",f"=IFERROR({col}8/{col}7,0)",CF,fmt=PCT,align=AR,border=True)
C(ws,"H9","=F9-G9",CF,fmt=PCT,align=AR,border=True)
C(ws,"B11","RESTE À TROUVER (EBITDA vs objectif) :",CB,align=AR); ws.merge_cells("B11:E11")
C(ws,"F11","=IF(G8-F8>0,G8-F8,0)",CFB,FRISK,fmt=EUR,align=AR,border=True); C(ws,"G11",'=IF(G8-F8>0,"à combler","atteint")',CIT,align=AC,border=True)
# ---- ② LEVIERS & HYPOTHÈSES (varier en %, 3 scénarios) ----
band(ws,13,"B","H","② Leviers & hypothèses  —  varie en % · bascule par scénario (colonne ACTIF = scénario choisi)")
for i,h in enumerate(["Paramètre","Unité","Base (atterr.)","Cadrage","Optimiste","Prudent","ACTIF"]): C(ws,f"{GL(2+i)}15",h,CHDR,FBLUE,align=AC,border=True)
C(ws,"H15","ACTIF",CHDR,FNAVY,align=AC,border=True)
MATCHSC='MATCH($D$3,$E$15:$G$15,0)'
levs=[("Variation du budget marketing (→ volume)","%",0,0.10,0.20,-0.05,PCT),
 ("Hausse tarifaire (prix)","%",0,0.03,0.04,0.02,PCT),
 ("Gain de conversion admissions","pts",0,0.015,0.04,0.0,PCT),
 ("Amélioration du taux de passage","pts",0,0.01,0.03,-0.01,PCT),
 ("Inflation des charges (groupe)","%",0,0.02,0.015,0.03,PCT),
 ("Politique salariale (groupe)","%",0,0.025,0.02,0.03,PCT)]
r=16
for lib,u,ba,cad,opt,pru,fmt in levs:
    C(ws,f"B{r}",lib,CREG,align=AL,border=True); C(ws,f"C{r}",u,CREG,align=AC,border=True); C(ws,f"D{r}",ba,CIT,fmt=fmt,align=AC,border=True)
    C(ws,f"E{r}",cad,CIN,fmt=fmt,align=AC,border=True); C(ws,f"F{r}",opt,CIN,fmt=fmt,align=AC,border=True); C(ws,f"G{r}",pru,CIN,fmt=fmt,align=AC,border=True)
    C(ws,f"H{r}",f"=INDEX(E{r}:G{r},{MATCHSC})",CFB,FLIGHT,fmt=fmt,align=AC,border=True); r+=1
band(ws,22,"B","H","Constantes de référence — base = atterrissage · les 3 scénarios sont SAISISSABLES (€ : variation en % · taux : valeur en %)")
# typ : "var" = € piloté par une VARIATION en % (ACTIF = base×(1+var)) ; "val" = taux/x saisi DIRECTEMENT (ACTIF = valeur)
consts=[("Coût chargé / ETP permanent","€",ETPC,EUR,"var"),("Frais de dossier / nouvel inscrit","€",FRAIS,EUR,"var"),
 ("Frais de structure & marketing groupe (FIXE)","€",STRUCT_FIXE,EUR,"var"),
 ("Autres charges d'exploitation / étudiant","€",AUTRES_ETU,EUR,"var"),
 ("Élasticité marketing (défaut)","x",ELAST_DEF,X2,"val"),("Conversion cand.→inscrit (défaut)","%",CONV_N1,PCT,"val")]
r=23
for lib,u,val,fmt,typ in consts:
    C(ws,f"B{r}",lib,CREG,align=AL,border=True); C(ws,f"C{r}",u,CREG,align=AC,border=True); C(ws,f"D{r}",val,CIN,FYEL,fmt=fmt,align=AC,border=True)
    if typ=="var":   # € : les 3 scénarios sont des VARIATIONS en % (défaut 0 %) → ACTIF = base × (1+variation)
        for col in ["E","F","G"]: C(ws,f"{col}{r}",0,CIN,fmt=PCT,align=AC,border=True)
        C(ws,f"H{r}",f"=D{r}*(1+INDEX(E{r}:G{r},{MATCHSC}))",CFB,FLIGHT,fmt=fmt,align=AC,border=True)
    else:            # taux/x : on saisit la VALEUR par scénario (défaut = base) → ACTIF = valeur
        for col in ["E","F","G"]: C(ws,f"{col}{r}",val,CIN,fmt=fmt,align=AC,border=True)
        C(ws,f"H{r}",f"=INDEX(E{r}:G{r},{MATCHSC})",CFB,FLIGHT,fmt=fmt,align=AC,border=True)
    r+=1
band(ws,32,"B","H","Drivers d'allocation de la structure — en cascade (un driver par niveau)")
dvmc=DataValidation(type="list",formula1='"Effectifs,Chiffre d\'affaires,Surface m2"',allow_blank=False); ws.add_data_validation(dvmc)
dvpr=DataValidation(type="list",formula1='"Effectifs,Chiffre d\'affaires,Nombre de classes"',allow_blank=False); ws.add_data_validation(dvpr)
C(ws,"B33","Groupe → MARQUE (par) :",CB,align=AR); ws.merge_cells("B33:C33"); C(ws,"D33","Chiffre d'affaires",CINB,FYEL,align=AC,border=True); dvmc.add(ws["D33"])
C(ws,"B34","Marque → CAMPUS (par) :",CB,align=AR); ws.merge_cells("B34:C34"); C(ws,"D34","Effectifs",CINB,FYEL,align=AC,border=True); dvmc.add(ws["D34"])
C(ws,"B35","Campus → PROGRAMME (par) :",CB,align=AR); ws.merge_cells("B35:C35"); C(ws,"D35","Nombre de classes",CINB,FYEL,align=AC,border=True); dvpr.add(ws["D35"])
C(ws,"E33","◄ pilote l'allocation par campus (feuille 09)",CIT,align=AL); ws.merge_cells("E33:H33")
C(ws,"E35","◄ pour l'allocation fine à la classe (feuille 11c)",CIT,align=AL); ws.merge_cells("E35:H35")
# ---- ③ COEFFICIENTS STRATÉGIQUES (marque × campus) ----
band(ws,5,"J","O","③ Coefficients stratégiques (marque × campus) — intensité marketing / prix")
for i,h in enumerate(["Marque","Ville","Marketing","Prix","Posture"]): C(ws,f"{GL(10+i)}6",h,CHDR,FBLUE,align=AC,border=True)
C(ws,"O6","Clé",CHDR,FBLUE,align=AC,border=True)
r=7
for marque,(dom,cv,cp,base,villes) in BRANDS.items():
    for ville in villes:
        C(ws,f"J{r}",marque,CREG,align=AL,border=True); C(ws,f"K{r}",ville,CREG,align=AC,border=True)
        C(ws,f"L{r}",cv,CIN,fmt=X2,align=AC,border=True); C(ws,f"M{r}",cp,CIN,fmt=X2,align=AC,border=True)
        C(ws,f"N{r}",f'=IF(L{r}>=1.15,"Pousser",IF(L{r}<=0.85,"Défendre","Maintenir"))',CF,align=AC,border=True)
        C(ws,f"O{r}",f'=J{r}&"|"&K{r}',CF,align=AC,border=True); r+=1
# ---- ④ GRAPHIQUES ----
bar=BarChart(); bar.type="col"; bar.title="CA & EBITDA par version (€)"; bar.height=7.5; bar.width=15; bar.style=10
data=Reference(ws,min_col=2,max_col=6,min_row=7,max_row=8); bar.add_data(data,titles_from_data=True,from_rows=True)
bar.set_categories(Reference(ws,min_col=3,max_col=6,min_row=6,max_row=6)); ws.add_chart(bar,"J23")
lc=BarChart(); lc.type="col"; lc.title="Marge EBITDA % par version"; lc.height=7.5; lc.width=15; lc.style=12
d2=Reference(ws,min_col=3,max_col=6,min_row=9,max_row=9); lc.add_data(d2,from_rows=True)
lc.set_categories(Reference(ws,min_col=3,max_col=6,min_row=6,max_row=6)); lc.legend=None; ws.add_chart(lc,"J40")

# ============================================================ 11_Simulateur (logique CFO 360° : dilution symétrique)
ws=wb.create_sheet("11_Simulateur"); ws.sheet_view.showGridLines=False
sc=["Marque","Ville","Programme","Année","Effectif","CA","Coûts directs\névitables","MARGE DE\nCONTRIBUTION","Rempl.","🤖 Reco","Motif","Décision",
 "Effectif\naprès","Contribution\naprès","Structure allouée\n(après, dilution)","Résultat tout\ncompris (après)","Δ EBITDA"]
for i,w in enumerate([13,9,15,6,8,11,12,13,7,19,40,17,9,12,12,12,11]): ws.column_dimensions[GL(1+i)].width=w
ws.merge_cells("A1:Q1"); C(ws,"A1","Simulateur & conseil de décisions — logique CFO 360° : décider sur la CONTRIBUTION ; ouvrir/fermer redilue la structure sur TOUS",CTIT,FNAVY,align=AL); ws.row_dimensions[1].height=22
DR0=6; last=DR0+N-1
STRUCT_TOT=f"({ALc('E')}+{ALc('F')}+{ALc('I')})"; STRUCT_ETU=f"{STRUCT_TOT}/{msum('AD')}"
EFF_AP=f"SUM(M{DR0}:M{last})"   # effectif total APRÈS décisions (pilote la dilution, dans les deux sens)
C(ws,"A2","EBITDA groupe AVANT :",CB,align=AR); ws.merge_cells("A2:C2"); C(ws,"D2",f"={ALc('J')}",CFB,fmt=EUR,align=AR,border=True)
C(ws,"E2","Σ Δ décisions :",CB,align=AR); ws.merge_cells("E2:F2"); C(ws,"G2",f"=SUM(Q{DR0}:Q{last})",CFB,fmt=EUR,align=AR,border=True); ws.merge_cells("G2:H2")
C(ws,"I2","EBITDA groupe APRÈS :",CB,align=AR); ws.merge_cells("I2:J2"); C(ws,"K2","=D2+G2",CFB,FTOT,fmt=EUR,align=AR,border=True); ws.merge_cells("K2:L2")
C(ws,"A3","Structure / étudiant — AVANT :",CB,align=AR); ws.merge_cells("A3:C3"); C(ws,"D3",f"={STRUCT_ETU}",CFB,fmt=EUR,align=AR,border=True)
C(ws,"E3","APRÈS :",CB,align=AR); ws.merge_cells("E3:F3"); C(ws,"G3",f"=IFERROR({STRUCT_TOT}/{EFF_AP},0)",CFB,FRISK,fmt=EUR,align=AR,border=True); ws.merge_cells("G3:H3")
C(ws,"I3","Effectif total APRÈS :",CB,align=AR); ws.merge_cells("I3:J3"); C(ws,"K3",f"={EFF_AP}",CFB,fmt=NB,align=AC,border=True)
C(ws,"M3","⚠ Fermer retire des étudiants → structure/étudiant ↑ (tous absorbent +). Ouvrir en capte → structure/étudiant ↓ (tous s'enrichissent). La structure totale ne bouge pas ; l'EBITDA groupe ne varie que des contributions.",CIT,align=ALW); ws.merge_cells("M3:Q3"); ws.row_dimensions[3].height=30
C(ws,"A4","🧪 BAC À SABLE : décisions à la maille d'UNE promo — impact SIMULÉ (encadré ci-dessus), sans effet sur le budget officiel (09/10/13), qui se pilote via les LEVIERS (02). « Regrouper » ici respecte la capacité de la promo ; pour mutualiser des sections parallèles d'un même CAMPUS, voir la feuille 11b_Mutualisation.",Font(name=F,italic=True,bold=True,color="C00000"),align=AL); ws.merge_cells("A4:Q4"); ws.row_dimensions[4].height=16
for i,h in enumerate(sc): C(ws,f"{GL(1+i)}5",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[5].height=34
# deux jeux de décisions : ENTRÉE (recrutement décidable) vs POURSUITE (cohorte déjà inscrite → seulement l'organisation des classes)
dv_ent=DataValidation(type="list",formula1='"Maintenir,Ne pas lancer,Ouvrir +1 classe,Regrouper (-1 classe)"',allow_blank=False); ws.add_data_validation(dv_ent)
dv_pou=DataValidation(type="list",formula1='"Maintenir,Regrouper (-1 classe)"',allow_blank=False); ws.add_data_validation(dv_pou)
for idx in range(N):
    r=DR0+idx; mr=MR0+idx; entry=rows[idx]["entry"]
    am=MO('AM')+str(mr); an=MO('AN')+str(mr); u=MO('U')+str(mr); ai=MO('AI')+str(mr); cap=MO('O')+str(mr)
    metu=f"({MO('AE')+str(mr)}*{MO('AF')+str(mr)}-({MO('R')+str(mr)}+{KAUTRES})*(1+{PINFL})-{MO('S')+str(mr)})"  # contribution marginale / étudiant capté
    C(ws,f"A{r}",f"={MO('A')+str(mr)}",CL,align=AL,border=True); C(ws,f"B{r}",f"={MO('B')+str(mr)}",CL,align=AC,border=True)
    C(ws,f"C{r}",f"={MO('C')+str(mr)}",CL,align=AL,border=True); C(ws,f"D{r}",f"={MO('D')+str(mr)}",CL,align=AC,border=True)
    C(ws,f"E{r}",f"={MO('AD')+str(mr)}",CL,fmt=NB,align=AC,border=True); C(ws,f"F{r}",f"={MO('AH')+str(mr)}",CL,fmt=EUR,align=AR,border=True)
    C(ws,f"G{r}",f"=F{r}-{am}",CF,fmt=EUR,align=AR,border=True)                       # coûts directs évitables
    C(ws,f"H{r}",f"={am}",CFB,FLIGHT,fmt=EUR,align=AR,border=True)                     # CONTRIBUTION (métrique de décision)
    C(ws,f"I{r}",f"={an}",CL,fmt=PCT,align=AC,border=True)
    if entry:  # année d'entrée : recrutement décidable
        reco=f'=IF(H{r}<0,"🔴 Ne pas lancer",IF(I{r}>=0.95,"🟢 Ouvrir +1 (capter+diluer)",IF(I{r}<0.55,"🟡 Surveiller","🟢 Maintenir")))'
        motif=f'=IF(H{r}<0,"Entrée à contribution négative → ne pas lancer cette cohorte",IF(I{r}>=0.95,"Saturé → ouvrir capte des étudiants ET dilue la structure pour tous",IF(I{r}<0.55,"Sous-rempli mais contribution positive → garder","Sain")))'
        dv=dv_ent
    else:      # année de poursuite : cohorte inscrite → seule l'organisation des classes est décidable
        reco=f'=IF(H{r}<0,"🔴 Restructurer (mutualiser)",IF(AND({ai}>=2,E{r}<=({ai}-1)*{cap}*1.1),"🟡 Regroupable (−1 classe tient en capacité)",IF(I{r}<0.55,"🟠 Sous-rempli mais −1 classe déborderait","🟢 Maintenir")))'
        motif=f'="Poursuite : cohorte déjà inscrite → seul levier = regrouper. −1 classe créditée uniquement si l\'effectif tient dans les classes restantes (capacité +10% tolérée). Remplissage si −1 classe : "&IF({ai}>=2,TEXT(IFERROR(E{r}/(({ai}-1)*{cap}),0),"0%"),"n/a (1 seule classe)")'
        dv=dv_pou
    C(ws,f"J{r}",reco,CF,align=AL,border=True)
    C(ws,f"K{r}",motif,CIT,align=ALW,border=True)
    C(ws,f"L{r}","Maintenir",CINB,FYEL,align=AC,border=True); dv.add(ws[f"L{r}"])
    C(ws,f"M{r}",f'=IF(L{r}="Ne pas lancer",0,IF(L{r}="Ouvrir +1 classe",E{r}+{cap},E{r}))',CF,fmt=NB,align=AC,border=True)   # effectif après
    C(ws,f"N{r}",f'=IF(L{r}="Ne pas lancer",0,IF(L{r}="Ouvrir +1 classe",H{r}+{cap}*{metu}-{u},IF(L{r}="Regrouper (-1 classe)",IF(AND({ai}>=2,E{r}<=({ai}-1)*{cap}*1.1),H{r}+{u},H{r}),H{r})))',CF,fmt=EUR,align=AR,border=True)  # contribution après : -1 classe créditée SEULEMENT si l'effectif tient dans les classes restantes (capacité +10%)
    C(ws,f"O{r}",f"=M{r}*IFERROR({STRUCT_TOT}/{EFF_AP},0)",CF,fmt=EUR,align=AR,border=True)   # structure allouée après (dilution symétrique)
    C(ws,f"P{r}",f"=N{r}-O{r}",CF,fmt=EUR,align=AR,border=True)                        # résultat tout compris après
    C(ws,f"Q{r}",f"=N{r}-H{r}",CFB,fmt=EUR,align=AR,border=True)                       # Δ EBITDA groupe (= Δ contribution)
ws.freeze_panes="E6"

# ============================================================ 11b_Mutualisation (inter-sections, maille CAMPUS × CYCLE)
CYCLES={"BAC":("Bachelor",["B1","B2","B3"]),"MAST":("Mastère",["M1","M2"]),"BTS":("BTS",["1","2"])}
mut_groups=[]; _seen=set()
for rr in rows:
    key=(rr["marque"],rr["ville"],rr["type"])
    if key in _seen: continue
    _seen.add(key); mut_groups.append(key)
def niv_mask(cyc): return "("+"+".join(f'({mrng("D")}="{n}")' for n in CYCLES[cyc][1])+")"
ws=wb.create_sheet("11b_Mutualisation"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":16,"C":11,"D":11,"E":10,"F":12,"G":11,"H":13,"I":14,"J":13,"K":15}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:K2"); C(ws,"B2","Mutualisation inter-sections — maille CAMPUS × CYCLE (un cran au-dessus du simulateur)",CTIT,FNAVY,align=AL); ws.row_dimensions[2].height=26
ws.merge_cells("B3:K3"); C(ws,"B3","Le simulateur (11) raisonne à la maille d'UNE promo. Ici on regroupe les promos d'un même CYCLE sur un campus (ex. Bachelor B1+B2+B3, qui partagent un tronc commun) : chacune arrondit au minimum de son côté → de la capacité se gaspille ; la mutualisation en récupère une partie.",CIT); ws.row_dimensions[3].height=28
C(ws,"B4","% d'heures mutualisables (tronc commun partagé) :",CB,align=AR); ws.merge_cells("B4:F4")
C(ws,"G4",0.25,CINB,FYEL,fmt=PCT,align=AC,border=True); MUT="$G$4"
C(ws,"H4","← hypothèse : part des heures réellement partageable au sein d'un cycle (le reste reste spécifique à chaque année).",CIT,align=AL); ws.merge_cells("H4:K4")
for i,h in enumerate(["Marque","Ville","Cycle","Effectif","Sections\naujourd'hui","Capacité\nmoy.","Sections si\nmutualisé","Sections\néconomisables","Coût moyen\n/ classe","Économie\npotentielle €"]):
    C(ws,f"{GL(2+i)}5",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[5].height=30
MU0=6
for idx,(m,vl,cyc) in enumerate(mut_groups):
    r=MU0+idx; base=f'({mrng("A")}="{m}")*({mrng("B")}="{vl}")*{niv_mask(cyc)}'
    C(ws,f"B{r}",m,CL,align=AL,border=True); C(ws,f"C{r}",vl,CL,align=AC,border=True); C(ws,f"D{r}",CYCLES[cyc][0],CL,align=AC,border=True)
    C(ws,f"E{r}",f"=SUMPRODUCT({base}*{mrng('AD')})",CF,fmt=NB,align=AC,border=True)          # effectif du cycle sur le campus
    C(ws,f"F{r}",f"=SUMPRODUCT({base}*{mrng('AI')})",CF,fmt=NB,align=AC,border=True)          # sections aujourd'hui = Σ classes (chaque année arrondit au minimum)
    C(ws,f"G{r}",f"=IFERROR(SUMPRODUCT({base}*{mrng('AI')}*{mrng('O')})/F{r},0)",CF,fmt=NB1,align=AC,border=True)  # capacité moyenne pondérée
    C(ws,f"H{r}",f"=IF(E{r}<=0,0,MAX(1,ROUNDUP(E{r}/G{r},0)))",CF,fmt=NB,align=AC,border=True)   # plancher capacité : jamais dépassé
    C(ws,f"I{r}",f"=MAX(0,F{r}-H{r})",CFB,fmt=NB,align=AC,border=True)                     # sections récupérables par mutualisation
    C(ws,f"J{r}",f"=IFERROR(SUMPRODUCT({base}*{mrng('AJ')})/F{r},0)",CF,fmt=EUR,align=AR,border=True)  # coût moyen d'une classe
    C(ws,f"K{r}",f"=I{r}*J{r}*{MUT}",CFB,fmt=EUR,align=AR,border=True)                     # économie potentielle = sections récup. × coût/classe × %mutualisable
MUN=MU0+len(mut_groups)-1; r=MU0+len(mut_groups)
C(ws,f"B{r}","TOTAL",CFB,FTOT,align=AL,border=True)
for col in ["C","D"]: C(ws,f"{col}{r}"," ",fill=FTOT,border=True)
for col,fmt in [("E",NB),("F",NB),("G",None),("H",NB),("I",NB),("J",None),("K",EUR)]:
    if fmt is None: C(ws,f"{col}{r}"," ",fill=FTOT,border=True)
    else: C(ws,f"{col}{r}",f"=SUM({col}{MU0}:{col}{MUN})",CFB,FTOT,fmt=fmt,align=(AR if fmt==EUR else AC),border=True)
MUTOT=r
C(ws,f"B{r+2}","Économie potentielle totale de mutualisation :",CB,align=AR); ws.merge_cells(f"B{r+2}:F{r+2}")
C(ws,f"G{r+2}",f"=K{MUTOT}",CFB,FRISK,fmt=EUR,align=AR,border=True); ws.merge_cells(f"G{r+2}:H{r+2}")
ws.merge_cells(f"B{r+4}:K{r+6}")
C(ws,f"B{r+4}","🧪 Indicateur de POTENTIEL (bac à sable), maille campus × cycle. « Sections économisables » = classes récupérables si l'on regroupe les étudiants d'un même cycle dans des classes pleines ; le plancher = ARRONDI.SUP(effectif du cycle / capacité) → la capacité n'est JAMAIS dépassée (même garde-fou que le simulateur, mais sur le cycle). On ne mélange PAS des cycles différents (un B1 et un M2 ne partagent pas de classe). L'économie ne porte que sur la part d'heures réellement mutualisable au sein du cycle (tronc commun) — le reste reste spécifique à chaque année. N'affecte PAS le P&L officiel ; à décliner dans Tagetik à la maille de consolidation.",CIT,align=ALW)
ws.freeze_panes="E6"

# ============================================================ 11c_Cascade (allocation en cascade sur un VRAI campus : MBway Lyon)
ws=wb.create_sheet("11c_Cascade"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":24,"C":11,"D":11,"E":9,"F":12,"G":13,"H":13,"I":13,"J":13,"K":11,"L":10,"M":13,"N":13}.items(): ws.column_dimensions[c].width=w
for c in ["P","Q","R","S","T"]: ws.column_dimensions[c].width=11
FM,FV="ISCOM","Toulouse"  # campus focus : petit campus → marges fines → l'effet d'entraînement (rouge) se voit
REDF=PatternFill("solid",fgColor="FFC7CE"); REDFONT=Font(name=F,color="9C0006",bold=True)
ws.merge_cells("B2:N2"); C(ws,"B2",f"Allocation EN CASCADE sur un VRAI campus ({FM} {FV}) & effet d'une fermeture",CTIT,FNAVY,align=AL); ws.row_dimensions[2].height=26
ws.merge_cells("B3:N3"); C(ws,"B3","OBJET DE L'ONGLET : montrer, sur un vrai campus, QUAND fermer une promo est bénéfique (ou pas). Chiffres RÉELS du moteur. La structure NON-ÉVITABLE descend en cascade : groupe → marque au CA → campus au nb d'étudiants → promo au NOMBRE DE CLASSES. On DÉCIDE sur la CONTRIBUTION (évitable) ; le « résultat tout compris » (après structure) n'est qu'informatif. Change l'effectif SCÉNARIO (bleu) : contribution et classes se recalculent.",CIT); ws.row_dimensions[3].height=40
# --- section 1 : structure non-évitable du campus focus = siège cascadé (CA→effectif) + loyer & permanents réels ---
band(ws,5,"B","N",f"1) Structure non-évitable de {FM} {FV}  —  siège groupe cascadé (CA marque → effectif campus) + loyer & permanents directs (réels, feuille 09)")
LYROW=AR0+[i for i,cc in enumerate(campus) if cc["marque"]==FM and cc["ville"]==FV][0]  # ligne 09_Allocation du campus focus
caM=f'SUMIFS({mrng("AH")},{mrng("A")},"{FM}")'; effM=f'SUMIFS({mrng("AD")},{mrng("A")},"{FM}")'
effLy=f'SUMIFS({mrng("AD")},{mrng("A")},"{FM}",{mrng("B")},"{FV}")'
C(ws,"B6","Siège / structure centrale GROUPE (réel) :",CB,align=AR); ws.merge_cells("B6:F6"); C(ws,"G6",f"={KSTRUCT}*(1+{PINFL})",CL,fmt=EUR,align=AR,border=True)
C(ws,"B7",f"× Part {FM} au CA :",CB,align=AR); ws.merge_cells("B7:F7"); C(ws,"G7",f"=IFERROR({caM}/{msum('AH')},0)",CL,fmt=PCT,align=AC,border=True)
C(ws,"H7",f"→ Siège {FM} :",CB,align=AR); C(ws,"I7","=G6*G7",CF,fmt=EUR,align=AR,border=True)
C(ws,"B8",f"× Part {FV} dans {FM} (aux effectifs) :",CB,align=AR); ws.merge_cells("B8:F8"); C(ws,"G8",f"=IFERROR({effLy}/{effM},0)",CL,fmt=PCT,align=AC,border=True)
C(ws,"H8",f"→ Siège {FV} :",CB,align=AR); C(ws,"I8","=I7*G8",CF,fmt=EUR,align=AR,border=True)
C(ws,"B9",f"Loyer {FV} (réel, feuille 09) :",CB,align=AR); ws.merge_cells("B9:H9"); C(ws,"I9",f"='09_Allocation'!E{LYROW}",CL,fmt=EUR,align=AR,border=True)
C(ws,"B10",f"Permanents {FV} (réel, feuille 09) :",CB,align=AR); ws.merge_cells("B10:H10"); C(ws,"I10",f"='09_Allocation'!F{LYROW}",CL,fmt=EUR,align=AR,border=True)
C(ws,"B11",f"ENVELOPE {FV} = siège + loyer + permanents (NON-ÉVITABLE) :",CB,align=AR); ws.merge_cells("B11:H11"); C(ws,"I11","=I8+I9+I10",CFB,FTOT,fmt=EUR,align=AR,border=True)
ENV="$I$11"
# --- section 2 : répartition campus -> promos AU NB DE CLASSES + décision (économie unitaire réelle, effectif éditable) ---
band(ws,13,"B","N","2) Répartition campus → promos AU NOMBRE DE CLASSES + décision  (change l'effectif bleu ; le rouge = résultat tout compris < 0)")
hh=[f"Promo ({FM} {FV})","Eff. réel\n(moteur)","Eff.\nscénario","Classes","CA","Contribution\n(évitable)","Struct/classe","Struct allouée","Résultat t.c.\n(avant)","Décision","Cl.\naprès","Struct après","Résultat t.c.\n(après)"]
for i,h in enumerate(hh): C(ws,f"{GL(2+i)}14",h,CHDR,FBLUE,align=AC,border=True)
C(ws,"P13","Paramètres unitaires (issus du modèle, feuille 05)",CIT,align=AL); ws.merge_cells("P13:T13")
for i,h in enumerate(["Capacité","Coût/classe","Tarif","Charge/étu","CAC/étu"]): C(ws,f"{GL(16+i)}14",h,CHDR,FBLUE,align=AC,border=True)
ws.row_dimensions[14].height=30
dv_c=DataValidation(type="list",formula1='"Maintenir,Fermer"',allow_blank=False); ws.add_data_validation(dv_c)
mrow_lyon={rr["niv"]:MR0+i for i,rr in enumerate(rows) if rr["marque"]==FM and rr["ville"]==FV}  # lignes moteur exactes (lien direct, sans SUMIFS)
# promo, niveau, eff_scenario_defaut, cap, cout/classe, tarif, charge/étu (pédago+autres), cac, décision — ISCOM Toulouse (réel, feuille 05)
promos=[("Bachelor 1 (INIT)","B1",51,32,37200,8500,1484,320,"Maintenir"),
        ("Bachelor 2 (ALT)","B2",44,32,34720,8000,1484,0,"Maintenir"),
        ("Bachelor 3 (ALT)","B3",38,32,32240,8000,1484,0,"Maintenir"),
        ("Mastère 1 (ALT — piège)","M1",31,26,35520,9000,1604,600,"Maintenir"),
        ("Mastère 2 — extinction","M2",4,26,32560,9000,1604,0,"Fermer")]
P0=15; PN=P0+len(promos)-1
for i,(nm,niv,effsc,cap,coutcl,tarif,charge,cac,dec) in enumerate(promos):
    r=P0+i
    C(ws,f"P{r}",cap,CIN,fmt=NB,align=AC,border=True); C(ws,f"Q{r}",coutcl,CIN,fmt=EUR,align=AR,border=True); C(ws,f"R{r}",tarif,CIN,fmt=EUR,align=AR,border=True); C(ws,f"S{r}",charge,CIN,fmt=EUR,align=AR,border=True); C(ws,f"T{r}",cac,CIN,fmt=EUR,align=AR,border=True)
    C(ws,f"B{r}",nm,CREG,align=AL,border=True)
    C(ws,f"C{r}",f"='08_Moteur'!AD{mrow_lyon[niv]}",CL,fmt=NB,align=AC,border=True)  # effectif réel du moteur (lien direct sur la ligne)
    C(ws,f"D{r}",effsc,CINB,FYEL,fmt=NB,align=AC,border=True)                                          # effectif scénario (éditable)
    C(ws,f"E{r}",f"=IF(D{r}<=0,0,MAX(1,ROUNDUP(D{r}/P{r},0)))",CF,fmt=NB,align=AC,border=True)          # classes
    C(ws,f"F{r}",f"=D{r}*R{r}",CF,fmt=EUR,align=AR,border=True)                                         # CA
    C(ws,f"G{r}",f"=F{r}-E{r}*Q{r}-D{r}*S{r}-D{r}*T{r}",CFB,fmt=EUR,align=AR,border=True)               # contribution (évitable)
    C(ws,f"H{r}",f"=IFERROR({ENV}/SUM($E${P0}:$E${PN}),0)",CF,fmt=EUR,align=AR,border=True)             # struct/classe avant
    C(ws,f"I{r}",f"=E{r}*H{r}",CF,fmt=EUR,align=AR,border=True)                                         # struct allouée avant
    C(ws,f"J{r}",f"=G{r}-I{r}",CFB,fmt=EUR,align=AR,border=True)                                        # résultat t.c. avant
    C(ws,f"K{r}",dec,CINB,FYEL,align=AC,border=True); dv_c.add(ws[f"K{r}"])
    C(ws,f"L{r}",f'=IF(K{r}="Fermer",0,E{r})',CF,fmt=NB,align=AC,border=True)                           # classes après
    C(ws,f"M{r}",f'=IF(K{r}="Fermer",0,L{r}*IFERROR({ENV}/SUM($L${P0}:$L${PN}),0))',CF,fmt=EUR,align=AR,border=True)  # struct après (redivision)
    C(ws,f"N{r}",f'=IF(K{r}="Fermer",0,G{r}-M{r})',CFB,fmt=EUR,align=AR,border=True)                    # résultat t.c. après
r=PN+1
C(ws,f"B{r}","TOTAL",CFB,FTOT,align=AL,border=True)
for col in ["C","D","E"]: C(ws,f"{col}{r}",f"=SUM({col}{P0}:{col}{PN})",CFB,FTOT,fmt=NB,align=AC,border=True)
for col in ["F","G","I","J"]: C(ws,f"{col}{r}",f"=SUM({col}{P0}:{col}{PN})",CFB,FTOT,fmt=EUR,align=AR,border=True)
C(ws,f"H{r}"," ",fill=FTOT,border=True); C(ws,f"K{r}"," ",fill=FTOT,border=True)
C(ws,f"L{r}",f"=SUM(L{P0}:L{PN})",CFB,FTOT,fmt=NB,align=AC,border=True)
for col in ["M","N"]: C(ws,f"{col}{r}",f"=SUM({col}{P0}:{col}{PN})",CFB,FTOT,fmt=EUR,align=AR,border=True)
ws.conditional_formatting.add(f"J{P0}:J{PN}",CellIsRule(operator="lessThan",formula=["0"],fill=REDF,font=REDFONT))
ws.conditional_formatting.add(f"N{P0}:N{PN}",CellIsRule(operator="lessThan",formula=["0"],fill=REDF,font=REDFONT))
kr=r+2
C(ws,f"B{kr}","EBITDA campus AVANT :",CB,align=AR); ws.merge_cells(f"B{kr}:E{kr}"); C(ws,f"F{kr}",f"=SUM(G{P0}:G{PN})-{ENV}",CFB,fmt=EUR,align=AR,border=True)
C(ws,f"H{kr}","EBITDA campus APRÈS :",CB,align=AR); ws.merge_cells(f"H{kr}:I{kr}"); C(ws,f"J{kr}",f'=SUMIFS(G{P0}:G{PN},K{P0}:K{PN},"<>Fermer")-{ENV}',CFB,FTOT,fmt=EUR,align=AR,border=True)
C(ws,f"M{kr}","Δ EBITDA :",CB,align=AR); C(ws,f"N{kr}",f"=J{kr}-F{kr}",CFB,FRISK,fmt=EUR,align=AR,border=True)
ws.merge_cells(f"B{kr+2}:N{kr+6}")
C(ws,f"B{kr+2}","🧪 LECTURE (vrai campus ISCOM Toulouse, bac à sable). On DÉCIDE sur la CONTRIBUTION (évitable), JAMAIS sur le résultat tout compris. Trois enseignements visibles ici : ① FERMER BÉNÉFIQUE — le Mastère 2 est en EXTINCTION (4 étudiants) : sa CONTRIBUTION est négative (une classe coûte plus qu'elle ne rapporte) → le fermer fait MONTER l'EBITDA campus (Δ = perte évitée). ② EFFET D'ENTRAÎNEMENT — en fermant, la structure non-évitable se redivise sur les classes restantes (÷ nb de classes) → le résultat tout compris du Bachelor 3 (déjà mince) bascule EN ROUGE, alors que l'EBITDA s'est amélioré : c'est un effet d'ALLOCATION comptable, pas une vraie perte. ③ LE PIÈGE — le Mastère 1 est EN ROUGE en résultat tout compris (il porte beaucoup de structure) MAIS sa CONTRIBUTION est POSITIVE (~+119 k€) → il NE FAUT PAS le fermer : le fermer FERAIT PERDRE cette contribution et alourdirait tout le monde. Essaie : mets le Mastère 1 sur « Fermer » → l'EBITDA CHUTE. N'affecte pas le P&L officiel.",CIT,align=ALW)
ws.freeze_panes="C15"

# ============================================================ 12_Sensibilite
ws=wb.create_sheet("12_Sensibilite"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":40,"C":16,"D":44}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:D2"); C(ws,"B2","Sensibilité — impact € sur l'EBITDA par levier",CTIT,FNAVY,align=AL); ws.row_dimensions[2].height=26
ws.merge_cells("B3:D3"); C(ws,"B3","Sens unique : tes leviers → EBITDA. Approximations vivantes au point courant. Pour combler l'écart de cadrage, lis quel levier pèse le plus.",CIT)
for i,h in enumerate(["Levier (pas)","Impact EBITDA","Lecture"]): C(ws,f"{GL(2+i)}5",h,CHDR,FBLUE,align=AC,border=True)
scol=f"SUMPRODUCT({mrng('AD')}*{mrng('AE')}*{mrng('AF')})"
altn=f'SUMPRODUCT(({mrng("E")}="ALT")*{mrng("AD")}*{mrng("AE")})'
newc=f'SUMPRODUCT(({mrng("F")}=1)*{mrng("AM")})'
varm=f'SUMPRODUCT(({mrng("F")}=1)*{mrng("H")}*{mrng("S")}*(1+{mrng("Y")}))'
reinb=f'SUMPRODUCT(({mrng("F")}=0)*{mrng("K")})'
ctb_etu=f"IFERROR({msum('AM')}/{msum('AD')},0)"
rows_sens=[("+1 % de hausse tarifaire",f"=0.01*{scol}","Tombe quasi intégralement en EBITDA."),
 ("+1 % de budget marketing",f"=0.01*{KELAST}*{newc}-0.01*{varm}","Plus de leads → plus d'inscrits, net du coût."),
 ("+1 point de taux de passage",f"=0.01*{reinb}*{ctb_etu}","Rétention : + d'étudiants qui poursuivent.")]
r=6
for lib,f2,lec in rows_sens:
    C(ws,f"B{r}",lib,CB,align=AL,border=True); C(ws,f"C{r}",f2,CF,fmt=EUR,align=AR,border=True); C(ws,f"D{r}",lec,CIT,align=ALW,border=True); ws.row_dimensions[r].height=26; r+=1

# ============================================================ 13_Simulation
ws=wb.create_sheet("13_Simulation"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":30,"C":16,"D":16,"E":14}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:E2"); C(ws,"B2","Tableau de bord — simulation",CTIT,FNAVY,align=AL); ws.row_dimensions[2].height=24
C(ws,"B3","Scénario :",CB,align=AR); C(ws,"C3","='01_Cadrage'!D3",CL,FYEL,align=AC,border=True)
for i,h in enumerate(["Indicateur","Budget","N-1","Évolution"]): C(ws,f"{GL(2+i)}5",h,CHDR,FBLUE,align=AC,border=True)
ebit=ALc('J'); ebn1=f"{n1_ctb}-{n1_loy}-{n1_perm}-{n1_str}"
altR=f'SUMPRODUCT(({mrng("E")}="ALT")*{mrng("AD")})'; altF=f'SUMPRODUCT(({mrng("E")}="ALT")*{mrng("G")})'
kp=[("Effectif total",f"={msum('AD')}",f"={n1_eff}",NB),("Chiffre d'affaires",f"={msum('AH')}",f"={n1_ca}",EUR),
 ("Marge de contribution",f"={msum('AM')}",f"={n1_ctb}",EUR),("EBITDA",f"={ebit}",f"={ebn1}",EUR),
 ("Marge EBITDA %",f"=IFERROR({ebit}/{msum('AH')},0)",f"=IFERROR(({ebn1})/{n1_ca},0)",PCT),
 ("Nombre de classes",f"={msum('AI')}",f"={msum('M')}",NB),("Nouveaux inscrits",f"={msum('AB')}",f"={msum('H')}",NB),
 ("Taux d'alternance",f"=IFERROR({altR}/{msum('AD')},0)",f"=IFERROR({altF}/{msum('G')},0)",PCT)]
r=6
for lib,bud,n1,fmt in kp:
    fill=None
    C(ws,f"B{r}",lib,CB,fill,align=AL,border=True); C(ws,f"C{r}",bud,CF,fill,fmt=fmt,align=AC,border=True); C(ws,f"D{r}",n1,CF,fill,fmt=fmt,align=AC,border=True)
    C(ws,f"E{r}",(f"=C{r}-D{r}" if fmt==PCT else f"=IFERROR(C{r}/D{r}-1,0)"),CF,fill,fmt=PCT,align=AC,border=True); r+=1

# ============================================================ 14_Mapping_Tagetik
ws=wb.create_sheet("14_Mapping_Tagetik"); ws.sheet_view.showGridLines=False
for c,w in {"A":2,"B":28,"C":28,"D":48}.items(): ws.column_dimensions[c].width=w
ws.merge_cells("B2:D2"); C(ws,"B2","Passerelle vers CCH Tagetik",CTIT,FNAVY,align=AL); ws.row_dimensions[2].height=24
band(ws,4,"B","D","Correspondance modèle → Tagetik")
for i,h in enumerate(["Concept","Objet Tagetik","Détail"]): C(ws,f"{GL(2+i)}5",h,CHDR,FBLUE,align=AC,border=True)
r=6
for a,b,c in [("Marque / Campus","Entity","Hiérarchie Groupe→Marque→Campus"),("Programme × Année × Modalité","Dimensions analytiques","Maille fine, attributs initial/alternance"),
 ("Params (capacité, heures…)","Attributs de dimension / driver","Données de référence par programme×année (feuille 05)"),
 ("Historique N-2/N-1","Category ACTUAL (multi-années)","Plus de profondeur = élasticité marketing robuste"),
 ("Cohortes / taux de passage","Règle de calcul","Progression B1→B2→B3"),
 ("Objectifs (cadrage)","Version cible + écart","Top-down vs bottom-up"),("Décisions ouvrir/fermer","Scénarios / versions","Simulateur → snapshot"),
 ("Allocation structure","Cost allocation (driver)","Effectif/CA/m²")]:
    C(ws,f"B{r}",a,CB,align=ALW,border=True); C(ws,f"C{r}",b,CREG,align=ALW,border=True); C(ws,f"D{r}",c,CREG,align=ALW,border=True); ws.row_dimensions[r].height=28; r+=1

# ============================================================ Codes de dimensions (Tagetik : CODE + DESCRIPTION)
MCODE={"MBway":"MBWAY","ISCOM":"ISCOM","Ipac Bachelor Factory":"IPAC","Pigier":"PIGIER","Tunon":"TUNON"}
VCODE={"Paris":"PAR","Lyon":"LYO","Nantes":"NAN","Bordeaux":"BOR","Lille":"LIL","Toulouse":"TLS","Rennes":"REN","Montpellier":"MTP"}
PCODE={"Bachelor Management":"BAC_MGT","Mastère Management":"MAS_MGT","Bachelor Communication":"BAC_COM","Mastère Communication":"MAS_COM","Bachelor Commerce":"BAC_CCE","BTS Gestion":"BTS_GES","Bachelor RH":"BAC_RH","Bachelor Tourisme":"BAC_TOU"}
CYCCODE={"BAC":"BACH","MAST":"MAST","BTS":"BTS"}; CYCLIB={"BAC":"Bachelor","MAST":"Mastère","BTS":"BTS"}
def ancode(niv): return {"1":"BTS1","2":"BTS2"}.get(niv,niv)
def campuscode(m,v): return f"{MCODE[m]}_{VCODE[v]}"
# comptes : code, description, type, unité, scalable-par-version
ACC=[("706100","Droits de scolarité","COMPTABLE","EUR",1),("708500","Frais de dossier / inscription","COMPTABLE","EUR",1),
 ("604100","Vacations & sous-traitance pédagogique","COMPTABLE","EUR",1),("606300","Fournitures & ressources pédagogiques","COMPTABLE","EUR",1),
 ("623000","Publicité & marketing (acquisition)","COMPTABLE","EUR",1),("613200","Locations immobilières (loyers)","COMPTABLE","EUR",1),
 ("641100","Rémunérations du personnel permanent","COMPTABLE","EUR",1),("622600","Honoraires & frais de siège (structure groupe)","COMPTABLE","EUR",1),
 ("606800","Autres achats & charges externes","COMPTABLE","EUR",1),("681100","Dotations aux amortissements","COMPTABLE","EUR",1),
 ("STA_EFF","Effectif inscrit","TECHNIQUE","NB",1),("STA_CAND","Candidatures reçues","TECHNIQUE","NB",1),("STA_ADMIS","Candidats admis","TECHNIQUE","NB",1),
 ("STA_NOUV","Nouveaux inscrits","TECHNIQUE","NB",1),("STA_REINS","Réinscrits","TECHNIQUE","NB",1),("STA_CLASS","Nombre de classes","TECHNIQUE","NB",1),
 ("STA_HEURE","Heures d'enseignement","TECHNIQUE","NB",1),("STA_CAPA","Capacité par classe","TECHNIQUE","NB",0),
 ("DRV_CONV","Taux de conversion candidat→inscrit","TECHNIQUE","PCT",0),("DRV_PASS","Taux de passage (réinscription)","TECHNIQUE","PCT",0),
 ("PRX_TARIF","Tarif moyen scolarité","TECHNIQUE","EUR",0),("CAC_LEAD","Coût d'acquisition par lead","TECHNIQUE","EUR",0)]
ACCLIB={a[0]:a for a in ACC}
VERS=[("2023ACT_VDEF","Réalisé 2023 (N-2) — version définitive","Actual",2023,0.90),
 ("2024ACT_VDEF","Réalisé 2024 (N-1) — version définitive","Actual",2024,0.95),
 ("2025FCST_V4","Atterrissage 2025 — Forecast V4","Forecast",2025,1.00)]
# ---- structure de coûts agrégée par campus (pour l'allouer jusqu'à la classe) ----
_totca=sum(c["ca"] for c in campus); _mca={}; _meff={}
for c in campus: _mca[c["marque"]]=_mca.get(c["marque"],0)+c["ca"]; _meff[c["marque"]]=_meff.get(c["marque"],0)+c["eff"]
_clcamp={}
for rr in rows: _clcamp[(rr["marque"],rr["ville"])]=_clcamp.get((rr["marque"],rr["ville"]),0)+rr["classes"]
_struct={}
for c in campus:
    k=(c["marque"],c["ville"]); siege=STRUCT_FIXE*(_mca[c["marque"]]/_totca)*(c["eff"]/_meff[c["marque"]])
    _struct[k]=dict(loyer=c["loyer"],perm=c["etp"]*ETPC,siege=siege,da=c["da"],cls=_clcamp[k])

# ============================================================ DATA_Referentiel (dimensions Tagetik : CODE · DESCRIPTION · hiérarchies)
wsr=wb.create_sheet("DATA_Referentiel",2); wsr.sheet_view.showGridLines=False
for c,w in {"A":2,"B":18,"C":42,"D":16,"E":14,"F":12}.items(): wsr.column_dimensions[c].width=w
wsr.merge_cells("B2:F2"); C(wsr,"B2","RÉFÉRENTIEL DES DIMENSIONS (à charger dans Tagetik) — CODE · DESCRIPTION · HIÉRARCHIE",CTIT,FNAVY,align=AL); wsr.row_dimensions[2].height=26
rf=4
def reftable(title,heads,data):
    global rf
    band(wsr,rf,"B","F",title); rf+=1
    for i,h in enumerate(heads): C(wsr,f"{GL(2+i)}{rf}",h,CHDR,FBLUE,align=AC,border=True)
    rf+=1
    for row in data:
        for i,val in enumerate(row): C(wsr,f"{GL(2+i)}{rf}",val,CREG,align=(AL if i<2 else AC),border=True)
        rf+=1
    rf+=1
# 1) Entité (hiérarchie Groupe→Marque→Campus)
ent=[["EDU","EDUSERVICES (Groupe)","","Groupe"]]
for m in BRANDS: ent.append([MCODE[m],m,"EDU","Marque"])
for c in campus: ent.append([campuscode(c["marque"],c["ville"]),f"{c['marque']} {c['ville']}",MCODE[c["marque"]],"Campus"])
reftable("1) Dimension ENTITÉ — hiérarchie Groupe → Marque → Campus",["Code","Description","Parent","Niveau"],ent)
# 2) Compte (hiérarchie + type)
acc=[["PL","Compte de résultat","","Noeud",""],["CA","Chiffre d'affaires","PL","Noeud","Produit"],
 ["COUTS_DIR","Coûts directs évitables","PL","Noeud","Charge"],["STRUCTURE","Structure (non-évitable)","PL","Noeud","Charge"],["AMORT","Amortissements","PL","Noeud","Charge"],
 ["STATS","Statistiques (effectifs/flux)","","Noeud","Technique"],["INDUCT","Inducteurs & prix unitaires","","Noeud","Technique"]]
par={"706100":"CA","708500":"CA","604100":"COUTS_DIR","606300":"COUTS_DIR","623000":"COUTS_DIR","606800":"COUTS_DIR",
 "613200":"STRUCTURE","641100":"STRUCTURE","622600":"STRUCTURE","681100":"AMORT"}
for code,desc,typ,unite,sc in ACC:
    p=par.get(code, ("STATS" if code.startswith("STA_") else "INDUCT")); sens=("Produit" if p=="CA" else ("Charge" if typ=="COMPTABLE" else "Technique"))
    acc.append([code,desc,p,typ,sens])
reftable("2) Dimension COMPTE — hiérarchie P&L + statistiques (comptes comptables réels · codes techniques)",["Code","Description","Parent","Type","Sens"],acc)
# 3) Version / scénario
ver=[[v[0],v[1],v[2],v[3]] for v in VERS]+[["2026BUD_V1","Budget 2026 — version 1 (construit par le modèle)","Budget",2026]]
reftable("3) Dimension VERSION / SCÉNARIO — code Tagetik (année + type + version)",["Code","Description","Type","Année"],ver)
# 4) Programme
prog=[]
for m in BRANDS:
    for pnom,ptype,_ in PROGS[m]:
        if not any(x[0]==PCODE[pnom] for x in prog): prog.append([PCODE[pnom],pnom,CYCLIB[ptype],BRANDS[m][0]])
reftable("4) Dimension PROGRAMME",["Code","Description","Cycle","Domaine"],prog)
# 5) Année d'études & 6) Modalité & 7) Période
reftable("5) Dimension ANNÉE D'ÉTUDES",["Code","Description","Cycle"],[["B1","Bachelor 1","Bachelor"],["B2","Bachelor 2","Bachelor"],["B3","Bachelor 3","Bachelor"],["M1","Mastère 1","Mastère"],["M2","Mastère 2","Mastère"],["BTS1","BTS 1re année","BTS"],["BTS2","BTS 2e année","BTS"]])
reftable("6) Dimension MODALITÉ",["Code","Description"],[["INIT","Formation initiale"],["ALT","Alternance"]])
reftable("7) Dimension PÉRIODE",["Code","Description","Type"],[["2023","Exercice 2023 (clos 31/08)","Annuel"],["2024","Exercice 2024","Annuel"],["2025","Exercice 2025","Annuel"],["2026","Exercice 2026","Annuel"]])

# ============================================================ DATA_Chargement (TABLE DE FAITS en long, format Tagetik)
wsd=wb.create_sheet("DATA_Chargement",3); wsd.sheet_view.showGridLines=False
for c,w in {"A":2,"B":13,"C":11,"D":8,"E":9,"F":11,"G":34,"H":13,"I":13,"J":9,"K":8,"L":14}.items(): wsd.column_dimensions[c].width=w
wsd.merge_cells("B1:L1"); C(wsd,"B1","TABLE DE FAITS À CHARGER DANS TAGETIK — format long (une ligne = un croisement dimension × compte × version)",CTIT,FNAVY,align=AL); wsd.row_dimensions[1].height=24
wsd.merge_cells("B2:L2"); C(wsd,"B2","Codes issus de DATA_Referentiel. Comptable = vrai compte PCG ; technique = code STA_/DRV_/PRX_/CAC_. Structure (loyers, permanents, siège, D&A) DÉJÀ ALLOUÉE jusqu'à la classe (par nb de classes). 3 versions : Réalisé N-2 · Réalisé N-1 · Atterrissage (Forecast V4).",CIT,align=ALW); wsd.row_dimensions[2].height=30
factH=["Entité","Programme","Année","Modalité","Compte","Libellé compte","Type","Version","Période","Unité","Montant"]
for i,h in enumerate(factH): C(wsd,f"{GL(2+i)}4",h,CHDR,FBLUE,align=AC,border=True)
UF={"EUR":EUR,"NB":NB,"PCT":PCT}
fr=5
for rr in rows:
    k=(rr["marque"],rr["ville"]); st=_struct[k]; share=(rr["classes"]/st["cls"] if st["cls"] else 0)
    conv=(round(rr["nouv"]/rr["cand"],4) if rr["cand"] else 0)
    base={"706100":rr["eff"]*rr["tarif"],"708500":rr["nouv"]*FRAIS,"604100":rr["classes"]*rr["heures"]*rr["taux"],
     "606300":rr["eff"]*rr["pedago"],"623000":rr["nouv"]*rr["cacv"],"613200":round(st["loyer"]*share),
     "641100":round(st["perm"]*share),"622600":round(st["siege"]*share),"606800":rr["eff"]*AUTRES_ETU,"681100":round(st["da"]*share),
     "STA_EFF":rr["eff"],"STA_CAND":rr["cand"],"STA_ADMIS":rr["admis"],"STA_NOUV":rr["nouv"],"STA_REINS":rr["rein"],
     "STA_CLASS":rr["classes"],"STA_HEURE":rr["classes"]*rr["heures"],"STA_CAPA":rr["cap"],
     "DRV_CONV":conv,"DRV_PASS":rr["passage"],"PRX_TARIF":rr["tarif"],"CAC_LEAD":rr["cacv"]}
    ent=campuscode(rr["marque"],rr["ville"]); pcode=PCODE[rr["prog"]]; an=ancode(rr["niv"])
    for code,desc,typ,unite,sc in ACC:
        v0=base[code]
        if not v0: continue
        for vcode,vlib,vtyp,per,fac in VERS:
            val=round(v0*fac) if (sc and unite in("EUR","NB")) else (round(v0*fac,3) if sc else v0)
            row=[ent,pcode,an,rr["mod"],code,desc,typ,vcode,per,unite,val]
            for i,x in enumerate(row): C(wsd,f"{GL(2+i)}{fr}",x,(CREG if i<7 else CF),fmt=(UF.get(unite) if i==10 else None),align=(AL if i in(0,5) else AC))
            fr+=1
wsd.freeze_panes="B5"

# ============================================================ NOTES explicatives sur chaque entête (info-bulles)

# ============================================================ REPORTING (dashboard CFO natif, sourcé sur le réalisé 07_Structure)
wsp=wb.create_sheet("REPORTING",4); wsp.sheet_view.showGridLines=False
for c in range(1,17): wsp.column_dimensions[GL(c)].width=(2 if c==1 else 9.5)
for c in range(19,26): wsp.column_dimensions[GL(c)].width=12
S7=f"'07_Structure'!"; A=f"{S7}$A$4:$A$17"
GCA=f"SUM({S7}$D$4:$D$17)"; GEFF=f"SUM({S7}$C$4:$C$17)"; GCON=f"SUM({S7}$E$4:$E$17)"
GEBP=f"SUM({S7}$K$4:$K$17)"; GDA=f"SUM({S7}$I$4:$I$17)"; GLOY=f"SUM({S7}$F$4:$F$17)"; GPER=f"SUM({S7}$H$4:$H$17)"
GEB=f"({GEBP}-{STRUCT_FIXE})"
C(wsp,"S2","Données (sourcées 07_Structure — réalisé)",CIT,align=AL); wsp.merge_cells("S2:Y2")
for i,h in enumerate(["Marque","CA","EBITDA","Marge%","CA/étu"]): C(wsp,f"{GL(19+i)}3",h,CHDR,FBLUE,align=AC,border=True)
MK=list(BRANDS.keys())
for i,m in enumerate(MK):
    r=4+i
    C(wsp,f"S{r}",m,CREG,align=AL,border=True)
    C(wsp,f"T{r}",f'=SUMIF({A},"{m}",{S7}$D$4:$D$17)',CF,fmt=EUR,align=AR,border=True)
    C(wsp,f"U{r}",f'=SUMIF({A},"{m}",{S7}$K$4:$K$17)-{STRUCT_FIXE}*T{r}/({GCA})',CF,fmt=EUR,align=AR,border=True)
    C(wsp,f"V{r}",f"=IFERROR(U{r}/T{r},0)",CF,fmt=PCT,align=AC,border=True)
    C(wsp,f"W{r}",f'=IFERROR(T{r}/SUMIF({A},"{m}",{S7}$C$4:$C$17),0)',CF,fmt=EUR,align=AR,border=True)
C(wsp,"S11","CA groupe",CREG,border=True); C(wsp,"T11",f"={GCA}",CF,fmt=EUR,align=AR,border=True)
C(wsp,"S12","EBITDA groupe",CREG,border=True); C(wsp,"T12",f"={GEB}",CF,fmt=EUR,align=AR,border=True)
C(wsp,"S13","EBIT groupe",CREG,border=True); C(wsp,"T13",f"={GEB}-{GDA}",CF,fmt=EUR,align=AR,border=True)
for i,(lab,f) in enumerate([("2023 (N-2)",0.90),("2024 (N-1)",0.95),("2025 (Atterr.)",1.00)]):
    r=15+i; C(wsp,f"S{r}",lab,CREG,border=True); C(wsp,f"T{r}",f"={GCA}*{f}",CF,fmt=EUR,align=AR,border=True); C(wsp,f"U{r}",f"={GEB}*{f}",CF,fmt=EUR,align=AR,border=True)
for i,(lab,v) in enumerate([("Candidatures",2807),("Admis",1738),("Inscrits",1044)]):
    r=19+i; C(wsp,f"S{r}",lab,CREG,border=True); C(wsp,f"T{r}",v,CF,fmt=NB,align=AR,border=True)
for i,(lab,f) in enumerate([("Coûts directs",f"={GCA}-{GCON}"),("Personnel permanent",f"={GPER}"),("Loyers",f"={GLOY}"),("Siège groupe",f"={STRUCT_FIXE}"),("D&A",f"={GDA}"),("EBITDA",f"={GEB}")]):
    r=23+i; C(wsp,f"S{r}",lab,CREG,border=True); C(wsp,f"T{r}",f,CF,fmt=EUR,align=AR,border=True)
wsp.merge_cells("B2:P2"); C(wsp,"B2","EDUSERVICES — Reporting CFO (réalisé, après chargement de l'historique)",CTIT,FNAVY,align=AL); wsp.row_dimensions[2].height=26
def tile(c0,lab,val,fmt,sub):
    a=GL(c0); b=GL(c0+1)
    wsp.merge_cells(f"{a}4:{b}4"); C(wsp,f"{a}4",lab,Font(name=F,size=9,color="5B6B7F"),FLIGHT,align=AC,border=True)
    wsp.merge_cells(f"{a}5:{b}5"); C(wsp,f"{a}5",val,Font(name=F,size=15,bold=True,color=NAVY),FLIGHT,fmt=fmt,align=AC,border=True)
    wsp.merge_cells(f"{a}6:{b}6"); C(wsp,f"{a}6",sub,Font(name=F,size=8,italic=True,color="5B6B7F"),FLIGHT,align=AC,border=True)
for r in (4,5,6): wsp.row_dimensions[r].height=(15 if r!=5 else 22)
tile(2,"Chiffre d'affaires",f"={GCA}",EUR,"réalisé")
tile(4,"Marge contribution",f"={GCON}",EUR,"contribution")
tile(6,"EBITDA",f"={GEB}",EUR,"14,6% (=conso)")
tile(8,"EBIT",f"={GEB}-{GDA}",EUR,"après D&A")
tile(10,"CA / étudiant",f"=IFERROR({GCA}/{GEFF},0)",EUR,"scolarité+frais")
tile(12,"Effectif",f"={GEFF}",NB,"82% alternance")
wsp.conditional_formatting.add("V4:V8",CellIsRule(operator="lessThan",formula=["0.1"],fill=PatternFill("solid",fgColor="FCE4E4"),font=Font(name=F,color="9C0006",bold=True)))
wsp.conditional_formatting.add("V4:V8",CellIsRule(operator="greaterThanOrEqual",formula=["0.16"],fill=PatternFill("solid",fgColor="E2EFDA"),font=Font(name=F,color="15803D",bold=True)))
def mkbar(title,dcol,dr0,drn,anchor,w=11,h=7.2,typ="col"):
    ch=BarChart(); ch.type=typ; ch.title=title; ch.width=w; ch.height=h; ch.legend=None; ch.style=10
    ch.add_data(Reference(wsp,min_col=dcol,max_col=dcol,min_row=dr0,max_row=drn))
    ch.set_categories(Reference(wsp,min_col=19,max_col=19,min_row=dr0,max_row=drn)); wsp.add_chart(ch,anchor)
mkbar("Marge EBITDA % par marque",22,4,8,"B8")
mkbar("CA par marque (€)",20,4,8,"I8")
mkbar("Chiffre d'affaires 3 ans (€)",20,15,17,"B23")
mkbar("Funnel d'admissions (nb)",20,19,21,"I23")
mkbar("Structure de coûts → EBITDA (€)",20,23,28,"B38",typ="bar")
wsp.merge_cells("I38:P41"); C(wsp,"I38","🔎 Lecture CFO : MBway porte la marge (~18 %), Pigier décroche (~5 %). Les petits campus sous-absorbent la structure fixe → marge fully-allocated faible. Le funnel (conv. ~37 %) et le CAC pilotent le volume — indicateur avancé du budget.",CIT,align=ALW)

# ============================================================ NOTES explicatives sur chaque entête (info-bulles)
from openpyxl.comments import Comment
def norm(s): return str(s).replace("\n"," ").replace("  "," ").strip()
GLOSS={
 # communs
 "Marque":"Marque / école du groupe (dimension Entity dans Tagetik).",
 "Marque / École":"Marque / école du groupe.",
 "Ville":"Campus (ville). Une entité = Marque + Ville.",
 "Programme":"Programme de formation (ex. Bachelor Management).",
 "Année":"Année d'études : B1/B2/B3 (Bachelor), M1/M2 (Mastère), 1/2 (BTS).",
 "Niveau":"Année d'études (B1, B2, B3, M1, M2…).",
 "Mod":"Modalité : INIT = initial (payé par la famille) · ALT = alternance (financé OPCO/NPEC).",
 "Mod.":"Modalité : INIT = initial · ALT = alternance (financée OPCO/NPEC).",
 "Modalité":"INIT = initial (famille) · ALT = alternance (OPCO/NPEC).",
 "Type":"Type de diplôme : BAC (Bachelor), MAST (Mastère), BTS.",
 "Domaine":"Domaine d'activité de la marque.",
 "Devise":"Devise de reporting.",
 # 05 Param_Prog_Annee
 "Clé (prog|année)":"Clé technique 'programme|année' pour relier les paramètres au moteur.",
 "Capacité":"Capacité cible d'une classe (nb d'étudiants).","Capacité cible":"Capacité cible par classe.",
 "Seuil ouverture":"Seuil historique d'ouverture (indicatif ; le point mort réel est calculé dans le moteur).",
 "Heures / classe":"Heures d'enseignement délivrées par classe et par an.",
 "Taux horaire":"Coût horaire chargé de l'enseignement (vacation), en €.",
 "Coût pédago / étu":"Coût pédagogique par étudiant/an, HORS salaires des intervenants (ceux-ci sont dans Enseignement = heures × taux). Couvre supports & ressources de cours, licences logicielles pédagogiques (ex. créa pour Communication), plateforme LMS, jury/examens, certifications, consommables/projets. Varie par domaine : Communication 550 € > Tourisme 450 > Management 400 > Commerce 380 ; BTS 350 ; Mastère +120 €.",
 "CAC variable":"Coût d'acquisition variable (achat de leads) par nouvel inscrit. N'existe qu'en année d'entrée.",
 "Taux de passage":"Taux de progression de l'année inférieure vers celle-ci (réinscription). Vide en année d'entrée.",
 # 06 Historique
 "Entrée":"1 = année d'entrée (on recrute) · 0 = année de poursuite (cohorte progressée).",
 "Candidatures N-1":"Candidatures reçues l'an dernier (haut de l'entonnoir admissions).",
 "Cand N-1":"Candidatures l'an dernier.","Cand Bud":"Candidatures budget = candidatures N-1 × (1 + élasticité × effort marketing).",
 "Nouv N-1":"Nouveaux inscrits l'an dernier (entrants).","Nouv Bud":"Nouveaux inscrits budget = candidatures budget × taux de conversion.",
 "Réins N-1":"Réinscrits l'an dernier (étudiants qui poursuivent).","Réins Bud":"Réinscrits budget = effectif de l'année inférieure (N-1) × (taux de passage + amélioration).",
 "Effectif N-1":"Effectif réel l'an dernier.","Eff N-1":"Effectif réel l'an dernier.",
 "Eff. année inf. N-1":"Effectif de l'année juste inférieure l'an dernier — base de la cohorte pour calculer la progression.",
 "EffInf N-1":"Effectif de l'année inférieure l'an dernier (base de progression de la cohorte).",
 "Tarif N-1":"Tarif moyen l'an dernier (€/étudiant/an).","Tarif Bud":"Tarif budget = tarif N-1 × (1 + hausse prix × coefficient marque).",
 "Classes N-1":"Nombre de classes l'an dernier.","Cl N-1":"Nombre de classes l'an dernier.",
 "Élasticité mesurée":"Élasticité marketing = %Δ candidatures ÷ %Δ marketing, mesurée sur N-2/N-1.",
 # 07 Structure
 "Effectif N-1 ":"Effectif réel l'an dernier.","CA N-1":"Chiffre d'affaires réel l'an dernier.",
 "Contribution N-1":"Marge de contribution réelle l'an dernier (CA − coûts directs).",
 "Loyer N-1":"Loyer du campus l'an dernier (coût de structure, fixe).",
 "ETP perm.":"Effectifs permanents (équivalent temps plein) du campus.",
 "Masse perm. N-1":"Masse salariale permanente = ETP × coût chargé par ETP.",
 "D&A N-1":"Dotations aux amortissements l'an dernier.","Surface m²":"Surface du campus (utilisée comme driver d'allocation possible).",
 "EBITDA campus N-1":"EBITDA du campus = contribution − loyer − masse salariale permanente.",
 # 08 Moteur
 "clé":"Clé programme|année pour retrouver les paramètres (feuille 05).",
 "Cap":"Capacité cible par classe (lue dans 05).","Heures":"Heures d'enseignement par classe/an (05).",
 "Taux":"Taux horaire chargé de l'enseignement (05).","Pédago":"Coût pédagogique par étudiant (05).",
 "CACvar":"Coût d'acquisition variable par nouvel inscrit (05).",
 "Passage":"Taux de passage de l'année inférieure (05). 0 en année d'entrée.",
 "Coût/cl":"Coût d'une classe = heures × taux horaire × (1 + politique salariale).",
 "cMkt":"Coefficient stratégique MARKETING de la marque (feuille 03) — module l'effort par marque.",
 "cPrix":"Coefficient stratégique PRIX de la marque (feuille 03) — module la hausse tarifaire par marque.",
 "Élast":"Élasticité marketing mesurée (06). Repli si historique manquant.",
 "Effort":"Effort marketing appliqué = levier 'variation budget marketing' × coefficient marque.",
 "Conv":"Taux de conversion candidature→inscrit (mesuré par cellule) + gain du levier conversion.",
 "Effectif Bud":"Effectif budget = nouveaux + réinscrits.",
 "CA Bud":"CA budget = effectif × tarif × facteur financement + frais de dossier.",
 "Cl besoin":"Classes nécessaires = arrondi supérieur (effectif / capacité) — le MINIMUM par cellule (promo). Ce minimum ne se réduit pas au sein d'une promo ; le regroupement de sections parallèles d'un même campus se traite un cran au-dessus (feuille 11b_Mutualisation).",
 "Enseign":"Coût d'enseignement = classes × coût par classe.",
 "Pédago€":"Coût pédagogique = effectif × coût pédago/étudiant × (1 + inflation).",
 "Mktg":"Marketing = nouveaux × CAC variable × (1 + effort). = 0 en année de poursuite.",
 "Contrib":"MARGE DE CONTRIBUTION = CA − enseignement − pédago − marketing − autres charges.",
 "Rempl":"Taux de remplissage = effectif / (classes × capacité).",
 "Rempl.":"Taux de remplissage = effectif / (classes × capacité).",
 "Contr/étu":"Contribution par étudiant.","Pt mort":"Point mort = nb d'étudiants pour couvrir le coût des classes.",
 # 09 Allocation
 "Contribution":"Marge de contribution (CA − coûts directs évitables).",
 "Loyer":"Loyer du campus (structure fixe).","Masse perm.":"Masse salariale permanente (structure fixe).",
 "Driver":"Valeur du driver d'allocation (effectif, CA ou m² selon 02_Leviers).",
 "Part":"Part du campus dans le driver total.","Alloué":"Frais de structure groupe alloués à ce campus.",
 "EBITDA campus":"EBITDA du campus = contribution − loyer − permanents − structure allouée.",
 "D&A":"Dotations aux amortissements.","EBIT":"Résultat d'exploitation = EBITDA − D&A.","EBIT campus":"EBIT du campus = EBITDA campus − D&A.",
 # 10 PnL
 "Rubrique":"Ligne du compte de résultat.","Réalisé N-1":"Valeur réelle de l'an dernier.","Budget N+1":"Valeur budgétée.",
 "Écart €":"Budget − Réalisé (en €).","Écart %":"Variation en % vs l'an dernier.","Effet":"Composante du pont de passage.","Montant":"Montant de l'effet (en €).",
 # 11 Simulateur
 "Effectif":"Effectif budget de la promo.","CA":"Chiffre d'affaires budget de la promo.",
 "Coûts directs évitables":"Coûts qui DISPARAISSENT si on ferme la promo (enseignement, pédago, marketing, autres variables).",
 "MARGE DE CONTRIBUTION":"CA − coûts directs évitables. C'est LA métrique de décision : on ne ferme que si elle est négative.",
 "Structure allouée (après, dilution)":"Quote-part de structure fixe (loyer, permanents, siège) portée par la promo, RECALCULÉE après décisions : moins d'étudiants au total → chacun en porte plus (dilution).",
 "Résultat tout compris (après)":"Contribution après − structure allouée. Vue INFORMATIVE — surtout PAS un critère de fermeture.",
 "🤖 Reco":"Recommandation automatique (contribution + remplissage). Dépend de l'année : entrée (lancer/capter) vs poursuite (regrouper).",
 "Motif":"Explication de la recommandation.","Décision":"Ton choix. Menu restreint selon l'année : entrée = lancer/capter ; poursuite = seulement regrouper.",
 "Effectif après":"Effectif après ta décision (0 si 'ne pas lancer' ; +1 classe captée si 'ouvrir').",
 "Contribution après":"Contribution recalculée selon ta décision. Regrouper (−1 classe) n'est crédité (économie d'une classe) QUE si l'effectif tient dans les classes restantes : effectif ≤ (classes−1) × capacité × 1,10 (tolérance salle +10%). Sinon aucune économie : le regroupement déborderait la capacité.",
 "Δ EBITDA":"Impact sur l'EBITDA groupe = contribution après − contribution avant.",
 # 11b Mutualisation
 "Cycle":"Cycle de diplôme (Bachelor / Mastère / BTS). On ne mutualise qu'au sein d'un même cycle : ses années partagent un tronc commun ; on ne mélange pas des cycles différents.",
 "Sections aujourd'hui":"Nombre de classes ouvertes aujourd'hui pour ce cycle sur le campus = somme des classes de chaque année (chacune arrondit au minimum de son côté).",
 "Capacité moy.":"Capacité moyenne d'une classe sur le campus (pondérée par le nombre de classes).",
 "Sections si mutualisé":"Plancher physique = ARRONDI.SUP(effectif total du campus / capacité moyenne). C'est le minimum de classes si on regroupait les étudiants dans des classes pleines — la capacité n'est jamais dépassée.",
 "Sections économisables":"Classes récupérables = sections aujourd'hui − plancher. Vient des arrondis : chaque promo arrondit au minimum, mais mises en commun elles remplissent mieux.",
 "Coût moyen / classe":"Coût d'enseignement moyen d'une classe du campus = coût enseignement total / nombre de classes.",
 "Économie potentielle €":"Sections économisables × coût moyen/classe × % d'heures mutualisables. Potentiel indicatif (bac à sable), pas un impact P&L officiel.",
 # 11c Cascade
 "Promo (ISCOM Toulouse)":"Promotion = programme × année du campus focus (ISCOM Toulouse).",
 "Eff. réel (moteur)":"Effectif budget RÉEL calculé par le moteur (feuille 08) pour cette promo — sert de référence en regard du scénario.",
 "Eff. scénario":"Effectif que TU testes (bleu, éditable). Baisse-le pour simuler une promo sous-remplie / en extinction et voir l'effet.",
 "Classes":"Nombre de classes = arrondi supérieur (effectif scénario / capacité).",
 "Contribution (évitable)":"CA − coûts directs évitables (enseignement + charge/étu + CAC). C'est LA métrique de décision : on ne ferme que si elle est négative.",
 "Struct/classe":"Structure non-évitable du campus (envelope) ÷ nombre total de classes. C'est ce qui se REDIVISE quand on ferme une promo.",
 "Struct allouée":"Structure portée par la promo = ses classes × structure/classe.",
 "Résultat t.c. (avant)":"Résultat TOUT COMPRIS = contribution − structure allouée. INFORMATIF : peut être négatif alors que la contribution est positive (le piège). JAMAIS un critère de fermeture.",
 "Cl. après":"Nombre de classes après ta décision (0 si Fermer).",
 "Struct après":"Structure allouée après décisions : l'envelope (inchangée, non-évitable) se redivise sur les classes RESTANTES → chacune en porte plus.",
 "Résultat t.c. (après)":"Résultat tout compris recalculé après décisions. En ROUGE si < 0 : rend visible l'effet d'entraînement.",
 "Coût/classe":"Coût d'enseignement d'une classe (heures × taux horaire), issu du modèle (feuille 05).",
 "Tarif":"Tarif de scolarité par étudiant et par an (feuille 05).",
 "Charge/étu":"Charges variables par étudiant = coût pédagogique + autres charges d'exploitation (feuille 05).",
 "CAC/étu":"Coût d'acquisition (achat de leads) par étudiant — uniquement en année d'entrée.",
 # divers entêtes
 "Clé (marque|campus)":"Clé technique marque|campus : le moteur s'en sert pour retrouver le bon coefficient stratégique du campus.",
 "Heures/classe":"Heures d'enseignement délivrées par classe et par an.",
 "Coût pédago/étu":"Coût pédagogique par étudiant/an, HORS salaires des intervenants (supports, licences, LMS, jury/examens, certifications). 350-670 € selon le domaine.",
 "Concept":"Notion du modèle Excel à transposer dans Tagetik.",
 "Objet Tagetik":"Objet CCH Tagetik correspondant (dimension, hiérarchie, règle de calcul, workflow…).",
 "Détail":"Précision sur la correspondance et le paramétrage Tagetik.",
 # 12 Sensibilite
 "Levier (pas)":"Levier testé et son incrément.","Impact EBITDA":"Effet € sur l'EBITDA d'un pas du levier (approximation vivante).","Lecture":"Interprétation.",
 # 13 Simulation
 "Indicateur":"Indicateur de pilotage.","Budget":"Valeur budgétée.","N-1":"Valeur réelle de l'an dernier.","Évolution":"Variation Budget vs N-1.",
 # 03 Coeff
 "Intensité MARKETING":"Coefficient appliqué à l'effort marketing du cadrage pour cette marque (>1 = on pousse).",
 "Intensité PRIX":"Coefficient appliqué à la hausse tarifaire du cadrage pour cette marque.","Posture":"Lecture du positionnement (pousser / défendre / maintenir).",
 # 01 Objectifs
 "Cible (direction)":"Objectif fixé par la direction (saisie).","Budget construit":"Résultat du budget construit par les drivers (moteur).","Écart":"Budget construit − cible.",
}
GLOSS={norm(k):v for k,v in GLOSS.items()}
HEADER_ROW={"04_Referentiel":5,"05_Param_Prog_Annee":3,"06_Historique":3,
 "07_Structure":3,"08_Moteur":2,"09_Allocation":5,"10_PnL":5,"11_Simulateur":5,"11b_Mutualisation":5,"11c_Cascade":14,"12_Sensibilite":5,"13_Simulation":5,"14_Mapping_Tagetik":5}
def add_note(cell,text):
    cm=Comment(text,"Guide"); cm.width=300; cm.height=120; cell.comment=cm
for shname,hr in HEADER_ROW.items():
    ws=wb[shname]
    for col in range(1,46):
        cell=ws.cell(row=hr,column=col)
        if cell.value is not None:
            key=norm(cell.value)
            if key in GLOSS: add_note(cell,GLOSS[key])
# 01_Cadrage : notes de guidage sur le poste de commande CFO
ws=wb["01_Cadrage"]
add_note(ws["D3"],"Scénario actif : change cette cellule (Cadrage/Optimiste/Prudent) et TOUT le budget se recalcule.")
add_note(ws["D4"],"Base de projection : ATTERRISSAGE N (réel YTD + reprévision) ou RÉALISÉ N-1 (année clôturée). Le budget construit projette depuis la base retenue (facteur en F4).")
for c,t in {"C6":"Réel de l'avant-dernière année (clôturé) — référence de tendance.","D6":"Réalisé de l'an dernier (année clôturée) — base possible de projection.","E6":"Dernier ATTERRISSAGE de l'année en cours (réel YTD + reprévision) — base par défaut.","F6":"Budget monté par le moteur (feuille 08), projeté depuis la BASE retenue (D4) — bouge avec les leviers.","G6":"Objectif fixé par la direction (saisie).","H6":"Écart = budget − objectif."}.items(): add_note(ws[c],t)
add_note(ws["G7"],"Objectif de CA fixé par la direction (saisie).")
add_note(ws["G8"],"Objectif d'EBITDA fixé par la direction (saisie).")
add_note(ws["F11"],"Reste à trouver = objectif EBITDA − budget construit (si positif). À combler via les leviers (voir 12_Sensibilite).")
for col,tx in {"D15":"Référence au dernier atterrissage (0 % = statu quo).","E15":"Valeurs du scénario Cadrage (central).","F15":"Valeurs du scénario Optimiste.","G15":"Valeurs du scénario Prudent.","H15":"Valeur ACTIVE = celle du scénario choisi en D3."}.items():
    add_note(ws[col],tx)
levn={16:"Variation du budget marketing → pilote le volume via l'élasticité mesurée.",17:"Hausse tarifaire moyenne, modulée par marque (coeff prix).",18:"Points de conversion candidature→inscrit gagnés (admissions).",19:"Amélioration du taux de passage (réinscription).",20:"Inflation appliquée aux charges (pédago, loyers, autres).",21:"Revalorisation des rémunérations chargées (permanents + enseignement)."}
for rr,tx in levn.items(): add_note(ws.cell(row=rr,column=2),tx)
add_note(ws["L6"],"Intensité marketing par campus (>1 = on pousse, <1 = on défend).")
add_note(ws["M6"],"Intensité prix par campus.")

# ---- OBJET DE L'ONGLET : une note explicite sur le titre de CHAQUE feuille (pour guider le consultant) ----
OBJET={
 "00_Notice":"Sommaire et mode d'emploi du classeur : rôle de chaque feuille, légende des couleurs, périmètre. Commence ici.",
 "01_Cadrage":"POSTE DE COMMANDE CFO. Cadrage top-down (N-2 · atterrissage N · objectif · budget construit · écart · reste à trouver), leviers en % avec scénarios Cadrage/Optimiste/Prudent, coefficients stratégiques par marque×campus, et graphes. Tout le budget se pilote d'ici.",
 "04_Referentiel":"Dimensions et plan de comptes du modèle (entités Groupe→Marque→Campus, comptes fixe/variable) — le socle de structuration.",
 "05_Param_Prog_Annee":"Données de référence UNITAIRES par programme × année (capacité, heures, taux, pédago, CAC, passage) qui alimentent le moteur. À charger avec le réel.",
 "06_Historique":"Réalisé N-1 par cellule (funnel + cohorte) et historique marketing N-2/N-1 servant à MESURER la conversion, le passage et l'élasticité (pas d'invention).",
 "07_Structure":"Réalisé N-1 par CAMPUS (loyers, ETP permanents, D&A, m²) : la base des coûts de structure.",
 "08_Moteur":"CŒUR DE CALCUL : construit le budget de CHAQUE cellule (cohortes, marketing→volume, tarif, financement alternance, point mort). Toutes les feuilles de résultat en découlent.",
 "09_Allocation":"Répartir les frais de structure FIXES sur les campus selon un driver (effectifs / CA / m²).",
 "10_PnL":"Compte de résultat consolidé N-1 vs Budget + PONT d'explication du CA (volume / tarif / frais).",
 "11_Simulateur":"BAC À SABLE de décisions à la maille PROMO (ouvrir / fermer / regrouper) → impact EBITDA AVANT/APRÈS. N'affecte PAS le budget officiel.",
 "11b_Mutualisation":"BAC À SABLE à la maille CAMPUS × CYCLE : combien de classes / € récupérables en mutualisant le tronc commun, sans dépasser la capacité.",
 "11c_Cascade":"BAC À SABLE : allocation de la structure EN CASCADE (marque→campus→classe) et démonstration de QUAND fermer une promo est bénéfique ou non, sur un vrai campus.",
 "12_Sensibilite":"Classer les LEVIERS par impact € sur l'EBITDA, pour savoir lequel actionner afin de combler l'écart de cadrage.",
 "13_Simulation":"TABLEAU DE BORD : KPIs clés Budget vs N-1 (effectif, CA, EBITDA, taux d'alternance).",
 "14_Mapping_Tagetik":"PASSERELLE : correspondance entre les objets du modèle Excel et leur implémentation dans CCH Tagetik.",
 "DATA_Referentiel":"RÉFÉRENTIEL des dimensions à charger dans Tagetik : chaque membre a un CODE et une DESCRIPTION, avec les HIÉRARCHIES (Entité Groupe→Marque→Campus ; Compte P&L + statistiques ; Version/scénario ; Programme ; Année ; Modalité ; Période).",
 "DATA_Chargement":"TABLE DE FAITS à charger dans Tagetik, en format LONG : une ligne = un croisement (entité × programme × année × modalité) × COMPTE × VERSION. Comptes comptables réels (PCG) pour le financier, codes techniques pour les stats. La structure (loyers, permanents, siège, D&A) est DÉJÀ ALLOUÉE jusqu'à la classe. 3 versions : Réalisé N-2, Réalisé N-1, Atterrissage (Forecast V4).",
 "REPORTING":"DASHBOARD CFO à montrer juste après le chargement : tuiles KPI (CA, EBITDA, marge, CA/étudiant), P&L, profitabilité par marque (mise en couleur), top/flop campus, funnel, trajectoire 3 ans. Graphiques natifs sourcés sur le réalisé (07_Structure).",
}
for shname,obj in OBJET.items():
    ws=wb[shname]; placed=False
    for r in range(1,4):
        for col in range(1,6):
            cell=ws.cell(row=r,column=col)
            if isinstance(cell.value,str) and len(cell.value)>12 and not cell.value.startswith("="):
                cm=Comment("OBJET DE L'ONGLET — "+obj,"Guide"); cm.width=340; cm.height=170; cell.comment=cm; placed=True; break
        if placed: break

try: wb.calculation.fullCalcOnLoad=True
except Exception:
    from openpyxl.workbook.properties import CalcProperties; wb.calculation=CalcProperties(fullCalcOnLoad=True)
wb.save(OUT); print("TOUTES LES FEUILLES OK (avec notes d'entêtes)")
