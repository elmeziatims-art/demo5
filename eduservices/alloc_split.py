#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Sort un pool 'FRAIS DE MARQUE' (compte 6236 pub/marque) du siege, avec sa
PROPRE cle de niveau 1 (ALLOC_GRP_MARQUE) -> il est reparti groupe->marque par
une logique differente du holding, puis reste dans la marque. Visible partout.
  - HOLDING = 6414,6226,626,6281,6331,6333  (cascade K1 = ALLOC_GRP_BRAND)
  - MARQUE  = 6236                           (cascade K4 = ALLOC_GRP_MARQUE)
Total marge groupe inchange (les 2 pools telescopent). Seule la repartition bouge.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import ColorScaleRule
F="CAD_SAAD_LIVE.xlsx"; wb=openpyxl.load_workbook(F)
ca=wb["_CALC_ALLOC"]; al=wb["Allocation"]; ps=wb["Pilotage"]
NM=2000
def cpt(acct): return 'SUMIFS(Compta!$F:$F,Compta!$A:$A,"GRP",Compta!$B:$B,"%s",Compta!$C:$C,$A{r})'%acct
HOLD="+".join(cpt(a) for a in ("6414","6226","626","6281","6331","6333"))
MARQ=cpt("6236")
def g(inner): return '=IF($A{r}="","",'+inner+')'
sel=lambda kcell,cca,ceff,ccls: 'IF(Pilotage!%s="Chiffre d\'affaires",%s,IF(Pilotage!%s="Effectif",%s,%s))'%(kcell,cca,kcell,ceff,ccls)
tpl={
 "AB":g(HOLD),                                        # HOLDING_TOT (6236 retire)
 "AO":g(MARQ),                                        # FRAIS_MARQUE_TOT (6236)
 "AP":g(sel("$C$55","$S{r}","$Q{r}","$R{r}")),        # D1M selon K4 (marque)
 "AQ":g(sel("$C$55","$V{r}","$T{r}","$U{r}")),        # D1G selon K4 (groupe)
 "AM":g('$AB{r}*IFERROR($AG{r}/$AH{r},0)*IFERROR($AE{r}/$AF{r},0)*IFERROR($AC{r}/$AD{r},0)'),  # COST_HOLDING
 "AR":g('$AO{r}*IFERROR($AP{r}/$AQ{r},0)*IFERROR($AE{r}/$AF{r},0)*IFERROR($AC{r}/$AD{r},0)'),  # COST_MARQUE
 "AN":g('$I{r}-($AI{r}+$AJ{r}+$AK{r}+$AL{r}+$AM{r}+$AR{r})'),  # MARGE
}
for r in range(2,NM+1):
    for col,t in tpl.items(): ca["%s%d"%(col,r)]=t.format(r=r)
ca["AO1"]="FRAIS_MARQUE_TOT"; ca["AP1"]="D1M_K4"; ca["AQ1"]="D1G_K4"; ca["AR1"]="COST_MARQUE"

# ---- Allocation a-cote : Z=holding, AC=marque, AA=marge ----
al["Z1"]="COST_HOLDING (live)"; al["AC1"]="COST_MARQUE (live)"
al["Z1"].font=Font(bold=True,color="843C0C"); al["AC1"].font=Font(bold=True,color="7030A0")
for r in range(2,NM+1):
    al["Z%d"%r]='=IF(Allocation!D%d="","",_CALC_ALLOC!AM%d)'%(r,r)     # holding
    al["AC%d"%r]='=IF(Allocation!D%d="","",_CALC_ALLOC!AR%d)'%(r,r)    # marque
    al["AA%d"%r]='=IF(Allocation!D%d="","",_CALC_ALLOC!AN%d)'%(r,r)    # marge
    al["AC%d"%r].number_format="# ##0"

# =========================================================== PILOTAGE
NAVY="1F3864";AMBER="BF8F00";INPUT="FFF2CC";GREY_L="F2F2F2";WHITE="FFFFFF";LIVE="C6E0B4"
PURP="7030A0";FCE="FCE4D6";GREEN="548235"
thin=Side(style="thin",color="BFBFBF");box=Border(thin,thin,thin,thin)
def fill(h):return PatternFill("solid",fgColor=h)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEF=Alignment(horizontal="left",vertical="center");RIG=Alignment(horizontal="right",vertical="center")
def S(ref,val=None,font=None,fillc=None,align=None,border=None,fmt=None):
    c=ps[ref]
    if val is not None:c.value=val
    if font:c.font=font
    if fillc:c.fill=fill(fillc)
    if align:c.alignment=align
    if border:c.border=border
    if fmt:c.number_format=fmt
H1=Font(size=10,bold=True,color=WHITE);LABB=Font(size=10,bold=True,color=NAVY)
INPF=Font(size=11,bold=True,color="7F6000");VAL=Font(size=9);LIVEF=Font(size=9,bold=True,color="375623")

# --- 4e cle : ALLOC_GRP_MARQUE en B55/C55 (remplace l'ancienne note) ---
for cc in ("B","C","D","E"): ps["%s55"%cc].value=None; ps["%s55"%cc].fill=PatternFill()
S("B55","ALLOC_GRP_MARQUE",font=LABB,align=LEF,border=box,fillc=GREY_L)
ps.merge_cells("C55:E55")
S("C55","Effectif",font=INPF,align=CEN,border=box,fillc=INPUT)
for cc in ("D","E"): ps["%s55"%cc].fill=fill(INPUT); ps["%s55"%cc].border=box
dv=DataValidation(type="list",formula1="=$A$1:$A$3",allow_blank=False); ps.add_data_validation(dv); dv.add("C55")
ps["B54"].value="ALLOC_CAMP_CLASS"  # ensure labels intact

# --- effet table 56-63 : split Frais marque / Frais holding (10 colonnes B-K) ---
heads=["Marque","VAC","PERM","ODIR","STRUCT campus","Frais marque","Frais holding","Cout complet","Marge complete","Marge %"]
for i,h in enumerate(heads):
    c=ps.cell(56,2+i,h); c.font=H1; c.fill=fill(AMBER); c.alignment=CEN; c.border=box
ps.row_dimensions[56].height=26
MARQUES=[("MBWAY","MBway","2E75B6"),("ISCOM","ISCOM","548235"),("IPAC","Ipac","BF8F00"),
         ("PIGIER","Pigier","843C0C"),("TUNON","Tunon","7030A0")]
alc={"VAC":"V","PERM":"W","ODIR":"X","STRUCT":"Y","MARQUE":"AC","HOLDING":"Z"}
def sif(col,code): return '=SUMIFS(Allocation!$%s:$%s,Allocation!$E:$E,"%s",Allocation!$C:$C,"2026")'%(col,col,code)
for i,(code,lab,_) in enumerate(MARQUES):
    r=57+i; band=WHITE if i%2==0 else GREY_L
    S("B%d"%r,lab,font=LABB,align=LEF,border=box,fillc=band)
    S("C%d"%r,sif("V",code),font=VAL,align=RIG,border=box,fillc=band,fmt="# ##0")
    S("D%d"%r,sif("W",code),font=VAL,align=RIG,border=box,fillc=band,fmt="# ##0")
    S("E%d"%r,sif("X",code),font=VAL,align=RIG,border=box,fillc=band,fmt="# ##0")
    S("F%d"%r,sif("Y",code),font=VAL,align=RIG,border=box,fillc=band,fmt="# ##0")
    S("G%d"%r,sif("AC",code),font=Font(size=9,bold=True,color=PURP),align=RIG,border=box,fillc="E9D9F2",fmt="# ##0")  # marque
    S("H%d"%r,sif("Z",code),font=VAL,align=RIG,border=box,fillc=band,fmt="# ##0")                                    # holding
    S("I%d"%r,"=SUM(C%d:H%d)"%(r,r),font=VAL,align=RIG,border=box,fillc=band,fmt="# ##0")
    S("J%d"%r,sif("AA",code),font=LIVEF,align=RIG,border=box,fillc=LIVE,fmt="# ##0")
    S("K%d"%r,'=IFERROR(J%d/SUMIFS(Allocation!$K:$K,Allocation!$E:$E,"%s",Allocation!$C:$C,"2026"),0)'%(r,code),
      font=VAL,align=RIG,border=box,fillc=band,fmt="0.0%")
# total row 62
S("B62","GROUPE",font=Font(bold=True,color=WHITE),fillc=NAVY,align=LEF,border=box)
for col in "CDEFGHIJ":
    S("%s62"%col,"=SUM(%s57:%s61)"%(col,col),font=Font(bold=True,color=WHITE),fillc=NAVY,align=RIG,border=box,fmt="# ##0")
S("K62",'=IFERROR(J62/SUMIFS(Allocation!$K:$K,Allocation!$C:$C,"2026"),0)',font=Font(bold=True,color=WHITE),fillc=NAVY,align=RIG,border=box,fmt="0.0%")

# --- rollup marge (col AA rows 67-71) -> pointe vers J (marge) ---
for i in range(5): ps["AA%d"%(67+i)]="=J%d"%(57+i)

# --- heatmaps effet : Frais marque(G), Frais holding(H), Cout complet(I), Marge%(K) ---
def heat(rng,hi="F8696B"):
    ps.conditional_formatting.add(rng, ColorScaleRule(start_type='min',start_color='63BE7B',
        mid_type='percentile',mid_value=50,mid_color='FFEB84',end_type='max',end_color=hi))
for rng in ["G57:G61","H57:H61","I57:I61"]: heat(rng)                 # couts lourds = rouge
heat("K57:K61","63BE7B")  # marge% haute = vert  (start rouge)
ps.conditional_formatting.add("K57:K61", ColorScaleRule(start_type='min',start_color='F8696B',
    mid_type='percentile',mid_value=50,mid_color='FFEB84',end_type='max',end_color='63BE7B'))

# --- charts : reconstruire marge/marque + decomposition (6 postes) ---
# retirer les 2 anciens charts d'allocation (garder cap/donut/CA&EBITDA)
keep=[c for c in ps._charts if getattr(c,'anchor',None) and str(getattr(c.anchor,'_from',None)) and c.anchor._from.row<45]
ps._charts=[c for c in ps._charts if c not in ps._charts[3:]]  # garder les 3 premiers (S4,S30,S45)
def dpts(): return [DataPoint(idx=i,spPr=GraphicalProperties(solidFill=m[2])) for i,m in enumerate(MARQUES)]
# marge complete par marque (reagit aux cles)
mm=BarChart(); mm.type="col"; mm.style=12; mm.title="Effet des cles : marge complete par marque"
mm.height=7.5; mm.width=13
mm.add_data(Reference(ps,min_col=10,min_row=56,max_row=61),titles_from_data=True)  # J marge
mm.set_categories(Reference(ps,min_col=2,min_row=57,max_row=61)); mm.legend=None
mm.series[0].data_points=dpts(); mm.y_axis.numFmt="# ##0"; ps.add_chart(mm,"L51")
# decomposition cout complet (6 postes incl marque + holding)
st=BarChart(); st.type="col"; st.grouping="stacked"; st.overlap=100
st.title="Decomposition du cout complet (dont frais marque vs holding)"
st.height=7.5; st.width=13
for j in range(6):  # C..H
    st.add_data(Reference(ps,min_col=3+j,min_row=56,max_row=61),titles_from_data=True)
st.set_categories(Reference(ps,min_col=2,min_row=57,max_row=61)); st.y_axis.numFmt="# ##0"
ps.add_chart(st,"L67")

# widths
ps.column_dimensions["G"].width=13; ps.column_dimensions["H"].width=13
ps.column_dimensions["I"].width=13; ps.column_dimensions["J"].width=14
wb.calculation.fullCalcOnLoad=True
wb.save(F)
print("OK split frais de marque (6236, cle ALLOC_GRP_MARQUE) / holding. Visible dans la table + charts.")
